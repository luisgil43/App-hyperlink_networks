# fleet/views_user.py
from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from typing import Dict, List, Optional, Tuple

from django.contrib.auth.decorators import login_required
from django.http import Http404, HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from usuarios.decoradores import rol_requerido

from .models import (Vehicle, VehicleAssignment, VehicleService,
                     VehicleServiceType)


def _badge_classes(status: str) -> str:
    if status == "vencido":
        return "bg-red-100 text-red-800"
    if status == "proximo":
        return "bg-yellow-100 text-yellow-800"
    if status == "vigente":
        return "bg-emerald-100 text-emerald-800"
    if status == "sin_config":
        return "bg-gray-100 text-gray-800"
    return "bg-gray-100 text-gray-800"


def _safe_int(v) -> Optional[int]:
    try:
        if v is None or v == "":
            return None
        return int(v)
    except Exception:
        return None


def _parse_steps_csv(csv_value: str | None) -> List[int]:
    if not csv_value:
        return []
    out: List[int] = []
    for part in str(csv_value).split(","):
        part = part.strip()
        if not part:
            continue
        iv = _safe_int(part)
        if iv is not None and iv > 0:
            out.append(iv)
    return sorted(set(out), reverse=True)


def _pick_steps(service_type: VehicleServiceType, kind: str) -> List[int]:
    """
    kind: 'km' o 'days'
    """
    if kind == "km":
        steps = _parse_steps_csv(
            getattr(service_type, "alert_before_km_steps", "") or ""
        )
        if steps:
            return steps
        legacy = _safe_int(getattr(service_type, "alert_before_km", None))
        return [legacy] if legacy and legacy > 0 else []
    else:
        steps = _parse_steps_csv(
            getattr(service_type, "alert_before_days_steps", "") or ""
        )
        if steps:
            return steps
        legacy = _safe_int(getattr(service_type, "alert_before_days", None))
        return [legacy] if legacy and legacy > 0 else []


@dataclass
class MaintenanceRow:
    service_type: VehicleServiceType
    last_service: Optional[VehicleService]

    next_due_km: Optional[int]
    next_due_date: Optional[date]

    remaining_km: Optional[int]
    remaining_days: Optional[int]

    status: str  # vigente | proximo | vencido | sin_config
    status_label: str
    badge_class: str


@dataclass
class HistoryGroup:
    key: str
    label: str
    total_count: int
    rows: List[VehicleService]


@dataclass
class AssignmentPanel:
    assignment: VehicleAssignment
    vehicle: Vehicle
    estado_label: str
    badge_class: str
    assigned_at_label: str
    closed_at_label: str

    maint_rows: List[MaintenanceRow]
    history_groups: List[HistoryGroup]


def _calc_status_for_row(
    *,
    today: date,
    service_type: VehicleServiceType,
    vehicle_km: Optional[int],
    next_due_km: Optional[int],
    next_due_date: Optional[date],
) -> Tuple[str, str]:
    if next_due_km is None and next_due_date is None:
        return ("sin_config", "No config")

    # overdue by miles
    if next_due_km is not None and vehicle_km is not None:
        remaining_km = next_due_km - vehicle_km
        if remaining_km <= 0:
            return ("vencido", "Overdue")

    # overdue by date
    if next_due_date is not None:
        remaining_days = (next_due_date - today).days
        if remaining_days <= 0:
            return ("vencido", "Overdue")

    # due soon by steps
    due_soon = False

    if next_due_km is not None and vehicle_km is not None:
        remaining_km = next_due_km - vehicle_km
        km_steps = _pick_steps(service_type, "km")
        if any(remaining_km <= s for s in km_steps):
            due_soon = True

    if next_due_date is not None:
        remaining_days = (next_due_date - today).days
        day_steps = _pick_steps(service_type, "days")
        if any(remaining_days <= s for s in day_steps):
            due_soon = True

    if due_soon:
        return ("proximo", "Due soon")

    return ("vigente", "OK")


def _service_group_key_and_label(s: VehicleService) -> Tuple[str, str]:
    if getattr(s, "service_type_obj_id", None):
        st = getattr(s, "service_type_obj", None)
        name = (st.name if st else "Service").strip()
        return (f"obj:{s.service_type_obj_id}", name)

    try:
        label = s.get_service_type_display()
    except Exception:
        label = str(getattr(s, "service_type", "Service"))

    return (f"legacy:{getattr(s, 'service_type', 'other')}", label)


def _parse_hist_limit(request) -> int:
    raw = (request.GET.get("hist") or "").strip()
    if raw in ("5", "10", "20"):
        return int(raw)
    return 5


def _fmt_dt(dt) -> str:
    if not dt:
        return "—"
    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)
    return dt.strftime("%d-%m-%Y %H:%M")


@login_required
@rol_requerido("usuario")
def my_vehicle_dashboard(request):
    user = request.user
    today = timezone.localdate()
    hist_limit = _parse_hist_limit(request)

    assignments = (
        VehicleAssignment.objects.select_related("vehicle", "vehicle__status")
        .filter(user=user)
        .order_by("-is_active", "-assigned_at", "-id")
    )

    panels: List[AssignmentPanel] = []

    for a in assignments:
        v = a.vehicle

        start_dt = a.assigned_at
        end_dt = a.closed_at if (not a.is_active and a.closed_at) else timezone.now()

        estado_label = "Active" if a.is_active else "Closed"
        estado_badge = (
            "bg-emerald-100 text-emerald-800"
            if a.is_active
            else "bg-gray-200 text-gray-800"
        )

        services_qs = (
            VehicleService.objects.filter(vehicle=v)
            .select_related("service_type_obj")
            .order_by("-created_at", "-id")
        )

        if start_dt:
            services_qs = services_qs.filter(created_at__gte=start_dt)
        if end_dt:
            services_qs = services_qs.filter(created_at__lte=end_dt)

        services_list = list(services_qs[:2000])

        vehicle_km = _safe_int(getattr(v, "kilometraje_actual", None))
        types_qs = VehicleServiceType.objects.filter(is_active=True).order_by("name")

        last_by_type: Dict[int, VehicleService] = {}
        for s in services_list:
            st = getattr(s, "service_type_obj", None)
            if st and st.id not in last_by_type:
                last_by_type[st.id] = s

        maint_rows: List[MaintenanceRow] = []
        for st in types_qs:
            last = last_by_type.get(st.id)

            next_due_km = (
                _safe_int(getattr(last, "next_due_km", None)) if last else None
            )
            next_due_date = getattr(last, "next_due_date", None) if last else None

            remaining_km = None
            if next_due_km is not None and vehicle_km is not None:
                remaining_km = next_due_km - vehicle_km

            remaining_days = None
            if next_due_date is not None:
                remaining_days = (next_due_date - today).days

            status, status_label = _calc_status_for_row(
                today=today,
                service_type=st,
                vehicle_km=vehicle_km,
                next_due_km=next_due_km,
                next_due_date=next_due_date,
            )

            maint_rows.append(
                MaintenanceRow(
                    service_type=st,
                    last_service=last,
                    next_due_km=next_due_km,
                    next_due_date=next_due_date,
                    remaining_km=remaining_km,
                    remaining_days=remaining_days,
                    status=status,
                    status_label=status_label,
                    badge_class=_badge_classes(status),
                )
            )

        rank = {"vencido": 0, "proximo": 1, "vigente": 2, "sin_config": 3}

        def sort_key(r: MaintenanceRow):
            urg_days = r.remaining_days if r.remaining_days is not None else 10**9
            urg_km = r.remaining_km if r.remaining_km is not None else 10**9
            return (
                rank.get(r.status, 9),
                urg_days,
                urg_km,
                (r.service_type.name or "").lower(),
            )

        maint_rows.sort(key=sort_key)

        buckets: Dict[str, Dict[str, object]] = {}
        for s in services_list:
            key, label = _service_group_key_and_label(s)
            if key not in buckets:
                buckets[key] = {"label": label, "rows": []}
            buckets[key]["rows"].append(s)

        groups: List[HistoryGroup] = []
        for key, obj in buckets.items():
            rows = obj["rows"]
            rows.sort(key=lambda x: (x.created_at or timezone.now()), reverse=True)
            groups.append(
                HistoryGroup(
                    key=key,
                    label=str(obj["label"]),
                    total_count=len(rows),
                    rows=rows[:hist_limit],
                )
            )

        groups.sort(
            key=lambda g: (
                g.rows[0].created_at
                if g.rows and g.rows[0].created_at
                else timezone.now()
            ),
            reverse=True,
        )

        panels.append(
            AssignmentPanel(
                assignment=a,
                vehicle=v,
                estado_label=estado_label,
                badge_class=estado_badge,
                assigned_at_label=_fmt_dt(a.assigned_at),
                closed_at_label=_fmt_dt(a.closed_at) if a.closed_at else "—",
                maint_rows=maint_rows,
                history_groups=groups,
            )
        )

    return render(
        request,
        "fleet/user_my_vehicle.html",
        {
            "today": today,
            "hist_limit": hist_limit,
            "panels": panels,
        },
    )


@login_required
@rol_requerido("usuario")
def my_vehicle_history_excel(request):
    user = request.user
    assignment_id = request.GET.get("assignment")
    group_key = (request.GET.get("group") or "").strip()

    if not assignment_id or not group_key:
        raise Http404("Missing parameters.")

    try:
        assignment_id = int(assignment_id)
    except Exception:
        raise Http404("Invalid assignment.")

    a = (
        VehicleAssignment.objects.select_related("vehicle")
        .filter(pk=assignment_id, user=user)
        .first()
    )
    if not a:
        raise Http404("Assignment not found.")

    v = a.vehicle
    start_dt = a.assigned_at
    end_dt = a.closed_at if (not a.is_active and a.closed_at) else timezone.now()

    qs = (
        VehicleService.objects.filter(vehicle=v)
        .select_related("service_type_obj")
        .order_by("-created_at", "-id")
    )
    if start_dt:
        qs = qs.filter(created_at__gte=start_dt)
    if end_dt:
        qs = qs.filter(created_at__lte=end_dt)

    rows: List[VehicleService] = []
    type_label = "Service"

    if group_key.startswith("obj:"):
        try:
            type_id = int(group_key.split(":", 1)[1])
        except Exception:
            raise Http404("Invalid group.")
        qs = qs.filter(service_type_obj_id=type_id)
        rows = list(qs[:10000])
        type_label = (
            VehicleServiceType.objects.filter(pk=type_id)
            .values_list("name", flat=True)
            .first()
            or f"Type_{type_id}"
        )
    elif group_key.startswith("legacy:"):
        legacy = group_key.split(":", 1)[1].strip()
        qs = qs.filter(service_type=legacy)
        rows = list(qs[:10000])
        type_label = legacy or "legacy"
    else:
        raise Http404("Invalid group.")

    wb = Workbook()
    ws = wb.active
    ws.title = "History"

    ws.append(
        [
            "Plate",
            "Vehicle",
            "Assignment status",
            "Assigned from",
            "Closed at",
            "Group",
            "Created at",
            "Odometer (miles)",
            "Amount",
            "Notes",
        ]
    )

    def safe_notes(x: str) -> str:
        return (x or "").replace("\r\n", "\n").strip()

    for s in rows:
        created_at = s.created_at
        if created_at and timezone.is_aware(created_at):
            created_at = timezone.localtime(created_at)

        ws.append(
            [
                v.patente,
                f"{(v.marca or '').strip()} {(v.modelo or '').strip()}".strip(),
                "Active" if a.is_active else "Closed",
                _fmt_dt(a.assigned_at),
                _fmt_dt(a.closed_at) if a.closed_at else "—",
                type_label,
                created_at.strftime("%d-%m-%Y %H:%M") if created_at else "—",
                s.kilometraje_declarado if s.kilometraje_declarado is not None else "",
                float(s.monto) if s.monto is not None else "",
                safe_notes(getattr(s, "notes", "")),
            ]
        )

    for col in range(1, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18

    filename = f"history_{v.patente}_{type_label}.xlsx".replace(" ", "_")

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
