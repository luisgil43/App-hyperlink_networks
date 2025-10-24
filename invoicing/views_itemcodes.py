import io
from decimal import Decimal, InvalidOperation

from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.http import FileResponse, HttpResponseNotAllowed, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse

from .forms_itemcodes import ItemCodeForm, ItemCodeImportForm
from .models import ItemCode


# -------- helpers import ----------
def _norm(s: str) -> str:
    """Normaliza encabezados: minúsculas, sin espacios extras, reemplaza guiones/underscores."""
    return (s or "").strip().lower().replace("_", " ").replace("-", " ")

HEADER_MAPS = {
    "city": "city",
    "project": "project",
    "office": "office",
    "client": "client",
    "work type": "work_type",
    "worktype": "work_type",
    "job code": "job_code",
    "jobcode": "job_code",
    "description": "description",
    "uom": "uom",
    "unit of measure": "uom",
    "rate": "rate",
    "rite": "rate",  # tolera typo
}

# ---------------- LIST ----------------
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator

from .models import ItemCode


@login_required
def itemcodes_list(request):
    f_city    = (request.GET.get("f_city") or "").strip()
    f_project = (request.GET.get("f_project") or "").strip()
    f_client  = (request.GET.get("f_client") or "").strip()
    f_code    = (request.GET.get("f_code") or "").strip()

    qs = ItemCode.objects.all().order_by("city", "project", "job_code")

    if f_city:
        qs = qs.filter(city__icontains=f_city)
    if f_project:
        qs = qs.filter(project__icontains=f_project)
    if f_client:
        qs = qs.filter(client__icontains=f_client)
    if f_code:
        qs = qs.filter(job_code__icontains=f_code)

    per_page = (request.GET.get("per") or "10").lower()
    page_num = int(request.GET.get("page") or 1)

    paginator = pagina = None
    items = qs
    if per_page != "all":
        try:
            per_int = max(1, int(per_page))
        except ValueError:
            per_int = 10
        paginator = Paginator(qs, per_int)
        pagina = paginator.get_page(page_num)
        items = pagina.object_list

    ctx = {
        "page_title": "Item Codes",
        "items": items,
        "pagina": pagina,
        "paginator": paginator,
        "per_page": per_page,
        "f_city": f_city,
        "f_project": f_project,
        "f_client": f_client,
        "f_code": f_code,
    }
    return render(request, "invoicing/itemcodes_list.html", ctx)



# ---------------- CREATE/EDIT ----------------
@login_required
def itemcodes_edit(request, pk=None):
    obj = get_object_or_404(ItemCode, pk=pk) if pk else None
    if request.method == "POST":
        form = ItemCodeForm(request.POST, instance=obj)
        if form.is_valid():
            form.save()
            return redirect("invoicing:itemcodes_list")
    else:
        form = ItemCodeForm(instance=obj)

    return render(request, "invoicing/itemcodes_edit.html", {
        "page_title": ("Edit Item Code" if obj else "New Item Code"),
        "form": form,
        "obj": obj,
    })

# ---------------- DELETE ----------------
@login_required
def itemcodes_delete(request):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])
    pk = request.POST.get("id")
    obj = get_object_or_404(ItemCode, pk=pk)
    obj.delete()
    return JsonResponse({"ok": True})

# ---------------- IMPORT ----------------
@login_required
def itemcodes_import(request):
    if request.method == "POST":
        form = ItemCodeImportForm(request.POST, request.FILES)
        if form.is_valid():
            f = form.cleaned_data["file"]
            # openpyxl lectura en memoria
            from openpyxl import load_workbook
            wb = load_workbook(filename=f, data_only=True)
            ws = wb.active

            # map headers
            header = {}
            for col_idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1):
                key = HEADER_MAPS.get(_norm(str(cell.value)), None)
                if key:
                    header[key] = col_idx

            required = ["job_code","description"]
            missing = [k for k in required if k not in header]
            if missing:
                return render(request, "invoicing/itemcodes_import.html", {
                    "form": form,
                    "error": f"Missing required headers: {', '.join(missing)}",
                })

            # iterate rows
            created = 0
            updated = 0
            for r in ws.iter_rows(min_row=2):
                def val(col_name):
                    c = header.get(col_name)
                    if not c: return ""
                    v = r[c-1].value
                    return "" if v is None else str(v).strip()

                job_code = val("job_code")
                if not job_code:
                    continue

                # convierto rate -> Decimal
                raw_rate = val("rate")
                rate = Decimal("0")
                if raw_rate != "":
                    try:
                        rate = Decimal(str(raw_rate).replace(",", ""))
                    except InvalidOperation:
                        rate = Decimal("0")

                data = {
                    "city":        val("city"),
                    "project":     val("project"),
                    "office":      val("office"),
                    "client":      val("client"),
                    "work_type":   val("work_type"),
                    "job_code":    job_code,
                    "description": val("description"),
                    "uom":         val("uom"),
                    "rate":        rate,
                }

                obj, created_flag = ItemCode.objects.update_or_create(
                    job_code=job_code,
                    defaults=data
                )
                if created_flag: created += 1
                else: updated += 1

            return render(request, "invoicing/itemcodes_import.html", {
                "form": ItemCodeImportForm(),
                "success": f"Import finished. Created: {created}, Updated: {updated}."
            })
    else:
        form = ItemCodeImportForm()

    return render(request, "invoicing/itemcodes_import.html", {"form": form})

# ---------------- TEMPLATE XLSX (descarga) ----------------
@login_required
def itemcodes_template(request):
    # genera un .xlsx con encabezados correctos y una fila de ejemplo
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "ItemCodes"

    headers = ["City","Project","Office","Client","Work Type","Job Code","Description","UOM","Rate"]
    ws.append(headers)
    ws.append(["Miami","Hyperlink 1","HQ","Planix","Labor","NET-ARCH","Network design & architecture","hr","150.00"])

    # auto ancho básico
    for col in range(1, len(headers)+1):
        ws.column_dimensions[chr(64+col)].width = 18

    buff = io.BytesIO()
    wb.save(buff)
    buff.seek(0)
    return FileResponse(buff, as_attachment=True, filename="ItemCodes_Template.xlsx")