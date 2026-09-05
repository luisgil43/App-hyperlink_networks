from __future__ import annotations

import json
from datetime import date
from zipfile import BadZipFile

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.core.paginator import Paginator
from django.db import transaction
from django.http import Http404, HttpRequest, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

from client_submissions.models import (ClientSubmission, ClientSubmissionBatch,
                                       ClientSubmissionEvent)
from client_submissions.services.billing_mapper import (
    build_billing_submission_snapshot, normalize_job_code)
from client_submissions.services.submission_builder import (
    EmptyBillingSelectionError, InvalidBatchConfigurationError,
    create_submission_batch, revalidate_batch, revalidate_submission)
from operaciones.models import SesionBilling
from usuarios.decoradores import rol_requerido

# ============================================================
# Configuración
# ============================================================


DEFAULT_SUBMITTED_BY_EMAIL = "l.suarez@hyperlink-networks.com"

DEFAULT_COPY_EMAIL = "l.suarez@hyperlink-networks.com"

DEFAULT_SUBCONTRACTOR_NAME = "Hyperlink"


# ============================================================
# Helpers de acceso
# ============================================================


ALLOWED_ROLES = (
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)


def _clean_text(
    value,
) -> str:
    if value is None:
        return ""

    return str(value).strip()


def _user_can_manage_client_submissions(
    user,
) -> bool:
    """
    Protección adicional.

    El decorador rol_requerido se mantiene en las vistas,
    pero este helper también permite proteger objetos concretos.
    """

    if not user or not user.is_authenticated:
        return False

    if user.is_superuser:
        return True

    role_candidates = {
        _clean_text(
            getattr(
                user,
                "rol",
                "",
            )
        ).lower(),
        _clean_text(
            getattr(
                user,
                "role",
                "",
            )
        ).lower(),
    }

    role_candidates.discard("")

    if role_candidates.intersection(
        ALLOWED_ROLES,
    ):
        return True

    try:
        roles = {
            _clean_text(
                role_name,
            ).lower()
            for role_name in user.roles.values_list(
                "name",
                flat=True,
            )
        }

        if roles.intersection(
            ALLOWED_ROLES,
        ):
            return True

    except Exception:
        pass

    return False


def _assert_manage_permission(
    request: HttpRequest,
) -> None:
    if not _user_can_manage_client_submissions(
        request.user,
    ):
        raise PermissionDenied(
            "You do not have permission to manage client submissions."
        )


# ============================================================
# Helpers de Billing / Client Submission
# ============================================================


def _get_snapshot_codes(
    billing_snapshot: dict,
) -> set[str]:
    """
    Obtiene los códigos reales del Billing snapshot.

    Ejemplos:

        C-108-UG
        C-108-AER
        C-108.1
        C-109
        C-110.2
    """

    items = billing_snapshot.get(
        "items",
        [],
    )

    if not isinstance(
        items,
        list,
    ):
        return set()

    codes = set()

    for item in items:
        if not isinstance(
            item,
            dict,
        ):
            continue

        code = normalize_job_code(
            item.get(
                "codigo_trabajo",
                "",
            )
        )

        if code:
            codes.add(
                code,
            )

    return codes


def _build_client_items_preview(
    billing_session,
) -> list[dict]:
    """
    Construye el resumen exacto de cantidades que serán enviadas
    al formulario Daily Production Reporting - SPLICING.

    Las cantidades salen de los códigos reales aprobados
    en el Invoice.

    Mapeo:

        C-108-UG
            -> C-108-UG - Splice Case Quantity

        C-108-AER
            -> C-108-AER - Splice Case Quantity

        C-109
            -> C-109 - HO-1 Fusion Splice Quantity

        C-110.2
            -> C-110 - DS Splitter Add - 1x2

        C-110.4
            -> C-110 - DS Splitter Add - 1x4

        C-110.8
            -> C-110 - DS Splitter Add - 1x8

        C-110.16
            -> C-110 - DS Splitter Add - 1x16

    C-108.1 no se muestra como cantidad del formulario.
    Únicamente activa Re-Entry = YES.
    """

    billing_snapshot = build_billing_submission_snapshot(
        billing_session,
    )

    fields = billing_snapshot.get(
        "fields",
        {},
    )

    if not isinstance(
        fields,
        dict,
    ):
        fields = {}

    codes = _get_snapshot_codes(
        billing_snapshot,
    )

    splice_case_code = ""

    if "C-108-AER" in codes:
        splice_case_code = "C-108-AER"

    elif "C-108-UG" in codes:
        splice_case_code = "C-108-UG"

    field_map = []

    if splice_case_code:
        field_map.append(
            (
                "splice_case_quantity",
                splice_case_code,
                "Splice Case Quantity",
            )
        )

    field_map.extend(
        [
            (
                "fusion_splice_quantity",
                "C-109",
                "HO-1 Fusion Splice Quantity",
            ),
            (
                "ds_splitter_1x2_quantity",
                "C-110.2",
                "DS Splitter Add - 1x2",
            ),
            (
                "ds_splitter_1x4_quantity",
                "C-110.4",
                "DS Splitter Add - 1x4",
            ),
            (
                "ds_splitter_1x8_quantity",
                "C-110.8",
                "DS Splitter Add - 1x8",
            ),
            (
                "ds_splitter_1x16_quantity",
                "C-110.16",
                "DS Splitter Add - 1x16",
            ),
        ]
    )

    client_items = []

    for (
        field_key,
        code,
        label,
    ) in field_map:
        quantity = fields.get(
            field_key,
        )

        if quantity in {
            None,
            "",
        }:
            continue

        try:
            if (
                float(
                    quantity,
                )
                <= 0
            ):
                continue

        except (
            TypeError,
            ValueError,
        ):
            continue

        client_items.append(
            {
                "key": field_key,
                "code": code,
                "label": label,
                "quantity": quantity,
            }
        )

    return client_items


def _build_billing_automatic_flags(
    billing_session,
) -> dict:
    """
    Determina valores automáticos del proyecto usando
    únicamente los códigos reales aprobados en Invoice.

    C-108-AER
        -> Aerial Case YES

    C-108.1
        -> Re-Entry YES
    """

    billing_snapshot = build_billing_submission_snapshot(
        billing_session,
    )

    codes = _get_snapshot_codes(
        billing_snapshot,
    )

    return {
        "aerial_case": ("C-108-AER" in codes),
        "re_entry": ("C-108.1" in codes),
        "codes": codes,
    }


def _resolve_common_finish_date(
    billings,
):
    """
    Retorna la fecha común si TODOS los Billings seleccionados
    tienen exactamente la misma finance_finish_date.

    Si existen fechas distintas o falta una fecha:

        None
    """

    billings = list(
        billings,
    )

    if not billings:
        return None

    finish_dates = [
        getattr(
            billing,
            "finance_finish_date",
            None,
        )
        for billing in billings
    ]

    if any(finish_date is None for finish_date in finish_dates):
        return None

    unique_dates = set(
        finish_dates,
    )

    if (
        len(
            unique_dates,
        )
        != 1
    ):
        return None

    return finish_dates[0]


# ============================================================
# Helpers de request
# ============================================================


def _parse_ids(
    raw_ids,
) -> list[int]:
    """
    Acepta:

        1,2,3
        ["1", "2", "3"]

    Elimina duplicados manteniendo el orden.
    """

    if raw_ids is None:
        return []

    if isinstance(
        raw_ids,
        (
            list,
            tuple,
            set,
        ),
    ):
        values = list(
            raw_ids,
        )

    else:
        values = str(
            raw_ids,
        ).split(",")

    result = []
    seen = set()

    for value in values:
        text = _clean_text(
            value,
        )

        if not text.isdigit():
            continue

        numeric_id = int(
            text,
        )

        if numeric_id in seen:
            continue

        seen.add(
            numeric_id,
        )

        result.append(
            numeric_id,
        )

    return result


def _parse_bool(
    value,
) -> bool:
    if isinstance(
        value,
        bool,
    ):
        return value

    value = _clean_text(
        value,
    ).lower()

    return value in {
        "1",
        "true",
        "yes",
        "on",
        "y",
    }


def _parse_date(
    value,
):
    value = _clean_text(
        value,
    )

    if not value:
        return None

    try:
        return date.fromisoformat(
            value,
        )

    except ValueError:
        return None


def _parse_additional_emails(
    value,
) -> list[str]:
    if not value:
        return []

    if isinstance(
        value,
        list,
    ):
        raw_values = value

    else:
        raw_values = (
            str(
                value,
            )
            .replace(
                ";",
                ",",
            )
            .split(",")
        )

    result = []
    seen = set()

    for raw_value in raw_values:
        email = _clean_text(
            raw_value,
        ).lower()

        if not email:
            continue

        if email in seen:
            continue

        seen.add(
            email,
        )

        result.append(
            email,
        )

    return result


def _parse_submission_configs(
    request: HttpRequest,
    billing_ids: list[int],
) -> tuple[
    str,
    dict[int, dict],
]:
    """
    Lee la configuración de Work Completed.

    La configuración común aplica únicamente a:

        Fiber Placed
        Splicing
        Testing

    Aerial Case y Re-Entry NO son configurables manualmente.

    Se obtienen automáticamente desde Billing:

        C-108-AER -> Aerial Case YES
        C-108.1   -> Re-Entry YES

    Aerial Sequential IN / OUT siempre se leen individualmente
    por proyecto.
    """

    configuration_mode = _clean_text(
        request.POST.get(
            "configuration_mode",
            "common",
        )
    ).lower()

    if configuration_mode not in {
        "common",
        "individual",
    }:
        configuration_mode = "common"

    # ========================================================
    # Configuración común de Work Types
    # ========================================================

    common_config = {
        "fiber_placed": _parse_bool(
            request.POST.get(
                "fiber_placed",
            )
        ),
        "splicing": _parse_bool(
            request.POST.get(
                "splicing",
            )
        ),
        "testing": _parse_bool(
            request.POST.get(
                "testing",
            )
        ),
    }

    # ========================================================
    # Crear configuración por proyecto
    # ========================================================

    submission_configs: dict[int, dict] = {}

    for billing_id in billing_ids:
        prefix = f"project_{billing_id}"

        if configuration_mode == "common":
            fiber_placed = common_config["fiber_placed"]

            splicing = common_config["splicing"]

            testing = common_config["testing"]

        else:
            fiber_placed = _parse_bool(
                request.POST.get(
                    f"{prefix}_fiber_placed",
                )
            )

            splicing = _parse_bool(
                request.POST.get(
                    f"{prefix}_splicing",
                )
            )

            testing = _parse_bool(
                request.POST.get(
                    f"{prefix}_testing",
                )
            )

        submission_configs[billing_id] = {
            "configuration_mode": configuration_mode,
            "fiber_placed": fiber_placed,
            "splicing": splicing,
            "testing": testing,
            # ================================================
            # Aerial Sequential siempre es individual.
            #
            # El builder decidirá mediante C-108-AER si estos
            # valores realmente aplican.
            # ================================================
            "aerial_case_value_1": _clean_text(
                request.POST.get(
                    f"{prefix}_aerial_case_value_1",
                )
            ),
            "aerial_case_value_2": _clean_text(
                request.POST.get(
                    f"{prefix}_aerial_case_value_2",
                )
            ),
        }

    return (
        configuration_mode,
        submission_configs,
    )


# ============================================================
# Billing selection
# ============================================================


def _get_selected_billings(
    ids: list[int],
):
    """
    Recupera los Billing seleccionados respetando el orden
    exacto de selección.
    """

    if not ids:
        return []

    billings = SesionBilling.objects.filter(
        id__in=ids,
        is_direct_discount=False,
    ).prefetch_related(
        "items",
        "tecnicos_sesion__tecnico",
        "tecnicos_sesion__evidencias__requisito",
    )

    by_id = {billing.pk: billing for billing in billings}

    return [by_id[billing_id] for billing_id in ids if billing_id in by_id]


# ============================================================
# Presentación / serialización
# ============================================================


def _submission_status_counts(
    batch: ClientSubmissionBatch,
) -> dict:
    submissions = batch.submissions.all()

    return {
        "total": submissions.count(),
        "pending_client_submission": submissions.filter(
            status=(ClientSubmission.Status.PENDING_CLIENT_SUBMISSION)
        ).count(),
        "preparing": submissions.filter(
            status=ClientSubmission.Status.PREPARING
        ).count(),
        "awaiting_verification": submissions.filter(
            status=(ClientSubmission.Status.AWAITING_VERIFICATION)
        ).count(),
        "submitting": submissions.filter(
            status=ClientSubmission.Status.SUBMITTING
        ).count(),
        "awaiting_email_confirmation": submissions.filter(
            status=(ClientSubmission.Status.AWAITING_EMAIL_CONFIRMATION)
        ).count(),
        "sent_to_client": submissions.filter(
            status=ClientSubmission.Status.SENT_TO_CLIENT
        ).count(),
        "dry_run_completed": submissions.filter(
            status=ClientSubmission.Status.DRY_RUN_COMPLETED
        ).count(),
        "failed": submissions.filter(status=ClientSubmission.Status.FAILED).count(),
        "cancelled": submissions.filter(
            status=ClientSubmission.Status.CANCELLED
        ).count(),
        "ready": submissions.filter(
            validation_ok=True,
        ).count(),
        "validation_errors": submissions.filter(
            validation_ok=False,
        ).count(),
    }


def _submission_to_dict(
    submission: ClientSubmission,
) -> dict:
    payload = submission.form_payload or {}

    zip_data = payload.get(
        "zip",
        {},
    )

    quantities = payload.get(
        "quantities",
        {},
    )

    return {
        "id": submission.pk,
        "public_id": str(
            submission.public_id,
        ),
        "sequence_number": (submission.sequence_number),
        "billing_session_id": (submission.billing_session_id),
        "project_id": (submission.project_id),
        "dfn_name": (submission.dfn_name),
        "access_point_id": (submission.access_point_id),
        "status": submission.status,
        "status_label": (submission.get_status_display()),
        "validation_ok": (submission.validation_ok),
        "validation_errors": (submission.validation_errors or []),
        "validation_warnings": (submission.validation_warnings or []),
        "zip_available": (submission.zip_available),
        "zip_filename": (submission.zip_filename),
        "evidence_count": int(
            zip_data.get(
                "evidence_count",
                0,
            )
            or 0
        ),
        "quantities": quantities,
        # ====================================================
        # Fecha individual del proyecto
        # ====================================================
        "production_completed_date": (
            payload.get(
                "production_completed_date",
                "",
            )
        ),
        # ====================================================
        # Configuración individual del proyecto
        # ====================================================
        "configuration_mode": (
            payload.get(
                "configuration_mode",
                "common",
            )
        ),
        "fiber_placed": bool(
            payload.get(
                "fiber_placed",
                False,
            )
        ),
        "splicing": bool(
            payload.get(
                "splicing",
                False,
            )
        ),
        "testing": bool(
            payload.get(
                "testing",
                False,
            )
        ),
        "aerial_case": bool(
            payload.get(
                "aerial_case",
                False,
            )
        ),
        "re_entry": bool(
            payload.get(
                "re_entry",
                False,
            )
        ),
        "aerial_case_value_1": (
            payload.get(
                "aerial_case_value_1",
                "",
            )
        ),
        "aerial_case_value_2": (
            payload.get(
                "aerial_case_value_2",
                "",
            )
        ),
        # ====================================================
        # Ejecución
        # ====================================================
        "attempt_count": (submission.attempt_count),
        "browser_confirmation_received": (submission.browser_confirmation_received),
        "email_confirmation_received": (submission.email_confirmation_received),
        "submitted_at": (
            submission.submitted_at.isoformat() if submission.submitted_at else None
        ),
        # ====================================================
        # Error actual
        # ====================================================
        "last_error_code": (submission.last_error_code),
        "last_error_message": (submission.last_error_message),
    }


def _batch_to_dict(
    batch: ClientSubmissionBatch,
) -> dict:
    return {
        "id": batch.pk,
        "public_id": str(
            batch.public_id,
        ),
        "name": batch.name,
        "status": batch.status,
        "status_label": (batch.get_status_display()),
        "execution_mode": (batch.execution_mode),
        "execution_mode_label": (batch.get_execution_mode_display()),
        "submitted_by_email": (batch.submitted_by_email),
        "send_copy_of_responses": (batch.send_copy_of_responses),
        "copy_email": (batch.copy_email),
        "additional_copy_emails": (batch.additional_copy_emails or []),
        "is_subcontractor": (batch.is_subcontractor),
        "subcontractor_name": (batch.subcontractor_name),
        "production_completed_date": (
            batch.production_completed_date.isoformat()
            if batch.production_completed_date
            else ""
        ),
        "market": batch.market,
        "fiber_placed": batch.fiber_placed,
        "splicing": batch.splicing,
        "testing": batch.testing,
        "aerial_case": batch.aerial_case,
        "re_entry": batch.re_entry,
        "created_at": (batch.created_at.isoformat()),
        "started_at": (batch.started_at.isoformat() if batch.started_at else None),
        "finished_at": (batch.finished_at.isoformat() if batch.finished_at else None),
        "counts": _submission_status_counts(
            batch,
        ),
    }


# ============================================================
# Crear Batch desde Invoices
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_GET
def create_batch_from_invoices(
    request: HttpRequest,
) -> HttpResponse:
    """
    Primera pantalla después de seleccionar invoices.

    Production Completed Date utiliza finance_finish_date.

    Si todos los proyectos tienen la misma fecha:
        se muestra una fecha común.

    Si tienen fechas distintas:
        la fecha común queda vacía y cada proyecto muestra
        su propia fecha.

    Aerial Case:
        C-108-AER -> YES

    Re-Entry:
        C-108.1 -> YES
    """

    _assert_manage_permission(
        request,
    )

    ids = _parse_ids(
        request.GET.get(
            "ids",
        )
    )

    if not ids:
        messages.warning(
            request,
            "Select at least one invoice.",
        )

        return redirect(
            "facturacion:invoices",
        )

    billings = _get_selected_billings(
        ids,
    )

    if not billings:
        messages.warning(
            request,
            "No valid invoices were found in the selection.",
        )

        return redirect(
            "facturacion:invoices",
        )

    found_ids = {billing.pk for billing in billings}

    missing_ids = [billing_id for billing_id in ids if billing_id not in found_ids]

    # ========================================================
    # Fecha común del Batch
    # ========================================================

    common_finish_date = _resolve_common_finish_date(
        billings,
    )

    has_mixed_finish_dates = common_finish_date is None

    # ========================================================
    # Proyectos seleccionados
    # ========================================================

    selected_rows = []

    for billing in billings:
        client_items = _build_client_items_preview(
            billing,
        )

        automatic_flags = _build_billing_automatic_flags(
            billing,
        )

        finish_date = getattr(
            billing,
            "finance_finish_date",
            None,
        )

        selected_rows.append(
            {
                "id": billing.pk,
                "project_id": billing.proyecto_id,
                "client": billing.cliente,
                "city": billing.ciudad,
                "project": billing.proyecto,
                "office": billing.oficina,
                "finance_status": billing.finance_status,
                "finance_status_label": (billing.get_finance_status_display()),
                # ============================================
                # Fecha individual real del Invoice
                # ============================================
                "finish_date": finish_date,
                # ============================================
                # Flags automáticos por código
                # ============================================
                "aerial_case": bool(
                    automatic_flags.get(
                        "aerial_case",
                        False,
                    )
                ),
                "re_entry": bool(
                    automatic_flags.get(
                        "re_entry",
                        False,
                    )
                ),
                # ============================================
                # Información del Billing
                # ============================================
                "items_count": billing.items.count(),
                "evidence_count": sum(
                    assignment.evidencias.count()
                    for assignment in billing.tecnicos_sesion.all()
                ),
                # ============================================
                # Cantidades Smartsheet
                # ============================================
                "client_items": client_items,
                "client_items_count": len(
                    client_items,
                ),
            }
        )

    initial_data = {
        "submitted_by_email": (DEFAULT_SUBMITTED_BY_EMAIL),
        "copy_email": (DEFAULT_COPY_EMAIL),
        "subcontractor_name": (DEFAULT_SUBCONTRACTOR_NAME),
        "production_completed_date": (
            common_finish_date.isoformat() if common_finish_date else ""
        ),
        "execution_mode": (ClientSubmissionBatch.ExecutionMode.LIVE),
        "send_copy_of_responses": True,
        "is_subcontractor": True,
        "configuration_mode": "common",
        "fiber_placed": False,
        "splicing": False,
        "testing": False,
    }

    aerial_selected_count = sum(
        1
        for row in selected_rows
        if row.get(
            "aerial_case",
            False,
        )
    )
    context = {
        "selected_ids": ids,
        "selected_ids_csv": ",".join(
            str(
                billing_id,
            )
            for billing_id in ids
        ),
        "selected_rows": selected_rows,
        "selected_count": len(
            selected_rows,
        ),
        "aerial_selected_count": aerial_selected_count,
        "missing_ids": missing_ids,
        "initial_data": initial_data,
        "execution_modes": (ClientSubmissionBatch.ExecutionMode.choices),
        # ====================================================
        # Información para UI de fechas
        # ====================================================
        "common_finish_date": (common_finish_date),
        "has_mixed_finish_dates": (has_mixed_finish_dates),
    }

    return render(
        request,
        "client_submissions/create_batch.html",
        context,
    )

# ============================================================
# Preview de Aerial Sequentials desde Excel
# ============================================================


def _normalize_excel_header(value) -> str:
    """
    Normaliza encabezados de la planilla para poder reconocerlos
    aunque tengan diferencias menores de espacios/mayúsculas.
    """

    return _clean_text(value).lower().replace("\n", " ").replace("\r", " ")


def _excel_value_to_text(value) -> str:
    """
    Convierte valores Excel a texto sin introducir '.0'
    en valores numéricos enteros.

    Ej:
        63002.0 -> "63002"
        63002   -> "63002"
    """

    if value is None:
        return ""

    if isinstance(value, bool):
        return "1" if value else "0"

    if isinstance(value, int):
        return str(value)

    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))

        return str(value).strip()

    return str(value).strip()


def _normalize_project_match_value(value) -> str:
    """
    Normalización SOLO para matching.

    No altera el valor original que después se muestra al usuario.
    """

    return _clean_text(value).upper().replace(" ", "")


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
def preview_aerial_sequentials(
    request: HttpRequest,
) -> JsonResponse:
    """
    Lee una planilla Excel de Aerial Sequentials y devuelve
    únicamente coincidencias contra los invoices seleccionados.

    Importante:
    - NO crea el Batch.
    - NO modifica SesionBilling.
    - NO guarda el Excel.
    - NO modifica ClientSubmission.
    - Solo devuelve un preview JSON.

    La planilla esperada contiene:

        A1:
            prefijo del proyecto
            Ej: 0913BA_06

        Columnas:
            HH Number
            Seq. #1(In)
            Seq. #1(Out)

    Project ID resultante:

        <A1>_<HH Number>

    Ej:
        0913BA_06 + 7000-027
        =
        0913BA_06_7000-027

    Si un mismo HH Number aparece varias veces:
    - NO se rechaza el Excel completo.
    - Si solamente existe una combinación IN/OUT válida, se usa.
    - Si existen varias combinaciones válidas, se devuelve como
      "ambiguous" para que el usuario seleccione cuál aplicar.
    """

    _assert_manage_permission(
        request,
    )

    # ========================================================
    # Invoices seleccionados
    # ========================================================

    ids = _parse_ids(
        request.POST.get(
            "ids",
        )
    )

    if not ids:
        return JsonResponse(
            {
                "ok": False,
                "error": "No invoices were selected.",
            },
            status=400,
        )

    billings = _get_selected_billings(
        ids,
    )

    if not billings:
        return JsonResponse(
            {
                "ok": False,
                "error": "No valid invoices were found.",
            },
            status=400,
        )

    # ========================================================
    # Archivo
    # ========================================================

    uploaded_file = request.FILES.get(
        "file",
    )

    if uploaded_file is None:
        return JsonResponse(
            {
                "ok": False,
                "error": "Select an Excel file first.",
            },
            status=400,
        )

    filename = _clean_text(
        getattr(
            uploaded_file,
            "name",
            "",
        )
    )

    if not filename.lower().endswith(
        ".xlsx",
    ):
        return JsonResponse(
            {
                "ok": False,
                "error": "The sequential file must be an .xlsx Excel file.",
            },
            status=400,
        )

    max_file_size = 10 * 1024 * 1024

    if uploaded_file.size > max_file_size:
        return JsonResponse(
            {
                "ok": False,
                "error": "The Excel file is too large. Maximum size is 10 MB.",
            },
            status=400,
        )

    # ========================================================
    # Solo proyectos realmente Aerial según Billing
    # ========================================================

    selected_aerial_by_project: dict[str, dict] = {}

    for billing in billings:
        automatic_flags = _build_billing_automatic_flags(
            billing,
        )

        if not bool(
            automatic_flags.get(
                "aerial_case",
                False,
            )
        ):
            continue

        project_id = _clean_text(
            billing.proyecto_id,
        )

        normalized_project_id = _normalize_project_match_value(
            project_id,
        )

        if not normalized_project_id:
            continue

        selected_aerial_by_project[normalized_project_id] = {
            "billing_id": billing.pk,
            "project_id": project_id,
        }

    if not selected_aerial_by_project:
        return JsonResponse(
            {
                "ok": False,
                "error": ("None of the selected invoices are Aerial Case projects."),
            },
            status=400,
        )

    # ========================================================
    # Leer workbook
    # ========================================================

    try:
        uploaded_file.seek(
            0,
        )

        workbook = load_workbook(
            uploaded_file,
            read_only=True,
            data_only=True,
        )

    except (
        InvalidFileException,
        BadZipFile,
        OSError,
        ValueError,
    ):
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "The Excel file could not be read. "
                    "Please verify that it is a valid .xlsx file."
                ),
            },
            status=400,
        )

    try:
        # ====================================================
        # Preferimos Client Reports.
        #
        # Si no existe, permitimos una única hoja.
        # ====================================================

        if "Client Reports" in workbook.sheetnames:
            worksheet = workbook["Client Reports"]

        elif (
            len(
                workbook.sheetnames,
            )
            == 1
        ):
            worksheet = workbook[workbook.sheetnames[0]]

        else:
            return JsonResponse(
                {
                    "ok": False,
                    "error": ('The workbook must contain a "Client Reports" sheet.'),
                },
                status=400,
            )

        # ====================================================
        # Prefijo del proyecto
        # ====================================================

        workbook_prefix = _clean_text(
            worksheet.cell(
                row=1,
                column=1,
            ).value
        )

        normalized_prefix = _normalize_project_match_value(
            workbook_prefix,
        )

        if not normalized_prefix:
            return JsonResponse(
                {
                    "ok": False,
                    "error": ("The project prefix was not found in cell A1."),
                },
                status=400,
            )

        # ====================================================
        # Buscar encabezados
        # ====================================================

        header_row = None

        hh_column = None
        sequential_in_column = None
        sequential_out_column = None
        notes_column = None

        for row_number in range(
            1,
            min(
                worksheet.max_row,
                20,
            )
            + 1,
        ):
            headers = {}

            for column_number in range(
                1,
                worksheet.max_column + 1,
            ):
                header = _normalize_excel_header(
                    worksheet.cell(
                        row=row_number,
                        column=column_number,
                    ).value
                )

                if header:
                    headers[header] = column_number

            # -----------------------------------------------
            # HH Number
            # -----------------------------------------------

            for candidate in (
                "hh number",
                "hh #",
                "hh",
            ):
                if candidate in headers:
                    hh_column = headers[candidate]
                    break

            # -----------------------------------------------
            # Sequential IN
            # -----------------------------------------------

            for candidate in (
                "seq. #1(in)",
                "seq #1(in)",
                "seq. #1 (in)",
                "sequential in",
                "seq in",
            ):
                if candidate in headers:
                    sequential_in_column = headers[candidate]
                    break

            # -----------------------------------------------
            # Sequential OUT
            # -----------------------------------------------

            for candidate in (
                "seq. #1(out)",
                "seq #1(out)",
                "seq. #1 (out)",
                "sequential out",
                "seq out",
            ):
                if candidate in headers:
                    sequential_out_column = headers[candidate]
                    break

            # -----------------------------------------------
            # Notes opcional
            # -----------------------------------------------

            for candidate in (
                "notes",
                "note",
                "comments",
                "comment",
            ):
                if candidate in headers:
                    notes_column = headers[candidate]
                    break

            if hh_column and sequential_in_column and sequential_out_column:
                header_row = row_number
                break

            hh_column = None
            sequential_in_column = None
            sequential_out_column = None
            notes_column = None

        if header_row is None:
            return JsonResponse(
                {
                    "ok": False,
                    "error": (
                        "Required columns were not found. "
                        'Expected "HH Number", "Seq. #1(In)" '
                        'and "Seq. #1(Out)".'
                    ),
                },
                status=400,
            )

        # ====================================================
        # Leer filas del Excel
        #
        # IMPORTANTE:
        # Un mismo Project ID puede aparecer varias veces.
        # Por eso almacenamos una LISTA de registros por Project.
        # ====================================================

        excel_records: dict[str, list[dict]] = {}

        valid_excel_rows = 0

        for row_number in range(
            header_row + 1,
            worksheet.max_row + 1,
        ):
            hh_number = _excel_value_to_text(
                worksheet.cell(
                    row=row_number,
                    column=hh_column,
                ).value
            )

            sequential_in = _excel_value_to_text(
                worksheet.cell(
                    row=row_number,
                    column=sequential_in_column,
                ).value
            )

            sequential_out = _excel_value_to_text(
                worksheet.cell(
                    row=row_number,
                    column=sequential_out_column,
                ).value
            )

            notes = ""

            if notes_column:
                notes = _excel_value_to_text(
                    worksheet.cell(
                        row=row_number,
                        column=notes_column,
                    ).value
                )

            # Fila completamente vacía.
            if not (hh_number or sequential_in or sequential_out or notes):
                continue

            # Sin HH Number no podemos hacer matching.
            if not hh_number:
                continue

            valid_excel_rows += 1

            full_project_id = f"{workbook_prefix}_{hh_number}"

            normalized_full_project_id = _normalize_project_match_value(
                full_project_id,
            )

            record = {
                "row_number": row_number,
                "hh_number": hh_number,
                "project_id": full_project_id,
                "sequential_in": sequential_in,
                "sequential_out": sequential_out,
                "notes": notes,
            }

            excel_records.setdefault(
                normalized_full_project_id,
                [],
            ).append(record)

        # ====================================================
        # Match SOLO contra Aerial seleccionados
        # ====================================================

        matches = []

        ambiguous = []

        missing = []

        for (
            normalized_project_id,
            selected_project,
        ) in selected_aerial_by_project.items():

            excel_options = excel_records.get(
                normalized_project_id,
                [],
            )

            # ------------------------------------------------
            # No existe en Excel
            # ------------------------------------------------

            if not excel_options:
                missing.append(
                    {
                        "billing_id": selected_project["billing_id"],
                        "project_id": selected_project["project_id"],
                        "reason": "Project not found in Excel.",
                    }
                )

                continue

            # ------------------------------------------------
            # Solo opciones con IN y OUT completos
            # ------------------------------------------------

            valid_options = [
                option
                for option in excel_options
                if option["sequential_in"] and option["sequential_out"]
            ]

            if not valid_options:
                missing.append(
                    {
                        "billing_id": selected_project["billing_id"],
                        "project_id": selected_project["project_id"],
                        "reason": ("Sequential IN or OUT is missing in Excel."),
                    }
                )

                continue

            # ------------------------------------------------
            # Eliminar duplicados EXACTOS.
            #
            # Si el mismo HH aparece dos veces con exactamente
            # el mismo IN/OUT, no tiene sentido pedir elección.
            # ------------------------------------------------

            unique_options = []

            seen_pairs = set()

            for option in valid_options:
                pair_key = (
                    option["sequential_in"],
                    option["sequential_out"],
                )

                if pair_key in seen_pairs:
                    continue

                seen_pairs.add(
                    pair_key,
                )

                unique_options.append(option)

            # ------------------------------------------------
            # Una sola combinación válida
            # ------------------------------------------------

            if (
                len(
                    unique_options,
                )
                == 1
            ):
                excel_record = unique_options[0]

                matches.append(
                    {
                        "billing_id": selected_project["billing_id"],
                        "project_id": selected_project["project_id"],
                        "hh_number": excel_record["hh_number"],
                        "sequential_in": excel_record["sequential_in"],
                        "sequential_out": excel_record["sequential_out"],
                        "excel_row": excel_record["row_number"],
                        "notes": excel_record["notes"],
                    }
                )

                continue

            # ------------------------------------------------
            # Más de una combinación válida:
            # debe decidir el usuario.
            # ------------------------------------------------

            ambiguous.append(
                {
                    "billing_id": selected_project["billing_id"],
                    "project_id": selected_project["project_id"],
                    "hh_number": unique_options[0]["hh_number"],
                    "options": [
                        {
                            "sequential_in": option["sequential_in"],
                            "sequential_out": option["sequential_out"],
                            "excel_row": option["row_number"],
                            "notes": option["notes"],
                        }
                        for option in unique_options
                    ],
                }
            )

        # ====================================================
        # Excel rows ignoradas
        #
        # Son filas que pertenecen a proyectos que NO están
        # dentro de los Aerial seleccionados en este Batch.
        #
        # Si un proyecto seleccionado aparece varias veces,
        # esas filas NO se consideran "ignored" porque forman
        # parte del resultado/elección del usuario.
        # ====================================================

        ignored_excel_rows = sum(
            len(records)
            for (
                excel_key,
                records,
            ) in excel_records.items()
            if excel_key not in selected_aerial_by_project
        )

        # ====================================================
        # Respuesta
        # ====================================================

        return JsonResponse(
            {
                "ok": True,
                "filename": filename,
                "sheet_name": worksheet.title,
                "workbook_prefix": workbook_prefix,
                "counts": {
                    "selected_aerial": len(
                        selected_aerial_by_project,
                    ),
                    "matched": len(
                        matches,
                    ),
                    "ambiguous": len(
                        ambiguous,
                    ),
                    "missing": len(
                        missing,
                    ),
                    "valid_excel_rows": valid_excel_rows,
                    "ignored_excel_rows": ignored_excel_rows,
                },
                "matches": matches,
                "ambiguous": ambiguous,
                "missing": missing,
            }
        )

    finally:
        workbook.close()


# ============================================================
# Confirmar creación
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
def create_batch_submit(
    request: HttpRequest,
) -> HttpResponse:
    """
    Crea:

        ClientSubmissionBatch
        ClientSubmission x N

    Production Completed Date no utiliza el valor enviado
    manualmente por el formulario.

    create_submission_batch vuelve a leer:

        SesionBilling.finance_finish_date

    para cada proyecto.

    Aerial Case y Re-Entry también serán resueltos desde
    los códigos reales del Billing.

    El modo de ejecución predeterminado es LIVE. Dry Run
    continúa disponible cuando el usuario lo selecciona
    explícitamente en el formulario.
    """

    _assert_manage_permission(
        request,
    )

    ids = _parse_ids(
        request.POST.get(
            "ids",
        )
    )

    if not ids:
        messages.error(
            request,
            "No invoices were selected.",
        )

        return redirect(
            "facturacion:invoices",
        )

    billings = _get_selected_billings(
        ids,
    )

    if not billings:
        messages.error(
            request,
            "No valid invoices were found.",
        )

        return redirect(
            "facturacion:invoices",
        )

    # ========================================================
    # IDs reales que sí serán procesados
    # ========================================================

    billing_ids = [billing.pk for billing in billings]

    # ========================================================
    # Modo de ejecución
    #
    # LIVE es el modo predeterminado.
    # DRY_RUN se conserva cuando el usuario lo selecciona
    # explícitamente desde el formulario.
    # ========================================================

    execution_mode = _clean_text(
        request.POST.get(
            "execution_mode",
            ClientSubmissionBatch.ExecutionMode.LIVE,
        )
    )

    valid_execution_modes = {
        value for value, _label in ClientSubmissionBatch.ExecutionMode.choices
    }

    if execution_mode not in valid_execution_modes:
        execution_mode = ClientSubmissionBatch.ExecutionMode.LIVE

    # ========================================================
    # Configuración común / individual
    # ========================================================

    (
        configuration_mode,
        submission_configs,
    ) = _parse_submission_configs(
        request,
        billing_ids,
    )

    # ========================================================
    # Configuración representativa del Batch
    #
    # Solo Work Types.
    #
    # Aerial Case y Re-Entry son automáticos por proyecto.
    # ========================================================

    all_configs = list(
        submission_configs.values(),
    )

    batch_fiber_placed = any(
        bool(
            config.get(
                "fiber_placed",
                False,
            )
        )
        for config in all_configs
    )

    batch_splicing = any(
        bool(
            config.get(
                "splicing",
                False,
            )
        )
        for config in all_configs
    )

    batch_testing = any(
        bool(
            config.get(
                "testing",
                False,
            )
        )
        for config in all_configs
    )

    # ========================================================
    # Crear Batch
    # ========================================================

    try:
        result = create_submission_batch(
            created_by=request.user,
            billing_sessions=billings,
            name=_clean_text(
                request.POST.get(
                    "name",
                )
            ),
            execution_mode=execution_mode,
            submitted_by_email=_clean_text(
                request.POST.get(
                    "submitted_by_email",
                )
            ),
            send_copy_of_responses=_parse_bool(
                request.POST.get(
                    "send_copy_of_responses",
                )
            ),
            copy_email=_clean_text(
                request.POST.get(
                    "copy_email",
                )
            ),
            additional_copy_emails=(
                _parse_additional_emails(
                    request.POST.get(
                        "additional_copy_emails",
                    )
                )
            ),
            is_subcontractor=_parse_bool(
                request.POST.get(
                    "is_subcontractor",
                )
            ),
            subcontractor_name=_clean_text(
                request.POST.get(
                    "subcontractor_name",
                )
            ),
            # ================================================
            # La fecha real será resuelta desde los Billings.
            # ================================================
            production_completed_date=None,
            market=_clean_text(
                request.POST.get(
                    "market",
                )
            ),
            fiber_placed=batch_fiber_placed,
            splicing=batch_splicing,
            testing=batch_testing,
            # ================================================
            # Compatibilidad del modelo Batch.
            #
            # No se utilizan para decidir el proyecto.
            # ================================================
            aerial_case=False,
            re_entry=False,
            submission_configs=submission_configs,
            notes=_clean_text(
                request.POST.get(
                    "notes",
                )
            ),
        )

    except (
        EmptyBillingSelectionError,
        InvalidBatchConfigurationError,
    ) as exc:
        messages.error(
            request,
            str(
                exc,
            ),
        )

        back_url = (
            reverse(
                "client_submissions:create_batch_from_invoices",
            )
            + "?ids="
            + ",".join(
                str(
                    billing_id,
                )
                for billing_id in billing_ids
            )
        )

        return redirect(
            back_url,
        )

    except Exception as exc:
        messages.error(
            request,
            ("The client submission batch " "could not be created. " f"{exc}"),
        )

        back_url = (
            reverse(
                "client_submissions:create_batch_from_invoices",
            )
            + "?ids="
            + ",".join(
                str(
                    billing_id,
                )
                for billing_id in billing_ids
            )
        )

        return redirect(
            back_url,
        )

    batch = get_object_or_404(
        ClientSubmissionBatch,
        pk=result.batch_id,
    )

    # ========================================================
    # Resultado
    # ========================================================

    if result.total_with_errors:
        messages.warning(
            request,
            (
                f"Batch created with "
                f"{result.total_ready} project(s) ready "
                f"and {result.total_with_errors} "
                "project(s) requiring review."
            ),
        )

    else:
        execution_label = (
            "Live submission"
            if execution_mode == ClientSubmissionBatch.ExecutionMode.LIVE
            else "Dry Run"
        )

        messages.success(
            request,
            (
                "Batch created successfully. "
                f"{result.total_ready} project(s) are ready. "
                f"Execution mode: {execution_label}."
            ),
        )

    return redirect(
        "client_submissions:batch_detail",
        public_id=batch.public_id,
    )


# ============================================================
# Lista de Batches
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_GET
def batch_list(
    request: HttpRequest,
) -> HttpResponse:
    _assert_manage_permission(
        request,
    )

    batches = (
        ClientSubmissionBatch.objects.select_related(
            "created_by",
            "current_submission",
        )
        .prefetch_related(
            "submissions",
        )
        .order_by(
            "-created_at",
            "-id",
        )
    )

    status_filter = _clean_text(
        request.GET.get(
            "status",
        )
    )

    if status_filter:
        batches = batches.filter(
            status=status_filter,
        )

    cantidad = _clean_text(
        request.GET.get(
            "cantidad",
            "10",
        )
    )

    if cantidad not in {
        "5",
        "10",
        "20",
        "50",
        "100",
    }:
        cantidad = "10"

    paginator = Paginator(
        batches,
        int(
            cantidad,
        ),
    )

    pagina = paginator.get_page(
        request.GET.get(
            "page",
        )
    )

    query_params = request.GET.copy()

    query_params.pop(
        "page",
        None,
    )

    qs_keep = query_params.urlencode()

    context = {
        "batches": pagina.object_list,
        "pagina": pagina,
        "cantidad": cantidad,
        "qs_keep": qs_keep,
        "status_filter": status_filter,
        "batch_status_choices": (ClientSubmissionBatch.Status.choices),
    }

    return render(
        request,
        "client_submissions/batch_list.html",
        context,
    )


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
@transaction.atomic
def batch_delete(
    request: HttpRequest,
    public_id,
) -> JsonResponse:
    _assert_manage_permission(
        request,
    )

    batch = get_object_or_404(
        ClientSubmissionBatch.objects.select_for_update(),
        public_id=public_id,
    )

    if batch.status == ClientSubmissionBatch.Status.RUNNING:
        return JsonResponse(
            {
                "ok": False,
                "error": ("A batch cannot be deleted while " "it is being processed."),
            },
            status=409,
        )

    batch_id = batch.pk

    batch.delete()

    return JsonResponse(
        {
            "ok": True,
            "batch_id": batch_id,
        }
    )


# ============================================================
# Detalle del Batch
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_GET
def batch_detail(
    request: HttpRequest,
    public_id,
) -> HttpResponse:
    _assert_manage_permission(
        request,
    )

    batch = get_object_or_404(
        ClientSubmissionBatch.objects.select_related(
            "created_by",
            "current_submission",
        ),
        public_id=public_id,
    )

    submissions = batch.submissions.select_related(
        "billing_session",
    ).order_by(
        "sequence_number",
        "id",
    )

    counts = _submission_status_counts(
        batch,
    )

    context = {
        "batch": batch,
        "submissions": submissions,
        "counts": counts,
        "can_start": (
            batch.status
            in {
                ClientSubmissionBatch.Status.DRAFT,
                ClientSubmissionBatch.Status.PENDING,
                ClientSubmissionBatch.Status.PAUSED,
            }
        ),
        "can_pause": (batch.status == ClientSubmissionBatch.Status.RUNNING),
        "can_cancel": batch.can_cancel,
        "has_validation_errors": (counts["validation_errors"] > 0),
        "is_dry_run": batch.is_dry_run,
        "is_live": batch.is_live,
    }

    return render(
        request,
        "client_submissions/batch_detail.html",
        context,
    )


# ============================================================
# Revalidar todo el Batch
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
def batch_revalidate(
    request: HttpRequest,
    public_id,
) -> HttpResponse:
    _assert_manage_permission(
        request,
    )

    batch = get_object_or_404(
        ClientSubmissionBatch,
        public_id=public_id,
    )

    if batch.status in {
        ClientSubmissionBatch.Status.RUNNING,
        ClientSubmissionBatch.Status.AWAITING_VERIFICATION,
    }:
        messages.error(
            request,
            ("The batch cannot be revalidated " "while it is being processed."),
        )

        return redirect(
            "client_submissions:batch_detail",
            public_id=batch.public_id,
        )

    results = revalidate_batch(
        batch,
    )

    ready_count = sum(1 for result in results if result.validation_ok)

    error_count = (
        len(
            results,
        )
        - ready_count
    )

    if error_count:
        messages.warning(
            request,
            (
                "Revalidation completed. "
                f"{ready_count} project(s) ready and "
                f"{error_count} project(s) still have errors."
            ),
        )

    else:
        messages.success(
            request,
            ("Revalidation completed. " f"All {ready_count} project(s) are ready."),
        )

    return redirect(
        "client_submissions:batch_detail",
        public_id=batch.public_id,
    )


# ============================================================
# Revalidar un proyecto
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
def submission_revalidate(
    request: HttpRequest,
    public_id,
) -> HttpResponse:
    _assert_manage_permission(
        request,
    )

    submission = get_object_or_404(
        ClientSubmission.objects.select_related(
            "batch",
            "billing_session",
        ),
        public_id=public_id,
    )

    if submission.batch.status in {
        ClientSubmissionBatch.Status.RUNNING,
        ClientSubmissionBatch.Status.AWAITING_VERIFICATION,
    }:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project cannot be revalidated while "
                    "the batch is being processed."
                ),
            },
            status=409,
        )

    result = revalidate_submission(
        submission,
    )

    return JsonResponse(
        {
            "ok": True,
            "submission": (
                _submission_to_dict(
                    submission,
                )
            ),
            "validation": result.as_dict(),
        }
    )

# ============================================================
# Reiniciar flujo de un proyecto
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
@transaction.atomic
def submission_restart_flow(
    request: HttpRequest,
    public_id,
) -> JsonResponse:
    """
    Reinicia de forma segura una submission que quedó detenida
    durante el flujo de Client Submission.

    Esta acción:

    - Nunca permite reiniciar SENT_TO_CLIENT.
    - Nunca toca otras submissions del Batch.
    - Impide reiniciar si existe otro proyecto actualmente
      en un estado operativo.
    - Usa reset_for_retry() para limpiar correctamente
      timestamps, errores y estado del navegador.
    - Devuelve el Batch a PENDING para que el worker
      pueda recoger nuevamente el proyecto.
    """

    _assert_manage_permission(
        request,
    )

    submission = get_object_or_404(
        ClientSubmission.objects.select_for_update().select_related(
            "batch",
            "billing_session",
        ),
        public_id=public_id,
    )

    batch = ClientSubmissionBatch.objects.select_for_update().get(
        pk=submission.batch_id,
    )

    # --------------------------------------------------------
    # Protección absoluta:
    # un proyecto enviado al cliente nunca puede reiniciarse.
    # --------------------------------------------------------

    if submission.status == ClientSubmission.Status.SENT_TO_CLIENT:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project has already been sent to the client "
                    "and cannot be restarted."
                ),
            },
            status=409,
        )

    # --------------------------------------------------------
    # Solo permitimos recuperación de estados donde realmente
    # puede existir un flujo detenido.
    # --------------------------------------------------------

    restartable_statuses = {
        ClientSubmission.Status.AWAITING_VERIFICATION,
        ClientSubmission.Status.FAILED,
        ClientSubmission.Status.CANCELLED,
    }

    if submission.status not in restartable_statuses:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project cannot be restarted from its "
                    f"current status: {submission.get_status_display()}."
                ),
            },
            status=409,
        )

    # --------------------------------------------------------
    # Verificar que no exista OTRO proyecto del mismo Batch
    # ejecutándose actualmente.
    #
    # El proyecto que estamos reiniciando se excluye de esta
    # comprobación porque precisamente queremos recuperar
    # ese flujo.
    # --------------------------------------------------------

    active_submission_statuses = {
        ClientSubmission.Status.PREPARING_CLIENT_SUBMISSION,
        ClientSubmission.Status.AWAITING_VERIFICATION,
        ClientSubmission.Status.SUBMITTING_TO_CLIENT,
        ClientSubmission.Status.AWAITING_EMAIL_CONFIRMATION,
    }

    other_active_submission = (
        ClientSubmission.objects.select_for_update()
        .filter(
            batch=batch,
            status__in=active_submission_statuses,
        )
        .exclude(
            pk=submission.pk,
        )
        .order_by(
            "sequence_number",
            "id",
        )
        .first()
    )

    if other_active_submission is not None:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "Another project in this batch is currently being "
                    "processed. Restart flow is not available until "
                    "that project finishes."
                ),
            },
            status=409,
        )

    previous_submission_status = submission.status
    previous_batch_status = batch.status

    # --------------------------------------------------------
    # reset_for_retry() es la lógica oficial del modelo.
    #
    # Devuelve la submission a:
    #
    #   pending_client_submission
    #
    # y limpia:
    # - timestamps de ejecución
    # - verificación
    # - browser session
    # - errores
    #
    # No modifica attempt_count.
    # --------------------------------------------------------

    submission.reset_for_retry()

    # --------------------------------------------------------
    # Reiniciar estado operativo del Batch.
    #
    # El worker solamente debe encontrar el Batch nuevamente
    # como PENDING.
    # --------------------------------------------------------

    now = timezone.now()

    batch.status = ClientSubmissionBatch.Status.PENDING
    batch.current_submission = None
    batch.paused_at = None
    batch.finished_at = None
    batch.last_activity_at = now

    batch.save(
        update_fields=[
            "status",
            "current_submission",
            "paused_at",
            "finished_at",
            "last_activity_at",
            "updated_at",
        ]
    )

    return JsonResponse(
        {
            "ok": True,
            "message": (f"Project {submission.project_id} was restarted successfully."),
            "previous_submission_status": previous_submission_status,
            "submission_status": submission.status,
            "previous_batch_status": previous_batch_status,
            "batch_status": batch.status,
            "submission": _submission_to_dict(
                submission,
            ),
        }
    )


# ============================================================
# Preparar Batch para procesamiento
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
@transaction.atomic
def batch_start(
    request: HttpRequest,
    public_id,
) -> HttpResponse:
    """
    Valida el Batch y lo deja pendiente para procesamiento.

    IMPORTANTE:

    La Web NO ejecuta Playwright.

    La Web únicamente cambia el Batch a:

        PENDING

    El Hyperlink background worker detectará el Batch
    y ejecutará:

        client_submissions.automation.worker.run_once()

    El worker procesa únicamente proyectos con estado:

        PENDING_CLIENT_SUBMISSION

    y con:

        validation_ok=True
    """

    _assert_manage_permission(
        request,
    )

    batch = get_object_or_404(
        ClientSubmissionBatch.objects.select_for_update(),
        public_id=public_id,
    )

    # ========================================================
    # Estado permitido
    # ========================================================

    if batch.status not in {
        ClientSubmissionBatch.Status.DRAFT,
        ClientSubmissionBatch.Status.PAUSED,
        ClientSubmissionBatch.Status.PENDING,
    }:
        messages.error(
            request,
            (
                "This batch cannot be started from its current status: "
                f"{batch.get_status_display()}."
            ),
        )

        return redirect(
            "client_submissions:batch_detail",
            public_id=batch.public_id,
        )

    # ========================================================
    # Validaciones
    # ========================================================

    invalid_count = batch.submissions.filter(
        validation_ok=False,
    ).count()

    if invalid_count:
        messages.error(
            request,
            (
                f"The batch cannot start because "
                f"{invalid_count} project(s) have validation errors."
            ),
        )

        return redirect(
            "client_submissions:batch_detail",
            public_id=batch.public_id,
        )

    # ========================================================
    # Proyectos que realmente puede procesar el worker
    # ========================================================

    processable_count = batch.submissions.filter(
        status=ClientSubmission.Status.PENDING_CLIENT_SUBMISSION,
        validation_ok=True,
    ).count()

    if processable_count == 0:
        messages.warning(
            request,
            (
                "There are no pending and validated projects "
                "left to process in this batch."
            ),
        )

        return redirect(
            "client_submissions:batch_detail",
            public_id=batch.public_id,
        )

    # ========================================================
    # Dejar Batch en la cola
    # ========================================================

    batch.status = ClientSubmissionBatch.Status.PENDING

    batch.paused_at = None

    batch.cancelled_at = None

    batch.finished_at = None

    batch.worker_identifier = ""

    batch.current_submission = None

    batch.last_error = ""

    batch.last_activity_at = timezone.now()

    batch.save(
        update_fields=[
            "status",
            "paused_at",
            "cancelled_at",
            "finished_at",
            "worker_identifier",
            "current_submission",
            "last_error",
            "last_activity_at",
            "updated_at",
        ]
    )

    # ========================================================
    # Evento
    # ========================================================

    ClientSubmissionEvent.objects.create(
        batch=batch,
        level=ClientSubmissionEvent.Level.INFO,
        event_type="batch_queued_for_processing",
        message=(
            f"Batch #{batch.pk} was queued for "
            "background processing by "
            f"{request.user.get_username()}."
        ),
        metadata={
            "execution_mode": batch.execution_mode,
            "processable_count": processable_count,
            "user_id": request.user.pk,
        },
    )

    # ========================================================
    # Resultado
    # ========================================================

    messages.success(
        request,
        (
            "Batch queued successfully. "
            "The background worker will start processing it."
        ),
    )

    return redirect(
        "client_submissions:batch_detail",
        public_id=batch.public_id,
    )


# ============================================================
# Pausar Batch
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
@transaction.atomic
def batch_pause(
    request: HttpRequest,
    public_id,
) -> HttpResponse:
    _assert_manage_permission(
        request,
    )

    batch = get_object_or_404(
        ClientSubmissionBatch.objects.select_for_update(),
        public_id=public_id,
    )

    if batch.status not in {
        ClientSubmissionBatch.Status.RUNNING,
        ClientSubmissionBatch.Status.PENDING,
    }:
        messages.warning(
            request,
            "This batch is not currently processing.",
        )

        return redirect(
            "client_submissions:batch_detail",
            public_id=batch.public_id,
        )

    batch.mark_paused()

    ClientSubmissionEvent.objects.create(
        batch=batch,
        level=ClientSubmissionEvent.Level.WARNING,
        event_type="batch_paused",
        message=(f"Batch #{batch.pk} was paused by " f"{request.user.get_username()}."),
        metadata={
            "user_id": request.user.pk,
        },
    )

    messages.success(
        request,
        "The batch was paused.",
    )

    return redirect(
        "client_submissions:batch_detail",
        public_id=batch.public_id,
    )


# ============================================================
# Cancelar Batch
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
@transaction.atomic
def batch_cancel(
    request: HttpRequest,
    public_id,
) -> HttpResponse:
    _assert_manage_permission(
        request,
    )

    batch = get_object_or_404(
        ClientSubmissionBatch.objects.select_for_update(),
        public_id=public_id,
    )

    if not batch.can_cancel:
        messages.warning(
            request,
            "This batch can no longer be cancelled.",
        )

        return redirect(
            "client_submissions:batch_detail",
            public_id=batch.public_id,
        )

    batch.mark_cancelled()

    batch.submissions.filter(
        status__in=[
            ClientSubmission.Status.PENDING_CLIENT_SUBMISSION,
            ClientSubmission.Status.PREPARING,
        ]
    ).update(
        status=ClientSubmission.Status.CANCELLED,
        finished_at=timezone.now(),
        updated_at=timezone.now(),
    )

    ClientSubmissionEvent.objects.create(
        batch=batch,
        level=ClientSubmissionEvent.Level.WARNING,
        event_type="batch_cancelled",
        message=(
            f"Batch #{batch.pk} was cancelled by " f"{request.user.get_username()}."
        ),
        metadata={
            "user_id": request.user.pk,
        },
    )

    messages.success(
        request,
        "The batch was cancelled.",
    )

    return redirect(
        "client_submissions:batch_detail",
        public_id=batch.public_id,
    )


# ============================================================
# Estado JSON para actualización en vivo
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_GET
def batch_status_json(
    request: HttpRequest,
    public_id,
) -> JsonResponse:
    _assert_manage_permission(
        request,
    )

    batch = get_object_or_404(
        ClientSubmissionBatch.objects.select_related(
            "current_submission",
        ),
        public_id=public_id,
    )

    submissions = batch.submissions.select_related(
        "billing_session",
    ).order_by(
        "sequence_number",
        "id",
    )

    return JsonResponse(
        {
            "ok": True,
            "batch": _batch_to_dict(
                batch,
            ),
            "current_submission": (
                _submission_to_dict(
                    batch.current_submission,
                )
                if batch.current_submission
                else None
            ),
            "submissions": [
                _submission_to_dict(
                    submission,
                )
                for submission in submissions
            ],
        }
    )


# ============================================================
# Estado JSON del listado de Batches
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_GET
def batch_list_status_json(
    request: HttpRequest,
) -> JsonResponse:
    _assert_manage_permission(
        request,
    )

    batches = ClientSubmissionBatch.objects.prefetch_related(
        "submissions",
    ).order_by(
        "-created_at",
    )[:50]

    rows = []

    for batch in batches:
        submissions = list(
            batch.submissions.all(),
        )

        total = len(
            submissions,
        )

        completed_statuses = {
            ClientSubmission.Status.SENT_TO_CLIENT,
            ClientSubmission.Status.DRY_RUN_COMPLETED,
            ClientSubmission.Status.FAILED,
            ClientSubmission.Status.CANCELLED,
        }

        completed = sum(
            1 for submission in submissions if submission.status in completed_statuses
        )

        active_submission = next(
            (
                submission
                for submission in submissions
                if submission.status
                in {
                    ClientSubmission.Status.PREPARING,
                    ClientSubmission.Status.AWAITING_VERIFICATION,
                    ClientSubmission.Status.SUBMITTING,
                    ClientSubmission.Status.AWAITING_EMAIL_CONFIRMATION,
                }
            ),
            None,
        )

        progress_percent = round((completed / total) * 100) if total else 0

        rows.append(
            {
                "id": batch.id,
                "public_id": str(
                    batch.public_id,
                ),
                "status": batch.status,
                "status_label": batch.get_status_display(),
                "total": total,
                "completed": completed,
                "progress_percent": progress_percent,
                "active_submission": (
                    {
                        "project_id": active_submission.project_id,
                        "sequence_number": (active_submission.sequence_number),
                        "status": active_submission.status,
                        "status_label": (active_submission.get_status_display()),
                    }
                    if active_submission
                    else None
                ),
            }
        )

    return JsonResponse(
        {
            "ok": True,
            "batches": rows,
        }
    )
# ============================================================
# Eliminar un proyecto de un Batch
# ============================================================


@login_required
@rol_requerido(
    "admin",
    "pm",
    "supervisor",
    "facturacion",
    "emision_facturacion",
)
@require_POST
@transaction.atomic
def submission_delete(
    request: HttpRequest,
    public_id,
) -> JsonResponse:
    """
    Elimina un ClientSubmission individual del Batch sin
    eliminar el Batch completo.

    Reglas:

    - Permite eliminar proyectos que todavía no han comenzado
      procesamiento real.
    - Permite eliminar PENDING_CLIENT_SUBMISSION, FAILED
      y CANCELLED.
    - No permite eliminar proyectos que ya estén activos,
      enviados al cliente o completados en Dry Run.
    - No permite eliminar el proyecto que actualmente está
      siendo procesado por el worker.
    - No permite eliminar mientras el Batch está RUNNING
      o AWAITING_VERIFICATION.
    - No permite eliminar el último proyecto del Batch.
    - Renumera sequence_number después de eliminar.
    - Registra auditoría mediante ClientSubmissionEvent.
    """

    _assert_manage_permission(
        request,
    )

    # ========================================================
    # Buscar Submission inicial
    # ========================================================

    submission = get_object_or_404(
        ClientSubmission.objects.select_related(
            "batch",
            "billing_session",
        ),
        public_id=public_id,
    )

    # ========================================================
    # Bloquear Batch
    # ========================================================

    batch = get_object_or_404(
        ClientSubmissionBatch.objects.select_for_update(),
        pk=submission.batch_id,
    )

    # ========================================================
    # Bloquear nuevamente Submission dentro de la transacción
    # ========================================================

    submission = get_object_or_404(
        ClientSubmission.objects.select_for_update().select_related(
            "billing_session",
        ),
        pk=submission.pk,
        batch=batch,
    )

    # ========================================================
    # Protección del Batch
    #
    # RUNNING:
    # El worker ya puede estar trabajando activamente.
    #
    # AWAITING_VERIFICATION:
    # Existe un proyecto detenido esperando intervención
    # humana / CAPTCHA.
    # ========================================================

    if batch.status == ClientSubmissionBatch.Status.RUNNING:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project cannot be deleted while the batch "
                    "is actively being processed. "
                    "Pause the batch before removing a project."
                ),
            },
            status=409,
        )

    if batch.status == ClientSubmissionBatch.Status.AWAITING_VERIFICATION:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project cannot be deleted while the batch "
                    "is awaiting human verification."
                ),
            },
            status=409,
        )

    # ========================================================
    # No eliminar el Submission actual del worker
    # ========================================================

    if batch.current_submission_id == submission.pk:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project is currently being processed by "
                    "the background worker and cannot be deleted."
                ),
            },
            status=409,
        )

    # ========================================================
    # Estados del Submission que NO pueden eliminarse
    #
    # Desde PREPARING en adelante ya comenzó procesamiento.
    # ========================================================

    protected_submission_statuses = {
        ClientSubmission.Status.PREPARING,
        ClientSubmission.Status.AWAITING_VERIFICATION,
        ClientSubmission.Status.SUBMITTING,
        ClientSubmission.Status.AWAITING_EMAIL_CONFIRMATION,
        ClientSubmission.Status.SENT_TO_CLIENT,
        ClientSubmission.Status.DRY_RUN_COMPLETED,
    }

    if submission.status in protected_submission_statuses:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project can no longer be deleted because "
                    "client processing has already started."
                ),
            },
            status=409,
        )

    # ========================================================
    # Estados que sí permitimos eliminar
    #
    # Normalmente:
    #
    # PENDING_CLIENT_SUBMISSION
    # FAILED
    # CANCELLED
    #
    # Esto evita borrar estados futuros accidentalmente.
    # ========================================================

    deletable_submission_statuses = {
        ClientSubmission.Status.PENDING_CLIENT_SUBMISSION,
        ClientSubmission.Status.FAILED,
        ClientSubmission.Status.CANCELLED,
    }

    if submission.status not in deletable_submission_statuses:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "This project cannot be deleted from its current "
                    f"status: '{submission.get_status_display()}'."
                ),
            },
            status=409,
        )

    # ========================================================
    # No permitir dejar el Batch vacío
    # ========================================================

    total_submissions = batch.submissions.count()

    if total_submissions <= 1:
        return JsonResponse(
            {
                "ok": False,
                "error": (
                    "The last project cannot be removed from the batch. "
                    "Delete the entire batch instead."
                ),
            },
            status=409,
        )

    # ========================================================
    # Snapshot para auditoría antes de eliminar
    # ========================================================

    deleted_submission_id = submission.pk

    deleted_public_id = str(
        submission.public_id,
    )

    deleted_project_id = submission.project_id

    deleted_sequence_number = submission.sequence_number

    deleted_billing_session_id = submission.billing_session_id

    deleted_status = submission.status

    deleted_status_label = submission.get_status_display()

    deleted_validation_ok = submission.validation_ok

    deleted_validation_errors = submission.validation_errors or []

    deleted_last_error_code = submission.last_error_code

    deleted_last_error_message = submission.last_error_message

    # ========================================================
    # Protección adicional:
    # limpiar current_submission si existiera una referencia
    # residual inesperada.
    #
    # Normalmente esta rama no debería ejecutarse porque arriba
    # ya bloqueamos ese caso, pero la dejamos por consistencia.
    # ========================================================

    if batch.current_submission_id == submission.pk:
        batch.current_submission = None

        batch.save(
            update_fields=[
                "current_submission",
                "updated_at",
            ]
        )

    # ========================================================
    # Eliminar Submission
    #
    # IMPORTANTE:
    # Esto NO elimina SesionBilling.
    # Solo elimina el proyecto dentro de Client Submissions.
    # ========================================================

    submission.delete()

    # ========================================================
    # Renumerar proyectos restantes
    # ========================================================

    remaining_submissions = list(
        batch.submissions.select_for_update().order_by(
            "sequence_number",
            "id",
        )
    )

    submissions_to_update = []

    for (
        new_sequence_number,
        remaining_submission,
    ) in enumerate(
        remaining_submissions,
        start=1,
    ):
        if remaining_submission.sequence_number == new_sequence_number:
            continue

        remaining_submission.sequence_number = new_sequence_number

        submissions_to_update.append(
            remaining_submission,
        )

    if submissions_to_update:
        ClientSubmission.objects.bulk_update(
            submissions_to_update,
            [
                "sequence_number",
                "updated_at",
            ],
        )

    # ========================================================
    # Auditoría
    # ========================================================

    ClientSubmissionEvent.objects.create(
        batch=batch,
        level=ClientSubmissionEvent.Level.WARNING,
        event_type="submission_removed_from_batch",
        message=(
            f"Project {deleted_project_id} was removed from "
            f"Batch #{batch.pk} by "
            f"{request.user.get_username()}."
        ),
        metadata={
            "user_id": request.user.pk,
            "username": request.user.get_username(),
            "submission_id": deleted_submission_id,
            "submission_public_id": deleted_public_id,
            "project_id": deleted_project_id,
            "billing_session_id": (deleted_billing_session_id),
            "sequence_number": (deleted_sequence_number),
            "status": deleted_status,
            "status_label": (deleted_status_label),
            "validation_ok": (deleted_validation_ok),
            "validation_errors": (deleted_validation_errors),
            "last_error_code": (deleted_last_error_code),
            "last_error_message": (deleted_last_error_message),
        },
    )

    # ========================================================
    # Actualizar actividad del Batch
    # ========================================================

    batch.last_activity_at = timezone.now()

    batch.save(
        update_fields=[
            "last_activity_at",
            "updated_at",
        ]
    )

    # ========================================================
    # Contadores después de eliminar
    # ========================================================

    counts = _submission_status_counts(
        batch,
    )

    # ========================================================
    # Respuesta
    # ========================================================

    return JsonResponse(
        {
            "ok": True,
            "message": (f"Project {deleted_project_id} was removed " "from the batch."),
            "deleted_submission_id": (deleted_submission_id),
            "deleted_project_id": (deleted_project_id),
            "batch": {
                "id": batch.pk,
                "public_id": str(
                    batch.public_id,
                ),
                "status": batch.status,
                "status_label": (batch.get_status_display()),
                "counts": counts,
            },
        }
    )
