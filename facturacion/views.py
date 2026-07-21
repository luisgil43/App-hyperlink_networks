# ajusta si tu decorador está en otro módulo
import datetime
import datetime as dt
import re
import traceback
from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

import openpyxl
import pdfplumber
import xlwt
from dateutil import parser
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required, user_passes_test
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (Case, DecimalField, ExpressionWrapper, F,
                              IntegerField, Prefetch, Q, Subquery, Sum, Value,
                              When)
from django.db.models.functions import Coalesce
from django.http import (HttpResponse, HttpResponseBadRequest,
                         HttpResponseForbidden, HttpResponseNotAllowed,
                         JsonResponse)
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.utils.module_loading import import_string
from django.utils.timezone import is_aware
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from facturacion.models import CartolaMovimiento, Proyecto
from operaciones.forms import MovimientoUsuarioForm
from operaciones.models import (EvidenciaFotoBilling, ItemBilling,
                                ItemBillingTecnico, SesionBilling,
                                SesionBillingTecnico)
from usuarios.decoradores import rol_requerido
from usuarios.models import ProyectoAsignacion

from .forms import (CartolaAbonoForm, CartolaGastoForm,
                    CartolaMovimientoCompletoForm, ProyectoForm, TipoGastoForm)
from .models import CartolaMovimiento, Proyecto, TipoGasto

User = get_user_model()

import re

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import IntegrityError, transaction
from django.db.models import IntegerField
from django.db.models.functions import Cast, Substr
from django.shortcuts import get_object_or_404, redirect
from django.urls import reverse

# ⬇️ agrega junto a tus imports
from core.decorators import project_object_access_required
from core.permissions import (filter_queryset_by_access, projects_ids_for_user,
                              user_has_project_access)

from .models import CartolaMovimiento, Proyecto


def suggest_next_project_code(prefix: str = "PRJ-", width: int = 6) -> str:
    pattern = rf"^{re.escape(prefix)}\d{{{width}}}$"
    start = len(prefix) + 1  # Substr es 1-indexado

    qs = (Proyecto.objects
          .filter(codigo__regex=pattern)
          .annotate(num=Cast(Substr('codigo', start, width), IntegerField()))
          .order_by('-num'))

    last = qs.values_list('num', flat=True).first() or 0
    return f"{prefix}{last + 1:0{width}d}"

def _next_project_code():
    # extrae los 6 dígitos y calcula el siguiente
    qs = (Proyecto.objects
          .filter(codigo__regex=r'^PRJ-\d{6}$')
          .annotate(n=Cast(Substr('codigo', 5, 6), IntegerField()))
          .order_by('-n'))
    last = qs.first().n if qs.exists() else 0
    return f"PRJ-{last + 1:06d}"

@login_required
@rol_requerido('facturacion', 'admin')
def listar_cartola(request):
    import json
    import re
    from datetime import datetime as py_datetime

    from django.contrib import messages
    from django.core.paginator import Paginator
    from django.db.models import Case, CharField, IntegerField, Q, Value, When
    from django.db.models.functions import Cast

    # --- Cantidad/paginación (mantén string para la UI)
    cantidad_param = request.GET.get('cantidad', '10')

    # Limitamos a máx. 100 (y "todos" también se interpreta como 100)
    if cantidad_param == 'todos':
        page_size = 100
    else:
        try:
            page_size = max(5, min(int(cantidad_param), 100))
        except ValueError:
            page_size = 10
            cantidad_param = '10'

    # ---------- Filtros tipo Excel recibidos por GET ----------
    params = request.GET.copy()
    excel_filters_raw = (params.get('excel_filters') or '').strip()
    try:
        excel_filters = json.loads(excel_filters_raw) if excel_filters_raw else {}
    except json.JSONDecodeError:
        excel_filters = {}

    # --- Filtros (string trimming)
    usuario = (request.GET.get('usuario') or '').strip()
    fecha_txt = (request.GET.get('fecha') or '').strip()
    proyecto = (request.GET.get('proyecto') or '').strip()
    categoria = (request.GET.get('categoria') or '').strip()
    tipo = (request.GET.get('tipo') or '').strip()
    rut_factura = (request.GET.get('rut_factura') or '').strip()
    estado = (request.GET.get('estado') or '').strip()

    # --- Base queryset (✅ FIX PERFORMANCE: evita N+1 con select_related)
    movimientos = (
        CartolaMovimiento.objects
        .select_related(
            "usuario",
            "proyecto",
            "tipo",
            "aprobado_por_supervisor",
            "aprobado_por_pm",
            "aprobado_por_finanzas",
        )
    )

    # Anotar fecha como texto ISO (YYYY-MM-DD...) para poder usar icontains
    movimientos = movimientos.annotate(
        fecha_iso=Cast('fecha', CharField())
    )

    # --- Filtro usuario: username / nombres / apellidos
    if usuario:
        movimientos = movimientos.filter(
            Q(usuario__username__icontains=usuario) |
            Q(usuario__first_name__icontains=usuario) |
            Q(usuario__last_name__icontains=usuario)
        )

    # --- Filtro fecha (permite parcial: DD, DD-MM, DD-MM-YYYY)
    if fecha_txt:
        # Permitimos 08-12-2025, 08/12/2025, 8-12, 29, etc.
        fecha_normalizada = fecha_txt.replace('/', '-').strip()

        # Patrones permitidos: "D", "DD", "DD-MM", "DD-M", "DD-MM-YYYY"
        m = re.match(r'^(\d{1,2})(?:-(\d{1,2}))?(?:-(\d{1,4}))?$', fecha_normalizada)
        if m:
            dia_str = m.group(1)
            mes_str = m.group(2)
            anio_str = m.group(3)

            try:
                q_fecha = Q()

                # Siempre filtramos por día
                dia = int(dia_str)
                q_fecha &= Q(fecha__day=dia)

                # Si hay mes, también filtramos por mes
                if mes_str:
                    mes = int(mes_str)
                    q_fecha &= Q(fecha__month=mes)

                # Si hay año completo (4 dígitos), filtramos por año
                if anio_str and len(anio_str) == 4:
                    anio = int(anio_str)
                    q_fecha &= Q(fecha__year=anio)

                movimientos = movimientos.filter(q_fecha)

            except ValueError:
                messages.warning(
                    request,
                    "Formato de fecha inválido. Use DD, DD-MM o DD-MM-YYYY."
                )
        else:
            messages.warning(
                request,
                "Formato de fecha inválido. Use DD, DD-MM o DD-MM-YYYY."
            )

    # --- Otros filtros
    if proyecto:
        movimientos = movimientos.filter(proyecto__nombre__icontains=proyecto)
    if categoria:
        movimientos = movimientos.filter(tipo__categoria__icontains=categoria)
    if tipo:
        movimientos = movimientos.filter(tipo__nombre__icontains=tipo)
    if rut_factura:
        movimientos = movimientos.filter(rut_factura__icontains=rut_factura)
    if estado:
        movimientos = movimientos.filter(status=estado)

    # ✅ SOLO MOSTRAR: pendientes por finanzas + aprobados por finanzas
    movimientos = movimientos.filter(status__in=['aprobado_pm', 'aprobado_finanzas'])

    # --- ORDEN personalizado (se mantiene igual)
    movimientos = movimientos.annotate(
        prioridad=Case(
            When(status='aprobado_pm', then=Value(0)),
            When(status__startswith='pendiente', then=Value(1)),
            When(status__startswith='rechazado', then=Value(2)),
            When(status__startswith='aprobado', then=Value(3)),
            default=Value(4),
            output_field=IntegerField(),
        )
    ).order_by('prioridad', '-fecha')

    # ============================================================
    #  Traemos TODO a memoria:
    #  - aplicar filtros Excel (Python)
    #  - construir excel_global (todos los registros filtrados)
    #  - luego paginar
    # ============================================================
    movimientos_list = list(movimientos)

    # Helpers de formato CLP (para que los valores del panel se vean igual que la tabla)
    def format_clp(n):
        try:
            n = 0 if n is None else n
            n_int = int(n)
            return f"${n_int:,}".replace(",", ".")
        except Exception:
            return "$0"

    # ---------- Aplicar filtros Excel ----------
    if excel_filters:
        def matches_excel_filters(m):
            for col, values in excel_filters.items():
                if not values:
                    continue
                values_set = set(values)

                # índices según tabla GZ:
                # 0 Usuario
                # 1 Fecha
                # 2 Fecha real del gasto
                # 3 Proyecto
                # 4 Categoría
                # 5 Tipo
                # 6 RUT factura
                # 7 Tipo de documento
                # 8 N° Documento
                # 9 Observaciones
                # 10 N° Transferencia
                # 11 Comprobante
                # 12 Cargos
                # 13 Abonos
                # 14 Status
                if col == "0":
                    label = str(m.usuario) if m.usuario else ""

                elif col == "1":
                    d = getattr(m, "fecha", None)
                    label = d.strftime("%d-%m-%Y") if d else ""

                elif col == "2":
                    d = getattr(m, "fecha_transaccion", None) or getattr(m, "fecha", None)
                    label = d.strftime("%d-%m-%Y") if d else ""

                elif col == "3":
                    label = str(m.proyecto) if m.proyecto else ""

                elif col == "4":
                    if m.tipo and getattr(m.tipo, "categoria", None):
                        label = (m.tipo.categoria or "").title()
                    else:
                        label = ""

                elif col == "5":
                    label = str(m.tipo) if m.tipo else ""

                elif col == "6":
                    label = (getattr(m, "rut_factura", None) or "—").strip() or "—"

                elif col == "7":
                    label = (getattr(m, "tipo_doc", None) or "—").strip() or "—"

                elif col == "8":
                    label = (getattr(m, "numero_doc", None) or "—").strip() or "—"

                elif col == "9":
                    label = (getattr(m, "observaciones", None) or "").strip()

                elif col == "10":
                    label = (getattr(m, "numero_transferencia", None) or "—").strip() or "—"

                elif col == "11":
                    label = "Ver" if getattr(m, "comprobante", None) else "—"

                elif col == "12":
                    label = format_clp(getattr(m, "cargos", 0) or 0)

                elif col == "13":
                    label = format_clp(getattr(m, "abonos", 0) or 0)

                elif col == "14":
                    label = m.get_status_display() if getattr(m, "status", None) else ""

                else:
                    continue

                if label not in values_set:
                    return False

            return True

        movimientos_list = [m for m in movimientos_list if matches_excel_filters(m)]

    # ---------- Distinct globales para filtros tipo Excel ----------
    excel_global = {}

    # 0 Usuario
    excel_global[0] = sorted({str(m.usuario) for m in movimientos_list if m.usuario})

    # 1 Fecha
    excel_global[1] = sorted({
        m.fecha.strftime("%d-%m-%Y")
        for m in movimientos_list
        if getattr(m, "fecha", None)
    })

    # 2 Fecha real del gasto
    excel_global[2] = sorted({
        (getattr(m, "fecha_transaccion", None) or getattr(m, "fecha", None)).strftime("%d-%m-%Y")
        for m in movimientos_list
        if (getattr(m, "fecha_transaccion", None) or getattr(m, "fecha", None))
    })

    # 3 Proyecto
    excel_global[3] = sorted({str(m.proyecto) for m in movimientos_list if m.proyecto})

    # 4 Categoría
    excel_global[4] = sorted({
        (m.tipo.categoria or "").title()
        for m in movimientos_list
        if m.tipo and getattr(m.tipo, "categoria", None)
    })

    # 5 Tipo
    excel_global[5] = sorted({str(m.tipo) for m in movimientos_list if m.tipo})

    # 6 RUT factura
    excel_global[6] = sorted({
        (getattr(m, "rut_factura", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    # 7 Tipo doc
    excel_global[7] = sorted({
        (getattr(m, "tipo_doc", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    # 8 N° doc
    excel_global[8] = sorted({
        (getattr(m, "numero_doc", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    # 9 Observaciones
    excel_global[9] = sorted({
        (getattr(m, "observaciones", None) or "").strip()
        for m in movimientos_list
    })

    # 10 N° transferencia
    excel_global[10] = sorted({
        (getattr(m, "numero_transferencia", None) or "—").strip() or "—"
        for m in movimientos_list
    })

    # 11 Comprobante
    excel_global[11] = sorted({
        "Ver" if getattr(m, "comprobante", None) else "—"
        for m in movimientos_list
    })

    # 12 Cargos
    excel_global[12] = sorted({
        format_clp(getattr(m, "cargos", 0) or 0)
        for m in movimientos_list
    })

    # 13 Abonos
    excel_global[13] = sorted({
        format_clp(getattr(m, "abonos", 0) or 0)
        for m in movimientos_list
    })

    # 14 Status (display)
    estado_map = dict(CartolaMovimiento.ESTADOS)
    status_codes = {m.status for m in movimientos_list if getattr(m, "status", None)}
    excel_global[14] = sorted(estado_map.get(c, c) for c in status_codes)

    excel_global_json = json.dumps(excel_global)

    # --- Paginación (después de filtros Excel)
    paginator = Paginator(movimientos_list, page_size)
    page_number = request.GET.get('page')
    pagina = paginator.get_page(page_number)

    # --- Estado choices y eco de filtros a la plantilla
    estado_choices = CartolaMovimiento.ESTADOS
    filtros = {
        'usuario': usuario,
        'fecha': fecha_txt,
        'proyecto': proyecto,
        'categoria': categoria,
        'tipo': tipo,
        'rut_factura': rut_factura,
        'estado': estado,
    }

    # qs helpers (igual estilo Hyperlink, por si luego quieres usar)
    params_no_page = params.copy()
    params_no_page.pop('page', None)
    base_qs = params_no_page.urlencode()
    full_qs = params.urlencode()

    return render(
        request,
        'facturacion/listar_cartola.html',
        {
            'pagina': pagina,
            'cantidad': cantidad_param,
            'estado_choices': estado_choices,
            'filtros': filtros,
            'excel_global_json': excel_global_json,
            'base_qs': base_qs,
            'full_qs': full_qs,
        }
    )


@login_required
@rol_requerido('facturacion', 'admin')
def registrar_abono(request):
    # Detectar el usuario “destinatario” del abono (viene en el form)
    target_user = None
    user_field_names = ('usuario', 'user', 'tecnico')
    if request.method == 'POST':
        for fn in user_field_names:
            uid = request.POST.get(fn)
            if uid:
                try:
                    target_user = User.objects.get(pk=uid)
                except User.DoesNotExist:
                    target_user = None
                break
    else:
        for fn in user_field_names:
            uid = request.GET.get(fn)
            if uid:
                try:
                    target_user = User.objects.get(pk=uid)
                except User.DoesNotExist:
                    target_user = None
                break

    form = CartolaAbonoForm(request.POST or None, request.FILES or None)

    # 🔒 Restringir el combo de proyectos del formulario
    if hasattr(form, 'fields') and 'proyecto' in form.fields:
        # Si ya eligieron un usuario destino, mostrar SOLO proyectos donde él participa.
        # Si no, mostrar (por defecto) los proyectos a los que el actor (tú) tiene acceso.
        allowed_ids = projects_ids_for_user(target_user) if target_user else projects_ids_for_user(request.user)
        form.fields['proyecto'].queryset = Proyecto.objects.filter(id__in=allowed_ids).order_by('nombre')

    if request.method == 'POST':
        if form.is_valid():
            movimiento = form.save(commit=False)

            # Si el form NO setea el usuario y lo detectamos arriba, lo fijamos.
            if target_user and not getattr(movimiento, 'usuario_id', None):
                movimiento.usuario = target_user

            proj_id = getattr(getattr(movimiento, 'proyecto', None), 'id', None)
            if not proj_id:
                messages.error(request, "You must choose a project.")
                return render(request, 'facturacion/registrar_abono.html', {'form': form})

            # 🔒 1) El actor debe tener acceso al proyecto
            if not user_has_project_access(request.user, proj_id):
                messages.error(request, "You don't have access to the selected project.")
                return render(request, 'facturacion/registrar_abono.html', {'form': form})

            # 🔒 2) El usuario destino debe participar en ese proyecto
            if getattr(movimiento, 'usuario_id', None):
                target_allowed = projects_ids_for_user(movimiento.usuario)
                if proj_id not in target_allowed:
                    messages.error(request, "The selected user is not assigned to that project.")
                    return render(request, 'facturacion/registrar_abono.html', {'form': form})

            # Forzar categoría como abono
            tipo_abono = TipoGasto.objects.filter(categoria='abono').first()
            movimiento.tipo = tipo_abono
            movimiento.cargos = 0

            if 'comprobante' in request.FILES:
                movimiento.comprobante = request.FILES['comprobante']

            movimiento.save()
            messages.success(request, "Transaction registered successfully.")
            return redirect('facturacion:listar_cartola')
        else:
            messages.error(request, "Please correct the errors before proceeding.")

    return render(request, 'facturacion/registrar_abono.html', {'form': form})


@login_required
@rol_requerido('facturacion', 'admin')
def crear_tipo(request):
    if request.method == 'POST':
        form = TipoGastoForm(request.POST)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                tipos = TipoGasto.objects.all().order_by('-id')
                html = render_to_string(
                    'facturacion/partials/tipo_gasto_table.html', {'tipos': tipos})
                return JsonResponse({'success': True, 'html': html})
            messages.success(request, "Expense type created successfully.")
            return redirect('facturacion:crear_tipo')
    else:
        form = TipoGastoForm()
    tipos = TipoGasto.objects.all().order_by('-id')
    return render(request, 'facturacion/crear_tipo.html', {'form': form, 'tipos': tipos})


@login_required
@rol_requerido('admin')
def editar_tipo(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    if request.method == 'POST':
        form = TipoGastoForm(request.POST, instance=tipo)
        if form.is_valid():
            form.save()
            if request.headers.get('x-requested-with') == 'XMLHttpRequest':
                tipos = TipoGasto.objects.all().order_by('-id')
                html = render_to_string(
                    'facturacion/partials/tipo_gasto_table.html', {'tipos': tipos})
                return JsonResponse({'success': True, 'html': html})
            messages.success(request, "Expense type updated successfully.")
            return redirect('facturacion:crear_tipo')
    else:
        form = TipoGastoForm(instance=tipo)
    tipos = TipoGasto.objects.all().order_by('-id')
    # Usamos el mismo template que crear
    return render(request, 'facturacion/crear_tipo.html', {'form': form, 'tipos': tipos, 'editando': True})


@login_required
@rol_requerido('admin')
def eliminar_tipo(request, pk):
    tipo = get_object_or_404(TipoGasto, pk=pk)
    tipo.delete()
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        tipos = TipoGasto.objects.all().order_by('-id')
        html = render_to_string(
            'facturacion/partials/tipo_gasto_table.html', {'tipos': tipos})
        return JsonResponse({'success': True, 'html': html})
    messages.success(request, "Expense type deleted successfully.")
    return redirect('facturacion:crear_tipo')


@login_required
@rol_requerido("admin")
@require_POST
def toggle_tipo(request, pk):
    """
    Activa/Desactiva un TipoGasto usando is_active.
    - POST only
    - Si es AJAX: devuelve html del tbody actualizado
    - Si no: redirect a crear_tipo
    """
    tipo = get_object_or_404(TipoGasto, pk=pk)
    tipo.is_active = not tipo.is_active
    tipo.save(update_fields=["is_active"])

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        tipos = TipoGasto.objects.all().order_by("-id")
        html = render_to_string(
            "facturacion/partials/tipo_gasto_table.html",
            {"tipos": tipos},
            request=request,
        )
        return JsonResponse(
            {
                "success": True,
                "html": html,
                "is_active": tipo.is_active,
                "msg": f'Expense type "{tipo.nombre}" {"activated" if tipo.is_active else "deactivated"} successfully.',
            }
        )

    messages.success(
        request,
        f'Expense type "{tipo.nombre}" {"activated" if tipo.is_active else "deactivated"} successfully.',
    )
    return redirect("facturacion:crear_tipo")


@login_required
@rol_requerido('facturacion', 'admin')
def crear_proyecto(request):
    if request.method == 'POST':
        form = ProyectoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Project created successfully.")
            return redirect('facturacion:crear_proyecto')
    else:
        next_code = _next_project_code()
        form = ProyectoForm(initial={'codigo': next_code, 'activo': '1'})  # preselecciona Active

    proyectos = Proyecto.objects.all().order_by('-id')
    return render(request, 'facturacion/crear_proyecto.html', {
        'form': form,
        'proyectos': proyectos,
        'next_code': next_code if request.method != 'POST' else None,
    })


# Editar
@login_required
@rol_requerido('admin')
def editar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    if request.method == 'POST':
        form = ProyectoForm(request.POST, instance=proyecto)
        if form.is_valid():
            form.save()
            messages.success(request, "Project updated successfully.")
            return redirect('facturacion:crear_proyecto')
    else:
        form = ProyectoForm(instance=proyecto)
    proyectos = Proyecto.objects.all().order_by('-id')
    return render(request, 'facturacion/crear_proyecto.html', {
        'form': form,
        'proyectos': proyectos
    })

# Eliminar


@login_required
@rol_requerido('admin')
def eliminar_proyecto(request, pk):
    proyecto = get_object_or_404(Proyecto, pk=pk)
    if request.method == 'POST':
        proyecto.delete()
        messages.success(request, "Project deleted successfully.")
        return redirect('facturacion:crear_proyecto')
    return redirect('facturacion:crear_proyecto')


@login_required
@rol_requerido('facturacion', 'supervisor', 'pm', 'admin')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def aprobar_movimiento(request, pk):
    """
    Aprueba un movimiento.

    - Si es petición AJAX → responde JSON (sin redirect).
    - Si es petición normal → mantiene el redirect a listar_cartola.
    """
    mov = get_object_or_404(CartolaMovimiento, pk=pk)

    ok = False
    prev_status = mov.status
    new_status = mov.status

    if mov.tipo and mov.tipo.categoria != "abono":
        # Asignar aprobador según el rol / estado actual
        if getattr(request.user, "es_supervisor", False) and mov.status == 'pendiente_supervisor':
            mov.status = 'aprobado_supervisor'
            mov.aprobado_por_supervisor = request.user
        elif getattr(request.user, "es_pm", False) and mov.status == 'aprobado_supervisor':
            mov.status = 'aprobado_pm'
            mov.aprobado_por_pm = request.user
        elif getattr(request.user, "es_facturacion", False) and mov.status == 'aprobado_pm':
            mov.status = 'aprobado_finanzas'
            mov.aprobado_por_finanzas = request.user

        if mov.status != prev_status:
            mov.motivo_rechazo = ''  # Limpiar cualquier rechazo previo
            mov.save()
            ok = True
            new_status = mov.status

    # --- Si es AJAX, devolvemos JSON ---
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        if not ok:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Movement cannot be approved in the current state or category.",
                    "status": mov.status,
                    "status_display": mov.get_status_display(),
                },
                status=400,
            )

        return JsonResponse(
            {
                "ok": True,
                "id": mov.pk,
                "status": new_status,
                "status_display": mov.get_status_display(),
            }
        )

    # --- Flujo normal (no AJAX) ---
    if ok:
        messages.success(request, "Expense approved successfully.")
    else:
        messages.error(request, "Expense cannot be approved in the current state.")
    return redirect('facturacion:listar_cartola')


@login_required
@rol_requerido('facturacion', 'supervisor', 'pm', 'admin')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def rechazar_movimiento(request, pk):
    mov = get_object_or_404(CartolaMovimiento, pk=pk)
    if request.method == 'POST':
        motivo = request.POST.get('motivo_rechazo', '').strip()
        if mov.tipo and mov.tipo.categoria != "abono":
            if request.user.es_supervisor and mov.status == 'pendiente_supervisor':
                mov.status = 'rechazado_supervisor'
                mov.aprobado_por_supervisor = request.user
            elif request.user.es_pm and mov.status == 'aprobado_supervisor':
                mov.status = 'rechazado_pm'
                mov.aprobado_por_pm = request.user
            elif request.user.es_facturacion and mov.status == 'aprobado_pm':
                mov.status = 'rechazado_finanzas'
                mov.aprobado_por_finanzas = request.user  # Usuario de finanzas

            mov.motivo_rechazo = motivo
            mov.save()
            messages.success(request, "Expense rejected successfully.")
    return redirect('facturacion:listar_cartola')

@login_required
@rol_requerido('admin')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def aprobar_abono_como_usuario(request, pk):
    mov = get_object_or_404(CartolaMovimiento, pk=pk)
    if mov.tipo and mov.tipo.categoria == "abono" and mov.status == "pendiente_abono_usuario":
        mov.status = "aprobado_abono_usuario"
        mov.save()
        messages.success(request, "Deposit approved as user.")
    next_url = request.GET.get('next') or request.META.get('HTTP_REFERER') or reverse('facturacion:listar_cartola')
    return redirect(next_url)


from django import forms  # 👈 agregar este import arriba


@login_required
@rol_requerido('facturacion', 'admin')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def editar_movimiento(request, pk):
    movimiento = get_object_or_404(CartolaMovimiento, pk=pk)

    # ¿Es abono o gasto?
    es_abono = bool(movimiento.tipo and movimiento.tipo.categoria == "abono")
    FormClass = CartolaAbonoForm if es_abono else MovimientoUsuarioForm
    estado_restaurado = 'pendiente_abono_usuario' if es_abono else 'pendiente_supervisor'

    def ensure_real_consumption_date_field(form):
        """
        Asegura que el form tenga real_consumption_date para precargar y guardar,
        sin depender de cómo esté declarado el ModelForm.
        """
        if hasattr(movimiento, "real_consumption_date") and "real_consumption_date" not in form.fields:
            form.fields["real_consumption_date"] = forms.DateField(
                required=False,
                widget=forms.DateInput(attrs={"type": "date"})
            )
            # precarga
            if movimiento.real_consumption_date:
                val = movimiento.real_consumption_date
                # por si fuera datetime
                if hasattr(val, "date"):
                    try:
                        val = timezone.localtime(val).date()
                    except Exception:
                        val = val.date()
                form.initial["real_consumption_date"] = val
        return form

    if request.method == 'POST':
        form = FormClass(request.POST, request.FILES, instance=movimiento)
        form = ensure_real_consumption_date_field(form)

        if form.is_valid():
            movimiento = form.save(commit=False)

            # ✅ guardar nuevo campo si viene en el form
            if "real_consumption_date" in form.cleaned_data:
                movimiento.real_consumption_date = form.cleaned_data.get("real_consumption_date")

            # Reemplazo explícito del comprobante si viene un archivo nuevo
            if 'comprobante' in request.FILES:
                movimiento.comprobante = request.FILES['comprobante']

            # Reemplazo explícito de la foto del tablero (solo si viene)
            if 'foto_tablero' in request.FILES:
                movimiento.foto_tablero = request.FILES['foto_tablero']

            # ----- Reset de estado -----
            if es_abono and movimiento.status == 'rechazado_abono_usuario':
                movimiento.status = 'pendiente_abono_usuario'
                movimiento.motivo_rechazo = ""
            elif form.changed_data:
                movimiento.status = estado_restaurado
                movimiento.motivo_rechazo = ""

            movimiento.save()
            messages.success(request, "Expense updated successfully.")

            next_url = (
                request.GET.get('next')
                or request.POST.get('next')
                or request.META.get('HTTP_REFERER')
                or reverse('facturacion:listar_cartola')
            )
            return redirect(next_url)
    else:
        form = FormClass(instance=movimiento)
        form = ensure_real_consumption_date_field(form)

    return render(request, 'facturacion/editar_movimiento.html', {
        'form': form,
        'movimiento': movimiento
    })


@login_required
@rol_requerido('admin')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def eliminar_movimiento(request, pk):
    movimiento = get_object_or_404(CartolaMovimiento, pk=pk)

    if request.method == 'POST':
        movimiento.delete()
        messages.success(request, "Expense deleted successfully.")

        # 👇 RESPETAR FILTROS (next con todos los params)
        next_url = (
            request.GET.get('next')
            or request.POST.get('next')
            or request.META.get('HTTP_REFERER')
            or reverse('facturacion:listar_cartola')
        )
        return redirect(next_url)

    # GET: muestra la pantalla de confirmación
    return render(request, 'facturacion/eliminar_movimiento.html', {'movimiento': movimiento})


@login_required
@rol_requerido('facturacion', 'admin')
def listar_saldos_usuarios(request):
    from decimal import Decimal

    from django.core.paginator import Paginator
    from django.db.models import (Case, DecimalField, ExpressionWrapper, F, Q,
                                  Sum, Value, When)
    from django.db.models.functions import Coalesce

    from core.permissions import filter_queryset_by_assignment_history

    from .models import CartolaMovimiento

    cantidad = request.GET.get('cantidad', '5')

    USER_PENDING = ['pendiente_abono_usuario']
    SUP_PENDING  = ['pendiente_supervisor']
    PM_PENDING   = ['aprobado_supervisor']
    FIN_PENDING  = ['aprobado_pm']

    DEC = DecimalField(max_digits=12, decimal_places=2)
    V0  = Value(Decimal('0.00'), output_field=DEC)

    pend_user_abonos = Sum(
        Case(
            When(Q(abonos__gt=0) & Q(status__in=USER_PENDING), then=F('abonos')),
            default=V0, output_field=DEC,
        )
    )
    pend_sup_abonos = Sum(
        Case(
            When(Q(abonos__gt=0) & Q(status__in=SUP_PENDING), then=F('abonos')),
            default=V0, output_field=DEC,
        )
    )
    pend_sup_cargos = Sum(
        Case(
            When(Q(cargos__gt=0) & Q(status__in=SUP_PENDING), then=F('cargos')),
            default=V0, output_field=DEC,
        )
    )
    pend_pm_abonos = Sum(
        Case(
            When(Q(abonos__gt=0) & Q(status__in=PM_PENDING), then=F('abonos')),
            default=V0, output_field=DEC,
        )
    )
    pend_pm_cargos = Sum(
        Case(
            When(Q(cargos__gt=0) & Q(status__in=PM_PENDING), then=F('cargos')),
            default=V0, output_field=DEC,
        )
    )
    pend_fin_abonos = Sum(
        Case(
            When(Q(abonos__gt=0) & Q(status__in=FIN_PENDING), then=F('abonos')),
            default=V0, output_field=DEC,
        )
    )
    pend_fin_cargos = Sum(
        Case(
            When(Q(cargos__gt=0) & Q(status__in=FIN_PENDING), then=F('cargos')),
            default=V0, output_field=DEC,
        )
    )

    base = CartolaMovimiento.objects.all()

    # ✅ BLINDADO: proyecto + include_history/start_at también para agregados
    base = filter_queryset_by_assignment_history(
        base,
        request.user,
        project_field="proyecto_id",
        date_field="fecha",
    )

    qs = (
        base
        .values('usuario__id', 'usuario__first_name', 'usuario__last_name', 'usuario__email')
        .annotate(
            monto_rendido = Coalesce(Sum('cargos'), V0, output_field=DEC),
            monto_asignado = Coalesce(Sum('abonos'), V0, output_field=DEC),

            pend_user = pend_user_abonos,

            _pend_sup_abonos = pend_sup_abonos,
            _pend_sup_cargos = pend_sup_cargos,
            _pend_pm_abonos  = pend_pm_abonos,
            _pend_pm_cargos  = pend_pm_cargos,
            _pend_fin_abonos = pend_fin_abonos,
            _pend_fin_cargos = pend_fin_cargos,
        )
        .annotate(
            pend_sup = ExpressionWrapper(
                Coalesce(F('_pend_sup_abonos'), V0, output_field=DEC) +
                Coalesce(F('_pend_sup_cargos'), V0, output_field=DEC),
                output_field=DEC,
            ),
            pend_pm = ExpressionWrapper(
                Coalesce(F('_pend_pm_abonos'), V0, output_field=DEC) +
                Coalesce(F('_pend_pm_cargos'), V0, output_field=DEC),
                output_field=DEC,
            ),
            pend_fin = ExpressionWrapper(
                Coalesce(F('_pend_fin_abonos'), V0, output_field=DEC) +
                Coalesce(F('_pend_fin_cargos'), V0, output_field=DEC),
                output_field=DEC,
            ),
            monto_disponible = ExpressionWrapper(
                Coalesce(F('monto_asignado'), V0, output_field=DEC) -
                Coalesce(F('monto_rendido'), V0, output_field=DEC),
                output_field=DEC,
            ),
        )
        .order_by('usuario__first_name', 'usuario__last_name')
    )

    paginator = Paginator(qs, qs.count() or 1) if cantidad == 'todos' else Paginator(qs, int(cantidad))
    pagina = paginator.get_page(request.GET.get('page'))

    return render(request, 'facturacion/listar_saldos_usuarios.html', {
        'saldos': pagina,
        'pagina': pagina,
        'cantidad': cantidad,
    })


from datetime import datetime, time, timedelta

from django.db import models
from django.db.models import Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.timezone import is_aware


def _parse_date_any(s: str):
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


@login_required
@rol_requerido("facturacion", "admin")
def exportar_cartola(request):
    import json
    from datetime import datetime, time, timedelta
    from io import BytesIO

    from django.db import models
    from django.db.models import Q
    from django.http import HttpResponse
    from django.utils import timezone
    from django.utils.timezone import is_aware
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    from core.permissions import filter_queryset_by_assignment_history

    from .models import CartolaMovimiento

    def _parse_date_any(s: str):
        for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    def _cell_value_for_excel_filter(mov, col_idx):
        if col_idx == 0:
            return str(mov.usuario or "").strip() or "(Vacías)"

        if col_idx == 1:
            fecha = getattr(mov, "fecha", None)
            if not fecha:
                return "(Vacías)"
            if isinstance(fecha, datetime):
                if is_aware(fecha):
                    fecha = timezone.localtime(fecha)
                fecha = fecha.date()
            return fecha.strftime("%d-%m-%Y")

        if col_idx == 2:
            rcd = getattr(mov, "real_consumption_date", None)
            if not rcd:
                return "(Vacías)"
            if isinstance(rcd, datetime):
                if is_aware(rcd):
                    rcd = timezone.localtime(rcd)
                rcd = rcd.date()
            return rcd.strftime("%d-%m-%Y")

        if col_idx == 3:
            return str(getattr(mov, "proyecto", "") or "").strip() or "(Vacías)"

        if col_idx == 4:
            cat = getattr(getattr(mov, "tipo", None), "categoria", "") or ""
            return str(cat).title().strip() or "(Vacías)"

        if col_idx == 5:
            return str(getattr(mov, "tipo", "") or "").strip() or "(Vacías)"

        if col_idx == 6:
            return str(getattr(mov, "rut_factura", "") or "—").strip() or "(Vacías)"

        if col_idx == 7:
            return str(getattr(mov, "tipo_doc", "") or "—").strip() or "(Vacías)"

        if col_idx == 8:
            return str(getattr(mov, "numero_doc", "") or "—").strip() or "(Vacías)"

        if col_idx == 9:
            return str(getattr(mov, "observaciones", "") or "").strip() or "(Vacías)"

        if col_idx == 10:
            return (
                str(getattr(mov, "numero_transferencia", "") or "—").strip()
                or "(Vacías)"
            )

        if col_idx == 11:
            return "Ver" if getattr(mov, "comprobante", None) else "—"

        if col_idx == 12:
            try:
                return f"${float(getattr(mov, 'cargos', 0) or 0):,.2f}"
            except Exception:
                return "$0.00"

        if col_idx == 13:
            try:
                return f"${float(getattr(mov, 'abonos', 0) or 0):,.2f}"
            except Exception:
                return "$0.00"

        if col_idx == 14:
            return str(mov.get_status_display() or "").strip() or "(Vacías)"

        return "(Vacías)"

    params = request.GET

    du = (params.get("du") or "").strip()
    fecha_str = (params.get("fecha") or "").strip()
    proyecto = (params.get("proyecto") or "").strip()
    categoria = (params.get("categoria") or "").strip()
    tipo = (params.get("tipo") or "").strip()
    estado = (params.get("estado") or "").strip()
    rut = (params.get("rut_factura") or "").strip()

    movimientos = (
        CartolaMovimiento.objects.all()
        .select_related("usuario", "proyecto", "tipo")
        .order_by("-fecha")
    )

    # ✅ BLINDADO: proyecto + include_history/start_at
    movimientos = filter_queryset_by_assignment_history(
        movimientos,
        request.user,
        project_field="proyecto_id",
        date_field="fecha",
    )

    if du:
        movimientos = movimientos.filter(
            Q(usuario__username__icontains=du)
            | Q(usuario__first_name__icontains=du)
            | Q(usuario__last_name__icontains=du)
        )

    if fecha_str:
        if fecha_str.isdigit() and 1 <= int(fecha_str) <= 31:
            dia = int(fecha_str)
            movimientos = movimientos.filter(fecha__day=dia)
        else:
            f = _parse_date_any(fecha_str)
            if f:
                campo_fecha = CartolaMovimiento._meta.get_field("fecha")
                if isinstance(campo_fecha, models.DateTimeField):
                    tz = timezone.get_current_timezone()
                    start = timezone.make_aware(datetime.combine(f, time.min), tz)
                    end = start + timedelta(days=1)
                    movimientos = movimientos.filter(fecha__gte=start, fecha__lt=end)
                else:
                    movimientos = movimientos.filter(fecha=f)

    if proyecto:
        movimientos = movimientos.filter(proyecto__nombre__icontains=proyecto)
    if categoria:
        movimientos = movimientos.filter(tipo__categoria__icontains=categoria)
    if tipo:
        movimientos = movimientos.filter(tipo__nombre__icontains=tipo)
    if rut:
        movimientos = movimientos.filter(rut_factura__icontains=rut)
    if estado:
        movimientos = movimientos.filter(status=estado)

    # ✅ Aplicar filtros tipo Excel al export
    excel_filters_raw = params.get("excel_filters") or ""
    if excel_filters_raw:
        try:
            raw_filters = json.loads(excel_filters_raw)
            parsed_filters = {}

            for col_key, values in raw_filters.items():
                try:
                    col_idx = int(col_key)
                except Exception:
                    continue

                if isinstance(values, list):
                    parsed_filters[col_idx] = set(str(v).strip() for v in values)

            if parsed_filters:
                ids_ok = []

                for mov in movimientos:
                    keep = True

                    for col_idx, allowed_values in parsed_filters.items():
                        if not allowed_values:
                            continue

                        current_value = _cell_value_for_excel_filter(mov, col_idx)

                        if current_value not in allowed_values:
                            keep = False
                            break

                    if keep:
                        ids_ok.append(mov.id)

                movimientos = movimientos.filter(id__in=ids_ok)

        except Exception:
            pass

    # ===== Excel XLSX real =====
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False

    header_fill = PatternFill("solid", fgColor="374151")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    left = Alignment(horizontal="left", vertical="center")
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    columns = [
        "User",
        "Date",
        "Project",
        "Real consumption date",
        "Category",
        "Type",
        "Remarks",
        "Transfer Number",
        "Odometer (km)",
        "Debits",
        "Credits",
        "Status",
    ]

    ws.append(columns)

    for col_num, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_all

    if not movimientos.exists():
        ws.cell(row=2, column=1, value="Sin resultados para los filtros aplicados.")
    else:
        for row_num, mov in enumerate(movimientos, start=2):
            fecha_excel = getattr(mov, "fecha", None)
            if isinstance(fecha_excel, datetime):
                if is_aware(fecha_excel):
                    fecha_excel = timezone.localtime(fecha_excel).replace(tzinfo=None)
                fecha_excel = fecha_excel.date()

            rcd = getattr(mov, "real_consumption_date", None)
            if isinstance(rcd, datetime):
                if is_aware(rcd):
                    rcd = timezone.localtime(rcd).replace(tzinfo=None)
                rcd = rcd.date()

            cat = getattr(getattr(mov, "tipo", None), "categoria", "") or ""
            tipo_txt = str(getattr(mov, "tipo", "") or "")

            ws.cell(row=row_num, column=1, value=str(mov.usuario))
            ws.cell(row=row_num, column=2, value=fecha_excel)
            ws.cell(
                row=row_num, column=3, value=str(getattr(mov, "proyecto", "") or "")
            )

            if rcd:
                ws.cell(row=row_num, column=4, value=rcd)
            else:
                ws.cell(row=row_num, column=4, value="")

            ws.cell(row=row_num, column=5, value=str(cat).title())
            ws.cell(row=row_num, column=6, value=tipo_txt)
            ws.cell(row=row_num, column=7, value=mov.observaciones or "")
            ws.cell(row=row_num, column=8, value=mov.numero_transferencia or "")

            try:
                ws.cell(
                    row=row_num,
                    column=9,
                    value=float(mov.kilometraje) if mov.kilometraje is not None else "",
                )
            except Exception:
                ws.cell(row=row_num, column=9, value="")

            ws.cell(row=row_num, column=10, value=float(mov.cargos or 0))
            ws.cell(row=row_num, column=11, value=float(mov.abonos or 0))
            ws.cell(row=row_num, column=12, value=mov.get_status_display())

            for col in range(1, 13):
                c = ws.cell(row=row_num, column=col)
                c.border = border_all

                if col in (2, 4):
                    c.number_format = "DD-MM-YYYY"
                    c.alignment = center
                elif col in (9, 10, 11):
                    c.alignment = right
                else:
                    c.alignment = left

            ws.cell(row=row_num, column=10).number_format = "$#,##0.00"
            ws.cell(row=row_num, column=11).number_format = "$#,##0.00"

    widths = {
        1: 28,
        2: 14,
        3: 28,
        4: 22,
        5: 16,
        6: 18,
        7: 36,
        8: 18,
        9: 16,
        10: 14,
        11: 14,
        12: 28,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    now_str = timezone.localtime().strftime("%Y%m%d_%H%M%S")

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        f'attachment; filename="transactions_ledger_{now_str}.xlsx"'
    )

    return response


@login_required
@rol_requerido("facturacion", "admin")
def exportar_saldos(request):
    from datetime import datetime
    from io import BytesIO

    from django.db.models import Case, DecimalField, F, Q, Sum, Value, When
    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    USER_PENDING = [
        "pendiente_usuario",
        "pendiente_aprobacion_usuario",
        "pendiente_abono_usuario",
    ]
    SUP_PENDING = ["pendiente_supervisor"]
    PM_PENDING = ["aprobado_supervisor", "pendiente_pm"]
    FIN_PENDING = ["aprobado_pm", "pendiente_finanzas"]

    def _sum_pending_abonos(status_list):
        return Sum(
            Case(
                When(Q(abonos__gt=0) & Q(status__in=status_list), then=F("abonos")),
                default=Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )

    def _sum_pending_cargos(status_list):
        return Sum(
            Case(
                When(Q(cargos__gt=0) & Q(status__in=status_list), then=F("cargos")),
                default=Value(0),
                output_field=DecimalField(max_digits=12, decimal_places=2),
            )
        )

    # 🔒 Limitar a proyectos con acceso del usuario
    base = CartolaMovimiento.objects.all()
    base = filter_queryset_by_access(base, request.user, "proyecto_id")

    # ✅ Limitar también por fecha (ventana ProyectoAsignacion)
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    can_view_legacy_history = request.user.is_superuser or getattr(
        request.user, "es_usuario_historial", False
    )

    if (ProyectoAsignacion is not None) and (not can_view_legacy_history):
        try:
            proyecto_ids_visibles = list(
                base.values_list("proyecto_id", flat=True).distinct()
            )
        except Exception:
            proyecto_ids_visibles = []

        try:
            asignaciones = list(
                ProyectoAsignacion.objects.filter(
                    usuario=request.user, proyecto_id__in=proyecto_ids_visibles
                )
            )
        except Exception:
            asignaciones = []

        if asignaciones:
            access_by_pk = {}

            for a in asignaciones:
                if a.include_history or not a.start_at:
                    access_by_pk[a.proyecto_id] = {
                        "include_history": True,
                        "start_at": None,
                    }
                else:
                    access_by_pk[a.proyecto_id] = {
                        "include_history": False,
                        "start_at": a.start_at,
                    }

            ids_ok = []

            for mid, pid, fecha in base.values_list("id", "proyecto_id", "fecha"):
                if pid is None:
                    continue

                access = access_by_pk.get(pid)

                if not access:
                    continue

                if access["include_history"] or access["start_at"] is None:
                    ids_ok.append(mid)
                    continue

                if not fecha:
                    continue

                start_at = access["start_at"]

                if isinstance(start_at, datetime):
                    start_date = start_at.date()
                else:
                    start_date = start_at

                if isinstance(fecha, datetime):
                    fecha_date = fecha.date()
                else:
                    fecha_date = fecha

                if fecha_date >= start_date:
                    ids_ok.append(mid)

            base = base.filter(id__in=ids_ok)

    balances = (
        base.values("usuario__first_name", "usuario__last_name")
        .annotate(
            rendered_amount=Sum("cargos", default=0),
            assigned_amount=Sum("abonos", default=0),
            available_amount=Sum(F("abonos") - F("cargos"), default=0),
            pending_user=_sum_pending_abonos(USER_PENDING),
            sup_abonos=_sum_pending_abonos(SUP_PENDING),
            sup_cargos=_sum_pending_cargos(SUP_PENDING),
            pm_abonos=_sum_pending_abonos(PM_PENDING),
            pm_cargos=_sum_pending_cargos(PM_PENDING),
            fin_abonos=_sum_pending_abonos(FIN_PENDING),
            fin_cargos=_sum_pending_cargos(FIN_PENDING),
        )
        .order_by("usuario__first_name", "usuario__last_name")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Available Balances"

    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False

    header_fill = PatternFill("solid", fgColor="374151")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")

    thin = Side(style="thin", color="D1D5DB")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    columns = [
        "User",
        "Amount Rendered",
        "Assigned Amount",
        "Available Amount",
        "Pending (User)",
        "Pending (Supervisor)",
        "Pending (PM)",
        "Pending (Finance)",
    ]

    ws.append(columns)

    for col_num, title in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border_all

    for r, b in enumerate(balances, start=2):
        pend_sup = float((b["sup_abonos"] or 0) + (b["sup_cargos"] or 0))
        pend_pm = float((b["pm_abonos"] or 0) + (b["pm_cargos"] or 0))
        pend_fin = float((b["fin_abonos"] or 0) + (b["fin_cargos"] or 0))

        ws.cell(
            row=r,
            column=1,
            value=f"{b['usuario__first_name']} {b['usuario__last_name']}",
        )
        ws.cell(row=r, column=2, value=float(b["rendered_amount"] or 0))
        ws.cell(row=r, column=3, value=float(b["assigned_amount"] or 0))
        ws.cell(row=r, column=4, value=float(b["available_amount"] or 0))
        ws.cell(row=r, column=5, value=float(b["pending_user"] or 0))
        ws.cell(row=r, column=6, value=pend_sup)
        ws.cell(row=r, column=7, value=pend_pm)
        ws.cell(row=r, column=8, value=pend_fin)

        for col in range(1, 9):
            c = ws.cell(row=r, column=col)
            c.border = border_all

            if col == 1:
                c.alignment = left
            else:
                c.alignment = right
                c.number_format = "$#,##0.00"

    widths = {
        1: 28,
        2: 18,
        3: 18,
        4: 18,
        5: 18,
        6: 22,
        7: 18,
        8: 22,
    }

    for col, width in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="available_balances.xlsx"'

    return response


def _parse_decimal(val: str | None) -> Decimal | None:
    if val is None:
        return None
    s = (str(val) or "").strip().replace(",", "")
    if s == "":
        return None
    try:
        return Decimal(s)
    except (InvalidOperation, ValueError):
        return None


def _can_edit_real_week(user) -> bool:
    """
    PM / Finance / Admin can edit the real pay week.
    Ajusta esto a tu lógica de roles si usas permisos/grupos.
    """
    try:
        from usuarios.utils import user_has_any_role  # opcional
        return user_has_any_role(user, ["pm", "facturacion", "admin"])
    except Exception:
        # Fallback: usa el decorador que ya aplicamos a la vista
        return True


def _limit_invoices_by_assignment_and_history(qs, user):
    """
    Lógica de historial igual que en producción admin, usando ProyectoAsignacion:

      - Toma las ProyectoAsignacion del usuario.
      - Si include_history=True -> ve TODO el historial de ese proyecto.
      - Si include_history=False y tiene start_at -> solo ve desde start_at.
      - Resultado: solo invoices de proyectos donde está asignado.

    Si el usuario no tiene asignaciones, devuelve qs.none().
    """
    asignaciones = ProyectoAsignacion.objects.filter(usuario=user)
    if not asignaciones.exists():
        return qs.none()

    cond = Q()
    for asig in asignaciones:
        pid_str = str(asig.proyecto_id)
        if asig.include_history or not asig.start_at:
            cond |= Q(proyecto_id=pid_str)
        else:
            cond |= Q(proyecto_id=pid_str, creado_en__gte=asig.start_at)

    return qs.filter(cond)


def _attach_project_label(page_obj):
    """
    Crea un atributo 'project_label' en cada SesionBilling de la página,
    usando la MISMA filosofía que en produccion_admin:

      - Usa Proyecto (id, código, nombre).
      - Intenta resolver primero por 'proyecto' (texto),
        luego por 'proyecto_id'.
      - Si encuentra Proyecto -> usa p.nombre.
      - Si no encuentra, hace fallback a s.proyecto o s.proyecto_id.
    """
    sessions = list(page_obj.object_list)
    if not sessions:
        return page_obj

    # --- Recolectamos textos/ids candidatos de esta página ---
    raw_texts = set()
    raw_ids = set()

    for s in sessions:
        proj_text = (getattr(s, "proyecto", "") or "").strip()
        proj_id = getattr(s, "proyecto_id", None)

        for raw in (proj_text, proj_id):
            if raw in (None, "", "-"):
                continue
            txt = str(raw).strip()
            if not txt:
                continue
            raw_texts.add(txt)
            try:
                raw_ids.add(int(txt))
            except (TypeError, ValueError):
                # no es un entero, puede ser código o nombre
                pass

    # Si no hay nada, solo ponemos fallback y salimos
    if not raw_texts and not raw_ids:
        for s in sessions:
            s.project_label = getattr(s, "proyecto", None) or getattr(s, "proyecto_id", "") or ""
        page_obj.object_list = sessions
        return page_obj

    # --- Cargamos solo los proyectos relevantes ---
    proj_q = Q()
    if raw_ids:
        proj_q |= Q(pk__in=raw_ids)
    if raw_texts:
        proj_q |= Q(codigo__in=raw_texts) | Q(nombre__in=raw_texts)

    proyectos = Proyecto.objects.filter(proj_q).distinct()

    by_id = {p.id: p for p in proyectos}
    by_code = {
        (p.codigo or "").strip().lower(): p
        for p in proyectos
        if getattr(p, "codigo", None)
    }
    by_name = {
        (p.nombre or "").strip().lower(): p
        for p in proyectos
        if getattr(p, "nombre", None)
    }

    def _resolve_for_session(s):
        proj_text = (getattr(s, "proyecto", "") or "").strip()
        proj_id = getattr(s, "proyecto_id", None)

        proyecto_sel = None

        # 1) intentar con proj_text (igual que en produccion_admin)
        if proj_text:
            try:
                pid = int(proj_text)
            except (TypeError, ValueError):
                key = proj_text.lower()
                proyecto_sel = by_code.get(key) or by_name.get(key)
            else:
                proyecto_sel = by_id.get(pid)

        # 2) si no, probar con proyecto_id
        if not proyecto_sel and proj_id not in (None, "", "-"):
            txt = str(proj_id).strip()
            try:
                pid2 = int(txt)
            except (TypeError, ValueError):
                key2 = txt.lower()
                proyecto_sel = by_code.get(key2) or by_name.get(key2)
            else:
                proyecto_sel = by_id.get(pid2)

        # 3) Construir label
        if proyecto_sel:
            return getattr(proyecto_sel, "nombre", str(proyecto_sel))

        # Fallback: lo que ya teníamos en la sesión
        if proj_text:
            return proj_text
        if proj_id not in (None, "", "-"):
            return str(proj_id)
        return ""

    for s in sessions:
        s.project_label = _resolve_for_session(s)

    page_obj.object_list = sessions
    return page_obj


@login_required
@rol_requerido("facturacion", "admin")
def invoices_list(request):
    """
    Finanzas:
      - 'open': todo lo que realmente está en Finanzas
      - 'paid': solo cobrados
      - 'all': todo lo de Finanzas

    PERFORMANCE:
      - Primero filtra con queryset liviano.
      - Luego pagina IDs.
      - Solo carga relaciones pesadas para la página visible.
    """
    import json
    from datetime import date as _date
    from datetime import datetime
    from urllib.parse import urlencode

    from django.core.paginator import Paginator
    from django.db.models import Prefetch, Q
    from django.utils import timezone

    from facturacion.models import Proyecto
    from operaciones.models import (BillingPayWeekSnapshot,
                                    EvidenciaFotoBilling, ItemBilling,
                                    ItemBillingTecnico, SesionBilling,
                                    SesionBillingTecnico)

    user = request.user
    scope = (request.GET.get("scope") or "open").strip()

    can_view_legacy_history = user.is_superuser or getattr(
        user,
        "es_usuario_historial",
        False,
    )

    try:
        proyectos_user = filter_queryset_by_access(
            Proyecto.objects.all(),
            user,
            "id",
        )
    except Exception:
        proyectos_user = Proyecto.objects.none()

    if proyectos_user.exists():
        allowed_keys = set()

        for p in proyectos_user:
            nombre = (getattr(p, "nombre", "") or "").strip()

            if nombre:
                allowed_keys.add(nombre)

            codigo = getattr(p, "codigo", None)

            if codigo:
                allowed_keys.add(str(codigo).strip())

            allowed_keys.add(str(p.id).strip())
    else:
        allowed_keys = set()

    # ============================================================
    # Query base liviana.
    # NO hacemos prefetch_related aquí para no cargar todo antes
    # de paginar.
    # ============================================================
    qs = SesionBilling.objects.annotate(
        finance_status_order=Case(
            When(finance_status="sent", then=0),
            When(finance_status="sent_to_client", then=1),
            When(finance_status="pending_invoice", then=2),
            When(finance_status="invoiced", then=3),
            When(finance_status="pending", then=3),
            When(finance_status="discount_applied", then=4),
            When(finance_status="review_discount", then=4),
            When(finance_status="in_review", then=4),
            When(finance_status="rejected", then=4),
            When(finance_status="paid", then=5),
            default=4,
            output_field=IntegerField(),
        )
    ).order_by(
        "finance_status_order",
        "-creado_en",
    )

    finance_open_base = [
        "discount_applied",
        "sent",
        "sent_to_client",
        "pending_invoice",
        "invoiced",
        "in_review",
        "pending",
        "rejected",
        "paid",
    ]

    if scope == "paid":
        qs = qs.filter(finance_status="paid")

    elif scope == "all":
        qs = qs.exclude(
            Q(finance_status__in=["none", ""])
            | Q(finance_status__isnull=True)
            | (Q(finance_status="review_discount") & Q(finance_sent_at__isnull=True))
        )

    else:
        qs = qs.filter(
            Q(finance_status__in=finance_open_base)
            | (Q(finance_status="review_discount") & Q(finance_sent_at__isnull=False))
        )

    qs = qs.filter(
        Q(estado__in=["aprobado_supervisor", "aprobado_pm"])
        | Q(is_direct_discount=True)
    )

    if not can_view_legacy_history:
        if allowed_keys:
            qs = qs.filter(proyecto__in=allowed_keys)
        else:
            qs = SesionBilling.objects.none()

    # ============================================================
    # Limitar por historial de ProyectoAsignacion
    # ============================================================
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    if ProyectoAsignacion is not None and not can_view_legacy_history:
        try:
            asignaciones = list(
                ProyectoAsignacion.objects.filter(
                    usuario=user,
                    proyecto__in=proyectos_user,
                ).select_related("proyecto")
            )
        except Exception:
            asignaciones = []

        if asignaciones:
            access_by_key = {}

            for a in asignaciones:
                p = getattr(a, "proyecto", None)

                if not p:
                    continue

                if getattr(a, "include_history", False) or not getattr(
                    a,
                    "start_at",
                    None,
                ):
                    access = {
                        "include_history": True,
                        "start_at": None,
                    }
                else:
                    access = {
                        "include_history": False,
                        "start_at": a.start_at,
                    }

                for k in (
                    getattr(p, "nombre", None),
                    getattr(p, "codigo", None),
                    getattr(p, "id", None),
                ):
                    if k is None:
                        continue

                    ks = str(k).strip()

                    if ks:
                        access_by_key[ks.lower()] = access

            ids_ok = []

            for sid, proj_txt, creado_en in qs.values_list(
                "id",
                "proyecto",
                "creado_en",
            ):
                key = str(proj_txt).strip().lower() if proj_txt else ""

                if not key:
                    continue

                access = access_by_key.get(key)

                if not access:
                    continue

                if access["include_history"] or access["start_at"] is None:
                    ids_ok.append(sid)
                    continue

                if not creado_en:
                    continue

                start_at = access["start_at"]

                start_date = (
                    start_at.date() if isinstance(start_at, datetime) else start_at
                )

                if isinstance(creado_en, datetime):
                    creado_date = (
                        timezone.localtime(creado_en).date()
                        if timezone.is_aware(creado_en)
                        else creado_en.date()
                    )
                else:
                    creado_date = creado_en

                if creado_date >= start_date:
                    ids_ok.append(sid)

            qs = qs.filter(id__in=ids_ok)

    # ============================================================
    # Filtros rápidos
    # ============================================================
    f = {
        "tech": (request.GET.get("f_tech") or "").strip(),
        "projid": (request.GET.get("f_projid") or "").strip(),
        "week": (request.GET.get("f_week") or "").strip(),
        "client": (request.GET.get("f_client") or "").strip(),
        "finish": (request.GET.get("f_finish") or "").strip(),
    }

    qs_filtered = qs

    if f["projid"]:
        qs_filtered = qs_filtered.filter(proyecto_id__icontains=f["projid"])

    if f["week"]:
        qs_filtered = qs_filtered.filter(
            Q(semana_pago_proyectada__icontains=f["week"])
            | Q(semana_pago_real__icontains=f["week"])
            | Q(discount_week__icontains=f["week"])
            | Q(pay_week_snapshots__semana_resultado__icontains=f["week"])
            | Q(pay_week_snapshots__semana_base__icontains=f["week"])
        )

    if f["tech"]:
        qs_filtered = qs_filtered.filter(
            Q(tecnicos_sesion__tecnico__first_name__icontains=f["tech"])
            | Q(tecnicos_sesion__tecnico__last_name__icontains=f["tech"])
            | Q(tecnicos_sesion__tecnico__username__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__first_name__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__last_name__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__username__icontains=f["tech"])
        )

    if f["client"]:
        qs_filtered = qs_filtered.filter(cliente__icontains=f["client"])

    if f["finish"]:
        try:
            d_fin = _date.fromisoformat(f["finish"])
            qs_filtered = qs_filtered.filter(finance_finish_date=d_fin)
        except ValueError:
            pass

    qs_filtered = qs_filtered.distinct()

    # ============================================================
    # Helpers livianos para labels
    # ============================================================
    def money_label(n):
        try:
            return f"${float(n or 0):.2f}"
        except Exception:
            return "$0.00"

    def session_status_label_light(s):
        if getattr(s, "is_direct_discount", False):
            return "Direct discount"

        if s.estado == "aprobado_pm":
            return "Approved by PM"

        if s.estado == "rechazado_pm":
            return "Rejected by PM"

        if s.estado == "aprobado_supervisor":
            return "Approved by supervisor"

        if s.estado == "rechazado_supervisor":
            return "Rejected by supervisor"

        if s.estado == "en_revision_supervisor":
            return "In supervisor review"

        if s.estado == "finalizado":
            return "Finished (pending review)"

        if s.estado == "en_proceso":
            return "In progress"

        return "Assigned"

    def finance_status_label_light(s):
        if s.finance_status == "review_discount":
            return "Review discount"

        if s.finance_status == "discount_applied":
            return "Discount applied"

        if s.finance_status == "sent":
            return "Pending to send to client"

        if s.finance_status == "sent_to_client":
            return "Sent to client"

        if s.finance_status == "pending_invoice":
            return "Pending invoicing"

        if s.finance_status == "invoiced":
            return "Invoiced"

        if s.finance_status == "in_review":
            return "In review"

        if s.finance_status == "rejected":
            return "Rejected"

        if s.finance_status == "pending":
            return "Pending payment"

        if s.finance_status == "paid":
            return "Collected"

        return "—"

    def resolve_project_labels_for_sessions(sessions):
        proj_ids = set()
        proj_texts = set()

        for s in sessions:
            raw_proyecto = getattr(s, "proyecto", None)

            if raw_proyecto not in (None, "", "-"):
                txt = str(raw_proyecto).strip()

                if txt:
                    proj_texts.add(txt)

                    try:
                        proj_ids.add(int(txt))
                    except Exception:
                        pass

            raw_proyecto_id = getattr(
                s,
                "proyecto_id",
                None,
            )

            if raw_proyecto_id not in (None, "", "-"):
                txt2 = str(raw_proyecto_id).strip()

                if txt2:
                    proj_texts.add(txt2)

                    try:
                        proj_ids.add(int(txt2))
                    except Exception:
                        pass

        proj_q = Q()

        if proj_ids:
            proj_q |= Q(id__in=proj_ids)

        if proj_texts:
            proj_q |= Q(nombre__in=proj_texts) | Q(codigo__in=proj_texts)

        if proj_q:
            proyectos = Proyecto.objects.filter(proj_q).only(
                "id",
                "nombre",
                "codigo",
            )
        else:
            proyectos = Proyecto.objects.none()

        by_id = {str(p.id): p.nombre for p in proyectos}

        by_code = {
            (p.codigo or "").strip().lower(): p.nombre
            for p in proyectos
            if getattr(p, "codigo", None)
        }

        by_name = {
            (p.nombre or "").strip().lower(): p.nombre
            for p in proyectos
            if getattr(p, "nombre", None)
        }

        for s in sessions:
            raw = str(getattr(s, "proyecto", "") or "").strip()

            raw_id = str(getattr(s, "proyecto_id", "") or "").strip()

            label = ""

            if raw:
                label = (
                    by_id.get(raw)
                    or by_code.get(raw.lower())
                    or by_name.get(raw.lower())
                    or raw
                )

            if not label and raw_id:
                label = (
                    by_id.get(raw_id)
                    or by_code.get(raw_id.lower())
                    or by_name.get(raw_id.lower())
                    or raw_id
                )

            s.project_label = label

        return sessions

    # ============================================================
    # Query liviana para filtros Excel y paginación
    # ============================================================
    light_qs = qs_filtered.only(
        "id",
        "creado_en",
        "proyecto_id",
        "direccion_proyecto",
        "semana_pago_proyectada",
        "estado",
        "is_direct_discount",
        "cliente",
        "ciudad",
        "proyecto",
        "oficina",
        "subtotal_tecnico",
        "subtotal_empresa",
        "finance_daily_number",
        "finance_finish_date",
        "real_company_billing",
        "finance_status",
        "semana_pago_real",
        "discount_week",
        "finance_note",
    )

    light_rows = list(light_qs)

    resolve_project_labels_for_sessions(light_rows)

    light_ids = [s.id for s in light_rows]

    technicians_by_session = {}
    comments_by_session = {}

    assignments_light = (
        SesionBillingTecnico.objects.filter(
            sesion_id__in=light_ids,
        )
        .select_related("tecnico")
        .only(
            "sesion_id",
            "porcentaje",
            "tecnico_comentario",
            "tecnico__first_name",
            "tecnico__last_name",
            "tecnico__username",
        )
    )

    for assignment in assignments_light:
        technician = assignment.tecnico

        technician_name = technician.get_full_name().strip() or technician.username

        technician_label = f"{technician_name} " f"({assignment.porcentaje:.2f}%)"

        technicians_by_session.setdefault(
            assignment.sesion_id,
            [],
        ).append(technician_label)

        comment = (assignment.tecnico_comentario or "").strip()

        if comment:
            comments_by_session.setdefault(
                assignment.sesion_id,
                [],
            ).append(f"{technician_name}: {comment}")

    for s in light_rows:
        s.excel_technicians = technicians_by_session.get(
            s.id,
            [],
        )

        s.excel_comments = comments_by_session.get(
            s.id,
            [],
        )

        finance_note = (getattr(s, "finance_note", "") or "").strip()

        if finance_note:
            s.excel_comments = list(s.excel_comments)

            s.excel_comments.append(f"Finance note: {finance_note}")

    def excel_value_for_invoice_light(s, col):
        if col == "0":
            d = getattr(s, "creado_en", None)

            return d.strftime("%Y-%m-%d %H:%M") if d else ""

        if col == "1":
            return str(getattr(s, "proyecto_id", "") or "")

        if col == "2":
            return str(
                getattr(
                    s,
                    "direccion_proyecto",
                    "",
                )
                or ""
            )

        if col == "3":
            return str(
                getattr(
                    s,
                    "semana_pago_proyectada",
                    "",
                )
                or "—"
            )

        if col == "4":
            return session_status_label_light(s)

        if col == "5":
            values = (
                getattr(
                    s,
                    "excel_technicians",
                    [],
                )
                or []
            )

            return " | ".join(values) if values else "—"

        if col == "6":
            return str(getattr(s, "cliente", "") or "")

        if col == "7":
            return str(getattr(s, "ciudad", "") or "")

        if col == "8":
            return str(
                getattr(
                    s,
                    "project_label",
                    "",
                )
                or ""
            )

        if col == "9":
            return str(getattr(s, "oficina", "") or "")

        if col == "10":
            return money_label(
                getattr(
                    s,
                    "subtotal_tecnico",
                    0,
                )
            )

        if col == "11":
            return money_label(
                getattr(
                    s,
                    "subtotal_empresa",
                    0,
                )
            )

        if col == "12":
            return str(
                getattr(
                    s,
                    "finance_daily_number",
                    "",
                )
                or "—"
            )

        if col == "13":
            d = getattr(
                s,
                "finance_finish_date",
                None,
            )

            return d.strftime("%Y-%m-%d") if d else "—"

        if col == "14":
            real = getattr(
                s,
                "real_company_billing",
                None,
            )

            return "—" if real is None else money_label(real)

        if col == "15":
            real = getattr(
                s,
                "real_company_billing",
                None,
            )

            subtotal = getattr(
                s,
                "subtotal_empresa",
                None,
            )

            if real is None or subtotal is None:
                return "—"

            try:
                diff = float(subtotal or 0) - float(real or 0)
            except Exception:
                return "—"

            if diff > 0:
                return f"+ ${abs(diff):.2f}"

            if diff < 0:
                return f"- ${abs(diff):.2f}"

            return "$0.00"

        if col == "16":
            return finance_status_label_light(s)

        if col == "17":
            week = (
                (
                    getattr(
                        s,
                        "semana_pago_real",
                        "",
                    )
                    or ""
                ).strip()
                or (
                    getattr(
                        s,
                        "discount_week",
                        "",
                    )
                    or ""
                ).strip()
                or (
                    getattr(
                        s,
                        "semana_pago_proyectada",
                        "",
                    )
                    or ""
                ).strip()
                or "—"
            )

            return week

        if col == "18":
            values = (
                getattr(
                    s,
                    "excel_comments",
                    [],
                )
                or []
            )

            return " | ".join(values) if values else "—"

        return ""

    def excel_values_for_invoice_light(s, col):
        if col == "5":
            return getattr(
                s,
                "excel_technicians",
                [],
            ) or ["—"]

        if col == "18":
            return getattr(
                s,
                "excel_comments",
                [],
            ) or ["—"]

        return [
            excel_value_for_invoice_light(
                s,
                col,
            )
        ]

    # ============================================================
    # Filtros Excel
    # ============================================================
    excel_filters_raw = (request.GET.get("excel_filters") or "").strip()

    try:
        excel_filters = json.loads(excel_filters_raw) if excel_filters_raw else {}
    except json.JSONDecodeError:
        excel_filters = {}

    if excel_filters:
        filtered_light_rows = []

        for s in light_rows:
            ok = True

            for col, values in excel_filters.items():
                values_set = set(values or [])

                if not values_set:
                    continue

                invoice_values = set(
                    excel_values_for_invoice_light(
                        s,
                        col,
                    )
                )

                if values_set.isdisjoint(invoice_values):
                    ok = False
                    break

            if ok:
                filtered_light_rows.append(s)

        light_rows = filtered_light_rows

    # ============================================================
    # Opciones globales para filtros Excel
    # ============================================================
    excel_global = {}

    for col in range(19):
        vals = set()

        for s in light_rows:
            invoice_values = excel_values_for_invoice_light(
                s,
                str(col),
            )

            for value in invoice_values:
                vals.add(value or "(Vacías)")

        excel_global[col] = sorted(vals)

    excel_global_json = json.dumps(excel_global)

    # ============================================================
    # Paginación sobre IDs
    # ============================================================
    raw_cantidad = request.GET.get(
        "cantidad",
        "10",
    )

    try:
        per_page = int(raw_cantidad)
    except Exception:
        per_page = 10

    if per_page < 5:
        per_page = 5

    if per_page > 100:
        per_page = 100

    cantidad = str(per_page)

    filtered_ids = [s.id for s in light_rows]

    paginator = Paginator(
        filtered_ids,
        per_page,
    )

    pagina_ids = paginator.get_page(request.GET.get("page"))

    page_ids = list(pagina_ids.object_list)

    order_map = {pk: idx for idx, pk in enumerate(page_ids)}

    # ============================================================
    # Cargar relaciones pesadas SOLO para la página visible
    # ============================================================
    page_rows = list(
        SesionBilling.objects.filter(id__in=page_ids).prefetch_related(
            Prefetch(
                "items",
                queryset=ItemBilling.objects.prefetch_related(
                    Prefetch(
                        "desglose_tecnico",
                        queryset=ItemBillingTecnico.objects.select_related("tecnico"),
                    )
                ),
            ),
            Prefetch(
                "tecnicos_sesion",
                queryset=SesionBillingTecnico.objects.select_related(
                    "tecnico"
                ).prefetch_related(
                    Prefetch(
                        "evidencias",
                        queryset=EvidenciaFotoBilling.objects.only(
                            "id",
                            "imagen",
                            "tecnico_sesion_id",
                            "requisito_id",
                        ).order_by("-id"),
                    )
                ),
            ),
            Prefetch(
                "pay_week_snapshots",
                queryset=BillingPayWeekSnapshot.objects.select_related(
                    "tecnico",
                    "item",
                )
                .filter(is_adjustment=False)
                .order_by(
                    "tecnico__first_name",
                    "tecnico__last_name",
                    "tecnico__username",
                    "tipo_trabajo",
                    "codigo_trabajo",
                    "id",
                ),
            ),
        )
    )

    page_rows.sort(
        key=lambda s: order_map.get(
            s.id,
            999999,
        )
    )

    resolve_project_labels_for_sessions(page_rows)

    # ============================================================
    # Helpers que usan relaciones, solo para página visible
    # ============================================================
    def legacy_paid_flag(s):
        note = getattr(s, "finance_note", "") or ""

        try:
            tech_ids = list(
                s.tecnicos_sesion.all().values_list(
                    "tecnico_id",
                    flat=True,
                )
            )
        except Exception:
            tech_ids = []

        possible_weeks = [
            (
                getattr(
                    s,
                    "semana_pago_real",
                    "",
                )
                or ""
            )
            .strip()
            .upper(),
            (
                getattr(
                    s,
                    "semana_pago_proyectada",
                    "",
                )
                or ""
            )
            .strip()
            .upper(),
            (
                getattr(
                    s,
                    "discount_week",
                    "",
                )
                or ""
            )
            .strip()
            .upper(),
        ]

        possible_weeks = [w for w in possible_weeks if w]

        for tech_id in tech_ids:
            for wk in possible_weeks:
                marker = f"[TECH_WEEKLY_PAYMENT_PAID:" f"{tech_id}:{wk}]"

                if marker in note:
                    return True

        return False

    def build_payweek_groups(s):
        groups_map = {}

        snaps = (
            list(
                getattr(
                    s,
                    "pay_week_snapshots",
                    [],
                ).all()
            )
            if hasattr(
                s,
                "pay_week_snapshots",
            )
            else []
        )

        if snaps:
            for snap in snaps:
                tech_name = (
                    snap.tecnico.get_full_name().strip()
                    if (
                        getattr(
                            snap,
                            "tecnico",
                            None,
                        )
                        and snap.tecnico.get_full_name()
                    )
                    else getattr(
                        snap.tecnico,
                        "username",
                        "",
                    )
                    or f"User {snap.tecnico_id}"
                )

                grp = groups_map.setdefault(
                    tech_name,
                    {
                        "tech_name": tech_name,
                        "weeks_summary": "",
                        "lines": [],
                    },
                )

                work_type = (
                    (snap.tipo_trabajo or "").strip()
                    or (
                        getattr(
                            snap.item,
                            "tipo_trabajo",
                            "",
                        )
                        or ""
                    ).strip()
                    or "Legacy"
                )

                week = (
                    (snap.semana_resultado or "").strip()
                    or (snap.semana_base or "").strip()
                    or (
                        getattr(
                            s,
                            "semana_pago_real",
                            "",
                        )
                        or ""
                    ).strip()
                    or (
                        getattr(
                            s,
                            "discount_week",
                            "",
                        )
                        or ""
                    ).strip()
                    or (
                        getattr(
                            s,
                            "semana_pago_proyectada",
                            "",
                        )
                        or ""
                    ).strip()
                    or "—"
                )

                is_paid_line = bool(
                    getattr(
                        snap,
                        "paid_at",
                        None,
                    )
                    or getattr(
                        snap,
                        "is_paid",
                        False,
                    )
                )

                grp["lines"].append(
                    {
                        "work_type": work_type,
                        "codigo_trabajo": (snap.codigo_trabajo or "").strip(),
                        "week": week,
                        "is_legacy": False,
                        "snapshot_id": snap.id,
                        "is_paid": is_paid_line,
                    }
                )

            groups = list(groups_map.values())

            for grp in groups:
                weeks = []

                for line in grp["lines"]:
                    wk = (line.get("week") or "").strip()

                    if wk and wk not in weeks:
                        weeks.append(wk)

                grp["weeks_summary"] = ", ".join(weeks) if weeks else "—"

            return groups

        asignaciones = (
            list(s.tecnicos_sesion.all())
            if hasattr(
                s,
                "tecnicos_sesion",
            )
            else []
        )

        base_week = (
            (
                getattr(
                    s,
                    "semana_pago_real",
                    "",
                )
                or ""
            ).strip()
            or (
                getattr(
                    s,
                    "discount_week",
                    "",
                )
                or ""
            ).strip()
            or (
                getattr(
                    s,
                    "semana_pago_proyectada",
                    "",
                )
                or ""
            ).strip()
            or "—"
        )

        legacy_is_paid = legacy_paid_flag(s)

        tech_names = []

        for asig in asignaciones:
            tech_name = (
                asig.tecnico.get_full_name().strip()
                if (
                    getattr(
                        asig,
                        "tecnico",
                        None,
                    )
                    and asig.tecnico.get_full_name()
                )
                else getattr(
                    asig.tecnico,
                    "username",
                    "",
                )
                or f"User {asig.tecnico_id}"
            )

            if tech_name and tech_name not in tech_names:
                tech_names.append(tech_name)

        tech_label = ", ".join(tech_names) if tech_names else "—"

        return [
            {
                "tech_name": tech_label,
                "weeks_summary": base_week,
                "lines": [
                    {
                        "work_type": "Legacy",
                        "codigo_trabajo": "",
                        "week": base_week,
                        "is_legacy": True,
                        "session_id": s.id,
                        "dom_id": f"{s.id}-legacy",
                        "is_paid": legacy_is_paid,
                    }
                ],
            }
        ]

    def payweek_snapshot_label(s):
        groups = build_payweek_groups(s)

        if not groups:
            return str(
                getattr(
                    s,
                    "semana_pago_real",
                    "",
                )
                or "—"
            )

        rows = []

        for grp in groups:
            tech_name = grp.get("tech_name") or "—"

            for line in grp.get(
                "lines",
                [],
            ):
                work_type = (line.get("work_type") or "").strip() or "Work type"

                week = (line.get("week") or "").strip() or "—"

                suffix = " [Paid]" if line.get("is_paid") else ""

                rows.append(f"{tech_name} — " f"{work_type} → " f"{week}{suffix}")

        return (
            " | ".join(rows)
            if rows
            else str(
                getattr(
                    s,
                    "semana_pago_real",
                    "",
                )
                or "—"
            )
        )

    # ============================================================
    # Preparar objetos visibles para template
    # ============================================================
    for s in page_rows:
        comentarios = []

        try:
            for st in s.tecnicos_sesion.all():
                txt = (
                    getattr(
                        st,
                        "tecnico_comentario",
                        "",
                    )
                    or ""
                ).strip()

                if txt:
                    comentarios.append(st)
        except Exception:
            pass

        s.comentarios_tecnicos = comentarios
        s.payweek_groups = build_payweek_groups(s)
        s.payweek_snapshot_label = payweek_snapshot_label(s)

    pagina_ids.object_list = page_rows
    pagina = pagina_ids

    # ============================================================
    # Mantener querystring
    # ============================================================
    keep_params = {
        "scope": scope,
        "cantidad": cantidad,
        "f_tech": f["tech"],
        "f_projid": f["projid"],
        "f_week": f["week"],
        "f_client": f["client"],
        "f_finish": f["finish"],
    }

    if excel_filters_raw:
        keep_params["excel_filters"] = excel_filters_raw

    qs_keep = urlencode({k: v for k, v in keep_params.items() if v not in ("", None)})

    return render(
        request,
        "facturacion/invoices_list.html",
        {
            "pagina": pagina,
            "cantidad": cantidad,
            "scope": scope,
            "can_edit_real_week": _can_edit_real_week(request.user),
            "f_tech": f["tech"],
            "f_projid": f["projid"],
            "f_week_input": f["week"],
            "f_client": f["client"],
            "f_finish": f["finish"],
            "qs_keep": qs_keep,
            "excel_global_json": excel_global_json,
        },
    )


@login_required
@rol_requerido("facturacion", "admin")
@require_POST
def invoice_update_real(request, pk):
    """
    Actualiza campos financieros del invoice.

    Flujo de Real Company Billing:

        pending_invoice
            ↓ al ingresar monto
        invoiced

    Restricciones:

    - Mientras esté en "sent", primero debe marcarse como enviado al cliente.
    - Mientras esté en "sent_to_client", primero debe pasar a Pending invoicing.
    - Los registros históricos en "pending" pueden seguir editándose.
    - Los registros cobrados no pueden modificar Real Company Billing.
    """
    from decimal import Decimal, InvalidOperation

    from django.db import transaction
    from django.http import HttpResponseForbidden, JsonResponse
    from django.shortcuts import get_object_or_404

    from operaciones.models import SesionBilling

    if request.headers.get("x-requested-with") != "XMLHttpRequest":
        return HttpResponseForbidden("AJAX only")

    real_raw = request.POST.get("real", None)
    week_raw = request.POST.get("week", None)
    daily_raw = request.POST.get("daily_number", None)

    with transaction.atomic():
        s = get_object_or_404(
            SesionBilling.objects.select_for_update(),
            pk=pk,
        )

        updated_fields = []

        # ========================================================
        # Real Company Billing
        # ========================================================
        if real_raw is not None:
            current_status = (getattr(s, "finance_status", "") or "").strip()

            if current_status == "paid":
                return JsonResponse(
                    {
                        "ok": False,
                        "error": (
                            "This invoice is already collected. "
                            "Real Company Billing cannot be modified."
                        ),
                    },
                    status=409,
                )

            if current_status == "sent":
                return JsonResponse(
                    {
                        "ok": False,
                        "error": (
                            "Mark the invoice as sent to the client "
                            "before entering Real Company Billing."
                        ),
                    },
                    status=409,
                )

            if current_status == "sent_to_client":
                return JsonResponse(
                    {
                        "ok": False,
                        "error": (
                            "Mark the invoice as Pending invoicing "
                            "before entering Real Company Billing."
                        ),
                    },
                    status=409,
                )

            allowed_real_statuses = {
                "pending_invoice",
                "invoiced",
                "pending",
                "in_review",
                "rejected",
            }

            if current_status not in allowed_real_statuses:
                return JsonResponse(
                    {
                        "ok": False,
                        "error": (
                            "Real Company Billing cannot be entered "
                            "from the current finance status."
                        ),
                        "current_status": current_status,
                    },
                    status=409,
                )

            raw = (real_raw or "").strip()

            if raw in ("", "-", "—", "null", "None"):
                s.real_company_billing = None
                updated_fields.append("real_company_billing")

                if current_status == "invoiced":
                    s.finance_status = "pending_invoice"
                    updated_fields.append("finance_status")

            else:
                txt = raw.replace("$", "").replace(",", "").replace(" ", "")

                try:
                    amount = Decimal(txt)
                except (InvalidOperation, ValueError):
                    return JsonResponse(
                        {
                            "ok": False,
                            "error": "Invalid amount.",
                        },
                        status=400,
                    )

                if amount < Decimal("0"):
                    return JsonResponse(
                        {
                            "ok": False,
                            "error": ("Real Company Billing cannot be negative."),
                        },
                        status=400,
                    )

                s.real_company_billing = amount
                updated_fields.append("real_company_billing")

                if current_status == "pending_invoice":
                    s.finance_status = "invoiced"
                    updated_fields.append("finance_status")

        # ========================================================
        # Real pay week
        # ========================================================
        if week_raw is not None:
            s.semana_pago_real = (week_raw or "").strip()
            updated_fields.append("semana_pago_real")

        # ========================================================
        # Daily Number
        # ========================================================
        if daily_raw is not None:
            value = (daily_raw or "").strip()
            s.finance_daily_number = value or None
            updated_fields.append("finance_daily_number")

        if updated_fields:
            updated_fields = list(dict.fromkeys(updated_fields))

            if hasattr(s, "finance_updated_at"):
                updated_fields.append("finance_updated_at")

            s.save(update_fields=updated_fields)

    difference = None

    if s.real_company_billing is not None:
        difference = (s.subtotal_empresa or Decimal("0")) - s.real_company_billing

    return JsonResponse(
        {
            "ok": True,
            "real": (
                None
                if s.real_company_billing is None
                else f"{s.real_company_billing:.2f}"
            ),
            "week": s.semana_pago_real or "",
            "daily_number": s.finance_daily_number or "",
            "difference": ("" if difference is None else f"{difference:.2f}"),
            "finance_status": s.finance_status,
        }
    )


@login_required
@rol_requerido("facturacion", "admin")
@require_POST
def invoice_mark_paid(request, pk: int):
    """
    Marca un invoice como Collected.

    Requisitos:

    - Debe tener Real Company Billing.
    - Si se está cobrando menos de lo esperado, pide confirmación.
    - Permite cobrar directamente cuando el monto ya está registrado.
    - Mantiene compatibilidad con registros históricos en "pending".
    """
    from decimal import Decimal

    from django.db import transaction
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from django.utils import timezone

    from operaciones.models import SesionBilling

    with transaction.atomic():
        s = get_object_or_404(
            SesionBilling.objects.select_for_update(),
            pk=pk,
        )

        if s.finance_status == "paid":
            return JsonResponse(
                {
                    "ok": True,
                    "message": "This invoice is already collected.",
                    "finance_status": "paid",
                    "finance_status_label": "Collected",
                }
            )

        if s.real_company_billing is None:
            return JsonResponse(
                {
                    "ok": False,
                    "error": (
                        "Real Company Billing is required before "
                        "marking the invoice as collected."
                    ),
                },
                status=400,
            )

        difference = (s.subtotal_empresa or Decimal("0")) - s.real_company_billing

        force = (request.POST.get("force") or "") == "1"

        if difference > 0 and not force:
            return JsonResponse(
                {
                    "ok": False,
                    "confirm": True,
                    "message": (
                        "You are collecting less than expected. "
                        "Do you still want to mark it as collected?"
                    ),
                },
                status=409,
            )

        s.finance_status = "paid"

        if not s.finance_finish_date:
            s.finance_finish_date = timezone.localdate()

        update_fields = [
            "finance_status",
            "finance_finish_date",
        ]

        if hasattr(s, "finance_updated_at"):
            s.finance_updated_at = timezone.now()
            update_fields.append("finance_updated_at")

        s.save(update_fields=update_fields)

    return JsonResponse(
        {
            "ok": True,
            "message": "Invoice marked as collected.",
            "finance_status": "paid",
            "finance_status_label": "Collected",
        }
    )


@login_required
@rol_requerido("facturacion", "admin")
def invoice_reject(request, pk: int):
    """
    Reject invoice back to Operations with a reason.
    Sets finance_status='rejected' and stores the note.
    """
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    reason = (request.POST.get("reason") or "").strip()
    if not reason:
        return HttpResponseBadRequest("A rejection reason is required.")

    s = get_object_or_404(SesionBilling, pk=pk)
    with transaction.atomic():
        s.finance_status = "rejected"
        s.finance_note = reason
        s.save(update_fields=["finance_status", "finance_note"])

    return JsonResponse({"ok": True})


@login_required
@rol_requerido("facturacion", "admin")
@require_POST
@transaction.atomic
def invoice_remove(request, pk: int):
    """
    Saca la sesión de la cola de Finanzas (NO borra el billing).

    Reglas:
    - Si es descuento directo -> vuelve a 'review_discount'
    - Si NO es descuento directo -> vuelve a 'none'
    - No toca status operativo
    - No borra markers legacy de pago técnico
    """
    s = (
        SesionBilling.objects.only(
            "id",
            "is_direct_discount",
            "finance_status",
            "finance_note",
        )
        .filter(pk=pk)
        .first()
    )

    if not s:
        if request.headers.get("x-requested-with") == "XMLHttpRequest":
            return JsonResponse({"ok": False, "error": "Not found"}, status=404)
        messages.error(request, "Billing not found.")
        return redirect(
            request.META.get("HTTP_REFERER") or reverse("facturacion:invoices")
        )

    new_status = "review_discount" if s.is_direct_discount else "none"

    # Preservar markers técnicos legacy
    old_note = getattr(s, "finance_note", "") or ""
    keep_lines = []

    for ln in old_note.splitlines():
        txt = (ln or "").strip()
        if txt.startswith("[TECH_WEEKLY_PAYMENT_PAID:") and txt.endswith("]"):
            keep_lines.append(txt)

    preserved_note = "\n".join(keep_lines).strip()

    SesionBilling.objects.filter(pk=s.pk).update(
        finance_status=new_status,
        finance_note=preserved_note,
        finance_sent_at=None,
        finance_finish_date=None,
        finance_updated_at=timezone.now(),
    )

    if request.headers.get("x-requested-with") == "XMLHttpRequest":
        return JsonResponse({"ok": True, "id": s.pk, "finance_status": new_status})

    if new_status == "review_discount":
        messages.success(request, f"Direct discount #{s.pk} returned to Operations.")
    else:
        messages.success(request, f"Billing #{s.pk} removed from Finance queue.")

    return redirect(request.META.get("HTTP_REFERER") or reverse("facturacion:invoices"))


@rol_requerido('facturacion', 'admin', 'pm')
@require_POST
def invoice_discount_verified(request, pk):
    s = get_object_or_404(SesionBilling, pk=pk)

    if not s.is_direct_discount:
        return JsonResponse({"ok": False, "error": "NOT_DIRECT_DISCOUNT"}, status=400)

    note = (request.POST.get("note") or "").strip()

    # Primer estado de descuentos → discount_applied
    s.finance_status = "discount_applied"
    if note:
        # apendea (igual que en otros flujos)
        prefix = f"{timezone.now():%Y-%m-%d %H:%M} Finance: "
        s.finance_note = ((s.finance_note + "\n")
                          if s.finance_note else "") + prefix + note

    s.save(update_fields=[
        "finance_status",
        *(["finance_note"] if note else []),
        "finance_updated_at",  # auto_now se actualiza
    ])

    return JsonResponse({"ok": True, "finance_status": s.finance_status})


def _to_excel_dt(value):
    if not value:
        return ""
    try:
        return timezone.make_naive(value) if timezone.is_aware(value) else value
    except Exception:
        return value


@rol_requerido("facturacion", "admin", "pm")
def invoices_export(request):
    """
    
Exporta a Excel los invoices de Finanzas.

- Respeta el scope: open | paid | all.
- Incluye los estados:
    * sent -> Pending to send to client
    * sent_to_client -> Sent to client
    * pending_invoice -> Pending invoicing
    * invoiced -> Invoiced
    * pending -> Pending payment (legacy)
    * paid -> Collected
- Mantiene el orden:
    1. Pending to send to client
    2. Sent to client
    3. Pending invoicing
    4. Invoiced / Pending payment legacy
    5. Otros estados

    """
    from datetime import datetime
    from decimal import Decimal

    from django.db.models import Case, IntegerField, Prefetch, Q, When
    from django.http import HttpResponse
    from django.utils import timezone
    from openpyxl import Workbook

    from facturacion.models import Proyecto
    from operaciones.models import (EvidenciaFotoBilling, ItemBilling,
                                    ItemBillingTecnico, SesionBilling,
                                    SesionBillingTecnico)

    user = request.user
    scope = (request.GET.get("scope") or "open").strip()

    # ============================================================
    # Usuarios privilegiados con acceso al historial completo
    # ============================================================
    can_view_legacy_history = user.is_superuser or getattr(
        user,
        "es_usuario_historial",
        False,
    )

    # ============================================================
    # Proyectos visibles para el usuario
    # Misma lógica utilizada por invoices_list
    # ============================================================
    try:
        proyectos_user = filter_queryset_by_access(
            Proyecto.objects.all(),
            user,
            "id",
        )
    except Exception:
        proyectos_user = Proyecto.objects.none()

    proyectos_list = list(proyectos_user)

    if proyectos_list:
        allowed_keys = set()

        for p in proyectos_list:
            nombre = (getattr(p, "nombre", "") or "").strip()
            if nombre:
                allowed_keys.add(nombre)

            codigo = getattr(p, "codigo", None)
            if codigo:
                allowed_keys.add(str(codigo).strip())

            allowed_keys.add(str(p.id).strip())
    else:
        allowed_keys = set()

    # ============================================================
    # Query base
    #
    # Orden:
    #   sent             -> Pending to send to client
    #   sent_to_client   -> Sent to client
    #   pending          -> Pending payment
    #   otros estados
    #   paid             -> Collected
    # ============================================================
    qs = (
        SesionBilling.objects.annotate(
            finance_status_order=Case(
                When(finance_status="sent", then=0),
                When(finance_status="sent_to_client", then=1),
                When(finance_status="pending_invoice", then=2),
                When(finance_status="invoiced", then=3),
                When(finance_status="pending", then=3),
                When(finance_status="discount_applied", then=4),
                When(finance_status="review_discount", then=4),
                When(finance_status="in_review", then=4),
                When(finance_status="rejected", then=4),
                When(finance_status="paid", then=5),
                default=4,
                output_field=IntegerField(),
            )
        )
        .prefetch_related(
            Prefetch(
                "items",
                queryset=ItemBilling.objects.prefetch_related(
                    Prefetch(
                        "desglose_tecnico",
                        queryset=ItemBillingTecnico.objects.select_related("tecnico"),
                    )
                ),
            ),
            Prefetch(
                "tecnicos_sesion",
                queryset=SesionBillingTecnico.objects.select_related(
                    "tecnico"
                ).prefetch_related(
                    Prefetch(
                        "evidencias",
                        queryset=EvidenciaFotoBilling.objects.only(
                            "id",
                            "imagen",
                            "tecnico_sesion_id",
                            "requisito_id",
                        ).order_by("-id"),
                    )
                ),
            ),
        )
        .order_by(
            "finance_status_order",
            "-creado_en",
        )
    )

    # ============================================================
    # Alcance financiero: open | paid | all
    #
    # IMPORTANTE:
    # paid también está dentro de open para que Collected permanezca
    # en la tabla y también aparezca en el export principal.
    # ============================================================
    finance_open_base = [
        "discount_applied",
        "sent",
        "sent_to_client",
        "pending_invoice",
        "invoiced",
        "in_review",
        "pending",
        "rejected",
        "paid",
    ]

    if scope == "paid":
        qs = qs.filter(finance_status="paid")

    elif scope == "all":
        qs = qs.exclude(
            Q(finance_status__in=["none", ""])
            | Q(finance_status__isnull=True)
            | (Q(finance_status="review_discount") & Q(finance_sent_at__isnull=True))
        )

    else:
        qs = qs.filter(
            Q(finance_status__in=finance_open_base)
            | (Q(finance_status="review_discount") & Q(finance_sent_at__isnull=False))
        )

    qs = qs.filter(
        Q(estado__in=["aprobado_supervisor", "aprobado_pm"])
        | Q(is_direct_discount=True)
    )

    # ============================================================
    # Limitar por proyectos asignados para usuarios sin historial
    # ============================================================
    if not can_view_legacy_history:
        if allowed_keys:
            qs = qs.filter(proyecto__in=allowed_keys)
        else:
            qs = SesionBilling.objects.none()

    # ============================================================
    # Limitar también por ventana histórica de ProyectoAsignacion
    #
    # include_history=True:
    #   ve todo el historial del proyecto.
    #
    # start_at:
    #   ve registros desde la fecha de asignación.
    # ============================================================
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    if ProyectoAsignacion is not None and not can_view_legacy_history:
        try:
            asignaciones = list(
                ProyectoAsignacion.objects.filter(
                    usuario=user,
                    proyecto__in=proyectos_user,
                ).select_related("proyecto")
            )
        except Exception:
            asignaciones = []

        if asignaciones:
            access_by_key = {}

            for asignacion in asignaciones:
                proyecto = getattr(asignacion, "proyecto", None)

                if not proyecto:
                    continue

                if getattr(asignacion, "include_history", False) or not getattr(
                    asignacion,
                    "start_at",
                    None,
                ):
                    access = {
                        "include_history": True,
                        "start_at": None,
                    }
                else:
                    access = {
                        "include_history": False,
                        "start_at": asignacion.start_at,
                    }

                project_keys = (
                    getattr(proyecto, "nombre", None),
                    getattr(proyecto, "codigo", None),
                    getattr(proyecto, "id", None),
                )

                for key in project_keys:
                    if key is None:
                        continue

                    normalized_key = str(key).strip()

                    if normalized_key:
                        access_by_key[normalized_key.lower()] = access

            ids_ok = []

            for session_id, proyecto_texto, creado_en in qs.values_list(
                "id",
                "proyecto",
                "creado_en",
            ):
                project_key = (
                    str(proyecto_texto).strip().lower() if proyecto_texto else ""
                )

                if not project_key:
                    continue

                access = access_by_key.get(project_key)

                if not access:
                    continue

                if access["include_history"] or access["start_at"] is None:
                    ids_ok.append(session_id)
                    continue

                if not creado_en:
                    continue

                start_at = access["start_at"]

                if isinstance(start_at, datetime):
                    start_date = start_at.date()
                else:
                    start_date = start_at

                if isinstance(creado_en, datetime):
                    if timezone.is_aware(creado_en):
                        creado_date = timezone.localtime(creado_en).date()
                    else:
                        creado_date = creado_en.date()
                else:
                    creado_date = creado_en

                if creado_date >= start_date:
                    ids_ok.append(session_id)

            qs = qs.filter(id__in=ids_ok)

    qs = qs.distinct()

    # ============================================================
    # Mapas para resolver el nombre de Proyecto
    # ============================================================
    if can_view_legacy_history:
        proyectos_list = list(Proyecto.objects.all())

    by_id = {p.id: p for p in proyectos_list}

    by_code = {
        (p.codigo or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "codigo", None)
    }

    by_name = {
        (p.nombre or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "nombre", None)
    }

    def resolve_project_label(session):
        """
        Devuelve el nombre legible del proyecto.

        Intenta resolver SesionBilling.proyecto y proyecto_id mediante:
        - ID de Proyecto.
        - Código.
        - Nombre.
        - Texto original como fallback.
        """
        project_text = str(getattr(session, "proyecto", "") or "").strip()

        project_id_value = getattr(
            session,
            "proyecto_id",
            None,
        )

        selected_project = None

        if project_text:
            try:
                numeric_project_id = int(project_text)
            except (TypeError, ValueError):
                normalized_project_text = project_text.lower()

                selected_project = by_code.get(normalized_project_text) or by_name.get(
                    normalized_project_text
                )
            else:
                selected_project = by_id.get(numeric_project_id)

        if not selected_project and project_id_value not in (None, "", "-"):
            try:
                numeric_project_id = int(project_id_value)
            except (TypeError, ValueError):
                normalized_project_id = str(project_id_value).strip().lower()

                selected_project = by_code.get(normalized_project_id) or by_name.get(
                    normalized_project_id
                )
            else:
                selected_project = by_id.get(numeric_project_id)

        if selected_project:
            return getattr(
                selected_project,
                "nombre",
                str(selected_project),
            )

        if project_text:
            return project_text

        if project_id_value not in (None, "", "-"):
            return str(project_id_value)

        return ""

    # ============================================================
    # Labels de estados
    # ============================================================
    status_map = {
        "aprobado_pm": "Approved by PM",
        "rechazado_pm": "Rejected by PM",
        "aprobado_supervisor": "Approved by supervisor",
        "rechazado_supervisor": "Rejected by supervisor",
        "en_revision_supervisor": "In supervisor review",
        "finalizado": "Finished (pending review)",
        "en_proceso": "In progress",
        "asignado": "Assigned",
    }

    finance_map = {
        "none": "—",
        "review_discount": "Review discount",
        "discount_applied": "Discount applied",
        "sent": "Pending to send to client",
        "sent_to_client": "Sent to client",
        "pending_invoice": "Pending invoicing",
        "invoiced": "Invoiced",
        "pending": "Pending payment",
        "in_review": "In review",
        "rejected": "Rejected",
        "paid": "Collected",
    }

    def techs_string(session):
        parts = []

        for session_technician in session.tecnicos_sesion.all():
            technician = session_technician.tecnico

            if technician:
                technician_name = technician.get_full_name() or technician.username
            else:
                technician_name = "—"

            parts.append(
                f"{technician_name} " f"({session_technician.porcentaje:.2f}%)"
            )

        return ", ".join(parts)

    def comments_string(session):
        parts = []

        for session_technician in session.tecnicos_sesion.all():
            comment = (
                getattr(
                    session_technician,
                    "tecnico_comentario",
                    "",
                )
                or ""
            ).strip()

            if not comment:
                continue

            technician = getattr(
                session_technician,
                "tecnico",
                None,
            )

            if technician:
                technician_name = technician.get_full_name() or technician.username
            else:
                technician_name = "—"

            parts.append(f"{technician_name}: {comment}")

        return "\n".join(parts)

    # ============================================================
    # Crear Excel
    # ============================================================
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Invoices"

    headers = [
        "Date",
        "Project ID",
        "Project address",
        "Projected week",
        "Status",
        "Technicians",
        "Client",
        "City",
        "Project",
        "Office",
        "Technical Billing",
        "Company Billing",
        "Daily Number",
        "Finish date",
        "Real Company Billing",
        "Difference",
        "Finance status",
        "Finance note",
        "Pay week / Discount week",
        "Comment",
        "Job Code",
        "Work Type",
        "Description",
        "UOM",
        "Quantity",
        "Technical Rate",
        "Company Rate",
        "Subtotal Technical",
        "Subtotal Company",
    ]

    worksheet.append(headers)

    for session in qs:
        if getattr(session, "is_direct_discount", False):
            status_label = "Direct discount"
        else:
            status_label = status_map.get(
                session.estado,
                "Assigned",
            )

        finance_label = finance_map.get(
            session.finance_status,
            "—",
        )

        real_week = getattr(session, "semana_pago_real", "") or ""

        discount_week = (
            getattr(session, "discount_week", "")
            or getattr(session, "semana_descuento", "")
            or ""
        )

        if real_week and discount_week:
            pay_or_discount_week = f"{real_week} / {discount_week}"
        else:
            pay_or_discount_week = real_week or discount_week

        project_label = resolve_project_label(session)
        comment_text = comments_string(session)

        head_common = [
            _to_excel_dt(session.creado_en),
            session.proyecto_id,
            session.direccion_proyecto,
            session.semana_pago_proyectada or "",
            status_label,
            techs_string(session),
            session.cliente or "",
            session.ciudad or "",
            project_label or "",
            session.oficina or "",
            float(session.subtotal_tecnico or 0),
            float(session.subtotal_empresa or 0),
            session.finance_daily_number or "",
            (
                _to_excel_dt(session.finance_finish_date)
                if getattr(
                    session,
                    "finance_finish_date",
                    None,
                )
                else ""
            ),
            (
                float(session.real_company_billing)
                if session.real_company_billing is not None
                else 0.0
            ),
            float(session.diferencia or 0),
            finance_label,
            session.finance_note or "",
            pay_or_discount_week,
            comment_text,
        ]

        items_manager = getattr(
            session,
            "items",
            None,
        )

        items = items_manager.all() if items_manager is not None else []

        if not items:
            worksheet.append(
                head_common
                + [
                    "",
                    "",
                    "",
                    "",
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                    0.0,
                ]
            )
            continue

        for item in items:
            total_quantity = Decimal(str(item.cantidad or 0))

            company_rate = Decimal(str(item.precio_empresa or 0))

            technical_breakdown_manager = getattr(
                item,
                "desglose_tecnico",
                None,
            )

            technical_breakdowns = (
                list(technical_breakdown_manager.all())
                if technical_breakdown_manager is not None
                else []
            )

            if technical_breakdowns:
                for breakdown in technical_breakdowns:
                    percentage = Decimal(str(breakdown.porcentaje or 0)) / Decimal(
                        "100"
                    )

                    technician_quantity = total_quantity * percentage

                    technical_rate = Decimal(
                        str(
                            getattr(
                                breakdown,
                                "tarifa_base",
                                0,
                            )
                            or 0
                        )
                    )

                    technical_subtotal = technical_rate * technician_quantity

                    company_subtotal = company_rate * technician_quantity

                    worksheet.append(
                        head_common
                        + [
                            item.codigo_trabajo or "",
                            item.tipo_trabajo or "",
                            item.descripcion or "",
                            item.unidad_medida or "",
                            float(technician_quantity),
                            float(technical_rate),
                            float(company_rate),
                            float(technical_subtotal),
                            float(company_subtotal),
                        ]
                    )

            else:
                technical_subtotal = Decimal(str(item.subtotal_tecnico or 0))

                company_subtotal = Decimal(
                    str(item.subtotal_empresa or (company_rate * total_quantity))
                )

                worksheet.append(
                    head_common
                    + [
                        item.codigo_trabajo or "",
                        item.tipo_trabajo or "",
                        item.descripcion or "",
                        item.unidad_medida or "",
                        float(total_quantity),
                        0.0,
                        float(company_rate),
                        float(technical_subtotal),
                        float(company_subtotal),
                    ]
                )

    now_string = timezone.now().strftime("%Y%m%d_%H%M%S")

    response = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument." "spreadsheetml.sheet"
        )
    )

    response["Content-Disposition"] = (
        f'attachment; filename="invoices_{now_string}.xlsx"'
    )

    workbook.save(response)

    return response


@login_required
@rol_requerido("facturacion", "admin", "pm")
@require_POST
def invoice_mark_sent_to_client(request, pk):
    """
    Marca un invoice como enviado al cliente.

        sent
        Pending to send to client

            ↓

        sent_to_client
        Sent to client
    """
    from django.db import transaction
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from django.utils import timezone

    from operaciones.models import SesionBilling

    with transaction.atomic():
        invoice = get_object_or_404(
            SesionBilling.objects.select_for_update(),
            pk=pk,
        )

        current_status = (
            getattr(invoice, "finance_status", "") or ""
        ).strip()

        if current_status == "sent_to_client":
            return JsonResponse(
                {
                    "ok": True,
                    "message": (
                        "The invoice was already marked as sent "
                        "to the client."
                    ),
                    "finance_status": "sent_to_client",
                    "finance_status_label": "Sent to client",
                }
            )

        if current_status == "paid":
            return JsonResponse(
                {
                    "ok": False,
                    "error": (
                        "This invoice is already collected and cannot "
                        "be marked as sent to the client."
                    ),
                },
                status=409,
            )

        if current_status != "sent":
            return JsonResponse(
                {
                    "ok": False,
                    "error": (
                        "Only invoices with status "
                        "'Pending to send to client' can be marked "
                        "as sent."
                    ),
                    "current_status": current_status,
                },
                status=409,
            )

        invoice.finance_status = "sent_to_client"

        update_fields = ["finance_status"]

        if hasattr(invoice, "finance_updated_at"):
            invoice.finance_updated_at = timezone.now()
            update_fields.append("finance_updated_at")

        invoice.save(update_fields=update_fields)

    return JsonResponse(
        {
            "ok": True,
            "message": "Invoice marked as sent to client.",
            "finance_status": "sent_to_client",
            "finance_status_label": "Sent to client",
        }
    )


@login_required
@rol_requerido("facturacion", "admin", "pm")
@require_POST
def invoice_mark_pending_invoice(request, pk):
    """
    Marca un invoice como pendiente por facturar.

        sent_to_client
        Sent to client

            ↓

        pending_invoice
        Pending invoicing
    """
    from django.db import transaction
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from django.utils import timezone

    from operaciones.models import SesionBilling

    with transaction.atomic():
        invoice = get_object_or_404(
            SesionBilling.objects.select_for_update(),
            pk=pk,
        )

        current_status = (getattr(invoice, "finance_status", "") or "").strip()

        if current_status == "pending_invoice":
            return JsonResponse(
                {
                    "ok": True,
                    "message": ("The invoice is already pending invoicing."),
                    "finance_status": "pending_invoice",
                    "finance_status_label": "Pending invoicing",
                }
            )

        if current_status == "paid":
            return JsonResponse(
                {
                    "ok": False,
                    "error": (
                        "This invoice is already collected and cannot "
                        "be marked as pending invoicing."
                    ),
                },
                status=409,
            )

        if current_status != "sent_to_client":
            return JsonResponse(
                {
                    "ok": False,
                    "error": (
                        "Only invoices with status 'Sent to client' "
                        "can be marked as Pending invoicing."
                    ),
                    "current_status": current_status,
                },
                status=409,
            )

        invoice.finance_status = "pending_invoice"

        update_fields = ["finance_status"]

        if hasattr(invoice, "finance_updated_at"):
            invoice.finance_updated_at = timezone.now()
            update_fields.append("finance_updated_at")

        invoice.save(update_fields=update_fields)

    return JsonResponse(
        {
            "ok": True,
            "message": "Invoice marked as pending invoicing.",
            "finance_status": "pending_invoice",
            "finance_status_label": "Pending invoicing",
        }
    )


@login_required
@rol_requerido("facturacion", "admin")
@require_POST
def invoice_bulk_update_status(request):
    """
    Cambia de forma masiva el estado financiero de invoices.

    Reglas:

    - Excluye siempre los Direct Discount.
    - Solo procesa invoices visibles para el usuario.
    - Solo procesa invoices aprobados por Supervisor o PM.
    - Permite cambiar directamente a:
        sent
        sent_to_client
        pending_invoice
        invoiced
        paid
    - Para invoiced y paid:
        * Si Real Company Billing está vacío, copia subtotal_empresa.
        * No sobrescribe un Real Company Billing existente.
    - Para paid:
        * Completa finance_finish_date con la fecha actual si está vacía.
    """
    import json
    from datetime import datetime
    from decimal import Decimal

    from django.db import transaction
    from django.http import JsonResponse
    from django.shortcuts import get_object_or_404
    from django.utils import timezone

    from facturacion.models import Proyecto
    from operaciones.models import SesionBilling

    allowed_statuses = {
        "sent": "Pending to send to client",
        "sent_to_client": "Sent to client",
        "pending_invoice": "Pending invoicing",
        "invoiced": "Invoiced",
        "paid": "Collected",
    }

    content_type = (request.content_type or "").lower()

    if "application/json" in content_type:
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except (json.JSONDecodeError, UnicodeDecodeError):
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Invalid JSON payload.",
                },
                status=400,
            )
    else:
        payload = request.POST

    raw_ids = payload.get("ids", [])
    target_status = str(payload.get("target_status", "") or "").strip()

    if isinstance(raw_ids, str):
        raw_ids = [
            value.strip()
            for value in raw_ids.replace(";", ",").split(",")
            if value.strip()
        ]

    if not isinstance(raw_ids, (list, tuple)):
        return JsonResponse(
            {
                "ok": False,
                "error": "Invalid invoice selection.",
            },
            status=400,
        )

    invoice_ids = []

    for value in raw_ids:
        try:
            invoice_id = int(value)
        except (TypeError, ValueError):
            continue

        if invoice_id > 0 and invoice_id not in invoice_ids:
            invoice_ids.append(invoice_id)

    if not invoice_ids:
        return JsonResponse(
            {
                "ok": False,
                "error": "Select at least one invoice.",
            },
            status=400,
        )

    if target_status not in allowed_statuses:
        return JsonResponse(
            {
                "ok": False,
                "error": "Select a valid finance status.",
            },
            status=400,
        )

    user = request.user

    can_view_legacy_history = user.is_superuser or getattr(
        user,
        "es_usuario_historial",
        False,
    )

    # ============================================================
    # Proyectos visibles para el usuario
    # ============================================================
    try:
        proyectos_user = filter_queryset_by_access(
            Proyecto.objects.all(),
            user,
            "id",
        )
    except Exception:
        proyectos_user = Proyecto.objects.none()

    proyectos_list = list(proyectos_user)

    allowed_project_keys = set()

    for project in proyectos_list:
        project_name = (getattr(project, "nombre", "") or "").strip()
        project_code = str(getattr(project, "codigo", "") or "").strip()
        project_id = str(getattr(project, "id", "") or "").strip()

        if project_name:
            allowed_project_keys.add(project_name.lower())

        if project_code:
            allowed_project_keys.add(project_code.lower())

        if project_id:
            allowed_project_keys.add(project_id.lower())

    # ============================================================
    # Ventana histórica de ProyectoAsignacion
    # ============================================================
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    access_by_project_key = {}

    if ProyectoAsignacion is not None and not can_view_legacy_history:
        try:
            assignments = list(
                ProyectoAsignacion.objects.filter(
                    usuario=user,
                    proyecto__in=proyectos_user,
                ).select_related("proyecto")
            )
        except Exception:
            assignments = []

        for assignment in assignments:
            project = getattr(assignment, "proyecto", None)

            if not project:
                continue

            include_history = bool(getattr(assignment, "include_history", False))

            start_at = getattr(assignment, "start_at", None)

            access_data = {
                "include_history": include_history or not start_at,
                "start_at": None if include_history or not start_at else start_at,
            }

            for key in (
                getattr(project, "nombre", None),
                getattr(project, "codigo", None),
                getattr(project, "id", None),
            ):
                if key is None:
                    continue

                normalized_key = str(key).strip().lower()

                if normalized_key:
                    access_by_project_key[normalized_key] = access_data

    def user_can_access_invoice(invoice):
        if can_view_legacy_history:
            return True

        project_key = str(getattr(invoice, "proyecto", "") or "").strip().lower()

        if not project_key:
            return False

        if project_key not in allowed_project_keys:
            return False

        if ProyectoAsignacion is None:
            return True

        access = access_by_project_key.get(project_key)

        if not access:
            return False

        if access["include_history"] or access["start_at"] is None:
            return True

        created_at = getattr(invoice, "creado_en", None)

        if not created_at:
            return False

        start_at = access["start_at"]

        start_date = start_at.date() if isinstance(start_at, datetime) else start_at

        if isinstance(created_at, datetime):
            created_date = (
                timezone.localtime(created_at).date()
                if timezone.is_aware(created_at)
                else created_at.date()
            )
        else:
            created_date = created_at

        return created_date >= start_date

    updated_count = 0
    filled_real_count = 0
    skipped_direct_discount_count = 0
    skipped_access_count = 0
    skipped_not_approved_count = 0
    not_found_count = 0

    today = timezone.localdate()

    with transaction.atomic():
        invoices = list(
            SesionBilling.objects.select_for_update().filter(
                id__in=invoice_ids,
            )
        )

        found_ids = {invoice.id for invoice in invoices}
        not_found_count = len(set(invoice_ids) - found_ids)

        for invoice in invoices:
            if getattr(invoice, "is_direct_discount", False):
                skipped_direct_discount_count += 1
                continue

            if invoice.estado not in {
                "aprobado_supervisor",
                "aprobado_pm",
            }:
                skipped_not_approved_count += 1
                continue

            if not user_can_access_invoice(invoice):
                skipped_access_count += 1
                continue

            update_fields = []

            if invoice.finance_status != target_status:
                invoice.finance_status = target_status
                update_fields.append("finance_status")

            if target_status in {"invoiced", "paid"}:
                if invoice.real_company_billing is None:
                    invoice.real_company_billing = (
                        invoice.subtotal_empresa
                        if invoice.subtotal_empresa is not None
                        else Decimal("0")
                    )
                    update_fields.append("real_company_billing")
                    filled_real_count += 1

            if target_status == "paid":
                if not invoice.finance_finish_date:
                    invoice.finance_finish_date = today
                    update_fields.append("finance_finish_date")

            if hasattr(invoice, "finance_updated_at"):
                invoice.finance_updated_at = timezone.now()
                update_fields.append("finance_updated_at")

            update_fields = list(dict.fromkeys(update_fields))

            if update_fields:
                invoice.save(update_fields=update_fields)

            updated_count += 1

    skipped_count = (
        skipped_direct_discount_count
        + skipped_access_count
        + skipped_not_approved_count
        + not_found_count
    )

    return JsonResponse(
        {
            "ok": True,
            "message": (
                f"{updated_count} invoice(s) updated to "
                f"{allowed_statuses[target_status]}."
            ),
            "target_status": target_status,
            "target_status_label": allowed_statuses[target_status],
            "updated_count": updated_count,
            "filled_real_count": filled_real_count,
            "skipped_count": skipped_count,
            "skipped_direct_discount_count": (skipped_direct_discount_count),
            "skipped_access_count": skipped_access_count,
            "skipped_not_approved_count": (skipped_not_approved_count),
            "not_found_count": not_found_count,
        }
    )
