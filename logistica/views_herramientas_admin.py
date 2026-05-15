from __future__ import annotations

import re
import unicodedata
from collections import defaultdict
from decimal import Decimal
from uuid import uuid4

from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Q
from django.http import (Http404, HttpResponse, HttpResponseForbidden,
                         JsonResponse)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone

from usuarios.decoradores import rol_requerido

from .forms_herramientas import (BodegaForm, HerramientaForm,
                                 InventarioReviewForm)
from .models import (Bodega, Herramienta, HerramientaAsignacion,
                     HerramientaInventario)


def _is_ajax(request) -> bool:
    return (request.headers.get("x-requested-with") or "").lower() == "xmlhttprequest"


def _user_label(u) -> str:
    if not u:
        return ""
    name = (u.get_full_name() or "").strip()
    return name or (u.username or "")


def _can_admin_logistica(user) -> bool:
    if getattr(user, "es_admin_general", False):
        return True
    if (
        getattr(user, "es_supervisor", False)
        or getattr(user, "es_pm", False)
        or getattr(user, "es_logistica", False)
    ):
        return True
    return False


def _build_inventory_payload_for_asignacion(request, a: HerramientaAsignacion) -> dict:
    last_inv = (
        HerramientaInventario.objects.filter(asignacion=a)
        .select_related("revisado_por")
        .order_by("-created_at", "-id")
        .first()
    )

    inv = None
    if last_inv:
        inv = {
            "id": last_inv.id,
            "estado": last_inv.estado or "",
            "motivo_rechazo": last_inv.motivo_rechazo or "",
            "foto_url": last_inv.foto.url if getattr(last_inv, "foto", None) else "",
            "puede_aprobar": last_inv.estado == "pendiente",
            "aprobar_url": reverse("logistica:aprobar_inventario", args=[last_inv.id]),
            "rechazar_url": reverse(
                "logistica:rechazar_inventario", args=[last_inv.id]
            ),
        }

    puede_solicitar = (not last_inv) or (last_inv.estado != "pendiente")

    return {
        "ok": True,
        "asig_id": a.id,
        "tool_id": a.herramienta_id,
        "inv": inv,
        "puede_solicitar": puede_solicitar,
        "solicitar_url": reverse(
            "logistica:solicitar_inventario_asignacion", args=[a.id]
        ),
        "historial_url": reverse(
            "logistica:inventario_historial_asignacion_admin", args=[a.id]
        ),
        "prox_due": (
            a.herramienta.next_inventory_due.strftime("%m/%d/%Y")
            if a.herramienta.next_inventory_due
            else ""
        ),
    }


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def bodegas_manage(request):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    tipo_id = request.GET.get("editar")
    edit = None

    if tipo_id:
        edit = get_object_or_404(Bodega, pk=tipo_id)
        form = BodegaForm(request.POST or None, instance=edit)
    else:
        form = BodegaForm(request.POST or None)

    if request.method == "POST":
        if form.is_valid():
            obj = form.save(commit=False)
            if not obj.pk:
                obj.creada_por = request.user
            obj.save()
            messages.success(request, "✅ Warehouse saved successfully.")
            return redirect("logistica:bodegas_manage")

        messages.error(request, "❌ Please review the warehouse fields.")

    bodegas = Bodega.objects.all().order_by("nombre")
    can_delete_bodega = bool(getattr(request.user, "es_admin_general", False))

    return render(
        request,
        "logistica/admin_bodegas_manage.html",
        {
            "form": form,
            "bodegas": bodegas,
            "edit": edit,
            "can_delete_bodega": can_delete_bodega,
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramientas_list(request):
    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()
    cantidad = (request.GET.get("cantidad") or "20").strip()
    page_number = (request.GET.get("page") or "1").strip()

    try:
        per_page = int(cantidad)
    except Exception:
        per_page = 20

    if per_page not in (5, 10, 20, 50, 100):
        per_page = 20

    qs = Herramienta.objects.select_related("bodega", "creada_por").order_by(
        "-created_at", "-id"
    )

    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(serial__icontains=q)
            | Q(descripcion__icontains=q)
        )

    if status:
        qs = qs.filter(status=status)

    paginator = Paginator(qs, per_page)
    pagina = paginator.get_page(page_number)

    tool_ids = [h.id for h in pagina.object_list]

    active_count_by_tool = defaultdict(int)
    active_names_by_tool = defaultdict(list)

    if tool_ids:
        active_asigs = (
            HerramientaAsignacion.objects.filter(
                herramienta_id__in=tool_ids, active=True
            )
            .select_related("asignado_a")
            .order_by("-asignado_at", "-id")
        )

        for a in active_asigs:
            active_count_by_tool[a.herramienta_id] += 1

            if len(active_names_by_tool[a.herramienta_id]) < 3:
                u = a.asignado_a
                name = (u.get_full_name() or u.username or "").strip()
                if name:
                    active_names_by_tool[a.herramienta_id].append(name)

    for h in pagina.object_list:
        h.active_asig_count = int(active_count_by_tool.get(h.id, 0) or 0)
        preview = active_names_by_tool.get(h.id, [])
        h.active_assignees_preview = ", ".join(preview) if preview else ""

    return render(
        request,
        "logistica/admin_herramientas_list.html",
        {
            "pagina": pagina,
            "q": q,
            "status": status,
            "cantidad": str(per_page),
            "STATUS_CHOICES": Herramienta.STATUS_CHOICES,
            "can_delete": bool(getattr(request.user, "es_admin_general", False)),
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def exportar_herramientas_excel(request):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    q = (request.GET.get("q") or "").strip()
    status = (request.GET.get("status") or "").strip()

    qs = Herramienta.objects.select_related("bodega", "creada_por").order_by(
        "-created_at"
    )

    if q:
        qs = qs.filter(
            Q(nombre__icontains=q)
            | Q(serial__icontains=q)
            | Q(descripcion__icontains=q)
        )

    if status:
        qs = qs.filter(status=status)

    herramientas = list(qs[:20000])
    tool_ids = [h.id for h in herramientas]

    active_asigs_txt_by_tool = defaultdict(list)
    active_asigs_count_by_tool = defaultdict(int)
    active_asigs_qty_by_tool = defaultdict(int)

    if tool_ids:
        active_asigs = (
            HerramientaAsignacion.objects.filter(
                active=True, herramienta_id__in=tool_ids
            )
            .select_related("asignado_a", "asignado_por")
            .order_by("herramienta_id", "-asignado_at", "-id")
        )

        for a in active_asigs:
            active_asigs_count_by_tool[a.herramienta_id] += 1
            qty = int(getattr(a, "cantidad_entregada", 0) or 0)
            active_asigs_qty_by_tool[a.herramienta_id] += qty

            u = a.asignado_a
            nombre = (u.get_full_name() or u.username or "").strip() or f"User#{u.id}"
            active_asigs_txt_by_tool[a.herramienta_id].append(f"{nombre} ({qty})")

    latest_inv_by_tool = {}
    if tool_ids:
        invs = (
            HerramientaInventario.objects.filter(herramienta_id__in=tool_ids)
            .select_related("revisado_por", "asignacion", "asignacion__asignado_a")
            .order_by("-created_at", "-id")
        )

        for inv in invs:
            if inv.herramienta_id not in latest_inv_by_tool:
                latest_inv_by_tool[inv.herramienta_id] = inv

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Tools"

    headers = [
        "Name",
        "Serial",
        "Description",
        "Commercial value",
        "Warehouse",
        "Tool status",
        "Active assignments",
        "Total assigned (active)",
        "# Active assignments",
        "Inventory (latest status)",
        "Inventory (latest submitted)",
        "Inventory (assigned to)",
        "Inventory (reviewed by)",
        "Inventory (review date)",
        "Next inventory due",
        "Created by",
        "Created at",
    ]
    ws.append(headers)

    for h in herramientas:
        asigs_txt = "; ".join(active_asigs_txt_by_tool.get(h.id, []))
        total_asig = int(active_asigs_qty_by_tool.get(h.id, 0) or 0)
        count_asig = int(active_asigs_count_by_tool.get(h.id, 0) or 0)

        last_inv = latest_inv_by_tool.get(h.id)

        inv_estado = ""
        inv_enviado = ""
        inv_asignado_a = ""
        inv_revisado_por = ""
        inv_revisado_at = ""

        if last_inv:
            inv_estado = last_inv.estado or ""
            inv_enviado = (
                timezone.localtime(last_inv.created_at).strftime("%m/%d/%Y %H:%M")
                if last_inv.created_at
                else ""
            )

            try:
                if last_inv.asignacion and last_inv.asignacion.asignado_a:
                    u = last_inv.asignacion.asignado_a
                    inv_asignado_a = (u.get_full_name() or u.username or "").strip()
            except Exception:
                inv_asignado_a = ""

            if last_inv.revisado_por:
                inv_revisado_por = (
                    last_inv.revisado_por.get_full_name()
                    or last_inv.revisado_por.username
                )

            if last_inv.revisado_at:
                inv_revisado_at = timezone.localtime(last_inv.revisado_at).strftime(
                    "%m/%d/%Y %H:%M"
                )

        creada_por = (
            h.creada_por.get_full_name() or h.creada_por.username
            if h.creada_por
            else ""
        )
        creada_el = (
            timezone.localtime(h.created_at).strftime("%m/%d/%Y %H:%M")
            if h.created_at
            else ""
        )
        prox_fecha = (
            h.next_inventory_due.strftime("%m/%d/%Y") if h.next_inventory_due else ""
        )

        ws.append(
            [
                h.nombre or "",
                h.serial or "",
                h.descripcion or "",
                float(h.valor_comercial) if h.valor_comercial is not None else "",
                h.bodega.nombre if h.bodega else "",
                (
                    h.get_status_display()
                    if hasattr(h, "get_status_display")
                    else (h.status or "")
                ),
                asigs_txt,
                total_asig,
                count_asig,
                inv_estado,
                inv_enviado,
                inv_asignado_a,
                inv_revisado_por,
                inv_revisado_at,
                prox_fecha,
                creada_por,
                creada_el,
            ]
        )

    widths = [
        28,
        18,
        40,
        18,
        20,
        18,
        55,
        22,
        20,
        24,
        24,
        26,
        24,
        24,
        20,
        24,
        20,
    ]

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"tools_{timezone.localdate().strftime('%Y-%m-%d')}.xlsx"
    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramienta_create(request):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    if request.method == "POST":
        form = HerramientaForm(request.POST, request.FILES)

        if form.is_valid():
            obj = form.save(commit=False)
            obj.creada_por = request.user

            if not obj.next_inventory_due:
                obj.mark_inventory_due_default()

            obj.save()
            messages.success(request, "✅ Tool created successfully.")
            return redirect("logistica:herramientas_list")

        messages.error(request, "❌ Please review the fields.")
    else:
        form = HerramientaForm()

    return render(
        request,
        "logistica/admin_herramienta_form.html",
        {
            "form": form,
            "mode": "create",
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramienta_edit(request, tool_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    obj = get_object_or_404(Herramienta, pk=tool_id)

    if request.method == "POST":
        form = HerramientaForm(request.POST, request.FILES, instance=obj)

        if form.is_valid():
            updated = form.save(commit=False)
            updated.save()
            messages.success(request, "✅ Tool updated successfully.")
            return redirect("logistica:herramientas_list")

        messages.error(request, "❌ Please review the fields.")
    else:
        form = HerramientaForm(instance=obj)

    return render(
        request,
        "logistica/admin_herramienta_form.html",
        {
            "form": form,
            "mode": "edit",
            "obj": obj,
            "can_delete": bool(getattr(request.user, "es_admin_general", False)),
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramienta_delete(request, tool_id: int):
    if not getattr(request.user, "es_admin_general", False):
        return HttpResponseForbidden("Only general admin can delete tools.")

    obj = get_object_or_404(Herramienta, pk=tool_id)

    if request.method == "POST":
        try:
            if obj.foto and obj.foto.name:
                obj.foto.delete(save=False)
        except Exception:
            pass

        invs = HerramientaInventario.objects.filter(herramienta=obj)
        for inv in invs:
            try:
                if inv.foto and inv.foto.name:
                    inv.foto.delete(save=False)
            except Exception:
                pass

        obj.delete()
        messages.success(request, "✅ Tool deleted successfully.")
        return redirect("logistica:herramientas_list")

    return render(
        request,
        "logistica/admin_herramienta_delete.html",
        {
            "obj": obj,
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramienta_reset_assignment_status(request, tool_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    tool = get_object_or_404(Herramienta, pk=tool_id)

    if request.method != "POST":
        raise Http404()

    a = HerramientaAsignacion.objects.filter(herramienta=tool, active=True).first()

    if not a:
        messages.error(request, "There is no active assignment to reset.")
        return redirect("logistica:herramientas_list")

    a.estado = "pendiente"
    a.comentario_rechazo = None
    a.rechazado_at = None
    a.aceptado_at = None
    a.save(
        update_fields=["estado", "comentario_rechazo", "rechazado_at", "aceptado_at"]
    )

    messages.success(request, "✅ Assignment status reset to Pending.")
    return redirect("logistica:herramientas_list")


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramienta_change_status(request, tool_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    tool = get_object_or_404(Herramienta, pk=tool_id)

    if request.method != "POST":
        raise Http404()

    new_status = (request.POST.get("status") or "").strip()
    just = (request.POST.get("justificacion") or "").strip()

    try:
        if new_status == "bodega":
            with transaction.atomic():
                a = HerramientaAsignacion.objects.filter(
                    herramienta=tool,
                    active=True,
                ).first()

                if a:
                    a.active = False
                    a.save(update_fields=["active"])

                tool.inventory_required = False
                tool.set_status(new_status, by_user=request.user, justification=just)
                tool.save(
                    update_fields=[
                        "status",
                        "inventory_required",
                        "status_justificacion",
                        "status_changed_at",
                        "status_changed_by",
                        "updated_at",
                    ]
                )
        else:
            tool.set_status(new_status, by_user=request.user, justification=just)
            tool.save(
                update_fields=[
                    "status",
                    "status_justificacion",
                    "status_changed_at",
                    "status_changed_by",
                    "updated_at",
                ]
            )

        messages.success(request, "✅ Status updated successfully.")
    except ValidationError as e:
        messages.error(request, f"❌ {e}")

    return redirect("logistica:herramientas_list")


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def solicitar_inventario(request, tool_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    tool = get_object_or_404(Herramienta, pk=tool_id)

    if request.method != "POST":
        raise Http404()

    tool.inventory_required = True
    tool.save(update_fields=["inventory_required", "updated_at"])

    messages.success(
        request, "📸 Inventory requested. The user can now upload a photo."
    )
    return redirect("logistica:herramientas_list")


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def aprobar_inventario(request, inv_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    if request.method != "POST":
        raise Http404()

    with transaction.atomic():
        inv = get_object_or_404(
            HerramientaInventario.objects.select_for_update().select_related(
                "asignacion", "asignacion__herramienta"
            ),
            pk=inv_id,
        )

        inv.approve(request.user)
        a = inv.asignacion

        if hasattr(inv, "asignacion") and hasattr(inv.asignacion, "herramienta"):
            a.herramienta = inv.asignacion.herramienta

    messages.success(request, "✅ Inventory approved.")

    if _is_ajax(request):
        payload = _build_inventory_payload_for_asignacion(request, a)
        payload["message"] = "Inventory approved."
        return JsonResponse(payload)

    nxt = (request.GET.get("next") or "").strip()

    if nxt:
        return redirect(nxt)

    return redirect("logistica:herramientas_list")


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def rechazar_inventario(request, inv_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    inv = get_object_or_404(
        HerramientaInventario.objects.select_related(
            "asignacion",
            "asignacion__herramienta",
        ),
        pk=inv_id,
    )

    nxt = (request.GET.get("next") or "").strip()

    if request.method == "POST" and _is_ajax(request):
        motivo = (request.POST.get("motivo_rechazo") or "").strip()

        if not motivo:
            return JsonResponse(
                {"ok": False, "error": "A rejection reason is required."},
                status=400,
            )

        with transaction.atomic():
            inv = get_object_or_404(
                HerramientaInventario.objects.select_for_update().select_related(
                    "asignacion", "asignacion__herramienta"
                ),
                pk=inv_id,
            )
            inv.reject(request.user, motivo)
            a = inv.asignacion

        payload = _build_inventory_payload_for_asignacion(request, a)
        payload["message"] = "Inventory rejected."
        return JsonResponse(payload)

    if request.method == "POST":
        form = InventarioReviewForm(request.POST)

        if form.is_valid():
            inv.reject(request.user, form.cleaned_data["motivo_rechazo"])
            messages.warning(
                request, "❌ Inventory rejected. The user can upload again."
            )

            if nxt:
                return redirect(nxt)

            return redirect("logistica:herramientas_list")

        messages.error(request, "A rejection reason is required.")
    else:
        form = InventarioReviewForm()

    return render(
        request,
        "logistica/admin_inventario_rechazar.html",
        {
            "inv": inv,
            "form": form,
            "next": nxt,
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def inventario_historial_admin(request, tool_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    tool = get_object_or_404(Herramienta, pk=tool_id)

    inventarios = (
        HerramientaInventario.objects.filter(herramienta=tool)
        .select_related(
            "asignacion",
            "asignacion__asignado_a",
            "asignacion__asignado_por",
            "revisado_por",
        )
        .order_by("-created_at", "-id")
    )

    return render(
        request,
        "logistica/admin_inventario_historial_asignacion.html",
        {
            "herramienta": tool,
            "inventarios": list(inventarios),
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def asignaciones_historial_admin(request, tool_id: int):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    tool = get_object_or_404(Herramienta, pk=tool_id)

    asignaciones = (
        HerramientaAsignacion.objects.filter(herramienta=tool)
        .select_related("asignado_a", "asignado_por")
        .order_by("-asignado_at", "-id")
    )

    return render(
        request,
        "logistica/admin_asignaciones_historial.html",
        {
            "herramienta": tool,
            "asignaciones": list(asignaciones),
        },
    )


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def bodega_delete(request, bodega_id: int):
    if not getattr(request.user, "es_admin_general", False):
        return HttpResponseForbidden("Only general admin can delete warehouses.")

    if request.method != "POST":
        raise Http404()

    bodega = get_object_or_404(Bodega, pk=bodega_id)

    tools_count = Herramienta.objects.filter(bodega=bodega).count()

    if tools_count > 0:
        messages.error(
            request,
            (
                f"❌ This warehouse cannot be deleted because it has "
                f"{tools_count} associated tool(s). Move those tools first."
            ),
        )
        return redirect("logistica:bodegas_manage")

    try:
        bodega.delete()
        messages.success(request, "✅ Warehouse deleted successfully.")
    except Exception as e:
        messages.error(request, f"❌ Warehouse could not be deleted: {e}")

    return redirect("logistica:bodegas_manage")


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramientas_importar_plantilla(request):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Tools"

    headers = [
        "name",
        "serial",
        "quantity",
        "commercial_value",
        "description",
        "status",
    ]
    ws.append(headers)

    ws.append(
        ["Hammer drill", "SN-AX12-8890", 3, 500.00, "Bosch brand, with case", "bodega"]
    )
    ws.append(["Gloves", "", 20, "", "Size M gloves", "bodega"])

    widths = [28, 22, 12, 18, 45, 16]

    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    filename = f"tools_import_template_{timezone.localdate().strftime('%Y-%m-%d')}.xlsx"

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(resp)
    return resp


@staff_member_required
@rol_requerido("admin", "pm", "supervisor", "logistica")
def herramientas_importar(request):
    if not _can_admin_logistica(request.user):
        return HttpResponseForbidden("You do not have permission.")

    if request.method != "POST":
        raise Http404()

    from openpyxl import load_workbook

    def _bool_post(v) -> bool:
        v = (v or "").strip().lower()
        return v in {"1", "true", "on", "yes"}

    def norm(s: str) -> str:
        s = s or ""
        s = s.replace("\u00a0", " ").replace("\u2007", " ").replace("\u202f", " ")
        s = s.strip().lower()
        s = unicodedata.normalize("NFKD", s)
        s = "".join(c for c in s if not unicodedata.combining(c))
        s = re.sub(r"\s+", " ", s).strip()
        return s

    def gen_serial_unico() -> str:
        for _ in range(80):
            s = f"AUTO-{uuid4().hex[:10].upper()}"
            if not Herramienta.objects.filter(serial=s).exists():
                return s
        return f"AUTO-{uuid4().hex[:10].upper()}"

    def to_int(v, default=1) -> int:
        try:
            if v is None:
                return default
            if isinstance(v, str) and not v.strip():
                return default
            n = int(v)
            return n if n > 0 else default
        except Exception:
            return default

    def to_decimal(v) -> Decimal:
        try:
            if v is None:
                return Decimal("0")
            if isinstance(v, str) and not v.strip():
                return Decimal("0")
            d = Decimal(str(v))
            return d if d >= 0 else Decimal("0")
        except Exception:
            return Decimal("0")

    confirmar = (request.POST.get("confirmar") or "").strip()

    if confirmar == "1":
        token = (request.POST.get("token") or "").strip()
        modo = (request.POST.get("modo") or "reemplazar").strip()

        if modo not in {"reemplazar", "sumar"}:
            modo = "reemplazar"

        session_key = f"logistica_import_tools:{token}"
        payload = request.session.get(session_key)

        if not payload:
            messages.error(
                request, "❌ The import session expired. Please upload the file again."
            )
            return redirect("logistica:herramientas_list")

        rows = payload.get("rows") or []

        if not rows:
            messages.error(request, "❌ There are no rows to import.")
            return redirect("logistica:herramientas_list")

        bodegas_by_id = {str(b.id): b for b in Bodega.objects.all()}

        bodega_default_id = (request.POST.get("bodega_default") or "").strip()
        bodega_default = (
            bodegas_by_id.get(bodega_default_id) if bodega_default_id else None
        )

        aplicar_default_a_existentes = _bool_post(
            request.POST.get("aplicar_default_a_existentes")
        )

        if aplicar_default_a_existentes and not bodega_default:
            messages.error(
                request,
                "❌ Select a default warehouse if you want to apply it to existing tools.",
            )
            return redirect("logistica:herramientas_list")

        created = 0
        updated = 0
        errores = 0

        with transaction.atomic():
            for i, r in enumerate(rows):
                try:
                    nombre = (r.get("nombre") or "").strip()

                    if not nombre:
                        continue

                    serial_final = (r.get("serial") or "").strip()
                    serial_excel = (r.get("serial_excel") or "").strip()

                    if not serial_final:
                        serial_final = gen_serial_unico()

                    cantidad = to_int(r.get("cantidad"), default=1)
                    valor = to_decimal(r.get("valor_comercial"))
                    descripcion = (r.get("descripcion") or "").strip() or None

                    status = (r.get("status") or "").strip() or None
                    if status and status not in dict(Herramienta.STATUS_CHOICES):
                        status = None

                    bodega_mode = (
                        (request.POST.get(f"bodega_mode_{i}") or "").strip().lower()
                    )
                    if bodega_mode not in {"keep", "default", "custom"}:
                        bodega_mode = ""

                    bodega_sel = (request.POST.get(f"bodega_sel_{i}") or "").strip()
                    bodega_custom = (
                        bodegas_by_id.get(bodega_sel) if bodega_sel else None
                    )

                    obj = Herramienta.objects.filter(serial=serial_final).first()

                    if obj:
                        obj.nombre = nombre
                        obj.valor_comercial = valor
                        obj.descripcion = descripcion

                        if status:
                            obj.status = status

                        if modo == "sumar":
                            obj.cantidad = int(obj.cantidad or 0) + cantidad
                        else:
                            obj.cantidad = cantidad

                        if bodega_mode == "custom":
                            obj.bodega = bodega_custom
                        elif bodega_mode == "default":
                            obj.bodega = bodega_default
                        elif bodega_mode == "keep":
                            pass
                        else:
                            if (
                                aplicar_default_a_existentes
                                and bodega_default is not None
                            ):
                                obj.bodega = bodega_default

                        if serial_excel and obj.serial != serial_excel:
                            if (
                                not Herramienta.objects.exclude(pk=obj.pk)
                                .filter(serial=serial_excel)
                                .exists()
                            ):
                                obj.serial = serial_excel

                        obj.save()
                        updated += 1

                    else:
                        if bodega_mode == "custom":
                            bodega_final = bodega_custom
                        else:
                            bodega_final = bodega_default

                        obj = Herramienta.objects.create(
                            nombre=nombre,
                            serial=serial_final,
                            cantidad=cantidad,
                            valor_comercial=valor,
                            descripcion=descripcion,
                            bodega=bodega_final,
                            status=status or "bodega",
                            creada_por=request.user,
                        )

                        if not obj.next_inventory_due:
                            obj.mark_inventory_due_default()
                            obj.save(update_fields=["next_inventory_due", "updated_at"])

                        created += 1

                except Exception:
                    errores += 1
                    continue

        try:
            del request.session[session_key]
        except Exception:
            pass

        messages.success(
            request,
            (
                f"✅ Import completed ({modo}). "
                f"Created: {created} • Updated: {updated} • Errors: {errores}"
            ),
        )
        return redirect("logistica:herramientas_list")

    f = request.FILES.get("archivo")

    if not f:
        messages.error(request, "❌ Please select a file.")
        return redirect("logistica:herramientas_list")

    name = (getattr(f, "name", "") or "").lower()

    if not (name.endswith(".xlsx") or name.endswith(".xlsm")):
        messages.error(request, "❌ The file must be .xlsx")
        return redirect("logistica:herramientas_list")

    try:
        wb = load_workbook(f, data_only=True)
        ws = wb.active
    except Exception as e:
        messages.error(request, f"❌ Could not read the Excel file: {e}")
        return redirect("logistica:herramientas_list")

    rows_xl = list(ws.iter_rows(values_only=True))

    if not rows_xl:
        messages.error(request, "❌ The Excel file is empty.")
        return redirect("logistica:herramientas_list")

    header = [str(x or "").strip().lower() for x in rows_xl[0]]

    aliases = {
        "name": "nombre",
        "tool": "nombre",
        "nombre": "nombre",
        "serial": "serial",
        "quantity": "cantidad",
        "qty": "cantidad",
        "cantidad": "cantidad",
        "commercial_value": "valor_comercial",
        "commercial value": "valor_comercial",
        "valor_comercial": "valor_comercial",
        "description": "descripcion",
        "descripcion": "descripcion",
        "status": "status",
    }

    normalized_header = [aliases.get(h, h) for h in header]
    required = {"nombre", "cantidad"}

    if not required.issubset(set(normalized_header)):
        messages.error(
            request,
            "❌ Invalid template. It must include at least: name, quantity.",
        )
        return redirect("logistica:herramientas_list")

    idx = {col: normalized_header.index(col) for col in normalized_header if col}

    def get_cell(r, col):
        i = idx.get(col)
        if i is None:
            return None
        return r[i] if i < len(r) else None

    existing_by_serial = {}
    for h in (
        Herramienta.objects.filter(serial__isnull=False)
        .exclude(serial="")
        .only("id", "serial", "nombre", "bodega_id", "cantidad")
        .select_related("bodega")
    ):
        s = (h.serial or "").strip()
        if s and s not in existing_by_serial:
            existing_by_serial[s] = h

    existing_by_name = {}
    for h in (
        Herramienta.objects.all()
        .only("id", "nombre", "serial", "bodega_id", "cantidad")
        .select_related("bodega")
        .order_by("-id")
    ):
        k = norm(h.nombre)
        if k and k not in existing_by_name:
            existing_by_name[k] = h

    preview_rows = []
    coincidencias = 0

    for r in rows_xl[1:]:
        try:
            nombre = str(get_cell(r, "nombre") or "").strip()

            if not nombre:
                continue

            serial_excel = str(get_cell(r, "serial") or "").strip()
            cantidad = to_int(get_cell(r, "cantidad"), default=1)
            valor = to_decimal(get_cell(r, "valor_comercial"))
            descripcion = str(get_cell(r, "descripcion") or "").strip() or ""

            status = str(get_cell(r, "status") or "").strip() or ""
            if status and status not in dict(Herramienta.STATUS_CHOICES):
                status = ""

            existe = None
            serial_final = ""

            if serial_excel:
                existe = existing_by_serial.get(serial_excel)
                serial_final = serial_excel
            else:
                cand = existing_by_name.get(norm(nombre))
                if cand:
                    existe = cand

                if existe and (existe.serial or "").strip():
                    serial_final = (existe.serial or "").strip()
                else:
                    serial_final = gen_serial_unico()

            if existe:
                coincidencias += 1

            preview_rows.append(
                {
                    "nombre": nombre,
                    "serial": serial_final,
                    "serial_excel": serial_excel,
                    "cantidad": cantidad,
                    "valor_comercial": str(valor),
                    "descripcion": descripcion,
                    "status": status,
                    "existe": bool(existe),
                    "existe_cantidad": int(existe.cantidad) if existe else 0,
                    "existe_bodega": (
                        existe.bodega.nombre
                        if existe and getattr(existe, "bodega", None)
                        else ""
                    ),
                }
            )

        except Exception:
            continue

    if not preview_rows:
        messages.warning(request, "⚠️ No valid rows were found in the file.")
        return redirect("logistica:herramientas_list")

    token = uuid4().hex
    session_key = f"logistica_import_tools:{token}"
    request.session[session_key] = {"rows": preview_rows}
    request.session.modified = True

    bodegas = list(Bodega.objects.all().order_by("nombre"))

    return render(
        request,
        "logistica/admin_herramientas_import_preview.html",
        {
            "token": token,
            "preview_rows": preview_rows,
            "bodegas": bodegas,
            "coincidencias": coincidencias,
            "bodegas_faltantes": 0,
        },
    )
