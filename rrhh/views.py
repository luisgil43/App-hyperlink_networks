from django.shortcuts import render, get_object_or_404, redirect
from django.utils.http import url_has_allowed_host_and_scheme
from django.http import HttpResponseForbidden, FileResponse, Http404
from django.shortcuts import get_object_or_404, render, redirect
from django.db import transaction
from .forms import SignatureCaptureForm  # asumiendo que ya lo tienes
# o usa tu campo CustomUser.firma_digital si migraste
from .models import UserSignature
from django.conf import settings
from django.utils.module_loading import import_string
import mimetypes
from django.http import HttpResponse, Http404
from .models import UserSignature
from .forms import SignatureCaptureForm
import fitz  # PyMuPDF
from PyPDF2 import PdfReader, PdfWriter
from usuarios.decoradores import rol_requerido  # tu decorador de roles
from .forms import RateSheetForm  # del punto 2
from .models import RateSheet
from django.views.decorators.http import require_http_methods
from django.shortcuts import get_object_or_404, redirect, render
from django.core.paginator import Paginator
from usuarios.models import CustomUser  # Asegúrate de importar esto
from usuarios.utils import crear_notificacion
from reportlab.lib.pagesizes import A4
from rrhh.forms import AprobacionAdelantoForm
from collections import Counter
from openpyxl.styles import Font
from django.urls import NoReverseMatch
from rrhh.forms import CronogramaPagoForm
from rrhh.models import CronogramaPago
from django.utils import timezone
from .models import CronogramaPago
from .forms import CronogramaPagoForm
from django.core.exceptions import ValidationError
from rrhh.utils import generar_pdf_solicitud_vacaciones
from rrhh.models import SolicitudVacaciones
from django.shortcuts import get_object_or_404, redirect
from rrhh.utils import generar_pdf_solicitud_vacaciones  # ⬅️ al inicio del archivo
from io import BytesIO
from rrhh.models import FichaIngreso
from reportlab.lib.pagesizes import LETTER
from reportlab.pdfgen import canvas
from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
import uuid
import os
from .models import ContratoTrabajo
from .forms import ContratoTrabajoForm
from django.shortcuts import get_object_or_404
import logging
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.http import FileResponse, Http404
from .forms import FichaIngresoForm
from .forms import SolicitudVacacionesForm
from datetime import date
from django.http import HttpResponseForbidden
from usuarios.decoradores import rol_requerido
from .forms import RevisionVacacionesForm
from rrhh.models import Feriado
import json
from rrhh.models import DiasVacacionesTomadosManualmente
from django.urls import reverse
from django.db.models import Q
from .utils import contar_dias_habiles
from .forms import DocumentoTrabajadorForm, TipoDocumentoForm
from .models import DocumentoTrabajador, TipoDocumento, CustomUser
from django.db.models import OuterRef, Subquery
from .forms import ReemplazoDocumentoForm
import cloudinary.uploader
from django.core.files.base import ContentFile
from django.utils.text import slugify
import openpyxl
from datetime import datetime
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.utils.encoding import smart_str
from .utils import calcular_estado_documento
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import io
from rrhh.utils import generar_ficha_ingreso_pdf
# from rrhh.utils import agregar_firma_trabajador_a_ficha
import requests
from django.core.files.base import ContentFile
from django.contrib.auth import get_user_model
# from .utils import agregar_firma_pm_a_ficha
from usuarios.models import CustomUser
from PIL import Image
from .forms import FirmaForm
import base64
from rrhh.forms import SolicitudAdelantoAdminForm
from rrhh.models import SolicitudAdelanto
from rrhh.models import FichaIngreso
from decimal import Decimal
import calendar
from django.db.models import Sum
from .forms import SolicitudAdelantoForm
from rrhh.utils import generar_pdf_solicitud_adelanto
from .utils import generar_pdf_solicitud_adelanto
User = get_user_model()


logger = logging.getLogger(__name__)


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def listar_contratos_admin(request):
    contratos = ContratoTrabajo.objects.select_related('tecnico')

    identidades = contratos.values_list(
        'tecnico__identidad', flat=True).distinct()
    nombres = contratos.values_list(
        'tecnico__first_name', 'tecnico__last_name').distinct()
    fechas_inicio = contratos.values_list('fecha_inicio', flat=True).distinct()

    fechas_termino_raw = contratos.values_list('fecha_termino', flat=True)
    fechas_termino = []
    for fecha in fechas_termino_raw:
        if fecha:
            fechas_termino.append(str(fecha))
        else:
            fechas_termino.append("Indefinido")
    fechas_termino = sorted(set(fechas_termino))

    nombres_completos = sorted(set(
        f"{n[0]} {n[1]}" for n in nombres if n[0] and n[1]
    ))

    return render(request, 'rrhh/listar_contratos_admin.html', {
        'contratos': contratos,
        'identidades': identidades,
        'nombres': nombres_completos,
        'fechas_inicio': fechas_inicio,
        'fechas_termino': fechas_termino,
    })


@login_required
def listar_contratos_usuario(request):
    try:
        usuario = request.user
        logger.info(f"🧪 Usuario: {usuario} - ID: {usuario.id}")

        contratos = ContratoTrabajo.objects.filter(tecnico=usuario)
        logger.info(f"🧪 Total contratos: {contratos.count()}")

        return render(request, 'rrhh/contratos_trabajo.html', {
            'contratos': contratos
        })
    except Exception as e:
        logger.error(f"❌ Error al cargar contratos usuario: {e}")
        raise e  # Deja que falle para ver en los logs de Render


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def crear_contrato(request):
    if request.method == 'POST':
        archivo = request.FILES.get('archivo')

        # Validar archivo antes de crear el formulario
        if not archivo:
            messages.error(request, '❌ Debes subir un archivo PDF.')
            return render(request, 'rrhh/crear_contrato.html', {'form': ContratoTrabajoForm(request.POST)})

        if archivo.content_type != 'application/pdf':
            messages.error(
                request, '❌ Estás intentando subir un documento no válido. El archivo debe estar en formato PDF.')
            return render(request, 'rrhh/crear_contrato.html', {'form': ContratoTrabajoForm(request.POST)})

        form = ContratoTrabajoForm(request.POST, request.FILES)

        if form.is_valid():
            contrato = form.save(commit=False)

            if request.POST.get('indefinido-check'):
                contrato.fecha_termino = None

            contrato.archivo = archivo
            contrato.save()

            crear_notificacion(
                usuario=contrato.tecnico,
                mensaje='Se ha generado un nuevo contrato de trabajo. Puedes revisarlo y firmarlo en la plataforma.',
                # Ajusta la URL si es distinta
                url=reverse('rrhh:mis_contratos'),
                tipo='info'
            )

            messages.success(request, '✅ Contrato creado correctamente.')
            return redirect('rrhh:contratos_trabajo')
        else:
            messages.error(
                request, '❌ Error al crear el contrato. Revisa los campos.')
    else:
        form = ContratoTrabajoForm()

    return render(request, 'rrhh/crear_contrato.html', {'form': form})


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def editar_contrato(request, contrato_id):
    contrato = get_object_or_404(ContratoTrabajo, id=contrato_id)

    form = ContratoTrabajoForm(
        request.POST or None, request.FILES or None, instance=contrato)

    if request.method == 'POST' and form.is_valid():
        contrato = form.save(commit=False)

        if request.POST.get('indefinido-check'):
            contrato.fecha_termino = None

        reemplazar = form.cleaned_data.get('reemplazar_archivo')
        archivo_nuevo = request.FILES.get('archivo')

        if reemplazar and archivo_nuevo:
            if archivo_nuevo.content_type != 'application/pdf':
                messages.error(request, '❌ El archivo debe ser un PDF válido.')
                return render(request, 'rrhh/editar_contrato.html', {'form': form, 'contrato': contrato})

            try:
                if contrato.archivo and contrato.archivo.name:
                    nombre_original = contrato.archivo.name.split(
                        '/')[-1]  # 🔁 Guardamos antes
                    # 🗑️ Eliminar archivo existente
                    contrato.archivo.delete(save=False)

                    archivo_nuevo.seek(0)
                    contenido = archivo_nuevo.read()
                    contrato.archivo.save(
                        nombre_original, ContentFile(contenido), save=False)
            except Exception as e:
                messages.error(
                    request, f"❌ Error al subir el nuevo archivo: {e}")
                return render(request, 'rrhh/editar_contrato.html', {'form': form, 'contrato': contrato})

        contrato.save()
        messages.success(request, '✅ Contrato actualizado correctamente.')
        return redirect('rrhh:contratos_trabajo')

    return render(request, 'rrhh/editar_contrato.html', {'form': form, 'contrato': contrato})


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def eliminar_contrato(request, contrato_id):
    contrato = get_object_or_404(ContratoTrabajo, id=contrato_id)

    if request.method == 'POST':
        try:
            if contrato.archivo and contrato.archivo.name:
                contrato.archivo.delete(save=False)
            contrato.delete()
            messages.success(request, "✅ Contrato eliminado correctamente.")
        except Exception as e:
            messages.error(
                request, f"❌ Ocurrió un error al eliminar el contrato: {e}")
        return redirect('rrhh:contratos_trabajo')

    return render(request, 'rrhh/eliminar_contrato.html', {'contrato': contrato})


@staff_member_required
def ver_contrato(request, contrato_id):
    try:
        contrato = get_object_or_404(ContratoTrabajo, id=contrato_id)

        try:
            archivo = contrato.archivo.open()
        except Exception as e:
            messages.error(request, f"❌ No se pudo acceder al archivo: {e}")
            return redirect('rrhh:contratos_trabajo')

        # ✅ Mostramos el PDF directamente en navegador
        return FileResponse(contrato.archivo.open(), content_type='application/pdf')

    except Exception as e:
        messages.error(request, f"❌ Error al mostrar el contrato: {e}")
        return redirect('rrhh:contratos_trabajo')


@login_required
def listar_fichas_ingreso_usuario(request):
    fichas = FichaIngreso.objects.filter(usuario=request.user)
    return render(request, 'rrhh/listar_fichas_ingreso_usuario.html', {
        'fichas': fichas
    })

# Vista admin para listar fichas de ingreso (reutiliza ContratoTrabajo)


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def listar_fichas_ingreso_admin(request):
    fichas = FichaIngreso.objects.all()
    return render(request, 'rrhh/listar_fichas_ingreso_admin.html', {'fichas': fichas})

# Crear ficha (reutilizando formulario)


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def crear_ficha_ingreso(request):
    if request.method == 'POST':
        form = FichaIngresoForm(request.POST)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha.creado_por = request.user

            # Buscar usuario por RUT
            rut_limpio = ficha.rut.replace('.', '').replace('-', '')
            try:
                usuario = CustomUser.objects.get(
                    identidad__icontains=rut_limpio)
                ficha.usuario = usuario
            except CustomUser.DoesNotExist:
                ficha.usuario = None
                messages.warning(
                    request, "⚠️ No se encontró ningún usuario con el RUT ingresado. Se guardará sin asignación de usuario."
                )

            ficha.save()
            generar_ficha_ingreso_pdf(ficha)

            # Notificar al PM solo si hay usuario asignado
            if ficha.usuario:
                nombre_usuario = ficha.usuario.get_full_name()
            else:
                nombre_usuario = "un trabajador sin asignación"

            pms = CustomUser.objects.filter(roles__nombre='pm', is_active=True)
            for pm in pms:
                crear_notificacion(
                    usuario=pm,
                    mensaje=f"Debes revisar la nueva ficha de ingreso de {nombre_usuario}",
                    url=reverse('rrhh:listar_fichas_ingreso_admin'),
                    tipo='info'
                )

            messages.success(
                request, "Ficha de ingreso guardada exitosamente.")
            return redirect('rrhh:listar_fichas_ingreso_admin')
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        form = FichaIngresoForm()

    return render(request, 'rrhh/crear_ficha_ingreso.html', {'form': form})


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def editar_ficha_ingreso(request, pk):
    ficha = get_object_or_404(FichaIngreso, pk=pk)
    estado_anterior = ficha.estado  # 👈 capturamos el estado antes de guardar

    if request.method == 'POST':
        form = FichaIngresoForm(request.POST, request.FILES, instance=ficha)
        if form.is_valid():
            ficha = form.save(commit=False)
            ficha_modificada = False

            # Si estaba firmada o rechazada, se reinicia
            if ficha.estado in ['rechazada_pm', 'rechazada_usuario', 'aprobada']:
                ficha.estado = 'pendiente_pm'
                ficha.firma_rrhh = None
                ficha.firma_pm = None
                ficha.firma_trabajador = None
                ficha_modificada = True

            ficha.save()
            generar_ficha_ingreso_pdf(ficha)

            # 🔔 Notificar al PM si corresponde
            if ficha.pm:
                print("🔔 Enviando notificación al PM...")
                crear_notificacion(
                    usuario=ficha.pm,
                    mensaje=f"La ficha de ingreso de {ficha.usuario.get_full_name()} ha sido modificada y requiere tu revisión.",
                    url=reverse('rrhh:listar_fichas_ingreso_admin'),
                    tipo='info'
                )

            messages.success(
                request, "Ficha actualizada correctamente y reiniciada para aprobación.")
            return redirect('rrhh:listar_fichas_ingreso_admin')
        else:
            print("❌ Formulario no válido:", form.errors)
    else:
        form = FichaIngresoForm(instance=ficha)

    return render(request, 'rrhh/editar_ficha_ingreso.html', {'form': form})


@login_required
@rol_requerido('admin', 'pm')
def rechazar_ficha_ingreso_pm(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id)

    if request.method == 'POST':
        motivo = request.POST.get('motivo', '').strip()

        if not motivo:
            messages.error(
                request, "Debes ingresar un motivo para rechazar la ficha.")
            return redirect('rrhh:listar_fichas_ingreso_admin')

        ficha.estado = 'rechazada_pm'
        ficha.motivo_rechazo_pm = motivo
        ficha.save()

        # 🔔 Notificar a RRHH (quien creó la ficha)
        if ficha.creado_por:
            crear_notificacion(
                usuario=ficha.creado_por,
                mensaje=f"La ficha de ingreso para {ficha.usuario.get_full_name()} fue rechazada por el PM. Motivo: {motivo}",
                url=reverse('rrhh:listar_fichas_ingreso_admin'),
                tipo='warning'
            )

        messages.warning(
            request, "❌ Has rechazado la ficha correctamente. RRHH ha sido notificado.")
        return redirect('rrhh:listar_fichas_ingreso_admin')

    return redirect('rrhh:listar_fichas_ingreso_admin')


@login_required
@rol_requerido('admin', 'usuario')
def aprobar_ficha_ingreso_trabajador(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id, usuario=request.user)

    if ficha.firma_trabajador:
        messages.info(request, "Ya has firmado esta ficha.")
        return redirect('dashboard:mis_fichas_ingreso')

    if not (ficha.firma_pm and ficha.firma_rrhh):
        messages.error(
            request, "No puedes firmar todavía. Aún falta la aprobación del PM o de RRHH.")
        return redirect('dashboard:mis_fichas_ingreso')

    ficha.firma_trabajador = request.user.firma_digital
    ficha.estado = 'completada'
    ficha.save()

    # ✅ Generar nuevo PDF con las tres firmas
    generar_ficha_ingreso_pdf(ficha)

    # 🔔 Notificar a RRHH
    rrhh_usuarios = CustomUser.objects.filter(roles__nombre='rrhh')
    for rrhh in rrhh_usuarios:
        crear_notificacion(
            usuario=rrhh,
            mensaje=f"El trabajador {request.user.get_full_name()} ha firmado su ficha de ingreso.",
            url=reverse('rrhh:listar_fichas_ingreso_admin', args=[ficha.id]),
            tipo='success'
        )

    # 🔔 Notificar al PM asignado
    if ficha.pm:
        crear_notificacion(
            usuario=ficha.pm,
            mensaje=f"{request.user.get_full_name()} ha firmado su ficha de ingreso.",
            url=reverse('rrhh:revisar_ficha_pm', args=[ficha.id]),
            tipo='info'
        )

    messages.success(request, "✅ Has aprobado y firmado tu ficha de ingreso.")
    return redirect('dashboard:mis_fichas_ingreso')


@login_required
@rol_requerido('usuario')
def rechazar_ficha_ingreso_trabajador(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id, usuario=request.user)

    if request.method == 'POST':
        motivo = request.POST.get('motivo', '').strip()

        if not motivo:
            messages.error(
                request, "Debes ingresar un motivo para rechazar la ficha.")
            return redirect('rrhh:mis_fichas_ingreso')

        ficha.estado = 'rechazada_usuario'
        ficha.motivo_rechazo_usuario = motivo
        ficha.save()

        # 🔔 Notificar a RRHH
        rrhh_usuarios = CustomUser.objects.filter(
            is_active=True, is_staff=True)

        for rrhh in rrhh_usuarios:
            crear_notificacion(
                usuario=rrhh,
                mensaje=f"El trabajador {request.user.get_full_name()} ha rechazado su ficha de ingreso.",
                # ✅ Sin argumentos
                url=reverse('rrhh:listar_fichas_ingreso_admin'),
                tipo='warning'
            )

        messages.warning(
            request, "❌ Has rechazado la ficha correctamente. RRHH ha sido notificado.")
        return redirect('rrhh:mis_fichas_ingreso')

    return redirect('rrhh:mis_fichas_ingreso')


@rol_requerido('admin', 'pm')
def revisar_ficha_pm(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id)

    if request.method == 'POST':
        accion = request.POST.get('accion')

        if accion == 'aprobar':
            ficha.estado = 'pendiente_usuario'  # ahora espera la aprobación del trabajador
            ficha.save()

            # 🔔 Notificar al trabajador
            if ficha.usuario:
                crear_notificacion(
                    usuario=ficha.usuario,
                    mensaje="Tu ficha de ingreso ha sido aprobada por el PM. Ahora debes revisarla y firmarla.",
                    url=reverse('rrhh:listar_fichas_ingreso_usuario'),
                    tipo='info'
                )

            messages.success(
                request, "Ficha aprobada y enviada al trabajador para su validación.")

        elif accion == 'rechazar':
            ficha.estado = 'rechazada_pm'
            ficha.save()

            # 🔔 Notificar a RRHH
            rrhh_usuarios = CustomUser.objects.filter(
                rol='rrhh')  # o usa es_rrhh=True
            for rrhh in rrhh_usuarios:
                crear_notificacion(
                    usuario=rrhh,
                    mensaje=f"La ficha de {ficha.usuario.get_full_name()} fue rechazada por el PM.",
                    # Ajusta si tienes detalle de ficha
                    url=reverse('rrhh:listar_fichas_ingreso_admin'),
                    tipo='warning'
                )

            messages.warning(
                request, "Ficha rechazada. Recursos Humanos ha sido notificado.")

        return redirect('rrhh:listar_fichas_ingreso_admin')

    return render(request, 'rrhh/revision_ficha_pm.html', {
        'ficha': ficha
    })


@login_required
@rol_requerido('pm', 'admin')
def firmar_ficha_pm(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id)

    # No permitir firmar si ya está finalizada o rechazada
    if ficha.estado in ['rechazada_pm', 'rechazada_usuario', 'aprobada']:
        messages.error(
            request, "No se puede firmar una ficha ya finalizada o rechazada.")
        return redirect('rrhh:listar_fichas_ingreso_admin')

    # Validar que RRHH tenga firma digital
    if not ficha.creado_por or not ficha.creado_por.firma_digital:
        messages.error(
            request, "El responsable de RRHH aún no ha registrado su firma. No se puede continuar.")
        return redirect('rrhh:listar_fichas_ingreso_admin')

    # Validar que el PM actual tenga firma
    if not request.user.firma_digital:
        messages.error(
            request, "Debes tener una firma digital registrada para aprobar la ficha.")
        return redirect('rrhh:listar_fichas_ingreso_admin')

    # Verificar si ya fue firmada por el PM
    if ficha.firma_pm:
        messages.info(request, "Ya has firmado esta ficha.")
        return redirect('rrhh:listar_fichas_ingreso_admin')

    # Guardar firma, PM y actualizar estado
    ficha.firma_pm = request.user.firma_digital
    ficha.pm = request.user
    ficha.estado = 'pendiente_usuario'
    ficha.motivo_rechazo = None  # Limpiar rechazo anterior si lo hubo
    ficha.save()

    # 🔔 Notificar al trabajador que debe revisar y firmar
    if ficha.usuario:
        crear_notificacion(
            usuario=ficha.usuario,
            mensaje="Tu ficha de ingreso ha sido firmada por el PM. Revisa y firma para completar el proceso.",
            url=reverse('rrhh:mis_fichas_ingreso'),
            tipo='info'
        )

    messages.success(
        request, "✅ Ficha firmada correctamente. Ahora debe revisarla el trabajador.")
    return redirect('rrhh:listar_fichas_ingreso_admin')


@login_required
@rol_requerido('usuario', 'admin')
def firmar_ficha_ingreso(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id, usuario=request.user)

    # Validar que el PM haya aprobado
    if not ficha.firma_pm:
        messages.error(
            request, "No puedes firmar esta ficha aún. Falta la aprobación del PM.")
        return redirect('rrhh:listar_fichas_ingreso_usuario')

    # Validar si falta alguna firma
    errores = []

    if not request.user.firma_digital or not hasattr(request.user.firma_digital, 'url'):
        errores.append(
            "Debes registrar tu firma digital antes de poder firmar la ficha.")
    if not ficha.pm or not ficha.pm.firma_digital:
        errores.append("El PM asignado aún no ha registrado su firma.")
    if not ficha.creado_por or not ficha.creado_por.firma_digital:
        errores.append("El encargado de RRHH aún no ha registrado su firma.")

    if errores:
        for e in errores:
            messages.error(request, e)

        fichas = FichaIngreso.objects.filter(
            usuario=request.user).order_by('-id')
        return render(request, 'rrhh/listar_fichas_ingreso_usuario.html', {'fichas': fichas})

    # Si es POST, firmamos
    if request.method == 'POST':
        ficha.firma_trabajador = request.user.firma_digital
        ficha.firma_rrhh = ficha.creado_por.firma_digital
        ficha.firma_pm = ficha.pm.firma_digital

        try:
            from rrhh.utils import firmar_ficha_ingreso_pdf
            firmar_ficha_ingreso_pdf(ficha)

            ficha.estado = 'aprobada'
            ficha.save()

            # 🔔 Notificar a RRHH que el trabajador firmó
            rrhh_usuarios = CustomUser.objects.filter(roles__nombre='rrhh')
            for rrhh in rrhh_usuarios:
                crear_notificacion(
                    usuario=rrhh,
                    mensaje=f"La ficha de ingreso de {request.user.get_full_name()} ha sido firmada por el trabajador.",
                    url=reverse('rrhh:listar_fichas_ingreso_admin'),
                    tipo='info'
                )

            messages.success(request, "✅ Ficha firmada correctamente.")
        except Exception as e:
            messages.error(
                request, f"No se pudo completar la firma del PDF: {e}")

        return redirect('rrhh:listar_fichas_ingreso_usuario')

    return redirect('rrhh:listar_fichas_ingreso_usuario')


@login_required
@rol_requerido('usuario', 'admin')
def rechazar_ficha_ingreso(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id, usuario=request.user)

    if ficha.firma_pm and not ficha.firma_trabajador:
        ficha.estado = 'rechazada_usuario'
        ficha.save()

        # 🔔 Notificar a RRHH
        rrhh_usuarios = CustomUser.objects.filter(roles__nombre='rrhh')
        for rrhh in rrhh_usuarios:
            crear_notificacion(
                usuario=rrhh,
                mensaje=f"La ficha de ingreso de {request.user.get_full_name()} fue rechazada por el trabajador.",
                url=reverse('rrhh:listar_fichas_ingreso_admin'),
                tipo='warning'
            )

        messages.warning(
            request, "❌ Has rechazado la ficha. RRHH ha sido notificado.")
    else:
        messages.error(
            request, "⚠️ No puedes rechazar esta ficha en su estado actual.")

    return redirect('rrhh:listar_fichas_ingreso_usuario')


@staff_member_required
@rol_requerido('admin', 'pm', 'rrhh')
def eliminar_ficha_ingreso(request, pk):
    ficha = get_object_or_404(FichaIngreso, pk=pk)

    if request.method == 'POST':
        try:
            if ficha.archivo and ficha.archivo.name:
                # ✅ Elimina el PDF de Cloudinary
                ficha.archivo.delete(save=False)

            ficha.delete()
            messages.success(request, "Ficha eliminada correctamente.")
        except Exception as e:
            messages.error(
                request, f"Ocurrió un error al eliminar la ficha: {e}")

        return redirect('rrhh:listar_fichas_ingreso_admin')

    return render(request, 'rrhh/eliminar_ficha_ingreso.html', {'ficha': ficha})


@login_required
@rol_requerido('admin', 'pm', 'rrhh')
def generar_ficha_pdf(request, ficha_id):
    ficha = get_object_or_404(FichaIngreso, id=ficha_id)

    # Generar PDF temporal
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("FICHA DE INGRESO DE PERSONAL", styles['Heading1']),
        Spacer(1, 12),
        Paragraph(
            f"Nombre: {ficha.nombres} {ficha.apellidos}", styles['Normal']),
        Paragraph(f"RUT: {ficha.rut}", styles['Normal']),
        Paragraph(f"Cargo: {ficha.cargo}", styles['Normal']),
        Paragraph(f"Proyecto: {ficha.faena}", styles['Normal']),
        # ... más campos según corresponda ...
    ]
    doc.build(elements)
    buffer.seek(0)

    # Nombre de archivo lógico para descarga
    nombre_archivo = f"FichaIngreso_{ficha.rut.replace('.', '').replace('-', '')}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=nombre_archivo)


def exportar_fichas_ingreso_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fichas de Ingreso"

    # Encabezados organizados por sección del formulario
    encabezados = [
        'Usuario', 'Nombres', 'Apellidos', 'RUT', 'Fecha de nacimiento', 'Edad',
        'Sexo', 'Estado civil', 'Nacionalidad', 'Hijos', 'Nivel estudios',
        'Profesión u oficio',
        'Dirección', 'Comuna', 'Región',
        'Teléfono', 'Correo electrónico',
        'Nombre contacto emergencia', 'Teléfono emergencia', 'Parentesco emergencia',
        'Dirección emergencia',
        'AFP', 'Salud',
        'Banco 1', 'Tipo cuenta 1', 'Número cuenta 1',
        'Banco 2', 'Tipo cuenta 2', 'Número cuenta 2',
        'Cargo', 'Jefe directo', 'Proyecto', 'Fecha inicio',
        'Tipo contrato', 'Jornada', 'Sueldo base', 'Bono', 'Colación', 'Movilización',
        'Talla polera', 'Talla pantalón', 'Talla zapato', 'Observaciones',
        'Estado ficha', 'Motivo rechazo PM', 'Motivo rechazo usuario',
    ]
    ws.append(encabezados)

    # Obtener fichas
    fichas = FichaIngreso.objects.select_related('usuario').all()

    # Agregar datos
    for ficha in fichas:
        ws.append([
            ficha.usuario.get_full_name() if ficha.usuario else '',
            ficha.nombres,
            ficha.apellidos,
            ficha.rut,
            ficha.fecha_nacimiento.strftime(
                '%d-%m-%Y') if ficha.fecha_nacimiento else '',
            ficha.edad,
            ficha.sexo,
            ficha.estado_civil,
            ficha.nacionalidad,
            ficha.hijos,
            ficha.nivel_estudios,
            ficha.profesion_u_oficio,
            ficha.direccion,
            ficha.comuna,
            ficha.region,
            ficha.telefono,
            ficha.email,
            ficha.nombre_contacto_emergencia,
            ficha.telefono_emergencia,
            ficha.parentesco_emergencia,
            ficha.direccion_emergencia,
            ficha.afp,
            ficha.salud,
            ficha.banco,
            ficha.tipo_cuenta,
            ficha.numero_cuenta,
            ficha.banco_2,
            ficha.tipo_cuenta_2,
            ficha.numero_cuenta_2,
            ficha.cargo,
            ficha.jefe_directo,
            ficha.proyecto,
            ficha.fecha_inicio.strftime(
                '%d-%m-%Y') if ficha.fecha_inicio else '',
            ficha.tipo_contrato,
            ficha.jornada,
            ficha.sueldo_base,
            ficha.bono,
            ficha.colacion,
            ficha.movilizacion,
            ficha.talla_polera,
            ficha.talla_pantalon,
            ficha.talla_zapato,
            ficha.observaciones,
            ficha.get_estado_display(),
            ficha.motivo_rechazo_pm or '',
            ficha.motivo_rechazo_usuario or '',
        ])

    # Crear hoja de resumen
    ws_resumen = wb.create_sheet(title="Resumen")

    resumenes = {
        "Sexo": Counter(ficha.sexo or "No definido" for ficha in fichas),
        "Nacionalidad": Counter(ficha.nacionalidad or "No definida" for ficha in fichas),
        "AFP": Counter(ficha.afp or "No definida" for ficha in fichas),
        "Salud": Counter(ficha.salud or "No definida" for ficha in fichas),
        "Tipo de Contrato": Counter(ficha.tipo_contrato or "No definido" for ficha in fichas),
        "Talla Polera": Counter(ficha.talla_polera or "No definida" for ficha in fichas),
        "Talla Pantalón": Counter(ficha.talla_pantalon or "No definida" for ficha in fichas),
        "Talla Zapato": Counter(ficha.talla_zapato or "No definida" for ficha in fichas),
    }

    fila = 1
    for titulo, conteo in resumenes.items():
        ws_resumen.cell(row=fila, column=1,
                        value=titulo).font = Font(bold=True)
        fila += 1
        ws_resumen.cell(row=fila, column=1,
                        value="Valor").font = Font(bold=True)
        ws_resumen.cell(row=fila, column=2,
                        value="Cantidad").font = Font(bold=True)
        fila += 1
        for valor, cantidad in conteo.items():
            ws_resumen.cell(row=fila, column=1, value=valor)
            ws_resumen.cell(row=fila, column=2, value=cantidad)
            fila += 1
        fila += 2

    # Generar respuesta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=FichasIngreso.xlsx'
    wb.save(response)
    return response


@login_required
@rol_requerido('usuario')
def mis_vacaciones(request):
    usuario = request.user

    # Llamar al método del modelo para obtener días disponibles
    dias_disponibles = usuario.obtener_dias_vacaciones_disponibles()

    # Obtener los días cargados manualmente si existen
    try:
        dias_manuales = usuario.vacaciones_manuales.cantidad_dias
    except DiasVacacionesTomadosManualmente.DoesNotExist:
        dias_manuales = 0

    if request.method == 'POST':
        form = SolicitudVacacionesForm(request.POST, usuario=usuario)
        if form.is_valid():
            solicitud = form.save(commit=False)
            solicitud.usuario = usuario
            solicitud.dias_solicitados = form.cleaned_data.get(
                'dias_solicitados', 0)
            solicitud.save()

            # Notificar a todos los supervisores si no hay uno asignado
            if usuario.supervisor:
                crear_notificacion(
                    usuario=usuario.supervisor,
                    mensaje=f"{usuario.get_full_name()} ha solicitado vacaciones.",
                    url=reverse('rrhh:revisar_supervisor'),
                    tipo='info'
                )
            else:
                supervisores = CustomUser.objects.filter(
                    roles__nombre='supervisor').distinct()
                for supervisor in supervisores:
                    crear_notificacion(
                        usuario=supervisor,
                        mensaje=f"{usuario.get_full_name()} ha solicitado vacaciones.",
                        url=reverse('rrhh:revisar_supervisor'),
                        tipo='info'
                    )

            messages.success(request, "Solicitud enviada correctamente.")
            return redirect('rrhh:mis_vacaciones')

    else:
        form = SolicitudVacacionesForm(usuario=usuario)

    solicitudes = SolicitudVacaciones.objects.filter(
        usuario=usuario).order_by('-fecha_solicitud')

    feriados = Feriado.objects.values_list('fecha', flat=True)
    feriados_json = json.dumps([f.strftime("%Y-%m-%d") for f in feriados])

    context = {
        'dias_disponibles': dias_disponibles,
        'form': form,
        'solicitudes': solicitudes,
        'feriados_json': feriados_json,
        'dias_manuales': dias_manuales
    }

    return render(request, 'rrhh/solicitud_vacaciones.html', context)


@login_required
def editar_solicitud_vacaciones(request, pk):
    solicitud = get_object_or_404(
        SolicitudVacaciones, pk=pk, usuario=request.user)

    if solicitud.estatus != 'pendiente_supervisor':
        mensajes_estado = {
            'rechazada_supervisor': "No puedes editar una solicitud que fue rechazada por el supervisor.",
            'pendiente_pm': "No puedes editar una solicitud que ya fue enviada al PM.",
            'rechazada_pm': "No puedes editar una solicitud que fue rechazada por el PM.",
            'pendiente_rrhh': "No puedes editar una solicitud que ya fue enviada a RRHH.",
            'rechazada_rrhh': "No puedes editar una solicitud que fue rechazada por RRHH.",
            'rechazada_admin': "No puedes editar una solicitud que fue rechazada por administración.",
            'aprobada': "No puedes editar una solicitud que ya fue aprobada.",
        }
        mensaje = mensajes_estado.get(
            solicitud.estatus, "No puedes editar esta solicitud.")
        messages.warning(request, mensaje)
        return redirect('rrhh:mis_vacaciones')

    if request.method == 'POST':
        form = SolicitudVacacionesForm(request.POST, instance=solicitud)
        if form.is_valid():
            solicitud = form.save(commit=False)

            # Recalcular los días hábiles
            fecha_inicio = form.cleaned_data['fecha_inicio']
            fecha_fin = form.cleaned_data['fecha_fin']
            dias = contar_dias_habiles(fecha_inicio, fecha_fin)

            solicitud.dias_solicitados = dias
            solicitud.save()
            messages.success(request, "Solicitud actualizada correctamente.")
            return redirect('rrhh:mis_vacaciones')
    else:
        form = SolicitudVacacionesForm(instance=solicitud)

    return render(request, 'rrhh/editar_solicitud_vacaciones.html', {
        'form': form,
        'solicitud': solicitud
    })


@login_required
def eliminar_solicitud_vacaciones(request, pk):
    solicitud = get_object_or_404(
        SolicitudVacaciones, pk=pk, usuario=request.user)

    if solicitud.estatus != 'pendiente_supervisor':
        mensajes_estado = {
            'rechazada_supervisor': "No puedes eliminar una solicitud rechazada por el supervisor.",
            'pendiente_pm': "No puedes eliminar una solicitud que ya fue enviada al PM.",
            'rechazada_pm': "No puedes eliminar una solicitud rechazada por el PM.",
            'pendiente_rrhh': "No puedes eliminar una solicitud que ya fue enviada a RRHH.",
            'rechazada_rrhh': "No puedes eliminar una solicitud rechazada por RRHH.",
            'rechazada_admin': "No puedes eliminar una solicitud rechazada por administración.",
            'aprobada': "No puedes eliminar una solicitud que ya fue aprobada.",
        }
        mensaje = mensajes_estado.get(
            solicitud.estatus, "No puedes eliminar esta solicitud.")
        messages.warning(request, mensaje)
        return redirect('rrhh:mis_vacaciones')

    if request.method == 'POST':
        solicitud.delete()
        messages.success(request, "Solicitud eliminada correctamente.")
        return redirect('rrhh:mis_vacaciones')

    return render(request, 'rrhh/confirmar_eliminacion.html', {
        'solicitud': solicitud
    })

# --- Vista para Supervisor ---


@staff_member_required
@rol_requerido('supervisor', 'admin')
def revisar_solicitudes_supervisor(request):
    if not request.user.es_supervisor and not request.user.es_admin_general:
        return HttpResponseForbidden("No tienes permiso para ver esta vista.")

    solicitudes = SolicitudVacaciones.objects.filter(
        estatus='pendiente_supervisor').order_by('fecha_solicitud')

    return render(request, 'rrhh/revisar_vacaciones_supervisor.html', {
        'solicitudes': solicitudes,
        'titulo': "Solicitudes Pendientes - Supervisor",
        'rol': 'supervisor',
    })

# --- Vista para PM ---


@staff_member_required
@rol_requerido('pm', 'admin')
def revisar_solicitudes_pm(request):
    if not request.user.es_pm and not request.user.es_admin_general:
        return HttpResponseForbidden("No tienes permiso para ver esta vista.")

    solicitudes = SolicitudVacaciones.objects.filter(
        estatus='pendiente_pm').order_by('fecha_solicitud')

    return render(request, 'rrhh/revisar_vacaciones_pm.html', {
        'solicitudes': solicitudes,
        'titulo': "Solicitudes Pendientes - PM",
        'rol': 'pm',
    })

# --- Vista para RRHH ---


@staff_member_required
@rol_requerido('rrhh', 'admin')
def revisar_solicitudes_rrhh(request):
    if not request.user.es_rrhh and not request.user.es_admin_general:
        return HttpResponseForbidden("No tienes permiso para ver esta vista.")

    # Recoger filtros desde GET
    identidad = request.GET.get('identidad', '').strip()
    nombre = request.GET.get('nombre', '').strip()
    estatus = request.GET.get('estatus', '').strip()

    # Filtros base
    solicitudes = SolicitudVacaciones.objects.all()

    if identidad:
        solicitudes = solicitudes.filter(
            usuario__identidad__icontains=identidad)

    if nombre:
        solicitudes = solicitudes.filter(
            Q(usuario__first_name__icontains=nombre) |
            Q(usuario__last_name__icontains=nombre)
        )

    if estatus:
        solicitudes = solicitudes.filter(estatus=estatus)

    solicitudes = solicitudes.order_by('-fecha_solicitud')

    return render(request, 'rrhh/revisar_vacaciones_rrhh.html', {
        'solicitudes': solicitudes,
        'filtros': {
            'identidad': identidad,
            'nombre': nombre,
            'estatus': estatus,
        }
    })


@staff_member_required
@rol_requerido('supervisor', 'pm', 'rrhh', 'admin')
def revisar_solicitud(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudVacaciones, pk=solicitud_id)

    # Detectar rol activo
    if request.user.es_supervisor:
        rol = 'supervisor'
    elif request.user.es_pm:
        rol = 'pm'
    elif request.user.es_rrhh or request.user.es_admin_general:
        rol = 'rrhh'
    else:
        return HttpResponseForbidden("No tienes permiso para revisar esta solicitud.")

    transiciones_validas = {
        'supervisor': 'pendiente_supervisor',
        'pm': 'pendiente_pm',
        'rrhh': 'pendiente_rrhh',
    }

    if transiciones_validas.get(rol) != solicitud.estatus:
        return HttpResponseForbidden("No puedes revisar esta solicitud.")

    if request.method == 'POST':
        form = RevisionVacacionesForm(request.POST)
        if form.is_valid():
            accion = request.POST.get('accion')
            observacion = form.cleaned_data['observacion']
            solicitud.observacion = observacion

            if accion == 'aprobar':
                if rol == 'supervisor':
                    solicitud.estatus = 'pendiente_pm'
                elif rol == 'pm':
                    solicitud.estatus = 'pendiente_rrhh'
                elif rol == 'rrhh':
                    solicitud.estatus = 'aprobada'
            elif accion == 'rechazar':
                solicitud.estatus = f'rechazada_{rol}'

            solicitud.save()
            messages.success(
                request, f"Solicitud {'aprobada' if accion == 'aprobar' else 'rechazada'} exitosamente.")
            return redirect('rrhh:revisar_' + rol)
    else:
        form = RevisionVacacionesForm()

    return render(request, 'rrhh/revisar_solicitud_vacaciones.html', {
        'solicitud': solicitud,
        'form': form
    })


@staff_member_required
# ✅ Solo admin general y RRHH pueden acceder
@rol_requerido('admin', 'rrhh')
def revisar_todas_vacaciones(request):
    if not request.user.es_admin_general and not request.user.es_rrhh:
        return HttpResponseForbidden("No tienes permiso para ver esta vista.")

    solicitudes = SolicitudVacaciones.objects.all().order_by('-fecha_solicitud')

    return render(request, 'rrhh/revisar_todas_vacaciones.html', {
        'solicitudes': solicitudes,
        'titulo': "Todas las Solicitudes de Vacaciones",
        'rol': 'rrhh' if request.user.es_rrhh else 'admin',
    })


@staff_member_required
@rol_requerido('supervisor', 'pm', 'rrhh', 'admin')
def rechazar_solicitud_vacaciones(request):
    if request.method == 'POST':
        solicitud_id = request.POST.get('solicitud_id')
        observacion = request.POST.get('observacion')
        solicitud = get_object_or_404(SolicitudVacaciones, pk=solicitud_id)

        # Detectar el rol de mayor privilegio permitido
        if request.user.es_supervisor:
            rol_usuario = 'supervisor'
        elif request.user.es_pm:
            rol_usuario = 'pm'
        elif request.user.es_rrhh:
            rol_usuario = 'rrhh'
        elif request.user.es_admin_general:
            rol_usuario = 'admin'
        else:
            return HttpResponseForbidden("Rol no válido para esta acción.")

        # Definir estatus de rechazo
        estatus_rechazo = {
            'supervisor': 'rechazada_supervisor',
            'pm': 'rechazada_pm',
            'rrhh': 'rechazada_rrhh',
            'admin': 'rechazada_admin'
        }[rol_usuario]

        solicitud.estatus = estatus_rechazo
        solicitud.observacion = observacion

        # Registrar quién rechazó
        if rol_usuario == 'supervisor':
            solicitud.aprobado_por_supervisor = request.user
        elif rol_usuario == 'pm':
            solicitud.aprobado_por_pm = request.user
        elif rol_usuario == 'rrhh':
            solicitud.aprobado_por_rrhh = request.user

        solicitud.save()

        # 🔔 Notificar al trabajador del rechazo
        crear_notificacion(
            usuario=solicitud.usuario,
            mensaje=f"❌ Tu solicitud de vacaciones fue rechazada por {rol_usuario.upper()}. Motivo: {observacion}",
            url=reverse('rrhh:mis_vacaciones'),
            tipo='error'
        )

        messages.success(
            request, f"Solicitud rechazada por {rol_usuario.upper()}.")

        redirecciones = {
            'supervisor': 'rrhh:revisar_supervisor',
            'pm': 'rrhh:revisar_pm',
            'rrhh': 'rrhh:revisar_rrhh',
            'admin': 'rrhh:revisar_todas_vacaciones'
        }

        return redirect(reverse(redirecciones[rol_usuario]))

    return HttpResponseForbidden("Acceso denegado.")


@rol_requerido('supervisor', 'admin')
def aprobar_vacacion_supervisor(request, pk):
    solicitud = get_object_or_404(SolicitudVacaciones, pk=pk)

    if solicitud.estatus == 'pendiente_supervisor':
        solicitud.estatus = 'pendiente_pm'
        solicitud.aprobado_por_supervisor = request.user
        solicitud.save()

        # 🔔 Notificar al PM (si hay uno asignado, o a todos los PM)
        mensaje_pm = f"{solicitud.usuario.get_full_name()} realizó una solicitud de vacaciones la cual está pendiente por ti."
        url_pm = reverse('rrhh:revisar_pm')

        if solicitud.usuario.pm:
            crear_notificacion(
                usuario=solicitud.usuario.pm,
                mensaje=mensaje_pm,
                url=url_pm,
                tipo='info'
            )
        else:
            pms = CustomUser.objects.filter(roles__nombre='pm').distinct()
            for pm in pms:
                crear_notificacion(
                    usuario=pm,
                    mensaje=mensaje_pm,
                    url=url_pm,
                    tipo='info'
                )

        # 🔔 Notificar al trabajador que avanzó al siguiente paso
        crear_notificacion(
            usuario=solicitud.usuario,
            mensaje="Tu solicitud de vacaciones fue aprobada por el Supervisor y está pendiente por aprobación del PM.",
            url=reverse('rrhh:mis_vacaciones'),
            tipo='success'
        )

        messages.success(request, "Solicitud aprobada y enviada al PM.")
    else:
        messages.warning(
            request, "La solicitud no está pendiente para Supervisor.")

    return redirect('rrhh:revisar_supervisor')


@rol_requerido('pm', 'admin')
def aprobar_vacacion_pm(request, pk):
    solicitud = get_object_or_404(SolicitudVacaciones, pk=pk)

    if solicitud.estatus == 'pendiente_pm':
        solicitud.estatus = 'pendiente_rrhh'
        solicitud.aprobado_por_pm = request.user
        solicitud.save()

        # 🔔 Notificar a todos los usuarios con rol 'rrhh'
        mensaje_rrhh = f"{solicitud.usuario.get_full_name()} realizó una solicitud de vacaciones la cual está pendiente por ti."
        url_rrhh = reverse('rrhh:revisar_rrhh')

        rrhhs = CustomUser.objects.filter(roles__nombre='rrhh').distinct()
        for rrhh in rrhhs:
            crear_notificacion(
                usuario=rrhh,
                mensaje=mensaje_rrhh,
                url=url_rrhh,
                tipo='info'
            )

        # 🔔 Notificar al trabajador
        crear_notificacion(
            usuario=solicitud.usuario,
            mensaje="Tu solicitud de vacaciones fue aprobada por el PM y está pendiente por RRHH.",
            url=reverse('rrhh:mis_vacaciones'),
            tipo='success'
        )

        messages.success(request, "Solicitud aprobada y enviada a RRHH.")
    else:
        messages.warning(request, "La solicitud no está pendiente para PM.")

    return redirect('rrhh:revisar_pm')


@staff_member_required
@rol_requerido('admin', 'rrhh')
def aprobar_vacacion_rrhh(request, pk):
    solicitud = get_object_or_404(SolicitudVacaciones, pk=pk)

    if solicitud.estatus != 'pendiente_rrhh':
        messages.error(request, "Esta solicitud ya fue revisada.")
        return redirect('rrhh:revisar_rrhh')

    trabajador = solicitud.usuario
    pm = solicitud.aprobado_por_pm
    rrhh = request.user

    # Validación de firmas
    faltantes = []
    if not trabajador.firma_digital:
        faltantes.append("del trabajador")
    if not pm or not pm.firma_digital:
        faltantes.append("del jefe directo")
    if not rrhh.firma_digital:
        faltantes.append("de Recursos Humanos")

    if faltantes:
        mensaje = "❌ No se puede completar la aprobación. Faltan las firmas " + \
            ", ".join(faltantes) + "."
        messages.error(request, mensaje)
        return redirect('rrhh:revisar_rrhh')

    # Aprobación y cambio de estado
    solicitud.estatus = 'aprobada'
    solicitud.aprobado_por_rrhh = rrhh
    solicitud.save()

    try:
        generar_pdf_solicitud_vacaciones(solicitud)
        messages.success(
            request, "✅ Solicitud aprobada y documento generado correctamente.")
    except Exception as e:
        print(f"⚠️ Error al generar el PDF: {e}")
        messages.warning(
            request, f"Solicitud aprobada, pero hubo un error al generar el documento PDF: {e}")

    # 🔔 Notificar al trabajador
    crear_notificacion(
        usuario=trabajador,
        mensaje="🎉 Tu solicitud de vacaciones fue aprobada por RRHH.",
        url=reverse('rrhh:mis_vacaciones'),
        tipo='success'
    )

    return redirect('rrhh:revisar_rrhh')


@staff_member_required
@rol_requerido('admin', 'rrhh')
def eliminar_solicitud_vacaciones_admin(request, pk):
    # Seguridad extra por si acaso alguien sin rol llega hasta aquí
    if not request.user.es_rrhh and not request.user.es_admin_general:
        messages.error(
            request, "No tienes permisos para eliminar esta solicitud.")
        return redirect('dashboard_admin:vacaciones_admin')

    solicitud = get_object_or_404(SolicitudVacaciones, pk=pk)

    if request.method == 'POST':
        try:
            # Elimina el archivo PDF si existe
            if solicitud.archivo_pdf and solicitud.archivo_pdf.name:
                solicitud.archivo_pdf.delete(save=False)

            # Elimina la solicitud
            solicitud.delete()
            messages.success(request, "Solicitud eliminada correctamente.")
        except Exception as e:
            messages.error(request, f"Ocurrió un error al eliminar: {e}")

        return redirect('dashboard_admin:vacaciones_admin')

    messages.warning(request, "La eliminación debe hacerse mediante POST.")
    return redirect('dashboard_admin:vacaciones_admin')


@rol_requerido('admin', 'rrhh')
def subir_documento_trabajador(request):
    if request.method == 'POST':
        form = DocumentoTrabajadorForm(request.POST, request.FILES)
        if form.is_valid():
            trabajador = form.cleaned_data['trabajador']
            tipo = form.cleaned_data['tipo_documento']
            archivo = form.cleaned_data['archivo']
            fecha_emision = form.cleaned_data['fecha_emision']
            fecha_vencimiento = form.cleaned_data['fecha_vencimiento']

            url = reverse('rrhh:mis_documentos')

            # Verificar si ya existe
            existente = DocumentoTrabajador.objects.filter(
                trabajador=trabajador, tipo_documento=tipo
            ).first()

            if existente:
                # Reemplazar el archivo
                existente.archivo = archivo
                existente.fecha_emision = fecha_emision
                existente.fecha_vencimiento = fecha_vencimiento
                existente.save()
                messages.success(
                    request, '📄 Documento reemplazado correctamente.')

                crear_notificacion(
                    usuario=trabajador,
                    mensaje=f"Se ha reemplazado tu documento '{tipo}'.",
                    url=url
                )
            else:
                form.save()
                messages.success(request, '📄 Documento subido correctamente.')

                crear_notificacion(
                    usuario=trabajador,
                    mensaje=f"Se ha subido un nuevo documento '{tipo}' a tu perfil.",
                    url=url
                )

            return redirect('rrhh:listado_documentos')
    else:
        form = DocumentoTrabajadorForm()

    return render(request, 'rrhh/subir_documento_trabajador.html', {'form': form})


def calcular_estado_documento(doc):
    if not doc or not doc.archivo or not doc.archivo.name:
        return "Faltante"

    if not doc.fecha_vencimiento:
        return "Faltante"

    hoy = date.today()

    if doc.fecha_vencimiento < hoy:
        return "Vencido"
    elif (doc.fecha_vencimiento - hoy).days <= 7:
        return "Por vencer"
    return "Vigente"


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def listado_documentos_trabajador(request):
    filtro_nombre = request.GET.get('trabajador', '').strip()
    filtro_tipo = request.GET.get('tipo', '')
    filtro_fecha = request.GET.get('fecha', '')
    filtro_estado = request.GET.get('estado', '')

    documentos = DocumentoTrabajador.objects.select_related(
        'trabajador', 'tipo_documento').all()

    # Aplicar filtros
    if filtro_nombre:
        documentos = documentos.filter(
            Q(trabajador__nombres__icontains=filtro_nombre) |
            Q(trabajador__apellidos__icontains=filtro_nombre)
        )
    if filtro_tipo:
        documentos = documentos.filter(tipo_documento__id=filtro_tipo)
    if filtro_fecha:
        documentos = documentos.filter(fecha_vencimiento=filtro_fecha)

    # Cargar todos los tipos de documentos para los filtros
    tipos = TipoDocumento.objects.all()

    data = []
    trabajadores_vistos = set()

    for doc in documentos:
        estado = calcular_estado_documento(doc)

        # Filtrar por estado después de calcularlo
        if filtro_estado and estado != filtro_estado:
            continue

        trabajador_id = doc.trabajador.id

        # Agrupar documentos por trabajador
        if trabajador_id not in trabajadores_vistos:
            trabajadores_vistos.add(trabajador_id)
            data.append({
                "trabajador": doc.trabajador,
                "documentos": []
            })

        # Agregar documento al trabajador correspondiente
        for entrada in data:
            if entrada["trabajador"].id == trabajador_id:
                entrada["documentos"].append({
                    "tipo": doc.tipo_documento,
                    "doc": doc,
                    "estado": estado
                })

    return render(request, 'rrhh/listado_documentos_trabajador.html', {
        "data": data,
        "tipos": tipos,
        "filtro_nombre": filtro_nombre,
        "filtro_tipo": filtro_tipo,
        "filtro_fecha": filtro_fecha,
        "filtro_estado": filtro_estado,
    })


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def crear_tipo_documento(request):
    tipo_id = request.GET.get('editar')
    tipo_a_editar = None

    if tipo_id:
        tipo_a_editar = get_object_or_404(TipoDocumento, pk=tipo_id)
        form = TipoDocumentoForm(request.POST or None, instance=tipo_a_editar)
    else:
        form = TipoDocumentoForm(request.POST or None)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(
                request, "Tipo de documento actualizado correctamente." if tipo_a_editar else "Tipo de documento creado correctamente.")
            return redirect('rrhh:crear_tipo_documento')
        else:
            messages.error(request, "Error al guardar el tipo de documento.")

    tipos = TipoDocumento.objects.all()
    return render(request, 'rrhh/crear_tipo_documento.html', {
        'form': form,
        'tipos': tipos,
        'editando': tipo_a_editar
    })


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def reemplazar_documento(request, documento_id):
    documento = get_object_or_404(DocumentoTrabajador, pk=documento_id)

    if request.method == 'POST':
        form = ReemplazoDocumentoForm(request.POST, request.FILES)
        if form.is_valid():
            nuevo_archivo = form.cleaned_data['archivo']

            # Eliminar archivo anterior en Cloudinary
            if documento.archivo:
                public_id = documento.archivo.name.rsplit('.', 1)[0]
                cloudinary.uploader.destroy(public_id, invalidate=True)

            # Guardar nuevo archivo
            identidad = documento.trabajador.identidad
            tipo_slug = slugify(documento.tipo_documento.nombre)
            filename = f"{tipo_slug}.pdf"
            path = f"Documentos de los trabajadores/{identidad}/{filename}"
            documento.archivo.save(path, ContentFile(nuevo_archivo.read()))

            # Actualizar fechas
            documento.fecha_emision = form.cleaned_data['fecha_emision']
            documento.fecha_vencimiento = form.cleaned_data['fecha_vencimiento']
            documento.save()

            messages.success(request, "Documento reemplazado correctamente.")
            return redirect('rrhh:listado_documentos')
        else:
            messages.error(request, "Hubo un error al subir el documento.")
    else:
        form = ReemplazoDocumentoForm()

    return render(request, 'rrhh/reemplazar_documento.html', {
        'form': form,
        'documento': documento
    })


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def exportar_documentos_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Documentos Trabajadores"

    # Cabeceras
    headers = [
        "Trabajador",
        "Identidad",
        "Correo",
        "Tipo de documento",
        "Fecha de carga",
        "Fecha de emisión",
        "Fecha de expiración",
        "Estado"
    ]
    ws.append(headers)

    documentos = DocumentoTrabajador.objects.select_related(
        'trabajador', 'tipo_documento')

    for doc in documentos:
        estado = calcular_estado_documento(doc)
        ws.append([
            doc.trabajador.get_full_name(),
            doc.trabajador.identidad,
            doc.trabajador.email,
            doc.tipo_documento.nombre,
            doc.creado.strftime('%Y-%m-%d') if doc.creado else "—",
            doc.fecha_emision.strftime(
                '%Y-%m-%d') if doc.fecha_emision else "—",
            doc.fecha_vencimiento.strftime(
                '%Y-%m-%d') if doc.fecha_vencimiento else "—",
            estado
        ])

    # Ajustar el ancho de columnas automáticamente
    for col_num, _ in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col_num)].width = 25

    # Preparar archivo Excel para descarga
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={smart_str("documentos_trabajadores.xlsx")}'
    wb.save(response)
    return response


@login_required(login_url='usuarios:login')
@rol_requerido('usuario')
def mis_documentos(request):
    documentos = DocumentoTrabajador.objects.filter(trabajador=request.user)
    documentos_info = []

    for doc in documentos:
        documentos_info.append({
            "tipo": doc.tipo_documento.nombre,
            "archivo": doc.archivo,
            "fecha_emision": doc.fecha_emision,
            "fecha_vencimiento": doc.fecha_vencimiento,
            "estado": calcular_estado_documento(doc)
        })

    return render(request, 'rrhh/mis_documentos.html', {
        "documentos": documentos_info
    })


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def eliminar_documento(request, id):
    documento = get_object_or_404(DocumentoTrabajador, id=id)

    if request.method == 'POST':
        try:
            if documento.archivo and documento.archivo.name:
                documento.archivo.delete(save=False)
            documento.delete()
            messages.success(
                request, "El documento fue eliminado correctamente.")
        except Exception as e:
            messages.error(
                request, f"Ocurrió un error al eliminar el documento: {e}")

        return redirect('rrhh:listado_documentos_trabajador')

    # Evita eliminar por GET accidental
    messages.error(request, "La eliminación debe hacerse mediante POST.")
    return redirect('rrhh:listado_documentos_trabajador')


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def eliminar_tipo_documento(request, pk):
    tipo = get_object_or_404(TipoDocumento, pk=pk)

    if request.method == 'POST':
        try:
            tipo.delete()
            messages.success(
                request, "✅ Tipo de documento eliminado correctamente.")
        except Exception as e:
            messages.error(request, f"❌ Error al eliminar: {e}")
        return redirect('rrhh:crear_tipo_documento')

    messages.warning(request, "⚠️ La eliminación debe hacerse mediante POST.")
    return redirect('rrhh:crear_tipo_documento')


@staff_member_required
@rol_requerido('rrhh', 'admin', 'pm')
def editar_cronograma_pago(request, usuario_id):
    usuario = get_object_or_404(CustomUser, id=usuario_id)
    cronograma, _ = CronogramaPago.objects.get_or_create(usuario=usuario)

    if request.method == 'POST':
        form = CronogramaPagoForm(request.POST, instance=cronograma)
        if form.is_valid():
            form.save()
            # Ajusta esta URL si es necesario
            return redirect('dashboard_admin:listar_usuarios')
    else:
        form = CronogramaPagoForm(instance=cronograma)

    return render(request, 'rrhh/editar_cronograma_pago.html', {
        'form': form,
        'usuario': usuario
    })


@login_required(login_url='usuarios:login')
@rol_requerido('usuario')
def ver_cronograma_pago(request):
    cronograma = CronogramaPago.objects.first()

    cronograma_mensual = []
    meses = [
        'enero', 'febrero', 'marzo', 'abril', 'mayo', 'junio',
        'julio', 'agosto', 'septiembre', 'octubre', 'noviembre', 'diciembre'
    ]

    for mes in meses:
        texto = getattr(cronograma, f"{mes}_texto")
        fecha = getattr(cronograma, f"{mes}_fecha")

        # Convertir texto a fecha si es string válido
        if texto:
            try:
                desde = datetime.strptime(texto, "%Y-%m-%d").date()
            except ValueError:
                desde = None
        else:
            desde = None

        cronograma_mensual.append({
            "mes": mes.capitalize(),
            "desde": desde,
            "hasta": fecha
        })

    context = {
        "cronograma": cronograma,
        "cronograma_mensual": cronograma_mensual
    }
    return render(request, "rrhh/ver_cronograma_pago.html", context)


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def cronograma_pago_admin(request):
    cronograma, _ = CronogramaPago.objects.get_or_create(id=1)
    meses = [
        "enero", "febrero", "marzo", "abril", "mayo", "junio",
        "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
    ]

    if request.method == "POST":
        for mes in meses:
            texto = request.POST.get(f"{mes}_texto", "").strip()
            fecha_str = request.POST.get(f"{mes}_fecha", "").strip()

            setattr(cronograma, f"{mes}_texto", texto)

            if fecha_str:
                try:
                    fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
                except ValueError:
                    fecha = None
            else:
                fecha = None

            setattr(cronograma, f"{mes}_fecha", fecha)

        cronograma.save()
        messages.success(request, "Cronograma actualizado correctamente.")
        return redirect('rrhh:cronograma_pago_admin')

    datos_meses = []
    for mes in meses:
        datos_meses.append({
            'nombre': mes,
            'texto': getattr(cronograma, f"{mes}_texto") or '',
            'fecha': getattr(cronograma, f"{mes}_fecha").strftime('%Y-%m-%d') if getattr(cronograma, f"{mes}_fecha") else '',
        })

    return render(request, 'rrhh/cronograma_pago_admin.html', {
        'cronograma': cronograma,
        'datos_meses': datos_meses
    })


@login_required(login_url='usuarios:login')
@rol_requerido('usuario')
def listar_adelantos_usuario(request):
    usuario = request.user

    ficha = FichaIngreso.objects.filter(usuario=usuario).first()
    hoy = date.today()
    sueldo_disponible = 0
    adelanto_mes = None

    if ficha and ficha.sueldo_base:
        maximo = ficha.sueldo_base * Decimal('0.5')
        total_aprobado = SolicitudAdelanto.objects.filter(
            trabajador=usuario,
            estado='aprobada',
            fecha_solicitud__month=hoy.month,
            fecha_solicitud__year=hoy.year
        ).aggregate(total=Sum('monto_aprobado'))['total'] or 0

        sueldo_disponible = int(maximo - total_aprobado)
        if sueldo_disponible < 0:
            sueldo_disponible = 0
    else:
        messages.error(
            request, "No tienes ficha de ingreso registrada con sueldo base.")
        sueldo_disponible = 0
        adelanto_mes = True  # bloquea el formulario

    # Bloquear si ya tiene solicitud aprobada este mes
    adelanto_mes = SolicitudAdelanto.objects.filter(
        trabajador=usuario,
        estado='aprobada',
        fecha_solicitud__month=hoy.month,
        fecha_solicitud__year=hoy.year
    ).exists()

    if request.method == 'POST':
        form = SolicitudAdelantoForm(
            request.POST, trabajador=usuario, monto_maximo=sueldo_disponible)
        # form = SolicitudAdelantoForm(request.POST, trabajador=usuario)
        if form.is_valid():
            monto_solicitado = form.cleaned_data['monto_solicitado']

            solicitud = form.save(commit=False)
            solicitud.trabajador = usuario
            solicitud.estado = 'pendiente_pm'
            solicitud.save()

            if monto_solicitado > sueldo_disponible:
                request.session['adelanto_supera_monto'] = True
            else:
                messages.success(request, "Solicitud enviada correctamente.")

            return redirect('rrhh:listar_adelantos_usuario')
    else:
        # form = SolicitudAdelantoForm(trabajador=usuario)
        form = SolicitudAdelantoForm(
            trabajador=usuario, monto_maximo=sueldo_disponible)

    if request.session.pop('adelanto_supera_monto', False):
        messages.warning(
            request,
            "El monto solicitado supera el disponible. Tu solicitud quedará sujeta a aprobación."
        )

    solicitudes = SolicitudAdelanto.objects.filter(
        trabajador=usuario).order_by('-fecha_solicitud')

    return render(request, 'rrhh/listar_adelantos_usuario.html', {
        'form': form,
        'sueldo_disponible': sueldo_disponible,
        'solicitudes': solicitudes,
        'bloqueado': adelanto_mes,
    })


@staff_member_required
@rol_requerido('pm', 'admin')
def aprobar_adelanto_pm(request, id):
    solicitud = get_object_or_404(SolicitudAdelanto, id=id)

    if solicitud.estado != 'pendiente_pm':
        messages.warning(
            request, "Esta solicitud ya fue gestionada por el PM.")
        return redirect('rrhh:listar_adelanto_admin')

    solicitud.estado = 'pendiente_rrhh'
    solicitud.aprobado_por_pm = request.user
    solicitud.save()

    messages.success(
        request, "Solicitud aprobada correctamente y enviada a RRHH.")
    return redirect('rrhh:listar_adelanto_admin')


@staff_member_required
@rol_requerido('pm', 'rrhh', 'admin')
def rechazar_adelanto_pm(request, id):
    solicitud = get_object_or_404(SolicitudAdelanto, id=id)

    if request.method == 'POST':
        motivo = request.POST.get('motivo_rechazo', '').strip()

        if not motivo:
            messages.error(request, "Debes indicar un motivo de rechazo.")
            return redirect(request.META.get('HTTP_REFERER', 'rrhh:listar_adelanto_admin'))

        if solicitud.estado == 'pendiente_pm':
            solicitud.estado = 'rechazada_pm'
        elif solicitud.estado == 'pendiente_rrhh':
            solicitud.estado = 'rechazada_rrhh'

        solicitud.motivo_rechazo = motivo
        solicitud.save()
        messages.success(request, "Solicitud rechazada correctamente.")
        return redirect('rrhh:listar_adelanto_admin')


@staff_member_required
@rol_requerido('rrhh', 'admin')
def aprobar_adelanto_rrhh(request, id):
    solicitud = get_object_or_404(SolicitudAdelanto, id=id)

    if request.method == 'POST':
        form = AprobacionAdelantoForm(request.POST, request.FILES)
        if form.is_valid():
            trabajador = solicitud.trabajador
            pm = solicitud.aprobado_por_pm
            rrhh = request.user

            # 🔒 Validar firmas
            errores = []
            if not trabajador.firma_digital:
                errores.append(
                    f"El trabajador {trabajador.get_full_name()} no tiene firma registrada.")
            if not pm or not pm.firma_digital:
                errores.append("El PM responsable no tiene firma registrada.")
            if not rrhh.firma_digital:
                errores.append("Tú como RRHH no tienes firma registrada.")

            if errores:
                for error in errores:
                    messages.error(request, error)
                return redirect(request.path)

            # ✅ Si todos tienen firma, continuar con la aprobación
            monto_final = form.cleaned_data['monto_aprobado']
            comprobante = form.cleaned_data['comprobante']

            solicitud.comprobante_transferencia = comprobante
            solicitud.monto_aprobado = monto_final
            solicitud.estado = 'aprobada'
            solicitud.aprobado_por_rrhh = rrhh
            solicitud.motivo_rechazo = ''  # Limpiar motivo anterior

            try:
                generar_pdf_solicitud_adelanto(solicitud)
            except Exception as e:
                messages.error(
                    request, f"No se pudo generar la planilla PDF: {e}")
                return redirect(request.path)

            solicitud.save()
            messages.success(
                request, "Solicitud aprobada y planilla generada correctamente.")
            return redirect('rrhh:listar_adelanto_admin')
        else:
            messages.error(request, "Formulario inválido. Revise los campos.")
    else:
        form = AprobacionAdelantoForm()

    return render(request, 'rrhh/aprobar_adelanto_rrhh.html', {
        'solicitud': solicitud,
        'form': form,
    })


@staff_member_required
@rol_requerido('admin', 'rrhh', 'pm')
def listar_adelanto_admin(request):
    solicitudes = SolicitudAdelanto.objects.select_related('trabajador')

    busqueda = request.GET.get('busqueda')
    mes = request.GET.get('mes')
    año = request.GET.get('año')

    if busqueda:
        solicitudes = solicitudes.filter(
            Q(trabajador__nombres__icontains=busqueda) |
            Q(trabajador__identidad__icontains=busqueda)
        )
    if mes:
        solicitudes = solicitudes.filter(fecha_solicitud__month=mes)
    if año:
        solicitudes = solicitudes.filter(fecha_solicitud__year=año)

    solicitudes = solicitudes.order_by('-fecha_solicitud')
    meses = [
        (1, 'enero'), (2, 'febrero'), (3, 'marzo'), (4, 'abril'),
        (5, 'mayo'), (6, 'junio'), (7, 'julio'), (8, 'agosto'),
        (9, 'septiembre'), (10, 'octubre'), (11, 'noviembre'), (12, 'diciembre')
    ]
    años = SolicitudAdelanto.objects.dates('fecha_solicitud', 'year').distinct(
    ).values_list('fecha_solicitud__year', flat=True)

    return render(request, 'rrhh/listar_adelanto_admin.html', {
        'solicitudes': solicitudes,
        'meses': meses,
        'años': años,
    })


@staff_member_required
@rol_requerido('admin', 'rrhh')
def eliminar_adelanto_admin(request, id):
    solicitud = get_object_or_404(SolicitudAdelanto, id=id)

    if request.user.rol == 'rrhh' and not solicitud.puede_editar_rrhh:
        messages.error(
            request, "No tienes permiso para eliminar esta solicitud.")
        return redirect('rrhh:listar_adelanto_admin')

    if request.method == 'POST':
        # 🧹 Eliminar archivos de Cloudinary si existen
        if solicitud.comprobante_transferencia and solicitud.comprobante_transferencia.name:
            solicitud.comprobante_transferencia.delete(save=False)

        if solicitud.planilla_pdf and solicitud.planilla_pdf.name:
            solicitud.planilla_pdf.delete(save=False)

        solicitud.delete()
        messages.success(
            request, "Solicitud y archivos eliminados correctamente.")
        return redirect('rrhh:listar_adelanto_admin')

    return render(request, 'rrhh/eliminar_adelanto_admin.html', {
        'solicitud': solicitud
    })


@staff_member_required
@rol_requerido('admin', 'rrhh')
def editar_adelanto_admin(request, id):
    solicitud = get_object_or_404(SolicitudAdelanto, id=id)

    if request.user.rol == 'rrhh' and not solicitud.puede_editar_rrhh:
        messages.error(
            request, "No tienes permiso para editar esta solicitud.")
        return redirect('rrhh:listar_adelanto_admin')

    monto_anterior = solicitud.monto_aprobado

    if request.method == 'POST':
        form = SolicitudAdelantoAdminForm(
            request.POST,
            request.FILES,
            instance=solicitud,
            trabajador=solicitud.trabajador,
            usuario_actual=request.user
        )

        reemplazar_comprobante = 'comprobante_transferencia' in request.FILES

        if form.is_valid():
            solicitud_editada = form.save(commit=False)
            nuevo_monto = solicitud_editada.monto_aprobado

            # ✅ Mantener estado y aprobadores actuales
            solicitud_editada.estado = solicitud.estado
            solicitud_editada.aprobado_por_pm = solicitud.aprobado_por_pm
            solicitud_editada.aprobado_por_rrhh = solicitud.aprobado_por_rrhh

            # 🧾 Reemplazar comprobante si se subió uno nuevo
            if reemplazar_comprobante:
                if solicitud.comprobante_transferencia and solicitud.comprobante_transferencia.storage.exists(solicitud.comprobante_transferencia.name):
                    solicitud.comprobante_transferencia.delete(save=False)

                solicitud_editada.comprobante_transferencia = request.FILES.get(
                    'comprobante_transferencia')

            # 🧾 Verificar si el monto cambió y debe regenerarse PDF
            regenerar_pdf = nuevo_monto != monto_anterior

            if regenerar_pdf:
                if solicitud.planilla_pdf and solicitud.planilla_pdf.storage.exists(solicitud.planilla_pdf.name):
                    solicitud.planilla_pdf.delete(save=False)
                solicitud_editada.planilla_pdf = None

            solicitud_editada.save()

            if regenerar_pdf:
                try:
                    generar_pdf_solicitud_adelanto(solicitud_editada)
                    messages.success(
                        request, "PDF regenerado con el nuevo monto.")
                except Exception as e:
                    messages.warning(
                        request, f"El nuevo PDF no pudo generarse: {e}"
                    )

            messages.success(request, "Solicitud actualizada correctamente.")
            return redirect('rrhh:listar_adelanto_admin')
        else:
            messages.error(request, "Formulario inválido.")
    else:
        form = SolicitudAdelantoAdminForm(
            instance=solicitud,
            trabajador=solicitud.trabajador,
            usuario_actual=request.user
        )

    return render(request, 'rrhh/editar_adelanto.html', {
        'form': form,
        'solicitud': solicitud
    })


@login_required(login_url='usuarios:login')
@rol_requerido('usuario')
def eliminar_adelanto_usuario(request, pk):
    solicitud = get_object_or_404(
        SolicitudAdelanto, pk=pk, trabajador=request.user)

    if request.method == 'POST':
        solicitud.delete()
        messages.success(request, "Solicitud eliminada correctamente.")
        return redirect('rrhh:listar_adelantos_usuario')

    return render(request, 'rrhh/eliminar_adelanto_usuario.html', {'solicitud': solicitud})


@login_required(login_url='usuarios:login')
@rol_requerido('usuario')
def editar_adelanto_usuario(request, pk):
    solicitud = get_object_or_404(
        SolicitudAdelanto, pk=pk, trabajador=request.user
    )

    # ✅ Permitimos editar si está en pendiente o ha sido rechazada en cualquier etapa
    estados_permitidos = ['pendiente_pm', 'rechazada_pm',
                          'rechazada_rrhh', 'rechazada_usuario']
    if solicitud.estado not in estados_permitidos:
        messages.error(
            request,
            "Solo puedes editar solicitudes en estado 'pendiente por aprobación PM' o que hayan sido rechazadas."
        )
        return redirect('rrhh:listar_adelantos_usuario')

    if request.method == 'POST':
        form = SolicitudAdelantoForm(request.POST, instance=solicitud)
        if form.is_valid():
            solicitud_editada = form.save(commit=False)

            # ✅ Si fue rechazada, borramos el motivo
            if solicitud.estado in ['rechazada_pm', 'rechazada_rrhh', 'rechazada_usuario']:
                solicitud_editada.motivo_rechazo = ""

            # Reiniciar el flujo
            solicitud_editada.estado = 'pendiente_pm'  # Reinicia el flujo
            solicitud_editada.monto_aprobado = None
            solicitud_editada.aprobado_por_pm = None
            solicitud_editada.aprobado_por_rrhh = None
            solicitud_editada.planilla_pdf = None
            solicitud_editada.comprobante_transferencia = None

            solicitud_editada.save()

            messages.success(
                request, "Solicitud editada correctamente. Vuelve a revisión del PM."
            )
            return redirect('rrhh:listar_adelantos_usuario')
    else:
        form = SolicitudAdelantoForm(instance=solicitud)

    return render(request, 'rrhh/editar_adelanto_usuario.html', {
        'form': form,
        'solicitud': solicitud
    })


@staff_member_required
@rol_requerido('rrhh', 'admin')
def rechazar_adelanto_rrhh(request, solicitud_id):
    solicitud = get_object_or_404(SolicitudAdelanto, id=solicitud_id)
    if request.method == 'POST':
        motivo = request.POST.get('motivo_rechazo', '')
        solicitud.estado = 'rechazada_rrhh'
        solicitud.motivo_rechazo = motivo
        solicitud.save()
        messages.success(request, "Solicitud rechazada por RRHH.")
    return redirect('rrhh:listar_adelanto_admin')


@staff_member_required
@rol_requerido('admin')
def activar_edicion_rrhh(request, id):
    solicitud = get_object_or_404(SolicitudAdelanto, id=id)
    solicitud.puede_editar_rrhh = not solicitud.puede_editar_rrhh
    solicitud.save()
    estado = "activada" if solicitud.puede_editar_rrhh else "desactivada"
    messages.success(request, f"Edición para RRHH {estado} correctamente.")
    return redirect('rrhh:listar_adelanto_admin')


# === Helpers para estampar firma en el PDF ===


def _stamp_signature_on_last_page(unsigned_pdf_bytes: bytes, signature_png_bytes: bytes, x=72, y=72, max_width=200) -> bytes:
    """
    Dibuja la firma PNG en la última página del PDF.
    - x, y: posición en puntos (72pt = 1 inch) desde la esquina inferior izquierda
    - max_width: ancho máximo de la firma
    Devuelve bytes del PDF firmado.
    """
    # 1) Leer PDF base
    base_reader = PdfReader(io.BytesIO(unsigned_pdf_bytes))
    writer = PdfWriter()
    for p in base_reader.pages:
        writer.add_page(p)

    # 2) Preparar overlay de una página del mismo tamaño que la última
    last_page = base_reader.pages[-1]
    w = float(last_page.mediabox.width)
    h = float(last_page.mediabox.height)

    packet = io.BytesIO()
    c = canvas.Canvas(packet, pagesize=(w, h))

    # Redimensionar firma manteniendo proporción
    sig_img = Image.open(io.BytesIO(signature_png_bytes)).convert("RGBA")
    ratio = min(1.0, max_width / sig_img.width)
    sig_w = sig_img.width * ratio
    sig_h = sig_img.height * ratio

    # ReportLab necesita la imagen como archivo/bytes
    sig_buf = io.BytesIO()
    sig_img.resize((int(sig_w), int(sig_h))).save(sig_buf, format="PNG")
    sig_buf.seek(0)

    # Dibuja firma en (x, y)
    c.drawImage(sig_buf, x, y, width=sig_w, height=sig_h, mask='auto')
    c.save()

    # 3) Combinar overlay con última página
    packet.seek(0)
    overlay_reader = PdfReader(packet)
    writer.pages[-1].merge_page(overlay_reader.pages[0])

    # 4) Escribir resultado
    out = io.BytesIO()
    writer.write(out)
    return out.getvalue()


# ==========================
#   ADMIN / RRHH VIEWS
# ==========================

@login_required(login_url='usuarios:login')
@rol_requerido('admin', 'rrhh')  # ajusta si PM también debe ver
def list_rate_sheets(request):
    """
    Lista administrativa de Rate Sheets (todas).
    """
    qs = RateSheet.objects.select_related('technician').order_by('-created_at')

    # Paginación
    cantidad = request.GET.get('cantidad', '10')
    cantidad = 1000000 if cantidad == 'todos' else int(cantidad)
    paginator = Paginator(qs, cantidad)
    page_number = request.GET.get('page')
    pagina = paginator.get_page(page_number)

    return render(request, 'rrhh/rate_sheet_list.html', {
        'pagina': pagina,
        'cantidad': request.GET.get('cantidad', '10'),
    })


@login_required(login_url='usuarios:login')
@rol_requerido('admin', 'rrhh')
@require_http_methods(["GET", "POST"])
def add_rate_sheet(request):
    """
    Cargar un nuevo Rate Sheet (PDF sin firmar).
    """
    if request.method == "POST":
        form = RateSheetForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, "Rate Sheet uploaded.")
            return redirect('rrhh:list_rate_sheets')
        messages.error(request, "Please review the form fields.")
    else:
        form = RateSheetForm()

    return render(request, 'rrhh/rate_sheet_create.html', {'form': form})


@login_required(login_url='usuarios:login')
@rol_requerido('admin', 'rrhh')
@require_http_methods(["GET", "POST"])
def edit_rate_sheet(request, pk):
    """
    Editar: normalmente solo cambiar el técnico o reemplazar el PDF sin firmar.
    """
    rs = get_object_or_404(RateSheet, pk=pk)
    if request.method == "POST":
        form = RateSheetForm(request.POST, request.FILES, instance=rs)
        if form.is_valid():
            form.save()
            messages.success(request, "Rate Sheet updated.")
            return redirect('rrhh:list_rate_sheets')
        messages.error(request, "Please review the form fields.")
    else:
        form = RateSheetForm(instance=rs)

    return render(request, 'rrhh/rate_sheet_edit.html', {'form': form, 'rs': rs})


@login_required(login_url='usuarios:login')
@rol_requerido('admin', 'rrhh')
@require_http_methods(["POST"])
def delete_rate_sheet(request, pk):
    """
    Eliminar un Rate Sheet (no borra archivos en Wasabi por defecto).
    """
    rs = get_object_or_404(RateSheet, pk=pk)
    rs.delete()
    messages.success(request, "Rate Sheet deleted.")
    return redirect('rrhh:list_rate_sheets')


# ==========================
#   TECHNICIAN VIEW
# ==========================
# rrhh/views.py


@staff_member_required
@rol_requerido('rrhh', 'admin', 'pm')
def eliminar_firma(request, user_id):
    if request.method != "POST":
        messages.error(request, "Invalid method.")
        return redirect('rrhh:listar_firmas')

    user = get_object_or_404(CustomUser, id=user_id)

    # Guardamos referencias antes de tocar el campo
    field = getattr(user, "firma_digital", None)
    name = getattr(field, "name", "") if field else ""
    storage = getattr(field, "storage", None)

    # También contemplamos el modelo auxiliar UserSignature (si lo estás usando)
    usig = None
    try:
        usig = UserSignature.objects.get(user=user)
    except Exception:
        usig = None

    deleted_any = False
    errors = []

    with transaction.atomic():
        # 1) Borrar en Wasabi el archivo del FileField en CustomUser
        if name and storage:
            try:
                # borra explícitamente en el bucket
                if storage.exists(name):
                    storage.delete(name)
                deleted_any = True
            except Exception as e:
                errors.append(f"FileField delete error: {e}")

        # 2) Limpiar el campo en BD
        try:
            if field:
                # evita que Django intente borrar de nuevo en save()
                field.delete(save=False)
            user.firma_digital = None
            user.save(update_fields=["firma_digital"])
        except Exception as e:
            errors.append(f"DB clear error: {e}")

        # 3) Si tienes UserSignature, bórralo (archivo + fila)
        if usig and usig.image and getattr(usig.image, "name", ""):
            try:
                img_name = usig.image.name
                img_storage = usig.image.storage
                if img_storage and img_storage.exists(img_name):
                    img_storage.delete(img_name)
                usig.delete()
                deleted_any = True
            except Exception as e:
                errors.append(f"UserSignature delete error: {e}")
        elif usig:
            # no había imagen, borra la fila para mantener consistencia
            try:
                usig.delete()
            except Exception:
                pass

    if errors:
        messages.error(
            request, "No se pudo eliminar completamente la firma: " + " | ".join(errors))
    elif deleted_any:
        messages.success(
            request, f"Firma eliminada para {user.get_full_name() or user.username}.")
    else:
        messages.info(request, "Este usuario no tiene una firma registrada.")

    return redirect('rrhh:listar_firmas')


@staff_member_required
@rol_requerido('rrhh', 'admin', 'pm')
def signature_preview_admin(request, user_id):
    """
    Devuelve la imagen de la firma desde Wasabi para <img>.
    - 404 si no hay firma o no existe el objeto en el bucket
    - sin caché para evitar imágenes “pegadas” tras borrar/reemplazar
    """
    usuario = get_object_or_404(CustomUser, id=user_id)

    name = getattr(getattr(usuario, "firma_digital", None), "name", "")
    if not name:
        raise Http404("Signature not found")

    StorageClass = import_string(settings.DEFAULT_FILE_STORAGE)
    storage = StorageClass()

    # Verifica existencia (evita stacktrace si se borró en el bucket)
    if not storage.exists(name):
        raise Http404("Signature not found")

    try:
        f = storage.open(name, "rb")
    except Exception:
        raise Http404("Signature not found")

    content_type = mimetypes.guess_type(name)[0] or "image/png"
    resp = FileResponse(f, content_type=content_type)

    # Desactivar caché del navegador/CDN/proxy
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"

    return resp


@staff_member_required
def register_signature_admin(request, user_id):
    """
    Registra/Reemplaza o Elimina la firma de un usuario específico (admin).
    Incluye manejo de Cancel.
    """
    usuario = get_object_or_404(CustomUser, id=user_id)
    next_url = request.GET.get("next") or reverse("rrhh:listar_firmas")

    StorageClass = import_string(settings.DEFAULT_FILE_STORAGE)
    storage = StorageClass()

    if request.method == "POST":
        # Cancelar
        if request.POST.get("cancel"):
            return redirect(next_url)

        # Eliminar firma
        if "eliminar_firma" in request.POST:
            name = getattr(getattr(usuario, "firma_digital", None), "name", "")
            if name:
                try:
                    if storage.exists(name):
                        storage.delete(name)
                except Exception:
                    pass
                usuario.firma_digital = None
                usuario.save(update_fields=["firma_digital"])
                messages.success(request, "Signature deleted.")
            else:
                messages.info(
                    request, "This user has no registered signature.")
            return redirect(next_url)

        # Guardar/Reemplazar
        dataurl = request.POST.get("signature_dataurl", "")
        if not dataurl or not dataurl.startswith("data:image"):
            messages.error(request, "Please draw a signature before saving.")
            return redirect(request.path)

        try:
            _, b64 = dataurl.split(",", 1)
            img_bytes = base64.b64decode(b64)
        except Exception:
            messages.error(request, "Invalid signature data.")
            return redirect(request.path)

        if not img_bytes:
            messages.error(request, "Signature image is empty.")
            return redirect(request.path)

        old_name = getattr(getattr(usuario, "firma_digital", None), "name", "")
        if old_name:
            try:
                if storage.exists(old_name):
                    storage.delete(old_name)
            except Exception:
                pass

        nombre = slugify(usuario.get_full_name()
                         or usuario.username) or f"user-{usuario.id}"
        key = f"RRHH/Signatures/{nombre}/signature.png"

        saved_name = storage.save(key, ContentFile(img_bytes))
        usuario.firma_digital.name = saved_name
        usuario.save(update_fields=["firma_digital"])

        messages.success(
            request, f"Signature saved for {usuario.get_full_name() or usuario.username}.")
        return redirect(next_url)

    return render(request, "rrhh/register_signature_unified.html", {
        "target": usuario,
        "base_template": "dashboard_admin/base.html",
        "can_replace": True,
        "has_signature": bool(getattr(getattr(usuario, "firma_digital", None), "name", "")),
        "preview_url": reverse("rrhh:signature_preview_admin", args=[usuario.id]) if getattr(getattr(usuario, "firma_digital", None), "name", "") else None,
        "next_url": next_url,
    })


@staff_member_required
@rol_requerido('rrhh', 'admin', 'pm')
def listar_firmas(request):
    qs = CustomUser.objects.all().order_by('first_name', 'last_name', 'username')

    cantidad_param = request.GET.get('cantidad', '10')
    per_page = 1000000 if cantidad_param == 'todos' else int(cantidad_param)

    paginator = Paginator(qs, per_page)
    page_number = request.GET.get('page')
    pagina = paginator.get_page(page_number)

    return render(request, 'rrhh/listar_firmas.html', {
        'pagina': pagina,
        'cantidad': cantidad_param,
    })


@login_required(login_url='usuarios:login')
def signature_preview(request):
    """
    Devuelve la imagen de la firma del usuario autenticado *desde Wasabi*.
    Sin helpers externos: resolvemos storage, abrimos por key guardada en el campo,
    validamos bytes > 0 y devolvemos con content-type correcto.
    """
    # 1) Key en el FileField
    sig_field = getattr(request.user, "firma_digital", None)
    name = getattr(sig_field, "name", "") if sig_field else ""
    if not name:
        raise Http404("No signature")

    # 2) Storage (debe ser Wasabi en tu settings)
    StorageClass = import_string(settings.DEFAULT_FILE_STORAGE)
    storage = StorageClass()

    # 3) Leer bytes
    try:
        with storage.open(name, "rb") as f:
            data = f.read()
    except Exception:
        raise Http404("No signature")

    if not data:
        raise Http404("No signature")

    # 4) Content-Type y cache
    ct = mimetypes.guess_type(name)[0] or "image/png"
    resp = HttpResponse(data, content_type=ct)
    resp["Cache-Control"] = "no-store"
    return resp


@login_required(login_url='usuarios:login')
def sign_rate_sheet(request, pk):
    rs = get_object_or_404(RateSheet, pk=pk)

    # Solo el propio técnico
    if rs.technician_id != request.user.id:
        messages.error(request, "You are not allowed to sign this Rate Sheet.")
        return redirect('rrhh:my_rate_sheets')

    # Evitar doble firma
    if rs.status != "pending":
        messages.info(request, "This Rate Sheet is already signed.")
        return redirect('rrhh:my_rate_sheets')

    # Storage
    StorageClass = import_string(settings.DEFAULT_FILE_STORAGE)
    storage = StorageClass()

    # ¿Firma válida?
    sig_field = getattr(request.user, "firma_digital", None)
    sig_name = getattr(sig_field, "name", "") if sig_field else ""
    has_signature = False
    if sig_name:
        try:
            with storage.open(sig_name, "rb") as f:
                has_signature = bool(f.read(10))
        except Exception:
            has_signature = False
            try:
                request.user.firma_digital = None
                request.user.save(update_fields=["firma_digital"])
            except Exception:
                pass
            sig_name = ""

    if request.method == "POST":
        if request.POST.get("cancel"):
            return redirect('rrhh:my_rate_sheets')

        # 1) Bytes de firma
        if not has_signature:
            dataurl = request.POST.get("signature_dataurl", "")
            if not dataurl.startswith("data:image"):
                messages.error(
                    request, "Please draw your signature to continue.")
                return redirect(request.path)
            try:
                _, b64 = dataurl.split(",", 1)
                sig_bytes = base64.b64decode(b64)
            except Exception as e:
                messages.error(request, f"Invalid signature data. ({e})")
                return redirect(request.path)
            if not sig_bytes:
                messages.error(request, "Empty signature.")
                return redirect(request.path)

            # Guardar en Wasabi
            nombre = slugify(request.user.get_full_name(
            ) or request.user.username) or f"user-{request.user.id}"
            key = f"RRHH/Signatures/{nombre}/signature.png"
            try:
                saved_name = storage.save(key, ContentFile(sig_bytes))
                request.user.firma_digital.name = saved_name
                request.user.save(update_fields=["firma_digital"])
                sig_name = saved_name
                has_signature = True
            except Exception as e:
                messages.error(request, f"Could not save your signature: {e}")
                return redirect(request.path)
        else:
            try:
                with storage.open(sig_name, "rb") as f:
                    sig_bytes = f.read()
            except Exception as e:
                messages.error(
                    request, f"Could not read your saved signature: {e}")
                return redirect(request.path)
            if not sig_bytes:
                messages.error(
                    request, "Your saved signature file is empty or unreadable. Please contact Human Resources.")
                return redirect(request.path)

        # 2) Leer PDF sin firmar
        with rs.file_unsigned.open("rb") as f:
            pdf_bytes = f.read()

        # 3) Estampar firma justo después del último texto
        try:
            doc = fitz.open(stream=pdf_bytes, filetype="pdf")
            page = doc[-1]
            page_w, page_h = page.rect.width, page.rect.height
            cx = page_w / 2.0

            # Última línea real de texto (más robusto que blocks)
            last_text_bottom = 72  # fallback: 1" desde arriba
            try:
                # (x0,y0,x1,y1,word,block_no,line_no,word_no)
                words = page.get_text("words")
                if words:
                    last_text_bottom = max(w[3] for w in words if w[4].strip())
                else:
                    blocks = page.get_text("blocks")
                    bottoms = [b[3] for b in blocks if len(
                        b) >= 5 and (b[4] or "").strip()]
                    if bottoms:
                        last_text_bottom = max(bottoms)
            except Exception:
                pass

            # Parámetros
            sig_w = 240
            sig_h = 90
            margin = 24
            gap_after_text = 10
            gap_line = 6
            gap_text = 4
            caption_h = 22
            line_w = 1.2

            # Posición de la firma
            y_sig_top = last_text_bottom + gap_after_text
            # no pasar el margen inferior
            y_sig_top = min(y_sig_top, page_h - margin - sig_h)
            # ni subir demasiado
            y_sig_top = max(margin, y_sig_top)

            rect_sig = fitz.Rect(cx - sig_w/2.0, y_sig_top,
                                 cx + sig_w/2.0, y_sig_top + sig_h)
            page.insert_image(rect_sig, stream=sig_bytes, keep_proportion=True)

            # Texto
            name = rs.technician.get_full_name() or rs.technician.username
            date_str = timezone.localtime().strftime("%Y-%m-%d %H:%M")
            caption = f"Signed by {name} — {date_str}"

            def draw_caption(y_top, fontsize=9.5, height=caption_h):
                rect_text = fitz.Rect(rect_sig.x0, y_top,
                                      rect_sig.x1, y_top + height)
                # insert_textbox devuelve 0 si no pudo dibujar
                status = page.insert_textbox(rect_text, caption,
                                             fontsize=fontsize, fontname="helv",
                                             align=1, color=(0, 0, 0))
                return status

            # Preferir debajo de la firma
            need_below = gap_line + line_w + gap_text + caption_h
            space_below = page_h - rect_sig.y1 - margin

            if space_below >= need_below:
                y_line = rect_sig.y1 + gap_line
                page.draw_line((rect_sig.x0, y_line), (rect_sig.x1,
                               y_line), color=(0, 0, 0), width=line_w)
                # intentar dibujar; si no entra, aumenta alto / baja fuente
                status = draw_caption(
                    y_line + gap_text, fontsize=9.5, height=caption_h)
                if status == 0:
                    status = draw_caption(
                        y_line + gap_text, fontsize=9, height=caption_h+6)
                    if status == 0:
                        draw_caption(y_line + gap_text,
                                     fontsize=8.5, height=caption_h+10)
            else:
                # Encima de la firma
                y_line = rect_sig.y0 - gap_line
                page.draw_line((rect_sig.x0, y_line), (rect_sig.x1,
                               y_line), color=(0, 0, 0), width=line_w)
                y_caption_top = max(margin, y_line - gap_text - caption_h)
                status = draw_caption(
                    y_caption_top, fontsize=9.5, height=caption_h)
                if status == 0:
                    y_caption_top = max(
                        margin, y_line - gap_text - (caption_h+6))
                    status = draw_caption(
                        y_caption_top, fontsize=9, height=caption_h+6)
                    if status == 0:
                        y_caption_top = max(
                            margin, y_line - gap_text - (caption_h+10))
                        draw_caption(y_caption_top, fontsize=8.5,
                                     height=caption_h+10)

            out_bytes = doc.tobytes()
            doc.close()
        except Exception as e:
            messages.error(request, f"Could not sign the document: {e}")
            return redirect(request.path)

        # 4) Guardar PDF firmado
        filename = f"signed-{slugify(request.user.get_full_name() or request.user.username)}.pdf"
        rs.file_signed.save(filename, ContentFile(out_bytes), save=False)
        rs.status = "signed"
        rs.signed_at = timezone.now()
        rs.save(update_fields=["file_signed", "status", "signed_at"])

        messages.success(request, "Rate Sheet signed successfully.")
        return redirect('rrhh:my_rate_sheets')

    # GET
    preview_url = reverse("rrhh:signature_preview") if has_signature else None
    return render(request, "rrhh/rate_sheet_sign.html", {
        "rs": rs,
        "has_signature": has_signature,
        "signature_preview_url": preview_url,
    })


@login_required(login_url='usuarios:login')
def my_rate_sheets(request):
    qs = RateSheet.objects.filter(
        technician=request.user
    ).order_by('-created_at')

    # --- has_signature (sin helpers) ---
    StorageClass = import_string(
        settings.DEFAULT_FILE_STORAGE)  # tu backend Wasabi
    storage = StorageClass()
    sig_field = getattr(request.user, "firma_digital", None)
    sig_name = getattr(sig_field, "name", "") if sig_field else ""
    if sig_name:
        try:
            has_signature = storage.exists(sig_name)
        except Exception:
            # si el storage no soporta exists() en local, damos por buena la existencia del nombre
            has_signature = True
    else:
        has_signature = False
    # -----------------------------------

    cantidad = request.GET.get('cantidad', '10')
    per_page = 1000000 if cantidad == 'todos' else int(cantidad)
    paginator = Paginator(qs, per_page)
    page_number = request.GET.get('page')
    pagina = paginator.get_page(page_number)

    return render(request, 'rrhh/my_rate_sheets.html', {
        'pagina': pagina,
        'cantidad': cantidad,
        'has_signature': has_signature,
    })


@login_required(login_url='usuarios:login')
def register_signature_unified(request, user_id=None):
    """
    - user_id = None  -> USUARIO (no reemplaza)
    - user_id = int   -> ADMIN   (puede reemplazar firma del otro)
    """
    is_admin_ctx = user_id is not None
    target = request.user if not is_admin_ctx else get_object_or_404(
        CustomUser, id=user_id)

    # Permisos en modo admin
    if is_admin_ctx and not (request.user.is_staff and request.user.tiene_rol('rrhh', 'admin', 'pm')):
        return HttpResponseForbidden("Not allowed")

    # next seguro
    raw_next = request.GET.get("next") or request.POST.get("next")
    if raw_next and url_has_allowed_host_and_scheme(raw_next, allowed_hosts={request.get_host()}):
        next_url = raw_next
    else:
        next_url = reverse("rrhh:listar_firmas") if is_admin_ctx else reverse(
            "rrhh:my_rate_sheets")

    base_template = "dashboard_admin/base.html" if is_admin_ctx else "dashboard/base.html"
    can_replace = bool(is_admin_ctx)

    # Storage (DEFAULT_FILE_STORAGE -> Wasabi)
    StorageClass = import_string(settings.DEFAULT_FILE_STORAGE)
    storage = StorageClass()

    # Detectar firma real (y limpiar si el key ya no existe)
    sig_field = getattr(target, "firma_digital", None)
    sig_name = getattr(sig_field, "name", "") if sig_field else ""
    try:
        exists_in_bucket = storage.exists(sig_name) if sig_name else False
    except Exception:
        # fallback si el backend no implementa exists()
        exists_in_bucket = bool(sig_name)
    if sig_name and not exists_in_bucket:
        try:
            target.firma_digital = None
            target.save(update_fields=["firma_digital"])
        except Exception:
            pass
        sig_name, exists_in_bucket = "", False

    has_signature = bool(sig_name and exists_in_bucket)
    preview_url = (
        reverse("rrhh:signature_preview_admin", args=[target.id]) if is_admin_ctx
        else reverse("rrhh:signature_preview")
    ) if has_signature else None

    if request.method == "POST":
        # Cancelar
        if request.POST.get("cancel"):
            return redirect(next_url)

        # Usuario no puede reemplazar si ya tiene
        if has_signature and not can_replace:
            messages.info(
                request, "You already have a registered signature. If you need to replace it, please contact Human Resources.")
            return redirect(request.path)

        # Captura
        dataurl = request.POST.get("signature_dataurl", "")
        if not dataurl.startswith("data:image"):
            messages.error(request, "Invalid signature data.")
            return redirect(request.path)

        try:
            _, b64 = dataurl.split(",", 1)
            img_bytes = base64.b64decode(b64)
        except Exception as e:
            messages.error(request, f"Could not decode signature. ({e})")
            return redirect(request.path)

        if not img_bytes:
            messages.error(request, "Empty signature.")
            return redirect(request.path)

        # Guardar en Wasabi igual que en admin
        nombre = slugify(target.get_full_name()
                         or target.username) or f"user-{target.id}"
        key = f"RRHH/Signatures/{nombre}/signature.png"

        try:
            saved_name = storage.save(key, ContentFile(img_bytes))
            target.firma_digital.name = saved_name
            target.save(update_fields=["firma_digital"])
        except Exception as e:
            messages.error(
                request, f"Could not save signature to storage: {e}")
            return redirect(request.path)

        # Verificación opcional
        try:
            if not storage.exists(target.firma_digital.name):
                messages.warning(
                    request, "Signature was set but not found in storage right after saving.")
        except Exception:
            pass

        messages.success(request, "Your digital signature was saved." if not is_admin_ctx
                         else f"Signature saved for {target.get_full_name() or target.username}.")
        return redirect(next_url)

    ctx = {
        "base_template": base_template,
        "target": target,
        "has_signature": has_signature,
        "preview_url": preview_url,
        "can_replace": can_replace,
        "next_url": next_url,
    }
    return render(request, "rrhh/register_signature_unified.html", ctx)


@login_required(login_url='usuarios:login')
def signature_preview(request):
    """Devuelve la firma del usuario logueado (no cacheable)."""
    u = request.user
    name = getattr(getattr(u, "firma_digital", None), "name", "")
    if not name:
        raise Http404("No signature")

    StorageClass = import_string(settings.DEFAULT_FILE_STORAGE)
    storage = StorageClass()

    if not storage.exists(name):
        raise Http404("No signature")

    try:
        f = storage.open(name, "rb")
    except Exception:
        raise Http404("No signature")

    content_type = mimetypes.guess_type(name)[0] or "image/png"
    resp = FileResponse(f, content_type=content_type)
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = "0"
    return resp
