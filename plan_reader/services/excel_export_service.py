import re
from io import BytesIO

from django.conf import settings
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from plan_reader.models import PlanReaderJob

# =============================================================================
# BULK BILLING TEMPLATE CONFIG
# =============================================================================

SHEET_BILLINGS = "Billings"
SHEET_TECHNICIANS = "Technicians"
SHEET_ITEMS = "Items"

BILLINGS_HEADERS = [
    "bulk_key",
    "project_id",
    "client",
    "city",
    "project",
    "office",
    "project_address",
    "projected_week",
    "tech_payment_mode",
    "direct_discount",
    "cable_installation",
    "requirement_type",
    "requirement_list",
]

TECHNICIANS_HEADERS = [
    "bulk_key",
    "technician_username",
    "primary_feed",
]

ITEMS_HEADERS = [
    "bulk_key",
    "job_code",
    "quantity",
]


# =============================================================================
# HELPERS
# =============================================================================


def _clean_cell(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _setting(name, default=""):
    return _clean_cell(getattr(settings, name, default))


def _default_projected_week():
    today = timezone.localdate()
    year, week, _ = today.isocalendar()
    return f"{year}-W{int(week):02d}"


def _safe_text_for_id(value):
    text = _clean_cell(value)

    if not text:
        return ""

    text = text.replace(" ", "_")
    text = re.sub(r"[^A-Za-z0-9_\-\.]", "", text)
    text = re.sub(r"_+", "_", text)

    return text.strip("_")


def _normalize_box_type(value):
    text = _clean_cell(value).upper()
    text = text.replace("×", "X")
    text = text.replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def _requirement_list_from_final_box_type(final_box_type):
    """
    Mapea el Final Box Type detectado por Plan Reader
    al nombre exacto de Requirement List existente en Hyperlink.

    Requirement Lists actuales:
    - A4 1X2
    - A4 1x4
    - B8G
    - B8G 1X4
    - B8G 1X8
    """
    normalized = _normalize_box_type(final_box_type)

    if normalized in {"A4 1X2", "A4 TYPE 2", "A4 1 2"}:
        return "A4 1X2"

    if normalized in {"A4 1X4", "A4 TYPE 1", "A4 1 4"}:
        return "A4 1x4"

    if normalized in {"B8G 1X4", "B8G TYPE 3", "B8G 1 4"}:
        return "B8G 1X4"

    if normalized in {"B8G 1X8", "B8G TYPE 2", "B8G 1 8"}:
        return "B8G 1X8"

    if normalized in {
        "BGP",
        "BBP",
        "BGP TYPE 1",
        "BGP TYPE 2",
        "BBP TYPE 1",
        "BBP TYPE 2",
    }:
        return "B8G"

    if normalized == "B8G":
        return "B8G"

    if normalized.startswith("A4") and "1X2" in normalized:
        return "A4 1X2"

    if normalized.startswith("A4") and "1X4" in normalized:
        return "A4 1x4"

    if normalized.startswith("B8G") and "1X4" in normalized:
        return "B8G 1X4"

    if normalized.startswith("B8G") and "1X8" in normalized:
        return "B8G 1X8"

    if normalized.startswith("B8G"):
        return "B8G"

    if normalized.startswith("BGP") or normalized.startswith("BBP"):
        return "B8G"

    return ""


def _build_project_id(job, item):
    """
    Project ID final para Bulk Billing.

    Regla:
    CO_DFN_PROJECT_ID

    Ejemplos:

    CO = 0913RA
    DFN = 04
    box = 5003-002-5

    Resultado:
    0913RA_04_5003-002-5
    """
    co = _safe_text_for_id(job.co)
    dfn = _safe_text_for_id(job.dfn)
    box = _safe_text_for_id(item.project_name)

    parts = [
        value
        for value in [
            co,
            dfn,
            box,
        ]
        if value
    ]

    if parts:
        return "_".join(parts)

    return f"PLAN_READER_ITEM_{item.id}"


def _unique_bulk_key(base_key, used_keys, item_id):
    """
    bulk_key debe ser único dentro del Excel.

    Se usa el número de caja como base.
    Solo agrega sufijo si el mismo bulk_key ya existe dentro del archivo.
    """
    base_key = _safe_text_for_id(base_key) or f"BILL-{item_id}"

    if base_key not in used_keys:
        used_keys.add(base_key)
        return base_key

    candidate = f"{base_key}-{item_id}"

    if candidate not in used_keys:
        used_keys.add(candidate)
        return candidate

    counter = 2

    while True:
        candidate = f"{base_key}-{item_id}-{counter}"

        if candidate not in used_keys:
            used_keys.add(candidate)
            return candidate

        counter += 1


def _positive_int(value):
    try:
        number = int(value or 0)
    except Exception:
        number = 0

    return max(number, 0)


def _write_sheet_header(ws, headers):
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"


def _autosize_sheet(ws):
    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = _clean_cell(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 14), 45)


def _style_data_rows(ws):
    gray_fill = PatternFill("solid", fgColor="F3F4F6")

    for row in range(2, ws.max_row + 1):
        for cell in ws[row]:
            cell.fill = gray_fill
            cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_instructions_sheet(ws_help):
    ws_help.sheet_view.showGridLines = False

    dark_fill = PatternFill("solid", fgColor="1F2937")
    blue_fill = PatternFill("solid", fgColor="DBEAFE")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    amber_fill = PatternFill("solid", fgColor="FEF3C7")
    red_fill = PatternFill("solid", fgColor="FEE2E2")

    title_font = Font(color="FFFFFF", bold=True, size=16)
    section_font = Font(color="111827", bold=True, size=12)
    bold_font = Font(bold=True)
    normal_font = Font(color="374151", size=11)
    error_font = Font(color="991B1B", bold=True)

    ws_help.merge_cells("A1:F1")
    ws_help["A1"] = "Bulk Billing Import Guide"
    ws_help["A1"].fill = dark_fill
    ws_help["A1"].font = title_font
    ws_help["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_help.row_dimensions[1].height = 28

    row = 3

    def section(title, fill):
        nonlocal row
        ws_help.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)

        cell = ws_help.cell(row=row, column=1)
        cell.value = title
        cell.fill = fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

        ws_help.row_dimensions[row].height = 22
        row += 1

    def line(label, value="", note=""):
        nonlocal row

        ws_help.cell(row=row, column=1).value = label
        ws_help.cell(row=row, column=1).font = bold_font

        ws_help.cell(row=row, column=2).value = value
        ws_help.cell(row=row, column=2).font = normal_font

        if note:
            ws_help.merge_cells(
                start_row=row,
                start_column=3,
                end_row=row,
                end_column=6,
            )
            ws_help.cell(row=row, column=3).value = note
            ws_help.cell(row=row, column=3).font = normal_font

        for col in range(1, 7):
            ws_help.cell(row=row, column=col).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        row += 1

    def blank():
        nonlocal row
        row += 1

    section("1. General workflow", blue_fill)
    line("Step 1", "Review the Billings sheet.", "One row per detected box.")
    line(
        "Step 2",
        "Fill the Technicians sheet.",
        "Technicians are intentionally left blank from Plan Reader.",
    )
    line(
        "Step 3",
        "Review the Items sheet.",
        "Rows are generated from C-108, C-109 and C-110 quantities.",
    )
    line(
        "Step 4",
        "Upload the file.",
        "The Bulk Billing preview will validate technicians, prices and requirement lists.",
    )
    blank()

    section("2. Requirement Lists", green_fill)
    line("A4 1x2", "A4 1X2", "Generated from Final Box Type.")
    line("A4 1x4", "A4 1x4", "Generated from Final Box Type.")
    line("B8G", "B8G", "Generated from Final Box Type.")
    line("B8G 1x4", "B8G 1X4", "Generated from Final Box Type.")
    line("B8G 1x8", "B8G 1X8", "Generated from Final Box Type.")
    blank()

    section("3. Billings sheet", green_fill)
    line("bulk_key", "Required", "Generated from the box number.")
    line("project_id", "Required", "Generated as box number only. Example: 7020-014.")
    line("client", "Required", "Filled from Upload DFN Plan form.")
    line("city", "Required", "Filled from Upload DFN Plan form.")
    line("project", "Required", "Filled from Upload DFN Plan form.")
    line("office", "Required", "Filled from Upload DFN Plan form.")
    line("project_address", "Optional", "Left blank.")
    line("projected_week", "Required", "Uses current ISO week unless configured.")
    line("tech_payment_mode", "Required", "Default full.")
    line("direct_discount", "Required", "Default NO.")
    line("cable_installation", "Required", "Default NO.")
    line("requirement_type", "Required", "fiber.")
    line("requirement_list", "Required", "Generated from Final Box Type.")
    blank()

    section("4. Technicians sheet", amber_fill)
    line(
        "technician_username",
        "Blank",
        "Fill manually before uploading to Bulk Billing.",
    )
    line(
        "Accepted format",
        "tech1, tech2, tech3",
        "You can use one row or multiple technicians in one cell.",
    )
    blank()

    section("5. Items sheet", blue_fill)
    line("bulk_key", "Required", "Matches Billings sheet.")
    line("job_code", "Required", "Generated from Plan Reader item quantities.")
    line("quantity", "Required", "Always positive for normal billing.")
    blank()

    section("6. Important", red_fill)
    ws_help.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws_help.cell(row=row, column=1).value = (
        "IMPORTANT: Do not rename sheets or headers. "
        "Before uploading, fill technician_username in the Technicians sheet. "
        "Client, city, project and office come from the Plan Reader upload form. "
        "Project ID is generated as the box number only. "
        "Requirement List is generated from Final Box Type."
    )
    ws_help.cell(row=row, column=1).fill = red_fill
    ws_help.cell(row=row, column=1).font = error_font
    ws_help.cell(row=row, column=1).alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    ws_help.row_dimensions[row].height = 50

    ws_help.column_dimensions["A"].width = 24
    ws_help.column_dimensions["B"].width = 32
    ws_help.column_dimensions["C"].width = 24
    ws_help.column_dimensions["D"].width = 24
    ws_help.column_dimensions["E"].width = 24
    ws_help.column_dimensions["F"].width = 24


def _safe_filename(value):
    text = _safe_text_for_id(value)

    if not text:
        return "plan_reader_bulk_billing"

    return text


# =============================================================================
# MAIN EXPORT
# =============================================================================


def build_plan_reader_excel_response(job_id):
    """
    Genera directamente el Excel compatible con Bulk Billing.

    Hojas:
    - Billings
    - Technicians
    - Items
    - Instructions

    Importante:
    - Solo exporta items incluidos: is_duplicate=False.
    - Los técnicos quedan en blanco.
    - client/city/project/office salen del Job creado al subir el PDF.
    - project_id sale solo como número de caja.
    - requirement_list se llena según Final Box Type.
    """

    job = (
        PlanReaderJob.objects.select_related("uploaded_by")
        .prefetch_related("items")
        .get(id=job_id)
    )

    included_items = list(
        job.items.filter(is_duplicate=False).order_by(
            "sheet",
            "project_name",
            "primary_feed",
            "id",
        )
    )

    default_week = (
        _setting("PLAN_READER_BULK_DEFAULT_WEEK", "") or _default_projected_week()
    )
    default_payment_mode = _setting(
        "PLAN_READER_BULK_DEFAULT_TECH_PAYMENT_MODE",
        "full",
    )
    default_direct_discount = _setting(
        "PLAN_READER_BULK_DEFAULT_DIRECT_DISCOUNT",
        "NO",
    )
    default_cable_installation = _setting(
        "PLAN_READER_BULK_DEFAULT_CABLE_INSTALLATION",
        "NO",
    )

    default_requirement_type = "fiber"

    job_code_c108 = _setting("PLAN_READER_BULK_JOB_CODE_C108", "C-108-UG")
    job_code_c109 = _setting("PLAN_READER_BULK_JOB_CODE_C109", "C-109")
    job_code_c110 = _setting("PLAN_READER_BULK_JOB_CODE_C110", "C-110")

    wb = Workbook()

    ws_b = wb.active
    ws_b.title = SHEET_BILLINGS

    ws_t = wb.create_sheet(SHEET_TECHNICIANS)
    ws_i = wb.create_sheet(SHEET_ITEMS)
    ws_help = wb.create_sheet("Instructions")

    _write_sheet_header(ws_b, BILLINGS_HEADERS)
    _write_sheet_header(ws_t, TECHNICIANS_HEADERS)
    _write_sheet_header(ws_i, ITEMS_HEADERS)

    used_bulk_keys = set()

    for item in included_items:
        project_id = _build_project_id(job, item)
        bulk_key = _unique_bulk_key(project_id, used_bulk_keys, item.id)

        requirement_list = _requirement_list_from_final_box_type(
            item.calculated_box_type
        )

        ws_b.append(
            [
                bulk_key,
                project_id,
                job.client or "",
                job.city or "",
                job.project or "",
                job.office or "",
                "",
                default_week,
                default_payment_mode,
                default_direct_discount,
                default_cable_installation,
                default_requirement_type,
                requirement_list,
            ]
        )

        ws_t.append(
            [
                bulk_key,
                "",
                item.primary_feed or "",
            ]
        )

        c108_qty = _positive_int(item.c108_ug)
        c109_qty = _positive_int(item.c109_splices)
        c110_qty = _positive_int(item.c110_splitters)

        if c108_qty:
            ws_i.append(
                [
                    bulk_key,
                    job_code_c108,
                    c108_qty,
                ]
            )

        if c109_qty:
            ws_i.append(
                [
                    bulk_key,
                    job_code_c109,
                    c109_qty,
                ]
            )

        if c110_qty:
            ws_i.append(
                [
                    bulk_key,
                    job_code_c110,
                    c110_qty,
                ]
            )

    _write_instructions_sheet(ws_help)

    for ws in [ws_b, ws_t, ws_i, ws_help]:
        _autosize_sheet(ws)

    for ws in [ws_b, ws_t, ws_i]:
        _style_data_rows(ws)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = job.original_filename or f"plan_reader_job_{job.id}.pdf"
    safe_name = safe_name.replace(".pdf", "")
    safe_name = _safe_filename(safe_name)

    filename = f"BulkBilling_{safe_name}_Job_{job.id}.xlsx"

    return filename, output.getvalue()
