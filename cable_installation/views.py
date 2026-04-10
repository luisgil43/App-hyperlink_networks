import csv
import io
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count, Prefetch, Q
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook, load_workbook

from operaciones.models import SesionBilling, SesionBillingTecnico
from usuarios.decoradores import rol_requerido

from .models import CableAssignmentRequirement, CableEvidence, CableRequirement


def _sync_cable_requirements_to_assignments(billing: SesionBilling):
    """
    Asegura que cada asignación tenga una fila por cada requisito maestro.
    No elimina filas históricas automáticamente.
    """
    assignments = list(billing.tecnicos_sesion.all())
    requirements = list(
        billing.cable_requirements.all().order_by("order", "sequence_no", "id")
    )

    if not assignments or not requirements:
        return

    existing_pairs = set(
        CableAssignmentRequirement.objects.filter(
            assignment__in=assignments,
            requirement__in=requirements,
        ).values_list("assignment_id", "requirement_id")
    )

    to_create = []
    for assignment in assignments:
        for requirement in requirements:
            key = (assignment.id, requirement.id)
            if key in existing_pairs:
                continue
            to_create.append(
                CableAssignmentRequirement(
                    assignment=assignment,
                    requirement=requirement,
                    status=CableAssignmentRequirement.STATUS_PENDING,
                )
            )

    if to_create:
        CableAssignmentRequirement.objects.bulk_create(to_create)


def _normalize_required_import_value(value, default=True):
    if value in (None, "", " "):
        return default

    text = str(value).strip().lower()

    yes_values = {"1", "true", "t", "yes", "y", "si", "sí", "s"}
    no_values = {"0", "false", "f", "no", "n"}

    if text in yes_values:
        return True
    if text in no_values:
        return False

    raise ValueError("Required must be yes/no, si/no, true/false, or 1/0.")


def _parse_required_flag(value, default=True):
    if value in (None, ""):
        return default

    text = str(value).strip().lower()

    yes_values = {"1", "true", "t", "yes", "y", "si", "sí", "s"}
    no_values = {"0", "false", "f", "no", "n"}

    if text in yes_values:
        return True
    if text in no_values:
        return False
    return default


def _parse_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, ValueError, TypeError):
        return None


def _decimal_to_input(value):
    """
    Para inputs HTML:
    30.00 -> '30'
    30.50 -> '30.5'
    None  -> ''
    """
    if value in (None, ""):
        return ""
    try:
        d = Decimal(str(value))
    except (InvalidOperation, ValueError, TypeError):
        return str(value)

    s = format(d, "f")
    if "." in s:
        s = s.rstrip("0").rstrip(".")
    return s


def _is_ops_reviewer(user):
    return (
        getattr(user, "es_supervisor", False)
        or getattr(user, "es_pm", False)
        or getattr(user, "es_admin_general", False)
        or user.is_superuser
    )


def _build_requirement_rows_from_db(billing: SesionBilling):
    rows = []
    requirements = list(
        billing.cable_requirements.all().order_by("order", "sequence_no", "id")
    )

    for r in requirements:
        rows.append(
            {
                "id": str(r.id),
                "pk_label": str(r.sequence_no),
                "handhole": r.handhole or "",
                "planned_reserve_ft": _decimal_to_input(r.planned_reserve_ft),
                "warning": r.warning or "",
                "order": r.order,
                "errors": {},
            }
        )
    return rows


@login_required
@rol_requerido("supervisor", "admin", "pm")
def configure_requirements(request, billing_id):
    billing = get_object_or_404(
        SesionBilling, pk=billing_id, is_cable_installation=True
    )

    def _to_bool(value, default=True):
        if value in (None, ""):
            return default

        text = str(value).strip().lower()
        yes_values = {"1", "true", "t", "yes", "y", "si", "sí", "s"}
        no_values = {"0", "false", "f", "no", "n"}

        if text in yes_values:
            return True
        if text in no_values:
            return False
        return default

    def _build_requirement_rows_from_db():
        rows = []
        requirements = list(
            billing.cable_requirements.all().order_by("order", "sequence_no", "id")
        )

        for r in requirements:
            rows.append(
                {
                    "id": str(r.id),
                    "pk_label": str(r.sequence_no),
                    "handhole": r.handhole or "",
                    "planned_reserve_ft": _decimal_to_input(r.planned_reserve_ft),
                    "warning": r.warning or "",
                    "required": bool(getattr(r, "required", True)),
                    "order": r.order,
                    "errors": {},
                }
            )
        return rows

    form_errors = []
    rows = _build_requirement_rows_from_db()

    if request.method == "POST":
        handholes = request.POST.getlist("handhole[]")
        reserves = request.POST.getlist("planned_reserve_ft[]")
        warnings_list = request.POST.getlist("warning[]")
        required_list = request.POST.getlist("required[]")
        orders = request.POST.getlist("order[]")
        req_ids = request.POST.getlist("id[]")
        delete_ids = set(x for x in request.POST.getlist("delete_id[]") if x)

        rows = []
        normalized = []

        total_rows = max(
            len(handholes),
            len(reserves),
            len(warnings_list),
            len(required_list),
            len(orders),
            len(req_ids),
        )

        current_by_id = {str(r.id): r for r in billing.cable_requirements.all()}

        for idx in range(total_rows):
            req_id = req_ids[idx].strip() if idx < len(req_ids) and req_ids[idx] else ""
            handhole = (handholes[idx] or "").strip() if idx < len(handholes) else ""
            reserve_raw = (reserves[idx] or "").strip() if idx < len(reserves) else ""
            warning = (
                (warnings_list[idx] or "").strip() if idx < len(warnings_list) else ""
            )
            required_raw = (
                (required_list[idx] or "").strip()
                if idx < len(required_list)
                else "yes"
            )
            order_raw = (orders[idx] or "").strip() if idx < len(orders) else ""

            if req_id and req_id in delete_ids:
                continue

            row_errors = {}

            if not handhole:
                row_errors["handhole"] = "Handhole is required."

            reserve_value = None
            if reserve_raw == "":
                reserve_value = Decimal("0")
            else:
                reserve_value = _parse_decimal(reserve_raw)
                if reserve_value is None:
                    row_errors["planned_reserve_ft"] = "Enter a valid number."
                elif reserve_value < 0:
                    row_errors["planned_reserve_ft"] = "Reserve cannot be negative."

            if order_raw == "":
                order_value = idx
            else:
                try:
                    order_value = int(order_raw)
                    if order_value < 0:
                        row_errors["order"] = "Order cannot be negative."
                except Exception:
                    row_errors["order"] = "Enter a valid integer."
                    order_value = idx

            required_value = _to_bool(required_raw, default=True)

            pk_label = "Auto"
            if req_id and req_id in current_by_id:
                pk_label = str(current_by_id[req_id].sequence_no)

            row_payload = {
                "id": req_id,
                "pk_label": pk_label,
                "handhole": handhole,
                "planned_reserve_ft": reserve_raw,
                "warning": warning,
                "required": required_value,
                "order": order_raw if order_raw != "" else idx,
                "errors": row_errors,
            }
            rows.append(row_payload)

            if row_errors:
                continue

            normalized.append(
                {
                    "id": req_id or None,
                    "handhole": handhole,
                    "planned_reserve_ft": reserve_value,
                    "warning": warning,
                    "required": required_value,
                    "order": order_value,
                }
            )

        has_inline_errors = any(row["errors"] for row in rows)

        if has_inline_errors:
            form_errors.append("Please fix the highlighted fields before saving.")
        else:
            try:
                with transaction.atomic():
                    current_by_id = {
                        str(r.id): r for r in billing.cable_requirements.all()
                    }

                    if delete_ids:
                        CableRequirement.objects.filter(
                            billing=billing,
                            id__in=delete_ids,
                        ).delete()

                    next_seq = CableRequirement.next_sequence_for_billing(billing)

                    for row in normalized:
                        req_id = row["id"]

                        if req_id and req_id in current_by_id:
                            obj = current_by_id[req_id]
                            obj.handhole = row["handhole"]
                            obj.planned_reserve_ft = row["planned_reserve_ft"]
                            obj.warning = row["warning"]
                            obj.required = row["required"]
                            obj.order = row["order"]
                            obj.save(
                                update_fields=[
                                    "handhole",
                                    "planned_reserve_ft",
                                    "warning",
                                    "required",
                                    "order",
                                    "updated_at",
                                ]
                            )
                        else:
                            CableRequirement.objects.create(
                                billing=billing,
                                sequence_no=next_seq,
                                handhole=row["handhole"],
                                planned_reserve_ft=row["planned_reserve_ft"],
                                warning=row["warning"],
                                required=row["required"],
                                order=row["order"],
                            )
                            next_seq += 1

                    _sync_cable_requirements_to_assignments(billing)

                messages.success(request, "Cable requirements saved.")
                return redirect(
                    "cable_installation:configure_requirements",
                    billing_id=billing.id,
                )

            except Exception as exc:
                form_errors.append(f"Could not save requirements: {exc}")

    return render(
        request,
        "cable_installation/configure_requirements.html",
        {
            "billing": billing,
            "form_rows": rows,
            "form_errors": form_errors,
        },
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_GET
def export_measurements_excel(request, billing_id):
    billing = get_object_or_404(
        SesionBilling, pk=billing_id, is_cable_installation=True
    )

    rows = (
        CableAssignmentRequirement.objects.filter(assignment__sesion=billing)
        .select_related("assignment", "assignment__tecnico", "requirement")
        .order_by("requirement__order", "requirement__sequence_no", "assignment__id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Cable Measurements"

    ws.append(
        [
            "PK",
            "Handhole",
            "Required",
            "Technician",
            "Planned reserve (ft)",
            "Start ft",
            "Expected end ft (low)",
            "Expected end ft (high)",
            "End ft",
            "Installed ft",
            "Technician status",
            "Warning",
            "Technician note",
            "Supervisor note",
            "Measurement warning",
        ]
    )

    for row in rows:
        req = row.requirement
        tech_name = (
            getattr(row.assignment.tecnico, "get_full_name", lambda: "")()
            or row.assignment.tecnico.username
        )
        ws.append(
            [
                req.sequence_no,
                req.handhole,
                "Yes" if req.required else "No",
                tech_name,
                float(req.planned_reserve_ft or 0),
                float(req.start_ft or 0) if req.start_ft is not None else "",
                (
                    float(req.expected_end_ft_low or 0)
                    if req.expected_end_ft_low is not None
                    else ""
                ),
                (
                    float(req.expected_end_ft_high or 0)
                    if req.expected_end_ft_high is not None
                    else ""
                ),
                float(req.end_ft or 0) if req.end_ft is not None else "",
                float(req.installed_ft or 0),
                row.get_status_display(),
                req.warning,
                row.note,
                row.supervisor_note,
                req.measurement_warning_text,
            ]
        )

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f"cable_measurements_billing_{billing.id}.xlsx"
    return FileResponse(
        bio,
        as_attachment=True,
        filename=filename,
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
def import_requirements_page(request, billing_id):
    billing = get_object_or_404(
        SesionBilling,
        pk=billing_id,
        is_cable_installation=True,
    )

    return render(
        request,
        "cable_installation/import_requirements.html",
        {
            "billing": billing,
        },
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
def download_requirements_template(request, billing_id, ext):
    billing = get_object_or_404(
        SesionBilling,
        pk=billing_id,
        is_cable_installation=True,
    )

    ext = (ext or "").lower()
    filename_base = f"cable_requirements_template_billing_{billing.id}"

    if ext == "csv":
        content = (
            "handhole,planned_reserve_ft,required,warning,order\n"
            "1000-044,30,Yes,,0\n"
            "1000-044-1,20,YES,,1\n"
            "1000-044-2,10,si,Low clearance,2\n"
            "1000-044-3,0,No,,3\n"
        )
        response = HttpResponse(content, content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
        return response

    if ext in ("xlsx", "xls"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Requirements"

        ws.append(["handhole", "planned_reserve_ft", "required", "warning", "order"])
        ws.append(["1000-044", 30, "Yes", "", 0])
        ws.append(["1000-044-1", 20, "YES", "", 1])
        ws.append(["1000-044-2", 10, "si", "Low clearance", 2])
        ws.append(["1000-044-3", 0, "No", "", 3])

        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)

        response = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    messages.error(request, "Unsupported format. Use csv or xlsx.")
    return redirect(
        "cable_installation:import_requirements_page", billing_id=billing.id
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def import_requirements_preview(request, billing_id):
    billing = get_object_or_404(
        SesionBilling,
        pk=billing_id,
        is_cable_installation=True,
    )

    uploaded = request.FILES.get("file")
    if not uploaded:
        messages.error(request, "Please select a CSV or XLSX file.")
        return redirect(
            "cable_installation:import_requirements_page", billing_id=billing.id
        )

    filename = (getattr(uploaded, "name", "") or "").strip()
    ext = (filename.rsplit(".", 1)[-1] or "").lower()

    parsed_rows = []
    warnings = []
    errors = []

    try:
        if ext == "csv":
            raw = uploaded.read().decode("utf-8", errors="ignore")
            if not raw.strip():
                messages.warning(request, "The file is empty.")
                return redirect(
                    "cable_installation:import_requirements_page", billing_id=billing.id
                )

            lines = raw.splitlines()
            header_line = lines[0].lower() if lines else ""
            has_header = "handhole" in header_line

            if has_header:
                reader = csv.DictReader(io.StringIO(raw))
                for row_num, row in enumerate(reader, start=2):
                    handhole = (row.get("handhole") or "").strip()
                    if not handhole:
                        continue

                    reserve_raw = (row.get("planned_reserve_ft") or "").strip()
                    warning = (row.get("warning") or "").strip()
                    required_raw = row.get("required")
                    order_raw = row.get("order")

                    parsed_rows.append(
                        {
                            "row": row_num,
                            "handhole": handhole,
                            "planned_reserve_ft_raw": reserve_raw,
                            "warning": warning,
                            "required_raw": required_raw,
                            "order_raw": order_raw,
                        }
                    )
            else:
                reader = csv.reader(io.StringIO(raw))
                for row_num, row in enumerate(reader, start=1):
                    if not row:
                        continue

                    handhole = (row[0] or "").strip()
                    if not handhole:
                        continue

                    reserve_raw = (
                        str(row[1]).strip()
                        if len(row) > 1 and row[1] is not None
                        else ""
                    )
                    required_raw = (
                        str(row[2]).strip()
                        if len(row) > 2 and row[2] is not None
                        else ""
                    )
                    warning = (
                        str(row[3]).strip()
                        if len(row) > 3 and row[3] is not None
                        else ""
                    )
                    order_raw = (
                        str(row[4]).strip()
                        if len(row) > 4 and row[4] is not None
                        else ""
                    )

                    parsed_rows.append(
                        {
                            "row": row_num,
                            "handhole": handhole,
                            "planned_reserve_ft_raw": reserve_raw,
                            "warning": warning,
                            "required_raw": required_raw,
                            "order_raw": order_raw,
                        }
                    )

        elif ext in ("xlsx", "xls"):
            wb = load_workbook(uploaded, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))

            if not rows:
                messages.warning(request, "The spreadsheet is empty.")
                return redirect(
                    "cable_installation:import_requirements_page", billing_id=billing.id
                )

            header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            has_header = "handhole" in header
            start_idx = 1 if has_header else 0

            if has_header:
                i_handhole = header.index("handhole")
                i_reserve = (
                    header.index("planned_reserve_ft")
                    if "planned_reserve_ft" in header
                    else None
                )
                i_required = header.index("required") if "required" in header else None
                i_warning = header.index("warning") if "warning" in header else None
                i_order = header.index("order") if "order" in header else None

                for row_num, r in enumerate(rows[start_idx:], start=2):
                    handhole = (
                        str(r[i_handhole]).strip()
                        if i_handhole < len(r) and r[i_handhole] is not None
                        else ""
                    )
                    if not handhole:
                        continue

                    reserve_raw = (
                        str(r[i_reserve]).strip()
                        if i_reserve is not None
                        and i_reserve < len(r)
                        and r[i_reserve] is not None
                        else ""
                    )
                    required_raw = (
                        str(r[i_required]).strip()
                        if i_required is not None
                        and i_required < len(r)
                        and r[i_required] is not None
                        else ""
                    )
                    warning = (
                        str(r[i_warning]).strip()
                        if i_warning is not None
                        and i_warning < len(r)
                        and r[i_warning] is not None
                        else ""
                    )
                    order_raw = (
                        str(r[i_order]).strip()
                        if i_order is not None
                        and i_order < len(r)
                        and r[i_order] is not None
                        else ""
                    )

                    parsed_rows.append(
                        {
                            "row": row_num,
                            "handhole": handhole,
                            "planned_reserve_ft_raw": reserve_raw,
                            "warning": warning,
                            "required_raw": required_raw,
                            "order_raw": order_raw,
                        }
                    )
            else:
                for row_num, r in enumerate(rows, start=1):
                    if not r:
                        continue

                    handhole = str(r[0]).strip() if r[0] is not None else ""
                    if not handhole:
                        continue

                    reserve_raw = (
                        str(r[1]).strip() if len(r) > 1 and r[1] is not None else ""
                    )
                    required_raw = (
                        str(r[2]).strip() if len(r) > 2 and r[2] is not None else ""
                    )
                    warning = (
                        str(r[3]).strip() if len(r) > 3 and r[3] is not None else ""
                    )
                    order_raw = (
                        str(r[4]).strip() if len(r) > 4 and r[4] is not None else ""
                    )

                    parsed_rows.append(
                        {
                            "row": row_num,
                            "handhole": handhole,
                            "planned_reserve_ft_raw": reserve_raw,
                            "warning": warning,
                            "required_raw": required_raw,
                            "order_raw": order_raw,
                        }
                    )
        else:
            messages.error(request, "Unsupported file type. Use .csv or .xlsx.")
            return redirect(
                "cable_installation:import_requirements_page", billing_id=billing.id
            )

    except Exception as exc:
        messages.error(request, f"Could not parse the file: {exc}")
        return redirect(
            "cable_installation:import_requirements_page", billing_id=billing.id
        )

    if not parsed_rows:
        messages.warning(request, "No valid rows found in the file.")
        return redirect(
            "cable_installation:import_requirements_page", billing_id=billing.id
        )

    seen_in_file = {}
    cleaned_rows = []
    file_duplicates = []

    for row in parsed_rows:
        key = (row["handhole"] or "").strip().lower()
        if not key:
            continue

        if key in seen_in_file:
            first = seen_in_file[key]
            warnings.append(
                f"Row {row['row']}: duplicated handhole in file — '{row['handhole']}' — duplicates row {first['row']}."
            )
            file_duplicates.append(
                {
                    "row": row["row"],
                    "handhole": row["handhole"],
                    "dup_of_row": first["row"],
                    "dup_of_handhole": first["handhole"],
                }
            )
            continue

        seen_in_file[key] = {
            "row": row["row"],
            "handhole": row["handhole"],
        }
        cleaned_rows.append(row)

    existing_by_key = {}
    existing_qs = billing.cable_requirements.all().order_by(
        "order", "sequence_no", "id"
    )

    for idx, req in enumerate(existing_qs, start=1):
        key = (req.handhole or "").strip().lower()
        if key and key not in existing_by_key:
            existing_by_key[key] = {
                "row_num": idx,
                "obj": req,
            }

    duplicates = []
    to_create = []

    for idx, row in enumerate(cleaned_rows):
        reserve_raw = row["planned_reserve_ft_raw"]
        order_raw = row["order_raw"]

        if reserve_raw == "":
            reserve_value = Decimal("0")
        else:
            reserve_value = _parse_decimal(reserve_raw)
            if reserve_value is None:
                errors.append(f"Row {row['row']}: invalid planned_reserve_ft.")
                continue
            if reserve_value < 0:
                errors.append(
                    f"Row {row['row']}: planned_reserve_ft cannot be negative."
                )
                continue

        if order_raw == "":
            order_value = idx
        else:
            try:
                order_value = int(str(order_raw).strip())
            except Exception:
                errors.append(f"Row {row['row']}: order must be an integer.")
                continue

            if order_value < 0:
                errors.append(f"Row {row['row']}: order cannot be negative.")
                continue

        try:
            required_value = _normalize_required_import_value(
                row["required_raw"], default=True
            )
        except ValueError as exc:
            errors.append(f"Row {row['row']}: {exc}")
            continue

        key = (row["handhole"] or "").strip().lower()
        existing = existing_by_key.get(key)

        if existing:
            req = existing["obj"]
            duplicates.append(
                {
                    "handhole": row["handhole"],
                    "existing_order": req.order,
                    "existing_required": bool(req.required),
                    "incoming_order": order_value,
                    "incoming_required": required_value,
                    "existing_warning": req.warning or "",
                    "incoming_warning": row["warning"] or "",
                }
            )
        else:
            to_create.append(
                {
                    "handhole": row["handhole"],
                    "planned_reserve_ft": str(reserve_value),
                    "required": required_value,
                    "warning": row["warning"] or "",
                    "order": order_value,
                }
            )

    request.session["cable_req_import_preview"] = {
        "billing_id": billing.id,
        "source_filename": filename,
        "to_create": to_create,
        "duplicates": duplicates,
        "file_duplicates": file_duplicates,
        "warnings": warnings,
        "errors": errors,
    }

    return render(
        request,
        "cable_installation/preview_import_requirements.html",
        {
            "billing": billing,
            "source_filename": filename,
            "to_create": to_create,
            "duplicates": duplicates,
            "file_duplicates": file_duplicates,
            "warnings": warnings,
            "errors": errors,
        },
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def confirm_import_requirements(request, billing_id):
    billing = get_object_or_404(
        SesionBilling,
        pk=billing_id,
        is_cable_installation=True,
    )

    payload = request.session.get("cable_req_import_preview") or {}
    if not payload or payload.get("billing_id") != billing.id:
        messages.error(
            request, "No preview data to confirm. Please upload the file again."
        )
        return redirect(
            "cable_installation:import_requirements_page", billing_id=billing.id
        )

    to_create = payload.get("to_create") or []
    if not isinstance(to_create, list) or not to_create:
        messages.info(request, "Nothing to create.")
        request.session.pop("cable_req_import_preview", None)
        return redirect(
            "cable_installation:configure_requirements", billing_id=billing.id
        )

    try:
        with transaction.atomic():
            next_seq = CableRequirement.next_sequence_for_billing(billing)
            created_count = 0

            existing_keys = {
                (x.handhole or "").strip().lower()
                for x in billing.cable_requirements.all()
            }

            for row in to_create:
                handhole = (row.get("handhole") or "").strip()
                if not handhole:
                    continue

                key = handhole.lower()
                if key in existing_keys:
                    continue

                reserve_value = _parse_decimal(row.get("planned_reserve_ft"))
                if reserve_value is None:
                    reserve_value = Decimal("0")

                CableRequirement.objects.create(
                    billing=billing,
                    sequence_no=next_seq,
                    handhole=handhole,
                    planned_reserve_ft=reserve_value,
                    required=bool(row.get("required", True)),
                    warning=(row.get("warning") or "").strip(),
                    order=int(row.get("order", 0)),
                )
                next_seq += 1
                created_count += 1
                existing_keys.add(key)

            _sync_cable_requirements_to_assignments(billing)

        messages.success(request, f"Created {created_count} new cable requirement(s).")
    except Exception as exc:
        messages.error(request, f"Could not apply imported requirements: {exc}")
        return redirect(
            "cable_installation:import_requirements_page", billing_id=billing.id
        )
    finally:
        request.session.pop("cable_req_import_preview", None)

    return redirect("cable_installation:configure_requirements", billing_id=billing.id)
