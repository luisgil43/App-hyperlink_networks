import os
import tempfile
from copy import copy
from decimal import Decimal, InvalidOperation
from io import BytesIO
from pathlib import Path

import xlsxwriter
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.files import File
from django.core.files.storage import default_storage as storage
from django.db import transaction
from django.http import FileResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.utils.http import http_date
from django.utils.text import slugify
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from operaciones.excel_images import tmp_jpeg_from_filefield
from operaciones.models import ReporteFotograficoJob, SesionBilling
from usuarios.decoradores import rol_requerido

from .models import CableAssignmentRequirement, CableEvidence, CableRequirement


def _parse_decimal(value):
    if value in (None, ""):
        return None
    try:
        return Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, ValueError, TypeError):
        return None


def _sync_cable_requirements_to_assignments(billing):
    assignments = list(billing.tecnicos_sesion.all())
    requirements = list(
        billing.cable_requirements.all().order_by("order", "sequence_no", "id")
    )

    if not assignments or not requirements:
        return

    existing_pairs = set(
        CableAssignmentRequirement.objects.filter(
            assignment__in=assignments,
            requirement__in=requirements,
        ).values_list("assignment_id", "requirement_id")
    )

    to_create = []
    for assignment in assignments:
        for requirement in requirements:
            key = (assignment.id, requirement.id)
            if key in existing_pairs:
                continue
            to_create.append(
                CableAssignmentRequirement(
                    assignment=assignment,
                    requirement=requirement,
                    status=CableAssignmentRequirement.STATUS_PENDING,
                )
            )

    if to_create:
        CableAssignmentRequirement.objects.bulk_create(to_create)


def _status_badge_label(status):
    mapping = {
        "asignado": "Assigned",
        "en_proceso": "In progress",
        "finalizado": "Finished",
        "en_revision_supervisor": "In supervisor review",
        "rechazado_supervisor": "Rejected by supervisor",
        "aprobado_supervisor": "Approved by supervisor",
        "rechazado_pm": "Rejected by PM",
        "aprobado_pm": "Approved by PM",
    }
    return mapping.get(status, status or "—")


def storage_file_exists(filefield) -> bool:
    if not filefield or not getattr(filefield, "name", ""):
        return False
    try:
        return filefield.storage.exists(filefield.name)
    except Exception:
        return False


def _required_shots():
    return [
        CableEvidence.SHOT_START_CABLE,
        CableEvidence.SHOT_END_CABLE,
        CableEvidence.SHOT_HANDHOLE,
    ]


def _shot_label(shot_type: str):
    if shot_type == CableEvidence.SHOT_START_CABLE:
        return "Start cable photo"
    if shot_type == CableEvidence.SHOT_END_CABLE:
        return "End cable photo"
    if shot_type == CableEvidence.SHOT_HANDHOLE:
        return "Handhole / Camera photo"
    return "Photo"


def _missing_measurement_fields(requirement):
    missing = []
    if requirement.start_ft is None:
        missing.append("Shared Start ft")
    if requirement.end_ft is None:
        missing.append("Shared End ft")
    return missing


def _requirement_measurement_complete(requirement):
    return requirement.start_ft is not None and requirement.end_ft is not None


def _requirement_has_any_progress(requirement):
    if requirement.start_ft is not None or requirement.end_ft is not None:
        return True

    return CableEvidence.objects.filter(
        assignment_requirement__requirement=requirement
    ).exists()


def _requirement_present_non_rejected_shots(requirement):
    return set(
        CableEvidence.objects.filter(assignment_requirement__requirement=requirement)
        .exclude(shot_type="")
        .exclude(review_status=CableEvidence.REVIEW_REJECTED)
        .values_list("shot_type", flat=True)
        .distinct()
    )


def _fmt_ft(value):
    if value is None:
        return ""
    value = Decimal(value)
    if value == value.to_integral():
        return str(int(value))
    return format(value.normalize(), "f")


def _client_report_note_text(requirement):
    notes = []

    evidences = (
        CableEvidence.objects.filter(
            assignment_requirement__requirement=requirement,
            shot_type=CableEvidence.SHOT_HANDHOLE,
        )
        .exclude(note__isnull=True)
        .exclude(note__exact="")
        .order_by("id")
    )

    for ev in evidences:
        txt = (ev.note or "").strip()
        if txt and txt not in notes:
            notes.append(txt)

    return "\n".join(notes)


def _client_report_work_date(billing):
    finished_dt = (
        billing.tecnicos_sesion.exclude(finalizado_en__isnull=True)
        .order_by("-finalizado_en")
        .values_list("finalizado_en", flat=True)
        .first()
    )

    dt = finished_dt or billing.creado_en

    if timezone.is_aware(dt):
        dt = timezone.localtime(dt)

    return dt.date()


def _billing_review_progress(billing):
    requirements = list(
        billing.cable_requirements.all().order_by("order", "sequence_no", "id")
    )

    details = {}
    completed = 0
    total = 0

    for req in requirements:
        present_shots = _requirement_present_non_rejected_shots(req)
        measurement_complete = _requirement_measurement_complete(req)
        missing_measurement_fields = _missing_measurement_fields(req)
        missing_shots = [
            shot for shot in _required_shots() if shot not in present_shots
        ]
        has_any_progress = _requirement_has_any_progress(req)

        # -----------------------------
        # REGLAS NUEVAS
        # -----------------------------
        if req.required:
            # Requerido:
            # - siempre cuenta
            # - completo solo si tiene medida completa + fotos requeridas
            is_visible = True
            counts_for_approval = True
            is_complete = measurement_complete and not missing_shots

        else:
            # No requerido:
            # - si no tiene medida ni fotos: no se muestra y no cuenta
            # - si tiene medida completa aunque no tenga fotos: cuenta como completo
            # - si tiene algo cargado pero medida incompleta: se muestra pendiente y bloquea aprobación
            if not has_any_progress:
                is_visible = False
                counts_for_approval = False
                is_complete = False
            else:
                is_visible = True
                counts_for_approval = True
                is_complete = measurement_complete

        if counts_for_approval:
            total += 1
            if is_complete:
                completed += 1

        details[req.id] = {
            "measurement_complete": measurement_complete,
            "missing_measurement_fields": missing_measurement_fields,
            "missing_shots": missing_shots,
            "is_complete": is_complete,
            "is_visible": is_visible,
            "counts_for_approval": counts_for_approval,
            "has_any_progress": has_any_progress,
        }

    percent = int(round((completed / total) * 100)) if total > 0 else 0

    return {
        "total": total,
        "completed": completed,
        "percent": percent,
        "can_approve": total > 0 and completed == total,
        "details": details,
    }


@login_required
@rol_requerido("supervisor", "admin", "pm")
def review_requirements(request, billing_id):
    billing = get_object_or_404(
        SesionBilling, pk=billing_id, is_cable_installation=True
    )

    _sync_cable_requirements_to_assignments(billing)

    rows = list(
        CableAssignmentRequirement.objects.filter(assignment__sesion=billing)
        .select_related("assignment", "assignment__tecnico", "requirement")
        .order_by(
            "requirement__order",
            "requirement__sequence_no",
            "assignment__id",
            "id",
        )
    )

    evidences = list(
        CableEvidence.objects.filter(
            assignment_requirement__assignment__sesion=billing
        ).select_related(
            "assignment_requirement",
            "assignment_requirement__assignment",
            "assignment_requirement__assignment__tecnico",
            "assignment_requirement__requirement",
        )
    )

    shot_order = {
        CableEvidence.SHOT_START_CABLE: 1,
        CableEvidence.SHOT_END_CABLE: 2,
        CableEvidence.SHOT_HANDHOLE: 3,
    }

    evidences.sort(
        key=lambda ev: (
            ev.assignment_requirement.requirement.order,
            ev.assignment_requirement.requirement.sequence_no,
            ev.assignment_requirement.requirement.id,
            shot_order.get(ev.shot_type, 99),
            ev.id,
        )
    )

    progress = _billing_review_progress(billing)

    grouped = {}
    for row in rows:
        grouped.setdefault(
            row.requirement_id,
            {
                "requirement": row.requirement,
                "rows": [],
                "evidences": [],
            },
        )
        grouped[row.requirement_id]["rows"].append(row)

    for ev in evidences:
        tech = ev.assignment_requirement.assignment.tecnico
        uploader_name = (
            getattr(tech, "get_full_name", lambda: "")() or tech.username or "—"
        ).strip()

        evidence_item = {
            "id": ev.id,
            "image": ev.image,
            "taken_at": ev.taken_at,
            "review_status": ev.review_status,
            "note": ev.note,
            "review_comment": ev.review_comment,
            "uploader_name": uploader_name,
            "shot_type": ev.shot_type,
            "shot_label": _shot_label(ev.shot_type),
        }

        grouped.setdefault(
            ev.assignment_requirement.requirement_id,
            {
                "requirement": ev.assignment_requirement.requirement,
                "rows": [],
                "evidences": [],
            },
        )
        grouped[ev.assignment_requirement.requirement_id]["evidences"].append(
            evidence_item
        )

    grouped_requirements = []
    for block in grouped.values():
        req = block["requirement"]
        req_progress = progress["details"].get(
            req.id,
            {
                "measurement_complete": False,
                "missing_measurement_fields": ["Shared Start ft", "Shared End ft"],
                "missing_shots": _required_shots(),
                "is_complete": False,
                "is_visible": True,
                "counts_for_approval": True,
                "has_any_progress": False,
            },
        )

        # No mostrar requerimientos NO obligatorios sin ningún avance
        if not req_progress["is_visible"]:
            continue

        block["is_complete"] = req_progress["is_complete"]
        block["measurement_complete"] = req_progress["measurement_complete"]
        block["missing_measurement_fields"] = req_progress["missing_measurement_fields"]
        block["missing_shots"] = [_shot_label(s) for s in req_progress["missing_shots"]]
        block["has_any_progress"] = req_progress["has_any_progress"]
        block["counts_for_approval"] = req_progress["counts_for_approval"]

        grouped_requirements.append(block)

    grouped_requirements.sort(
        key=lambda x: (
            x["requirement"].order,
            x["requirement"].sequence_no,
            x["requirement"].id,
        )
    )

    billing_status_label = _status_badge_label(getattr(billing, "estado", ""))

    can_review_project = getattr(billing, "estado", "") not in {
        "aprobado_supervisor",
        "aprobado_pm",
    }

    can_approve_project = can_review_project and progress["can_approve"]

    return render(
        request,
        "cable_installation/review_requirements.html",
        {
            "billing": billing,
            "grouped_requirements": grouped_requirements,
            "billing_status_label": billing_status_label,
            "can_review_project": can_review_project,
            "can_approve_project": can_approve_project,
            "progress_percent": progress["percent"],
            "progress_completed": progress["completed"],
            "progress_total": progress["total"],
        },
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def reviewer_update_shared_requirement(request, requirement_id):
    requirement = get_object_or_404(CableRequirement, pk=requirement_id)

    start_ft = _parse_decimal(request.POST.get("start_ft"))
    end_ft = _parse_decimal(request.POST.get("end_ft"))

    reserve = requirement.planned_reserve_ft or Decimal("0.00")
    mismatch = False
    mismatch_message = ""
    expected_low = None
    expected_high = None
    installed_ft = Decimal("0.00")
    override_confirmed = (request.POST.get("override_confirmed") or "").strip() in (
        "1",
        "true",
        "True",
    )

    if start_ft is not None:
        expected_low = start_ft - reserve
        expected_high = start_ft + reserve

    if start_ft is not None and end_ft is not None:
        installed_ft = abs(end_ft - start_ft)
        if installed_ft != reserve:
            mismatch = True
            mismatch_message = (
                f"Expected End ft could be {expected_low} or {expected_high} "
                f"(difference {reserve} ft). You entered {end_ft} "
                f"(difference {installed_ft} ft)."
            )

    requirement.start_ft = start_ft
    requirement.end_ft = end_ft
    requirement.save()

    if mismatch and not override_confirmed:
        return JsonResponse(
            {
                "ok": False,
                "requires_confirmation": True,
                "error": mismatch_message,
                "requirement": {
                    "id": requirement.id,
                    "start_ft": (
                        ""
                        if requirement.start_ft is None
                        else str(requirement.start_ft)
                    ),
                    "end_ft": (
                        "" if requirement.end_ft is None else str(requirement.end_ft)
                    ),
                    "installed_ft": str(requirement.installed_ft),
                    "expected_low": (
                        ""
                        if requirement.expected_end_ft_low is None
                        else str(requirement.expected_end_ft_low)
                    ),
                    "expected_high": (
                        ""
                        if requirement.expected_end_ft_high is None
                        else str(requirement.expected_end_ft_high)
                    ),
                    "warning_text": requirement.measurement_warning_text,
                },
            },
            status=409,
        )

    return JsonResponse(
        {
            "ok": True,
            "message": "Shared measurement saved.",
            "requirement": {
                "id": requirement.id,
                "start_ft": (
                    "" if requirement.start_ft is None else str(requirement.start_ft)
                ),
                "end_ft": "" if requirement.end_ft is None else str(requirement.end_ft),
                "installed_ft": str(requirement.installed_ft),
                "expected_low": (
                    ""
                    if requirement.expected_end_ft_low is None
                    else str(requirement.expected_end_ft_low)
                ),
                "expected_high": (
                    ""
                    if requirement.expected_end_ft_high is None
                    else str(requirement.expected_end_ft_high)
                ),
                "warning_text": requirement.measurement_warning_text,
                "end_ft_overridden": requirement.end_ft_overridden,
            },
        }
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def reviewer_update_requirement(request, row_id):
    row = get_object_or_404(
        CableAssignmentRequirement.objects.select_related(
            "requirement", "assignment", "assignment__sesion"
        ),
        pk=row_id,
    )

    status = (request.POST.get("status") or "").strip()
    note = (request.POST.get("note") or "").strip()
    supervisor_note = (request.POST.get("supervisor_note") or "").strip()

    if status and status not in dict(CableAssignmentRequirement.STATUS_CHOICES):
        return JsonResponse({"ok": False, "error": "Invalid status."}, status=400)

    if status:
        row.status = status
    row.note = note
    row.supervisor_note = supervisor_note
    row.last_reviewed_at = timezone.now()
    row.last_reviewed_by = request.user
    row.save(
        update_fields=[
            "status",
            "note",
            "supervisor_note",
            "last_reviewed_at",
            "last_reviewed_by",
            "updated_at",
        ]
    )

    return JsonResponse(
        {
            "ok": True,
            "row": {
                "id": row.id,
                "status": row.status,
                "note": row.note,
                "supervisor_note": row.supervisor_note,
            },
        }
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def approve_evidence(request, evidence_id):
    ev = get_object_or_404(
        CableEvidence.objects.select_related(
            "assignment_requirement",
            "assignment_requirement__assignment",
            "assignment_requirement__assignment__sesion",
        ),
        pk=evidence_id,
    )
    ev.approve(request.user)

    row = ev.assignment_requirement
    if row.status == CableAssignmentRequirement.STATUS_REJECTED:
        row.status = CableAssignmentRequirement.STATUS_PENDING
        row.save(update_fields=["status", "updated_at"])

    return JsonResponse({"ok": True})


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def reject_evidence(request, evidence_id):
    ev = get_object_or_404(
        CableEvidence.objects.select_related(
            "assignment_requirement",
            "assignment_requirement__assignment",
            "assignment_requirement__assignment__sesion",
        ),
        pk=evidence_id,
    )
    comment = (request.POST.get("comment") or "").strip()
    if not comment:
        return JsonResponse({"ok": False, "error": "Comment is required."}, status=400)

    row = ev.assignment_requirement
    assignment = row.assignment
    billing = assignment.sesion
    now = timezone.now()

    with transaction.atomic():
        row.status = CableAssignmentRequirement.STATUS_REJECTED
        row.supervisor_note = comment
        row.last_reviewed_at = now
        row.last_reviewed_by = request.user
        row.save(
            update_fields=[
                "status",
                "supervisor_note",
                "last_reviewed_at",
                "last_reviewed_by",
                "updated_at",
            ]
        )

        billing.estado = "rechazado_supervisor"
        billing.save(update_fields=["estado"])

        for asg in billing.tecnicos_sesion.all():
            update_fields = ["estado"]
            asg.estado = "rechazado_supervisor"

            if hasattr(asg, "supervisor_revisado_en"):
                asg.supervisor_revisado_en = now
                update_fields.append("supervisor_revisado_en")

            if hasattr(asg, "supervisor_comentario"):
                asg.supervisor_comentario = comment
                update_fields.append("supervisor_comentario")

            if hasattr(asg, "reintento_habilitado"):
                asg.reintento_habilitado = True
                update_fields.append("reintento_habilitado")

            asg.save(update_fields=update_fields)

        ev.delete()

    return JsonResponse({"ok": True})


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def bulk_approve_evidences(request, billing_id):
    billing = get_object_or_404(
        SesionBilling, pk=billing_id, is_cable_installation=True
    )

    ids = request.POST.getlist("ids[]")
    if not ids:
        return JsonResponse(
            {"ok": False, "error": "No evidences selected."}, status=400
        )

    evidences = list(
        CableEvidence.objects.filter(
            id__in=ids,
            assignment_requirement__assignment__sesion=billing,
        )
    )

    for ev in evidences:
        ev.approve(request.user)
        row = ev.assignment_requirement
        if row.status == CableAssignmentRequirement.STATUS_REJECTED:
            row.status = CableAssignmentRequirement.STATUS_PENDING
            row.save(update_fields=["status", "updated_at"])

    return JsonResponse({"ok": True, "approved": len(evidences)})


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def approve_project_review(request, billing_id):
    billing = get_object_or_404(
        SesionBilling, pk=billing_id, is_cable_installation=True
    )

    progress = _billing_review_progress(billing)
    if not progress["can_approve"]:
        messages.error(
            request,
            "You cannot approve this project yet. All required handholes must be fully completed first.",
        )
        return redirect("cable_installation:review_requirements", billing_id=billing.id)

    now = timezone.now()

    with transaction.atomic():
        billing.estado = "aprobado_supervisor"
        billing.save(update_fields=["estado"])

        for assignment in billing.tecnicos_sesion.all():
            update_fields = ["estado"]
            assignment.estado = "aprobado_supervisor"

            if hasattr(assignment, "supervisor_revisado_en"):
                assignment.supervisor_revisado_en = now
                update_fields.append("supervisor_revisado_en")

            if hasattr(assignment, "reintento_habilitado"):
                assignment.reintento_habilitado = False
                update_fields.append("reintento_habilitado")

            assignment.save(update_fields=update_fields)

    messages.success(request, "Project approved by supervisor.")
    return redirect("cable_installation:review_requirements", billing_id=billing.id)


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def reject_project_review(request, billing_id):
    billing = get_object_or_404(
        SesionBilling, pk=billing_id, is_cable_installation=True
    )

    comment = (request.POST.get("comment") or "").strip()
    if not comment:
        messages.error(request, "Comment is required to reject the project.")
        return redirect("cable_installation:review_requirements", billing_id=billing.id)

    now = timezone.now()

    with transaction.atomic():
        billing.estado = "rechazado_supervisor"
        billing.save(update_fields=["estado"])

        for assignment in billing.tecnicos_sesion.all():
            update_fields = ["estado"]
            assignment.estado = "rechazado_supervisor"

            if hasattr(assignment, "supervisor_revisado_en"):
                assignment.supervisor_revisado_en = now
                update_fields.append("supervisor_revisado_en")

            if hasattr(assignment, "supervisor_comentario"):
                assignment.supervisor_comentario = comment
                update_fields.append("supervisor_comentario")

            if hasattr(assignment, "reintento_habilitado"):
                assignment.reintento_habilitado = True
                update_fields.append("reintento_habilitado")

            assignment.save(update_fields=update_fields)

    messages.warning(request, "Project rejected by supervisor.")
    return redirect("cable_installation:review_requirements", billing_id=billing.id)


@login_required
@rol_requerido("supervisor", "admin", "pm")
def export_client_excel(request, billing_id):
    billing = get_object_or_404(
        SesionBilling, pk=billing_id, is_cable_installation=True
    )

    requirements = list(
        billing.cable_requirements.all().order_by("order", "sequence_no", "id")
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Report"

    # Quitar líneas de división del resto de la hoja
    ws.sheet_view.showGridLines = False

    # Anchos
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 40
    ws.column_dimensions["F"].width = 16

    thin_side = Side(style="thin", color="000000")
    no_side = Side(style=None)

    def make_border(left=False, right=False, top=False, bottom=False):
        return Border(
            left=thin_side if left else no_side,
            right=thin_side if right else no_side,
            top=thin_side if top else no_side,
            bottom=thin_side if bottom else no_side,
        )

    title_font = Font(bold=True, size=12, color="000000")
    red_font = Font(bold=True, size=12, color="FF0000")
    notes_red_font = Font(size=12, color="FF0000")
    normal_font = Font(size=12, color="000000")

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # ---------------------------
    # Header row 1
    # ---------------------------
    ws["A1"] = billing.proyecto_id or ""
    ws["A1"].font = red_font
    ws["A1"].alignment = left
    ws["A1"].border = make_border(left=True, right=True, top=True, bottom=True)

    ws.merge_cells("B1:D1")
    ws["B1"] = "FIBER"
    ws["B1"].font = title_font
    ws["B1"].alignment = center

    # Bordes del merge B1:D1
    ws["B1"].border = make_border(left=True, top=True, bottom=True)
    ws["C1"].border = make_border(top=True, bottom=True)
    ws["D1"].border = make_border(right=True, top=True, bottom=True)

    # Completar líneas verticales arriba en E y F
    ws["E1"] = ""
    ws["E1"].border = make_border(left=True, right=True, top=True)

    ws["F1"] = ""
    ws["F1"].border = make_border(left=True, right=True, top=True)

    # ---------------------------
    # Header row 2
    # ---------------------------
    ws["A2"] = "HH Number"
    ws["B2"] = "Seq. #1(In)"
    ws["C2"] = "Seq. #1(Out)"
    ws["D2"] = "Slack"
    ws["E2"] = "Notes"
    ws["F2"] = "Week Ending"

    ws["A2"].font = title_font
    ws["B2"].font = title_font
    ws["C2"].font = title_font
    ws["D2"].font = title_font
    ws["E2"].font = title_font
    ws["F2"].font = red_font

    ws["A2"].alignment = left
    ws["B2"].alignment = left
    ws["C2"].alignment = left
    ws["D2"].alignment = left
    ws["E2"].alignment = center
    ws["F2"].alignment = center

    ws["A2"].border = make_border(left=True, right=True, top=True, bottom=True)
    ws["B2"].border = make_border(left=True, right=True, top=True, bottom=True)
    ws["C2"].border = make_border(left=True, right=True, top=True, bottom=True)
    ws["D2"].border = make_border(left=True, right=True, top=True, bottom=True)
    ws["E2"].border = make_border(left=True, right=True, top=True, bottom=True)
    ws["F2"].border = make_border(left=True, right=True, top=True, bottom=True)

    # ---------------------------
    # Data
    # ---------------------------
    week_ending = _client_report_work_date(billing)
    start_row = 3
    last_row = start_row + len(requirements) - 1 if requirements else start_row

    for row_idx, req in enumerate(requirements, start=start_row):
        # Nota: solo la nota de la foto handhole / camera
        handhole_evidence = (
            CableEvidence.objects.filter(
                assignment_requirement__requirement=req,
                shot_type=CableEvidence.SHOT_HANDHOLE,
            )
            .exclude(review_status=CableEvidence.REVIEW_REJECTED)
            .order_by("-taken_at", "-id")
            .first()
        )

        note_text = (handhole_evidence.note or "").strip() if handhole_evidence else ""
        is_last = row_idx == last_row

        ws.row_dimensions[row_idx].height = 24 if not note_text else 30

        # A = HH Number
        a = ws.cell(row_idx, 1, req.handhole or "")
        a.font = normal_font
        a.alignment = left
        a.border = make_border(left=True, bottom=True)

        # B = Seq. #1(In)
        b = ws.cell(row_idx, 2)
        b.value = float(req.start_ft) if req.start_ft is not None else ""
        b.font = normal_font
        b.alignment = center
        b.border = make_border(bottom=True)

        # C = Seq. #1(Out)
        c = ws.cell(row_idx, 3)
        c.value = float(req.end_ft) if req.end_ft is not None else ""
        c.font = normal_font
        c.alignment = center
        c.border = make_border(bottom=True)

        # D = Slack
        d = ws.cell(row_idx, 4)
        d.value = (
            float(req.planned_reserve_ft) if req.planned_reserve_ft is not None else ""
        )
        d.font = normal_font
        d.alignment = center
        d.border = make_border(bottom=True)

        # E = Notes (rojo)
        e = ws.cell(row_idx, 5, note_text)
        e.font = notes_red_font
        e.alignment = left_wrap
        e.border = make_border(bottom=True)

        # F = Week Ending
        f = ws.cell(row_idx, 6)
        f.value = week_ending.strftime("%m-%d-%y") if week_ending else ""
        f.font = normal_font
        f.alignment = center
        f.border = make_border(right=True, bottom=True)

        # Última fila: cerrar borde inferior
        if is_last:
            a.border = make_border(left=True, bottom=True)
            b.border = make_border(bottom=True)
            c.border = make_border(bottom=True)
            d.border = make_border(bottom=True)
            e.border = make_border(bottom=True)
            f.border = make_border(right=True, bottom=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"{billing.proyecto_id}_CLIENT_REPORT.xlsx"
    return FileResponse(
        output,
        as_attachment=True,
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


class CableReportCancelled(Exception):
    pass


def _cable_report_project_key(billing: SesionBilling) -> str:
    proj_slug = (
        slugify(billing.proyecto_id or f"billing-{billing.id}")
        or f"billing-{billing.id}"
    )
    sess_tag = f"{proj_slug}-{billing.id}"
    return f"cable_installation/reports/{sess_tag}/{sess_tag}.xlsx"


def _cable_report_evidences_qs(billing):
    return (
        CableEvidence.objects.filter(assignment_requirement__assignment__sesion=billing)
        .select_related(
            "assignment_requirement",
            "assignment_requirement__requirement",
            "assignment_requirement__assignment",
            "assignment_requirement__assignment__tecnico",
        )
        .order_by(
            "assignment_requirement__requirement__order",
            "assignment_requirement__requirement__sequence_no",
            "assignment_requirement__requirement__id",
            "id",
        )
    )


def _cable_report_sort_key(ev):
    shot_order = {
        CableEvidence.SHOT_START_CABLE: 1,
        CableEvidence.SHOT_END_CABLE: 2,
        CableEvidence.SHOT_HANDHOLE: 3,
    }
    req = ev.assignment_requirement.requirement
    return (
        req.order,
        req.sequence_no,
        req.id,
        shot_order.get(ev.shot_type, 99),
        ev.id,
    )


def _fmt_ft_compact(value):
    if value is None:
        return ""
    value = Decimal(value)
    if value == value.to_integral():
        return str(int(value))
    return format(value.normalize(), "f")


def _cable_report_block_title(ev):
    req = ev.assignment_requirement.requirement
    handhole = (req.handhole or "").strip()

    if ev.shot_type == CableEvidence.SHOT_START_CABLE:
        start_txt = _fmt_ft_compact(req.start_ft)
        return f"{handhole} | Start {start_txt}" if start_txt else handhole

    if ev.shot_type == CableEvidence.SHOT_END_CABLE:
        end_txt = _fmt_ft_compact(req.end_ft)
        return f"{handhole} | End {end_txt}" if end_txt else handhole

    return handhole


def _cable_report_note(ev):
    return (ev.note or "").strip()


def _xlsx_path_cable_photo_report(
    billing: SesionBilling, progress_cb=None, should_cancel=None
):
    import io
    import tempfile

    import xlsxwriter
    from PIL import Image

    evs = list(_cable_report_evidences_qs(billing))
    evs.sort(key=_cable_report_sort_key)

    tmp_xlsx = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_xlsx.close()

    wb = xlsxwriter.Workbook(tmp_xlsx.name, {"in_memory": False})
    ws = wb.add_worksheet("CABLE REPORT")
    ws.hide_gridlines(2)

    fmt_title = wb.add_format(
        {
            "bold": True,
            "align": "center",
            "valign": "vcenter",
            "border": 1,
            "bg_color": "#E8EEF7",
            "font_size": 12,
        }
    )
    fmt_head = wb.add_format(
        {
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "bold": True,
            "text_wrap": True,
            "bg_color": "#F5F7FB",
            "font_size": 11,
        }
    )
    fmt_box = wb.add_format({"border": 1})
    fmt_note = wb.add_format(
        {
            "border": 1,
            "align": "center",
            "valign": "vcenter",
            "text_wrap": True,
            "font_size": 9,
        }
    )

    BLOCK_COLS = 6
    SEP_COLS = 1
    LEFT_COL = 0
    RIGHT_COL = LEFT_COL + BLOCK_COLS + SEP_COLS

    HEAD_ROWS = 1
    ROWS_IMG = 12
    ROW_NOTE = 1
    ROW_SPACE = 1
    BLOCK_ROWS = HEAD_ROWS + ROWS_IMG + ROW_NOTE

    COL_W = 13
    IMG_ROW_H = 18

    def col_px(w):
        return int(w * 7 + 5)

    def row_px(h):
        return int(h * 4 / 3)

    max_w_px = BLOCK_COLS * col_px(COL_W)
    max_h_px = ROWS_IMG * row_px(IMG_ROW_H)

    for c in range(LEFT_COL, LEFT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)
    ws.set_column(LEFT_COL + BLOCK_COLS, LEFT_COL + BLOCK_COLS, 2)
    for c in range(RIGHT_COL, RIGHT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)

    ws.merge_range(
        0, 0, 0, RIGHT_COL + BLOCK_COLS - 1, billing.proyecto_id or "", fmt_title
    )
    cur_row = 2

    def draw_block(r, c, ev):
        title_txt = _cable_report_block_title(ev)
        note_txt = _cable_report_note(ev)

        ws.merge_range(r, c, r + HEAD_ROWS - 1, c + BLOCK_COLS - 1, title_txt, fmt_head)
        for rr in range(r, r + HEAD_ROWS):
            ws.set_row(rr, 22)

        img_top = r + HEAD_ROWS
        for rr in range(img_top, img_top + ROWS_IMG):
            ws.set_row(rr, IMG_ROW_H)
        ws.merge_range(
            img_top, c, img_top + ROWS_IMG - 1, c + BLOCK_COLS - 1, "", fmt_box
        )

        try:
            ev.image.open("rb")
            raw = ev.image.read()
            ev.image.close()

            image_data = io.BytesIO(raw)
            image_data.seek(0)

            with Image.open(io.BytesIO(raw)) as im:
                im = im.convert("RGB")
                w, h = im.size

            sx = max_w_px / float(w)
            sy = max_h_px / float(h)
            scale = min(sx, sy, 1.0)

            scaled_w = int(w * scale)
            scaled_h = int(h * scale)
            x_off = max((max_w_px - scaled_w) // 2, 0)
            y_off = max((max_h_px - scaled_h) // 2, 0)

            ws.insert_image(
                img_top,
                c,
                "image.jpg",
                {
                    "image_data": image_data,
                    "x_scale": scale,
                    "y_scale": scale,
                    "x_offset": x_off,
                    "y_offset": y_off,
                    "object_position": 1,
                },
            )
        except Exception as e:
            print(f"[CABLE REPORT] image insert failed for evidence {ev.id}: {e}")

        note_row = img_top + ROWS_IMG
        ws.merge_range(note_row, c, note_row, c + BLOCK_COLS - 1, note_txt, fmt_note)
        ws.set_row(note_row, 28 if note_txt else 20)

    idx = 0
    total = len(evs)

    for ev in evs:
        if callable(should_cancel) and should_cancel(idx):
            wb.close()
            raise CableReportCancelled()

        if idx % 2 == 0:
            draw_block(cur_row, LEFT_COL, ev)
        else:
            draw_block(cur_row, RIGHT_COL, ev)
            cur_row += BLOCK_ROWS + ROW_SPACE

        idx += 1

        if callable(progress_cb):
            progress_cb(idx, total)

    if idx % 2 == 1:
        cur_row += BLOCK_ROWS + ROW_SPACE

    wb.close()
    return tmp_xlsx.name


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def generate_cable_photo_report(request, billing_id):
    from usuarios.schedulers import enqueue_cable_photo_report

    billing = get_object_or_404(
        SesionBilling,
        pk=billing_id,
        is_cable_installation=True,
    )

    total_photos = _cable_report_evidences_qs(billing).count()
    if total_photos == 0:
        messages.warning(
            request,
            "This project has no photos yet. Upload at least one photo before generating the report.",
        )
        return redirect("cable_installation:review_requirements", billing_id=billing.id)

    last_job = (
        ReporteFotograficoJob.objects.filter(sesion=billing)
        .exclude(log__icontains="[partial]")
        .order_by("-creado_en")
        .first()
    )

    if last_job and last_job.estado in ("pendiente", "procesando"):
        messages.info(
            request,
            "The report is already being generated in background.",
        )
        return redirect("cable_installation:review_requirements", billing_id=billing.id)

    job = ReporteFotograficoJob.objects.create(
        sesion=billing,
        log="[cable-final] queued\n",
        total=0,
        procesadas=0,
    )

    transaction.on_commit(lambda: enqueue_cable_photo_report(job.id))

    messages.info(
        request,
        "Generating cable photo report in background. It will be available when ready.",
    )
    return redirect("cable_installation:review_requirements", billing_id=billing.id)


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_GET
def cable_photo_report_status(request, billing_id):
    billing = get_object_or_404(
        SesionBilling,
        pk=billing_id,
        is_cable_installation=True,
    )

    job = (
        ReporteFotograficoJob.objects.filter(sesion=billing)
        .exclude(log__icontains="[partial]")
        .order_by("-creado_en")
        .first()
    )

    ready = bool(
        billing.reporte_fotografico and storage_file_exists(billing.reporte_fotografico)
    )

    if not job:
        return JsonResponse(
            {
                "state": "ok" if ready else "none",
                "processed": 0,
                "total": 0,
                "error": "",
                "ready": ready,
            }
        )

    state_map = {
        "pendiente": "pending",
        "procesando": "processing",
        "ok": "ok",
        "error": "error",
    }

    return JsonResponse(
        {
            "state": state_map.get(job.estado, job.estado),
            "processed": job.procesadas or 0,
            "total": job.total or 0,
            "error": job.error or "",
            "ready": ready,
        }
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
def download_cable_photo_report(request, billing_id):
    billing = get_object_or_404(
        SesionBilling,
        pk=billing_id,
        is_cable_installation=True,
    )

    if not billing.reporte_fotografico or not storage_file_exists(
        billing.reporte_fotografico
    ):
        messages.warning(request, "The cable photo report is not ready yet.")
        return redirect("cable_installation:review_requirements", billing_id=billing.id)

    f = billing.reporte_fotografico.open("rb")
    filename = f"{billing.proyecto_id}_CABLE_PHOTO_REPORT.xlsx"

    resp = FileResponse(f, as_attachment=True, filename=filename)
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = http_date(0)
    return resp
