# operaciones/views.py

import calendar
import csv
import io
import json
import locale
import logging
import os
import re
import shutil
import xml.etree.ElementTree as ET
import zipfile
from copy import copy as _copy
from datetime import date
from datetime import date as _date
from datetime import datetime, timedelta
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO
from tempfile import NamedTemporaryFile
from typing import Optional
from urllib.parse import urlencode
from uuid import uuid4

import boto3
import pandas as pd
import requests
import xlsxwriter
import xlwt
from botocore.client import Config
from botocore.exceptions import ClientError
from django.conf import settings
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.core.exceptions import FieldError, ValidationError
from django.core.files.storage import default_storage
from django.core.files.storage import default_storage as storage
from django.core.paginator import Paginator
from django.db import models
from django.db import models as dj_models
from django.db import transaction
from django.db.models import (Case, Count, DecimalField, Exists, F, FloatField,
                              IntegerField, OuterRef, Prefetch, Q, Sum, Value,
                              When)
from django.db.models.functions import Coalesce, Length, Substr, Upper
from django.http import (FileResponse, HttpResponse, HttpResponseBadRequest,
                         HttpResponseForbidden, HttpResponseNotAllowed,
                         HttpResponseRedirect, HttpResponseServerError,
                         JsonResponse)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.encoding import force_str
from django.utils.html import escape
from django.utils.http import urlencode
from django.utils.text import slugify
from django.utils.timezone import is_aware, now
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_exempt, csrf_protect
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image as XLImage  # para copiar imágenes
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.pdfgen import canvas
from reportlab.platypus import (Image, Paragraph, SimpleDocTemplate, Spacer,
                                Table, TableStyle)

from access_control.services import user_can as access_user_can
from core.decorators import project_object_access_required
from core.permissions import (filter_queryset_by_access, projects_ids_for_user,
                              user_has_project_access)
from facturacion.models import CartolaMovimiento, Proyecto
from fleet.models import VehicleOdometerEvent, VehicleService
from operaciones.forms import PaymentApproveForm, PaymentRejectForm
from operaciones.models import AdjustmentEntry  # <-- IMPORTA EL MODELO
from operaciones.models import ItemBillingTecnico, SesionBilling, WeeklyPayment
from usuarios.decoradores import rol_requerido
from usuarios.models import CustomUser  # ajusta si tu user model es otro
from usuarios.utils import \
    crear_notificacion  # asegúrate de tener esta función

from .forms import MovimientoUsuarioForm  # crearemos este form
from .forms import PrecioActividadTecnicoForm  # lo definimos abajo
from .forms import (ImportarPreciosForm, PaymentApproveForm,  # <-- TUS FORMS
                    PaymentMarkPaidForm, PaymentRejectForm)
from .models import PrecioActividadTecnico  # <-- TU MODELO DE PRECIOS
from .models import SesionBilling  # ajusta a tu ruta real
from .models import (AdjustmentEntry, BillingPayWeekSnapshot,
                     EvidenciaFotoBilling, ItemBilling, ItemBillingTecnico,
                     RequisitoFotoBilling, SesionBillingTecnico, WeeklyPayment)
from .services.weekly import \
    materialize_week_for_payments  # crea/actualiza solo la semana indicada
from .services.weekly import \
    sync_weekly_totals_no_create  # versión que NO crea

try:
    from operaciones.models import AdjustmentEntry
except Exception:
    AdjustmentEntry = None

from collections import OrderedDict
from datetime import date, timedelta
from decimal import Decimal

from django.db import transaction
from django.db.models import Q
from django.views.decorators.http import require_GET, require_POST

# 👇 nuevo
from facturacion.models import Proyecto  # ajusta el app si está en otro lado
from operaciones.models import BillingPayWeekSnapshot, PrecioActividadTecnico

# type: ignore


WEEK_RE = re.compile(r"^\d{4}-W\d{2}$")
# --- Direct upload (receipts/rendiciones) ---


RECEIPT_MAX_MB = int(getattr(settings, "RECEIPT_DIRECT_UPLOADS_MAX_MB", 25))
RECEIPTS_SAFE_PREFIX = getattr(
    settings, "DIRECT_UPLOADS_RECEIPTS_PREFIX", "operaciones/rendiciones/"
)


def _billing_access_context(user):
    """
    Permisos visuales configurados desde Access Matrix.
    """

    can_create_billing = access_user_can(user, "billing.create_billing")
    can_edit_billing = access_user_can(user, "billing.edit_billing")

    can_view_technical_billing_amounts = access_user_can(
        user,
        "billing.view_technical_amounts",
    )

    can_view_company_billing_amounts = access_user_can(
        user,
        "billing.view_company_amounts",
    )

    can_view_real_company_billing = access_user_can(
        user,
        "billing.view_real_company_billing",
    )

    can_view_billing_difference = access_user_can(
        user,
        "billing.view_billing_difference",
    )

    can_edit_real_week = access_user_can(
        user,
        "billing.edit_real_week",
    )
    can_delete_billing = access_user_can(
        user,
        "billing.delete_billing",
    )

    return {
        "can_create_billing": can_create_billing,
        "can_edit_billing": can_edit_billing,
        "can_edit_items": can_edit_billing,
        "can_edit_real_week": can_edit_real_week,
        "can_delete_billing": can_delete_billing,
        "can_view_technical_billing_amounts": can_view_technical_billing_amounts,
        "can_view_company_billing_amounts": can_view_company_billing_amounts,
        "can_view_real_company_billing": can_view_real_company_billing,
        "can_view_billing_difference": can_view_billing_difference,
        # Alias por retrocompatibilidad
        "can_view_technical_amounts": can_view_technical_billing_amounts,
        "can_view_company_amounts": can_view_company_billing_amounts,
    }


def _attach_accounting_lock_flags_to_sessions(sessions):
    """
    Optimización para listados.

    Marca en cada sesión:
      - has_paid_work_type_lines
      - can_reopen_billing_fast
      - can_delete_billing_fast
      - can_edit_billing_fast

    Importante:
    - NO reemplaza _session_is_paid_locked() para acciones POST.
    - Solo evita N consultas por fila en listar_billing().
    - Usa relaciones prefetched:
        pay_week_snapshots
        tecnicos_sesion
        weekly_payment dentro del snapshot
    """

    sessions = list(sessions or [])
    if not sessions:
        return sessions

    # ------------------------------------------------------------
    # 1) Primera pasada en memoria usando snapshots ya prefetched
    # ------------------------------------------------------------
    by_session_id = {}
    candidate_tech_ids = set()
    candidate_weeks = set()
    session_candidate_pairs = {}

    for s in sessions:
        by_session_id[s.id] = s

        is_locked = False
        snaps = []

        try:
            snaps = list(s.pay_week_snapshots.all())
        except Exception:
            snaps = []

        # IDs de técnicos de la sesión, ya prefetched
        tech_ids = []
        try:
            tech_ids = [
                int(x)
                for x in s.tecnicos_sesion.all().values_list("tecnico_id", flat=True)
                if x
            ]
        except Exception:
            try:
                tech_ids = [
                    int(a.tecnico_id)
                    for a in s.tecnicos_sesion.all()
                    if getattr(a, "tecnico_id", None)
                ]
            except Exception:
                tech_ids = []

        weeks = set()

        for snap in snaps:
            # En listar_billing ya prefetcheas is_adjustment=False,
            # pero dejamos esta defensa por si cambias el prefetch.
            if getattr(snap, "is_adjustment", False):
                continue

            if getattr(snap, "adjustment_of_id", None):
                continue

            payment_status = (
                (getattr(snap, "payment_status", "") or "").strip().lower()
            )

            if payment_status == "paid":
                is_locked = True
                break

            if getattr(snap, "paid_at", None):
                is_locked = True
                break

            wp = getattr(snap, "weekly_payment", None)
            if wp and (getattr(wp, "status", "") or "").strip().lower() == "paid":
                is_locked = True
                break

            week_value = (
                (getattr(snap, "semana_resultado", "") or "").strip().upper()
            )
            if week_value:
                weeks.add(week_value)

        # Legacy / fallback si no hay semanas desde snapshots
        if not weeks:
            possible_weeks = [
                (getattr(s, "semana_pago_real", "") or "").strip().upper(),
                (getattr(s, "semana_pago_proyectada", "") or "").strip().upper(),
                (getattr(s, "discount_week", "") or "").strip().upper(),
            ]
            weeks = {w for w in possible_weeks if w}

        s.has_paid_work_type_lines = bool(is_locked)

        # Si todavía no quedó bloqueada por snapshot directo,
        # guardamos pares técnico/semana para una consulta global.
        if not is_locked and tech_ids and weeks:
            pairs = set()
            for tid in tech_ids:
                candidate_tech_ids.add(tid)
                for wk in weeks:
                    candidate_weeks.add(wk)
                    pairs.add((tid, wk))

            session_candidate_pairs[s.id] = pairs

    # ------------------------------------------------------------
    # 2) Consulta global única a WeeklyPayment paid
    # ------------------------------------------------------------
    paid_pairs = set()

    if candidate_tech_ids and candidate_weeks:
        paid_pairs = set(
            WeeklyPayment.objects.filter(
                technician_id__in=candidate_tech_ids,
                week__in=candidate_weeks,
                status="paid",
            ).values_list("technician_id", "week")
        )

    # ------------------------------------------------------------
    # 3) Aplicar paid_pairs + markers legacy en finance_note
    # ------------------------------------------------------------
    for s in sessions:
        if getattr(s, "has_paid_work_type_lines", False):
            locked = True
        else:
            locked = False

            pairs = session_candidate_pairs.get(s.id, set())

            if paid_pairs and pairs:
                if any(pair in paid_pairs for pair in pairs):
                    locked = True

            if not locked:
                note = getattr(s, "finance_note", "") or ""
                for tech_id, wk in pairs:
                    marker = f"[TECH_WEEKLY_PAYMENT_PAID:{tech_id}:{wk}]"
                    if marker in note:
                        locked = True
                        break

        s.has_paid_work_type_lines = bool(locked)
        s.can_reopen_billing_fast = (
            s.estado in ("aprobado_supervisor", "aprobado_pm", "aprobado_finanzas")
            and not locked
        )
        s.can_delete_billing_fast = not locked
        s.can_edit_billing_fast = not locked

    return sessions

def _session_is_paid_locked(sesion) -> bool:
    """
    Bloqueo contable del Billing.

    Regla:
    - Si existe al menos una línea productiva de BillingPayWeekSnapshot
      marcada como paid para esta sesión, el Billing queda bloqueado.
    - También queda bloqueado si existe un WeeklyPayment paid asociado
      a algún técnico + semana efectiva de esta sesión.
    - Compatibilidad legacy:
      si no hay snapshots productivos, usa semana_pago_real /
      semana_pago_proyectada / discount_week y también revisa markers antiguos
      en finance_note.

    Importante:
    - Este bloqueo aplica para reabrir, eliminar y editar.
    - No depende de que finance_status del Billing sea "paid".
    """

    if not sesion:
        return False

    snapshot_model = BillingPayWeekSnapshot
    snapshot_fields = {f.name for f in snapshot_model._meta.get_fields()}

    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields
    has_payment_status = "payment_status" in snapshot_fields
    has_paid_at = "paid_at" in snapshot_fields
    has_weekly_payment = "weekly_payment" in snapshot_fields

    # ==========================================================
    # 1) Snapshots productivos de esta sesión
    # ==========================================================
    snaps_qs = snapshot_model.objects.filter(sesion=sesion)

    if has_is_adjustment:
        snaps_qs = snaps_qs.filter(is_adjustment=False)
    elif has_adjustment_of:
        snaps_qs = snaps_qs.filter(adjustment_of__isnull=True)

    # 1.1) Bloqueo directo por payment_status='paid'
    if has_payment_status:
        if snaps_qs.filter(payment_status="paid").exists():
            return True

    # 1.2) Bloqueo por paid_at
    if has_paid_at:
        if snaps_qs.exclude(paid_at__isnull=True).exists():
            return True

    # 1.3) Bloqueo por WeeklyPayment relacionado al snapshot
    if has_weekly_payment:
        if snaps_qs.filter(weekly_payment__status="paid").exists():
            return True

    # ==========================================================
    # 2) WeeklyPayment paid por técnico + semana efectiva
    # ==========================================================
    tech_ids = list(
        sesion.tecnicos_sesion.values_list("tecnico_id", flat=True)
    )

    if not tech_ids:
        return False

    weeks = list(
        snaps_qs.exclude(semana_resultado__isnull=True)
        .exclude(semana_resultado__exact="")
        .values_list("semana_resultado", flat=True)
        .distinct()
    )

    # Compatibilidad legacy: si no hay snapshots productivos
    if not weeks:
        possible_weeks = [
            (getattr(sesion, "semana_pago_real", "") or "").strip().upper(),
            (getattr(sesion, "semana_pago_proyectada", "") or "").strip().upper(),
            (getattr(sesion, "discount_week", "") or "").strip().upper(),
            
        ]
        weeks = [w for w in possible_weeks if w]

    if weeks:
        if WeeklyPayment.objects.filter(
            technician_id__in=tech_ids,
            week__in=weeks,
            status="paid",
        ).exists():
            return True

    # ==========================================================
    # 3) Compatibilidad legacy por marker en finance_note
    # ==========================================================
    note = getattr(sesion, "finance_note", "") or ""

    for tech_id in tech_ids:
        for wk in weeks:
            marker = f"[TECH_WEEKLY_PAYMENT_PAID:{tech_id}:{wk}]"
            if marker in note:
                return True

    return False


def _serialize_decimal_for_json(value) -> str:
    try:
        if isinstance(value, Decimal):
            return format(value.quantize(Decimal("0.01")), "f")
        return format(Decimal(str(value or 0)).quantize(Decimal("0.01")), "f")
    except Exception:
        return "0.00"

def _visible_tech_ids_for_user(user):
    """
    Devuelve:
      - None  => sin restricción, ve a todos los técnicos
      - set() => IDs de técnicos visibles para el usuario

    Reglas:
      - superuser / admin general: todos
      - facturación pura: todos
      - PM / supervisor: técnicos que comparten proyecto con él + él mismo
      - otros: solo él mismo
    """
    if not user or not getattr(user, "is_authenticated", False):
        return set()

    # Siempre puede verse a sí mismo
    ids = {user.id}

    # Admin real ve todo
    if getattr(user, "is_superuser", False) or getattr(user, "es_admin_general", False):
        return None

    tiene_rol = getattr(user, "tiene_rol", None)

    def has_role(role_name):
        if callable(tiene_rol):
            try:
                return bool(tiene_rol(role_name))
            except Exception:
                return False
        return bool(getattr(user, f"es_{role_name}", False))

    is_pm = has_role("pm") or getattr(user, "es_pm", False)
    is_supervisor = has_role("supervisor") or getattr(user, "es_supervisor", False)
    is_facturacion = has_role("facturacion") or getattr(user, "es_facturacion", False)

    # Facturación pura ve todo
    if is_facturacion and not (is_pm or is_supervisor):
        return None

    # Si no es PM ni supervisor, solo se ve a sí mismo
    if not (is_pm or is_supervisor):
        return ids

    # PM / supervisor: técnicos que comparten proyectos asignados
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        return ids

    my_project_ids = list(
        ProyectoAsignacion.objects.filter(usuario=user)
        .values_list("proyecto_id", flat=True)
        .distinct()
    )

    if not my_project_ids:
        return ids

    shared_user_ids = (
        ProyectoAsignacion.objects.filter(proyecto_id__in=my_project_ids)
        .values_list("usuario_id", flat=True)
        .distinct()
    )

    ids.update(shared_user_ids)
    return ids

def _deserialize_decimal_from_json(value) -> Decimal:
    try:
        return Decimal(str(value or "0")).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _build_paid_breakdown_snapshot_for_weekly_payment(wp) -> dict:
    """
    Congela el desglose del pago semanal al momento de marcarlo como PAID.

    Guarda:
      {
        "week": "2026-W19",
        "technician_id": 123,
        "technician_name": "Luis...",
        "lines": [
          {"project_label": "NB1234", "subtotal": "200.00", "source": "snapshot"},
          {"project_label": "Direct discount", "subtotal": "-10.00", "source": "legacy"},
          {"project_label": "Bonus", "subtotal": "50.00", "source": "adjustment"},
        ],
        "total": "240.00",
      }

    IMPORTANTE:
    - No depende de permisos ni visibilidad del usuario.
    - Debe servir aunque luego borren o modifiquen billings.
    """
    week = (getattr(wp, "week", "") or "").strip().upper()
    technician_id = getattr(wp, "technician_id", None)

    tech_name = ""
    try:
        tech = getattr(wp, "technician", None)
        if tech:
            tech_name = (tech.get_full_name() or tech.username or "").strip()
    except Exception:
        tech_name = ""

    if not week or not technician_id:
        return {
            "week": week,
            "technician_id": technician_id,
            "technician_name": tech_name,
            "lines": [],
            "total": "0.00",
        }

    lines = []
    total = Decimal("0.00")

    snapshot_model = BillingPayWeekSnapshot
    snapshot_fields = {f.name for f in snapshot_model._meta.get_fields()}

    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    # ==========================================================
    # 1) SNAPSHOTS productivos
    # ==========================================================
    snap_qs = snapshot_model.objects.filter(
        tecnico_id=technician_id,
        semana_resultado=week,
        sesion__isnull=False,
    )

    if has_is_adjustment:
        snap_qs = snap_qs.filter(is_adjustment=False)
    elif has_adjustment_of:
        snap_qs = snap_qs.filter(adjustment_of__isnull=True)

    snap_qs = snap_qs.filter(
        Q(sesion__estado__in={"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"})
        | Q(sesion__is_direct_discount=True)
        | Q(subtotal__lt=0)
    )

    det_prod = (
        snap_qs.values(
            "sesion__proyecto_id",
            "sesion__is_direct_discount",
        )
        .annotate(
            subtotal=Coalesce(
                Sum("subtotal"),
                Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2)),
            )
        )
        .order_by("sesion__proyecto_id")
    )

    for row in det_prod:
        subtotal = row["subtotal"] or Decimal("0.00")
        project_label = (
            "Direct discount"
            if row["sesion__is_direct_discount"]
            else (row["sesion__proyecto_id"] or "—")
        )

        lines.append(
            {
                "project_label": str(project_label),
                "subtotal": _serialize_decimal_for_json(subtotal),
                "source": "snapshot",
            }
        )
        total += subtotal

    # ==========================================================
    # 2) LEGACY sin snapshots productivos
    # ==========================================================
    legacy = _legacy_weekly_payment_details(
        tech_ids=[technician_id],
        weeks=[week],
        allowed_project_keys=None,
        can_view_all_projects=True,
    )

    for row in legacy.get((technician_id, week), []):
        subtotal = row.get("subtotal") or Decimal("0.00")
        lines.append(
            {
                "project_label": str(row.get("project_label") or "—"),
                "subtotal": _serialize_decimal_for_json(subtotal),
                "source": "legacy",
            }
        )
        total += subtotal

    # ==========================================================
    # 3) AJUSTES
    # ==========================================================
    try:
        from operaciones.models import AdjustmentEntry
    except Exception:
        AdjustmentEntry = None

    if AdjustmentEntry is not None:
        ADJ_LABEL = {
            "fixed_salary": "Fixed salary",
            "bonus": "Bonus",
            "advance": "Advance",
        }

        adj_rows = (
            AdjustmentEntry.objects.filter(
                technician_id=technician_id,
                week=week,
            )
            .values("adjustment_type")
            .annotate(
                total=Coalesce(
                    Sum("amount"),
                    Value(Decimal("0.00"), output_field=DecimalField(max_digits=18, decimal_places=2)),
                )
            )
            .order_by("adjustment_type")
        )

        for row in adj_rows:
            subtotal = row["total"] or Decimal("0.00")
            label = ADJ_LABEL.get(row["adjustment_type"], row["adjustment_type"] or "Adjustment")

            lines.append(
                {
                    "project_label": str(label),
                    "subtotal": _serialize_decimal_for_json(subtotal),
                    "source": "adjustment",
                }
            )
            total += subtotal

    return {
        "week": week,
        "technician_id": technician_id,
        "technician_name": tech_name,
        "lines": lines,
        "total": _serialize_decimal_for_json(total),
    }

def _snapshot_is_paid(snap) -> bool:
    """
    Determina si un snapshot ya quedó pagado.
    Compatible con varios estados/campos históricos.
    """
    try:
        if bool(getattr(snap, "is_paid", False)):
            return True
    except Exception:
        pass

    try:
        if getattr(snap, "paid_at", None):
            return True
    except Exception:
        pass

    try:
        if (getattr(snap, "payment_status", "") or "").strip().lower() == "paid":
            return True
    except Exception:
        pass

    try:
        wp = getattr(snap, "weekly_payment", None)
        if wp and (getattr(wp, "status", "") or "").strip().lower() == "paid":
            return True
    except Exception:
        pass

    return False


def _session_has_productive_snapshots(sesion) -> bool:
    snapshot_model = BillingPayWeekSnapshot
    snapshot_fields = {f.name for f in snapshot_model._meta.get_fields()}

    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    qs = snapshot_model.objects.filter(sesion=sesion)

    if has_is_adjustment:
        qs = qs.filter(is_adjustment=False)
    elif has_adjustment_of:
        qs = qs.filter(adjustment_of__isnull=True)

    return qs.exists()


def _item_is_paid_locked(item) -> bool:
    """
    Un item queda bloqueado si:
    - tiene snapshots productivos marcados como paid, o
    - es una sesión legacy sin snapshots productivos pero la sesión ya está pagada.
    """
    sesion = getattr(item, "sesion", None)
    if not sesion:
        return False

    snapshot_model = BillingPayWeekSnapshot
    snapshot_fields = {f.name for f in snapshot_model._meta.get_fields()}

    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    qs = snapshot_model.objects.filter(item=item, sesion=sesion)

    if has_is_adjustment:
        qs = qs.filter(is_adjustment=False)
    elif has_adjustment_of:
        qs = qs.filter(adjustment_of__isnull=True)

    snaps = list(qs.select_related("weekly_payment"))
    if snaps:
        return any(_snapshot_is_paid(snap) for snap in snaps)

    # fallback legacy: si no hay snapshots productivos en la sesión y la sesión está pagada,
    # se considera bloqueado el item completo.
    if not _session_has_productive_snapshots(sesion) and _session_is_paid_locked(
        sesion
    ):
        return True

    return False


def _normalize_qty_for_lock_compare(value) -> Decimal:
    try:
        return Decimal(str(value or "0")).quantize(Decimal("0.01"))
    except Exception:
        return Decimal("0.00")


def _validate_paid_locked_existing_items(sesion, filas) -> str | None:
    """
    Regla:
    - Los items ya pagados no pueden:
        * desaparecer
        * cambiar cantidad
        * duplicarse con el mismo código
    - Solo se pueden agregar códigos NUEVOS.
    """
    existing_items = list(
        sesion.items.all().only("id", "codigo_trabajo", "cantidad", "sesion_id")
    )

    locked_items = [it for it in existing_items if _item_is_paid_locked(it)]
    if not locked_items:
        return None

    locked_map = {}
    for it in locked_items:
        code = (getattr(it, "codigo_trabajo", "") or "").strip()
        qty = _normalize_qty_for_lock_compare(getattr(it, "cantidad", 0))
        locked_map.setdefault(code, []).append(qty)

    posted_map = {}
    for fila in filas:
        code = (fila.get("codigo") or "").strip()
        qty = _normalize_qty_for_lock_compare(fila.get("cantidad"))
        posted_map.setdefault(code, []).append(qty)

    for code in locked_map:
        locked_map[code] = sorted(locked_map[code])
    for code in posted_map:
        posted_map[code] = sorted(posted_map[code])

    for code, locked_qtys in locked_map.items():
        posted_qtys = posted_map.get(code)
        if not posted_qtys:
            return (
                f"The code '{code}' is already paid and cannot be removed. "
                f"Use Unpay first if you need to change it."
            )

        if posted_qtys != locked_qtys:
            return (
                f"The code '{code}' is already paid and its quantity cannot be changed. "
                f"Use Unpay first if you need to modify it."
            )

    return None


def _unmark_billings_paid_for_weekly_payment(wp):
    """

    Revierte SOLO las marcas de pago técnico (pay week).

    NO toca finance_status ni notas visibles de finanzas.

    """

    week = (getattr(wp, "week", "") or "").strip().upper()

    technician_id = getattr(wp, "technician_id", None)

    if not week or not technician_id:

        return

    snapshot_qs = BillingPayWeekSnapshot.objects.filter(
        tecnico_id=technician_id,
        semana_resultado=week,
    )

    snapshot_field_names = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}

    if "is_adjustment" in snapshot_field_names:

        snapshot_qs = snapshot_qs.filter(is_adjustment=False)

    elif "adjustment_of" in snapshot_field_names:

        snapshot_qs = snapshot_qs.filter(adjustment_of__isnull=True)

    # 1) flujo nuevo: snapshots

    for snap in snapshot_qs:

        changed = []

        if hasattr(snap, "is_paid") and getattr(snap, "is_paid", False):

            snap.is_paid = False

            changed.append("is_paid")

        if hasattr(snap, "paid_at") and getattr(snap, "paid_at", None):

            snap.paid_at = None

            changed.append("paid_at")

        if changed:

            snap.save(update_fields=changed)

    # 2) fallback legacy

    marker = f"[TECH_WEEKLY_PAYMENT_PAID:{technician_id}:{week}]"

    legacy_sessions = SesionBilling.objects.filter(
        tecnicos_sesion__tecnico_id=technician_id
    ).distinct()

    for s in legacy_sessions:

        note = (getattr(s, "finance_note", "") or "").strip()

        if not note or marker not in note:

            continue

        lines = [ln for ln in note.splitlines() if ln.strip() and ln.strip() != marker]

        s.finance_note = "\n".join(lines).strip()

        s.save(update_fields=["finance_note", "finance_updated_at"])


def _mark_billings_paid_for_weekly_payment(wp):
    """
    Marca como pagados los snapshots / legacy vinculados al WeeklyPayment,
    SOLO para producción válida.

    Regla:
    - Billing normal: solo si estado está aprobado por supervisor/PM/finanzas.
    - Direct discount: sí puede marcarse.
    - Nunca marca snapshots de billings en asignado, en proceso,
      finalizado, en revisión supervisor, rechazado, etc.
    """

    from django.utils import timezone

    ESTADOS_OK_SYNC = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}

    week = (getattr(wp, "week", "") or "").strip().upper()
    technician_id = getattr(wp, "technician_id", None)

    if not week or not technician_id:
        return

    now = timezone.now()

    snapshot_qs = BillingPayWeekSnapshot.objects.filter(
        tecnico_id=technician_id,
        semana_resultado=week,
        sesion__isnull=False,
    )

    snapshot_field_names = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}

    if "is_adjustment" in snapshot_field_names:
        snapshot_qs = snapshot_qs.filter(is_adjustment=False)
    elif "adjustment_of" in snapshot_field_names:
        snapshot_qs = snapshot_qs.filter(adjustment_of__isnull=True)

    # ✅ FIX CLAVE:
    # Solo marcar como paid snapshots de billings aprobados o direct discount.
    snapshot_qs = snapshot_qs.filter(
        Q(sesion__estado__in=ESTADOS_OK_SYNC)
        | Q(sesion__is_direct_discount=True)
    )

    # 1) flujo nuevo: snapshots
    for snap in snapshot_qs.select_related("sesion"):
        changed = []

        if hasattr(snap, "is_paid") and not getattr(snap, "is_paid", False):
            snap.is_paid = True
            changed.append("is_paid")

        if hasattr(snap, "paid_at") and not getattr(snap, "paid_at", None):
            snap.paid_at = now
            changed.append("paid_at")

        if hasattr(snap, "payment_status"):
            if (getattr(snap, "payment_status", "") or "").strip().lower() != "paid":
                snap.payment_status = "paid"
                changed.append("payment_status")

        if changed:
            snap.save(update_fields=changed)

    # 2) fallback legacy: sesiones sin snapshots productivos
    legacy_sessions = (
        SesionBilling.objects
        .filter(
            tecnicos_sesion__tecnico_id=technician_id,
        )
        .filter(
            Q(estado__in=ESTADOS_OK_SYNC)
            | Q(is_direct_discount=True)
        )
        .distinct()
    )

    for s in legacy_sessions:
        has_productive_snapshot = BillingPayWeekSnapshot.objects.filter(
            sesion=s,
            tecnico_id=technician_id,
            semana_resultado=week,
        ).exists()

        if has_productive_snapshot:
            continue

        session_week = (
            (getattr(s, "semana_pago_real", "") or "").strip().upper()
            or (getattr(s, "semana_pago_proyectada", "") or "").strip().upper()
        )

        if session_week != week:
            continue

        note = (getattr(s, "finance_note", "") or "").strip()
        marker = f"[TECH_WEEKLY_PAYMENT_PAID:{technician_id}:{week}]"

        if marker not in note:
            s.finance_note = f"{note}\n{marker}".strip() if note else marker
            s.save(update_fields=["finance_note", "finance_updated_at"])


def _legacy_weekly_payment_details(
    *,
    tech_ids,
    weeks=None,
    allowed_project_keys=None,
    can_view_all_projects=False,
):
    """
    Devuelve detalle legacy por (technician_id, week) para sesiones antiguas que:
    - están aprobadas o son direct discount
    - tienen semana_pago_real
    - NO tienen snapshots productivos

    Salida:
      {
        (tecnico_id, "2026-W19"): [
            {
                "project_label": "NB3231",
                "project_lookup_text": "Underground",
                "project_lookup_id": "NB3231",
                "subtotal": Decimal("200.00"),
            },
            ...
        ],
        ...
      }

    Reglas:
    - Si la sesión es direct discount -> label = "Direct discount"
    - Si no -> label = sesion.proyecto_id
    - project_lookup_text usa sesion.proyecto (el nombre real guardado en la sesión)
    - project_lookup_id usa sesion.proyecto_id
    - Si no hay desglose_tecnico, hace fallback por porcentaje de la sesión
    """
    from collections import defaultdict

    ESTADOS_OK_SYNC = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}

    snapshot_fields = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    tech_ids = {int(t) for t in (tech_ids or []) if str(t).isdigit()}
    if not tech_ids:
        return {}

    base_qs = (
        SesionBilling.objects
        .filter(Q(estado__in=ESTADOS_OK_SYNC) | Q(is_direct_discount=True))
        .exclude(semana_pago_real__isnull=True)
        .exclude(semana_pago_real__exact="")
        .prefetch_related(
            "items__desglose_tecnico",
            "tecnicos_sesion",
            "pay_week_snapshots",
        )
    )

    if weeks:
        weeks = {str(w).strip().upper() for w in weeks if str(w).strip()}
        if weeks:
            base_qs = base_qs.filter(semana_pago_real__in=weeks)

    if not can_view_all_projects:
        allowed_project_keys = {
            str(x).strip() for x in (allowed_project_keys or set()) if str(x).strip()
        }
        if not allowed_project_keys:
            return {}
        base_qs = base_qs.filter(proyecto__in=allowed_project_keys)

    details = defaultdict(list)

    for sesion in base_qs:
        snaps_qs = sesion.pay_week_snapshots.all()

        productive_exists = False
        for snap in snaps_qs:
            if has_is_adjustment and bool(getattr(snap, "is_adjustment", False)):
                continue
            if has_adjustment_of and getattr(snap, "adjustment_of_id", None):
                continue
            productive_exists = True
            break

        # Si ya tiene snapshots productivos, NO entra como legacy
        if productive_exists:
            continue

        week_real = (getattr(sesion, "semana_pago_real", "") or "").strip().upper()
        if not week_real:
            continue

        project_label = (
            "Direct discount"
            if getattr(sesion, "is_direct_discount", False)
            else (str(getattr(sesion, "proyecto_id", "") or "—").strip() or "—")
        )

        project_lookup_text = (str(getattr(sesion, "proyecto", "") or "").strip() or "")
        project_lookup_id = (str(getattr(sesion, "proyecto_id", "") or "").strip() or "")

        tech_totals = {}

        # 1) Intentar por desglose real
        try:
            for item in sesion.items.all():
                for bd in item.desglose_tecnico.all():
                    tid = getattr(bd, "tecnico_id", None)
                    if not tid or int(tid) not in tech_ids:
                        continue

                    subtotal = getattr(bd, "subtotal", None)
                    try:
                        subtotal = (
                            subtotal
                            if isinstance(subtotal, Decimal)
                            else Decimal(str(subtotal or 0))
                        )
                    except Exception:
                        subtotal = Decimal("0.00")

                    tech_totals[int(tid)] = tech_totals.get(int(tid), Decimal("0.00")) + subtotal
        except Exception:
            tech_totals = {}

        # 2) Fallback por porcentaje
        if not tech_totals:
            try:
                subtotal_tecnico = getattr(sesion, "subtotal_tecnico", None)
                subtotal_tecnico = (
                    subtotal_tecnico
                    if isinstance(subtotal_tecnico, Decimal)
                    else Decimal(str(subtotal_tecnico or 0))
                )
            except Exception:
                subtotal_tecnico = Decimal("0.00")

            try:
                asignaciones = list(sesion.tecnicos_sesion.all())
            except Exception:
                asignaciones = []

            for asig in asignaciones:
                tid = getattr(asig, "tecnico_id", None)
                if not tid or int(tid) not in tech_ids:
                    continue

                try:
                    pct = Decimal(str(getattr(asig, "porcentaje", 0) or 0))
                except Exception:
                    pct = Decimal("0.00")

                tech_totals[int(tid)] = (subtotal_tecnico * (pct / Decimal("100"))).quantize(
                    Decimal("0.01")
                )

        for tid, amount in tech_totals.items():
            if amount == 0:
                continue

            details[(tid, week_real)].append(
                {
                    "project_label": project_label,
                    "project_lookup_text": project_lookup_text,
                    "project_lookup_id": project_lookup_id,
                    "subtotal": amount,
                }
            )

    return dict(details)


def _legacy_session_totals_without_snapshots(
    week: str | None = None, tech_ids=None
) -> dict:
    """
    Retorna totales legacy por (technician_id, week) para sesiones antiguas que:
    - sí tienen semana_pago_real
    - están aprobadas o son direct discount
    - NO tienen snapshots productivos

    Optimización:
    - Si se pasa tech_ids, solo calcula esos técnicos.
    - Evita cargar legacy de todos los usuarios cuando solo se necesita uno.

    Salida:
      {
        (tecnico_id, "2026-W21"): Decimal("47.25"),
        ...
      }
    """
    from collections import defaultdict

    ESTADOS_OK_SYNC = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}

    snapshot_fields = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    tech_ids_set = None
    if tech_ids is not None:
        tech_ids_set = {int(t) for t in (tech_ids or []) if str(t).isdigit()}
        if not tech_ids_set:
            return {}

    base_qs = (
        SesionBilling.objects.filter(
            Q(estado__in=ESTADOS_OK_SYNC) | Q(is_direct_discount=True)
        )
        .exclude(semana_pago_real__isnull=True)
        .exclude(semana_pago_real__exact="")
    )

    if week:
        base_qs = base_qs.filter(semana_pago_real=week)

    if tech_ids_set is not None:
        base_qs = base_qs.filter(tecnicos_sesion__tecnico_id__in=tech_ids_set)

    base_qs = base_qs.prefetch_related(
        "items__desglose_tecnico",
        "tecnicos_sesion",
        "pay_week_snapshots",
    ).distinct()

    totals = defaultdict(lambda: Decimal("0.00"))

    for sesion in base_qs:
        snaps_qs = sesion.pay_week_snapshots.all()

        productive_exists = False
        for snap in snaps_qs:
            if has_is_adjustment and bool(getattr(snap, "is_adjustment", False)):
                continue
            if has_adjustment_of and getattr(snap, "adjustment_of_id", None):
                continue
            productive_exists = True
            break

        # Si ya tiene snapshots productivos, ya entra por el flujo nuevo
        if productive_exists:
            continue

        week_real = (getattr(sesion, "semana_pago_real", "") or "").strip().upper()
        if not week_real:
            continue

        tech_totals = {}

        # 1) Intentar usar desglose real por item/técnico
        try:
            for item in sesion.items.all():
                for bd in item.desglose_tecnico.all():
                    tid = getattr(bd, "tecnico_id", None)
                    if not tid:
                        continue

                    if tech_ids_set is not None and int(tid) not in tech_ids_set:
                        continue

                    subtotal = getattr(bd, "subtotal", None)
                    try:
                        subtotal = (
                            subtotal
                            if isinstance(subtotal, Decimal)
                            else Decimal(str(subtotal or 0))
                        )
                    except Exception:
                        subtotal = Decimal("0.00")

                    tech_totals[int(tid)] = (
                        tech_totals.get(int(tid), Decimal("0.00")) + subtotal
                    )
        except Exception:
            tech_totals = {}

        # 2) Fallback extremo: repartir subtotal_tecnico por porcentaje
        if not tech_totals:
            try:
                subtotal_tecnico = getattr(sesion, "subtotal_tecnico", None)
                subtotal_tecnico = (
                    subtotal_tecnico
                    if isinstance(subtotal_tecnico, Decimal)
                    else Decimal(str(subtotal_tecnico or 0))
                )
            except Exception:
                subtotal_tecnico = Decimal("0.00")

            try:
                asignaciones = list(sesion.tecnicos_sesion.all())
            except Exception:
                asignaciones = []

            for asig in asignaciones:
                tid = getattr(asig, "tecnico_id", None)
                if not tid:
                    continue

                if tech_ids_set is not None and int(tid) not in tech_ids_set:
                    continue

                try:
                    pct = Decimal(str(getattr(asig, "porcentaje", 0) or 0))
                except Exception:
                    pct = Decimal("0.00")

                tech_totals[int(tid)] = (
                    subtotal_tecnico * (pct / Decimal("100"))
                ).quantize(Decimal("0.01"))

        for tid, amount in tech_totals.items():
            if amount:
                totals[(int(tid), week_real)] += amount

    return dict(totals)


def build_payweek_groups(sesion):
    """
    Estructura para el template:
    [
      {
        "tech_name": "...",
        "weeks_summary": "2026-W18, 2026-W19",
        "lines": [
          {
            "snapshot_id": 1,
            "work_type": "Cable",
            "week": "2026-W19",
            "payment_status": "paid",
            "is_paid": True,
          },
          ...
        ]
      },
      ...
    ]

    REGLAS:
    - Nuevo modelo:
        agrupa por técnico + work_type a partir de snapshots
    - Compatibilidad legacy:
        si NO hay snapshots productivos pero la sesión sí tiene semana_pago_real,
        se arma una salida fallback por técnico usando la semana legacy

    IMPORTANTE:
    - En legacy se agrega "dom_id" único por línea para evitar IDs repetidos
      en el template/JS.
    - Cada línea trae:
        * payment_status
        * is_paid
    - Si la sesión ya tiene finance_status='paid', toda la sesión queda bloqueada
      y visible como Paid aunque no exista WeeklyPayment pagado todavía.
    """
    from collections import OrderedDict

    try:
        snaps = list(sesion.pay_week_snapshots.all())
    except Exception:
        snaps = []

    snapshot_fields = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    finance_paid = (
        getattr(sesion, "finance_status", "") or ""
    ).strip().lower() == "paid"

    def _tech_name_from_user(user_obj, fallback_id=None):
        try:
            if user_obj:
                return user_obj.get_full_name() or user_obj.username
        except Exception:
            pass
        return f"Tech {fallback_id or ''}".strip()

    def _week_sort_value(week_str: str):
        y, w = _parse_iso_week((week_str or "").strip())
        if y and w:
            return (y, w)
        return (0, 0)

    def _resolve_week_for_snap(snap):
        return (
            (getattr(snap, "semana_resultado", "") or "").strip().upper()
            or (getattr(snap, "semana_base", "") or "").strip().upper()
            or (getattr(sesion, "semana_pago_real", "") or "").strip().upper()
            or (getattr(sesion, "discount_week", "") or "").strip().upper()
            or (getattr(sesion, "semana_pago_proyectada", "") or "").strip().upper()
            or ""
        )

    def _legacy_line_is_paid(tecnico_id, week_value):
        """
        Legacy: marca paid por línea usando:
        1) finance_status='paid' en la sesión
        2) WeeklyPayment paid del técnico en esa semana
        """
        if finance_paid:
            return True

        week_value = (week_value or "").strip().upper()
        if not tecnico_id or not week_value:
            return False

        try:
            return WeeklyPayment.objects.filter(
                technician_id=tecnico_id,
                week=week_value,
                status="paid",
            ).exists()
        except Exception:
            return False

    # ==========================================================
    # 1) MODO NUEVO: snapshots productivos
    # ==========================================================
    productive_snaps = []
    for snap in snaps:
        if has_is_adjustment and bool(getattr(snap, "is_adjustment", False)):
            continue
        if has_adjustment_of and getattr(snap, "adjustment_of_id", None):
            continue
        productive_snaps.append(snap)

    if productive_snaps:
        grouped = OrderedDict()

        for snap in productive_snaps:
            tech_name = _tech_name_from_user(
                getattr(snap, "tecnico", None),
                getattr(snap, "tecnico_id", None),
            )

            tech_bucket = grouped.setdefault(
                tech_name,
                {
                    "tech_name": tech_name,
                    "weeks": set(),
                    "lines_map": OrderedDict(),
                },
            )

            work_type = (
                (getattr(snap, "tipo_trabajo", "") or "").strip()
                or (
                    getattr(getattr(snap, "item", None), "tipo_trabajo", "") or ""
                ).strip()
                or "Legacy"
            )

            week_value = _resolve_week_for_snap(snap)

            snap_payment_status = (
                (getattr(snap, "payment_status", "") or "").strip().lower()
            )
            snap_is_paid = (
                bool(getattr(snap, "is_paid", False)) or snap_payment_status == "paid"
            )

            if finance_paid:
                payment_status = "paid"
                is_paid = True
            else:
                payment_status = snap_payment_status or "pending"
                is_paid = snap_is_paid

            if week_value:
                tech_bucket["weeks"].add(week_value)

            line_key = work_type.lower()

            line_bucket = tech_bucket["lines_map"].setdefault(
                line_key,
                {
                    "snapshot_id": snap.id,
                    "work_type": work_type,
                    "weeks": set(),
                    "is_legacy": False,
                    "payment_status": payment_status,
                    "is_paid": is_paid,
                },
            )

            if week_value:
                line_bucket["weeks"].add(week_value)

            if not line_bucket.get("snapshot_id"):
                line_bucket["snapshot_id"] = snap.id

            if finance_paid:
                line_bucket["payment_status"] = "paid"
                line_bucket["is_paid"] = True
            else:
                if snap_is_paid:
                    line_bucket["payment_status"] = "paid"
                    line_bucket["is_paid"] = True
                elif not line_bucket.get("payment_status"):
                    line_bucket["payment_status"] = payment_status or "pending"

        out = []

        for _, grp in grouped.items():
            weeks_sorted = sorted(list(grp["weeks"]), key=_week_sort_value)
            weeks_summary = ", ".join(weeks_sorted) if weeks_sorted else "—"

            lines = []
            for _, line in grp["lines_map"].items():
                line_weeks_sorted = sorted(list(line["weeks"]), key=_week_sort_value)
                line_week = line_weeks_sorted[-1] if line_weeks_sorted else "—"

                lines.append(
                    {
                        "snapshot_id": line["snapshot_id"],
                        "work_type": line["work_type"],
                        "week": line_week,
                        "is_legacy": False,
                        "payment_status": (
                            "paid"
                            if line.get("is_paid")
                            else (line.get("payment_status") or "pending")
                        ),
                        "is_paid": bool(line.get("is_paid")),
                    }
                )

            lines.sort(key=lambda x: (x["work_type"] or "").lower())

            out.append(
                {
                    "tech_name": grp["tech_name"],
                    "weeks_summary": weeks_summary,
                    "lines": lines,
                }
            )

        out.sort(key=lambda g: (g["tech_name"] or "").lower())
        return out

    # ==========================================================
    # 2) MODO LEGACY: sin snapshots productivos
    # ==========================================================
    legacy_week = (
        (getattr(sesion, "semana_pago_real", "") or "").strip().upper()
        or (getattr(sesion, "discount_week", "") or "").strip().upper()
        or (getattr(sesion, "semana_pago_proyectada", "") or "").strip().upper()
    )

    if not legacy_week:
        return []

    try:
        asignaciones = list(
            sesion.tecnicos_sesion.select_related("tecnico").all().order_by("id")
        )
    except Exception:
        asignaciones = []

    if not asignaciones:
        return []

    out = []
    for idx, asig in enumerate(asignaciones, start=1):
        tech_name = _tech_name_from_user(
            getattr(asig, "tecnico", None),
            getattr(asig, "tecnico_id", None),
        )

        line_is_paid = _legacy_line_is_paid(
            getattr(asig, "tecnico_id", None),
            legacy_week,
        )

        out.append(
            {
                "tech_name": tech_name,
                "weeks_summary": legacy_week,
                "lines": [
                    {
                        "session_id": sesion.id,
                        "snapshot_id": 0,
                        "dom_id": f"{sesion.id}_{asig.tecnico_id}_{idx}",
                        "work_type": "Legacy",
                        "week": legacy_week,
                        "is_legacy": True,
                        "payment_status": "paid" if line_is_paid else "pending",
                        "is_paid": line_is_paid,
                    }
                ],
            }
        )

    out.sort(key=lambda g: (g["tech_name"] or "").lower())
    return out


def _billing_project_week_lock_message(sesion) -> str | None:
    """
    Si el billing ya está marcado como pagado, no se permite modificar
    la semana de pago proyectada.

    Devuelve:
      - str con el mensaje de bloqueo, si está bloqueado
      - None, si sí se puede editar
    """
    finance_status = (getattr(sesion, "finance_status", "") or "").strip().lower()
    if finance_status != "paid":
        return None

    paid_week = (
        (getattr(sesion, "semana_pago_real", "") or "").strip()
        or (getattr(sesion, "semana_pago_proyectada", "") or "").strip()
        or "the registered week"
    )

    return (
        f"This billing cannot be modified because it was already paid in week "
        f"'{paid_week}'. Please contact the administrator."
    )

def _parse_iso_week(value: str):
    """
    Recibe 'YYYY-W##' y devuelve (year, week).
    Si no es válido, retorna (None, None).
    """
    s = (value or "").strip().upper()
    m = re.match(r"^(\d{4})-W(\d{2})$", s)
    if not m:
        return None, None

    year = int(m.group(1))
    week = int(m.group(2))

    try:
        date.fromisocalendar(year, week, 1)
    except ValueError:
        return None, None

    return year, week

def _iso_week_monday(year: int, week: int):
    return date.fromisocalendar(year, week, 1)


def _format_iso_week(d: date):
    y, w, _ = d.isocalendar()
    return f"{y}-W{w:02d}"


def _add_weeks_to_iso_week(base_week: str, weeks_to_add: int) -> str:
    year, week = _parse_iso_week(base_week)
    if not year or not week:
        return ""
    monday = _iso_week_monday(year, week)
    result = monday + timedelta(weeks=int(weeks_to_add or 0))
    return _format_iso_week(result)


def rebuild_billing_payweek_snapshot(sesion):
    """
    Reconstruye el snapshot de semanas reales por técnico/item SIN borrar
    líneas ya pagadas ni líneas de ajuste.

    Regla:
      - semana_base = sesion.semana_pago_proyectada
      - payment_weeks = PrecioActividadTecnico.payment_weeks
      - semana_resultado = semana_base + payment_weeks

    Conserva:
      - snapshots pagados
      - snapshots de ajuste

    Actualiza:
      - snapshots normales no pagados
      - crea snapshots faltantes
      - elimina snapshots normales no pagados que ya no correspondan

    También congela:
      - tarifa_base
      - porcentaje
      - tarifa_efectiva
      - subtotal

    Retorna:
      {
        "created": <int>,
        "updated": <int>,
        "deleted": <int>,
        "summary_week": <str>,
      }
    """
    model = BillingPayWeekSnapshot
    field_names = {f.name for f in model._meta.get_fields()}

    has_payment_status = "payment_status" in field_names
    has_is_adjustment = "is_adjustment" in field_names
    has_adjustment_of = "adjustment_of" in field_names
    has_weekly_payment = "weekly_payment" in field_names
    has_paid_at = "paid_at" in field_names

    def _is_paid_snapshot(obj):
        if has_payment_status:
            return (getattr(obj, "payment_status", "") or "").strip().lower() == "paid"
        if has_paid_at and getattr(obj, "paid_at", None):
            return True
        if has_weekly_payment and getattr(obj, "weekly_payment_id", None):
            return True
        return False

    def _is_adjustment_snapshot(obj):
        if has_is_adjustment and bool(getattr(obj, "is_adjustment", False)):
            return True
        if has_adjustment_of and getattr(obj, "adjustment_of_id", None):
            return True
        return False

    def _parse_week_safe(value):
        try:
            y, wk = _parse_iso_week((value or "").strip())
            if y and wk:
                return (y, wk)
        except Exception:
            pass
        return None

    def _compute_summary_week_from_db():
        qs = model.objects.filter(sesion=sesion)

        if has_is_adjustment:
            qs = qs.filter(is_adjustment=False)
        elif has_adjustment_of:
            qs = qs.filter(adjustment_of__isnull=True)

        weeks = []
        for w in (
            qs.exclude(semana_resultado__isnull=True)
            .exclude(semana_resultado__exact="")
            .values_list("semana_resultado", flat=True)
        ):
            parsed = _parse_week_safe(w)
            if parsed:
                weeks.append((parsed[0], parsed[1], w))

        if not weeks:
            return ""

        weeks.sort(key=lambda x: (x[0], x[1]))
        return weeks[-1][2]

    semana_base = (sesion.semana_pago_proyectada or "").strip().upper()

    # Si no hay semana base, NO borrar snapshots pagados/ajustes.
    # Solo recalcular resumen desde lo que siga existiendo.
    if not semana_base:
        summary_week = _compute_summary_week_from_db()
        if (sesion.semana_pago_real or "") != summary_week:
            sesion.semana_pago_real = summary_week
            sesion.save(update_fields=["semana_pago_real"])
        return {"created": 0, "updated": 0, "deleted": 0, "summary_week": summary_week}

    proyecto_pk = _resolve_proyecto_pk_from_sesion(sesion)
    if not proyecto_pk:
        summary_week = _compute_summary_week_from_db()
        if (sesion.semana_pago_real or "") != summary_week:
            sesion.semana_pago_real = summary_week
            sesion.save(update_fields=["semana_pago_real"])
        return {"created": 0, "updated": 0, "deleted": 0, "summary_week": summary_week}

    created = 0
    updated = 0
    deleted = 0

    # ------------------------------------------------------------------
    # 1) Armar mapa deseado desde items actuales
    # ------------------------------------------------------------------
    desired_rows = {}

    items = sesion.items.prefetch_related("desglose_tecnico").all()

    for item in items:
        for bd in item.desglose_tecnico.all():
            tecnico_id = bd.tecnico_id

            precio = (
                PrecioActividadTecnico.objects.filter(
                    tecnico_id=tecnico_id,
                    ciudad=sesion.ciudad,
                    proyecto_id=proyecto_pk,
                    oficina=sesion.oficina,
                    cliente=sesion.cliente,
                    codigo_trabajo=item.codigo_trabajo,
                )
                .only("payment_weeks")
                .first()
            )

            if not precio:
                precio = (
                    PrecioActividadTecnico.objects.filter(
                        tecnico_id=tecnico_id,
                        ciudad=sesion.ciudad,
                        proyecto_id=proyecto_pk,
                        oficina=sesion.oficina,
                        cliente=sesion.cliente,
                        tipo_trabajo=item.tipo_trabajo,
                    )
                    .only("payment_weeks")
                    .first()
                )

            payment_weeks = int(getattr(precio, "payment_weeks", 0) or 0)
            semana_resultado = _add_weeks_to_iso_week(semana_base, payment_weeks)

            key = (item.id, tecnico_id)

            desired_rows[key] = {
                "sesion": sesion,
                "tecnico_id": tecnico_id,
                "item": item,
                "codigo_trabajo": (item.codigo_trabajo or "").strip(),
                "tipo_trabajo": (item.tipo_trabajo or "").strip(),
                "payment_weeks": payment_weeks,
                "semana_base": semana_base,
                "semana_resultado": semana_resultado,
                "tarifa_base": getattr(bd, "tarifa_base", Decimal("0.00"))
                or Decimal("0.00"),
                "porcentaje": getattr(bd, "porcentaje", Decimal("0.00"))
                or Decimal("0.00"),
                "tarifa_efectiva": getattr(bd, "tarifa_efectiva", Decimal("0.00"))
                or Decimal("0.00"),
                "subtotal": getattr(bd, "subtotal", Decimal("0.00")) or Decimal("0.00"),
            }

    # ------------------------------------------------------------------
    # 2) Cargar snapshots existentes y separar protegidos vs editables
    # ------------------------------------------------------------------
    existing = list(model.objects.filter(sesion=sesion).select_related("item"))

    reusable_by_key = {}
    protected_ids = set()

    for snap in existing:
        if _is_paid_snapshot(snap) or _is_adjustment_snapshot(snap):
            protected_ids.add(snap.id)
            continue

        key = (snap.item_id, snap.tecnico_id)
        reusable_by_key.setdefault(key, []).append(snap)

    touched_ids = set()

    # ------------------------------------------------------------------
    # 3) Update/Create de snapshots normales no pagados
    # ------------------------------------------------------------------
    for key, payload in desired_rows.items():
        bucket = reusable_by_key.get(key) or []

        snap = bucket.pop(0) if bucket else None

        if snap is None:
            create_kwargs = dict(payload)

            if has_payment_status:
                create_kwargs["payment_status"] = "pending"
            if has_is_adjustment:
                create_kwargs["is_adjustment"] = False
            if has_adjustment_of:
                create_kwargs["adjustment_of"] = None

            model.objects.create(**create_kwargs)
            created += 1
            continue

        changed = []

        for field in (
            "codigo_trabajo",
            "tipo_trabajo",
            "payment_weeks",
            "semana_base",
            "semana_resultado",
            "tarifa_base",
            "porcentaje",
            "tarifa_efectiva",
            "subtotal",
        ):
            new_value = payload[field]
            old_value = getattr(snap, field)
            if old_value != new_value:
                setattr(snap, field, new_value)
                changed.append(field)

        if snap.item_id != payload["item"].id:
            snap.item = payload["item"]
            changed.append("item")

        if snap.tecnico_id != payload["tecnico_id"]:
            snap.tecnico_id = payload["tecnico_id"]
            changed.append("tecnico")

        if (
            has_payment_status
            and (getattr(snap, "payment_status", "") or "").strip().lower() != "pending"
        ):
            snap.payment_status = "pending"
            changed.append("payment_status")

        if has_is_adjustment and bool(getattr(snap, "is_adjustment", False)):
            snap.is_adjustment = False
            changed.append("is_adjustment")

        if has_adjustment_of and getattr(snap, "adjustment_of_id", None) is not None:
            snap.adjustment_of = None
            changed.append("adjustment_of")

        if changed:
            snap.save(update_fields=changed)
            updated += 1

        touched_ids.add(snap.id)

    # ------------------------------------------------------------------
    # 4) Eliminar snapshots normales no pagados que ya no correspondan
    # ------------------------------------------------------------------
    stale_ids = []

    for snap in existing:
        if snap.id in protected_ids:
            continue

        key = (snap.item_id, snap.tecnico_id)
        if key not in desired_rows and snap.id not in touched_ids:
            stale_ids.append(snap.id)

    if stale_ids:
        deleted, _ = model.objects.filter(id__in=stale_ids).delete()

    # ------------------------------------------------------------------
    # 5) Recalcular semana resumen de la sesión
    # ------------------------------------------------------------------
    summary_week = _compute_summary_week_from_db()

    if (sesion.semana_pago_real or "") != summary_week:
        sesion.semana_pago_real = summary_week
        sesion.save(update_fields=["semana_pago_real"])

    return {
        "created": created,
        "updated": updated,
        "deleted": deleted,
        "summary_week": summary_week,
    }


def _iso_week_start(year: int, week: int):
    """
    Retorna el lunes de una ISO week.
    """
    return date.fromisocalendar(year, week, 1)


def _norm_text_key(value):
    return " ".join(str(value or "").strip().lower().split())


def _build_payment_weeks_consistency_key(
    *, tecnico_id, ciudad, proyecto_id, oficina, cliente, tipo_trabajo
):
    """
    La consistencia de payment_weeks se valida por:
    Technician + City + Project + Office + Client + Work Type
    """
    return (
        int(tecnico_id),
        _norm_text_key(ciudad),
        int(proyecto_id),
        _norm_text_key(oficina),
        _norm_text_key(cliente),
        _norm_text_key(tipo_trabajo),
    )


def _validate_preview_payment_weeks_consistency(preview_data, proyecto_id):
    """
    Valida consistencia dentro del preview importado.
    Si para la misma llave lógica existen distintos payment_weeks, marca error.
    """
    seen = {}

    for row in preview_data:
        tecnico_ids = row.get("tecnico") or []
        for tecnico_id in tecnico_ids:
            key = _build_payment_weeks_consistency_key(
                tecnico_id=tecnico_id,
                ciudad=row.get("ciudad"),
                proyecto_id=proyecto_id,
                oficina=row.get("oficina"),
                cliente=row.get("cliente"),
                tipo_trabajo=row.get("tipo_trabajo"),
            )
            val = row.get("payment_weeks")
            seen.setdefault(key, set()).add(0 if val in (None, "") else int(val))

    invalid_keys = {k: v for k, v in seen.items() if len(v) > 1}
    if not invalid_keys:
        return preview_data

    for row in preview_data:
        tecnico_ids = row.get("tecnico") or []
        row_invalid = False
        found_vals = None

        for tecnico_id in tecnico_ids:
            key = _build_payment_weeks_consistency_key(
                tecnico_id=tecnico_id,
                ciudad=row.get("ciudad"),
                proyecto_id=proyecto_id,
                oficina=row.get("oficina"),
                cliente=row.get("cliente"),
                tipo_trabajo=row.get("tipo_trabajo"),
            )
            if key in invalid_keys:
                row_invalid = True
                found_vals = sorted(list(invalid_keys[key]))
                break

        if row_invalid:
            msg = (
                "Inconsistent Payment Weeks for the same Technician / City / Project / "
                "Office / Client / Work Type. Found values: "
                + ", ".join(str(x) for x in found_vals)
            )
            row["error"] += (" | " if row["error"] else "") + msg

    return preview_data


def _collect_import_family_map(preview_data, proyecto_id):
    """
    Construye un mapa por familia lógica:
      key -> {
        "weeks": {1,2,...},
        "codes": {"BM2F","HO-1(<48)",...}
      }
    """
    family_map = {}

    for row in preview_data:
        tecnico_ids = row.get("tecnico") or []
        codigo = (row.get("codigo_trabajo") or "").strip()
        payment_weeks = row.get("payment_weeks")
        if payment_weeks in (None, ""):
            payment_weeks = 0
        payment_weeks = int(payment_weeks)

        for tecnico_id in tecnico_ids:
            key = _build_payment_weeks_consistency_key(
                tecnico_id=tecnico_id,
                ciudad=row.get("ciudad"),
                proyecto_id=proyecto_id,
                oficina=row.get("oficina"),
                cliente=row.get("cliente"),
                tipo_trabajo=row.get("tipo_trabajo"),
            )
            bucket = family_map.setdefault(key, {"weeks": set(), "codes": set()})
            bucket["weeks"].add(payment_weeks)
            if codigo:
                bucket["codes"].add(codigo)

    return family_map


def _validate_db_payment_weeks_consistency_for_import(
    *, tecnico, proyecto_id, family_map, replace=False
):
    """
    Valida por técnico TODO el lote importado contra BD.

    Regla:
    - Una familia lógica = Technician + City + Project + Office + Client + Work Type
    - Debe terminar con un único payment_weeks

    Si replace=True:
    - Se ignoran en BD los Job Codes que vienen en el mismo import para esa familia,
      porque esos registros serán reemplazados.
    """
    errors = []

    for key, payload in family_map.items():
        tecnico_id, ciudad_n, proyecto_id_key, oficina_n, cliente_n, tipo_n = key

        if int(tecnico.id) != int(tecnico_id):
            continue

        imported_weeks = set(int(x) for x in (payload.get("weeks") or set()))
        imported_codes = set((payload.get("codes") or set()))

        qs = PrecioActividadTecnico.objects.filter(
            tecnico=tecnico,
            proyecto_id=proyecto_id_key,
            tipo_trabajo__iexact=tipo_n,
        )

        db_rows = []
        for obj in qs.only(
            "id",
            "ciudad",
            "oficina",
            "cliente",
            "tipo_trabajo",
            "codigo_trabajo",
            "payment_weeks",
        ):
            if _norm_text_key(obj.ciudad) != ciudad_n:
                continue
            if _norm_text_key(obj.oficina) != oficina_n:
                continue
            if _norm_text_key(obj.cliente) != cliente_n:
                continue
            if _norm_text_key(obj.tipo_trabajo) != tipo_n:
                continue
            db_rows.append(obj)

        if replace and imported_codes:
            db_rows = [
                r
                for r in db_rows
                if (r.codigo_trabajo or "").strip() not in imported_codes
            ]

        db_weeks = set(int(getattr(r, "payment_weeks", 0) or 0) for r in db_rows)
        final_weeks = set(imported_weeks) | set(db_weeks)

        if len(final_weeks) > 1:
            sample_tipo = tipo_n or "-"
            sample_city = ciudad_n or "-"
            sample_office = oficina_n or "-"
            sample_client = cliente_n or "-"
            imported_codes_txt = (
                ", ".join(sorted(imported_codes)) if imported_codes else "-"
            )
            final_weeks_txt = ", ".join(str(x) for x in sorted(final_weeks))

            errors.append(
                "Inconsistent Payment Weeks for technician "
                f"'{tecnico.get_full_name() or tecnico.username}', "
                f"City '{sample_city}', Office '{sample_office}', Client '{sample_client}', "
                f"Work Type '{sample_tipo}'. "
                f"Imported Job Code(s): {imported_codes_txt}. "
                f"Final value(s) after import would be: {final_weeks_txt}."
            )

    return errors


def _build_receipt_key(user_id: int, filename: str) -> str:
    base = RECEIPTS_SAFE_PREFIX.rstrip("/")  # ej: operaciones/rendiciones
    ext = (filename.rsplit(".", 1)[-1] if "." in filename else "pdf").lower()
    today = timezone.now()
    # carpeta por usuario y fecha para que quede ordenado
    return f"{base}/{user_id}/{today:%Y/%m/%d}/rcpt_{uuid4().hex}.{ext}"

def _resolve_proyecto_pk_from_sesion(sesion, proyectos_qs=None):
    """
    Intenta resolver el PK de Proyecto (facturacion.Proyecto) desde SesionBilling.
    sesion.proyecto puede ser: id, nombre, código.
    sesion.proyecto_id suele ser el code/ID PROJECT.
    """
    # opcional: limitar a proyectos visibles si pasas proyectos_qs
    qs = proyectos_qs if proyectos_qs is not None else Proyecto.objects.all()

    raw_proj = (getattr(sesion, "proyecto", "") or "").strip()
    raw_pid  = (getattr(sesion, "proyecto_id", "") or "").strip()

    # 1) si sesion.proyecto es un id numérico
    if raw_proj:
        try:
            pid = int(raw_proj)
            p = qs.filter(pk=pid).first()
            if p:
                return p.id
        except (TypeError, ValueError):
            pass

        # 2) si sesion.proyecto es nombre o código
        p = qs.filter(Q(nombre__iexact=raw_proj) | Q(codigo__iexact=raw_proj)).first()
        if p:
            return p.id

    # 3) fallback: usar sesion.proyecto_id como código o id
    if raw_pid:
        p = qs.filter(Q(codigo__iexact=raw_pid) | Q(nombre__iexact=raw_pid)).first()
        if p:
            return p.id
        try:
            pid2 = int(raw_pid)
            p2 = qs.filter(pk=pid2).first()
            if p2:
                return p2.id
        except (TypeError, ValueError):
            pass

    return None


@login_required
@rol_requerido('usuario', 'facturacion', 'pm', 'admin')
@require_POST
def presign_rendicion(request, pk=None):
    """
    Pre-firma para subir DIRECTO el comprobante de rendición a Wasabi via POST.
    Request JSON: { filename, contentType, sizeBytes }
    Devuelve: {"post": {...}, "key": "<s3_key>"}  (url path-style)
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    filename = (data.get("filename") or "").strip()
    ctype = (data.get("contentType") or "").strip()
    size_b = int(data.get("sizeBytes") or 0)

    if not filename or ctype not in RECEIPT_ALLOWED_MIME:
        return HttpResponseBadRequest("Invalid file type.")
    if size_b <= 0 or size_b > RECEIPT_MAX_MB * 1024 * 1024:
        return HttpResponseBadRequest("File too large.")

    key = _build_receipt_key(request.user.id, filename)

    s3 = _s3_client()
    fields = {
        "acl": "private",
        "success_action_status": "201",
        # TIP: si quieres forzar Content-Type, puedes incluirlo aquí y en Conditions.
        # "Content-Type": ctype,
    }
    conditions = [
        {"acl": "private"},
        {"success_action_status": "201"},
        ["starts-with", "$key", key.rsplit("/", 1)[0] + "/"],
        ["content-length-range", 1, RECEIPT_MAX_MB * 1024 * 1024],
        # Si decides forzar Content-Type:
        # {"Content-Type": ctype},
    ]

    post = s3.generate_presigned_post(
        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
        Key=key,
        Fields=fields,
        Conditions=conditions,
        ExpiresIn=600,
    )

    # Forzar URL path-style (coincide con lo que ya usas)
    endpoint = settings.AWS_S3_ENDPOINT_URL.rstrip("/")
    bucket = settings.AWS_STORAGE_BUCKET_NAME
    post["url"] = f"{endpoint}/{bucket}"

    return JsonResponse({"post": post, "key": key})


def verificar_archivo_wasabi(ruta):
    """Verifica si un archivo existe en el bucket Wasabi."""
    s3 = _s3_client()  # ← usa el cliente único
    try:
        s3.head_object(Bucket=settings.AWS_STORAGE_BUCKET_NAME, Key=ruta)
        return True
    except ClientError:
        return False


# ==========================================================
# Helpers Odometer (miles) - declaración inmediata + aprobación posterior
# ==========================================================

# operaciones/views.py  (solo lo nuevo + mis_rendiciones completa + validar_odometro_vehicle_ajax)

# views.py (o donde la tengas)

import json

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponseBadRequest, JsonResponse
from django.utils import timezone
from django.views.decorators.http import require_POST

# ✅ fleet
try:
    from fleet.models import Vehicle, VehicleService
except Exception:  # pragma: no cover
    Vehicle = None
    VehicleService = None


def _normalizar_odometro(valor):
    """
    Convierte odómetro (miles) a int.
    Acepta: 1200, "1,200", "1.200", "1200 mi", "1200 miles"
    """
    if valor in (None, ""):
        return None
    try:
        s = str(valor).strip()
        s = s.replace(",", "").replace(" ", "")
        s = "".join(ch for ch in s if (ch.isdigit() or ch == "."))
        if not s:
            return None
        v = int(float(s))
        return v if v >= 0 else None
    except Exception:
        return None


from django.db.models import Q


def _validar_odometro_vecinos(vehicle_id, service_date, service_time, odo_nuevo):
    """
    Vecinos en la línea de tiempo:
    - vecino anterior (estrictamente antes): odo_nuevo >= odo_anterior
    - vecino posterior (estrictamente después): odo_nuevo <= odo_posterior
    """
    if VehicleService is None:
        return True, None

    if not vehicle_id or service_date is None or service_time is None or odo_nuevo is None:
        return True, None

    qs = (
        VehicleService.objects
        .filter(vehicle_id=vehicle_id)
        .exclude(kilometraje_declarado__isnull=True)
        .exclude(service_date__isnull=True)
        .exclude(service_time__isnull=True)
    )

    # ✅ vecino anterior: estrictamente antes (NO <=)
    anterior = (
        qs.filter(
            Q(service_date__lt=service_date) |
            Q(service_date=service_date, service_time__lt=service_time)
        )
        .order_by("-service_date", "-service_time", "-id")
        .first()
    )

    if anterior and anterior.kilometraje_declarado is not None:
        km_anterior = int(anterior.kilometraje_declarado)
        if odo_nuevo < km_anterior:
            return False, (
                f"The odometer ({odo_nuevo} miles) cannot be lower than the previous "
                f"record ({km_anterior} miles) on {anterior.service_date.strftime('%d/%m/%Y')} "
                f"at {anterior.service_time.strftime('%H:%M')}."
            )

    # ✅ vecino posterior: estrictamente después (NO >=)
    posterior = (
        qs.filter(
            Q(service_date__gt=service_date) |
            Q(service_date=service_date, service_time__gt=service_time)
        )
        .order_by("service_date", "service_time", "id")
        .first()
    )

    if posterior and posterior.kilometraje_declarado is not None:
        km_posterior = int(posterior.kilometraje_declarado)
        if odo_nuevo > km_posterior:
            return False, (
                f"The odometer ({odo_nuevo} miles) cannot be greater than a later "
                f"record ({km_posterior} miles) on {posterior.service_date.strftime('%d/%m/%Y')} "
                f"at {posterior.service_time.strftime('%H:%M')}."
            )

    return True, None

@login_required
@require_POST
def validar_odometro_ajax(request):
    """
    Validación en línea del odómetro (miles) para Services:
    - regla por vecinos (fecha real + hora) en VehicleService
    - devuelve current_odometer del vehículo para mostrarlo en UI

    Espera JSON:
      {
        "vehicle_id": "123",
        "kilometraje": "900",
        "real_consumption_date": "2026-04-06",
        "service_time": "14:03"
      }
    """
    if Vehicle is None:
        return JsonResponse({"ok": True, "current_odometer": None})

    try:
        data = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    vehicle_id_raw = (data.get("vehicle_id") or "").strip()
    km_raw = data.get("kilometraje")
    date_raw = (data.get("real_consumption_date") or "").strip()
    time_raw = (data.get("service_time") or "").strip()

    if not vehicle_id_raw:
        return JsonResponse({"ok": True, "current_odometer": None})

    try:
        vehicle_id = int(vehicle_id_raw)
    except Exception:
        return JsonResponse({"ok": False, "message": "Invalid vehicle."})

    v = Vehicle.objects.filter(pk=vehicle_id).only("id", "kilometraje_actual").first()
    if not v:
        return JsonResponse({"ok": False, "message": "Vehicle not found."})

    try:
        current_odo = int(getattr(v, "kilometraje_actual", 0) or 0)
    except Exception:
        current_odo = 0

    # Si aún no hay km ingresado, igual devolvemos el current_odo para el hint
    odo_nuevo = _normalizar_odometro(km_raw)
    if odo_nuevo is None or not date_raw or not time_raw:
        return JsonResponse({"ok": True, "current_odometer": current_odo})

    # Parse date/time
    try:
        service_date = timezone.datetime.strptime(date_raw, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse({"ok": False, "message": "Invalid date."})

    try:
        service_time = timezone.datetime.strptime(time_raw, "%H:%M").time()
    except Exception:
        return JsonResponse({"ok": False, "message": "Invalid time."})

    # no future (client-side hint)
    now_local = timezone.localtime(timezone.now())
    if service_date > now_local.date():
        return JsonResponse({"ok": False, "message": "You cannot register a Service with a future date.", "current_odometer": current_odo})
    if service_date == now_local.date() and service_time > now_local.time().replace(second=0, microsecond=0):
        return JsonResponse({"ok": False, "message": "You cannot register a Service with a future time.", "current_odometer": current_odo})

    # neighbors validation
    ok, msg = _validar_odometro_vecinos(vehicle_id, service_date, service_time, odo_nuevo)
    if not ok:
        return JsonResponse({"ok": False, "message": msg, "current_odometer": current_odo})

    return JsonResponse({"ok": True, "current_odometer": current_odo})

import time

from django.conf import settings
from django.contrib import messages
from django.core.exceptions import ValidationError
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Sum
from django.shortcuts import redirect, render
from django.utils import timezone

from facturacion.models import CartolaMovimiento

from .forms import MovimientoUsuarioForm

# ✅ fleet (solo para crear VehicleService)
try:
    from fleet.models import VehicleService
except Exception:  # pragma: no cover
    VehicleService = None


@login_required
@rol_requerido("usuario")
def mis_rendiciones(request):
    user = request.user

    # --- Query + paginación ---
    cantidad_str = request.GET.get("cantidad", "10")
    try:
        per_page = 1000000 if cantidad_str == "todos" else int(cantidad_str)
    except (TypeError, ValueError):
        per_page = 10
        cantidad_str = "10"

    movimientos_qs = (
        CartolaMovimiento.objects.filter(usuario=user)
        .select_related(
            "proyecto",
            "tipo",
            "service_type_obj",
            "vehicle",
            "aprobado_por_supervisor",
            "aprobado_por_pm",
            "aprobado_por_finanzas",
        )
        .order_by("-fecha")
    )

    paginator = Paginator(movimientos_qs, per_page)
    pagina = paginator.get_page(request.GET.get("page"))

    # --- Saldos ---
    saldo_disponible = (
        movimientos_qs.filter(
            tipo__categoria="abono", status="aprobado_abono_usuario"
        ).aggregate(total=Sum("abonos"))["total"]
        or 0
    ) - (
        movimientos_qs.exclude(tipo__categoria="abono")
        .filter(status="aprobado_finanzas")
        .aggregate(total=Sum("cargos"))["total"]
        or 0
    )
    saldo_pendiente = (
        movimientos_qs.filter(tipo__categoria="abono")
        .exclude(status="aprobado_abono_usuario")
        .aggregate(total=Sum("abonos"))["total"]
        or 0
    )
    saldo_rendido = (
        movimientos_qs.exclude(tipo__categoria="abono")
        .exclude(status="aprobado_finanzas")
        .aggregate(total=Sum("cargos"))["total"]
        or 0
    )

    # === claves presign (si vienen de un intento anterior fallido) ===
    wasabi_key_post = (request.POST.get("wasabi_key") or "").strip()
    wasabi_key_odo_post = (request.POST.get("wasabi_key_foto_tablero") or "").strip()

    # 🔒 proyectos permitidos para ESTE usuario
    allowed_ids = projects_ids_for_user(user)

    if request.method == "POST":
        form = MovimientoUsuarioForm(request.POST, request.FILES, user=request.user)

        # 🔒 Limitar proyectos del combo
        if hasattr(form, "fields") and "proyecto" in form.fields:
            form.fields["proyecto"].queryset = (
                form.fields["proyecto"]
                .queryset.filter(id__in=allowed_ids)
                .order_by("nombre")
            )

        if form.is_valid():
            cd = form.cleaned_data

            tipo = cd.get("tipo")
            tipo_nombre = (
                (getattr(tipo, "nombre", "") or str(tipo or "")).strip().lower()
            )
            es_service = tipo_nombre.startswith("service")

            try:
                with transaction.atomic():
                    mov: CartolaMovimiento = form.save(commit=False)
                    mov.usuario = user
                    mov.fecha = timezone.now()
                    mov.status = (
                        "pendiente_abono_usuario"
                        if (mov.tipo and mov.tipo.categoria == "abono")
                        else "pendiente_supervisor"
                    )

                    # 🔒 Validación servidor: proyecto permitido
                    proj = cd.get("proyecto")
                    if not proj or proj.id not in allowed_ids:
                        form.add_error(
                            "proyecto", "You are not assigned to that project."
                        )
                        raise ValidationError("Project not allowed.")

                    # ====== recibo (comprobante) ======
                    if wasabi_key_post:
                        mov.comprobante.name = wasabi_key_post
                    else:
                        mov.comprobante = cd.get("comprobante") or mov.comprobante

                    # ====== foto tablero (odómetro) ======
                    if wasabi_key_odo_post:
                        mov.foto_tablero.name = wasabi_key_odo_post
                    else:
                        mov.foto_tablero = cd.get("foto_tablero") or mov.foto_tablero

                    # ====== odómetro (miles) ======
                    mov.kilometraje = cd.get("kilometraje")

                    # ✅ NUEVO: guardar Service fields EN CARTOLA (para mostrar directo en tabla)
                    if es_service:
                        mov.vehicle = cd.get("vehicle")
                        mov.service_date = cd.get("real_consumption_date")
                        mov.service_time = cd.get("service_time")
                    else:
                        mov.vehicle = None
                        mov.service_date = None
                        mov.service_time = None

                    # Guardar rendición
                    mov.save()

                    # ✅ Si es Service: crear VehicleService (fleet) también
                    if es_service and VehicleService is not None:
                        v = cd.get("vehicle")
                        st = cd.get("service_type_obj")
                        service_time = cd.get("service_time")
                        real_date = cd.get("real_consumption_date")
                        odo_nuevo = cd.get("kilometraje")
                        amount = cd.get("cargos") or 0
                        notes = (cd.get("observaciones") or "").strip()

                        VehicleService.objects.create(
                            vehicle=v,
                            service_type_obj=st,
                            service_type="otro",
                            title=f"Expense Report #{mov.pk}",
                            service_date=real_date or timezone.localdate(),
                            service_time=service_time,
                            kilometraje_declarado=(
                                odo_nuevo if odo_nuevo is not None else None
                            ),
                            monto=amount,
                            notes=notes,
                        )

                        # ✅✅ IMPORTANTE:
                        # VehicleService.save() YA creó el VehicleOdometerEvent via vehicle.update_kilometraje(...)
                        # Aquí SOLO lo actualizamos para ponerle project (y marker), sin crear otro.
                        try:
                            if (
                                VehicleOdometerEvent is not None
                                and v
                                and odo_nuevo is not None
                            ):
                                odo_int = int(odo_nuevo)

                                # event_at consistente con service_date + time
                                if real_date and service_time:
                                    dt_naive = timezone.datetime.combine(
                                        real_date, service_time
                                    )
                                    event_at = timezone.make_aware(
                                        dt_naive, timezone.get_current_timezone()
                                    )
                                else:
                                    event_at = timezone.now()

                                marker = f"CartolaMovimiento#{mov.pk}"

                                ev = (
                                    VehicleOdometerEvent.objects.filter(
                                        vehicle=v, odometer=odo_int, source="service"
                                    )
                                    .order_by("-event_at", "-id")
                                    .first()
                                )

                                if ev:
                                    changed = []

                                    if ev.project_id is None and mov.proyecto_id:
                                        ev.project = mov.proyecto
                                        changed.append("project")

                                    if ev.event_at != event_at:
                                        ev.event_at = event_at
                                        changed.append("event_at")

                                    # agrega marker a notes solo si no está
                                    if marker and marker not in (ev.notes or ""):
                                        ev.notes = (ev.notes or "").strip()
                                        ev.notes = (
                                            f"{ev.notes} [{marker}]".strip()
                                            if ev.notes
                                            else f"[{marker}]"
                                        )
                                        changed.append("notes")

                                    if changed:
                                        ev.save(update_fields=changed)
                        except Exception:
                            pass

                # Verificación opcional en Wasabi cuando hubo subida directa
                if wasabi_key_post:
                    for _ in range(3):
                        if verificar_archivo_wasabi(mov.comprobante.name):
                            break
                        time.sleep(1)
                    else:
                        mov.delete()
                        messages.error(
                            request, "Error uploading the receipt. Please try again."
                        )
                        return redirect("operaciones:mis_rendiciones")

                if wasabi_key_odo_post:
                    for _ in range(3):
                        if verificar_archivo_wasabi(mov.foto_tablero.name):
                            break
                        time.sleep(1)
                    else:
                        mov.delete()
                        messages.error(
                            request,
                            "Error uploading the odometer photo. Please try again.",
                        )
                        return redirect("operaciones:mis_rendiciones")

                messages.success(request, "Expense report registered successfully.")
                return redirect("operaciones:mis_rendiciones")

            except Exception:
                ctx = {
                    "pagina": pagina,
                    "cantidad": cantidad_str,
                    "saldo_disponible": saldo_disponible,
                    "saldo_pendiente": saldo_pendiente,
                    "saldo_rendido": saldo_rendido,
                    "form": form,
                    "direct_uploads_receipts_enabled": True,
                    "receipt_max_mb": int(
                        getattr(settings, "RECEIPT_DIRECT_UPLOADS_MAX_MB", 25)
                    ),
                    "wasabi_key": wasabi_key_post,
                    "wasabi_key_foto_tablero": wasabi_key_odo_post,
                }
                return render(request, "operaciones/mis_rendiciones.html", ctx)

        # form inválido
        ctx = {
            "pagina": pagina,
            "cantidad": cantidad_str,
            "saldo_disponible": saldo_disponible,
            "saldo_pendiente": saldo_pendiente,
            "saldo_rendido": saldo_rendido,
            "form": form,
            "direct_uploads_receipts_enabled": True,
            "receipt_max_mb": int(
                getattr(settings, "RECEIPT_DIRECT_UPLOADS_MAX_MB", 25)
            ),
            "wasabi_key": wasabi_key_post,
            "wasabi_key_foto_tablero": wasabi_key_odo_post,
        }
        return render(request, "operaciones/mis_rendiciones.html", ctx)

    # GET
    form = MovimientoUsuarioForm(user=request.user)
    if hasattr(form, "fields") and "proyecto" in form.fields:
        form.fields["proyecto"].queryset = (
            form.fields["proyecto"]
            .queryset.filter(id__in=allowed_ids)
            .order_by("nombre")
        )

    return render(
        request,
        "operaciones/mis_rendiciones.html",
        {
            "pagina": pagina,
            "cantidad": cantidad_str,
            "saldo_disponible": saldo_disponible,
            "saldo_pendiente": saldo_pendiente,
            "saldo_rendido": saldo_rendido,
            "form": form,
            "direct_uploads_receipts_enabled": True,
            "receipt_max_mb": int(
                getattr(settings, "RECEIPT_DIRECT_UPLOADS_MAX_MB", 25)
            ),
            "wasabi_key": "",
            "wasabi_key_foto_tablero": "",
        },
    )


# Cerca de donde defines MULTIPART_EXPIRES_SECONDS
MULTIPART_EXPIRES_SECONDS = 900  # 15 min

RECEIPT_ALLOWED_MIME = set(getattr(
    settings,
    "RECEIPT_ALLOWED_MIME",
    {
        "application/pdf",
        "image/jpeg", "image/jpg",
        "image/png",
        "image/webp",
        "image/heic", "image/heif",
    }
))
# (Opcional) compatibilidad si en otro punto quedó el nombre viejo
ALLOWED_MIME = RECEIPT_ALLOWED_MIME


@login_required
@rol_requerido('usuario', 'facturacion', 'pm', 'admin')
@require_POST
def multipart_create(request):
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    filename = (data.get("filename") or "").strip()
    ctype = (data.get("contentType") or "").strip()

    # ✅ Validación correcta + indentación correcta
    if not filename or (ctype and ctype not in RECEIPT_ALLOWED_MIME):
        return HttpResponseBadRequest("Invalid file type.")

    key = _build_receipt_key(request.user.id, filename)
    s3 = _s3_client()
    try:
        resp = s3.create_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            ACL="private",
            ContentType=ctype or "application/octet-stream",
        )
    except ClientError as e:
        return HttpResponseBadRequest(str(e))

    return JsonResponse({
        "uploadId": resp["UploadId"],
        "key": key,
        "bucket": settings.AWS_STORAGE_BUCKET_NAME
    })
# --- 2) Firmar una parte ---


@login_required
@rol_requerido('usuario', 'facturacion', 'pm', 'admin')
@require_POST
def multipart_sign_part(request):
    """
    Body: { "key": "...", "uploadId": "...", "partNumber": 1 }
    Resp: { "url": "https://...presigned...", "partNumber": 1, "expiresIn": 900 }
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    key = (data.get("key") or "").strip()
    upload_id = (data.get("uploadId") or "").strip()
    part_number = int(data.get("partNumber") or 0)
    if not key or not upload_id or part_number <= 0:
        return HttpResponseBadRequest("Missing params.")

    s3 = _s3_client()
    try:
        url = s3.generate_presigned_url(
            ClientMethod="upload_part",
            Params={
                "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                "Key": key,
                "UploadId": upload_id,
                "PartNumber": part_number,
            },
            ExpiresIn=MULTIPART_EXPIRES_SECONDS,
        )
    except ClientError as e:
        return HttpResponseBadRequest(str(e))

    return JsonResponse({"url": url, "partNumber": part_number, "expiresIn": MULTIPART_EXPIRES_SECONDS})

# --- 3) Completar upload ---


@login_required
@rol_requerido('usuario', 'facturacion', 'pm', 'admin')
@require_POST
def multipart_complete(request):
    """
    Body: { "key": "...", "uploadId": "...", "parts": [{"ETag":"...", "PartNumber":1}, ...] }
    Resp: { "ok": true }
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    key = (data.get("key") or "").strip()
    upload_id = (data.get("uploadId") or "").strip()
    parts = data.get("parts") or []
    if not key or not upload_id or not parts:
        return HttpResponseBadRequest("Missing params.")

    s3 = _s3_client()
    try:
        s3.complete_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            MultipartUpload={"Parts": sorted(
                parts, key=lambda p: p["PartNumber"])},
            UploadId=upload_id,
        )
    except ClientError as e:
        return HttpResponseBadRequest(str(e))

    return JsonResponse({"ok": True})

# --- 4) Abortar upload (por si algo falla) ---


@login_required
@rol_requerido('usuario', 'facturacion', 'pm', 'admin')
@require_POST
def multipart_abort(request):
    """
    Body: { "key": "...", "uploadId": "..." }
    """
    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON")

    key = (data.get("key") or "").strip()
    upload_id = (data.get("uploadId") or "").strip()
    if not key or not upload_id:
        return HttpResponseBadRequest("Missing params.")

    s3 = _s3_client()
    try:
        s3.abort_multipart_upload(
            Bucket=settings.AWS_STORAGE_BUCKET_NAME,
            Key=key,
            UploadId=upload_id,
        )
    except ClientError:
        pass  # idempotente

    return JsonResponse({"ok": True})


@login_required
@rol_requerido('usuario')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def aprobar_abono(request, pk):
    mov = get_object_or_404(CartolaMovimiento, pk=pk, usuario=request.user)
    if mov.tipo.categoria == "abono" and mov.status == "pendiente_abono_usuario":
        mov.status = "aprobado_abono_usuario"
        mov.save()
        messages.success(request, "Deposit approved successfully.")
    return redirect('operaciones:mis_rendiciones')


@login_required
@rol_requerido('usuario')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def rechazar_abono(request, pk):
    mov = get_object_or_404(CartolaMovimiento, pk=pk, usuario=request.user)
    if request.method == "POST":
        motivo = request.POST.get("motivo", "")
        mov.status = "rechazado_abono_usuario"
        mov.motivo_rechazo = motivo
        mov.save()
        messages.error(request, "Deposit rejected and sent to Finance for review.")
    return redirect('operaciones:mis_rendiciones')


@login_required
@rol_requerido('usuario')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def editar_rendicion(request, pk):
    rendicion = get_object_or_404(CartolaMovimiento, pk=pk, usuario=request.user)

    if rendicion.status in ['aprobado_abono_usuario', 'aprobado_finanzas']:
        messages.error(request, "You cannot edit an already approved expense report.")
        return redirect('operaciones:mis_rendiciones')

    if request.method == 'POST':
        form = MovimientoUsuarioForm(request.POST, request.FILES, instance=rendicion)

        if form.is_valid():
            campos_editados = [f for f in form.changed_data if f not in ['status', 'actualizado']]
            if campos_editados and rendicion.status in [
                'rechazado_abono_usuario', 'rechazado_supervisor', 'rechazado_pm', 'rechazado_finanzas'
            ]:
                rendicion.status = 'pendiente_supervisor'

            form.save()
            messages.success(request, "Expense report successfully updated.")
            return redirect('operaciones:mis_rendiciones')
    else:
        form = MovimientoUsuarioForm(instance=rendicion)

    return render(request, 'operaciones/editar_rendicion.html', {'form': form})


@login_required
@rol_requerido('usuario')
@project_object_access_required(model='facturacion.CartolaMovimiento', object_kw='pk', project_attr='proyecto_id')
def eliminar_rendicion(request, pk):
    rendicion = get_object_or_404(CartolaMovimiento, pk=pk, usuario=request.user)

    if rendicion.status in ['aprobado_abono_usuario', 'aprobado_finanzas']:
        messages.error(request, "You cannot delete an already approved expense report.")
        return redirect('operaciones:mis_rendiciones')

    if request.method == 'POST':
        rendicion.delete()
        messages.success(request, "Expense report deleted successfully.")
        return redirect('operaciones:mis_rendiciones')

    return render(request, 'operaciones/eliminar_rendicion.html', {'rendicion': rendicion})


def _parse_fecha_fragmento(s: str):
    s = (s or "").strip()
    if not s:
        return {}

    parts = s.replace("/", "-").split("-")
    try:
        if len(parts) == 3 and all(p.isdigit() for p in parts):
            d, m, y = parts
            return {
                "fecha__date__day": int(d),
                "fecha__date__month": int(m),
                "fecha__date__year": int(y),
            }
        if len(parts) == 2 and all(p.isdigit() for p in parts):
            a, b = parts
            # dd-mm
            if len(a) <= 2 and len(b) <= 2:
                return {"fecha__date__day": int(a), "fecha__date__month": int(b)}
            # mm-yyyy
            if len(a) <= 2 and len(b) == 4:
                return {"fecha__date__month": int(a), "fecha__date__year": int(b)}
        if s.isdigit():
            val = int(s)
            if len(s) == 4:
                return {"fecha__date__year": val}
            return {"_day_or_month": val}
    except Exception:
        return {}
    return {}


@login_required
@rol_requerido('pm', 'admin', 'supervisor')
def vista_rendiciones(request):
    user = request.user

    # --------- Base visible según rol ---------
    if user.is_superuser:
        # Súper admin ve todo
        movimientos = CartolaMovimiento.objects.all()
    else:
        base = Q()
        # Rechazados (de cualquier etapa) visibles para supervisor y PM
        q_rechazados = Q(status__startswith='rechazado')

        # Supervisor: ve solo lo pendiente para él + todos los rechazados
        if getattr(user, 'es_supervisor', False):
            base |= Q(status='pendiente_supervisor') | q_rechazados

        # PM: ve lo pendiente para él (lo aprobado por supervisor) + todos los rechazados
        if getattr(user, 'es_pm', False):
            base |= Q(status='aprobado_supervisor') | q_rechazados

        movimientos = CartolaMovimiento.objects.filter(base) if base else CartolaMovimiento.objects.none()

    # 🔒 Limitar por proyectos asignados al usuario
    movimientos = filter_queryset_by_access(movimientos, request.user, 'proyecto_id')

    # ✅ Limitar también por fecha (ventana ProyectoAsignacion) - SOLO SE AGREGA ESTO
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    can_view_legacy_history = (
        request.user.is_superuser or
        getattr(request.user, "es_usuario_historial", False)
    )

    if ProyectoAsignacion is not None and (not can_view_legacy_history):
        try:
            # Proyectos que ya pasaron por filter_queryset_by_access (ojo: aquí ya está limitado)
            # Sacamos IDs desde el queryset actual para cruzar contra asignaciones
            proyecto_ids_visibles = list(
                movimientos.values_list("proyecto_id", flat=True).distinct()
            )
        except Exception:
            proyecto_ids_visibles = []

        try:
            asignaciones = list(
                ProyectoAsignacion.objects
                .filter(usuario=request.user, proyecto_id__in=proyecto_ids_visibles)
            )
        except Exception:
            asignaciones = []

        if asignaciones:
            access_by_pk = {}
            for a in asignaciones:
                if a.include_history or not a.start_at:
                    access_by_pk[a.proyecto_id] = {"include_history": True, "start_at": None}
                else:
                    access_by_pk[a.proyecto_id] = {"include_history": False, "start_at": a.start_at}

            # Filtramos por fecha usando el campo "fecha" del movimiento (tu campo principal)
            ids_ok = []
            for m in movimientos.only("id", "proyecto_id", "fecha"):
                pk = getattr(m, "proyecto_id", None)
                if pk is None:
                    continue
                access = access_by_pk.get(pk)
                if not access:
                    continue
                if access["include_history"] or access["start_at"] is None:
                    ids_ok.append(m.id)
                    continue
                # si el movimiento no tiene fecha, lo excluimos (mismo criterio defensivo)
                if not getattr(m, "fecha", None):
                    continue
                if m.fecha >= access["start_at"]:
                    ids_ok.append(m.id)

            movimientos = movimientos.filter(id__in=ids_ok)

    # ---------- Filtros ----------
    du = request.GET.get('du', '').strip()
    fecha_txt = request.GET.get('fecha', '').strip()
    real_fecha_txt = request.GET.get('real_fecha', '').strip()  # ✅ NUEVO
    proyecto = request.GET.get('proyecto', '').strip()
    tipo_txt = request.GET.get('tipo', '').strip()
    estado = request.GET.get('estado', '').strip()

    q = Q()
    if du:
        q &= (Q(usuario__first_name__icontains=du) |
              Q(usuario__last_name__icontains=du) |
              Q(usuario__username__icontains=du))

    if proyecto:
        q &= Q(proyecto__nombre__icontains=proyecto)

    if tipo_txt:
        q &= Q(tipo__nombre__icontains=tipo_txt)

    if estado:
        q &= Q(status=estado)

    # Fecha flexible (campo "fecha" = created_at / fecha del movimiento)
    if fecha_txt:
        fd = _parse_fecha_fragmento(fecha_txt)
        if fd:
            day_or_month = fd.pop("_day_or_month", None)
            if fd:
                q &= Q(**fd)
            if day_or_month is not None:
                q &= (Q(fecha__day=day_or_month) | Q(fecha__month=day_or_month))

    # ✅ NUEVO: Real consumption date flexible (DateField)
    if real_fecha_txt:
        fd = _parse_fecha_fragmento(real_fecha_txt)
        if fd:
            day_or_month = fd.pop("_day_or_month", None)

            # normaliza posibles claves antiguas y mapea fecha__* -> real_consumption_date__*
            new_fd = {}
            for k, v in fd.items():
                k2 = k.replace('fecha__date__', 'fecha__')
                if k2.startswith('fecha__'):
                    k2 = k2.replace('fecha__', 'real_consumption_date__', 1)
                new_fd[k2] = v

            if new_fd:
                q &= Q(**new_fd)

            if day_or_month is not None:
                q &= (Q(real_consumption_date__day=day_or_month) |
                      Q(real_consumption_date__month=day_or_month))

    if q:
        movimientos = movimientos.filter(q)

    # Orden personalizado
    movimientos = movimientos.annotate(
        orden_status=Case(
            When(status__startswith='pendiente', then=Value(1)),
            When(status__startswith='rechazado', then=Value(2)),
            When(status__startswith='aprobado',  then=Value(3)),
            default=Value(4),
            output_field=IntegerField(),
        )
    ).order_by('orden_status', '-fecha')

    # Totales (solo sobre lo que ve el usuario)
    total = movimientos.aggregate(total=Sum('cargos'))['total'] or 0
    pendientes = movimientos.filter(status__startswith='pendiente').aggregate(total=Sum('cargos'))['total'] or 0
    rechazados = movimientos.filter(status__startswith='rechazado').aggregate(total=Sum('cargos'))['total'] or 0

    # Paginación (sin "todos", máximo 100)
    raw_cantidad = request.GET.get("cantidad", "10")
    try:
        cantidad_pag = int(raw_cantidad)
    except (TypeError, ValueError):
        cantidad_pag = 10

    if cantidad_pag < 5:
        cantidad_pag = 5
    if cantidad_pag > 100:
        cantidad_pag = 100

    cantidad = str(cantidad_pag)

    paginator = Paginator(movimientos, cantidad_pag)
    page_number = request.GET.get("page")
    pagina = paginator.get_page(page_number)

    # Choices del modelo
    estado_choices = CartolaMovimiento._meta.get_field('status').choices

    base_qs = request.GET.copy()
    base_qs.pop('page', None)
    base_qs = base_qs.urlencode()

    return render(request, 'operaciones/vista_rendiciones.html', {
        'pagina': pagina,
        'cantidad': cantidad,
        'total': total,
        'pendientes': pendientes,
        'rechazados': rechazados,
        'filtros': {
            'du': du,
            'fecha': fecha_txt,
            'real_fecha': real_fecha_txt,  # ✅ NUEVO
            'proyecto': proyecto,
            'tipo': tipo_txt,
            'estado': estado
        },
        'estado_choices': estado_choices,
        'base_qs': base_qs,
    })


@login_required
@rol_requerido('pm', 'admin', 'supervisor', 'facturacion')
@project_object_access_required(model='facturacion.CartolaMovimiento',object_kw='pk',project_attr='proyecto_id')
def aprobar_rendicion(request, pk):
    mov = get_object_or_404(CartolaMovimiento, pk=pk)
    user = request.user

    if getattr(user, 'es_supervisor', False) and mov.status == 'pendiente_supervisor':
        mov.status = 'aprobado_supervisor'
        mov.aprobado_por_supervisor = user
    elif getattr(user, 'es_pm', False) and mov.status == 'aprobado_supervisor':
        mov.status = 'aprobado_pm'
        mov.aprobado_por_pm = user
    elif getattr(user, 'es_facturacion', False) and mov.status == 'aprobado_pm':
        mov.status = 'aprobado_finanzas'
        mov.aprobado_por_finanzas = user

    mov.motivo_rechazo = ''
    mov.save()
    messages.success(request, "Expense report approved successfully.")
    return redirect('operaciones:vista_rendiciones')


@login_required
@rol_requerido('pm', 'admin', 'supervisor', 'facturacion')
@project_object_access_required(model='facturacion.CartolaMovimiento',object_kw='pk',project_attr='proyecto_id')
def rechazar_rendicion(request, pk):
    movimiento = get_object_or_404(CartolaMovimiento, pk=pk)
    if request.method == 'POST':
        motivo = request.POST.get('motivo_rechazo')
        if motivo:
            movimiento.motivo_rechazo = motivo
            if request.user.es_supervisor and movimiento.status == 'pendiente_supervisor':
                movimiento.status = 'rechazado_supervisor'
                movimiento.aprobado_por_supervisor = request.user
            elif request.user.es_pm and movimiento.status == 'aprobado_supervisor':
                movimiento.status = 'rechazado_pm'
                movimiento.aprobado_por_pm = request.user
            elif request.user.es_facturacion and movimiento.status == 'aprobado_pm':
                movimiento.status = 'rechazado_finanzas'
                movimiento.aprobado_por_finanzas = request.user
            movimiento.save()
            messages.success(request, "Expense report rejected successfully.")
        else:
            messages.error(request, "Please enter the rejection reason.")
    return redirect('operaciones:vista_rendiciones')


@login_required
@rol_requerido('pm', 'admin')  # Si quieres, agrega 'supervisor', 'facturacion'
def exportar_rendiciones(request):
    from datetime import datetime

    import xlwt
    from django.db.models import Case, IntegerField, Q, Value, When
    from django.http import HttpResponse
    from django.utils.timezone import is_aware

    # ===== Base visible (misma lógica que vista_rendiciones) =====
    if request.user.is_superuser:
        base = CartolaMovimiento.objects.all()
    else:
        u = request.user
        visible_q = Q()

        # ✅ MISMA lógica que vista_rendiciones:
        # - Rechazados (cualquier etapa) visibles para supervisor y PM
        q_rechazados = Q(status__startswith='rechazado')

        if getattr(u, 'es_supervisor', False):
            visible_q |= Q(status='pendiente_supervisor') | q_rechazados

        if getattr(u, 'es_pm', False):
            visible_q |= Q(status='aprobado_supervisor') | q_rechazados

        # Si facturación está entrando a este export, mantenemos su lógica SIN romper la de la vista:
        if getattr(u, 'es_facturacion', False):
            visible_q |= Q(status='aprobado_pm') | q_rechazados

        base = CartolaMovimiento.objects.filter(visible_q) if visible_q else CartolaMovimiento.objects.none()

    # Limitar SIEMPRE a proyectos asignados al usuario
    base = filter_queryset_by_access(
        base.select_related('usuario', 'proyecto', 'tipo'),
        request.user,
        'proyecto_id'
    )

    # ✅ Limitar también por fecha (ventana ProyectoAsignacion) - SOLO SE AGREGA ESTO
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    can_view_legacy_history = (
        request.user.is_superuser or
        getattr(request.user, "es_usuario_historial", False)
    )

    if ProyectoAsignacion is not None and (not can_view_legacy_history):
        try:
            proyecto_ids_visibles = list(
                base.values_list("proyecto_id", flat=True).distinct()
            )
        except Exception:
            proyecto_ids_visibles = []

        try:
            asignaciones = list(
                ProyectoAsignacion.objects
                .filter(usuario=request.user, proyecto_id__in=proyecto_ids_visibles)
            )
        except Exception:
            asignaciones = []

        if asignaciones:
            access_by_pk = {}
            for a in asignaciones:
                if a.include_history or not a.start_at:
                    access_by_pk[a.proyecto_id] = {"include_history": True, "start_at": None}
                else:
                    access_by_pk[a.proyecto_id] = {"include_history": False, "start_at": a.start_at}

            ids_ok = []

            # ✅ ÚNICO CAMBIO: antes era base.only("id","proyecto_id","fecha")
            for mid, pid, fecha in base.values_list("id", "proyecto_id", "fecha"):
                if pid is None:
                    continue
                access = access_by_pk.get(pid)
                if not access:
                    continue
                if access["include_history"] or access["start_at"] is None:
                    ids_ok.append(mid)
                    continue
                if not fecha:
                    continue
                if fecha >= access["start_at"]:
                    ids_ok.append(mid)

            base = base.filter(id__in=ids_ok)

    # --------- Filtros ----------
    du        = (request.GET.get('du') or '').strip()
    fecha_txt = (request.GET.get('fecha') or '').strip()
    real_fecha_txt = (request.GET.get('real_fecha') or '').strip()  # ✅ NUEVO
    proyecto  = (request.GET.get('proyecto') or '').strip()
    tipo_txt  = (request.GET.get('tipo') or '').strip()
    estado    = (request.GET.get('estado') or '').strip()

    q = Q()
    if du:
        q &= (Q(usuario__first_name__icontains=du) |
              Q(usuario__last_name__icontains=du) |
              Q(usuario__username__icontains=du))
    if proyecto:
        q &= Q(proyecto__nombre__icontains=proyecto)
    if tipo_txt:
        q &= Q(tipo__nombre__icontains=tipo_txt)
    if estado:
        q &= Q(status=estado)

    # Fecha flexible (fecha del movimiento)
    if fecha_txt:
        fd = _parse_fecha_fragmento(fecha_txt)
        if fd:
            day_or_month = fd.pop("_day_or_month", None)
            if any(k.startswith('fecha__date__') for k in fd.keys()):
                fd = {k.replace('fecha__date__', 'fecha__'): v for k, v in fd.items()}
            if fd:
                q &= Q(**fd)
            if day_or_month is not None:
                q &= (Q(fecha__day=day_or_month) | Q(fecha__month=day_or_month))

    # ✅ NUEVO: Real consumption date flexible
    if real_fecha_txt:
        fd = _parse_fecha_fragmento(real_fecha_txt)
        if fd:
            day_or_month = fd.pop("_day_or_month", None)

            new_fd = {}
            for k, v in fd.items():
                k2 = k.replace('fecha__date__', 'fecha__')
                if k2.startswith('fecha__'):
                    k2 = k2.replace('fecha__', 'real_consumption_date__', 1)
                new_fd[k2] = v

            if new_fd:
                q &= Q(**new_fd)

            if day_or_month is not None:
                q &= (Q(real_consumption_date__day=day_or_month) |
                      Q(real_consumption_date__month=day_or_month))

    movimientos = base.filter(q) if q else base

    # ===== Orden FINAL =====
    movimientos = movimientos.annotate(
        orden_status=Case(
            When(status__startswith='pendiente', then=Value(1)),
            When(status__startswith='rechazado', then=Value(2)),
            When(status__startswith='aprobado',  then=Value(3)),
            default=Value(4),
            output_field=IntegerField(),
        )
    ).order_by('orden_status', '-fecha', '-id')

    # ----- Excel -----
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="expense_reports.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('Expense Reports')

    header_style = xlwt.easyxf('font: bold on; align: horiz center')
    date_style   = xlwt.easyxf(num_format_str='DD-MM-YYYY')

    columns = ["User", "Date", "Real consumption date", "Project", "Type", "Remarks", "Amount", "Status", "Odometer (km)"]
    for col_num, title in enumerate(columns):
        ws.write(0, col_num, title, header_style)

    for row_num, mov in enumerate(movimientos, start=1):
        ws.write(row_num, 0, str(mov.usuario))

        fecha_excel = mov.fecha
        if isinstance(fecha_excel, datetime):
            if is_aware(fecha_excel):
                fecha_excel = fecha_excel.astimezone().replace(tzinfo=None)
            fecha_excel = fecha_excel.date()
        ws.write(row_num, 1, fecha_excel, date_style)

        if mov.real_consumption_date:
            ws.write(row_num, 2, mov.real_consumption_date, date_style)
        else:
            ws.write(row_num, 2, "")

        ws.write(row_num, 3, str(getattr(mov.proyecto, "nombre", mov.proyecto or "")))
        ws.write(row_num, 4, str(getattr(mov.tipo, "nombre", mov.tipo or "")))
        ws.write(row_num, 5, mov.observaciones or "")
        ws.write(row_num, 6, float(mov.cargos or 0))
        ws.write(row_num, 7, mov.get_status_display())
        ws.write(row_num, 8, int(mov.kilometraje) if mov.kilometraje is not None else "")

    wb.save(response)
    return response

@login_required
@rol_requerido('usuario')
def exportar_mis_rendiciones(request):
    from datetime import datetime

    import xlwt
    from django.http import HttpResponse
    from django.utils.timezone import is_aware

    user = request.user

    # Base: solo mis movimientos
    base = (
        CartolaMovimiento.objects
        .filter(usuario=user)
        .select_related('usuario', 'proyecto', 'tipo')
        .order_by('-fecha')
    )
    # Limitar a proyectos donde el usuario tiene acceso
    movimientos = filter_queryset_by_access(base, user, 'proyecto_id')

    # Crear archivo Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="my_expense_reports.xls"'

    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('My Expense Reports')

    header_style = xlwt.easyxf('font: bold on; align: horiz center')
    date_style = xlwt.easyxf(num_format_str='DD-MM-YYYY')

    # ✅ NUEVO: agregamos "Real consumption date"
    columns = [
        "User",
        "Date",
        "Real consumption date",
        "Project",
        "Type",
        "Expenses (USD)",
        "Credits (USD)",
        "Remarks",
        "Status",
        "Odometer (km)",
    ]
    for col_num, column_title in enumerate(columns):
        ws.write(0, col_num, column_title, header_style)

    # Datos
    for row_num, mov in enumerate(movimientos, start=1):
        # Fecha: naive y solo date
        fecha_excel = mov.fecha
        if isinstance(fecha_excel, datetime):
            if is_aware(fecha_excel):
                fecha_excel = fecha_excel.astimezone().replace(tzinfo=None)
            fecha_excel = fecha_excel.date()

        ws.write(row_num, 0, mov.usuario.get_full_name())
        ws.write(row_num, 1, fecha_excel, date_style)

        # ✅ NUEVO: Real consumption date
        if mov.real_consumption_date:
            ws.write(row_num, 2, mov.real_consumption_date, date_style)
        else:
            ws.write(row_num, 2, "")

        ws.write(row_num, 3, str(mov.proyecto or ""))
        ws.write(row_num, 4, str(mov.tipo or ""))
        ws.write(row_num, 5, float(mov.cargos or 0))
        ws.write(row_num, 6, float(mov.abonos or 0))
        ws.write(row_num, 7, mov.observaciones or "")
        ws.write(row_num, 8, mov.get_status_display())
        ws.write(row_num, 9, int(mov.kilometraje) if mov.kilometraje is not None else "")

    wb.save(response)
    return response


@login_required(login_url="usuarios:login")
@rol_requerido("admin", "pm", "facturacion")
def listar_precios_tecnico(request):
    # ---- Cantidad por página (solo estos valores) ----
    cantidad_str = request.GET.get("cantidad", "10")
    allowed_page_sizes = {"5", "10", "20", "50", "100"}

    if cantidad_str not in allowed_page_sizes:
        cantidad_str = "10"

    cantidad = int(cantidad_str)

    # ---- Filtros (GET) ----
    f_tecnico = (request.GET.get("f_tecnico") or "").strip()
    f_ciudad = (request.GET.get("f_ciudad") or "").strip()
    f_proy = (request.GET.get("f_proyecto") or "").strip()
    f_codigo = (request.GET.get("f_codigo") or "").strip()

    qs = PrecioActividadTecnico.objects.select_related("tecnico", "proyecto").order_by(
        "-fecha_creacion"
    )

    # 🔒 Limitar por proyectos asignados al usuario actual SOLO si 'proyecto' es FK
    is_fk_proyecto = False
    try:
        f = PrecioActividadTecnico._meta.get_field("proyecto")
        if isinstance(f, dj_models.ForeignKey):
            is_fk_proyecto = True
            qs = filter_queryset_by_access(qs, request.user, "proyecto_id")
    except Exception:
        pass

    # ✅ Limitar también por fecha (ventana ProyectoAsignacion)
    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    can_view_legacy_history = request.user.is_superuser or getattr(
        request.user, "es_usuario_historial", False
    )

    if ProyectoAsignacion is not None and (not can_view_legacy_history):
        # --- Caso A: proyecto es FK ---
        if is_fk_proyecto:
            try:
                proyecto_ids_visibles = list(
                    qs.values_list("proyecto_id", flat=True).distinct()
                )
            except Exception:
                proyecto_ids_visibles = []

            try:
                asignaciones = list(
                    ProyectoAsignacion.objects.filter(
                        usuario=request.user, proyecto_id__in=proyecto_ids_visibles
                    ).select_related("proyecto")
                )
            except Exception:
                asignaciones = []

            if asignaciones:
                access_by_pk = {}
                for a in asignaciones:
                    if a.include_history or not a.start_at:
                        access_by_pk[a.proyecto_id] = {
                            "include_history": True,
                            "start_at": None,
                        }
                    else:
                        access_by_pk[a.proyecto_id] = {
                            "include_history": False,
                            "start_at": a.start_at,
                        }

                ids_ok = []
                for pid, rid, fcrea in qs.values_list(
                    "proyecto_id", "id", "fecha_creacion"
                ):
                    if pid is None:
                        continue
                    access = access_by_pk.get(pid)
                    if not access:
                        continue
                    if access["include_history"] or access["start_at"] is None:
                        ids_ok.append(rid)
                        continue
                    if not fcrea:
                        continue

                    start_at = access["start_at"]
                    fcrea_cmp = fcrea.date() if hasattr(fcrea, "date") else fcrea
                    start_cmp = (
                        start_at.date() if hasattr(start_at, "date") else start_at
                    )

                    if fcrea_cmp >= start_cmp:
                        ids_ok.append(rid)

                qs = qs.filter(id__in=ids_ok)

        # --- Caso B: proyecto es TEXTO (legacy) ---
        else:
            try:
                from facturacion.models import Proyecto
            except Exception:
                Proyecto = None

            if Proyecto is not None:
                try:
                    proyectos_user = filter_queryset_by_access(
                        Proyecto.objects.all(), request.user, "id"
                    )
                except Exception:
                    proyectos_user = Proyecto.objects.none()

                if proyectos_user.exists():
                    allowed_keys = set()
                    for p in proyectos_user:
                        nombre = (getattr(p, "nombre", "") or "").strip()
                        if nombre:
                            allowed_keys.add(nombre)

                        codigo = getattr(p, "codigo", None)
                        if codigo:
                            allowed_keys.add(str(codigo).strip())

                        allowed_keys.add(str(p.id).strip())
                else:
                    allowed_keys = set()

                if allowed_keys:
                    qs = qs.filter(proyecto__in=allowed_keys)
                else:
                    qs = PrecioActividadTecnico.objects.none()

                try:
                    asignaciones = list(
                        ProyectoAsignacion.objects.filter(
                            usuario=request.user, proyecto__in=list(proyectos_user)
                        ).select_related("proyecto")
                    )
                except Exception:
                    asignaciones = []

                if asignaciones:
                    access_by_key = {}
                    for a in asignaciones:
                        p = getattr(a, "proyecto", None)
                        if not p:
                            continue

                        if a.include_history or not a.start_at:
                            access = {"include_history": True, "start_at": None}
                        else:
                            access = {"include_history": False, "start_at": a.start_at}

                        for k in (
                            getattr(p, "nombre", None),
                            getattr(p, "codigo", None),
                            getattr(p, "id", None),
                        ):
                            if k is None:
                                continue
                            ks = str(k).strip()
                            if ks:
                                access_by_key[ks.lower()] = access

                    ids_ok = []
                    for rid, proj_txt, fcrea in qs.values_list(
                        "id", "proyecto", "fecha_creacion"
                    ):
                        key = str(proj_txt).strip().lower() if proj_txt else ""
                        if not key:
                            continue
                        access = access_by_key.get(key)
                        if not access:
                            continue
                        if access["include_history"] or access["start_at"] is None:
                            ids_ok.append(rid)
                            continue
                        if not fcrea:
                            continue

                        start_at = access["start_at"]
                        fcrea_cmp = fcrea.date() if hasattr(fcrea, "date") else fcrea
                        start_cmp = (
                            start_at.date() if hasattr(start_at, "date") else start_at
                        )

                        if fcrea_cmp >= start_cmp:
                            ids_ok.append(rid)

                    qs = qs.filter(id__in=ids_ok)

    if f_tecnico:
        qs = qs.filter(
            Q(tecnico__first_name__icontains=f_tecnico)
            | Q(tecnico__last_name__icontains=f_tecnico)
            | Q(tecnico__username__icontains=f_tecnico)
        )
    if f_ciudad:
        qs = qs.filter(ciudad__icontains=f_ciudad)
    if f_proy:
        try:
            qs = qs.filter(
                Q(proyecto__nombre__icontains=f_proy)
                | Q(proyecto__codigo__icontains=f_proy)
            )
        except FieldError:
            qs = qs.filter(proyecto__icontains=f_proy)
    if f_codigo:
        qs = qs.filter(codigo_trabajo__icontains=f_codigo)

    paginator = Paginator(qs, cantidad)
    page_number = request.GET.get("page")
    pagina = paginator.get_page(page_number)

    ctx = {
        "pagina": pagina,
        "cantidad": cantidad_str,
        "f_tecnico": f_tecnico,
        "f_ciudad": f_ciudad,
        "f_proyecto": f_proy,
        "f_codigo": f_codigo,
    }
    return render(request, "operaciones/listar_precios_tecnico.html", ctx)


try:
    from usuarios.models import \
        ProyectoAsignacion  # usuario, proyecto, include_history, start_at
except Exception:
    ProyectoAsignacion = None

def _to2(val):
    try:
        return float(Decimal(str(val)).quantize(Decimal("0.01")))
    except (InvalidOperation, ValueError, TypeError):
        return None


def _tecnicos_de_proyecto_qs(proyecto: Optional[Proyecto]):
    """
    Devuelve usuarios asignados al proyecto (sin filtrar por rol):
    1) via ProyectoAsignacion (preferido)
    2) via M2M User.proyectos
    3) via sesiones (SesionBillingTecnico -> sesion.proyecto_id) comparando contra
       [proyecto.codigo, str(proyecto.id), proyecto.nombre]
    """
    User = get_user_model()
    if not proyecto:
        return User.objects.none()

    # 1) Through table
    try:
        user_ids = proyecto.asignaciones.values_list("usuario_id", flat=True)
        qs_pa = User.objects.filter(id__in=user_ids).order_by("first_name", "last_name", "username")
        if qs_pa.exists():
            return qs_pa
    except Exception:
        pass

    # 2) M2M directo
    try:
        qs_m2m = User.objects.filter(proyectos=proyecto).order_by("first_name", "last_name", "username")
        if qs_m2m.exists():
            return qs_m2m
    except Exception:
        pass

    # 3) Fallback por sesiones (usa posibles llaves)
    keys = []
    for k in (getattr(proyecto, "codigo", None), getattr(proyecto, "id", None), getattr(proyecto, "nombre", None)):
        if k is not None and str(k).strip():
            keys.append(str(k).strip())

    if not keys:
        return User.objects.none()

    tech_ids = (
        SesionBillingTecnico.objects
        .filter(sesion__proyecto_id__in=keys)
        .values_list("tecnico_id", flat=True)
        .distinct()
    )
    return User.objects.filter(id__in=tech_ids).order_by("first_name", "last_name", "username")


@login_required(login_url="usuarios:login")
@rol_requerido("admin", "pm")
def importar_precios(request):
    """
    GET  -> muestra form; si viene ?proyecto_id, filtra técnicos.
    POST -> valida, arma preview y muestra conflictos.
    Regla nueva:
    Para un mismo Technician + City + Project + Office + Client + Work Type
    debe existir un único payment_weeks.
    """
    proyectos_qs = filter_queryset_by_access(Proyecto.objects.all(), request.user, "id")

    # ---------------- GET ----------------
    if request.method == "GET":
        form = ImportarPreciosForm()
        proyecto_id_get = (request.GET.get("proyecto_id") or "").strip()

        proyecto_sel = None
        if proyecto_id_get and proyectos_qs.filter(pk=proyecto_id_get).exists():
            proyecto_sel = proyectos_qs.get(pk=proyecto_id_get)

        form.fields["tecnicos"].queryset = _tecnicos_de_proyecto_qs(proyecto_sel)

        return render(
            request,
            "operaciones/importar_precios.html",
            {
                "form": form,
                "proyectos": proyectos_qs,
                "proyecto_sel": proyecto_sel,
            },
        )

    # ---------------- POST ----------------
    form = ImportarPreciosForm(request.POST, request.FILES)

    proyecto_id = (request.POST.get("proyecto_id") or "").strip()
    if not proyecto_id:
        messages.error(request, "Please select a Project.")
        return redirect("operaciones:importar_precios")

    if not proyectos_qs.filter(pk=proyecto_id).exists():
        messages.error(request, "Selected Project not found or not allowed.")
        return redirect("operaciones:importar_precios")

    proyecto = proyectos_qs.get(pk=proyecto_id)

    form.fields["tecnicos"].queryset = _tecnicos_de_proyecto_qs(proyecto)

    if not form.is_valid():
        messages.error(request, "Invalid form.")
        return redirect(
            f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto.id}"
        )

    try:
        archivo = request.FILES["archivo"]
        tecnicos = form.cleaned_data["tecnicos"]

        if not archivo.name.endswith(".xlsx"):
            messages.error(request, "The file must be in .xlsx format.")
            return redirect(
                f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto.id}"
            )

        df = pd.read_excel(archivo, header=0)
        if df.empty:
            messages.error(request, "The uploaded Excel file is empty.")
            return redirect(
                f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto.id}"
            )

        df.columns = (
            df.columns.str.strip().str.lower().str.replace(r"\s+", "_", regex=True)
        )

        colmap = {
            "city": ["city", "ciudad"],
            "office": ["office", "oficina", "oficce"],
            "client": ["client", "cliente"],
            "work_type": ["work_type", "tipo_trabajo", "tipo_de_trabajo"],
            "code": ["code", "job_code", "codigo", "codigo_trabajo"],
            "description": ["description", "descripcion", "descripción"],
            "uom": ["uom", "unidad_medida", "unidad", "unit"],
            "technical_price": [
                "technical_price",
                "tech_price",
                "precio_tecnico",
                "precio_técnico",
            ],
            "company_price": ["company_price", "precio_empresa", "companyprice"],
            "payment_weeks": [
                "payment_weeks",
                "payment_week",
                "weeks_to_pay",
                "weeks_payment",
                "payment_delay_weeks",
                "semanas_pago",
                "semana_pago",
            ],
        }

        def resolve(colkey, required=True):
            for cand in colmap[colkey]:
                if cand in df.columns:
                    return cand
            if required:
                raise KeyError(
                    f"Required column not found for '{colkey}'. Available columns: {list(df.columns)}"
                )
            return None

        def parse_payment_weeks(value):
            if value in (None, ""):
                return 0
            try:
                if pd.isna(value):
                    return 0
            except Exception:
                pass

            try:
                num = int(float(str(value).strip()))
            except Exception:
                return None

            if num < 0:
                return None
            return num

        c_city = resolve("city")
        c_code = resolve("code")
        c_desc = resolve("description")
        c_uom = resolve("uom")
        c_tp = resolve("technical_price")
        c_cp = resolve("company_price")
        c_pw = resolve("payment_weeks", required=False)
        c_office = resolve("office", required=False)
        c_client = resolve("client", required=False)
        c_wtype = resolve("work_type", required=False)

        preview_data = []
        for _, row in df.iterrows():
            payment_weeks = parse_payment_weeks(row.get(c_pw)) if c_pw else 0

            r = {
                "ciudad": row.get(c_city),
                "proyecto": proyecto.nombre,
                "codigo_trabajo": row.get(c_code),
                "descripcion": row.get(c_desc),
                "uom": row.get(c_uom),
                "precio_tecnico": _to2(row.get(c_tp)),
                "precio_empresa": _to2(row.get(c_cp)),
                "payment_weeks": payment_weeks,
                "oficina": row.get(c_office) if c_office else "",
                "cliente": row.get(c_client) if c_client else "",
                "tipo_trabajo": row.get(c_wtype) if c_wtype else "",
                "tecnico": [t.id for t in tecnicos],
                "error": "",
            }

            missing = []
            if not r["ciudad"]:
                missing.append("city")
            if not r["codigo_trabajo"]:
                missing.append("code")
            if not r["descripcion"]:
                missing.append("description")
            if not r["uom"]:
                missing.append("uom")
            if not r["tipo_trabajo"]:
                missing.append("work_type")

            if r["precio_tecnico"] is None:
                r["error"] += (" | " if r["error"] else "") + "Invalid Technical Price"
            if r["precio_empresa"] is None:
                r["error"] += (" | " if r["error"] else "") + "Invalid Company Price"
            if r["payment_weeks"] is None:
                r["error"] += (" | " if r["error"] else "") + "Invalid Payment Weeks"
            if missing:
                r["error"] += (
                    " | " if r["error"] else ""
                ) + f"Missing fields: {', '.join(missing)}"

            preview_data.append(r)

        # ✅ NUEVO: validar consistencia dentro del archivo importado
        preview_data = _validate_preview_payment_weeks_consistency(
            preview_data=preview_data,
            proyecto_id=proyecto.id,
        )

        request.session["preview_data"] = preview_data
        request.session["selected_proyecto_id"] = proyecto.id

        codes = {r["codigo_trabajo"] for r in preview_data if r.get("codigo_trabajo")}
        has_conflicts = False
        conflicts_by_tech = {}

        for t in tecnicos:
            qs_conf = PrecioActividadTecnico.objects.filter(
                tecnico=t, codigo_trabajo__in=codes
            )

            try:
                qs_conf = qs_conf.filter(proyecto_id=proyecto.id)
            except FieldError:
                nome = str(getattr(proyecto, "nombre", "")).strip()
                cod = str(getattr(proyecto, "codigo", "")).strip()
                cond = dj_models.Q()
                if nome:
                    cond |= dj_models.Q(proyecto__iexact=nome)
                if cod:
                    cond |= dj_models.Q(proyecto__iexact=cod)
                if cond:
                    qs_conf = qs_conf.filter(cond)
                else:
                    qs_conf = qs_conf.none()

            conflicts = list(
                qs_conf.values_list("codigo_trabajo", flat=True).distinct()
            )
            conflicts_by_tech[t.id] = conflicts
            if conflicts:
                has_conflicts = True

        return render(
            request,
            "operaciones/preview_import.html",
            {
                "preview_data": preview_data,
                "tecnicos": tecnicos,
                "has_conflicts": has_conflicts,
                "conflicts_by_tech": conflicts_by_tech,
                "proyecto_sel": proyecto,
            },
        )

    except KeyError as ke:
        messages.error(request, f"Column not found or incorrectly assigned: {ke}")
        return redirect(
            f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto.id}"
        )
    except Exception as e:
        messages.error(request, f"Error during import: {str(e)}")
        return redirect(
            f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto.id}"
        )


@login_required(login_url="usuarios:login")
@rol_requerido("admin", "pm")
def descargar_template_precios(request):
    """
    Genera el template Excel de precios desde cero, sin depender de un archivo base.
    Mantiene el estilo visual del template viejo y agrega la columna Payment Weeks
    después de UOM.
    """
    from io import BytesIO

    from django.http import HttpResponse
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Technician Prices"

    headers = [
        "City",  # A
        "Office",  # B
        "Client",  # C
        "Work Type",  # D
        "Job Code",  # E
        "Description",  # F
        "UOM",  # G
        "Payment\nWeeks",  # H
        "Technical Price",  # I
        "Company Price",  # J
    ]

    # ---- estilos ----
    header_fill = PatternFill(fill_type="solid", fgColor="DCE6F1")
    header_font = Font(name="Calibri", size=12, bold=True, color="000000")
    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    thin_black = Side(style="thin", color="000000")
    all_border = Border(
        left=thin_black,
        right=thin_black,
        top=thin_black,
        bottom=thin_black,
    )

    body_alignment = Alignment(
        horizontal="left",
        vertical="center",
        wrap_text=False,
    )

    # ---- dimensiones ----
    widths = {
        "A": 14,  # City
        "B": 14,  # Office
        "C": 16,  # Client
        "D": 16,  # Work Type
        "E": 14,  # Job Code
        "F": 20,  # Description
        "G": 10,  # UOM
        "H": 16,  # Payment Weeks
        "I": 18,  # Technical Price
        "J": 18,  # Company Price
    }

    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 34

    # ---- header ----
    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = all_border

    # ---- grilla visible hasta fila 15 y columna J ----
    for row in range(2, 16):
        for col in range(1, 11):  # A:J
            cell = ws.cell(row=row, column=col)
            cell.border = all_border
            cell.alignment = body_alignment

    # ---- formato numérico sugerido para precios ----
    for row in range(2, 16):
        ws.cell(row=row, column=9).number_format = "0.00"  # I Technical Price
        ws.cell(row=row, column=10).number_format = "0.00"  # J Company Price

    # ---- freeze pane ----
    ws.freeze_panes = "A2"

    # ---- respuesta ----
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="Technician_Prices_Template.xlsx"'
    )
    return response


@login_required(login_url='usuarios:login')
@rol_requerido('admin', 'pm')
def api_tecnicos_por_proyecto(request):
    """
    Devuelve en JSON los técnicos asignados a un proyecto visible para el usuario.

    Respuesta:
      {
        "tecnicos": [
          { "id": 1, "name": "Juan Pérez", "username": "jperez" },
          ...
        ]
      }
    """
    proyectos_qs = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        'id'
    )

    pid = (request.GET.get('proyecto_id') or '').strip()
    data = {"tecnicos": []}

    if pid and proyectos_qs.filter(pk=pid).exists():
        proyecto = proyectos_qs.get(pk=pid)
        for u in _tecnicos_de_proyecto_qs(proyecto):
            full_name = (u.get_full_name() or "").strip()
            label = full_name or u.username or f"User {u.id}"
            data["tecnicos"].append({
                "id": u.id,
                "name": label,
                "username": u.username,
            })

    return JsonResponse(data)


@login_required
@rol_requerido("admin", "pm")
def confirmar_importar_precios(request):
    if request.method != "POST":
        return redirect("operaciones:importar_precios")

    try:
        preview_data = request.session.get("preview_data", [])
        proyecto_id = request.session.get("selected_proyecto_id")

        if not preview_data or not proyecto_id:
            messages.error(request, "No data to save. Please try again.")
            return redirect("operaciones:importar_precios")

        replace = request.POST.get("replace") == "yes"
        created_total = 0
        updated_total = 0
        skipped_total = 0

        User = get_user_model()

        # Seguridad: no continuar si el preview ya trae errores
        preview_errors = [row for row in preview_data if row.get("error")]
        if preview_errors:
            messages.error(
                request,
                "Import cannot continue because the preview contains errors. Please correct the file and try again.",
            )
            return redirect(
                f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto_id}"
            )

        # ===== Validación por lote contra BD =====
        family_map = _collect_import_family_map(preview_data, int(proyecto_id))

        tecnico_ids_all = set()
        for row in preview_data:
            for tid in row.get("tecnico") or []:
                try:
                    tecnico_ids_all.add(int(tid))
                except Exception:
                    pass

        tecnicos_all = User.objects.filter(id__in=tecnico_ids_all)
        validation_errors = []

        for tecnico in tecnicos_all:
            errs = _validate_db_payment_weeks_consistency_for_import(
                tecnico=tecnico,
                proyecto_id=int(proyecto_id),
                family_map=family_map,
                replace=replace,
            )
            validation_errors.extend(errs)

        if validation_errors:
            for err in validation_errors:
                messages.error(request, err)
            return redirect(
                f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto_id}"
            )

        # ===== Guardado =====
        with transaction.atomic():
            for row in preview_data:
                tecnico_ids = row.get("tecnico", [])
                tecnicos = User.objects.filter(id__in=tecnico_ids)

                for tecnico in tecnicos:
                    lookup = dict(
                        tecnico=tecnico,
                        proyecto_id=proyecto_id,
                        ciudad=(row.get("ciudad") or "").strip(),
                        oficina=(row.get("oficina") or "").strip(),
                        cliente=(row.get("cliente") or "").strip(),
                        codigo_trabajo=(row.get("codigo_trabajo") or "").strip(),
                    )

                    defaults = dict(
                        tipo_trabajo=(row.get("tipo_trabajo") or "").strip(),
                        descripcion=(row.get("descripcion") or "").strip(),
                        unidad_medida=(row.get("uom") or "").strip(),
                        payment_weeks=int(row.get("payment_weeks") or 0),
                        precio_tecnico=row.get("precio_tecnico") or 0,
                        precio_empresa=row.get("precio_empresa") or 0,
                    )

                    if replace:
                        obj, created = PrecioActividadTecnico.objects.update_or_create(
                            **lookup,
                            defaults=defaults,
                        )
                        if created:
                            created_total += 1
                        else:
                            updated_total += 1
                    else:
                        obj, created = PrecioActividadTecnico.objects.get_or_create(
                            **lookup,
                            defaults=defaults,
                        )
                        if created:
                            created_total += 1
                        else:
                            skipped_total += 1

        msg = f"Import completed. Created: {created_total}, updated: {updated_total}"
        if skipped_total:
            msg += f", skipped (already existing): {skipped_total}"
        messages.success(request, msg)

        request.session.pop("preview_data", None)
        request.session.pop("selected_proyecto_id", None)

        return redirect("operaciones:listar_precios_tecnico")

    except ValidationError as e:
        messages.error(request, str(e))
        return redirect(
            f"{reverse('operaciones:importar_precios')}?proyecto_id={proyecto_id}"
        )
    except Exception as e:
        messages.error(request, f"An error occurred during the import: {str(e)}")
        return redirect("operaciones:importar_precios")


from core.permissions import project_object_access_required  # <-- NUEVO


@login_required
@rol_requerido('admin', 'pm')
@project_object_access_required(model='operaciones.PrecioActividadTecnico', object_kw='pk', project_attr='proyecto_id')
def editar_precio(request, pk):
    precio = get_object_or_404(PrecioActividadTecnico, pk=pk)
    if request.method == 'POST':
        form = PrecioActividadTecnicoForm(request.POST, instance=precio)
        if form.is_valid():
            form.save()
            messages.success(request, "Price updated successfully.")
            return redirect('operaciones:listar_precios_tecnico')
    else:
        form = PrecioActividadTecnicoForm(instance=precio)
    return render(request, 'operaciones/editar_precio.html', {'form': form, 'precio': precio})


@login_required
@rol_requerido('admin', 'pm')
@project_object_access_required(model='operaciones.PrecioActividadTecnico', object_kw='pk', project_attr='proyecto_id')
def eliminar_precio(request, pk):
    precio = get_object_or_404(PrecioActividadTecnico, pk=pk)
    precio.delete()
    messages.success(request, "Price deleted successfully.")
    return redirect('operaciones:listar_precios_tecnico')

# --- BILLING DE AQUI PARA ABAJO ---
#
# Ajusta si tu modelo de precios está en otra app


Usuario = get_user_model()


def money(x):  # redondeo
    return Decimal(x).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)


@login_required(login_url='usuarios:login')
@rol_requerido('admin', 'pm')
def bulk_delete_precios(request):
    if request.method != "POST":
        messages.error(request, "Invalid request.")
        return redirect('operaciones:listar_precios_tecnico')

    ids = request.POST.getlist("ids")
    return_page = request.POST.get("return_page") or ""
    return_cantidad = request.POST.get("return_cantidad") or ""

    if not ids:
        messages.info(request, "No prices selected.")
        return redirect('operaciones:listar_precios_tecnico')

    # 🔒 Solo puede borrar los que están dentro de sus proyectos
    qs = filter_queryset_by_access(
        PrecioActividadTecnico.objects.filter(id__in=ids).select_related('proyecto'),
        request.user,
        'proyecto_id'
    )

    deleted_count = qs.count()
    qs.delete()

    messages.success(request, f"{deleted_count} price(s) deleted successfully.")

    # reconstruye URL de retorno preservando filtros/paginación...
    # (tu código original aquí, igual)
    base = reverse('operaciones:listar_precios_tecnico')
    params = []
    if return_cantidad:
        params.append(f"cantidad={return_cantidad}")
    if return_page and return_cantidad != "todos":
        params.append(f"page={return_page}")

    for key in ("f_tecnico", "f_ciudad", "f_proyecto", "f_codigo"):
        val = (request.POST.get(key) or '').strip()
        if val:
            params.append(f"{key}={val}")

    url = f"{base}?{'&'.join(params)}" if params else base
    return redirect(url)

# ===== Listado =====


def repartir_100(n):
    if n <= 0:
        return []
    base = (Decimal("100.00")/Decimal(n)).quantize(Decimal("0.01"))
    partes = [base]*n
    diff = Decimal("100.00") - sum(partes)
    if diff and partes:
        partes[-1] = (partes[-1]+diff).quantize(Decimal("0.01"))
    return partes


# ===== Descuento directo =====

def recomputar_estado_desde_asignaciones(self, save: bool = True) -> str:
    # NUEVO: si es descuento directo, no tocar el estado
    if self.is_direct_discount:
        return self.estado

    estados = list(self.tecnicos_sesion.values_list("estado", flat=True))
    nuevo = "asignado"
    if estados:
        if any(e == "en_revision_supervisor" for e in estados):
            nuevo = "en_revision_supervisor"
        elif any(e == "en_proceso" for e in estados):
            nuevo = "en_proceso"
        elif all(e == "aprobado_pm" for e in estados):
            nuevo = "aprobado_pm"
        elif any(e == "rechazado_pm" for e in estados):
            nuevo = "rechazado_pm"
        elif all(e == "aprobado_supervisor" for e in estados):
            nuevo = "aprobado_supervisor"
        elif any(e == "rechazado_supervisor" for e in estados):
            nuevo = "rechazado_supervisor"

    if self.estado != nuevo:
        self.estado = nuevo
        if save:
            self.save(update_fields=["estado"])
    return self.estado


@login_required
@require_POST
def exportar_billing_excel(request):
    """
    Exporta a XLSX con columnas:
    Project ID, Date, Week, Project Address, City, Work Type, Job Code,
    Description, Qty, Subtotal Company.
    - Encabezado con color y filtros
    - Bordes finos en toda la tabla
    - Bandas alternadas (gris/ blanco) en filas de datos
    - Fila Total: 'Total' en Qty (col I) y monto en Subtotal Company (col J)
    - Líneas de cuadricula DESACTIVADAS
    """

    # ========= Estilos locales =========
    HDR_FILL = PatternFill("solid", fgColor="374151")   # gris oscuro
    HDR_FONT = Font(bold=True, color="FFFFFF")
    HDR_ALIGN = Alignment(horizontal="center", vertical="center")
    CELL_ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
    CELL_ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)
    CELL_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")
    THIN = Side(style="thin", color="D1D5DB")
    BORDER_ALL = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    ZEBRA_GRAY = "E5E7EB"   # gris clarito
    ZEBRA_WHITE = "FFFFFF"  # blanco

    # ========= Helpers locales =========
    def _export_headers():
        return [
            "Project ID", "Date", "Week", "Project Address", "City",
            "Work Type", "Job Code", "Description", "Qty", "Subtotal Company",
        ]

    def _get_address_from_session_only(s):
        return (
            getattr(s, "direccion_proyecto", None)
            or getattr(s, "direccion", None)
            or getattr(s, "project_address", None)
            or getattr(s, "direccion_obra", None)
            or ""
        )

    def _get_address_from_item_or_session(it, s):
        return (
            getattr(it, "direccion", None)
            or getattr(it, "project_address", None)
            or getattr(it, "direccion_obra", None)
            or getattr(s, "direccion_proyecto", None)
            or getattr(s, "direccion", None)
            or getattr(s, "project_address", None)
            or getattr(s, "direccion_obra", None)
            or ""
        )

    def _xlsx_response(workbook):
        from io import BytesIO
        bio = BytesIO()
        workbook.save(bio)
        bio.seek(0)
        ts = timezone.now().strftime("%Y%m%d_%H%M%S")
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="billing_export_{ts}.xlsx"'
        return resp

    def _format_money(ws, cols):
        money_fmt = '$#,##0.00'
        for col in cols:
            for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2, values_only=False):
                for c in col_cells:
                    c.number_format = money_fmt

    def _format_number(ws, cols):
        num_fmt = '#,##0.00'
        for col in cols:
            for col_cells in ws.iter_cols(min_col=col, max_col=col, min_row=2, values_only=False):
                for c in col_cells:
                    c.number_format = num_fmt

    def _set_widths(ws, mapping):
        for idx, width in mapping.items():
            ws.column_dimensions[get_column_letter(idx)].width = width

    def _apply_table_borders(ws):
        """Bordes + alineaciones en todo el rango con datos (incluye encabezado)."""
        max_r, max_c = ws.max_row, ws.max_column
        for r in range(1, max_r + 1):
            for c in range(1, max_c + 1):
                cell = ws.cell(row=r, column=c)
                cell.border = BORDER_ALL
                if c in (9, 10):                    # Qty / Subtotal
                    cell.alignment = CELL_ALIGN_RIGHT
                elif c in (4, 8):                   # Address / Description -> wrap
                    cell.alignment = CELL_ALIGN_LEFT_WRAP
                else:
                    cell.alignment = CELL_ALIGN_LEFT

    def _apply_zebra(ws, start_row: int, end_row: int, gray_hex: str, white_hex: str):
        """Relleno alternado (gris/blanco) desde start_row hasta end_row."""
        if end_row < start_row:
            return
        fill_gray = PatternFill("solid", fgColor=gray_hex)
        fill_white = PatternFill("solid", fgColor=white_hex)
        max_c = ws.max_column
        for r in range(start_row, end_row + 1):
            fill = fill_gray if (r - start_row) % 2 == 0 else fill_white
            for c in range(1, max_c + 1):
                ws.cell(row=r, column=c).fill = fill

    def _style_after_fill(ws):
        """Header gris + filtros + congelar panes."""
        for col, _ in enumerate(_export_headers(), start=1):
            cell = ws.cell(row=1, column=col)
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
            cell.alignment = HDR_ALIGN
            cell.border = BORDER_ALL
        ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
        ws.freeze_panes = "A2"

    # ========= 1) Parseo de IDs =========
    raw = (request.POST.get("ids") or "").strip()
    ids = [int(x) for x in raw.split(",") if x.strip().isdigit()]

    # ========= 2) Workbook / hoja =========
    wb = Workbook()
    ws = wb.active
    ws.title = "Billing"

    # 👉 Desactivar líneas de cuadricula (en pantalla y también en impresión)
    ws.sheet_view.showGridLines = False
    ws.print_options.gridLines = False

    # Encabezados
    headers = _export_headers()
    ws.append(headers)

    if not ids:
        _style_after_fill(ws)
        return _xlsx_response(wb)

    # ========= 3) Prefetch =========
    sesiones = (
        SesionBilling.objects.filter(id__in=ids)
        .prefetch_related(
            Prefetch("items", queryset=ItemBilling.objects.order_by("id")),
            Prefetch(
                "tecnicos_sesion",
                queryset=SesionBillingTecnico.objects.select_related("tecnico"),
            ),
        )
        .order_by("id")
    )

    # --- Seguridad extra: solo exportar billings de proyectos visibles para este usuario ---
    # Admin/superuser puede exportar todo; el resto queda limitado a sus proyectos.
    sesiones = list(sesiones)
    if not request.user.is_superuser:
        proyectos_visibles = filter_queryset_by_access(
            Proyecto.objects.all(),
            request.user,
            "id",
        )
        if proyectos_visibles.exists():
            allowed_proj_ids = {
                str(pk) for pk in proyectos_visibles.values_list("id", flat=True)
            }

            # ✅ NUEVO: ventana de visibilidad por ProyectoAsignacion (por proyecto y fecha)
            # Si include_history=True -> sin corte por fecha
            # Si include_history=False y start_at existe -> solo desde start_at en adelante
            asignaciones = []
            try:
                if ProyectoAsignacion is not None:
                    asignaciones = list(
                        ProyectoAsignacion.objects
                        .filter(usuario=request.user, proyecto__in=proyectos_visibles)
                        .select_related("proyecto")
                    )
            except Exception:
                asignaciones = []

            if asignaciones:
                asign_by_pid = {}
                for a in asignaciones:
                    p = getattr(a, "proyecto", None)
                    if not p:
                        continue
                    asign_by_pid[str(getattr(p, "id", "")).strip()] = a

                sesiones_filtradas = []
                for s in sesiones:
                    sp = getattr(s, "proyecto", None)
                    if sp not in allowed_proj_ids:
                        continue

                    a = asign_by_pid.get(str(sp).strip())
                    if not a:
                        # si no hay asignación, mantenemos solo por proyecto (no cambiamos lógica base)
                        sesiones_filtradas.append(s)
                        continue

                    if getattr(a, "include_history", False) or not getattr(a, "start_at", None):
                        sesiones_filtradas.append(s)
                    else:
                        if getattr(s, "creado_en", None) and s.creado_en >= a.start_at:
                            sesiones_filtradas.append(s)

                sesiones = sesiones_filtradas
            else:
                sesiones = [
                    s for s in sesiones
                    if getattr(s, "proyecto", None) in allowed_proj_ids
                ]
        else:
            # Si no tiene proyectos visibles, no exportamos nada
            sesiones = []

    # ========= 4) Filas =========
    total_subtotal_company = 0.0
    tz = timezone.get_current_timezone()

    for s in sesiones:
        dt = s.creado_en
        date_str = timezone.localtime(dt, tz).strftime("%d-%b").lower() if dt else ""
        week_str = getattr(s, "semana_pago_proyectada", "") or getattr(s, "week", "")
        city = getattr(s, "ciudad", "") or getattr(s, "city", "")
        project_id = getattr(s, "proyecto_id", "") or getattr(s, "project_id", "")

        if not s.items.all():
            addr_session = _get_address_from_session_only(s)
            ws.append([project_id, date_str, week_str, addr_session, city, "", "", "", 0.0, 0.0])
            continue

        for it in s.items.all():
            project_address = _get_address_from_item_or_session(it, s)
            qty = float(getattr(it, "cantidad", 0) or 0)
            sub_company = float(getattr(it, "subtotal_empresa", 0) or 0)
            ws.append([
                project_id,                  # A
                date_str,                    # B
                week_str,                    # C
                project_address,             # D
                city,                        # E
                getattr(it, "tipo_trabajo", "") or getattr(it, "work_type", ""),  # F
                getattr(it, "codigo_trabajo", "") or getattr(it, "job_code", ""),  # G
                getattr(it, "descripcion", "") or getattr(it, "description", ""),  # H
                qty,                         # I
                sub_company                  # J
            ])
            total_subtotal_company += sub_company

    # ========= 5) Formatos / estilos =========
    _format_money(ws, cols=[10])   # J: Subtotal Company
    _format_number(ws, cols=[9])   # I: Qty

    _set_widths(ws, {
        1: 12, 2: 10, 3: 12, 4: 36, 5: 14, 6: 14, 7: 12, 8: 34, 9: 6, 10: 16
    })

    _apply_table_borders(ws)

    # Zebra desde la fila 2 (datos) hasta la última fila de datos
    data_end = ws.max_row
    _apply_zebra(ws, start_row=2, end_row=data_end, gray_hex=ZEBRA_GRAY, white_hex=ZEBRA_WHITE)

    _style_after_fill(ws)

    # ========= 6) Fila Total =========
    ws.append([""] * 10)  # separador opcional
    total_row = ws.max_row
    ws.cell(row=total_row, column=9, value="Total").font = Font(bold=True)  # I
    ws.cell(row=total_row, column=10, value=total_subtotal_company).font = Font(bold=True)  # J
    ws.cell(row=total_row, column=10).number_format = '$#,##0.00'
    for col in range(1, 11):
        c = ws.cell(row=total_row, column=col)
        c.border = BORDER_ALL
        c.alignment = CELL_ALIGN_RIGHT if col in (9, 10) else (
            CELL_ALIGN_LEFT_WRAP if col in (4, 8) else CELL_ALIGN_LEFT
        )

    return _xlsx_response(wb)


def _norm(txt: str) -> str:
    """minúsculas + sin espacios/guiones/underscores (para comparar estados)."""
    if not txt:
        return ""
    t = txt.strip().lower()
    return "".join(ch for ch in t if ch.isalnum())


from datetime import datetime  # 👈 si no lo tienes ya, agrégalo arriba


@require_POST
@login_required
@rol_requerido("admin", "pm")
@transaction.atomic
def billing_send_finance(request):
    """
    Enviar a Finanzas SOLO si:
      - is_direct_discount == True             -> finance_status = 'review_discount'
      - estado == 'aprobado_supervisor' (normalizado) -> finance_status = 'sent'

    Nunca 400 por mezcla: procesa lo permitido y devuelve 'skipped' con motivo.
    Responde SIEMPRE JSON (nada de HTML).

    FIXES:
      - Sellar finance_sent_at también cuando new_status='review_discount'.
      - Permitir re-sellar si ya está en 'review_discount' PERO sin finance_sent_at (intentos previos).
    """
    # ---- parseo ids + nota + daily_number + finish_date ----
    ids, note, daily_number = [], "", ""
    finish_date = None
    ctype = (request.content_type or "").lower()

    if "application/json" in ctype:
        try:
            payload = json.loads(request.body.decode("utf-8") or "{}")
        except Exception:
            return JsonResponse({"ok": False, "error": "INVALID_JSON"}, status=400)
        ids = [int(x) for x in (payload.get("ids") or []) if str(x).isdigit()]
        note = (payload.get("note") or "").strip()
        daily_number = (payload.get("daily_number") or "").strip()
        finish_str = (payload.get("finish_date") or "").strip()
    else:
        raw = (request.POST.get("ids") or "").strip()
        ids = [int(x) for x in raw.split(",") if x.isdigit()]
        note = (request.POST.get("note") or "").strip()
        daily_number = (request.POST.get("daily_number") or "").strip()
        finish_str = (request.POST.get("finish_date") or "").strip()

    # 👉 parsear finish_date si viene
    if finish_str:
        try:
            # formato esperado: YYYY-MM-DD (lo que envía el JS con flatpickr)
            finish_date = datetime.fromisoformat(finish_str).date()
        except ValueError:
            try:
                # fallback por si alguna vez llega como dd/mm/YYYY
                finish_date = datetime.strptime(finish_str, "%d/%m/%Y").date()
            except ValueError:
                return JsonResponse(
                    {"ok": False, "error": "INVALID_FINISH_DATE"},
                    status=400,
                )

    if not ids:
        return JsonResponse({"ok": False, "error": "NO_IDS"}, status=400)

    # ---- reglas permitidas ----
    allowed_supervisor_norms = {
        "aprobadosupervisor",
        "approvedsupervisor",
        "approvedbysupervisor",
        "aprobadoporsupervisor",
    }

    # Estados de finanzas que BLOQUEAN reenvío.
    # OJO: dejaremos pasar 'review_discount' SI NO TIENE finance_sent_at (para reestampar).
    blocked_fin = {
        "sent", "senttofinance",
        "reviewdiscount", "discountapplied",
        "inreview", "pending", "readyforpayment",
        "paid", "rejected", "cancelled", "canceled",
        "enviado", "enrevision", "pendiente", "listoparapago",
        "pagado", "rechazado", "cancelado",
    }

    rows = list(SesionBilling.objects.filter(id__in=ids))
    now = timezone.now()

    updated = 0
    updated_rows = []
    skipped = []
    plan = []  # (id, new_finance_status)

    # --- Seguridad extra: solo billings de proyectos a los que el usuario tiene acceso ---
    # Admin/superuser puede enviar todo; PM queda restringido a sus proyectos visibles.
    forbidden_ids = set()
    if not request.user.is_superuser:
        proyectos_visibles = filter_queryset_by_access(
            Proyecto.objects.all(),
            request.user,
            "id",
        )
        if proyectos_visibles.exists():
            allowed_proj_ids = {
                str(pk) for pk in proyectos_visibles.values_list("id", flat=True)
            }

            # ✅ NUEVO: ventana de visibilidad por ProyectoAsignacion (por proyecto y fecha)
            asignaciones = []
            try:
                if ProyectoAsignacion is not None:
                    asignaciones = list(
                        ProyectoAsignacion.objects
                        .filter(usuario=request.user, proyecto__in=proyectos_visibles)
                        .select_related("proyecto")
                    )
            except Exception:
                asignaciones = []

            asign_by_pid = {}
            if asignaciones:
                for a in asignaciones:
                    p = getattr(a, "proyecto", None)
                    if not p:
                        continue
                    asign_by_pid[str(getattr(p, "id", "")).strip()] = a

            filtered_rows = []
            for s in rows:
                # En SesionBilling guardamos el PK del proyecto en s.proyecto (como string)
                sp = getattr(s, "proyecto", None)
                if sp in allowed_proj_ids:
                    # Si hay asignaciones, aplicar corte por fecha
                    if asign_by_pid:
                        a = asign_by_pid.get(str(sp).strip())
                        if a and (not getattr(a, "include_history", False)) and getattr(a, "start_at", None):
                            if getattr(s, "creado_en", None) and s.creado_en < a.start_at:
                                forbidden_ids.add(s.id)
                                continue
                    filtered_rows.append(s)
                else:
                    forbidden_ids.add(s.id)
            rows = filtered_rows
        else:
            # Si el usuario no tiene proyectos visibles, ningún billing es enviable
            forbidden_ids = {s.id for s in rows}
            rows = []

    def _norm(s: str) -> str:
        return (s or "").strip().lower().replace("_", "").replace("-", "").replace(" ", "")

    # Primera pasada: decidir plan
    for s in rows:
        estado_norm = _norm(getattr(s, "estado", ""))
        fin_norm = _norm(getattr(s, "finance_status", ""))
        fin_sent = getattr(s, "finance_sent_at", None)

        # Si está bloqueado en finanzas...
        if fin_norm in blocked_fin:
            # ...EXCEPCIÓN: permitir reestampar si está en review_discount PERO sin finance_sent_at
            if fin_norm == "reviewdiscount" and not fin_sent:
                plan.append((s.id, "review_discount"))
                continue

            skipped.append({
                "id": s.id, "estado": s.estado,
                "is_direct_discount": bool(s.is_direct_discount),
                "finance_status": s.finance_status,
                "skip_reason": "FINANCE_STATUS_BLOCKED",
            })
            continue

        # Flujo normal
        if getattr(s, "is_direct_discount", False) is True:
            plan.append((s.id, "review_discount"))
        elif estado_norm in allowed_supervisor_norms:
            plan.append((s.id, "sent"))
        else:
            skipped.append({
                "id": s.id, "estado": s.estado,
                "is_direct_discount": bool(s.is_direct_discount),
                "finance_status": s.finance_status,
                "skip_reason": "NOT_ALLOWED_STATUS",
            })

    # Agregar también los billings que se intentaron enviar pero pertenecen a proyectos no autorizados
    for bid in forbidden_ids:
        skipped.append({
            "id": bid,
            "estado": None,
            "is_direct_discount": None,
            "finance_status": None,
            "skip_reason": "FORBIDDEN_PROJECT",
        })

    # aplicar updates con lock
    by_id_new = {i: st for (i, st) in plan}
    if by_id_new:
        for s in SesionBilling.objects.select_for_update().filter(id__in=by_id_new.keys()):
            new_status = by_id_new[s.id]
            s.finance_status = new_status

            touched_fields = ["finance_status"]

            if hasattr(s, "finance_updated_at"):
                s.finance_updated_at = now
                touched_fields.append("finance_updated_at")

            if hasattr(s, "finance_sent_at") and new_status in ("sent", "review_discount"):
                s.finance_sent_at = now
                touched_fields.append("finance_sent_at")

            # 👇 Guardar Daily Number (mismo para todos los seleccionados)
            if daily_number:
                s.finance_daily_number = daily_number
                touched_fields.append("finance_daily_number")

            # 👇 NUEVO: guardar fecha de término si viene
            if finish_date is not None:
                s.finance_finish_date = finish_date
                touched_fields.append("finance_finish_date")

            if note:
                prefix = f"{now:%Y-%m-%d %H:%M} Ops: "
                s.finance_note = ((s.finance_note + "\n") if s.finance_note else "") + prefix + note
                touched_fields.append("finance_note")

            s.save(update_fields=touched_fields)
            updated += 1
            updated_rows.append({"id": s.id, "finance_status": s.finance_status})

    return JsonResponse({"ok": True, "count": updated, "updated": updated_rows, "skipped": skipped})

@login_required
@rol_requerido('admin', 'pm')
@require_POST
@transaction.atomic
def billing_mark_in_review(request, pk: int):
    s = get_object_or_404(SesionBilling, pk=pk)
    if s.finance_status != "rejected":
        messages.info(request, "Only applies when Finance has rejected it.")
        return redirect("operaciones:listar_billing")

    note = (request.POST.get("reason")
            or request.POST.get("note") or "").strip()
    now = timezone.now()

    # Lo dejamos como "in_review" (aparece en Finanzas con scope=open)
    s.finance_status = "in_review"
    s.finance_updated_at = now

    if note:
        prefix = f"{now:%Y-%m-%d %H:%M} Ops: "
        s.finance_note = (
            s.finance_note + "\n" if s.finance_note else "") + prefix + note

    s.save(update_fields=["finance_status",
           "finance_updated_at", "finance_note"])
    messages.success(request, "Marked as 'In review' for Finance.")
    return redirect("operaciones:listar_billing")


@login_required
@require_POST
def billing_reopen_asignado(request, pk):
    obj = get_object_or_404(SesionBilling, pk=pk)

    # 🔒 Bloqueo contable absoluto:
    # Si al menos una línea de algún trabajador ya fue pagada,
    # el Billing no se puede reabrir.
    if _session_is_paid_locked(obj):
        messages.error(
            request,
            (
                "This billing cannot be reopened because at least one worker/work type "
                "line has already been marked as paid."
            ),
        )
        return HttpResponseRedirect(
            request.META.get("HTTP_REFERER", "/operaciones/billing/listar/")
        )

    if obj.estado in ("aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"):
        with transaction.atomic():
            obj.estado = "asignado"
            obj.save(update_fields=["estado"])

            obj.tecnicos_sesion.all().update(
                estado="asignado",
                aceptado_en=None,
                finalizado_en=None,
                supervisor_revisado_en=None,
                supervisor_comentario="",
                pm_revisado_en=None,
                pm_comentario="",
                reintento_habilitado=True,
            )

            rebuild_billing_payweek_snapshot(obj)

        messages.success(
            request,
            f"Billing #{obj.pk} has been reopened to 'Assigned' and all assignments were reactivated.",
        )
    else:
        messages.info(request, "This record is not in an approved state.")

    return HttpResponseRedirect(
        request.META.get("HTTP_REFERER", "/operaciones/billing/listar/")
    )

@login_required
def billing_excel_options(request):
    """
    Devuelve los valores de filtros tipo Excel para Billing List.

    Esta vista se llama por AJAX después de cargar la tabla.
    No renderiza template.
    Respeta los filtros rápidos normales:
      - date
      - projid
      - week
      - tech
      - client
      - status

    Importante:
      - NO aplica excel_filters para construir las opciones.
      - Así el usuario puede cambiar filtros Excel sin perder opciones.
    """
    import json
    from datetime import date as _date
    from decimal import Decimal

    from django.db.models import Prefetch, Q
    from django.http import JsonResponse

    from core.permissions import filter_queryset_by_access
    from facturacion.models import Proyecto
    from operaciones.models import (BillingPayWeekSnapshot, SesionBilling,
                                    SesionBillingTecnico)

    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    user = request.user

    can_view_legacy_history = user.is_superuser or getattr(
        user, "es_usuario_historial", False
    )

    visible_filter = (
        Q(is_direct_discount=True)
        & Q(finance_sent_at__isnull=True)
        & ~Q(finance_status="paid")
    ) | (
        Q(is_direct_discount=False)
        & ~Q(finance_status__in=["sent", "pending", "paid", "in_review"])
    )

    qs = SesionBilling.objects.filter(visible_filter).order_by("-creado_en")

    # ============================================================
    # Restricción por proyectos
    # ============================================================
    if not can_view_legacy_history:
        try:
            proyectos_user = filter_queryset_by_access(
                Proyecto.objects.all(),
                user,
                "id",
            )
        except Exception:
            proyectos_user = Proyecto.objects.none()

        proyectos_user_list = list(proyectos_user)

        if proyectos_user_list:
            allowed_keys = set()

            for p in proyectos_user_list:
                nombre = (getattr(p, "nombre", "") or "").strip()
                if nombre:
                    allowed_keys.add(nombre)

                codigo = getattr(p, "codigo", None)
                if codigo:
                    allowed_keys.add(str(codigo).strip())

                allowed_keys.add(str(p.id).strip())

            qs = qs.filter(proyecto__in=allowed_keys)

            asignaciones = []

            if ProyectoAsignacion is not None:
                try:
                    asignaciones = list(
                        ProyectoAsignacion.objects.filter(
                            usuario=user,
                            proyecto__in=proyectos_user_list,
                        ).select_related("proyecto")
                    )
                except Exception:
                    asignaciones = []

            if asignaciones:
                window_q = Q()
                has_window_q = False

                for a in asignaciones:
                    p = getattr(a, "proyecto", None)
                    if not p:
                        continue

                    keys = set()

                    nombre = (getattr(p, "nombre", "") or "").strip()
                    if nombre:
                        keys.add(nombre)

                    codigo = getattr(p, "codigo", None)
                    if codigo:
                        keys.add(str(codigo).strip())

                    keys.add(str(p.id).strip())

                    if not keys:
                        continue

                    if getattr(a, "include_history", False) or not getattr(
                        a, "start_at", None
                    ):
                        window_q |= Q(proyecto__in=keys)
                    else:
                        window_q |= Q(proyecto__in=keys, creado_en__gte=a.start_at)

                    has_window_q = True

                qs = qs.filter(window_q) if has_window_q else qs.none()
        else:
            qs = qs.none()

    # ============================================================
    # Filtros rápidos normales
    # ============================================================
    f = {
        "date": (request.GET.get("date") or "").strip(),
        "projid": (request.GET.get("projid") or "").strip(),
        "week": (request.GET.get("week") or "").strip(),
        "tech": (request.GET.get("tech") or "").strip(),
        "client": (request.GET.get("client") or "").strip(),
        "status": (request.GET.get("status") or "").strip(),
    }

    qs_filtered = qs

    if f["date"]:
        try:
            d = _date.fromisoformat(f["date"])
            qs_filtered = qs_filtered.filter(creado_en__date=d)
        except ValueError:
            pass

    if f["projid"]:
        qs_filtered = qs_filtered.filter(proyecto_id__icontains=f["projid"])

    if f["week"]:
        qs_filtered = qs_filtered.filter(
            Q(semana_pago_proyectada__icontains=f["week"])
            | Q(semana_pago_real__icontains=f["week"])
            | Q(discount_week__icontains=f["week"])
            | Q(pay_week_snapshots__semana_resultado__icontains=f["week"])
            | Q(pay_week_snapshots__semana_base__icontains=f["week"])
        )

    if f["tech"]:
        qs_filtered = qs_filtered.filter(
            Q(tecnicos_sesion__tecnico__first_name__icontains=f["tech"])
            | Q(tecnicos_sesion__tecnico__last_name__icontains=f["tech"])
            | Q(tecnicos_sesion__tecnico__username__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__first_name__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__last_name__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__username__icontains=f["tech"])
        )

    if f["client"]:
        qs_filtered = qs_filtered.filter(cliente__icontains=f["client"])

    if f["status"]:
        status_txt = f["status"].lower().strip()

        if any(k in status_txt for k in ("direct", "descuento", "discount")):
            qs_filtered = qs_filtered.filter(is_direct_discount=True)
        else:
            mapping = [
                (
                    ("aprobado supervisor", "approved by supervisor"),
                    Q(estado="aprobado_supervisor"),
                ),
                (
                    ("rechazado supervisor", "rejected by supervisor"),
                    Q(estado="rechazado_supervisor"),
                ),
                (
                    ("en revision", "supervisor review", "in supervisor review"),
                    Q(estado="en_revision_supervisor"),
                ),
                (("finalizado", "finished"), Q(estado="finalizado")),
                (("en proceso", "in progress"), Q(estado="en_proceso")),
                (("asignado", "assigned"), Q(estado="asignado")),
                (("aprobado pm", "approved by pm"), Q(estado="aprobado_pm")),
                (("rechazado pm", "rejected by pm"), Q(estado="rechazado_pm")),
            ]

            applied = False

            for keys, cond in mapping:
                if any(k in status_txt for k in keys):
                    qs_filtered = qs_filtered.filter(cond)
                    applied = True
                    break

            if not applied:
                if "aprobado" in status_txt or "approved" in status_txt:
                    qs_filtered = qs_filtered.filter(
                        estado__in=["aprobado_supervisor", "aprobado_pm"]
                    )
                elif "rechazado" in status_txt or "rejected" in status_txt:
                    qs_filtered = qs_filtered.filter(
                        estado__in=["rechazado_supervisor", "rechazado_pm"]
                    )

    qs_filtered = qs_filtered.distinct()

    # ============================================================
    # Helpers para valores Excel
    # ============================================================
    def money_value(value):
        if value in (None, ""):
            return "—"

        try:
            return f"${Decimal(value):.2f}"
        except Exception:
            return str(value)

    def status_label(s):
        if getattr(s, "is_direct_discount", False):
            return "Direct discount"

        estado = getattr(s, "estado", "") or ""

        labels = {
            "aprobado_pm": "Approved by PM",
            "rechazado_pm": "Rejected by PM",
            "aprobado_supervisor": "Approved by supervisor",
            "rechazado_supervisor": "Rejected by supervisor",
            "en_revision_supervisor": "In supervisor review",
            "finalizado": "Finished (pending review)",
            "en_proceso": "In progress",
            "asignado": "Assigned",
        }

        return labels.get(estado, "Assigned")

    def finance_status_label(s):
        finance_status = getattr(s, "finance_status", "") or ""

        labels = {
            "sent": "Sent to Finance",
            "in_review": "In review",
            "rejected": "Rejected",
            "pending": "Pending payment",
            "paid": "Paid",
            "review_discount": "Review discount",
            "discount_applied": "Discount applied",
            "none": "—",
            "": "—",
            None: "—",
        }

        return labels.get(finance_status, "—")

    def diff_label(s):
        real = getattr(s, "real_company_billing", None)
        subtotal = getattr(s, "subtotal_empresa", None)

        if real in (None, "") or subtotal in (None, ""):
            return "—"

        try:
            real_d = Decimal(real)
            sub_d = Decimal(subtotal)
            diff = real_d - sub_d
        except Exception:
            return "—"

        if diff == 0:
            return "$0.00"

        if diff < 0:
            return f"- ${abs(diff):.2f}"

        return f"+ ${diff:.2f}"

    def techs_label(sesion):
        vals = []

        try:
            for st in sesion.tecnicos_sesion.all():
                if not getattr(st, "tecnico", None):
                    continue

                vals.append(st.tecnico.get_full_name() or st.tecnico.username)
        except Exception:
            pass

        return ", ".join(v for v in vals if v) or "—"

    def comments_label(s):
        vals = []

        try:
            for a in s.tecnicos_sesion.all():
                txt = (getattr(a, "tecnico_comentario", "") or "").strip()

                if not txt:
                    continue

                tech_name = (
                    a.tecnico.get_full_name() if getattr(a, "tecnico", None) else ""
                ) or (
                    getattr(a.tecnico, "username", "")
                    if getattr(a, "tecnico", None)
                    else ""
                )

                if tech_name:
                    vals.append(f"{tech_name}: {txt}")
                else:
                    vals.append(txt)
        except Exception:
            pass

        return " | ".join(vals) if vals else "—"

    def build_payweek_label(s):
        vals = []

        try:
            snaps = list(s.pay_week_snapshots.all())
        except Exception:
            snaps = []

        if snaps:
            for snap in snaps:
                tech_name = (
                    snap.tecnico.get_full_name().strip()
                    if getattr(snap, "tecnico", None) and snap.tecnico.get_full_name()
                    else getattr(snap.tecnico, "username", "")
                    or f"User {snap.tecnico_id}"
                )

                work_type = (
                    (snap.tipo_trabajo or "").strip()
                    or (getattr(snap.item, "tipo_trabajo", "") or "").strip()
                    or "Legacy"
                )

                week = (
                    (getattr(snap, "semana_resultado", "") or "").strip()
                    or (getattr(snap, "semana_base", "") or "").strip()
                    or (getattr(s, "semana_pago_real", "") or "").strip()
                    or (getattr(s, "discount_week", "") or "").strip()
                    or (getattr(s, "semana_pago_proyectada", "") or "").strip()
                    or "—"
                )

                vals.append(f"{tech_name} — {work_type} → {week}")

        if vals:
            return " | ".join(vals)

        return (
            (getattr(s, "semana_pago_real", "") or "").strip()
            or (getattr(s, "discount_week", "") or "").strip()
            or (getattr(s, "semana_pago_proyectada", "") or "").strip()
            or "—"
        )

    def resolve_project_labels_for_sessions(sessions):
        proj_ids = set()
        proj_texts = set()

        for s in sessions:
            raw_proyecto = getattr(s, "proyecto", None)
            if raw_proyecto not in (None, "", "-"):
                txt = str(raw_proyecto).strip()
                if txt:
                    proj_texts.add(txt)
                    try:
                        proj_ids.add(int(txt))
                    except Exception:
                        pass

            raw_proyecto_id = getattr(s, "proyecto_id", None)
            if raw_proyecto_id not in (None, "", "-"):
                txt2 = str(raw_proyecto_id).strip()
                if txt2:
                    proj_texts.add(txt2)
                    try:
                        proj_ids.add(int(txt2))
                    except Exception:
                        pass

        proj_q = Q()

        if proj_ids:
            proj_q |= Q(id__in=proj_ids)

        if proj_texts:
            proj_q |= Q(nombre__in=proj_texts) | Q(codigo__in=proj_texts)

        proyectos = (
            Proyecto.objects.filter(proj_q).only("id", "nombre", "codigo")
            if proj_q
            else Proyecto.objects.none()
        )

        by_id = {str(p.id): p.nombre for p in proyectos}
        by_code = {
            (p.codigo or "").strip().lower(): p.nombre
            for p in proyectos
            if getattr(p, "codigo", None)
        }
        by_name = {
            (p.nombre or "").strip().lower(): p.nombre
            for p in proyectos
            if getattr(p, "nombre", None)
        }

        for s in sessions:
            raw = str(getattr(s, "proyecto", "") or "").strip()
            raw_id = str(getattr(s, "proyecto_id", "") or "").strip()

            label = ""

            if raw:
                label = (
                    by_id.get(raw)
                    or by_code.get(raw.lower())
                    or by_name.get(raw.lower())
                    or raw
                )

            if not label and raw_id:
                label = (
                    by_id.get(raw_id)
                    or by_code.get(raw_id.lower())
                    or by_name.get(raw_id.lower())
                    or raw_id
                )

            s.proyecto_nombre = label
            s.project_label = label

        return sessions

    def excel_value_for_session(s, key):
        key = str(key)

        if key == "0":
            return (
                s.creado_en.strftime("%Y-%m-%d")
                if getattr(s, "creado_en", None)
                else "—"
            )

        if key == "1":
            return str(getattr(s, "proyecto_id", "") or "—")

        if key == "2":
            return str(getattr(s, "direccion_proyecto", "") or "—")

        if key == "3":
            return str(getattr(s, "semana_pago_proyectada", "") or "—")

        if key == "4":
            return status_label(s)

        if key == "5":
            return techs_label(s)

        if key == "6":
            return str(getattr(s, "cliente", "") or "—")

        if key == "7":
            return str(getattr(s, "ciudad", "") or "—")

        if key == "8":
            return str(
                getattr(s, "proyecto_nombre", "")
                or getattr(s, "project_label", "")
                or getattr(s, "proyecto", "")
                or "—"
            )

        if key == "9":
            return str(getattr(s, "oficina", "") or "—")

        if key == "10":
            return money_value(getattr(s, "subtotal_tecnico", None))

        if key == "11":
            return money_value(getattr(s, "subtotal_empresa", None))

        if key == "12":
            return money_value(getattr(s, "real_company_billing", None))

        if key == "13":
            return diff_label(s)

        if key == "14":
            return finance_status_label(s)

        if key == "15":
            return build_payweek_label(s)

        if key == "16":
            return comments_label(s)

        return "—"

    # ============================================================
    # Construcción de opciones
    # ============================================================
    excel_qs = qs_filtered.only(
        "id",
        "creado_en",
        "proyecto_id",
        "direccion_proyecto",
        "semana_pago_proyectada",
        "semana_pago_real",
        "discount_week",
        "estado",
        "is_direct_discount",
        "cliente",
        "ciudad",
        "proyecto",
        "oficina",
        "subtotal_tecnico",
        "subtotal_empresa",
        "real_company_billing",
        "finance_status",
        "finance_note",
    ).prefetch_related(
        Prefetch(
            "tecnicos_sesion",
            queryset=SesionBillingTecnico.objects.select_related("tecnico"),
        ),
        Prefetch(
            "pay_week_snapshots",
            queryset=BillingPayWeekSnapshot.objects.select_related(
                "tecnico",
                "item",
                "weekly_payment",
            )
            .filter(is_adjustment=False)
            .order_by(
                "tecnico__first_name",
                "tecnico__last_name",
                "tecnico__username",
                "tipo_trabajo",
                "codigo_trabajo",
                "id",
            ),
        ),
    )

    sessions = list(excel_qs)
    resolve_project_labels_for_sessions(sessions)

    excel_global = {str(i): set() for i in range(17)}

    for s in sessions:
        for key in excel_global.keys():
            val = excel_value_for_session(s, key)
            excel_global[key].add(str(val or "—"))

    data = {
        k: sorted(list(v), key=lambda x: x.lower())
        for k, v in excel_global.items()
    }

    return JsonResponse(
        {
            "ok": True,
            "excel_global": data,
        }
    )


@login_required
def listar_billing(request):
    """
    Visibilidad en Operaciones:
      - Descuento directo (is_direct_discount=True):
          mostrar SOLO si AÚN NO se ha enviado -> finance_sent_at IS NULL.
      - Resto:
          ocultar si finance_status ∈ {'sent','pending','paid','in_review'}.

    Mantiene filtros Excel del template:
      - NO construye excel_global_json pesado en la carga inicial.
      - Lee excel_filters desde GET.
      - Aplica filtros Excel antes de paginar solo cuando existen.
      - Mantiene paginación AJAX.
    """
    import json
    from datetime import date as _date
    from decimal import Decimal
    from urllib.parse import urlencode

    from django.core.paginator import Paginator
    from django.db.models import Prefetch, Q

    from core.permissions import filter_queryset_by_access
    from facturacion.models import Proyecto
    from operaciones.models import (BillingPayWeekSnapshot, ItemBilling,
                                    ItemBillingTecnico, SesionBilling,
                                    SesionBillingTecnico)

    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    user = request.user

    # ============================================================
    # Usuarios privilegiados
    # ============================================================
    can_view_legacy_history = user.is_superuser or getattr(
        user, "es_usuario_historial", False
    )

    # ============================================================
    # Visibilidad Operaciones
    # ============================================================
    visible_filter = (
        Q(is_direct_discount=True)
        & Q(finance_sent_at__isnull=True)
        & ~Q(finance_status="paid")
    ) | (
        Q(is_direct_discount=False)
        & ~Q(finance_status__in=["sent", "pending", "paid", "in_review"])
    )

    qs = SesionBilling.objects.filter(visible_filter).order_by("-creado_en")

    # ============================================================
    # Restricción por proyectos
    # ============================================================
    if not can_view_legacy_history:
        try:
            proyectos_user = filter_queryset_by_access(
                Proyecto.objects.all(),
                user,
                "id",
            )
        except Exception:
            proyectos_user = Proyecto.objects.none()

        proyectos_user_list = list(proyectos_user)

        if proyectos_user_list:
            allowed_keys = set()

            for p in proyectos_user_list:
                nombre = (getattr(p, "nombre", "") or "").strip()
                if nombre:
                    allowed_keys.add(nombre)

                codigo = getattr(p, "codigo", None)
                if codigo:
                    allowed_keys.add(str(codigo).strip())

                allowed_keys.add(str(p.id).strip())

            qs = qs.filter(proyecto__in=allowed_keys)

            asignaciones = []

            if ProyectoAsignacion is not None:
                try:
                    asignaciones = list(
                        ProyectoAsignacion.objects.filter(
                            usuario=user,
                            proyecto__in=proyectos_user_list,
                        ).select_related("proyecto")
                    )
                except Exception:
                    asignaciones = []

            if asignaciones:
                window_q = Q()
                has_window_q = False

                for a in asignaciones:
                    p = getattr(a, "proyecto", None)
                    if not p:
                        continue

                    keys = set()

                    nombre = (getattr(p, "nombre", "") or "").strip()
                    if nombre:
                        keys.add(nombre)

                    codigo = getattr(p, "codigo", None)
                    if codigo:
                        keys.add(str(codigo).strip())

                    keys.add(str(p.id).strip())

                    if not keys:
                        continue

                    if getattr(a, "include_history", False) or not getattr(
                        a, "start_at", None
                    ):
                        window_q |= Q(proyecto__in=keys)
                    else:
                        window_q |= Q(proyecto__in=keys, creado_en__gte=a.start_at)

                    has_window_q = True

                qs = qs.filter(window_q) if has_window_q else qs.none()
        else:
            qs = qs.none()

    # ============================================================
    # Filtros rápidos normales
    # ============================================================
    f = {
        "date": (request.GET.get("date") or "").strip(),
        "projid": (request.GET.get("projid") or "").strip(),
        "week": (request.GET.get("week") or "").strip(),
        "tech": (request.GET.get("tech") or "").strip(),
        "client": (request.GET.get("client") or "").strip(),
        "status": (request.GET.get("status") or "").strip(),
    }

    qs_filtered = qs

    if f["date"]:
        try:
            d = _date.fromisoformat(f["date"])
            qs_filtered = qs_filtered.filter(creado_en__date=d)
        except ValueError:
            pass

    if f["projid"]:
        qs_filtered = qs_filtered.filter(proyecto_id__icontains=f["projid"])

    if f["week"]:
        qs_filtered = qs_filtered.filter(
            Q(semana_pago_proyectada__icontains=f["week"])
            | Q(semana_pago_real__icontains=f["week"])
            | Q(discount_week__icontains=f["week"])
            | Q(pay_week_snapshots__semana_resultado__icontains=f["week"])
            | Q(pay_week_snapshots__semana_base__icontains=f["week"])
        )

    if f["tech"]:
        qs_filtered = qs_filtered.filter(
            Q(tecnicos_sesion__tecnico__first_name__icontains=f["tech"])
            | Q(tecnicos_sesion__tecnico__last_name__icontains=f["tech"])
            | Q(tecnicos_sesion__tecnico__username__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__first_name__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__last_name__icontains=f["tech"])
            | Q(pay_week_snapshots__tecnico__username__icontains=f["tech"])
        )

    if f["client"]:
        qs_filtered = qs_filtered.filter(cliente__icontains=f["client"])

    if f["status"]:
        status_txt = f["status"].lower().strip()

        if any(k in status_txt for k in ("direct", "descuento", "discount")):
            qs_filtered = qs_filtered.filter(is_direct_discount=True)
        else:
            mapping = [
                (
                    ("aprobado supervisor", "approved by supervisor"),
                    Q(estado="aprobado_supervisor"),
                ),
                (
                    ("rechazado supervisor", "rejected by supervisor"),
                    Q(estado="rechazado_supervisor"),
                ),
                (
                    ("en revision", "supervisor review", "in supervisor review"),
                    Q(estado="en_revision_supervisor"),
                ),
                (("finalizado", "finished"), Q(estado="finalizado")),
                (("en proceso", "in progress"), Q(estado="en_proceso")),
                (("asignado", "assigned"), Q(estado="asignado")),
                (("aprobado pm", "approved by pm"), Q(estado="aprobado_pm")),
                (("rechazado pm", "rejected by pm"), Q(estado="rechazado_pm")),
            ]

            applied = False

            for keys, cond in mapping:
                if any(k in status_txt for k in keys):
                    qs_filtered = qs_filtered.filter(cond)
                    applied = True
                    break

            if not applied:
                if "aprobado" in status_txt or "approved" in status_txt:
                    qs_filtered = qs_filtered.filter(
                        estado__in=["aprobado_supervisor", "aprobado_pm"]
                    )
                elif "rechazado" in status_txt or "rejected" in status_txt:
                    qs_filtered = qs_filtered.filter(
                        estado__in=["rechazado_supervisor", "rechazado_pm"]
                    )

    qs_filtered = qs_filtered.distinct()

    # ============================================================
    # Helpers generales
    # ============================================================
    def money_value(value):
        if value in (None, ""):
            return "—"

        try:
            return f"${Decimal(value):.2f}"
        except Exception:
            return str(value)

    def status_label(s):
        if getattr(s, "is_direct_discount", False):
            return "Direct discount"

        estado = getattr(s, "estado", "") or ""

        labels = {
            "aprobado_pm": "Approved by PM",
            "rechazado_pm": "Rejected by PM",
            "aprobado_supervisor": "Approved by supervisor",
            "rechazado_supervisor": "Rejected by supervisor",
            "en_revision_supervisor": "In supervisor review",
            "finalizado": "Finished (pending review)",
            "en_proceso": "In progress",
            "asignado": "Assigned",
        }

        return labels.get(estado, "Assigned")

    def finance_status_label(s):
        finance_status = getattr(s, "finance_status", "") or ""

        labels = {
            "sent": "Sent to Finance",
            "in_review": "In review",
            "rejected": "Rejected",
            "pending": "Pending payment",
            "paid": "Paid",
            "review_discount": "Review discount",
            "discount_applied": "Discount applied",
            "none": "—",
            "": "—",
            None: "—",
        }

        return labels.get(finance_status, "—")

    def diff_label(s):
        real = getattr(s, "real_company_billing", None)
        subtotal = getattr(s, "subtotal_empresa", None)

        if real in (None, "") or subtotal in (None, ""):
            return "—"

        try:
            real_d = Decimal(real)
            sub_d = Decimal(subtotal)
            diff = real_d - sub_d
        except Exception:
            return "—"

        if diff == 0:
            return "$0.00"

        if diff < 0:
            return f"- ${abs(diff):.2f}"

        return f"+ ${diff:.2f}"

    def techs_label(sesion):
        vals = []

        try:
            for st in sesion.tecnicos_sesion.all():
                if not getattr(st, "tecnico", None):
                    continue

                vals.append(st.tecnico.get_full_name() or st.tecnico.username)
        except Exception:
            pass

        return ", ".join(v for v in vals if v) or "—"

    def legacy_paid_flag(s):
        note = getattr(s, "finance_note", "") or ""

        try:
            tech_ids = list(
                s.tecnicos_sesion.all().values_list("tecnico_id", flat=True)
            )
        except Exception:
            tech_ids = []

        possible_weeks = [
            (getattr(s, "semana_pago_real", "") or "").strip().upper(),
            (getattr(s, "semana_pago_proyectada", "") or "").strip().upper(),
            (getattr(s, "discount_week", "") or "").strip().upper(),
        ]
        possible_weeks = [w for w in possible_weeks if w]

        for tech_id in tech_ids:
            for wk in possible_weeks:
                marker = f"[TECH_WEEKLY_PAYMENT_PAID:{tech_id}:{wk}]"
                if marker in note:
                    return True

        return False

    def build_payweek_groups(s):
        groups_map = {}

        snaps = (
            list(getattr(s, "pay_week_snapshots", []).all())
            if hasattr(s, "pay_week_snapshots")
            else []
        )

        if snaps:
            for snap in snaps:
                tech_name = (
                    snap.tecnico.get_full_name().strip()
                    if getattr(snap, "tecnico", None) and snap.tecnico.get_full_name()
                    else getattr(snap.tecnico, "username", "")
                    or f"User {snap.tecnico_id}"
                )

                grp = groups_map.setdefault(
                    tech_name,
                    {
                        "tech_name": tech_name,
                        "weeks_summary": "",
                        "lines": [],
                    },
                )

                work_type = (
                    (snap.tipo_trabajo or "").strip()
                    or (getattr(snap.item, "tipo_trabajo", "") or "").strip()
                    or "Legacy"
                )

                week = (
                    (getattr(snap, "semana_resultado", "") or "").strip()
                    or (getattr(snap, "semana_base", "") or "").strip()
                    or (getattr(s, "semana_pago_real", "") or "").strip()
                    or (getattr(s, "discount_week", "") or "").strip()
                    or (getattr(s, "semana_pago_proyectada", "") or "").strip()
                    or "—"
                )

                is_paid_line = (
                    getattr(snap, "payment_status", "") == "paid"
                    or bool(getattr(snap, "paid_at", None))
                    or (
                        getattr(snap, "weekly_payment", None)
                        and getattr(snap.weekly_payment, "status", "") == "paid"
                    )
                    or getattr(s, "finance_status", "") == "paid"
                )

                grp["lines"].append(
                    {
                        "work_type": work_type,
                        "codigo_trabajo": (snap.codigo_trabajo or "").strip(),
                        "week": week,
                        "is_legacy": False,
                        "snapshot_id": snap.id,
                        "is_paid": is_paid_line,
                    }
                )

            groups = list(groups_map.values())

            for grp in groups:
                weeks = []

                for line in grp["lines"]:
                    wk = (line.get("week") or "").strip()

                    if wk and wk not in weeks:
                        weeks.append(wk)

                grp["weeks_summary"] = ", ".join(weeks) if weeks else "—"

            return groups

        asignaciones = (
            list(s.tecnicos_sesion.all()) if hasattr(s, "tecnicos_sesion") else []
        )

        base_week = (
            (getattr(s, "semana_pago_real", "") or "").strip()
            or (getattr(s, "discount_week", "") or "").strip()
            or (getattr(s, "semana_pago_proyectada", "") or "").strip()
            or "—"
        )

        legacy_is_paid = legacy_paid_flag(s) or (
            getattr(s, "finance_status", "") == "paid"
        )

        tech_names = []

        for asig in asignaciones:
            tech_name = (
                asig.tecnico.get_full_name().strip()
                if getattr(asig, "tecnico", None) and asig.tecnico.get_full_name()
                else getattr(asig.tecnico, "username", "") or f"User {asig.tecnico_id}"
            )

            if tech_name and tech_name not in tech_names:
                tech_names.append(tech_name)

        tech_label = ", ".join(tech_names) if tech_names else "—"

        return [
            {
                "tech_name": tech_label,
                "weeks_summary": base_week,
                "lines": [
                    {
                        "work_type": "Legacy",
                        "codigo_trabajo": "",
                        "week": base_week,
                        "is_legacy": True,
                        "session_id": s.id,
                        "dom_id": f"{s.id}-legacy",
                        "is_paid": legacy_is_paid,
                    }
                ],
            }
        ]

    def payweek_snapshot_label(sesion, groups=None):
        groups = groups if groups is not None else build_payweek_groups(sesion)

        if not groups:
            return str(getattr(sesion, "semana_pago_real", "") or "—")

        rows = []

        for grp in groups:
            tech_name = grp.get("tech_name") or "—"

            for line in grp.get("lines", []):
                work_type = (line.get("work_type") or "").strip() or "Work type"
                week = (line.get("week") or "").strip() or "—"
                suffix = " [Paid]" if line.get("is_paid") else ""
                rows.append(f"{tech_name} — {work_type} → {week}{suffix}")

        return (
            " | ".join(rows)
            if rows
            else str(getattr(sesion, "semana_pago_real", "") or "—")
        )

    def comments_label(s):
        vals = []

        try:
            for a in s.tecnicos_sesion.all():
                txt = (getattr(a, "tecnico_comentario", "") or "").strip()

                if not txt:
                    continue

                tech_name = (
                    a.tecnico.get_full_name() if getattr(a, "tecnico", None) else ""
                ) or (
                    getattr(a.tecnico, "username", "")
                    if getattr(a, "tecnico", None)
                    else ""
                )

                if tech_name:
                    vals.append(f"{tech_name}: {txt}")
                else:
                    vals.append(txt)
        except Exception:
            pass

        return " | ".join(vals) if vals else "—"

    def resolve_project_labels_for_sessions(sessions):
        proj_ids = set()
        proj_texts = set()

        for s in sessions:
            raw_proyecto = getattr(s, "proyecto", None)
            if raw_proyecto not in (None, "", "-"):
                txt = str(raw_proyecto).strip()
                if txt:
                    proj_texts.add(txt)
                    try:
                        proj_ids.add(int(txt))
                    except Exception:
                        pass

            raw_proyecto_id = getattr(s, "proyecto_id", None)
            if raw_proyecto_id not in (None, "", "-"):
                txt2 = str(raw_proyecto_id).strip()
                if txt2:
                    proj_texts.add(txt2)
                    try:
                        proj_ids.add(int(txt2))
                    except Exception:
                        pass

        proj_q = Q()

        if proj_ids:
            proj_q |= Q(id__in=proj_ids)

        if proj_texts:
            proj_q |= Q(nombre__in=proj_texts) | Q(codigo__in=proj_texts)

        proyectos = (
            Proyecto.objects.filter(proj_q).only("id", "nombre", "codigo")
            if proj_q
            else Proyecto.objects.none()
        )

        by_id = {str(p.id): p.nombre for p in proyectos}
        by_code = {
            (p.codigo or "").strip().lower(): p.nombre
            for p in proyectos
            if getattr(p, "codigo", None)
        }
        by_name = {
            (p.nombre or "").strip().lower(): p.nombre
            for p in proyectos
            if getattr(p, "nombre", None)
        }

        for s in sessions:
            raw = str(getattr(s, "proyecto", "") or "").strip()
            raw_id = str(getattr(s, "proyecto_id", "") or "").strip()

            label = ""

            if raw:
                label = (
                    by_id.get(raw)
                    or by_code.get(raw.lower())
                    or by_name.get(raw.lower())
                    or raw
                )

            if not label and raw_id:
                label = (
                    by_id.get(raw_id)
                    or by_code.get(raw_id.lower())
                    or by_name.get(raw_id.lower())
                    or raw_id
                )

            s.proyecto_nombre = label
            s.project_label = label

        return sessions

    def excel_value_for_session(s, key):
        key = str(key)

        if key == "0":
            return (
                s.creado_en.strftime("%Y-%m-%d")
                if getattr(s, "creado_en", None)
                else "—"
            )

        if key == "1":
            return str(getattr(s, "proyecto_id", "") or "—")

        if key == "2":
            return str(getattr(s, "direccion_proyecto", "") or "—")

        if key == "3":
            return str(getattr(s, "semana_pago_proyectada", "") or "—")

        if key == "4":
            return status_label(s)

        if key == "5":
            return techs_label(s)

        if key == "6":
            return str(getattr(s, "cliente", "") or "—")

        if key == "7":
            return str(getattr(s, "ciudad", "") or "—")

        if key == "8":
            return str(
                getattr(s, "proyecto_nombre", "")
                or getattr(s, "project_label", "")
                or getattr(s, "proyecto", "")
                or "—"
            )

        if key == "9":
            return str(getattr(s, "oficina", "") or "—")

        if key == "10":
            return money_value(getattr(s, "subtotal_tecnico", None))

        if key == "11":
            return money_value(getattr(s, "subtotal_empresa", None))

        if key == "12":
            return money_value(getattr(s, "real_company_billing", None))

        if key == "13":
            return diff_label(s)

        if key == "14":
            return finance_status_label(s)

        if key == "15":
            groups = getattr(s, "payweek_groups", None)
            return payweek_snapshot_label(s, groups)

        if key == "16":
            return comments_label(s)

        return "—"

    # ============================================================
    # Filtros Excel
    # ============================================================
    excel_filters_raw = (request.GET.get("excel_filters") or "").strip()

    try:
        parsed_excel_filters = (
            json.loads(excel_filters_raw) if excel_filters_raw else {}
        )
    except Exception:
        parsed_excel_filters = {}

    excel_filters = {}

    if isinstance(parsed_excel_filters, dict):
        for k, vals in parsed_excel_filters.items():
            if isinstance(vals, list):
                clean_vals = set(str(v) for v in vals if str(v).strip() != "")
                if clean_vals:
                    excel_filters[str(k)] = clean_vals

    # Carga inicial liviana:
    # los valores globales de filtros Excel se cargan después por AJAX.
    excel_global_json = "{}"

    # Solo hacemos el cálculo pesado si YA hay filtros Excel aplicados.
    # Esto permite que el filtro siga funcionando al paginar o recargar,
    # pero evita cargar todo en la primera apertura de la vista.
    if excel_filters:
        excel_qs = qs_filtered.only(
            "id",
            "creado_en",
            "proyecto_id",
            "direccion_proyecto",
            "semana_pago_proyectada",
            "semana_pago_real",
            "discount_week",
            "estado",
            "is_direct_discount",
            "cliente",
            "ciudad",
            "proyecto",
            "oficina",
            "subtotal_tecnico",
            "subtotal_empresa",
            "real_company_billing",
            "finance_status",
            "finance_note",
        ).prefetch_related(
            Prefetch(
                "tecnicos_sesion",
                queryset=SesionBillingTecnico.objects.select_related("tecnico"),
            ),
            Prefetch(
                "pay_week_snapshots",
                queryset=BillingPayWeekSnapshot.objects.select_related(
                    "tecnico",
                    "item",
                    "weekly_payment",
                )
                .filter(is_adjustment=False)
                .order_by(
                    "tecnico__first_name",
                    "tecnico__last_name",
                    "tecnico__username",
                    "tipo_trabajo",
                    "codigo_trabajo",
                    "id",
                ),
            ),
        )

        excel_sessions = list(excel_qs)
        resolve_project_labels_for_sessions(excel_sessions)

        for s in excel_sessions:
            s.payweek_groups = build_payweek_groups(s)
            s.payweek_snapshot_label = payweek_snapshot_label(s, s.payweek_groups)
            s.techs_label = techs_label(s)

        allowed_ids = []

        for s in excel_sessions:
            keep = True

            for key, allowed_values in excel_filters.items():
                current_value = str(excel_value_for_session(s, key) or "—")

                if current_value not in allowed_values:
                    keep = False
                    break

            if keep:
                allowed_ids.append(s.id)

        qs_filtered = qs_filtered.filter(id__in=allowed_ids)

    qs_filtered = qs_filtered.distinct()

    # ============================================================
    # Query liviana para paginar
    # ============================================================
    light_qs = qs_filtered.only(
        "id",
        "creado_en",
        "proyecto_id",
        "direccion_proyecto",
        "semana_pago_proyectada",
        "semana_pago_real",
        "discount_week",
        "estado",
        "is_direct_discount",
        "cliente",
        "ciudad",
        "proyecto",
        "oficina",
        "subtotal_tecnico",
        "subtotal_empresa",
        "real_company_billing",
        "finance_status",
        "finance_note",
    ).order_by("-creado_en")

    # ============================================================
    # Paginación
    # ============================================================
    cantidad = request.GET.get("cantidad", "10")

    try:
        per_page = int(cantidad)
    except (TypeError, ValueError):
        per_page = 10

    if per_page < 5:
        per_page = 5

    if per_page > 50:
        per_page = 50

    cantidad = str(per_page)

    paginator = Paginator(light_qs, per_page)
    pagina_light = paginator.get_page(request.GET.get("page"))

    page_ids = [s.id for s in pagina_light.object_list]
    order_map = {pk: idx for idx, pk in enumerate(page_ids)}

    # ============================================================
    # Cargar relaciones pesadas solo para página visible.
    # IMPORTANTE: no se cargan evidencias/fotos aquí.
    # ============================================================
    page_rows = list(
        SesionBilling.objects.filter(id__in=page_ids).prefetch_related(
            Prefetch(
                "items",
                queryset=ItemBilling.objects.prefetch_related(
                    Prefetch(
                        "desglose_tecnico",
                        queryset=ItemBillingTecnico.objects.select_related("tecnico"),
                    )
                ),
            ),
            Prefetch(
                "tecnicos_sesion",
                queryset=SesionBillingTecnico.objects.select_related("tecnico"),
            ),
            Prefetch(
                "pay_week_snapshots",
                queryset=BillingPayWeekSnapshot.objects.select_related(
                    "tecnico",
                    "item",
                    "weekly_payment",
                )
                .filter(is_adjustment=False)
                .order_by(
                    "tecnico__first_name",
                    "tecnico__last_name",
                    "tecnico__username",
                    "tipo_trabajo",
                    "codigo_trabajo",
                    "id",
                ),
            ),
        )
    )

    page_rows.sort(key=lambda s: order_map.get(s.id, 999999))
    resolve_project_labels_for_sessions(page_rows)

    # ============================================================
    # Extras para template SOLO página visible
    # ============================================================
    _attach_accounting_lock_flags_to_sessions(page_rows)

    for s in page_rows:
        comentarios = []

        try:
            for a in s.tecnicos_sesion.all():
                txt = (getattr(a, "tecnico_comentario", "") or "").strip()

                if txt:
                    comentarios.append(a)
        except Exception:
            comentarios = []

        s.comentarios_tecnicos = comentarios
        s.payweek_groups = build_payweek_groups(s)
        s.payweek_snapshot_label = payweek_snapshot_label(s, s.payweek_groups)
        s.techs_label = techs_label(s)

    pagina_light.object_list = page_rows
    pagina = pagina_light

    # ============================================================
    # Permisos
    # ============================================================
    access_ctx = _billing_access_context(request.user)

    can_edit_real_week = access_ctx["can_edit_real_week"]

    can_edit_items = access_ctx["can_edit_items"]

    # ============================================================
    # Mantener filtros al paginar / cambiar cantidad
    # ============================================================
    keep_params = {}

    if f["date"]:
        keep_params["date"] = f["date"]

    if f["projid"]:
        keep_params["projid"] = f["projid"]

    if f["week"]:
        keep_params["week"] = f["week"]

    if f["tech"]:
        keep_params["tech"] = f["tech"]

    if f["client"]:
        keep_params["client"] = f["client"]

    if f["status"]:
        keep_params["status"] = f["status"]

    if excel_filters_raw:
        keep_params["excel_filters"] = excel_filters_raw

    if cantidad:
        keep_params["cantidad"] = cantidad

    qs_keep = urlencode(keep_params)

    context = {
        "pagina": pagina,
        "cantidad": cantidad,
        "can_edit_real_week": can_edit_real_week,
        "can_edit_items": can_edit_items,
        "f": f,
        "qs_keep": qs_keep,
        "excel_global_json": excel_global_json,
    }

    context.update(access_ctx)

    return render(
        request,
        "operaciones/billing_listar.html",
        context,
    )


@login_required
@rol_requerido("admin", "pm", "facturacion")
@csrf_protect
@require_POST
def billing_update_snapshot_week(request, snapshot_id: int):
    """
    Actualiza la semana_resultado de TODOS los snapshots
      de la misma sesión
    para el mismo técnico + work type.

    Reglas:
    - Si hay pagos PAID relacionados, SOLO admin puede modificar.
    - Excluye snapshots de ajuste si el modelo soporta esos campos.
    - NO modifica semana_base aquí; solo semana_resultado.
    - Recalcula semana_pago_real resumen de la sesión.
    - Re-sincroniza WeeklyPayment para semanas viejas y nuevas.
    - Responde SIEMPRE JSON.
    """
    try:
        snap = get_object_or_404(
            BillingPayWeekSnapshot.objects.select_related("sesion"),
            pk=snapshot_id,
        )

        semana = (request.POST.get("semana") or "").strip().upper()
        if not semana:
            return JsonResponse({"ok": False, "error": "Week is required."}, status=400)

        if not WEEK_RE.match(semana):
            return JsonResponse(
                {"ok": False, "error": "Invalid format. Use YYYY-W##"},
                status=400,
            )

        year, week = _parse_iso_week(semana)
        if not year or not week:
            return JsonResponse(
                {"ok": False, "error": "Invalid ISO week."},
                status=400,
            )

        sesion = snap.sesion

        if not access_user_can(request.user, "billing.edit_real_week"):
            return JsonResponse(
                {
                    "ok": False,
                    "error": "FORBIDDEN",
                    "message": "You do not have permission to edit the real pay week.",
                },
                status=403,
            )

        is_admin = _is_admin(request.user)

        # 🔒 Si hay pagos PAID relacionados, solo admin puede mover semanas
        if _session_is_paid_locked(sesion) and not is_admin:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "LOCKED_PAID",
                    "message": "This session has PAID weekly payments. Only admins can change the real pay week.",
                },
                status=403,
            )

        tecnico_id = snap.tecnico_id
        tipo_trabajo = (getattr(snap, "tipo_trabajo", "") or "").strip()

        snapshot_fields = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
        has_is_adjustment = "is_adjustment" in snapshot_fields
        has_adjustment_of = "adjustment_of" in snapshot_fields

        # Semanas antiguas afectadas por este grupo
        old_weeks = set(
            BillingPayWeekSnapshot.objects.filter(
                sesion=sesion,
                tecnico_id=tecnico_id,
                tipo_trabajo=tipo_trabajo,
            )
            .exclude(semana_resultado__isnull=True)
            .exclude(semana_resultado__exact="")
            .values_list("semana_resultado", flat=True)
        )

        qs = BillingPayWeekSnapshot.objects.filter(
            sesion=sesion,
            tecnico_id=tecnico_id,
            tipo_trabajo=tipo_trabajo,
        )

        if has_is_adjustment:
            qs = qs.filter(is_adjustment=False)
        elif has_adjustment_of:
            qs = qs.filter(adjustment_of__isnull=True)

        snaps = list(qs)

        changed_weeks = set()
        updated_count = 0

        with transaction.atomic():
            for obj in snaps:
                update_fields = []

                if (obj.semana_resultado or "") != semana:
                    obj.semana_resultado = semana
                    update_fields.append("semana_resultado")

                # ✅ IMPORTANTE:
                # NO tocar semana_base aquí.
                # semana_base representa la proyectada/origen, no la real editada línea a línea.

                if update_fields:
                    obj.save(update_fields=update_fields)
                    updated_count += 1

                if semana:
                    changed_weeks.add(semana)

            # Recalcular semana_pago_real resumen de la sesión
            snaps_all = BillingPayWeekSnapshot.objects.filter(sesion=sesion)

            if has_is_adjustment:
                snaps_all = snaps_all.filter(is_adjustment=False)
            elif has_adjustment_of:
                snaps_all = snaps_all.filter(adjustment_of__isnull=True)

            weeks_all = list(
                snaps_all.exclude(semana_resultado__isnull=True)
                .exclude(semana_resultado__exact="")
                .values_list("semana_resultado", flat=True)
            )

            summary_week = ""
            parsed = []

            for w in weeks_all:
                y, wk = _parse_iso_week(w)
                if y and wk:
                    parsed.append((y, wk, w))

            if parsed:
                parsed.sort(key=lambda x: (x[0], x[1]))
                summary_week = parsed[-1][2]

            update_fields = ["semana_pago_real"]
            sesion.semana_pago_real = summary_week

            if hasattr(sesion, "updated_at"):
                update_fields.append("updated_at")

            sesion.save(update_fields=update_fields)

        # Re-sincronizar semanas afectadas
        try:
            weeks_to_sync = set(filter(None, old_weeks | changed_weeks))
            for wk in weeks_to_sync:
                _sync_weekly_totals(week=wk)
        except Exception:
            pass

        return JsonResponse(
            {
                "ok": True,
                "semana": semana,
                "summary_week": summary_week,
                "updated_count": updated_count,
            }
        )

    except Exception as e:
        return JsonResponse(
            {
                "ok": False,
                "error": "SERVER_ERROR",
                "message": str(e),
            },
            status=500,
        )


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion")
@require_POST
@transaction.atomic
def billing_item_update_qty(request, item_id: int):
    """
    Actualiza Quantity en línea desde Billing List.

    Permitido para:
      - admin
      - pm
      - supervisor
      - facturacion

    Recalcula:
      - ItemBilling.cantidad
      - ItemBilling.subtotal_empresa
      - ItemBilling.subtotal_tecnico
      - ItemBillingTecnico.tarifa_efectiva
      - ItemBillingTecnico.subtotal
      - SesionBilling.subtotal_empresa
      - SesionBilling.subtotal_tecnico
      - BillingPayWeekSnapshot

    Bloquea:
      - sesiones con líneas ya pagadas
      - cantidades negativas salvo direct discount
    """

    if not access_user_can(request.user, "billing.edit_billing"):
        return JsonResponse(
            {
                "ok": False,
                "error": "FORBIDDEN",
                "message": "You do not have permission to edit billing quantities.",
            },
            status=403,
        )

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except json.JSONDecodeError:
        return JsonResponse(
            {
                "ok": False,
                "error": "INVALID_JSON",
                "message": "Invalid payload.",
            },
            status=400,
        )

    cantidad_raw = payload.get("cantidad", None)

    if cantidad_raw in (None, ""):
        return JsonResponse(
            {
                "ok": False,
                "error": "MISSING_QUANTITY",
                "message": "Missing quantity.",
            },
            status=400,
        )

    try:
        cantidad = Decimal(str(cantidad_raw)).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError, TypeError):
        return JsonResponse(
            {
                "ok": False,
                "error": "INVALID_QUANTITY",
                "message": "Invalid quantity.",
            },
            status=400,
        )

    try:
        item = (
            ItemBilling.objects.select_related("sesion")
            .prefetch_related("desglose_tecnico")
            .select_for_update()
            .get(pk=item_id)
        )
    except ItemBilling.DoesNotExist:
        return JsonResponse(
            {
                "ok": False,
                "error": "ITEM_NOT_FOUND",
                "message": "Item does not exist.",
            },
            status=404,
        )

    sesion = item.sesion

    # Seguridad por proyecto visible
    proyectos_visibles = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        "id",
    )

    proyecto_pk = _resolve_proyecto_pk_from_sesion(
        sesion,
        proyectos_qs=proyectos_visibles,
    )

    if not proyecto_pk:
        return JsonResponse(
            {
                "ok": False,
                "error": "FORBIDDEN_PROJECT",
                "message": "You do not have access to this billing project.",
            },
            status=403,
        )

    # Bloqueo contable
    if _session_is_paid_locked(sesion):
        return JsonResponse(
            {
                "ok": False,
                "error": "LOCKED_PAID",
                "message": (
                    "This billing cannot be modified because at least one "
                    "worker/work type line has already been marked as paid."
                ),
            },
            status=403,
        )

    is_direct_discount = bool(getattr(sesion, "is_direct_discount", False))

    if is_direct_discount:
        if cantidad > 0:
            cantidad = -cantidad
    else:
        if cantidad < 0:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "INVALID_QUANTITY",
                    "message": "Quantity cannot be negative unless this is a direct discount.",
                },
                status=400,
            )

    def _money(v) -> Decimal:
        try:
            return Decimal(str(v or "0")).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError, TypeError):
            return Decimal("0.00")

    payment_mode = (getattr(sesion, "tech_payment_mode", "") or "split").strip().lower()

    if payment_mode not in ("split", "full"):
        payment_mode = "split"

    subtotal_empresa = _money((item.precio_empresa or Decimal("0.00")) * cantidad)

    subtotal_tecnico = Decimal("0.00")

    for bd in item.desglose_tecnico.all():
        base = _money(getattr(bd, "tarifa_base", 0) or 0)
        pct = _money(getattr(bd, "porcentaje", 0) or 0)

        if payment_mode == "full":
            tarifa_efectiva = _money(base)
        else:
            tarifa_efectiva = _money(base * (pct / Decimal("100")))

        sub = _money(tarifa_efectiva * cantidad)

        ItemBillingTecnico.objects.filter(pk=bd.pk).update(
            tarifa_efectiva=tarifa_efectiva,
            subtotal=sub,
        )

        subtotal_tecnico += sub

    ItemBilling.objects.filter(pk=item.pk).update(
        cantidad=cantidad,
        subtotal_empresa=subtotal_empresa,
        subtotal_tecnico=subtotal_tecnico,
    )

    items_qs = ItemBilling.objects.filter(sesion=sesion).only(
        "subtotal_tecnico",
        "subtotal_empresa",
    )

    total_tecnico = items_qs.aggregate(s=Sum("subtotal_tecnico"))["s"] or Decimal(
        "0.00"
    )
    total_empresa = items_qs.aggregate(s=Sum("subtotal_empresa"))["s"] or Decimal(
        "0.00"
    )

    SesionBilling.objects.filter(pk=sesion.pk).update(
        subtotal_tecnico=_money(total_tecnico),
        subtotal_empresa=_money(total_empresa),
    )

    sesion_refrescada = SesionBilling.objects.get(pk=sesion.pk)

    snapshot_info = rebuild_billing_payweek_snapshot(sesion_refrescada)

    diff_text = "—"

    if sesion_refrescada.real_company_billing is not None:
        diff = _money(sesion_refrescada.real_company_billing) - _money(
            sesion_refrescada.subtotal_empresa
        )

        if diff < 0:
            diff_text = (
                f"<span class='font-semibold text-red-600'>"
                f"- ${abs(diff):.2f}</span>"
            )
        elif diff > 0:
            diff_text = (
                f"<span class='font-semibold text-green-600'>" f"+ ${diff:.2f}</span>"
            )
        else:
            diff_text = "<span class='text-gray-700'>$0.00</span>"

    return JsonResponse(
        {
            "ok": True,
            "item_id": item.pk,
            "cantidad": float(cantidad),
            "subtotal_tecnico": float(subtotal_tecnico),
            "subtotal_empresa": float(subtotal_empresa),
            "parent": {
                "id": sesion_refrescada.pk,
                "subtotal_tecnico": float(sesion_refrescada.subtotal_tecnico or 0),
                "subtotal_empresa": float(sesion_refrescada.subtotal_empresa or 0),
                "real_company_billing": (
                    float(sesion_refrescada.real_company_billing)
                    if sesion_refrescada.real_company_billing is not None
                    else None
                ),
                "semana_pago_real": snapshot_info.get("summary_week", "") or "",
                "diferencia_text": diff_text,
            },
        }
    )


# ===== Crear / Editar =====


@login_required
def crear_billing(request):
    if not access_user_can(request.user, "billing.create_billing"):
        messages.error(request, "You do not have permission to create billings.")
        return redirect("operaciones:listar_billing")

    if request.method == "POST":
        return _guardar_billing(request)

    clientes = (
        PrecioActividadTecnico.objects.values_list("cliente", flat=True)
        .distinct()
        .order_by("cliente")
    )

    proyectos_visibles = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        "id",
    )

    if proyectos_visibles.exists():
        tecnicos = (
            Usuario.objects.filter(
                is_active=True,
                precioactividadtecnico__isnull=False,
                precioactividadtecnico__proyecto_id__in=proyectos_visibles.values_list(
                    "id", flat=True
                ),
            )
            .distinct()
            .order_by("first_name", "last_name", "username")
        )
    else:
        tecnicos = Usuario.objects.none()

    context = {
        "is_edit": False,
        "sesion": None,
        "clientes": list(clientes),
        "tecnicos": tecnicos,
        "items": [],
        "ids_tecnicos": [],
        "proyecto_value": "",
        "proyecto_label": "",
    }

    context.update(_billing_access_context(request.user))

    return render(
        request,
        "operaciones/billing_editar.html",
        context,
    )


@login_required
def editar_billing(request, sesion_id: int):
    if not access_user_can(request.user, "billing.edit_billing"):
        messages.error(request, "You do not have permission to edit billings.")
        return redirect("operaciones:listar_billing")

    sesion = get_object_or_404(SesionBilling, pk=sesion_id)

    if _session_is_paid_locked(sesion):
        messages.error(
            request,
            (
                "This billing cannot be edited because at least one worker/work type "
                "line has already been marked as paid."
            ),
        )
        return redirect("operaciones:listar_billing")

    if request.method == "POST":
        return _guardar_billing(request, sesion=sesion)

    clientes = (
        PrecioActividadTecnico.objects.values_list("cliente", flat=True)
        .distinct()
        .order_by("cliente")
    )

    tecnicos = (
        Usuario.objects.filter(precioactividadtecnico__isnull=False, is_active=True)
        .distinct()
        .order_by("first_name", "last_name", "username")
    )

    items = sesion.items.prefetch_related("desglose_tecnico__tecnico").order_by("id")

    ids_tecnicos = list(
        sesion.tecnicos_sesion.filter(is_active=True).values_list(
            "tecnico_id", flat=True
        )
    )

    proyectos_qs = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        "id",
    )

    proyecto_sel = None

    raw_label = (
        getattr(sesion, "proyecto", None) or getattr(sesion, "proyecto_id", None) or ""
    )
    raw_label = str(raw_label).strip()

    proyecto_value = raw_label
    proyecto_label = raw_label

    raw = (getattr(sesion, "proyecto", "") or "").strip()

    if raw:
        try:
            pid = int(raw)
        except (TypeError, ValueError):
            proyecto_sel = proyectos_qs.filter(
                Q(nombre__iexact=raw) | Q(codigo__iexact=raw)
            ).first()
        else:
            proyecto_sel = proyectos_qs.filter(pk=pid).first()

    if not proyecto_sel and getattr(sesion, "proyecto_id", None):
        code = str(sesion.proyecto_id).strip()

        if code:
            proyecto_sel = proyectos_qs.filter(
                Q(codigo__iexact=code)
                | Q(nombre__iexact=code)
                | Q(nombre__icontains=code)
            ).first()

            if not proyecto_sel:
                try:
                    pid2 = int(code)
                except (TypeError, ValueError):
                    pid2 = None

                if pid2 is not None:
                    proyecto_sel = proyectos_qs.filter(pk=pid2).first()

    if proyecto_sel:
        proyecto_value = str(proyecto_sel.id)
        proyecto_label = (
            getattr(proyecto_sel, "nombre", "") or str(proyecto_sel)
        ).strip()

    context = {
        "is_edit": True,
        "sesion": sesion,
        "clientes": list(clientes),
        "tecnicos": tecnicos,
        "items": items,
        "ids_tecnicos": ids_tecnicos,
        "proyectos": proyectos_qs,
        "proyecto_sel": proyecto_sel,
        "proyecto_value": proyecto_value,
        "proyecto_label": proyecto_label,
    }

    context.update(_billing_access_context(request.user))

    return render(
        request,
        "operaciones/billing_editar.html",
        context,
    )


def _money(v) -> Decimal:
    try:
        return Decimal(str(v or "0")).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


def _tarifa_tecnico_lookup(tid, sesion, codigo) -> Decimal:
    """
    Busca la tarifa técnica usando SIEMPRE el PK real de Proyecto.
    """
    proyecto_pk = _resolve_proyecto_pk_from_sesion(sesion)
    if not proyecto_pk:
        return _money(0)

    val = (
        PrecioActividadTecnico.objects
        .filter(
            tecnico_id=tid,
            cliente=sesion.cliente,
            ciudad=sesion.ciudad,
            proyecto_id=proyecto_pk,
            oficina=sesion.oficina,
            codigo_trabajo=codigo,
        )
        .values_list("precio_tecnico", flat=True)
        .first()
    )

    if val is None:
        val = (
            PrecioActividadTecnico.objects
            .filter(
                cliente=sesion.cliente,
                ciudad=sesion.ciudad,
                proyecto_id=proyecto_pk,
                oficina=sesion.oficina,
                codigo_trabajo=codigo,
            )
            .values_list("precio_tecnico", flat=True)
            .first()
        )

    return _money(val or 0)


@transaction.atomic
def _recalcular_items_sesion(sesion):
    """
    ✅ Recalcula:
      - porcentajes ya vienen actualizados en SesionBillingTecnico
      - asegura ItemBillingTecnico para todos
      - recalcula tarifa_base, tarifa_efectiva, subtotal por técnico (según payment mode)
      - recalcula subtotal_tecnico/subtotal_empresa por item
      - recalcula subtotal_tecnico/subtotal_empresa por sesión
    ⚠️ No toca evidencias / requisitos.
    ⚠️ No toca estado (usamos update()).
    """
    tecnicos = list(
        SesionBillingTecnico.objects.filter(sesion=sesion)
        .values_list("tecnico_id", "porcentaje")
        .order_by("id")
    )
    if not tecnicos:
        return

    def _money(v) -> Decimal:
        try:
            return Decimal(str(v or "0")).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError, TypeError):
            return Decimal("0.00")

    pct_by_tid = {tid: _money(pct) for tid, pct in tecnicos}
    tech_ids = [tid for tid, _ in tecnicos]

    payment_mode = (getattr(sesion, "tech_payment_mode", "") or "split").strip().lower()
    if payment_mode not in ("split", "full"):
        payment_mode = "split"

    items = list(
        ItemBilling.objects.filter(sesion=sesion)
        .prefetch_related("desglose_tecnico")
        .order_by("id")
    )

    total_emp = Decimal("0.00")
    total_tec = Decimal("0.00")

    for it in items:
        qty = _money(it.cantidad)

        sub_emp = _money((it.precio_empresa or Decimal("0")) * qty)

        existing = {bd.tecnico_id: bd for bd in it.desglose_tecnico.all()}
        for tid in tech_ids:
            if tid not in existing:
                ItemBillingTecnico.objects.create(
                    item=it,
                    tecnico_id=tid,
                    tarifa_base=Decimal("0.00"),
                    porcentaje=Decimal("0.00"),
                    tarifa_efectiva=Decimal("0.00"),
                    subtotal=Decimal("0.00"),
                )

        desglose = list(ItemBillingTecnico.objects.filter(item=it).order_by("id"))

        sub_tecs = Decimal("0.00")

        for bd in desglose:
            tid = bd.tecnico_id
            pct = pct_by_tid.get(tid, Decimal("0.00"))

            base = _tarifa_tecnico_lookup(tid, sesion, it.codigo_trabajo)

            if payment_mode == "full":
                efectiva = _money(base)
            else:
                efectiva = _money(base * (pct / Decimal("100")))

            subtotal = _money(efectiva * qty)

            ItemBillingTecnico.objects.filter(pk=bd.pk).update(
                tarifa_base=base,
                porcentaje=pct,
                tarifa_efectiva=efectiva,
                subtotal=subtotal,
            )
            sub_tecs += subtotal

        ItemBilling.objects.filter(pk=it.pk).update(
            subtotal_empresa=sub_emp,
            subtotal_tecnico=sub_tecs,
        )

        total_emp += sub_emp
        total_tec += sub_tecs

    SesionBilling.objects.filter(pk=sesion.pk).update(
        subtotal_empresa=_money(total_emp),
        subtotal_tecnico=_money(total_tec),
    )


@login_required
@rol_requerido("admin", "pm", "supervisor")
@require_GET
def billing_tecnicos_disponibles(request, sesion_id: int):
    sesion = get_object_or_404(SesionBilling, pk=sesion_id)

    cliente = (getattr(sesion, "cliente", "") or "").strip()
    ciudad  = (getattr(sesion, "ciudad", "") or "").strip()
    oficina = (getattr(sesion, "oficina", "") or "").strip()

    if not (cliente and ciudad and oficina):
        return JsonResponse({"ok": True, "results": []})

    # ✅ proyecto FK
    proyectos_visibles = filter_queryset_by_access(Proyecto.objects.all(), request.user, "id")
    proyecto_pk = _resolve_proyecto_pk_from_sesion(sesion, proyectos_qs=proyectos_visibles)
    if not proyecto_pk:
        return JsonResponse({"ok": True, "results": []})

    existing_ids = set(sesion.tecnicos_sesion.filter(is_active=True).values_list("tecnico_id", flat=True)
    )

    tecnicos_qs = (
        Usuario.objects
        .filter(
            is_active=True,
            precioactividadtecnico__cliente=cliente,
            precioactividadtecnico__ciudad=ciudad,
            precioactividadtecnico__proyecto_id=proyecto_pk,  # ✅ FK
            precioactividadtecnico__oficina=oficina,
        )
        .distinct()
        .order_by("first_name", "last_name", "username")
    )

    results = []
    for u in tecnicos_qs:
        results.append({
            "id": u.id,
            "name": (u.get_full_name() or u.username or "").strip(),
            "username": (u.username or "").strip(),
            "email": (u.email or "").strip(),
            "checked": u.id in existing_ids,
            "disabled": u.id in existing_ids,
        })

    return JsonResponse({"ok": True, "results": results})


@login_required
@rol_requerido("admin", "pm", "supervisor")
@require_POST
@transaction.atomic
def billing_add_tecnico(request, sesion_id: int):
    sesion = get_object_or_404(SesionBilling, pk=sesion_id)

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse({"ok": False, "error": "INVALID_JSON"}, status=400)

    tid = payload.get("tecnico_id")
    if not tid or not str(tid).isdigit():
        return JsonResponse({"ok": False, "error": "Missing tecnico_id."}, status=400)
    tid = int(tid)

    tech = Usuario.objects.filter(pk=tid, is_active=True).first()
    if not tech:
        return JsonResponse({"ok": False, "error": "Technician not found."}, status=404)

    if SesionBillingTecnico.objects.filter(sesion=sesion, tecnico_id=tid).exists():
        return JsonResponse({"ok": False, "error": "Technician already in this billing."}, status=400)

    cliente = (getattr(sesion, "cliente", "") or "").strip()
    ciudad  = (getattr(sesion, "ciudad", "") or "").strip()
    oficina = (getattr(sesion, "oficina", "") or "").strip()
    if not (cliente and ciudad and oficina):
        return JsonResponse({"ok": False, "error": "SESSION_HEADER_INCOMPLETE"}, status=400)

    proyectos_visibles = filter_queryset_by_access(Proyecto.objects.all(), request.user, "id")
    proyecto_pk = _resolve_proyecto_pk_from_sesion(sesion, proyectos_qs=proyectos_visibles)
    if not proyecto_pk:
        return JsonResponse({"ok": False, "error": "SESSION_PROJECT_UNRESOLVED"}, status=400)

    has_prices = PrecioActividadTecnico.objects.filter(
        tecnico_id=tid,
        cliente=cliente,
        ciudad=ciudad,
        proyecto_id=proyecto_pk,  # ✅ FK
        oficina=oficina,
    ).exists()

    if not has_prices:
        return JsonResponse({
            "ok": False,
            "error": "TECH_NOT_ELIGIBLE_FOR_SESSION",
            "message": "This technician has no prices for this exact Client/City/Project/Office.",
        }, status=400)

    current_ids = list(
        SesionBillingTecnico.objects.filter(sesion=sesion).values_list("tecnico_id", flat=True)
    )
    final_ids = _actualizar_tecnicos_preservando_fotos(sesion, current_ids + [tid], request=request)

    # ⚠️ IMPORTANTE: usa SOLO UNA versión de _recalcular_items_sesion (ver punto 2)
    _recalcular_items_sesion(sesion)

    return JsonResponse({"ok": True, "final_ids": final_ids})


@login_required
@require_POST
@transaction.atomic
def eliminar_billing(request, sesion_id: int):
    if not access_user_can(request.user, "billing.delete_billing"):
        messages.error(request, "You do not have permission to delete billings.")
        return HttpResponseRedirect(
            request.META.get("HTTP_REFERER", "/operaciones/billing/listar/")
        )

    sesion = get_object_or_404(SesionBilling, pk=sesion_id)

    # 🔒 Bloqueo contable absoluto:
    # Si al menos una línea de algún trabajador ya fue pagada,
    # el Billing no se puede eliminar.
    if _session_is_paid_locked(sesion):
        messages.error(
            request,
            (
                "This billing cannot be deleted because at least one worker/work type "
                "line has already been marked as paid."
            ),
        )
        return HttpResponseRedirect(
            request.META.get("HTTP_REFERER", "/operaciones/billing/listar/")
        )

    sesion.delete()

    messages.success(request, "Billing deleted.")
    return redirect("operaciones:listar_billing")


@login_required
@transaction.atomic
def reasignar_tecnicos(request, sesion_id: int):
    """
    Safe reassignment:
    - NO borra asignaciones existentes (porque CASCADE borra requisitos/evidencias).
    - Actualiza set final de técnicos usando _actualizar_tecnicos_preservando_fotos()
      (mantiene a los que ya tienen evidencias aunque intenten sacarlos).
    - Redistribuye porcentajes a 100% entre los técnicos finales.
    - Recalcula items (tarifas/porcentajes) sin tocar fotos ni requisitos existentes.
    """
    sesion = get_object_or_404(SesionBilling, pk=sesion_id)

    if request.method != "POST":
        return HttpResponseBadRequest("POST required")

    ids = [int(x) for x in request.POST.getlist("tech_ids[]") if str(x).isdigit()]
    if not ids:
        return HttpResponseBadRequest("Select at least one technician.")

    # ✅ CLAVE: NO borrar sesion.tecnicos_sesion.all().delete()
    # Usar el helper que preserva al que ya tiene evidencias.
    final_ids = _actualizar_tecnicos_preservando_fotos(sesion, ids, request=request)

    if not final_ids:
        return HttpResponseBadRequest("No technicians left after applying evidence rules.")

    # ✅ Recalcular items con los nuevos % (esto ya lo tienes y NO borra fotos)
    _recalcular_items_sesion(sesion)

    messages.success(request, "Technicians reassigned and totals recalculated.")
    return redirect("operaciones:editar_billing", sesion_id=sesion.id)


# ===== Persistencia =====


def _actualizar_tecnicos_preservando_fotos(sesion, nuevos_ids, request=None):
    """
    Regla:
    - Puedes QUITAR un técnico aunque tenga evidencias.
    - Sus evidencias se mantienen.
    - Si el técnico sale y tiene evidencias, se deja INACTIVO (si existe is_active).
    - Si no tiene evidencias, se elimina.
    - En modo:
        * split -> reparte 100% entre técnicos activos
        * full  -> cada técnico activo queda con 100%
    - Retorna final_ids_activos en orden.
    """

    nuevos_ids = [int(x) for x in (nuevos_ids or []) if str(x).isdigit()]
    nuevos_ids = list(dict.fromkeys(nuevos_ids))
    nuevos_set = set(nuevos_ids)

    if not nuevos_ids:
        return []

    existentes = list(sesion.tecnicos_sesion.select_related("tecnico").all())
    por_tid = {ts.tecnico_id: ts for ts in existentes}

    def _has_is_active(obj) -> bool:
        return hasattr(obj, "is_active")

    def _to_decimal_2(v):
        try:
            return Decimal(str(v)).quantize(Decimal("0.01"))
        except Exception:
            return Decimal("0.00")

    payment_mode = (getattr(sesion, "tech_payment_mode", "") or "split").strip().lower()
    if payment_mode not in ("split", "full"):
        payment_mode = "split"

    if payment_mode == "full":
        partes_dec = [Decimal("100.00") for _ in nuevos_ids]
    else:
        partes = repartir_100(len(nuevos_ids))
        partes_dec = [_to_decimal_2(p) for p in partes]

    # 1) activar/crear los que quedan
    for tid, pct in zip(nuevos_ids, partes_dec):
        if tid in por_tid:
            ts = por_tid[tid]
            update_fields = []

            if ts.porcentaje != pct:
                ts.porcentaje = pct
                update_fields.append("porcentaje")

            if _has_is_active(ts) and getattr(ts, "is_active", True) is not True:
                ts.is_active = True
                update_fields.append("is_active")

            if update_fields:
                ts.save(update_fields=update_fields)

        else:
            kwargs = dict(
                sesion=sesion,
                tecnico_id=tid,
                porcentaje=pct,
            )

            try:
                SesionBillingTecnico._meta.get_field("is_active")
                kwargs["is_active"] = True
            except Exception:
                pass

            SesionBillingTecnico.objects.create(**kwargs)

    # 2) procesar los que salen
    for ts in existentes:
        if ts.tecnico_id in nuevos_set:
            continue

        has_evs = EvidenciaFotoBilling.objects.filter(tecnico_sesion=ts).exists()

        if has_evs:
            if _has_is_active(ts):
                if getattr(ts, "is_active", True) is not False:
                    ts.is_active = False
                    ts.porcentaje = Decimal("0.00")
                    ts.save(update_fields=["is_active", "porcentaje"])

            if request is not None:
                try:
                    tech_name = ts.tecnico.get_full_name() or ts.tecnico.username
                except Exception:
                    tech_name = str(ts.tecnico_id)

                messages.info(
                    request,
                    f"'{tech_name}' was removed from the billing but their photos were kept.",
                )
        else:
            ts.delete()

    return nuevos_ids


@transaction.atomic
def _guardar_billing(request, sesion=None):
    import re

    if sesion is None:

        if not access_user_can(request.user, "billing.create_billing"):

            messages.error(request, "You do not have permission to create billings.")

            return redirect("operaciones:listar_billing")

    else:

        if not access_user_can(request.user, "billing.create_billing"):

            messages.error(request, "You do not have permission to edit billings.")

            return redirect("operaciones:listar_billing")

    if sesion is not None and _session_is_paid_locked(sesion):

        messages.error(
            request,
            (
                "This billing cannot be modified because at least one worker/work type "
                "line has already been marked as paid."
            ),
        )

        return redirect("operaciones:listar_billing")

    WEEK_RE = re.compile(r"^\d{4}-W\d{2}$")

    def money(v) -> Decimal:
        try:
            return Decimal(str(v or "0")).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError, TypeError):
            return Decimal("0.00")

    def meta_codigo(cliente, ciudad, proyecto_pk, oficina, codigo):
        return (
            PrecioActividadTecnico.objects.filter(
                cliente=cliente,
                ciudad=ciudad,
                proyecto_id=proyecto_pk,
                oficina=oficina,
                codigo_trabajo=codigo,
            )
            .values("tipo_trabajo", "descripcion", "unidad_medida")
            .first()
        )

    def precio_empresa(cliente, ciudad, proyecto_pk, oficina, codigo) -> Decimal:
        val = (
            PrecioActividadTecnico.objects.filter(
                cliente=cliente,
                ciudad=ciudad,
                proyecto_id=proyecto_pk,
                oficina=oficina,
                codigo_trabajo=codigo,
            )
            .values_list("precio_empresa", flat=True)
            .first()
        )
        return money(val or 0)

    def tarifa_tecnico(tid, cliente, ciudad, proyecto_pk, oficina, codigo) -> Decimal:
        val = (
            PrecioActividadTecnico.objects.filter(
                tecnico_id=tid,
                cliente=cliente,
                ciudad=ciudad,
                proyecto_id=proyecto_pk,
                oficina=oficina,
                codigo_trabajo=codigo,
            )
            .values_list("precio_tecnico", flat=True)
            .first()
        )
        if val is None:
            val = (
                PrecioActividadTecnico.objects.filter(
                    cliente=cliente,
                    ciudad=ciudad,
                    proyecto_id=proyecto_pk,
                    oficina=oficina,
                    codigo_trabajo=codigo,
                )
                .values_list("precio_tecnico", flat=True)
                .first()
            )
        return money(val or 0)

    project_code = (request.POST.get("project_id") or "").strip()
    cliente = (request.POST.get("client") or "").strip()
    ciudad = (request.POST.get("city") or "").strip()
    project_pk_raw = (request.POST.get("project") or "").strip()
    oficina = (request.POST.get("office") or "").strip()
    ids = [int(x) for x in request.POST.getlist("tech_ids[]") if str(x).isdigit()]

    direccion_proyecto = (request.POST.get("direccion_proyecto") or "").strip()
    semana_pago_proyectada = (request.POST.get("semana_pago_proyectada") or "").strip()
    if semana_pago_proyectada and not WEEK_RE.match(semana_pago_proyectada):
        semana_pago_proyectada = ""

    is_direct_discount = request.POST.get("direct_discount") == "1"
    is_cable_installation = request.POST.get("is_cable_installation") == "1"

    tech_payment_mode = (
        (request.POST.get("tech_payment_mode") or "split").strip().lower()
    )
    if tech_payment_mode not in ("split", "full"):
        tech_payment_mode = "split"

    def build_post_items_context():
        items_ctx = []

        row_codes = request.POST.getlist("row_code[]")
        row_amounts = request.POST.getlist("row_amount[]")

        for code, amount in zip(row_codes, row_amounts):
            code = (code or "").strip()

            row_meta = None
            company_price_val = ""
            subtotal_company_val = ""
            subtotal_tecnico_val = ""

            if project_pk_raw.isdigit() and cliente and ciudad and oficina and code:
                try:
                    proyecto_pk_tmp = int(project_pk_raw)
                    row_meta = meta_codigo(
                        cliente, ciudad, proyecto_pk_tmp, oficina, code
                    )
                    if row_meta:
                        cp = precio_empresa(
                            cliente, ciudad, proyecto_pk_tmp, oficina, code
                        )
                        qty = money(amount)
                        company_price_val = f"{cp:.2f}"
                        subtotal_company_val = f"{money(cp * qty):.2f}"
                except Exception:
                    row_meta = None

            items_ctx.append(
                {
                    "codigo_trabajo": code,
                    "tipo_trabajo": row_meta["tipo_trabajo"] if row_meta else "",
                    "descripcion": row_meta["descripcion"] if row_meta else "",
                    "unidad_medida": row_meta["unidad_medida"] if row_meta else "",
                    "cantidad": amount,
                    "precio_empresa": company_price_val,
                    "subtotal_empresa": subtotal_company_val,
                    "subtotal_tecnico": subtotal_tecnico_val,
                    "desglose_tecnico": [],
                }
            )
        return items_ctx

    def render_with_data(error_msg=None):
        if error_msg:
            messages.error(request, error_msg)

        clientes = (
            PrecioActividadTecnico.objects.values_list("cliente", flat=True)
            .distinct()
            .order_by("cliente")
        )

        proyectos_visibles = filter_queryset_by_access(
            Proyecto.objects.all(),
            request.user,
            "id",
        )

        if proyectos_visibles.exists():
            tecnicos = (
                Usuario.objects.filter(
                    is_active=True,
                    precioactividadtecnico__isnull=False,
                    precioactividadtecnico__proyecto_id__in=proyectos_visibles.values_list(
                        "id", flat=True
                    ),
                )
                .distinct()
                .order_by("first_name", "last_name", "username")
            )
        else:
            tecnicos = Usuario.objects.none()

        proyecto_value = project_pk_raw
        proyecto_label = ""

        if project_pk_raw.isdigit():
            p = proyectos_visibles.filter(pk=int(project_pk_raw)).first()
            if p:
                proyecto_label = (getattr(p, "nombre", "") or str(p)).strip()

        sesion_ctx = {
            "id": getattr(sesion, "id", None),
            "proyecto_id": project_code,
            "cliente": cliente,
            "ciudad": ciudad,
            "proyecto": project_pk_raw,
            "oficina": oficina,
            "direccion_proyecto": direccion_proyecto,
            "semana_pago_proyectada": semana_pago_proyectada,
            "is_direct_discount": is_direct_discount,
            "is_cable_installation": is_cable_installation,
            "tech_payment_mode": tech_payment_mode,
        }

        context = {
            "is_edit": bool(sesion),
            "sesion": sesion_ctx,
            "clientes": list(clientes),
            "tecnicos": tecnicos,
            "items": build_post_items_context(),
            "ids_tecnicos": ids,
            "proyecto_value": proyecto_value,
            "proyecto_label": proyecto_label,
        }

        context.update(_billing_access_context(request.user))

        return render(
            request,
            "operaciones/billing_editar.html",
            context,
        )

    if not (project_code and cliente and ciudad and project_pk_raw and oficina):
        return render_with_data("Complete all header fields.")

    if not ids:
        return render_with_data("Select at least one technician.")

    try:
        proyecto_pk = int(project_pk_raw)
    except (TypeError, ValueError):
        return render_with_data("Invalid project.")

    proyecto_obj = Proyecto.objects.filter(pk=proyecto_pk).first()
    if not proyecto_obj:
        return render_with_data("Selected project does not exist.")

    proyecto_nombre = (getattr(proyecto_obj, "nombre", "") or "").strip()

    row_codes = request.POST.getlist("row_code[]")
    row_amounts = request.POST.getlist("row_amount[]")

    if not row_codes:
        return render_with_data("Please add at least one item.")

    if len(row_codes) != len(row_amounts):
        return render_with_data("Invalid items data.")

    filas = []
    for cod, amt in zip(row_codes, row_amounts):
        cod = (cod or "").strip()

        if not cod or amt in ("", None):
            return render_with_data("Each row requires Job Code and Quantity.")

        try:
            qty = Decimal(str(amt))
        except Exception:
            return render_with_data(f"Invalid quantity for code '{cod}'.")

        if is_direct_discount and qty > 0:
            qty = -qty

        meta = meta_codigo(cliente, ciudad, proyecto_pk, oficina, cod)
        if not meta:
            return render_with_data(
                f"Code '{cod}' does not exist with the selected filters."
            )

        filas.append(
            {
                "codigo": cod,
                "cantidad": qty,
                "meta": meta,
            }
        )

    if sesion is None:
        sesion = SesionBilling.objects.create(
            proyecto_id=project_code,
            cliente=cliente,
            ciudad=ciudad,
            proyecto=proyecto_nombre,
            oficina=oficina,
            direccion_proyecto=direccion_proyecto,
            semana_pago_proyectada=semana_pago_proyectada,
            semana_pago_real=semana_pago_proyectada if is_direct_discount else "",
            is_direct_discount=is_direct_discount,
            is_cable_installation=is_cable_installation,
            tech_payment_mode=tech_payment_mode,
        )
    else:
        sesion.proyecto_id = project_code
        sesion.cliente = cliente
        sesion.ciudad = ciudad
        sesion.proyecto = proyecto_nombre
        sesion.oficina = oficina
        sesion.direccion_proyecto = direccion_proyecto
        sesion.semana_pago_proyectada = semana_pago_proyectada
        if is_direct_discount:
            sesion.semana_pago_real = semana_pago_proyectada
        else:
            sesion.semana_pago_real = ""
        sesion.is_direct_discount = is_direct_discount
        sesion.is_cable_installation = is_cable_installation
        sesion.tech_payment_mode = tech_payment_mode
        sesion.save()

    final_ids = _actualizar_tecnicos_preservando_fotos(sesion, ids, request=request)

    ts_rows = list(
        sesion.tecnicos_sesion.filter(tecnico_id__in=final_ids).values_list(
            "tecnico_id", "porcentaje"
        )
    )
    pct_by_tid = {tid: money(pct) for tid, pct in ts_rows}

    ids_def = final_ids
    partes_def = [pct_by_tid.get(tid, Decimal("0.00")) for tid in ids_def]

    sesion.items.all().delete()

    total_emp = Decimal("0.00")
    total_tec = Decimal("0.00")

    for fila in filas:
        cod = fila["codigo"]
        qty = money(fila["cantidad"])
        meta = fila["meta"]

        p_emp = precio_empresa(cliente, ciudad, proyecto_pk, oficina, cod)
        sub_emp = money(p_emp * qty)

        item = ItemBilling.objects.create(
            sesion=sesion,
            codigo_trabajo=cod,
            tipo_trabajo=meta["tipo_trabajo"],
            descripcion=meta["descripcion"],
            unidad_medida=meta["unidad_medida"],
            cantidad=qty,
            precio_empresa=p_emp,
            subtotal_empresa=sub_emp,
            subtotal_tecnico=Decimal("0.00"),
        )

        sub_tecs = Decimal("0.00")

        for tid, pct in zip(ids_def, partes_def):
            base = tarifa_tecnico(tid, cliente, ciudad, proyecto_pk, oficina, cod)

            if tech_payment_mode == "full":
                efectiva = money(base)
                pct_guardado = Decimal("100.00")
            else:
                efectiva = money(base * (pct / Decimal("100")))
                pct_guardado = pct

            subtotal = money(efectiva * item.cantidad)

            ItemBillingTecnico.objects.create(
                item=item,
                tecnico_id=tid,
                tarifa_base=base,
                porcentaje=pct_guardado,
                tarifa_efectiva=efectiva,
                subtotal=subtotal,
            )
            sub_tecs += subtotal

        item.subtotal_tecnico = sub_tecs
        item.save(update_fields=["subtotal_tecnico"])

        total_emp += sub_emp
        total_tec += sub_tecs

    sesion.subtotal_empresa = money(total_emp)
    sesion.subtotal_tecnico = money(total_tec)
    sesion.save(
        update_fields=[
            "subtotal_empresa",
            "subtotal_tecnico",
            "semana_pago_real",
            "is_direct_discount",
            "is_cable_installation",
            "tech_payment_mode",
        ]
    )

    rebuild_billing_payweek_snapshot(sesion)

    messages.success(
        request,
        (
            "Direct discount saved and linked to the selected technician(s)."
            if is_direct_discount
            else "Billing saved successfully."
        ),
    )
    return redirect("operaciones:listar_billing")


# ===== Búsquedas / AJAX =====
def _precio_empresa(cliente, ciudad, proyecto, oficina, codigo):
    """
    proyecto = PK real de facturacion.Proyecto
    """
    q = PrecioActividadTecnico.objects.filter(
        cliente__iexact=cliente,
        ciudad__iexact=ciudad,
        proyecto_id=proyecto,
        oficina__iexact=oficina or "-",
        codigo_trabajo__iexact=codigo,
    ).first()
    return money(q.precio_empresa if q else 0)


def _tarifa_tecnico(tecnico_id, cliente, ciudad, proyecto, oficina, codigo):
    """
    proyecto = PK real de facturacion.Proyecto
    """
    q = PrecioActividadTecnico.objects.filter(
        tecnico_id=tecnico_id,
        cliente__iexact=cliente,
        ciudad__iexact=ciudad,
        proyecto_id=proyecto,
        oficina__iexact=oficina or "-",
        codigo_trabajo__iexact=codigo,
    ).first()
    return money(q.precio_tecnico if q else 0)


def _meta_codigo(cliente, ciudad, proyecto, oficina, codigo):
    """
    proyecto = PK real de facturacion.Proyecto
    """
    p = PrecioActividadTecnico.objects.filter(
        cliente__iexact=cliente,
        ciudad__iexact=ciudad,
        proyecto_id=proyecto,
        oficina__iexact=oficina or "-",
        codigo_trabajo__iexact=codigo,
    ).first()
    if not p:
        return None
    return {
        "tipo_trabajo": p.tipo_trabajo,
        "descripcion": p.descripcion,
        "unidad_medida": p.unidad_medida,
    }


@login_required
def ajax_clientes(request):
    data = list(
        PrecioActividadTecnico.objects
        .values_list("cliente", flat=True)
        .distinct()
        .order_by("cliente")
    )
    return JsonResponse({"results": data})


@login_required
def ajax_ciudades(request):
    cliente = request.GET.get("client", "")
    data = list(
        PrecioActividadTecnico.objects.filter(cliente__iexact=cliente)
        .values_list("ciudad", flat=True)
        .distinct()
        .order_by("ciudad")
    ) if cliente else []
    return JsonResponse({"results": data})


@login_required
def ajax_proyectos(request):
    from django.db.models import Count

    cliente = (request.GET.get("client") or "").strip()
    ciudad = (request.GET.get("city") or "").strip()
    tech_ids = [int(x) for x in request.GET.getlist("tech_ids[]") if str(x).isdigit()]

    data = []

    if not (cliente and ciudad):
        return JsonResponse({"results": data})

    proyectos_visibles_qs = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        "id",
    )
    visible_ids = list(proyectos_visibles_qs.values_list("id", flat=True))

    if not visible_ids:
        return JsonResponse({"results": []})

    qs = PrecioActividadTecnico.objects.filter(
        cliente__iexact=cliente,
        ciudad__iexact=ciudad,
        proyecto_id__in=visible_ids,
    )

    if tech_ids:
        needed = len(set(tech_ids))

        qs = (
            qs.filter(tecnico_id__in=tech_ids)
            .values("proyecto_id", "proyecto__nombre")
            .annotate(match_count=Count("tecnico_id", distinct=True))
            .filter(match_count=needed)
            .order_by("proyecto__nombre", "proyecto_id")
        )

        for row in qs:
            if not row["proyecto_id"]:
                continue
            data.append(
                {
                    "id": row["proyecto_id"],
                    "label": (
                        row["proyecto__nombre"] or str(row["proyecto_id"])
                    ).strip(),
                }
            )
    else:
        qs = (
            qs.values("proyecto_id", "proyecto__nombre")
            .distinct()
            .order_by("proyecto__nombre", "proyecto_id")
        )
        for row in qs:
            if not row["proyecto_id"]:
                continue
            data.append(
                {
                    "id": row["proyecto_id"],
                    "label": (
                        row["proyecto__nombre"] or str(row["proyecto_id"])
                    ).strip(),
                }
            )

    return JsonResponse({"results": data})


@login_required
def ajax_oficinas(request):
    from django.db.models import Count

    cliente = (request.GET.get("client") or "").strip()
    ciudad = (request.GET.get("city") or "").strip()
    proyecto_id = (request.GET.get("project") or "").strip()
    tech_ids = [int(x) for x in request.GET.getlist("tech_ids[]") if str(x).isdigit()]

    if not (cliente and ciudad and proyecto_id):
        return JsonResponse({"results": []})

    qs = PrecioActividadTecnico.objects.filter(
        cliente__iexact=cliente,
        ciudad__iexact=ciudad,
        proyecto_id=proyecto_id,
    )

    data = []

    if tech_ids:
        needed = len(set(tech_ids))
        qs = (
            qs.filter(tecnico_id__in=tech_ids)
            .values("oficina")
            .annotate(match_count=Count("tecnico_id", distinct=True))
            .filter(match_count=needed)
            .order_by("oficina")
        )
        data = [row["oficina"] for row in qs if (row["oficina"] or "").strip()]
    else:
        data = list(qs.values_list("oficina", flat=True).distinct().order_by("oficina"))

    return JsonResponse({"results": data})

@login_required
@require_POST
def ajax_validate_billing_selection(request):
    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        return JsonResponse({"ok": False, "error": "INVALID_JSON"}, status=400)

    client = (payload.get("client") or "").strip()
    city = (payload.get("city") or "").strip()
    project = str(payload.get("project") or "").strip()
    office = (payload.get("office") or "").strip()
    tech_ids = [int(x) for x in (payload.get("tech_ids") or []) if str(x).isdigit()]
    items = payload.get("items") or []

    if not client or not city or not project or not office:
        return JsonResponse(
            {
                "ok": False,
                "error": "MISSING_HEADER",
                "message": "Client, City, Project and Office are required.",
            },
            status=400,
        )

    if not tech_ids:
        return JsonResponse(
            {
                "ok": False,
                "error": "MISSING_TECHS",
                "message": "Select at least one technician.",
            },
            status=400,
        )

    tech_map = {
        u.id: (u.get_full_name() or u.username or f"Tech {u.id}").strip()
        for u in Usuario.objects.filter(id__in=tech_ids)
    }

    missing = []

    # Validación de header: todos deben tener al menos una tarifa en ese header exacto
    for tid in tech_ids:
        exists_header = PrecioActividadTecnico.objects.filter(
            tecnico_id=tid,
            cliente__iexact=client,
            ciudad__iexact=city,
            proyecto_id=project,
            oficina__iexact=office,
        ).exists()

        if not exists_header:
            missing.append(
                {
                    "type": "header",
                    "tecnico_id": tid,
                    "tecnico_nombre": tech_map.get(tid, f"Tech {tid}"),
                    "message": f"{tech_map.get(tid, f'Tech {tid}')} has no exact prices for Client / City / Project / Office.",
                }
            )

    # Validación item por item: todos deben tener Work Type + Job Code exactos
    for idx, item in enumerate(items, start=1):
        code = (item.get("code") or "").strip()
        work_type = (item.get("work_type") or "").strip()

        if not code and not work_type:
            continue

        for tid in tech_ids:
            exists_item = PrecioActividadTecnico.objects.filter(
                tecnico_id=tid,
                cliente__iexact=client,
                ciudad__iexact=city,
                proyecto_id=project,
                oficina__iexact=office,
                tipo_trabajo__iexact=work_type,
                codigo_trabajo__iexact=code,
            ).exists()

            if not exists_item:
                missing.append(
                    {
                        "type": "item",
                        "row": idx,
                        "code": code,
                        "work_type": work_type,
                        "tecnico_id": tid,
                        "tecnico_nombre": tech_map.get(tid, f"Tech {tid}"),
                        "message": (
                            f"Row {idx}: {tech_map.get(tid, f'Tech {tid}')} "
                            f"has no exact match for City / Project / Office / Client / Work Type / Job Code."
                        ),
                    }
                )

    if missing:
        return JsonResponse(
            {
                "ok": False,
                "error": "MISSING_EXACT_MATCHES",
                "message": "One or more selected technicians do not have exact matching prices.",
                "missing": missing,
            },
            status=400,
        )

    return JsonResponse({"ok": True})

@login_required
def ajax_buscar_codigos(request):
    cliente = request.GET.get("client", "")
    ciudad = request.GET.get("city", "")
    proyecto_id = request.GET.get("project", "")   # ← es el PK
    oficina = request.GET.get("office", "")
    q = (request.GET.get("q") or "").strip()

    if not (cliente and ciudad and proyecto_id and oficina):
        return JsonResponse({"error": "missing_filters"}, status=400)

    qs = PrecioActividadTecnico.objects.filter(
        cliente__iexact=cliente,
        ciudad__iexact=ciudad,
        proyecto_id=proyecto_id,              # ← ahora por FK
        oficina__iexact=oficina or "-",
    )
    if q:
        qs = qs.filter(codigo_trabajo__istartswith=q)

    data = list(
        qs.values("codigo_trabajo", "tipo_trabajo", "descripcion", "unidad_medida")
          .distinct()
          .order_by("codigo_trabajo")[:20]
    )
    return JsonResponse({"results": data})


@login_required
def ajax_detalle_codigo(request):
    cliente = (request.GET.get("client") or "").strip()
    ciudad = (request.GET.get("city") or "").strip()
    proyecto = (request.GET.get("project") or "").strip()
    oficina = (request.GET.get("office") or "").strip()
    codigo = (request.GET.get("code") or "").strip()

    if not (cliente and ciudad and proyecto and oficina and codigo):
        return JsonResponse({"error": "missing_filters"}, status=400)

    meta = _meta_codigo(cliente, ciudad, proyecto, oficina, codigo)
    if not meta:
        return JsonResponse({"error": "not_found"}, status=404)

    tech_ids = [int(x) for x in request.GET.getlist("tech_ids[]") if str(x).isdigit()]
    if not tech_ids:
        return JsonResponse({"error": "missing_technicians"}, status=400)

    payment_mode = (
        (
            request.GET.get("payment_mode")
            or request.GET.get("tech_payment_mode")
            or "split"
        )
        .strip()
        .lower()
    )
    if payment_mode not in ("split", "full"):
        payment_mode = "split"

    tech_map = {
        u.id: (u.get_full_name() or u.username or f"Tech {u.id}").strip()
        for u in Usuario.objects.filter(id__in=tech_ids)
    }

    exact_rows = []
    missing = []

    for tid in tech_ids:
        row = PrecioActividadTecnico.objects.filter(
            tecnico_id=tid,
            cliente__iexact=cliente,
            ciudad__iexact=ciudad,
            proyecto_id=proyecto,
            oficina__iexact=oficina,
            codigo_trabajo__iexact=codigo,
            tipo_trabajo__iexact=(meta["tipo_trabajo"] or "").strip(),
        ).first()

        if not row:
            missing.append(
                {
                    "tecnico_id": tid,
                    "tecnico_nombre": tech_map.get(tid, f"Tech {tid}"),
                    "message": (
                        f"{tech_map.get(tid, f'Tech {tid}')} has no exact price for "
                        f"Client / City / Project / Office / Work Type / Job Code."
                    ),
                }
            )
            continue

        exact_rows.append(row)

    if missing:
        return JsonResponse(
            {
                "error": "missing_exact_prices",
                "message": "Exact prices are missing for one or more selected technicians.",
                "missing": missing,
            },
            status=400,
        )

    precio_emp = money(exact_rows[0].precio_empresa or 0)
    partes = repartir_100(len(tech_ids)) if tech_ids else []

    desglose = []
    for row, pct in zip(exact_rows, partes):
        base = money(row.precio_tecnico or 0)
        payment_weeks = int(getattr(row, "payment_weeks", 0) or 0)

        if payment_mode == "full":
            porcentaje = Decimal("100.00")
            efectiva = base
        else:
            porcentaje = Decimal(str(pct)).quantize(Decimal("0.01"))
            efectiva = (base * (porcentaje / Decimal("100"))).quantize(Decimal("0.01"))

        desglose.append(
            {
                "tecnico_id": row.tecnico_id,
                "tecnico_nombre": tech_map.get(
                    row.tecnico_id, f"Tech {row.tecnico_id}"
                ),
                "tarifa_base": f"{base:.2f}",
                "porcentaje": f"{porcentaje:.2f}",
                "tarifa_efectiva": f"{efectiva:.2f}",
                "payment_weeks": payment_weeks,
            }
        )

    return JsonResponse(
        {
            "tipo_trabajo": meta["tipo_trabajo"],
            "descripcion": meta["descripcion"],
            "unidad_medida": meta["unidad_medida"],
            "precio_empresa": f"{precio_emp:.2f}",
            "desglose_tecnico": desglose,
        }
    )


ESTADOS_OK = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}


@login_required
@rol_requerido("admin", "supervisor", "pm", "facturacion")
def produccion_admin(request):
    """
    Producción por técnico (vista Admin) con filtros + paginación.

    Incluye:
    - sesiones con snapshots productivos
    - sesiones legacy aprobadas sin snapshots productivos
    - direct discounts
    - ajustes manuales

    PERFORMANCE:
    - Snapshots se leen directo desde BillingPayWeekSnapshot.
    - Legacy NO carga items/desglose antes de paginar.
    - El detalle legacy se arma SOLO para las filas visibles de la página.
    """
    import re
    from decimal import Decimal
    from urllib.parse import urlencode

    from django.core.paginator import Paginator
    from django.db.models import CharField, Exists, OuterRef, Prefetch, Q
    from django.db.models.functions import Cast
    from django.utils import timezone

    from core.permissions import filter_queryset_by_access
    from facturacion.models import Proyecto
    from operaciones.models import (BillingPayWeekSnapshot, ItemBilling,
                                    ItemBillingTecnico, SesionBilling,
                                    SesionBillingTecnico)

    try:
        from operaciones.models import AdjustmentEntry
    except Exception:
        AdjustmentEntry = None

    try:
        from usuarios.models import ProyectoAsignacion
    except Exception:
        ProyectoAsignacion = None

    # ============================================================
    # Helpers
    # ============================================================
    def _iso_week_str(dt):
        y, w, _ = dt.isocalendar()
        return f"{y}-W{int(w):02d}"

    def parse_week_query(q: str):
        """
        Acepta: '34', 'w34', 'W34', '2025-W34', '2025W34'
        Retorna (exact_iso, week_token)
        """
        if not q:
            return (None, None)

        s = q.strip().upper().replace("WEEK", "W").replace(" ", "")

        m = re.fullmatch(r"(\d{4})-?W(\d{1,2})", s)
        if m:
            year, ww = int(m.group(1)), int(m.group(2))
            return (f"{year}-W{ww:02d}", None)

        m = re.fullmatch(r"(?:W)?(\d{1,2})", s)
        if m:
            ww = int(m.group(1))
            return (None, f"W{ww:02d}")

        return (None, None)

    def _normalize_week_str(s: str) -> str:
        if not s:
            return ""

        s = str(s).replace("\u2013", "-").replace("\u2014", "-")
        s = re.sub(r"\s+", "", s)

        return s.upper()

    def _week_sort_key(week_str: str):
        if not week_str:
            return (-1, -1)

        s = str(week_str).upper().replace("WEEK", "W").replace(" ", "")

        m = re.search(r"(\d{4})-?W(\d{1,2})", s)
        if m:
            return (int(m.group(1)), int(m.group(2)))

        m = re.search(r"W(\d{1,2})", s)
        if m:
            return (0, int(m.group(1)))

        return (-1, -1)

    def _match_week_filter(
        week_real: str, exact_week: str | None, week_token: str | None
    ) -> bool:
        wr = (week_real or "").strip().upper()

        if not wr:
            return False

        if exact_week:
            token = exact_week.split("-", 1)[-1].upper()
            return wr == exact_week or token in wr

        if week_token:
            return week_token in wr

        return True

    def _match_tech_filter(user_obj, f_tech_value: str) -> bool:
        if not f_tech_value:
            return True

        target = f_tech_value.lower()

        full_name = (
            (
                (getattr(user_obj, "first_name", "") or "")
                + " "
                + (getattr(user_obj, "last_name", "") or "")
            )
            .strip()
            .lower()
        )

        username = (getattr(user_obj, "username", "") or "").lower()

        return target in full_name or target in username

    def _to_decimal(value):
        try:
            if isinstance(value, Decimal):
                return value
            return Decimal(str(value or 0))
        except Exception:
            return Decimal("0.00")

    # ============================================================
    # Configuración / filtros
    # ============================================================
    estados_ok = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}
    current_week = _iso_week_str(timezone.now())

    user = request.user
    can_view_legacy_history = user.is_superuser or getattr(
        user, "es_usuario_historial", False
    )

    f_project = (request.GET.get("f_project") or "").strip()
    f_week_input = (request.GET.get("f_week") or "").strip()
    f_tech = (request.GET.get("f_tech") or "").strip()
    f_client = (request.GET.get("f_client") or "").strip()

    exact_week, week_token = parse_week_query(f_week_input)

    # ============================================================
    # Proyectos visibles
    # ============================================================
    try:
        base_proyectos = Proyecto.objects.all()

        if can_view_legacy_history:
            proyectos_user = base_proyectos
        else:
            proyectos_user = filter_queryset_by_access(
                base_proyectos,
                user,
                "id",
            )
    except Exception:
        proyectos_user = Proyecto.objects.none()

    proyectos_list = list(proyectos_user)

    if proyectos_list:
        allowed_keys = set()

        for p in proyectos_list:
            nombre = (getattr(p, "nombre", "") or "").strip()
            if nombre:
                allowed_keys.add(nombre)

            codigo = getattr(p, "codigo", None)
            if codigo:
                allowed_keys.add(str(codigo).strip())

            allowed_keys.add(str(p.id).strip())
    else:
        allowed_keys = set()

    by_id = {p.id: p for p in proyectos_list}
    by_code = {
        (p.codigo or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "codigo", None)
    }
    by_name = {
        (p.nombre or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "nombre", None)
    }

    # ============================================================
    # Detectar campos opcionales del snapshot
    # ============================================================
    snapshot_field_names = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_field_names
    has_adjustment_of = "adjustment_of" in snapshot_field_names

    def _productive_snapshot_filter():
        q = Q()

        if has_is_adjustment:
            q &= Q(is_adjustment=False)

        if has_adjustment_of:
            q &= Q(adjustment_of__isnull=True)

        return q

    # ============================================================
    # 1) Flujo nuevo: snapshots productivos
    # ============================================================
    filas = []

    snap_qs = (
        BillingPayWeekSnapshot.objects.select_related("sesion", "tecnico", "item")
        .filter(
            _productive_snapshot_filter(),
            Q(sesion__estado__in=estados_ok) | Q(sesion__is_direct_discount=True),
        )
        .only(
            "id",
            "sesion_id",
            "tecnico_id",
            "item_id",
            "codigo_trabajo",
            "tipo_trabajo",
            "semana_base",
            "semana_resultado",
            "tarifa_efectiva",
            "subtotal",
            "sesion__id",
            "sesion__proyecto_id",
            "sesion__estado",
            "sesion__is_direct_discount",
            "sesion__cliente",
            "sesion__ciudad",
            "sesion__proyecto",
            "sesion__oficina",
            "sesion__semana_pago_proyectada",
            "tecnico__id",
            "tecnico__first_name",
            "tecnico__last_name",
            "tecnico__username",
            "item__id",
            "item__descripcion",
            "item__unidad_medida",
            "item__cantidad",
        )
        .order_by(
            "-semana_resultado",
            "tecnico__first_name",
            "tecnico__last_name",
            "tecnico__username",
            "id",
        )
    )

    # Visibilidad por proyectos
    if not can_view_legacy_history:
        if allowed_keys:
            snap_qs = snap_qs.filter(sesion__proyecto__in=allowed_keys)
        else:
            snap_qs = BillingPayWeekSnapshot.objects.none()

    # Filtro project
    if f_project:
        snap_qs = snap_qs.annotate(
            sesion_proyecto_id_str=Cast("sesion__proyecto_id", CharField())
        ).filter(
            Q(sesion_proyecto_id_str__icontains=f_project)
            | Q(sesion__proyecto__icontains=f_project)
        )

    # Filtro client
    if f_client:
        snap_qs = snap_qs.filter(sesion__cliente__icontains=f_client)

    # Filtro technician
    if f_tech:
        snap_qs = snap_qs.filter(
            Q(tecnico__first_name__icontains=f_tech)
            | Q(tecnico__last_name__icontains=f_tech)
            | Q(tecnico__username__icontains=f_tech)
        )

    # Filtro week
    if exact_week:
        token = exact_week.split("-", 1)[-1].upper()
        snap_qs = snap_qs.filter(
            Q(semana_resultado__iexact=exact_week)
            | Q(semana_resultado__icontains=token)
        )
    elif week_token:
        snap_qs = snap_qs.filter(semana_resultado__icontains=week_token)

    tech_rows = {}

    for snap in snap_qs.iterator(chunk_size=1000):
        s = getattr(snap, "sesion", None)
        tecnico = getattr(snap, "tecnico", None)

        if not s or not tecnico:
            continue

        week_real = (getattr(snap, "semana_resultado", "") or "").strip().upper()

        if not week_real:
            continue

        if not _match_week_filter(week_real, exact_week, week_token):
            continue

        if not _match_tech_filter(tecnico, f_tech):
            continue

        key = (s.id, tecnico.id, week_real)

        if key not in tech_rows:
            tech_rows[key] = {
                "_source": "snapshot",
                "sesion": s,
                "tecnico": tecnico,
                "project_id": s.proyecto_id,
                "week": week_real or "—",
                "status": s.estado,
                "is_discount": bool(getattr(s, "is_direct_discount", False)),
                "client": s.cliente,
                "city": s.ciudad,
                "project": s.proyecto,
                "office": s.oficina,
                "real_week": week_real or "—",
                "proj_week": s.semana_pago_proyectada or "—",
                "total_tecnico": Decimal("0.00"),
                "detalle": [],
                "adjustment_type": "",
            }

        subtotal = _to_decimal(getattr(snap, "subtotal", 0))
        tech_rows[key]["total_tecnico"] += subtotal

        if subtotal < 0:
            tech_rows[key]["is_discount"] = True

        item = getattr(snap, "item", None)
        rate_tec = _to_decimal(getattr(snap, "tarifa_efectiva", 0))

        tech_rows[key]["detalle"].append(
            {
                "codigo": getattr(snap, "codigo_trabajo", "") or "",
                "tipo": getattr(snap, "tipo_trabajo", "") or "",
                "desc": getattr(item, "descripcion", "") if item else "",
                "uom": getattr(item, "unidad_medida", "") if item else "",
                "qty": getattr(item, "cantidad", None) if item else None,
                "rate_tec": rate_tec,
                "subtotal_tec": subtotal,
            }
        )

    filas.extend(tech_rows.values())

    # ============================================================
    # 2) Legacy: sesiones SIN snapshots productivos
    #    Importante: NO carga items/desglose aquí.
    # ============================================================
    productive_snap_exists = BillingPayWeekSnapshot.objects.filter(
        _productive_snapshot_filter(),
        sesion_id=OuterRef("pk"),
    )

    legacy_qs = (
        SesionBilling.objects.filter(
            Q(estado__in=estados_ok) | Q(is_direct_discount=True)
        )
        .annotate(has_productive_snap=Exists(productive_snap_exists))
        .filter(has_productive_snap=False)
        .only(
            "id",
            "proyecto_id",
            "estado",
            "is_direct_discount",
            "cliente",
            "ciudad",
            "proyecto",
            "oficina",
            "semana_pago_real",
            "semana_pago_proyectada",
            "subtotal_tecnico",
            "creado_en",
        )
        .prefetch_related(
            Prefetch(
                "tecnicos_sesion",
                queryset=SesionBillingTecnico.objects.select_related("tecnico").only(
                    "id",
                    "sesion_id",
                    "tecnico_id",
                    "porcentaje",
                    "tecnico__id",
                    "tecnico__first_name",
                    "tecnico__last_name",
                    "tecnico__username",
                ),
            )
        )
        .order_by("-creado_en")
        .distinct()
    )

    if not can_view_legacy_history:
        if allowed_keys:
            legacy_qs = legacy_qs.filter(proyecto__in=allowed_keys)
        else:
            legacy_qs = SesionBilling.objects.none()

    if f_project:
        legacy_qs = legacy_qs.annotate(
            proyecto_id_str=Cast("proyecto_id", CharField())
        ).filter(
            Q(proyecto_id_str__icontains=f_project) | Q(proyecto__icontains=f_project)
        )

    if f_client:
        legacy_qs = legacy_qs.filter(cliente__icontains=f_client)

    if f_tech:
        legacy_qs = legacy_qs.filter(
            Q(tecnicos_sesion__tecnico__first_name__icontains=f_tech)
            | Q(tecnicos_sesion__tecnico__last_name__icontains=f_tech)
            | Q(tecnicos_sesion__tecnico__username__icontains=f_tech)
        )

    if exact_week:
        token = exact_week.split("-", 1)[-1].upper()
        legacy_qs = legacy_qs.filter(
            Q(semana_pago_real__iexact=exact_week)
            | Q(semana_pago_real__icontains=token)
            | Q(semana_pago_proyectada__iexact=exact_week)
            | Q(semana_pago_proyectada__icontains=token)
        )
    elif week_token:
        legacy_qs = legacy_qs.filter(
            Q(semana_pago_real__icontains=week_token)
            | Q(semana_pago_proyectada__icontains=week_token)
        )

    def _build_legacy_light_rows_for_session(s):
        """
        Legacy liviano:
        - NO arma detalle aquí.
        - Calcula total inicial por subtotal_tecnico * porcentaje.
        - En la página visible se recalcula con items/desglose real.
        """
        legacy_week = (getattr(s, "semana_pago_real", "") or "").strip().upper() or (
            getattr(s, "semana_pago_proyectada", "") or ""
        ).strip().upper()

        if not legacy_week:
            return []

        if not _match_week_filter(legacy_week, exact_week, week_token):
            return []

        try:
            asignaciones = list(s.tecnicos_sesion.all())
        except Exception:
            asignaciones = []

        if not asignaciones:
            return []

        subtotal_sesion = _to_decimal(getattr(s, "subtotal_tecnico", 0))

        out = []

        for asig in asignaciones:
            tecnico = getattr(asig, "tecnico", None)

            if not tecnico:
                continue

            if not _match_tech_filter(tecnico, f_tech):
                continue

            pct = _to_decimal(getattr(asig, "porcentaje", 0))
            total_tecnico = (subtotal_sesion * (pct / Decimal("100"))).quantize(
                Decimal("0.01")
            )

            out.append(
                {
                    "_source": "legacy",
                    "sesion": s,
                    "tecnico": tecnico,
                    "project_id": s.proyecto_id,
                    "week": legacy_week or "—",
                    "status": s.estado,
                    "is_discount": bool(getattr(s, "is_direct_discount", False)),
                    "client": s.cliente,
                    "city": s.ciudad,
                    "project": s.proyecto,
                    "office": s.oficina,
                    "real_week": legacy_week or "—",
                    "proj_week": s.semana_pago_proyectada or "—",
                    "total_tecnico": total_tecnico,
                    "detalle": [],
                    "adjustment_type": "",
                }
            )

        return out

    for s in legacy_qs.iterator(chunk_size=500):
        filas.extend(_build_legacy_light_rows_for_session(s))

    # ============================================================
    # 3) Ajustes manuales
    # ============================================================
    if AdjustmentEntry is not None:
        adj_qs = AdjustmentEntry.objects.select_related("technician")

        if not can_view_legacy_history:
            if allowed_keys:
                adj_qs = adj_qs.filter(
                    Q(project__in=allowed_keys) | Q(project_id__in=allowed_keys)
                )
            else:
                adj_qs = AdjustmentEntry.objects.none()

        if exact_week:
            token = exact_week.split("-", 1)[-1].upper()
            adj_qs = adj_qs.filter(
                Q(week__iexact=exact_week) | Q(week__icontains=token)
            )
        elif week_token:
            adj_qs = adj_qs.filter(week__icontains=week_token)

        if f_project:
            adj_qs = adj_qs.annotate(
                project_id_str=Cast("project_id", CharField())
            ).filter(
                Q(project_id_str__icontains=f_project) | Q(project__icontains=f_project)
            )

        if f_client:
            adj_qs = adj_qs.filter(client__icontains=f_client)

        if f_tech:
            adj_qs = adj_qs.filter(
                Q(technician__first_name__icontains=f_tech)
                | Q(technician__last_name__icontains=f_tech)
                | Q(technician__username__icontains=f_tech)
            )

        for a in adj_qs.iterator(chunk_size=500):
            t = a.technician

            amt = _to_decimal(getattr(a, "amount", 0))
            signed_amount = amt.copy_abs()

            filas.append(
                {
                    "_source": "adjustment",
                    "sesion": None,
                    "tecnico": t,
                    "project_id": "-",
                    "week": a.week or "—",
                    "status": "",
                    "is_discount": False,
                    "client": a.client,
                    "city": a.city,
                    "project": a.project,
                    "office": a.office,
                    "real_week": a.week or "—",
                    "proj_week": a.week or "—",
                    "total_tecnico": signed_amount,
                    "detalle": [],
                    "adjustment_type": a.adjustment_type,
                    "adjustment_id": a.id,
                }
            )

    # ============================================================
    # Resolver label de proyecto
    # ============================================================
    def _resolve_project_label(row):
        s = row.get("sesion")
        proj_text = None
        proj_id = None

        if s is not None:
            proj_text = (getattr(s, "proyecto", "") or "").strip()
            proj_id = getattr(s, "proyecto_id", None)
        else:
            proj_text = (row.get("project") or "").strip()
            proj_id = row.get("project_id", None)

        proyecto_sel = None

        if proj_text:
            try:
                pid = int(proj_text)
            except (TypeError, ValueError):
                key = proj_text.lower()
                proyecto_sel = by_code.get(key) or by_name.get(key)
            else:
                proyecto_sel = by_id.get(pid)

        if not proyecto_sel and proj_id not in (None, "", "-"):
            try:
                pid2 = int(proj_id)
            except (TypeError, ValueError):
                key2 = str(proj_id).strip().lower()
                proyecto_sel = by_code.get(key2) or by_name.get(key2)
            else:
                proyecto_sel = by_id.get(pid2)

        if proyecto_sel:
            return getattr(proyecto_sel, "nombre", str(proyecto_sel))

        if proj_text:
            return proj_text

        if proj_id not in (None, "", "-"):
            return str(proj_id)

        return ""

    for row in filas:
        row["project_label"] = _resolve_project_label(row)

    # ============================================================
    # Filtro adicional por texto de Project
    # ============================================================
    if f_project:
        needle = f_project.lower()

        def _match_project_text(row):
            return (
                needle in str(row.get("project_id") or "").lower()
                or needle in str(row.get("project") or "").lower()
                or needle in str(row.get("project_label") or "").lower()
            )

        filas = [r for r in filas if _match_project_text(r)]

    # ============================================================
    # Ventana de visibilidad por ProyectoAsignacion
    # ============================================================
    asignaciones = []

    if ProyectoAsignacion is not None:
        try:
            asignaciones = list(
                ProyectoAsignacion.objects.filter(
                    usuario=user,
                    proyecto__in=proyectos_list,
                ).select_related("proyecto")
            )
        except Exception:
            asignaciones = []

    if asignaciones and not can_view_legacy_history:
        access_by_pk = {}

        for a in asignaciones:
            if a.include_history or not a.start_at:
                access_by_pk[a.proyecto_id] = {
                    "include_history": True,
                    "start_week": None,
                }
            else:
                access_by_pk[a.proyecto_id] = {
                    "include_history": False,
                    "start_week": _iso_week_str(a.start_at),
                }

        def _project_pk_from_row(row):
            s = row.get("sesion")

            if s is not None:
                raw = getattr(s, "proyecto_id", None)

                if raw not in (None, "", "-"):
                    try:
                        return int(raw)
                    except (TypeError, ValueError):
                        pass

            text = (
                str(row.get("project_label") or "").strip()
                or str(row.get("project") or "").strip()
            )

            key = text.lower()

            if key:
                p = by_name.get(key)
                if p:
                    return p.id

                p = by_code.get(key)
                if p:
                    return p.id

            return None

        def _row_allowed(row):
            pk = _project_pk_from_row(row)

            if pk is None:
                return False

            access = access_by_pk.get(pk)

            if not access:
                return False

            if access["include_history"] or access["start_week"] is None:
                return True

            week_str = _normalize_week_str(row.get("real_week"))

            if not week_str:
                return False

            return _week_sort_key(week_str) >= _week_sort_key(access["start_week"])

        filas = [r for r in filas if _row_allowed(r)]

    # ============================================================
    # Orden final
    # ============================================================
    filas.sort(
        key=lambda r: (
            _week_sort_key(r.get("real_week") or r.get("week") or ""),
            str(getattr(r.get("tecnico"), "first_name", "") or ""),
            str(getattr(r.get("tecnico"), "last_name", "") or ""),
            str(getattr(r.get("tecnico"), "username", "") or ""),
        ),
        reverse=True,
    )

    # ============================================================
    # Paginación
    # ============================================================
    cantidad = request.GET.get("cantidad", "10")
    allowed_page_sizes = {"5", "10", "20", "50", "100"}

    if cantidad not in allowed_page_sizes:
        cantidad = "10"

    try:
        per_page = int(cantidad)
    except ValueError:
        per_page = 10

    paginator = Paginator(filas, per_page)
    page_number = request.GET.get("page") or 1
    pagina = paginator.get_page(page_number)

    # ============================================================
    # Enriquecer SOLO legacy visible con detalle real
    # ============================================================
    visible_rows = list(pagina.object_list)

    legacy_session_ids = {
        row["sesion"].id
        for row in visible_rows
        if row.get("_source") == "legacy" and row.get("sesion") is not None
    }

    legacy_sessions_map = {}

    if legacy_session_ids:
        legacy_full_qs = SesionBilling.objects.filter(
            id__in=legacy_session_ids
        ).prefetch_related(
            Prefetch(
                "items",
                queryset=ItemBilling.objects.prefetch_related(
                    Prefetch(
                        "desglose_tecnico",
                        queryset=ItemBillingTecnico.objects.select_related("tecnico"),
                    )
                ).order_by("id"),
            )
        )

        legacy_sessions_map = {s.id: s for s in legacy_full_qs}

    def _rebuild_legacy_detail_for_row(row):
        s_light = row.get("sesion")
        tecnico = row.get("tecnico")

        if not s_light or not tecnico:
            return row

        s = legacy_sessions_map.get(s_light.id)
        if not s:
            return row

        detalle = []
        total_tecnico = Decimal("0.00")

        try:
            items = list(s.items.all())
        except Exception:
            items = []

        for item in items:
            try:
                desglose_rows = [
                    bd
                    for bd in item.desglose_tecnico.all()
                    if getattr(bd, "tecnico_id", None) == tecnico.id
                ]
            except Exception:
                desglose_rows = []

            for bd in desglose_rows:
                subtotal = _to_decimal(getattr(bd, "subtotal", 0))
                tarifa_efectiva = _to_decimal(getattr(bd, "tarifa_efectiva", 0))

                detalle.append(
                    {
                        "codigo": getattr(item, "codigo_trabajo", "") or "",
                        "tipo": getattr(item, "tipo_trabajo", "") or "",
                        "desc": getattr(item, "descripcion", "") or "",
                        "uom": getattr(item, "unidad_medida", "") or "",
                        "qty": getattr(item, "cantidad", None),
                        "rate_tec": tarifa_efectiva,
                        "subtotal_tec": subtotal,
                    }
                )

                total_tecnico += subtotal

        if detalle:
            row["detalle"] = detalle
            row["total_tecnico"] = total_tecnico

        return row

    for row in visible_rows:
        if row.get("_source") == "legacy":
            _rebuild_legacy_detail_for_row(row)

    pagina.object_list = visible_rows

    # ============================================================
    # Querystring filtros
    # ============================================================
    filters_dict = {
        "f_project": f_project,
        "f_week": f_week_input,
        "f_tech": f_tech,
        "f_client": f_client,
        "cantidad": cantidad,
    }

    filters_qs = urlencode({k: v for k, v in filters_dict.items() if v})

    return render(
        request,
        "operaciones/produccion_admin.html",
        {
            "current_week": current_week,
            "pagina": pagina,
            "cantidad": cantidad,
            "f_project": f_project,
            "f_week_input": f_week_input,
            "f_tech": f_tech,
            "f_client": f_client,
            "filters_qs": filters_qs,
        },
    )


@login_required
@rol_requerido("admin", "supervisor", "pm", "facturacion")
def Exportar_produccion_admin(request):
    """
    Exporta exactamente la misma lógica de produccion_admin:
    - snapshots productivos
    - legacy sin snapshots
    - ajustes
    """
    import re
    from decimal import Decimal

    from django.db.models import CharField, Q
    from django.db.models.functions import Cast
    from django.http import HttpResponse
    from django.utils import timezone
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    from facturacion.models import Proyecto
    from operaciones.models import BillingPayWeekSnapshot, SesionBilling

    try:
        from operaciones.models import AdjustmentEntry
    except Exception:
        AdjustmentEntry = None

    from usuarios.models import ProyectoAsignacion

    def _iso_week_str(dt):
        y, w, _ = dt.isocalendar()
        return f"{y}-W{int(w):02d}"

    def parse_week_query(q: str):
        if not q:
            return (None, None)
        s = q.strip().upper().replace("WEEK", "W").replace(" ", "")
        m = re.fullmatch(r"(\d{4})-?W(\d{1,2})", s)
        if m:
            year, ww = int(m.group(1)), int(m.group(2))
            return (f"{year}-W{ww:02d}", None)
        m = re.fullmatch(r"(?:W)?(\d{1,2})", s)
        if m:
            ww = int(m.group(1))
            return (None, f"W{ww:02d}")
        return (None, None)

    def _normalize_week_str(s: str) -> str:
        if not s:
            return ""
        s = s.replace("\u2013", "-").replace("\u2014", "-")
        s = re.sub(r"\s+", "", s)
        return s.upper()

    def _week_sort_key(week_str: str):
        if not week_str:
            return (-1, -1)
        s = str(week_str).upper().replace("WEEK", "W").replace(" ", "")
        m = re.search(r"(\d{4})-?W(\d{1,2})", s)
        if m:
            return (int(m.group(1)), int(m.group(2)))
        m = re.search(r"W(\d{1,2})", s)
        if m:
            return (0, int(m.group(1)))
        return (-1, -1)

    def _match_week_filter(
        week_real: str, exact_week: str | None, week_token: str | None
    ) -> bool:
        wr = (week_real or "").strip().upper()
        if not wr:
            return False
        if exact_week:
            token = exact_week.split("-", 1)[-1].upper()
            return wr == exact_week or token in wr
        if week_token:
            return week_token in wr
        return True

    def _match_tech_filter(user_obj, f_tech_value: str) -> bool:
        if not f_tech_value:
            return True
        target = f_tech_value.lower()
        full_name = (
            (
                (getattr(user_obj, "first_name", "") or "")
                + " "
                + (getattr(user_obj, "last_name", "") or "")
            )
            .strip()
            .lower()
        )
        username = (getattr(user_obj, "username", "") or "").lower()
        return target in full_name or target in username

    def _status_label_export(
        sesion_estado: str, is_discount: bool, adjustment_type: str = ""
    ) -> str:
        if adjustment_type:
            return {
                "bonus": "Bonus",
                "advance": "Advance",
                "fixed_salary": "Fixed salary",
            }.get(adjustment_type, adjustment_type)

        if is_discount:
            return "Direct discount"

        mapping = {
            "aprobado_pm": "Approved by PM",
            "aprobado_supervisor": "Approved by Supervisor",
            "aprobado_finanzas": "Approved by Finance",
            "rechazado_pm": "Rejected by PM",
            "rechazado_supervisor": "Rejected by Supervisor",
            "en_revision_supervisor": "In Supervisor Review",
            "finalizado": "Finished (pending review)",
            "en_proceso": "In Progress",
            "asignado": "Assigned",
        }
        return mapping.get((sesion_estado or "").lower(), (sesion_estado or ""))

    def _build_legacy_rows_for_session(s):
        legacy_week = (getattr(s, "semana_pago_real", "") or "").strip().upper() or (
            getattr(s, "semana_pago_proyectada", "") or ""
        ).strip().upper()

        if not legacy_week:
            return []

        if not _match_week_filter(legacy_week, exact_week, week_token):
            return []

        try:
            asignaciones = list(
                s.tecnicos_sesion.select_related("tecnico").all().order_by("id")
            )
        except Exception:
            asignaciones = []

        if not asignaciones:
            return []

        out = []

        for asig in asignaciones:
            tecnico = getattr(asig, "tecnico", None)
            if not tecnico:
                continue

            if not _match_tech_filter(tecnico, f_tech):
                continue

            total_tecnico = Decimal("0.00")

            try:
                items = list(
                    s.items.prefetch_related("desglose_tecnico").all().order_by("id")
                )
            except Exception:
                items = []

            for item in items:
                try:
                    desglose_rows = [
                        bd
                        for bd in item.desglose_tecnico.all()
                        if getattr(bd, "tecnico_id", None) == tecnico.id
                    ]
                except Exception:
                    desglose_rows = []

                for bd in desglose_rows:
                    subtotal = getattr(bd, "subtotal", None)
                    try:
                        subtotal = (
                            subtotal
                            if isinstance(subtotal, Decimal)
                            else Decimal(str(subtotal or 0))
                        )
                    except Exception:
                        subtotal = Decimal("0.00")
                    total_tecnico += subtotal

            if total_tecnico == 0:
                try:
                    subtotal_sesion = getattr(s, "subtotal_tecnico", None)
                    subtotal_sesion = (
                        subtotal_sesion
                        if isinstance(subtotal_sesion, Decimal)
                        else Decimal(str(subtotal_sesion or 0))
                    )
                except Exception:
                    subtotal_sesion = Decimal("0.00")

                try:
                    pct = Decimal(str(getattr(asig, "porcentaje", 0) or 0))
                except Exception:
                    pct = Decimal("0.00")

                total_tecnico = (subtotal_sesion * (pct / Decimal("100"))).quantize(
                    Decimal("0.01")
                )

            out.append(
                {
                    "sesion": s,
                    "tecnico": tecnico,
                    "project_id": s.proyecto_id,
                    "week": legacy_week or "—",
                    "status": s.estado,
                    "is_discount": bool(getattr(s, "is_direct_discount", False)),
                    "client": s.cliente,
                    "city": s.ciudad,
                    "project": s.proyecto,
                    "office": s.oficina,
                    "real_week": legacy_week or "—",
                    "total_tecnico": total_tecnico,
                    "adjustment_type": "",
                }
            )

        return out

    estados_ok = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}

    user = request.user
    can_view_legacy_history = user.is_superuser or getattr(
        user, "es_usuario_historial", False
    )

    f_project = (request.GET.get("f_project") or "").strip()
    f_week_input = (request.GET.get("f_week") or "").strip()
    f_tech = (request.GET.get("f_tech") or "").strip()
    f_client = (request.GET.get("f_client") or "").strip()

    exact_week, week_token = parse_week_query(f_week_input)

    try:
        base_proyectos = Proyecto.objects.all()
        if can_view_legacy_history:
            proyectos_user = base_proyectos
        else:
            proyectos_user = filter_queryset_by_access(
                base_proyectos,
                request.user,
                "id",
            )
    except Exception:
        proyectos_user = Proyecto.objects.none()

    if proyectos_user.exists():
        allowed_keys = set()
        for p in proyectos_user:
            nombre = (getattr(p, "nombre", "") or "").strip()
            if nombre:
                allowed_keys.add(nombre)

            codigo = getattr(p, "codigo", None)
            if codigo:
                allowed_keys.add(str(codigo).strip())

            allowed_keys.add(str(p.id).strip())
    else:
        allowed_keys = set()

    qs = (
        SesionBilling.objects.filter(
            Q(estado__in=estados_ok) | Q(is_direct_discount=True)
        )
        .order_by("-creado_en")
        .prefetch_related(
            "tecnicos_sesion__tecnico",
            "pay_week_snapshots__tecnico",
            "pay_week_snapshots__item",
            "items__desglose_tecnico",
        )
        .distinct()
    )

    if not can_view_legacy_history:
        if allowed_keys:
            qs = qs.filter(proyecto__in=allowed_keys)
        else:
            qs = SesionBilling.objects.none()

    if f_project:
        qs = qs.annotate(proyecto_id_str=Cast("proyecto_id", CharField()))
        qs = qs.filter(
            Q(proyecto_id_str__icontains=f_project) | Q(proyecto__icontains=f_project)
        )

    if f_client:
        qs = qs.filter(cliente__icontains=f_client)

    qs = qs.distinct()

    filas = []

    snapshot_field_names = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_field_names
    has_adjustment_of = "adjustment_of" in snapshot_field_names

    for s in qs:
        snaps = list(s.pay_week_snapshots.all())

        productive_snaps = []
        for snap in snaps:
            if has_is_adjustment and bool(getattr(snap, "is_adjustment", False)):
                continue
            if has_adjustment_of and getattr(snap, "adjustment_of_id", None):
                continue
            productive_snaps.append(snap)

        if productive_snaps:
            tech_rows = {}

            for snap in productive_snaps:
                tecnico = getattr(snap, "tecnico", None)
                if not tecnico:
                    continue

                if not _match_tech_filter(tecnico, f_tech):
                    continue

                week_real = (
                    (getattr(snap, "semana_resultado", "") or "").strip().upper()
                )
                if not week_real:
                    continue

                if not _match_week_filter(week_real, exact_week, week_token):
                    continue

                key = (tecnico.id, week_real)

                if key not in tech_rows:
                    tech_rows[key] = {
                        "sesion": s,
                        "tecnico": tecnico,
                        "project_id": s.proyecto_id,
                        "week": week_real or "—",
                        "status": s.estado,
                        "is_discount": bool(getattr(s, "is_direct_discount", False)),
                        "client": s.cliente,
                        "city": s.ciudad,
                        "project": s.proyecto,
                        "office": s.oficina,
                        "real_week": week_real or "—",
                        "total_tecnico": Decimal("0"),
                        "adjustment_type": "",
                    }

                subtotal = (
                    snap.subtotal
                    if isinstance(snap.subtotal, Decimal)
                    else Decimal(str(snap.subtotal or 0))
                )
                tech_rows[key]["total_tecnico"] += subtotal

                if subtotal < 0:
                    tech_rows[key]["is_discount"] = True

            filas.extend(tech_rows.values())

        else:
            filas.extend(_build_legacy_rows_for_session(s))

    if AdjustmentEntry is not None:
        adj_qs = AdjustmentEntry.objects.select_related("technician")

        if not can_view_legacy_history:
            if allowed_keys:
                adj_qs = adj_qs.filter(
                    Q(project__in=allowed_keys) | Q(project_id__in=allowed_keys)
                )
            else:
                adj_qs = AdjustmentEntry.objects.none()

        if exact_week:
            token = exact_week.split("-", 1)[-1].upper()
            adj_qs = adj_qs.filter(
                Q(week__iexact=exact_week) | Q(week__icontains=token)
            )
        elif week_token:
            adj_qs = adj_qs.filter(week__icontains=week_token)

        if f_project:
            adj_qs = adj_qs.annotate(project_id_str=Cast("project_id", CharField()))
            adj_qs = adj_qs.filter(
                Q(project_id_str__icontains=f_project) | Q(project__icontains=f_project)
            )

        if f_client:
            adj_qs = adj_qs.filter(client__icontains=f_client)

        if f_tech:
            target = f_tech
            adj_qs = adj_qs.filter(
                Q(technician__first_name__icontains=target)
                | Q(technician__last_name__icontains=target)
                | Q(technician__username__icontains=target)
            )

        for a in adj_qs:
            t = a.technician
            amt = (
                a.amount
                if isinstance(a.amount, Decimal)
                else Decimal(str(a.amount or 0))
            )
            signed_amount = amt.copy_abs()

            filas.append(
                {
                    "sesion": None,
                    "tecnico": t,
                    "project_id": "-",
                    "week": a.week or "—",
                    "status": "",
                    "is_discount": False,
                    "client": a.client,
                    "city": a.city,
                    "project": a.project,
                    "office": a.office,
                    "real_week": a.week or "—",
                    "total_tecnico": signed_amount,
                    "adjustment_type": a.adjustment_type,
                    "adjustment_id": a.id,
                }
            )

    proyectos_list = list(proyectos_user)
    by_id = {p.id: p for p in proyectos_list}
    by_code = {
        (p.codigo or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "codigo", None)
    }
    by_name = {
        (p.nombre or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "nombre", None)
    }

    def _resolve_project_label(row):
        s = row.get("sesion")
        proj_text = None
        proj_id = None

        if s is not None:
            proj_text = (getattr(s, "proyecto", "") or "").strip()
            proj_id = getattr(s, "proyecto_id", None)
        else:
            proj_text = (row.get("project") or "").strip()
            proj_id = row.get("project_id", None)

        proyecto_sel = None

        if proj_text:
            try:
                pid = int(proj_text)
            except (TypeError, ValueError):
                key = proj_text.lower()
                proyecto_sel = by_code.get(key) or by_name.get(key)
            else:
                proyecto_sel = by_id.get(pid)

        if not proyecto_sel and proj_id not in (None, "", "-"):
            try:
                pid2 = int(proj_id)
            except (TypeError, ValueError):
                key2 = str(proj_id).strip().lower()
                proyecto_sel = by_code.get(key2) or by_name.get(key2)
            else:
                proyecto_sel = by_id.get(pid2)

        if proyecto_sel:
            return getattr(proyecto_sel, "nombre", str(proyecto_sel))

        if proj_text:
            return proj_text
        if proj_id not in (None, "", "-"):
            return str(proj_id)
        return ""

    for row in filas:
        row["project_label"] = _resolve_project_label(row)

    if f_project:
        needle = f_project.lower()

        def _match_project_text(row):
            return (
                needle in str(row.get("project_id") or "").lower()
                or needle in str(row.get("project") or "").lower()
                or needle in str(row.get("project_label") or "").lower()
            )

        filas = [r for r in filas if _match_project_text(r)]

    try:
        asignaciones = list(
            ProyectoAsignacion.objects.filter(
                usuario=request.user, proyecto__in=proyectos_list
            ).select_related("proyecto")
        )
    except Exception:
        asignaciones = []

    if asignaciones and not can_view_legacy_history:
        access_by_pk = {}
        for a in asignaciones:
            if a.include_history or not a.start_at:
                access_by_pk[a.proyecto_id] = {
                    "include_history": True,
                    "start_week": None,
                }
            else:
                access_by_pk[a.proyecto_id] = {
                    "include_history": False,
                    "start_week": _iso_week_str(a.start_at),
                }

        def _project_pk_from_row(row):
            s = row.get("sesion")

            if s is not None:
                raw = getattr(s, "proyecto_id", None)
                if raw not in (None, "", "-"):
                    try:
                        return int(raw)
                    except (TypeError, ValueError):
                        pass

            text = (
                str(row.get("project_label") or "").strip()
                or str(row.get("project") or "").strip()
            )
            key = text.lower()
            if key:
                p = by_name.get(key)
                if p:
                    return p.id
                p = by_code.get(key)
                if p:
                    return p.id
            return None

        def _row_allowed(row):
            pk = _project_pk_from_row(row)
            if pk is None:
                return False

            access = access_by_pk.get(pk)
            if not access:
                return False

            if access["include_history"] or access["start_week"] is None:
                return True

            week_str = _normalize_week_str(row.get("real_week"))
            if not week_str:
                return False

            return _week_sort_key(week_str) >= _week_sort_key(access["start_week"])

        filas = [r for r in filas if _row_allowed(r)]

    filas.sort(key=lambda r: _week_sort_key(r["real_week"]), reverse=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Production"

    headers = [
        "Project ID",
        "Real pay week",
        "Status",
        "Technician",
        "Client",
        "City",
        "Project",
        "Office",
        "Technical Billing",
    ]
    ws.append(headers)

    for r in filas:
        tech = r["tecnico"]
        try:
            tech_name = tech.get_full_name() or tech.username
        except Exception:
            tech_name = getattr(tech, "username", "") or ""

        status = _status_label_export(
            r.get("status", ""),
            r.get("is_discount", False),
            r.get("adjustment_type", ""),
        )

        project_cell = r.get("project_label") or r.get("project") or "-"

        ws.append(
            [
                r.get("project_id", "-") or "-",
                r.get("week", "") or r.get("real_week", ""),
                status or "",
                tech_name,
                r.get("client", "-") or "-",
                r.get("city", "-") or "-",
                project_cell,
                r.get("office", "-") or "-",
                float(r.get("total_tecnico") or 0.0),
            ]
        )

    for col in ws.columns:
        max_len = 0
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(10, max_len + 2), 50)

    last_col = len(headers)
    for col_cells in ws.iter_cols(
        min_col=last_col, max_col=last_col, min_row=2, values_only=False
    ):
        for c in col_cells:
            c.number_format = "#,##0.00"

    resp = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    resp["Content-Disposition"] = 'attachment; filename="production_export.xlsx"'
    wb.save(resp)
    return resp


@login_required
@rol_requerido("usuario")
def produccion_usuario(request):
    """
    Producción del técnico logueado.

    Optimizado:
    - Lee snapshots productivos directo desde BillingPayWeekSnapshot.
    - No carga todas las sesiones con todos sus items/snapshots al abrir.
    - Legacy se carga aparte y solo para el técnico logueado.
    - Ajustes se cargan directo.
    """
    import re
    from decimal import Decimal
    from urllib.parse import urlencode

    from django.core.paginator import Paginator
    from django.db.models import Exists, OuterRef, Q
    from django.utils import timezone

    from facturacion.models import Proyecto
    from operaciones.models import (AdjustmentEntry, BillingPayWeekSnapshot,
                                    SesionBilling)

    tecnico = request.user

    def _iso_week_str(dt):
        y, w, _ = dt.isocalendar()
        return f"{y}-W{int(w):02d}"

    def _parse_iso_week_local(s: str):
        if not s:
            return None
        s = s.strip().upper().replace(" ", "")
        m = re.fullmatch(r"(\d{4})-?W(\d{1,2})", s)
        if not m:
            return None
        return (int(m.group(1)), int(m.group(2)))

    def _week_match(week_value: str, weeks_wanted_set):
        if weeks_wanted_set is None:
            return True
        return (week_value or "").strip().upper() in weeks_wanted_set

    def _build_legacy_rows_for_session(s, weeks_wanted_set=None):
        """
        Fallback legacy para sesiones sin snapshots productivos.
        Solo arma fila del técnico logueado.
        """
        legacy_week = (getattr(s, "semana_pago_real", "") or "").strip().upper() or (
            getattr(s, "semana_pago_proyectada", "") or ""
        ).strip().upper()

        if not legacy_week:
            return []

        if not _week_match(legacy_week, weeks_wanted_set):
            return []

        try:
            asig = (
                s.tecnicos_sesion.select_related("tecnico")
                .filter(tecnico_id=tecnico.id)
                .first()
            )
        except Exception:
            asig = None

        if not asig:
            return []

        detalle = []
        total_tecnico = Decimal("0.00")

        try:
            items = list(
                s.items.prefetch_related("desglose_tecnico").all().order_by("id")
            )
        except Exception:
            items = []

        # 1) Intentar por desglose real
        for item in items:
            try:
                desglose_rows = [
                    bd
                    for bd in item.desglose_tecnico.all()
                    if getattr(bd, "tecnico_id", None) == tecnico.id
                ]
            except Exception:
                desglose_rows = []

            for bd in desglose_rows:
                subtotal = getattr(bd, "subtotal", None)
                try:
                    subtotal = (
                        subtotal
                        if isinstance(subtotal, Decimal)
                        else Decimal(str(subtotal or 0))
                    )
                except Exception:
                    subtotal = Decimal("0.00")

                tarifa_efectiva = getattr(bd, "tarifa_efectiva", None)
                try:
                    tarifa_efectiva = (
                        tarifa_efectiva
                        if isinstance(tarifa_efectiva, Decimal)
                        else Decimal(str(tarifa_efectiva or 0))
                    )
                except Exception:
                    tarifa_efectiva = Decimal("0.00")

                detalle.append(
                    {
                        "codigo": getattr(item, "codigo_trabajo", "") or "",
                        "tipo": getattr(item, "tipo_trabajo", "") or "",
                        "desc": getattr(item, "descripcion", "") or "",
                        "uom": getattr(item, "unidad_medida", "") or "",
                        "qty": getattr(item, "cantidad", None),
                        "rate_tec": tarifa_efectiva,
                        "subtotal_tec": subtotal,
                    }
                )
                total_tecnico += subtotal

        # 2) Fallback por porcentaje
        if not detalle:
            try:
                subtotal_sesion = getattr(s, "subtotal_tecnico", None)
                subtotal_sesion = (
                    subtotal_sesion
                    if isinstance(subtotal_sesion, Decimal)
                    else Decimal(str(subtotal_sesion or 0))
                )
            except Exception:
                subtotal_sesion = Decimal("0.00")

            try:
                pct = Decimal(str(getattr(asig, "porcentaje", 0) or 0))
            except Exception:
                pct = Decimal("0.00")

            total_tecnico = (subtotal_sesion * (pct / Decimal("100"))).quantize(
                Decimal("0.01")
            )

        return [
            {
                "sesion": s,
                "project_id": s.proyecto_id,
                "week": legacy_week or "—",
                "status": s.estado,
                "is_discount": bool(getattr(s, "is_direct_discount", False)),
                "client": s.cliente,
                "city": s.ciudad,
                "project": s.proyecto,
                "office": s.oficina,
                "real_week": legacy_week or "—",
                "total_tecnico": total_tecnico,
                "detalle": detalle,
                "adjustment_type": "",
                "adjustment_label": "",
            }
        ]

    estados_ok = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}
    current_week = _iso_week_str(timezone.now())
    current_tuple = _parse_iso_week_local(current_week)

    week_filter = (request.GET.get("week") or "all").strip()
    weeks_wanted = None if week_filter.lower() == "all" else {week_filter.upper()}

    filas = []
    total_semana_actual = Decimal("0.00")

    snapshot_field_names = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_field_names
    has_adjustment_of = "adjustment_of" in snapshot_field_names

    def _productive_snapshot_filter():
        q = Q()
        if has_is_adjustment:
            q &= Q(is_adjustment=False)
        if has_adjustment_of:
            q &= Q(adjustment_of__isnull=True)
        return q

    # ==========================================================
    # 1) Flujo nuevo: snapshots del técnico directo
    # ==========================================================
    snap_qs = (
        BillingPayWeekSnapshot.objects.select_related("sesion", "item")
        .filter(
            _productive_snapshot_filter(),
            tecnico_id=tecnico.id,
            sesion__isnull=False,
        )
        .filter(
            Q(sesion__estado__in=estados_ok)
            | Q(sesion__is_direct_discount=True)
            | Q(subtotal__lt=0)
        )
        .order_by("-semana_resultado", "-id")
    )

    if weeks_wanted is not None:
        snap_qs = snap_qs.filter(semana_resultado__in=weeks_wanted)

    week_rows = {}

    for snap in snap_qs:
        s = getattr(snap, "sesion", None)
        if not s:
            continue

        rw = (getattr(snap, "semana_resultado", "") or "").strip().upper()
        if not rw:
            continue

        if not _week_match(rw, weeks_wanted):
            continue

        key = (s.id, rw)

        if key not in week_rows:
            week_rows[key] = {
                "sesion": s,
                "project_id": s.proyecto_id,
                "week": rw or "—",
                "status": s.estado,
                "is_discount": bool(getattr(s, "is_direct_discount", False)),
                "client": s.cliente,
                "city": s.ciudad,
                "project": s.proyecto,
                "office": s.oficina,
                "real_week": rw or "—",
                "total_tecnico": Decimal("0.00"),
                "detalle": [],
                "adjustment_type": "",
                "adjustment_label": "",
            }

        try:
            subtotal = (
                snap.subtotal
                if isinstance(snap.subtotal, Decimal)
                else Decimal(str(snap.subtotal or 0))
            )
        except Exception:
            subtotal = Decimal("0.00")

        week_rows[key]["total_tecnico"] += subtotal

        if subtotal < 0:
            week_rows[key]["is_discount"] = True

        item = getattr(snap, "item", None)

        try:
            rate_tec = (
                snap.tarifa_efectiva
                if isinstance(getattr(snap, "tarifa_efectiva", 0), Decimal)
                else Decimal(str(getattr(snap, "tarifa_efectiva", 0) or 0))
            )
        except Exception:
            rate_tec = Decimal("0.00")

        week_rows[key]["detalle"].append(
            {
                "codigo": getattr(snap, "codigo_trabajo", "") or "",
                "tipo": getattr(snap, "tipo_trabajo", "") or "",
                "desc": getattr(item, "descripcion", "") if item else "",
                "uom": getattr(item, "unidad_medida", "") if item else "",
                "qty": getattr(item, "cantidad", None) if item else None,
                "rate_tec": rate_tec,
                "subtotal_tec": subtotal,
            }
        )

    for row in week_rows.values():
        if row["real_week"] == current_week:
            total_semana_actual += row["total_tecnico"]
        filas.append(row)

    # ==========================================================
    # 2) Legacy: solo sesiones sin snapshots productivos y del técnico
    # ==========================================================
    productive_snap_exists = BillingPayWeekSnapshot.objects.filter(
        _productive_snapshot_filter(),
        sesion_id=OuterRef("pk"),
    )

    legacy_qs = (
        SesionBilling.objects.filter(
            Q(estado__in=estados_ok) | Q(is_direct_discount=True)
        )
        .filter(tecnicos_sesion__tecnico_id=tecnico.id)
        .annotate(has_productive_snap=Exists(productive_snap_exists))
        .filter(has_productive_snap=False)
        .prefetch_related(
            "tecnicos_sesion__tecnico",
            "items__desglose_tecnico",
        )
        .order_by("-creado_en")
        .distinct()
    )

    if weeks_wanted is not None:
        legacy_qs = legacy_qs.filter(
            Q(semana_pago_real__in=weeks_wanted)
            | Q(semana_pago_proyectada__in=weeks_wanted)
        )

    for s in legacy_qs:
        legacy_rows = _build_legacy_rows_for_session(s, weeks_wanted)

        for row in legacy_rows:
            if row["real_week"] == current_week:
                total_semana_actual += row["total_tecnico"]
            filas.append(row)

    # ==========================================================
    # 3) Ajustes del técnico
    # ==========================================================
    adj_qs = AdjustmentEntry.objects.filter(technician=tecnico)

    if weeks_wanted is not None:
        adj_qs = adj_qs.filter(week__in=weeks_wanted)

    for a in adj_qs:
        try:
            amt = (
                a.amount
                if isinstance(a.amount, Decimal)
                else Decimal(str(a.amount or 0))
            )
        except Exception:
            amt = Decimal("0.00")

        amt_pos = abs(amt)
        rw = (a.week or "—").upper()

        if rw == current_week:
            total_semana_actual += amt_pos

        filas.append(
            {
                "sesion": None,
                "project_id": a.project_id or "",
                "week": a.week or "—",
                "status": "",
                "is_discount": False,
                "client": a.client or "-",
                "city": a.city or "-",
                "project": a.project or "-",
                "office": a.office or "-",
                "real_week": rw,
                "total_tecnico": amt_pos,
                "detalle": [],
                "adjustment_type": a.adjustment_type,
                "adjustment_label": a.get_adjustment_type_display(),
            }
        )

    # ==========================================================
    # Resolver Project label
    # ==========================================================
    proyectos_list = list(Proyecto.objects.all())
    by_id = {p.id: p for p in proyectos_list}
    by_code = {
        (p.codigo or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "codigo", None)
    }
    by_name = {
        (p.nombre or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "nombre", None)
    }

    def _resolve_project_label(row):
        s = row.get("sesion")
        proj_text = None
        proj_id = None

        if s is not None:
            proj_text = (getattr(s, "proyecto", "") or "").strip()
            proj_id = getattr(s, "proyecto_id", None)
        else:
            proj_text = (row.get("project") or "").strip()
            proj_id = row.get("project_id", None)

        proyecto_sel = None

        if proj_text:
            try:
                pid = int(proj_text)
            except (TypeError, ValueError):
                key = proj_text.lower()
                proyecto_sel = by_code.get(key) or by_name.get(key)
            else:
                proyecto_sel = by_id.get(pid)

        if not proyecto_sel and proj_id not in (None, "", "-"):
            try:
                pid2 = int(proj_id)
            except (TypeError, ValueError):
                key2 = str(proj_id).strip().lower()
                proyecto_sel = by_code.get(key2) or by_name.get(key2)
            else:
                proyecto_sel = by_id.get(pid2)

        if proyecto_sel:
            return getattr(proyecto_sel, "nombre", str(proyecto_sel))

        if proj_text:
            return proj_text
        if proj_id not in (None, "", "-"):
            return str(proj_id)
        return ""

    for row in filas:
        row["project_label"] = _resolve_project_label(row)

    def sort_key(row):
        t = _parse_iso_week_local(row["real_week"])

        if t is None:
            return (3, 9999, 99)

        if current_tuple and t == current_tuple:
            return (0, 0, 0)

        if current_tuple and t < current_tuple:
            return (1, -t[0], -t[1])

        return (2, t[0], t[1])

    filas.sort(key=sort_key)

    cantidad = (request.GET.get("cantidad") or "10").strip().lower()

    if cantidad != "todos":
        try:
            per_page = max(5, min(int(cantidad), 100))
        except ValueError:
            per_page = 10
            cantidad = "10"

        paginator = Paginator(filas, per_page)
        page_number = request.GET.get("page") or 1
        pagina = paginator.get_page(page_number)
    else:

        class _OnePage:
            number = 1
            has_previous = False
            has_next = False
            object_list = filas

            @property
            def paginator(self):
                class P:
                    num_pages = 1

                return P()

        pagina = _OnePage()

    keep = {"week": week_filter, "cantidad": cantidad}
    filters_qs = urlencode({k: v for k, v in keep.items() if v})

    return render(
        request,
        "operaciones/produccion_usuario.html",
        {
            "pagina": pagina,
            "cantidad": cantidad,
            "current_week": current_week,
            "total_semana_actual": total_semana_actual,
            "week_filter": week_filter,
            "filters_qs": filters_qs,
        },
    )


def _s3_client():
    """
    Wasabi S3 en path-style para evitar problemas de CORS/SSL.
    Usa el endpoint REGIONAL del bucket (p.ej. us-east-1).
    """
    return boto3.client(
        "s3",
        endpoint_url=getattr(settings, "AWS_S3_ENDPOINT_URL",
                             "https://s3.us-east-1.wasabisys.com"),
        region_name=getattr(settings, "AWS_S3_REGION_NAME", "us-east-1"),
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        config=Config(signature_version="s3v4", s3={
                      "addressing_style": "path"}),
        verify=getattr(settings, "AWS_S3_VERIFY", True),
    )


ESTADOS_OK = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}


@transaction.atomic
def _sync_weekly_totals(week: str | None = None, create_missing: bool = False) -> dict:
    """
    Sincroniza WeeklyPayment usando:
    - BillingPayWeekSnapshot SOLO si el billing está aprobado por supervisor/PM/finanzas
      o si es direct discount.
    - sesiones legacy sin snapshots productivos aprobadas o direct discount
    - ajustes manuales

    Regla importante:
    - Un billing normal con semana real Wxx NO cuenta como producción si todavía está:
      asignado, en_proceso, finalizado, en_revision_supervisor, rechazado_*, etc.
    """

    dec0 = Value(
        Decimal("0.00"),
        output_field=DecimalField(max_digits=18, decimal_places=2),
    )

    ESTADOS_OK_SYNC = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}

    snapshot_model = BillingPayWeekSnapshot
    snapshot_fields = {f.name for f in snapshot_model._meta.get_fields()}

    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    # ==========================================================
    # 1) PRODUCCIÓN NUEVA (BillingPayWeekSnapshot)
    #    SOLO billings aprobados o direct discount
    # ==========================================================
    snap_qs = snapshot_model.objects.filter(
        semana_resultado__gt="",
        sesion__isnull=False,
    )

    if week:
        snap_qs = snap_qs.filter(semana_resultado=week)

    if has_is_adjustment:
        snap_qs = snap_qs.filter(is_adjustment=False)
    elif has_adjustment_of:
        snap_qs = snap_qs.filter(adjustment_of__isnull=True)

    # ✅ FIX CLAVE:
    # No contar snapshots de billings normales si el billing no está aprobado.
    snap_qs = snap_qs.filter(
        Q(sesion__estado__in=ESTADOS_OK_SYNC) | Q(sesion__is_direct_discount=True)
    )

    agg_snaps = snap_qs.values("tecnico_id", "semana_resultado").annotate(
        total=Coalesce(
            Sum("subtotal"),
            dec0,
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )
    )

    items_totals = {
        (r["tecnico_id"], r["semana_resultado"]): (r["total"] or Decimal("0.00"))
        for r in agg_snaps
    }

    # ==========================================================
    # 2) PRODUCCIÓN LEGACY
    #    Tu helper ya filtra aprobados o direct discount
    # ==========================================================
    legacy_totals = _legacy_session_totals_without_snapshots(week=week)

    # ==========================================================
    # 3) AJUSTES MANUALES
    # ==========================================================
    try:
        from operaciones.models import AdjustmentEntry
    except Exception:
        AdjustmentEntry = None

    adj_totals = {}
    if AdjustmentEntry is not None:
        adj_qs = AdjustmentEntry.objects.all()

        if week:
            adj_qs = adj_qs.filter(week=week)

        agg_adj = adj_qs.values("technician_id", "week").annotate(
            total=Coalesce(
                Sum("amount"),
                dec0,
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )

        adj_totals = {
            (r["technician_id"], r["week"]): (r["total"] or Decimal("0.00"))
            for r in agg_adj
        }

    # ==========================================================
    # 4) SUMA FINAL POR (technician, week)
    # ==========================================================
    from collections import defaultdict

    merged = defaultdict(lambda: Decimal("0.00"))

    for k, v in items_totals.items():
        merged[k] += v

    for k, v in legacy_totals.items():
        merged[k] += v

    for k, v in adj_totals.items():
        merged[k] += v

    prod_totals = {k: v for k, v in merged.items() if v != 0}

    updated = 0
    deleted = 0
    created = 0

    # ==========================================================
    # 5) ACTUALIZA / ELIMINA EXISTENTES
    # ==========================================================
    weekly_qs = WeeklyPayment.objects.select_for_update()

    if week:
        weekly_qs = weekly_qs.filter(week=week)

    for wp in weekly_qs:
        key = (wp.technician_id, wp.week)

        if key not in prod_totals:
            # Si todavía no está pagado, se elimina porque ya no tiene producción válida.
            if wp.status != "paid":
                wp.delete()
                deleted += 1
            continue

        total = prod_totals.pop(key)

        if wp.amount != total:
            wp.amount = total
            save_fields = ["amount", "updated_at"]

            if wp.status == "approved_user":
                wp.status = "pending_payment"
                save_fields.append("status")

            wp.save(update_fields=save_fields)
            updated += 1

    # ==========================================================
    # 6) CREA FALTANTES
    # ==========================================================
    if create_missing and prod_totals:
        to_create = [
            WeeklyPayment(
                technician_id=tech_id,
                week=w,
                amount=total,
                status="pending_user",
            )
            for (tech_id, w), total in prod_totals.items()
            if (not week) or (w == week)
        ]

        WeeklyPayment.objects.bulk_create(to_create, ignore_conflicts=True)
        created = len(to_create)

    return {
        "updated": updated,
        "deleted": deleted,
        "created": created,
    }


@login_required
@rol_requerido("admin", "pm", "facturacion")
@never_cache
def admin_weekly_payments(request):
    """
    Pagos semanales:
    - TOP: NO pagados (hasta la semana actual, incluye semanas anteriores). Muestra desglose dinámico.
    - Bottom (Paid): historial con filtros + paginación.

    IMPORTANTE:
    - Para PAID history, el monto y desglose deben salir del snapshot congelado
      del WeeklyPayment, no de billings vivos.
    """
    import re
    from decimal import Decimal
    from urllib.parse import urlencode

    from django.core.paginator import Paginator
    from django.db.models import DecimalField, Q, Sum, Value
    from django.db.models.functions import Coalesce
    from django.utils import timezone

    from facturacion.models import Proyecto
    from operaciones.models import BillingPayWeekSnapshot, WeeklyPayment
    from usuarios.models import ProyectoAsignacion

    def _norm_week_input(raw: str) -> str:
        s = (raw or "").strip().upper()
        if not s:
            return ""
        m_year = re.match(r"^(\d{4})[- ]?W?(\d{1,2})$", s)
        if m_year:
            yy = int(m_year.group(1))
            ww = int(m_year.group(2))
            return f"{yy}-W{ww:02d}"
        y, w, _ = timezone.localdate().isocalendar()
        m_now = re.match(r"^W?(\d{1,2})$", s)
        if m_now:
            ww = int(m_now.group(1))
            return f"{y}-W{ww:02d}"
        return s

    def _dec0():
        return Value(
            Decimal("0.00"),
            output_field=DecimalField(max_digits=18, decimal_places=2),
        )

    def _iso_week_str(dt):
        y, w, _ = dt.isocalendar()
        return f"{y}-W{int(w):02d}"

    def _normalize_week_str(s: str) -> str:
        if not s:
            return ""
        s = s.replace("\u2013", "-").replace("\u2014", "-")
        s = re.sub(r"\s+", "", s)
        return s.upper()

    def _week_sort_key(week_str: str):
        if not week_str:
            return (-1, -1)

        s = str(week_str).upper().replace("WEEK", "W").replace(" ", "")
        m = re.search(r"(\d{4})-?W(\d{1,2})", s)
        if m:
            return (int(m.group(1)), int(m.group(2)))
        m = re.search(r"W(\d{1,2})", s)
        if m:
            return (0, int(m.group(1)))
        return (-1, -1)

    ADJ_LABEL = {
        "fixed_salary": "Fixed salary",
        "bonus": "Bonus",
        "advance": "Advance",
    }
    ESTADOS_OK = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}

    y, w, _ = timezone.localdate().isocalendar()
    current_week = f"{y}-W{int(w):02d}"

    _sync_weekly_totals(week=current_week, create_missing=True)

    visible_tech_ids = _visible_tech_ids_for_user(request.user)

    user = request.user
    can_view_legacy_history = user.is_superuser or getattr(
        user, "es_usuario_historial", False
    )

    try:
        base_proyectos = Proyecto.objects.all()
        if can_view_legacy_history:
            proyectos_user = base_proyectos
        else:
            proyectos_user = filter_queryset_by_access(
                base_proyectos,
                request.user,
                "id",
            )
    except Exception:
        proyectos_user = Proyecto.objects.none()

    if proyectos_user.exists():
        allowed_keys = set()
        for p in proyectos_user:
            nombre = (getattr(p, "nombre", "") or "").strip()
            if nombre:
                allowed_keys.add(nombre)
            codigo = getattr(p, "codigo", None)
            if codigo:
                allowed_keys.add(str(codigo).strip())
            allowed_keys.add(str(p.id).strip())
    else:
        allowed_keys = set()

    proyectos_list = list(proyectos_user)
    by_id = {p.id: p for p in proyectos_list}
    by_code = {
        (p.codigo or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "codigo", None)
    }
    by_name = {
        (p.nombre or "").strip().lower(): p
        for p in proyectos_list
        if getattr(p, "nombre", None)
    }

    try:
        asignaciones = list(
            ProyectoAsignacion.objects.filter(
                usuario=request.user, proyecto__in=proyectos_list
            ).select_related("proyecto")
        )
    except Exception:
        asignaciones = []

    access_by_pk = {}
    if asignaciones and not can_view_legacy_history:
        for a in asignaciones:
            if a.include_history or not a.start_at:
                access_by_pk[a.proyecto_id] = {
                    "include_history": True,
                    "start_week": None,
                }
            else:
                access_by_pk[a.proyecto_id] = {
                    "include_history": False,
                    "start_week": _iso_week_str(a.start_at),
                }

    def _project_pk_from_any(raw_proj_id, raw_proj_text):
        if raw_proj_id not in (None, "", "-"):
            try:
                pid = int(raw_proj_id)
            except (TypeError, ValueError):
                pid = None
            else:
                if pid in by_id:
                    return pid

        key = (str(raw_proj_text or "").strip()).lower()
        if key:
            p = by_name.get(key) or by_code.get(key)
            if p:
                return p.id
        return None

    def _allowed_by_window(project_pk: int | None, week_str: str) -> bool:
        if can_view_legacy_history:
            return True
        if not access_by_pk:
            return False
        if project_pk is None:
            return False
        access = access_by_pk.get(project_pk)
        if not access:
            return False
        if access["include_history"] or access["start_week"] is None:
            return True
        wk = _normalize_week_str(week_str)
        if not wk:
            return False
        return _week_sort_key(wk) >= _week_sort_key(access["start_week"])

    snapshot_field_names = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_field_names
    has_adjustment_of = "adjustment_of" in snapshot_field_names

    # ==========================================================
    # TOP (NO PAID) -> dinámico
    # ==========================================================
    top_qs = (
        WeeklyPayment.objects.filter(week__lte=current_week, amount__gt=0)
        .exclude(status="paid")
        .select_related("technician")
        .order_by("-week", "status", "technician__first_name", "technician__last_name")
    )

    if visible_tech_ids is not None:
        top_qs = top_qs.filter(technician_id__in=visible_tech_ids)

    top = list(top_qs)

    tech_ids_top = {wp.technician_id for wp in top}
    weeks_top = {wp.week for wp in top}
    details_map_top: dict[tuple[int, str], list] = {}

    if tech_ids_top and weeks_top:
        snap_qs = BillingPayWeekSnapshot.objects.filter(
            tecnico_id__in=tech_ids_top,
            semana_resultado__in=weeks_top,
            sesion__isnull=False,
        ).select_related("sesion")

        if has_is_adjustment:
            snap_qs = snap_qs.filter(is_adjustment=False)
        elif has_adjustment_of:
            snap_qs = snap_qs.filter(adjustment_of__isnull=True)

        snap_qs = snap_qs.filter(
            Q(sesion__estado__in=ESTADOS_OK)
            | Q(sesion__is_direct_discount=True)
            | Q(subtotal__lt=0)
        )

        if not can_view_legacy_history:
            if allowed_keys:
                snap_qs = snap_qs.filter(sesion__proyecto__in=allowed_keys)
            else:
                snap_qs = BillingPayWeekSnapshot.objects.none()

        det_prod = (
            snap_qs.values(
                "tecnico_id",
                "semana_resultado",
                "sesion__proyecto_id",
                "sesion__proyecto",
                "sesion__is_direct_discount",
            )
            .annotate(
                subtotal=Coalesce(
                    Sum("subtotal"),
                    _dec0(),
                    output_field=DecimalField(max_digits=18, decimal_places=2),
                )
            )
            .order_by("semana_resultado", "sesion__proyecto_id")
        )

        for r in det_prod:
            week_real = r["semana_resultado"]
            proj_pk = _project_pk_from_any(
                r.get("sesion__proyecto_id"), r.get("sesion__proyecto")
            )
            if not _allowed_by_window(proj_pk, week_real):
                continue

            key = (r["tecnico_id"], week_real)
            label = (
                "Direct discount"
                if r["sesion__is_direct_discount"]
                else (r["sesion__proyecto_id"] or "—")
            )
            details_map_top.setdefault(key, []).append(
                {
                    "project_label": str(label),
                    "subtotal": r["subtotal"] or Decimal("0.00"),
                }
            )

    if tech_ids_top and weeks_top:
        legacy_top = _legacy_weekly_payment_details(
            tech_ids=tech_ids_top,
            weeks=weeks_top,
            allowed_project_keys=allowed_keys,
            can_view_all_projects=can_view_legacy_history,
        )

        for key, rows in legacy_top.items():
            tech_id, wk = key
            visible_rows = []

            for row in rows:
                proj_pk = _project_pk_from_any(
                    row.get("project_lookup_id"),
                    row.get("project_lookup_text"),
                )

                if row.get("project_label") == "Direct discount" or _allowed_by_window(
                    proj_pk, wk
                ):
                    visible_rows.append(
                        {
                            "project_label": row.get("project_label"),
                            "subtotal": row.get("subtotal") or Decimal("0.00"),
                        }
                    )

            if visible_rows:
                details_map_top.setdefault(key, []).extend(visible_rows)

    try:
        from operaciones.models import AdjustmentEntry
    except Exception:
        AdjustmentEntry = None

    if AdjustmentEntry is not None and tech_ids_top and weeks_top:
        det_adj = AdjustmentEntry.objects.filter(
            technician_id__in=tech_ids_top, week__in=weeks_top
        )
        if not can_view_legacy_history:
            if allowed_keys:
                det_adj = det_adj.filter(
                    Q(project__in=allowed_keys) | Q(project_id__in=allowed_keys)
                )
            else:
                det_adj = AdjustmentEntry.objects.none()

        det_adj = det_adj.values(
            "technician_id", "week", "adjustment_type", "project_id", "project"
        ).annotate(
            total=Coalesce(
                Sum("amount"),
                _dec0(),
                output_field=DecimalField(max_digits=18, decimal_places=2),
            )
        )

        for r in det_adj:
            wk = r["week"]
            proj_pk = _project_pk_from_any(r.get("project_id"), r.get("project"))
            if not _allowed_by_window(proj_pk, wk):
                continue

            key = (r["technician_id"], wk)
            label = ADJ_LABEL.get(r["adjustment_type"], r["adjustment_type"])
            details_map_top.setdefault(key, []).append(
                {
                    "project_label": label,
                    "subtotal": r["total"] or Decimal("0.00"),
                }
            )

    top_filtered = []
    for wp in top:
        wp.details = details_map_top.get((wp.technician_id, wp.week), [])
        visible_total = (
            sum((d.get("subtotal") or Decimal("0.00")) for d in wp.details)
            if wp.details
            else Decimal("0.00")
        )
        wp.amount = visible_total
        if visible_total > 0:
            top_filtered.append(wp)
    top = top_filtered

    # ==========================================================
    # BOTTOM (PAID HISTORY) -> snapshot congelado
    # ==========================================================
    f_tech = (request.GET.get("f_tech") or "").strip()
    f_week_input = (request.GET.get("f_week") or "").strip()
    f_paid_week_input = (request.GET.get("f_paid_week") or "").strip()
    f_receipt = (request.GET.get("f_receipt") or "").strip()

    f_week = _norm_week_input(f_week_input)
    f_paid_week = _norm_week_input(f_paid_week_input)

    bottom_qs = WeeklyPayment.objects.filter(status="paid").select_related("technician")

    if visible_tech_ids is not None:
        bottom_qs = bottom_qs.filter(technician_id__in=visible_tech_ids)

    if f_tech:
        bottom_qs = bottom_qs.filter(
            Q(technician__first_name__icontains=f_tech)
            | Q(technician__last_name__icontains=f_tech)
            | Q(technician__username__icontains=f_tech)
        )
    if f_week:
        bottom_qs = bottom_qs.filter(week=f_week)
    if f_paid_week:
        bottom_qs = bottom_qs.filter(paid_week=f_paid_week)
    if f_receipt == "with":
        bottom_qs = bottom_qs.exclude(Q(receipt__isnull=True) | Q(receipt=""))
    elif f_receipt == "without":
        bottom_qs = bottom_qs.filter(Q(receipt__isnull=True) | Q(receipt=""))

    bottom_qs = bottom_qs.order_by(
        "-paid_week",
        "-week",
        "technician__first_name",
        "technician__last_name",
    )

    cantidad = (request.GET.get("cantidad") or "10").strip().lower()
    page_number = request.GET.get("page") or "1"

    if cantidad == "todos":
        pagina = list(bottom_qs)
    else:
        try:
            per_page = max(1, min(100, int(cantidad)))
        except ValueError:
            per_page = 10
            cantidad = "10"
        paginator = Paginator(bottom_qs, per_page)
        pagina = paginator.get_page(page_number)

    wp_list = list(pagina) if not isinstance(pagina, list) else pagina

    for wp in wp_list:
        snapshot_lines = []
        if hasattr(wp, "paid_breakdown_snapshot") and wp.paid_breakdown_snapshot:
            for row in wp.paid_breakdown_snapshot or []:
                snapshot_lines.append(
                    {
                        "project_label": row.get("project_label") or "—",
                        "subtotal": _deserialize_decimal_from_json(row.get("subtotal")),
                        "source": row.get("source") or "",
                    }
                )

        wp.details = snapshot_lines

        if hasattr(wp, "paid_amount_snapshot") and wp.paid_amount_snapshot is not None:
            wp.amount = wp.paid_amount_snapshot
        else:
            # fallback defensivo para registros viejos sin snapshot
            wp.amount = wp.amount or Decimal("0.00")

    if isinstance(pagina, list):
        pagina = wp_list
    else:
        pagina.object_list = wp_list

    keep = {
        "f_tech": f_tech,
        "f_week": f_week_input,
        "f_paid_week": f_paid_week_input,
        "f_receipt": f_receipt,
        "cantidad": cantidad,
    }
    filters_qs = urlencode({k: v for k, v in keep.items() if v})

    return render(
        request,
        "operaciones/pagos_admin_list.html",
        {
            "current_week": current_week,
            "top": top,
            "pagina": pagina,
            "cantidad": cantidad,
            "filters_qs": filters_qs,
            "f_tech": f_tech,
            "f_week_input": f_week_input,
            "f_paid_week_input": f_paid_week_input,
            "f_receipt": f_receipt,
        },
    )


@login_required
@rol_requerido("admin", "pm", "facturacion")
@require_POST
@transaction.atomic
def admin_unpay(request, pk: int):
    wp = get_object_or_404(WeeklyPayment, pk=pk)

    if wp.status != "paid":
        messages.info(request, "Only PAID items can be reverted.")
        return redirect("operaciones:admin_weekly_payments")

    _unmark_billings_paid_for_weekly_payment(wp)

    try:
        if wp.receipt:
            wp.receipt.delete(save=False)
    except Exception:
        pass

    wp.receipt = None
    wp.paid_week = ""
    wp.status = "pending_payment"
    wp.save(update_fields=["receipt", "paid_week", "status", "updated_at"])

    messages.success(request, "Payment reverted. It is now pending again.")
    return redirect("operaciones:admin_weekly_payments")


def _is_admin(user) -> bool:
    """
    Admin real del sistema.
    Compatible con:
    - superuser
    - es_admin_general
    - rol == 'admin'
    """
    if not user:
        return False

    if getattr(user, "is_superuser", False):
        return True

    if getattr(user, "es_admin_general", False):
        return True

    if (getattr(user, "rol", "") or "").strip().lower() == "admin":
        return True

    return False


@login_required
@rol_requerido("admin", "pm", "facturacion")
@csrf_protect
@require_POST
def billing_set_real_week(request, pk: int):
    """
    Actualiza la semana real efectiva de una sesión moviendo los
    BillingPayWeekSnapshot.semana_resultado de la sesión.

    - Si hay pagos PAID relacionados, SOLO admin puede modificar.
    - Recalcula sesion.semana_pago_real como semana resumen.
    - Re-sincroniza WeeklyPayment alrededor del cambio.
    """
    sesion = get_object_or_404(SesionBilling, pk=pk)

    if not access_user_can(request.user, "billing.edit_real_week"):
        return JsonResponse(
            {
                "ok": False,
                "error": "FORBIDDEN",
                "message": "You do not have permission to edit the real pay week.",
            },
            status=403,
        )

    new_week = (request.POST.get("week") or "").strip().upper()
    if not new_week:
        return JsonResponse({"ok": False, "error": "MISSING_WEEK"}, status=400)

    if not WEEK_RE.match(new_week):
        return JsonResponse(
            {
                "ok": False,
                "error": "INVALID_WEEK_FORMAT",
                "message": "Invalid format. Use YYYY-W##.",
            },
            status=400,
        )

    year, week = _parse_iso_week(new_week)
    if not year or not week:
        return JsonResponse(
            {"ok": False, "error": "INVALID_ISO_WEEK", "message": "Invalid ISO week."},
            status=400,
        )

    is_admin = _is_admin(request.user)

    # ¿Bloqueada por pagos 'PAID'?
    if _session_is_paid_locked(sesion) and not is_admin:
        return JsonResponse(
            {
                "ok": False,
                "error": "LOCKED_PAID",
                "message": "This session has PAID weekly payments. Only admins can change the real pay week.",
            },
            status=403,
        )

    snapshot_model = BillingPayWeekSnapshot
    snapshot_fields = {f.name for f in snapshot_model._meta.get_fields()}

    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    def _productivo_qs():
        qs = snapshot_model.objects.filter(sesion=sesion)
        if has_is_adjustment:
            qs = qs.filter(is_adjustment=False)
        elif has_adjustment_of:
            qs = qs.filter(adjustment_of__isnull=True)
        return qs

    # semanas anteriores afectadas para resincronizar
    old_weeks = set(
        snapshot_model.objects.filter(sesion=sesion)
        .exclude(semana_resultado__isnull=True)
        .exclude(semana_resultado__exact="")
        .values_list("semana_resultado", flat=True)
    )

    snaps = list(_productivo_qs())

    # Compatibilidad legacy si no existen snapshots productivos
    if not snaps:
        old_week = (sesion.semana_pago_real or "").upper()
        sesion.semana_pago_real = new_week
        sesion.save(update_fields=["semana_pago_real"])

        try:
            weeks_to_sync = set(filter(None, [old_week, new_week]))
            for wk in weeks_to_sync:
                _sync_weekly_totals(week=wk)
        except Exception:
            pass

        return JsonResponse({"ok": True, "week": new_week})

    changed_weeks = set()

    with transaction.atomic():
        for snap in snaps:
            payment_weeks = int(getattr(snap, "payment_weeks", 0) or 0)
            recalculated_week = _add_weeks_to_iso_week(new_week, payment_weeks)

            update_fields = []

            if (snap.semana_base or "") != new_week:
                snap.semana_base = new_week
                update_fields.append("semana_base")

            if (snap.semana_resultado or "") != recalculated_week:
                snap.semana_resultado = recalculated_week
                update_fields.append("semana_resultado")

            if update_fields:
                snap.save(update_fields=update_fields)

            if recalculated_week:
                changed_weeks.add(recalculated_week)

        # semana resumen = semana_resultado más alta de snapshots productivos
        summary_week = ""
        parsed = []

        all_result_weeks = list(
            _productivo_qs()
            .exclude(semana_resultado__isnull=True)
            .exclude(semana_resultado__exact="")
            .values_list("semana_resultado", flat=True)
        )

        for w in all_result_weeks:
            y, wk = _parse_iso_week(w)
            if y and wk:
                parsed.append((y, wk, w))

        if parsed:
            parsed.sort(key=lambda x: (x[0], x[1]))
            summary_week = parsed[-1][2]

        sesion.semana_pago_real = summary_week
        sesion.save(update_fields=["semana_pago_real"])

    # resincronizar semanas viejas y nuevas
    try:
        weeks_to_sync = set(filter(None, old_weeks | changed_weeks))
        for wk in weeks_to_sync:
            _sync_weekly_totals(week=wk)
    except Exception:
        pass

    return JsonResponse(
        {
            "ok": True,
            "week": new_week,
            "summary_week": sesion.semana_pago_real or "",
        }
    )


@require_POST
def presign_receipt(request, pk: int):
    """
    Presigned POST directo a Wasabi (path-style):
    - Sin Content-Type en condiciones (evita mismatches).
    - success_action_status=201.
    - Fuerza URL path-style: https://s3.<region>.wasabisys.com/<bucket>
    """
    wp = get_object_or_404(WeeklyPayment, pk=pk)

    filename = request.POST.get("filename") or "receipt"
    _base, ext = os.path.splitext(filename)
    ext = (ext or ".pdf").lower()

    key = f"operaciones/pagos/{wp.week}/{wp.technician_id}/receipt_{uuid4().hex}{ext}"

    s3 = _s3_client()
    fields = {
        "acl": "private",
        "success_action_status": "201",
    }
    conditions = [
        {"acl": "private"},
        {"success_action_status": "201"},
        ["content-length-range", 0, 25 * 1024 * 1024],
        # NOTA: no metemos Content-Type en conditions para evitar CORS/preflight raros
    ]

    post = s3.generate_presigned_post(
        Bucket=settings.AWS_STORAGE_BUCKET_NAME,
        Key=key,
        Fields=fields,
        Conditions=conditions,
        ExpiresIn=600,
    )

    # 👇 Forzar URL path-style (algunos entornos devuelven virtual-hosted)
    endpoint = settings.AWS_S3_ENDPOINT_URL.rstrip("/")
    bucket = settings.AWS_STORAGE_BUCKET_NAME
    post["url"] = f"{endpoint}/{bucket}"

    return JsonResponse({"post": post, "key": key})


@login_required
@rol_requerido('admin', 'pm', 'facturacion')
@transaction.atomic
def confirm_receipt(request, pk: int):
    """
    Confirma la subida directa: guarda key en FileField y marca 'paid'.
    No re-sube el archivo; solo enlaza el objeto S3 ya subido.

    IMPORTANTE:
    - Congela snapshot del desglose pagado para que el historial no dependa
      de billings/snapshots vivos.
    """
    wp = get_object_or_404(WeeklyPayment, pk=pk)
    key = request.POST.get("key")
    if not key:
        return HttpResponseBadRequest("Missing key")

    if wp.status not in ("approved_user", "pending_payment"):
        messages.error(request, "This item is not approved by the worker yet.")
        return redirect("operaciones:admin_weekly_payments")

    paid_snapshot = _build_paid_breakdown_snapshot_for_weekly_payment(wp)

    wp.receipt.name = key
    y, w, _ = timezone.localdate().isocalendar()
    wp.paid_week = f"{y}-W{int(w):02d}"
    wp.status = "paid"

    update_fields = ["receipt", "paid_week", "status", "updated_at"]

    if hasattr(wp, "paid_breakdown_snapshot"):
        wp.paid_breakdown_snapshot = paid_snapshot.get("lines", [])
        update_fields.append("paid_breakdown_snapshot")

    if hasattr(wp, "paid_amount_snapshot"):
        wp.paid_amount_snapshot = _deserialize_decimal_from_json(
            paid_snapshot.get("total")
        )
        update_fields.append("paid_amount_snapshot")

    if hasattr(wp, "paid_snapshot_at"):
        wp.paid_snapshot_at = timezone.now()
        update_fields.append("paid_snapshot_at")

    wp.save(update_fields=update_fields)

    _mark_billings_paid_for_weekly_payment(wp)

    messages.success(request, "Payment marked as PAID.")
    return redirect("operaciones:admin_weekly_payments")


@login_required
@rol_requerido("admin", "pm", "facturacion")
@transaction.atomic
def admin_mark_paid(request, pk: int):
    """
    Alternativa si no quieres presigned: sube via Django, guarda y marca 'paid'.

    IMPORTANTE:
    - Congela snapshot del desglose pagado para que el historial no dependa
      de billings/snapshots vivos.
    """
    wp = get_object_or_404(WeeklyPayment, pk=pk)

    if wp.status not in ("approved_user", "pending_payment"):
        messages.error(request, "This item is not approved by the worker yet.")
        return redirect("operaciones:admin_weekly_payments")

    form = PaymentMarkPaidForm(request.POST, request.FILES, instance=wp)
    if not form.is_valid():
        messages.error(request, "Receipt is required.")
        return redirect("operaciones:admin_weekly_payments")

    form.save()

    paid_snapshot = _build_paid_breakdown_snapshot_for_weekly_payment(wp)

    y, w, _ = timezone.localdate().isocalendar()
    wp.paid_week = f"{y}-W{int(w):02d}"
    wp.status = "paid"

    update_fields = ["paid_week", "status", "updated_at"]

    if hasattr(wp, "paid_breakdown_snapshot"):
        wp.paid_breakdown_snapshot = paid_snapshot.get("lines", [])
        update_fields.append("paid_breakdown_snapshot")

    if hasattr(wp, "paid_amount_snapshot"):
        wp.paid_amount_snapshot = _deserialize_decimal_from_json(
            paid_snapshot.get("total")
        )
        update_fields.append("paid_amount_snapshot")

    if hasattr(wp, "paid_snapshot_at"):
        wp.paid_snapshot_at = timezone.now()
        update_fields.append("paid_snapshot_at")

    wp.save(update_fields=update_fields)

    _mark_billings_paid_for_weekly_payment(wp)

    messages.success(request, "Payment marked as PAID.")
    return redirect("operaciones:admin_weekly_payments")


# ================================= USUARIO ================================= #


@login_required
@never_cache
def user_weekly_payments(request):
    """
    Vista del trabajador:
    - Sincroniza sus registros.
    - Crea/actualiza WeeklyPayment por semanas con producción nueva, legacy o ajustes.
    - NO borra weeks legacy por error.
    - Adjunta details y display_amount por semana.

    Optimización:
    - El legacy se calcula solo para el usuario actual.
    - Evita _legacy_session_totals_without_snapshots(None) para todos los técnicos.
    """
    from decimal import Decimal

    from django.db.models import Exists, F, OuterRef, Q, Sum

    from operaciones.models import (AdjustmentEntry, BillingPayWeekSnapshot,
                                    WeeklyPayment)

    user_id = request.user.id

    snapshot_field_names = {f.name for f in BillingPayWeekSnapshot._meta.get_fields()}
    has_is_adjustment = "is_adjustment" in snapshot_field_names
    has_adjustment_of = "adjustment_of" in snapshot_field_names

    # ---------------------------------------------------------
    # 1) Sync / create de weeks reales
    # ---------------------------------------------------------
    legacy_totals_user = _legacy_session_totals_without_snapshots(
        week=None,
        tech_ids=[user_id],
    )

    legacy_weeks = {
        wk
        for (tid, wk), amt in legacy_totals_user.items()
        if int(tid) == int(user_id) and (amt or Decimal("0.00")) != 0
    }

    if request.GET.get("skip_sync") != "1":
        sync_weekly_totals_no_create(technician_id=user_id)

        weeks_prod_qs = BillingPayWeekSnapshot.objects.filter(
            tecnico_id=user_id,
            semana_resultado__gt="",
            sesion__isnull=False,
        ).filter(
            Q(sesion__estado__in=ESTADOS_OK)
            | Q(sesion__is_direct_discount=True)
            | Q(subtotal__lt=0)
        )

        if has_is_adjustment:
            weeks_prod_qs = weeks_prod_qs.filter(is_adjustment=False)
        elif has_adjustment_of:
            weeks_prod_qs = weeks_prod_qs.filter(adjustment_of__isnull=True)

        weeks_prod = set(
            weeks_prod_qs.values_list("semana_resultado", flat=True).distinct()
        )

        weeks_adj_raw = set(
            AdjustmentEntry.objects.filter(technician_id=user_id)
            .exclude(amount=0)
            .values_list("week", flat=True)
            .distinct()
        )

        weeks_to_sync = set(
            filter(None, weeks_prod.union(legacy_weeks).union(weeks_adj_raw))
        )

        for wk in weeks_to_sync:
            try:
                _sync_weekly_totals(week=str(wk), create_missing=True)
            except Exception:
                pass

    # ---------------------------------------------------------
    # 2) Semana actual
    # ---------------------------------------------------------
    y, w, _ = timezone.localdate().isocalendar()
    current_week = f"{y}-W{int(w):02d}"

    # ---------------------------------------------------------
    # 3) Detectores de existencia real
    # ---------------------------------------------------------
    prod_exists = BillingPayWeekSnapshot.objects.filter(
        tecnico_id=user_id,
        semana_resultado=OuterRef("week"),
        sesion__isnull=False,
    ).filter(
        Q(sesion__estado__in=ESTADOS_OK)
        | Q(sesion__is_direct_discount=True)
        | Q(subtotal__lt=0)
    )

    if has_is_adjustment:
        prod_exists = prod_exists.filter(is_adjustment=False)
    elif has_adjustment_of:
        prod_exists = prod_exists.filter(adjustment_of__isnull=True)

    prod_exists = (
        prod_exists.values("tecnico_id")
        .annotate(total=Sum("subtotal"))
        .exclude(total=0)
    )

    adj_exists = (
        AdjustmentEntry.objects.filter(technician_id=user_id, week=OuterRef("week"))
        .exclude(amount=0)
        .values("id")[:1]
    )

    # ---------------------------------------------------------
    # 4) Limpieza: NO borrar weeks legacy
    # ---------------------------------------------------------
    delete_qs = (
        WeeklyPayment.objects.filter(technician_id=user_id)
        .annotate(has_prod=Exists(prod_exists), has_adj=Exists(adj_exists))
        .filter(has_prod=False, has_adj=False)
        .exclude(status="paid")
    )

    if legacy_weeks:
        delete_qs = delete_qs.exclude(week__in=legacy_weeks)

    delete_qs.delete()

    # ---------------------------------------------------------
    # 5) Lista principal
    # ---------------------------------------------------------
    mine_qs = (
        WeeklyPayment.objects.filter(technician_id=user_id)
        .annotate(has_prod=Exists(prod_exists), has_adj=Exists(adj_exists))
        .filter(
            Q(has_prod=True)
            | Q(has_adj=True)
            | Q(amount__gt=0)
            | Q(week__in=legacy_weeks)
        )
        .select_related("technician")
        .order_by("-week")
    )

    mine = list(mine_qs)

    weeks = {wp.week for wp in mine}
    details_map = {}
    totals_map = {}

    # ---------------------------------------------------------
    # 6) Detalle producción nueva
    # ---------------------------------------------------------
    if weeks:
        snap_qs = BillingPayWeekSnapshot.objects.filter(
            tecnico_id=user_id,
            semana_resultado__in=weeks,
            sesion__isnull=False,
        )

        if has_is_adjustment:
            snap_qs = snap_qs.filter(is_adjustment=False)
        elif has_adjustment_of:
            snap_qs = snap_qs.filter(adjustment_of__isnull=True)

        snap_qs = snap_qs.filter(
            Q(sesion__estado__in=ESTADOS_OK)
            | Q(sesion__is_direct_discount=True)
            | Q(subtotal__lt=0)
        )

        det_prod = (
            snap_qs.values(
                "semana_resultado",
                project_id=F("sesion__proyecto_id"),
                is_discount=F("sesion__is_direct_discount"),
            )
            .annotate(subtotal=Sum("subtotal"))
            .order_by("semana_resultado", "project_id")
        )

        for r in det_prod:
            week = r["semana_resultado"]
            sub = r["subtotal"] or Decimal("0.00")
            project_label = "Direct discount" if r["is_discount"] else r["project_id"]

            details_map.setdefault(week, []).append(
                {
                    "project_id": project_label,
                    "subtotal": sub,
                }
            )

            totals_map[week] = totals_map.get(week, Decimal("0.00")) + sub

        # ---------------------------------------------------------
        # 7) Detalle legacy solo del usuario
        # ---------------------------------------------------------
        legacy_details = _legacy_weekly_payment_details(
            tech_ids=[user_id],
            weeks=weeks,
            allowed_project_keys=None,
            can_view_all_projects=True,
        )

        for (tid, week), rows in legacy_details.items():
            if int(tid) != int(user_id):
                continue

            for row in rows:
                subtotal = row.get("subtotal") or Decimal("0.00")

                details_map.setdefault(week, []).append(
                    {
                        "project_id": row.get("project_label") or "—",
                        "subtotal": subtotal,
                    }
                )

                totals_map[week] = totals_map.get(week, Decimal("0.00")) + subtotal

    # ---------------------------------------------------------
    # 8) Ajustes
    # ---------------------------------------------------------
    LABEL = {
        "bonus": "Bonus",
        "fixed_salary": "Fixed salary",
        "advance": "Advance",
    }

    if weeks:
        det_adj = list(
            AdjustmentEntry.objects.filter(technician_id=user_id, week__in=weeks)
            .exclude(amount=0)
            .values("week", "adjustment_type", "amount", "project_id")
        )
    else:
        det_adj = []

    for a in det_adj:
        week = a["week"]

        try:
            amt = Decimal(a["amount"] or 0)
        except Exception:
            amt = Decimal("0.00")

        amt = abs(amt)

        details_map.setdefault(week, []).append(
            {
                "project_id": a.get("project_id") or "-",
                "label": LABEL.get(a["adjustment_type"], a["adjustment_type"]),
                "subtotal": amt,
            }
        )

        totals_map[week] = totals_map.get(week, Decimal("0.00")) + amt

    # ---------------------------------------------------------
    # 9) Aplicar display_amount final
    # ---------------------------------------------------------
    filtered_mine = []

    for wp in mine:
        wp.details = details_map.get(wp.week, [])
        wp.display_amount = totals_map.get(wp.week, Decimal("0.00"))

        if wp.display_amount > 0:
            filtered_mine.append(wp)

    mine = filtered_mine

    return render(
        request,
        "operaciones/pagos_user_list.html",
        {
            "current_week": current_week,
            "mine": mine,
            "approve_form": PaymentApproveForm(),
            "reject_form": PaymentRejectForm(),
        },
    )


@login_required
@transaction.atomic
def user_approve_payment(request, pk: int):
    wp = get_object_or_404(WeeklyPayment, pk=pk, technician=request.user)

    if wp.status != "pending_user":
        messages.info(
            request, "You can only approve when status is 'Pending my approval'.")
        return redirect("operaciones:user_weekly_payments")

    wp.reject_reason = ""
    wp.status = "pending_payment"  # aprobado -> queda esperando pago
    wp.save(update_fields=["status", "reject_reason", "updated_at"])

    messages.success(request, "Amount approved. Waiting for payment.")
    # 👇 evitamos que la siguiente vista vuelva a sincronizar y deshaga el cambio
    return redirect(f"{reverse('operaciones:user_weekly_payments')}?skip_sync=1")


@login_required
@transaction.atomic
def user_reject_payment(request, pk: int):
    wp = get_object_or_404(WeeklyPayment, pk=pk, technician=request.user)

    if wp.status != "pending_user":
        messages.info(
            request, "You can only reject when status is 'Pending my approval'.")
        return redirect("operaciones:user_weekly_payments")

    form = PaymentRejectForm(request.POST, instance=wp)
    if not form.is_valid():
        messages.error(request, "Please provide a reason.")
        return redirect("operaciones:user_weekly_payments")

    wp = form.save(commit=False)
    wp.status = "rejected_user"
    wp.save(update_fields=["status", "reject_reason", "updated_at"])

    messages.success(request, "Amount rejected. Your reason is visible now.")
    # ⬇️ Evita que la vista de lista vuelva a sincronizar y revierta el estado
    return redirect(f"{reverse('operaciones:user_weekly_payments')}?skip_sync=1")


def admin_reset_payment_status(request, pk: int):
    """
    Vuelve un registro RECHAZADO a 'pending_user' para que el técnico lo vuelva a aprobar.
    """
    wp = get_object_or_404(WeeklyPayment, pk=pk)

    if wp.status != "rejected_user":
        messages.info(
            request, "Only items rejected by the worker can be reset.")
        return redirect("operaciones:admin_weekly_payments")

    wp.status = "pending_user"
    wp.reject_reason = ""  # si prefieres conservar el motivo, comenta esta línea
    wp.save(update_fields=["status", "reject_reason", "updated_at"])

    messages.success(request, "Status reset to 'Pending worker approval'.")
    return redirect("operaciones:admin_weekly_payments")


# -------------------------------------------------------------------
# LOGGIN
# -------------------------------------------------------------------

logger = logging.getLogger("merge_xlsx")

# ===== Namespaces =====
_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
_NS_REL_PKG = "http://schemas.openxmlformats.org/package/2006/relationships"        # *.rels
_NS_REL_DOC = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"  # r:id en XML
_NS_CT = "http://schemas.openxmlformats.org/package/2006/content-types"
_NS_APP = "http://schemas.openxmlformats.org/officeDocument/2006/extended-properties"
_NS_VT = "http://schemas.openxmlformats.org/officeDocument/2006/docPropsVTypes"
_XML_NS = "http://www.w3.org/XML/1998/namespace"

# ¡IMPORTANTE! 'r' = namespace DOC (para r:id en XPaths)
NS = {"m": _NS_MAIN, "r": _NS_REL_DOC}
CT = {"ct": _NS_CT}

# ===== XML helpers =====


def _read_xml(zf: zipfile.ZipFile, path: str) -> ET.Element:
    return ET.fromstring(zf.read(path))


def _write_xml(zf: zipfile.ZipFile, path: str, root: ET.Element):
    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    zf.writestr(path, data)


def _fetch_to_temp(django_filefield) -> str:
    tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    with django_filefield.open("rb") as fsrc, open(tmp.name, "wb") as fdst:
        shutil.copyfileobj(fsrc, fdst, length=1024*1024)
    return tmp.name


def _max_index_from_paths(paths, prefix, suffix):
    mx = 0
    for p in paths:
        if p.startswith(prefix) and p.endswith(suffix):
            m = re.search(r"(\d+)", p[len(prefix):-len(suffix)])
            if m:
                mx = max(mx, int(m.group(1)))
    return mx


def _next_rid(wb_rels_root: ET.Element) -> str:
    ids = [
        int(e.attrib["Id"][3:])
        for e in wb_rels_root.findall(f".//{{{_NS_REL_PKG}}}Relationship")
        if e.attrib.get("Id", "").startswith("rId") and e.attrib["Id"][3:].isdigit()
    ]
    return f"rId{(max(ids)+1) if ids else 1}"

# ===== Content_Types helpers =====


def _read_ct(zf: zipfile.ZipFile): return ET.fromstring(
    zf.read("[Content_Types].xml"))


def _write_ct(zf: zipfile.ZipFile, root: ET.Element):
    data = ET.tostring(root, encoding="utf-8", xml_declaration=True)
    zf.writestr("[Content_Types].xml", data)


def _ensure_default(ct_root: ET.Element, ext: str, ctype: str):
    if ct_root.find(f".//ct:Default[@Extension='{ext}']", CT) is None:
        el = ET.SubElement(ct_root, "{%s}Default" % _NS_CT)
        el.set("Extension", ext)
        el.set("ContentType", ctype)


def _ensure_override(ct_root: ET.Element, partname: str, ctype: str):
    if ct_root.find(f".//ct:Override[@PartName='/{partname}']", CT) is None:
        el = ET.SubElement(ct_root, "{%s}Override" % _NS_CT)
        el.set("PartName", "/" + partname)
        el.set("ContentType", ctype)

# ===== .rels target normalizer =====


def _rels_target_to_zip_path(target: str) -> str:
    p = (target or "").replace("\\", "/")
    while p.startswith("../"):
        p = p[3:]
    if p.startswith("/"):
        p = p[1:]
    if not p.startswith("xl/"):
        p = "xl/" + p
    return p

# ===== sharedStrings → inlineStr =====


def _get_shared_strings(src_zip: zipfile.ZipFile):
    p = "xl/sharedStrings.xml"
    if p not in src_zip.namelist():
        return None, []
    root = _read_xml(src_zip, p)
    out = []
    for si in root.findall("{%s}si" % _NS_MAIN):
        out.append("".join((t.text or "")
                   for t in si.findall(".//{%s}t" % _NS_MAIN)))
    return root, out


def _inline_shared_strings(sheet_root: ET.Element, sst_list):
    if not sst_list:
        return
    for c in sheet_root.findall(".//m:c", NS):
        if c.get("t") != "s":
            continue
        v_el = c.find("m:v", NS)
        if v_el is None or v_el.text is None:
            c.set("t", "inlineStr")
            for ch in list(c):
                if ch.tag.endswith("v") or ch.tag.endswith("is"):
                    c.remove(ch)
            ET.SubElement(c, "{%s}is" % _NS_MAIN)
            continue
        try:
            idx = int(v_el.text)
        except:
            idx = -1
        text = sst_list[idx] if 0 <= idx < len(sst_list) else ""
        for ch in list(c):
            if ch.tag.endswith("v") or ch.tag.endswith("is"):
                c.remove(ch)
        c.set("t", "inlineStr")
        is_el = ET.SubElement(c, "{%s}is" % _NS_MAIN)
        t_el = ET.SubElement(is_el, "{%s}t" % _NS_MAIN)
        if text and (text.startswith(" ") or text.endswith(" ")):
            t_el.set("{%s}space" % _XML_NS, "preserve")
        t_el.text = text
        # NO tocamos el atributo 's' (estilo)

# ===== SOLO quitamos <extLst> (mantenemos estilos) =====


def _strip_extlst_only(sheet_root: ET.Element):
    for ch in list(sheet_root):
        if ch.tag.endswith("extLst"):
            sheet_root.remove(ch)

# ===== nombres de hoja seguros =====


def _safe_sheet_name(name: str) -> str:
    n = re.sub(r'[\\/:*?\[\]]', ' ', (name or '').strip())
    if n.startswith("'"):
        n = n[1:]
    if n.endswith("'"):
        n = n[:-1]
    n = re.sub(r'\s+', ' ', n)[:31]
    return n or 'Sheet'

# ===== app.xml =====


def _read_app_xml(zf: zipfile.ZipFile):
    p = "docProps/app.xml"
    return _read_xml(zf, p) if p in zf.namelist() else None


def _rewrite_app_xml(app_root: ET.Element, sheet_titles: list[str]):
    if app_root is None:
        return None
    app_root.set("xmlns", _NS_APP)
    app_root.set("xmlns:vt", _NS_VT)
    for tag in ("HeadingPairs", "TitlesOfParts"):
        n = app_root.find(f"{{{_NS_APP}}}{tag}")
        if n is not None:
            app_root.remove(n)
    hp = ET.SubElement(app_root, f"{{{_NS_APP}}}HeadingPairs")
    v = ET.SubElement(hp, f"{{{_NS_VT}}}vector", size="2", baseType="variant")
    var1 = ET.SubElement(v, f"{{{_NS_VT}}}variant")
    ET.SubElement(var1, f"{{{_NS_VT}}}lpstr").text = "Worksheets"
    var2 = ET.SubElement(v, f"{{{_NS_VT}}}variant")
    ET.SubElement(var2, f"{{{_NS_VT}}}i4").text = str(len(sheet_titles))
    top = ET.SubElement(app_root, f"{{{_NS_APP}}}TitlesOfParts")
    v2 = ET.SubElement(top, f"{{{_NS_VT}}}vector", size=str(
        len(sheet_titles)), baseType="lpstr")
    for nm in sheet_titles:
        ET.SubElement(v2, f"{{{_NS_VT}}}lpstr").text = nm
    return app_root

# ===== limpieza si falta .rels =====


def _strip_relationship_bound_elements(sheet_root: ET.Element):
    for xp in [".//m:drawing", ".//m:legacyDrawing", ".//m:legacyDrawingHF",
               ".//m:hyperlinks", ".//m:tableParts", ".//m:controls"]:
        for el in sheet_root.findall(xp, NS):
            try:
                sheet_root.remove(el)
            except:
                pass
    for el in sheet_root.findall(".//*[@r:id]", NS):
        try:
            el.attrib.pop("{%s}id" % _NS_REL_DOC, None)
        except:
            pass

# ===================================================================
# MERGE (con FIX de r:id y conservando estilos)
# ===================================================================


def merge_xlsx_files_preserving_images(src_paths, out_path, sheet_names=None):
    if not src_paths:
        raise ValueError("No hay archivos de entrada")
    if len(src_paths) == 1:
        shutil.copyfile(src_paths[0], out_path)
        return

    with zipfile.ZipFile(src_paths[0], "r") as base:
        existing = set(base.namelist())
        wb_xml_path = "xl/workbook.xml"
        wb_rels_path = "xl/_rels/workbook.xml.rels"
        wb_root = _read_xml(base, wb_xml_path)
        wb_rels_root = _read_xml(base, wb_rels_path)
        ct_root = _read_ct(base)
        app_root = _read_app_xml(base)

        # mínimos
        _ensure_default(
            ct_root, "rels", "application/vnd.openxmlformats-package.relationships+xml")
        _ensure_default(ct_root, "xml", "application/xml")
        _ensure_default(ct_root, "png", "image/png")
        _ensure_default(ct_root, "jpg", "image/jpeg")
        _ensure_default(ct_root, "jpeg", "image/jpeg")
        _ensure_default(
            ct_root, "vml", "application/vnd.openxmlformats-officedocument.vmlDrawing")
        _ensure_default(
            ct_root, "bin", "application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings")

        with zipfile.ZipFile(out_path, "w", compression=zipfile.ZIP_DEFLATED) as dst:
            skip = {wb_xml_path, wb_rels_path,
                    "[Content_Types].xml", "xl/calcChain.xml", "docProps/app.xml"}
            for name in existing:
                if name not in skip:
                    dst.writestr(name, base.read(name))

            used_names = set()
            sheets_node = wb_root.find("m:sheets", NS) or ET.SubElement(
                wb_root, "{%s}sheets" % _NS_MAIN)
            base_sheets = sheets_node.findall("m:sheet", NS)
            for s in base_sheets:
                nm = (s.get("name") or "").strip()
                if nm:
                    used_names.add(nm.lower())

            def unique_name(raw: str) -> str:
                base_nm = _safe_sheet_name(raw)
                name = base_nm or "Sheet"
                k = 2
                while name.lower() in used_names:
                    cut = 31 - (len(str(k))+3)
                    cut = 1 if cut < 1 else cut
                    name = f"{(base_nm or 'Sheet')[:cut]} ({k})"
                    k += 1
                used_names.add(name.lower())
                return name

            if sheet_names and len(sheet_names) >= 1 and base_sheets:
                base_sheets[0].set("name", unique_name(sheet_names[0]))

            def _idx(pref, suf): return _max_index_from_paths(
                existing, pref, suf) + 1
            next_sheet_idx = _idx("xl/worksheets/sheet", ".xml")
            next_drawing_idx = _idx("xl/drawings/drawing", ".xml")
            next_vml_idx = _idx("xl/drawings/vmlDrawing", ".vml")
            next_comments_idx = _idx("xl/comments", ".xml")
            next_table_idx = _idx("xl/tables/table", ".xml")
            next_ps_idx = _idx("xl/printerSettings/printerSettings", ".bin")

            img_nums = [int(m.group(1)) for p in existing for m in [
                re.search(r"xl/media/image(\d+)\.(?:png|jpe?g)$", p)] if m]
            chart_nums = [int(m.group(1)) for p in existing for m in [
                re.search(r"xl/charts/chart(\d+)\.xml$", p)] if m]
            next_image_idx = (max(img_nums)+1) if img_nums else 1
            next_chart_idx = (max(chart_nums)+1) if chart_nums else 1
            next_cstyle_idx = 1
            next_ccolor_idx = 1

            sheet_ids = []
            for s in wb_root.findall("m:sheets/m:sheet", NS):
                try:
                    sheet_ids.append(int(s.get("sheetId", "0")))
                except:
                    pass
            next_sheet_id = (max(sheet_ids)+1) if sheet_ids else 1

            def _add_sheet_from_src(src_path, i):
                nonlocal next_sheet_idx, next_sheet_id, next_drawing_idx, next_vml_idx
                nonlocal next_comments_idx, next_table_idx, next_ps_idx
                nonlocal next_image_idx, next_chart_idx, next_cstyle_idx, next_ccolor_idx

                with zipfile.ZipFile(src_path, "r") as src:
                    src_names = set(src.namelist())
                    if "xl/workbook.xml" not in src_names or "xl/_rels/workbook.xml.rels" not in src_names:
                        return

                    swb = _read_xml(src, "xl/workbook.xml")
                    swb_rels = _read_xml(
                        src, "xl/_rels/workbook.xml.rels")  # PACKAGE

                    src_sheets = swb.findall("m:sheets/m:sheet", NS)
                    if not src_sheets:
                        return

                    first_sheet = src_sheets[0]
                    src_rid = first_sheet.get("{%s}id" % _NS_REL_DOC) or first_sheet.get(
                        "r:id") or first_sheet.get("id")
                    src_sheet_path = None
                    if src_rid:
                        rel = swb_rels.find(
                            f".//{{{_NS_REL_PKG}}}Relationship[@Id='{src_rid}']")
                        if rel is not None:
                            t = rel.get("Target") or ""
                            p1 = "xl/"+t
                            src_sheet_path = p1 if p1 in src_names else (_rels_target_to_zip_path(
                                t) if _rels_target_to_zip_path(t) in src_names else None)
                    if not src_sheet_path:
                        cands = []
                        for p in src_names:
                            if p.startswith("xl/worksheets/sheet") and p.endswith(".xml"):
                                m = re.search(r"sheet(\d+)\.xml$", p)
                                idx = int(m.group(1)) if m else 9999
                                cands.append((idx, p))
                        if not cands:
                            return
                        cands.sort()
                        src_sheet_path = cands[0][1]

                    sheet_root = _read_xml(src, src_sheet_path)
                    _, sst_list = _get_shared_strings(src)
                    _inline_shared_strings(sheet_root, sst_list)
                    _strip_extlst_only(sheet_root)   # <<< mantenemos estilos

                    dst_sheet_name = f"worksheets/sheet{next_sheet_idx}.xml"
                    dst_sheet_path = "xl/" + dst_sheet_name
                    dst_sheet_rels_path = f"xl/worksheets/_rels/sheet{next_sheet_idx}.xml.rels"

                    rels_map = {}
                    drel_root = None

                    used_rids = {
                        el.get("{%s}id" % _NS_REL_DOC)
                        for el in sheet_root.findall(".//*[@r:id]", NS)
                        if el.get("{%s}id" % _NS_REL_DOC)
                    }

                    src_sheet_rels_path = f"xl/worksheets/_rels/{os.path.basename(src_sheet_path)}.rels"
                    if src_sheet_rels_path in src_names:
                        srel_root = _read_xml(src, src_sheet_rels_path)

                        drel_root = ET.Element("Relationships")
                        drel_root.set("xmlns", _NS_REL_PKG)

                        def _add_rel(_type, _target, _mode=None):
                            rel = ET.SubElement(drel_root, "Relationship")
                            rel.set("Id", f"rId{len(list(drel_root)) + 1}")
                            rel.set("Type", _type)
                            rel.set("Target", _target)
                            if _mode:
                                rel.set("TargetMode", _mode)
                            return rel.get("Id")

                        for r in srel_root.findall("{%s}Relationship" % _NS_REL_PKG):
                            rId = r.get("Id")
                            rTyp = (r.get("Type") or "")
                            rTgt = (r.get("Target") or "")
                            rMode = r.get("TargetMode")

                            if used_rids and rId not in used_rids:
                                continue

                            if rTyp.endswith("/drawing"):
                                src_draw_path = _rels_target_to_zip_path(rTgt)
                                if src_draw_path in src_names:
                                    new_draw_name = f"drawing{next_drawing_idx}.xml"
                                    dst_draw_path = "xl/drawings/" + new_draw_name
                                    draw_xml = _read_xml(src, src_draw_path)

                                    src_draw_rels = f"xl/drawings/_rels/{os.path.basename(src_draw_path)}.rels"
                                    if src_draw_rels in src_names:
                                        drels_xml = _read_xml(
                                            src, src_draw_rels)
                                        for ir in drels_xml.findall("{%s}Relationship" % _NS_REL_PKG):
                                            ityp = (ir.get("Type") or "")
                                            itgt = (ir.get("Target") or "")
                                            if ityp.endswith("/image"):
                                                src_img = _rels_target_to_zip_path(
                                                    itgt)
                                                if src_img in src_names:
                                                    ext = os.path.splitext(
                                                        src_img)[1].lower()
                                                    new_img = f"image{next_image_idx}{ext}"
                                                    dst.writestr(
                                                        "xl/media/"+new_img, src.read(src_img))
                                                    if ext == ".png":
                                                        _ensure_default(
                                                            ct_root, "png", "image/png")
                                                    elif ext in (".jpg", ".jpeg"):
                                                        _ensure_default(
                                                            ct_root, ext[1:], "image/jpeg")
                                                    ir.set(
                                                        "Target", "../media/"+new_img)
                                                    next_image_idx += 1
                                            elif ityp.endswith("/chart"):
                                                src_chart = _rels_target_to_zip_path(
                                                    itgt)
                                                if src_chart in src_names:
                                                    new_chart = f"chart{next_chart_idx}.xml"
                                                    dst.writestr(
                                                        "xl/charts/"+new_chart, src.read(src_chart))
                                                    _ensure_override(ct_root, "xl/charts/"+new_chart,
                                                                     "application/vnd.openxmlformats-officedocument.drawingml.chart+xml")
                                                    src_chart_rels = f"xl/charts/_rels/{os.path.basename(src_chart)}.rels"
                                                    if src_chart_rels in src_names:
                                                        crels_xml = _read_xml(
                                                            src, src_chart_rels)
                                                        for cr in crels_xml.findall("{%s}Relationship" % _NS_REL_PKG):
                                                            ctyp = (
                                                                cr.get("Type") or "")
                                                            ctgt = (
                                                                cr.get("Target") or "")
                                                            if ctyp.endswith("/image"):
                                                                cimg = _rels_target_to_zip_path(
                                                                    ctgt)
                                                                if cimg in src_names:
                                                                    ext = os.path.splitext(
                                                                        cimg)[1].lower()
                                                                    new_img = f"image{next_image_idx}{ext}"
                                                                    dst.writestr(
                                                                        "xl/media/"+new_img, src.read(cimg))
                                                                    if ext == ".png":
                                                                        _ensure_default(
                                                                            ct_root, "png", "image/png")
                                                                    elif ext in (".jpg", ".jpeg"):
                                                                        _ensure_default(
                                                                            ct_root, ext[1:], "image/jpeg")
                                                                    cr.set(
                                                                        "Target", "../media/"+new_img)
                                                                    next_image_idx += 1
                                                            elif ctyp.endswith("/chartStyle"):
                                                                s = _rels_target_to_zip_path(
                                                                    ctgt)
                                                                if s in src_names:
                                                                    new = f"style{next_cstyle_idx}.xml"
                                                                    dst.writestr(
                                                                        "xl/charts/"+new, src.read(s))
                                                                    _ensure_override(
                                                                        ct_root, "xl/charts/"+new, "application/vnd.ms-office.chartstyle+xml")
                                                                    cr.set(
                                                                        "Target", new)
                                                                    next_cstyle_idx += 1
                                                            elif ctyp.endswith("/chartColorStyle"):
                                                                s = _rels_target_to_zip_path(
                                                                    ctgt)
                                                                if s in src_names:
                                                                    new = f"colors{next_ccolor_idx}.xml"
                                                                    dst.writestr(
                                                                        "xl/charts/"+new, src.read(s))
                                                                    _ensure_override(
                                                                        ct_root, "xl/charts/"+new, "application/vnd.ms-office.chartcolorstyle+xml")
                                                                    cr.set(
                                                                        "Target", new)
                                                                    next_ccolor_idx += 1
                                                        _write_xml(
                                                            dst, f"xl/charts/_rels/{new_chart}.rels", crels_xml)
                                                    next_chart_idx += 1
                                        _write_xml(
                                            dst, f"xl/drawings/_rels/{new_draw_name}.rels", drels_xml)

                                    _write_xml(dst, dst_draw_path, draw_xml)
                                    new_rel_id = _add_rel(
                                        rTyp, "../drawings/" + new_draw_name)
                                    _ensure_override(ct_root, "xl/drawings/" + new_draw_name,
                                                     "application/vnd.openxmlformats-officedocument.drawing+xml")
                                    # remapeo r:id del drawing
                                    rels_map[rId] = new_rel_id
                                    next_drawing_idx += 1
                                continue

                            if rTyp.endswith("/hyperlink"):
                                new_rel_id = _add_rel(rTyp, rTgt, _mode=rMode)
                                rels_map[rId] = new_rel_id
                                continue

                            if rTyp.endswith("/table"):
                                s = _rels_target_to_zip_path(rTgt)
                                if s in src_names:
                                    new = f"table{next_table_idx}.xml"
                                    dst.writestr("xl/tables/"+new, src.read(s))
                                    _ensure_override(
                                        ct_root, "xl/tables/"+new, "application/vnd.openxmlformats-officedocument.spreadsheetml.table+xml")
                                    new_rel_id = _add_rel(
                                        rTyp, "../tables/"+new)
                                    rels_map[rId] = new_rel_id
                                    next_table_idx += 1
                                continue

                            if rTyp.endswith("/comments"):
                                s = _rels_target_to_zip_path(rTgt)
                                if s in src_names:
                                    new = f"comments{next_comments_idx}.xml"
                                    dst.writestr("xl/"+new, src.read(s))
                                    _ensure_override(
                                        ct_root, "xl/"+new, "application/vnd.openxmlformats-officedocument.spreadsheetml.comments+xml")
                                    new_rel_id = _add_rel(rTyp, "../"+new)
                                    rels_map[rId] = new_rel_id
                                    next_comments_idx += 1
                                continue

                            if rTyp.endswith("/vmlDrawing"):
                                s = _rels_target_to_zip_path(rTgt)
                                if s in src_names:
                                    new = f"vmlDrawing{next_vml_idx}.vml"
                                    dst.writestr("xl/drawings/" +
                                                 new, src.read(s))
                                    new_rel_id = _add_rel(
                                        rTyp, "../drawings/"+new)
                                    rels_map[rId] = new_rel_id
                                    next_vml_idx += 1
                                continue

                            if rTyp.endswith("/printerSettings"):
                                s = _rels_target_to_zip_path(rTgt)
                                if s in src_names:
                                    new = f"printerSettings{next_ps_idx}.bin"
                                    dst.writestr(
                                        "xl/printerSettings/"+new, src.read(s))
                                    _ensure_override(ct_root, "xl/printerSettings/"+new,
                                                     "application/vnd.openxmlformats-officedocument.spreadsheetml.printerSettings")
                                    new_rel_id = _add_rel(
                                        rTyp, "../printerSettings/"+new)
                                    rels_map[rId] = new_rel_id
                                    next_ps_idx += 1
                                continue

                            new_rel_id = _add_rel(rTyp, rTgt, _mode=rMode)
                            rels_map[rId] = new_rel_id

                        if list(drel_root):
                            _write_xml(dst, dst_sheet_rels_path, drel_root)
                    else:
                        if used_rids:
                            _strip_relationship_bound_elements(sheet_root)

                    for el in sheet_root.findall(".//*[@r:id]", NS):
                        rid = el.get("{%s}id" % _NS_REL_DOC)
                        if rid in rels_map:
                            el.set("{%s}id" % _NS_REL_DOC, rels_map[rid])

                    _write_xml(dst, dst_sheet_path, sheet_root)
                    _ensure_override(ct_root, dst_sheet_name,
                                     "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml")

                    new_rid = _next_rid(wb_rels_root)
                    sheet_el = ET.SubElement(
                        sheets_node, "{%s}sheet" % _NS_MAIN)
                    raw_name = (sheet_names[i] if (
                        sheet_names and i < len(sheet_names)) else f"Report {i+1}")
                    nm = _safe_sheet_name(raw_name)
                    name = nm or "Sheet"
                    k = 2
                    while name.lower() in used_names:
                        cut = 31 - (len(str(k))+3)
                        cut = 1 if cut < 1 else cut
                        name = f"{(nm or 'Sheet')[:cut]} ({k})"
                        k += 1
                    used_names.add(name.lower())
                    sheet_el.set("name", name)
                    sheet_el.set("sheetId", str(next_sheet_id))
                    sheet_el.set("{%s}id" % _NS_REL_DOC, new_rid)  # DOC NS

                    wb_rel = ET.SubElement(
                        wb_rels_root, "{%s}Relationship" % _NS_REL_PKG)
                    wb_rel.set("Id", new_rid)
                    wb_rel.set(
                        "Type", "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet")
                    wb_rel.set("Target", dst_sheet_name)

                    next_sheet_idx += 1
                    next_sheet_id += 1

            for i, src_path in enumerate(src_paths[1:], start=1):
                _add_sheet_from_src(src_path, i)

            titles = [(s.get("name") or "Sheet")
                      for s in wb_root.findall("m:sheets/m:sheet", NS)]
            if app_root is not None:
                app_fixed = _rewrite_app_xml(app_root, titles)
                if app_fixed is not None:
                    _write_xml(dst, "docProps/app.xml", app_fixed)

            _write_xml(dst, wb_xml_path, wb_root)
            _write_xml(dst, wb_rels_path, wb_rels_root)
            _write_ct(dst, ct_root)

            return titles

# ===== VISTA: merge y descarga =====


@login_required
@rol_requerido("supervisor", "pm", "admin", "facturacion")
def billing_merge_excel(request):
    run_id = datetime.now().strftime("%Y%m%d-%H%M%S")
    ids = []
    if request.method == "POST" and request.body:
        try:
            payload = json.loads(request.body.decode("utf-8"))
            ids = [int(x)
                   for x in (payload.get("ids") or []) if str(x).isdigit()]
        except Exception:
            logger.exception("MERGE RUN %s | invalid JSON", run_id)
    if not ids:
        qs = (request.GET.get("ids") or "").strip()
        if qs:
            ids = [int(x) for x in qs.split(",") if x.isdigit()]
    if not ids:
        return HttpResponseBadRequest("Debes indicar ids, ej: ?ids=53,59 o POST JSON {'ids':[...]}")
    logger.info("MERGE RUN %s | ids=%s", run_id, ids)

    sesiones = {s.id: s for s in SesionBilling.objects.filter(id__in=ids)}
    ordered = [sesiones[i] for i in ids if i in sesiones]

    src_paths, sheet_names, skipped = [], [], []
    for s in ordered:
        rf = getattr(s, "reporte_fotografico", None)
        if not rf:
            skipped.append(str(s.id))
            continue
        try:
            rf.open("rb")
            rf.close()
        except Exception:
            skipped.append(str(s.id))
            continue
        try:
            tmp = _fetch_to_temp(rf)
            src_paths.append(tmp)
            sheet_names.append(
                (f"{(s.proyecto_id or '').strip()}-{s.id}")[:31] or f"proj-{s.id}")
        except Exception:
            skipped.append(str(s.id))

    if not src_paths:
        return JsonResponse({"error": "Ninguno de los proyectos seleccionados tiene un reporte XLSX disponible."}, status=400)

    out_tmp = NamedTemporaryFile(delete=False, suffix=".xlsx")
    out_tmp.close()
    final_titles = merge_xlsx_files_preserving_images(
        src_paths, out_tmp.name, sheet_names=sheet_names)

    f = open(out_tmp.name, "rb")
    resp = FileResponse(
        f,
        as_attachment=True,
        filename="reportes_fotograficos_merged.xlsx",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Length"] = os.path.getsize(out_tmp.name)
    from django.utils.http import http_date
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = http_date(0)

    resp["X-Debug-Run"] = run_id
    try:
        resp["X-Merged-Count"] = str(len(final_titles))
        resp["X-Merged-Sheets"] = ",".join(final_titles)
    except Exception:
        pass
    if skipped:
        resp["X-Skipped-Ids"] = ",".join(skipped)
    return resp

import json

from django.contrib.auth.decorators import login_required
from django.core.files.storage import default_storage
from django.db import transaction
from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.views.decorators.http import require_POST

from operaciones.models import EvidenciaFotoBilling, SesionBilling


@login_required
@require_POST
@rol_requerido("supervisor", "admin", "pm")
def bulk_delete_evidencias(request, sesion_id):
    """
    Borra evidencias en MASIVO (1 request) para la vista de supervisor review.

    Payload JSON esperado:
      { "ids": [1,2,3,...] }

    Reglas:
      - Solo si la sesión NO está aprobada (aprobado_supervisor / aprobado_pm).
      - Solo evidencias que pertenezcan a tecnico_sesion__sesion_id = sesion_id.
      - Borra DB en batch y luego intenta borrar archivos (best-effort).
      - Optimizado: intenta borrar en lote (S3/Wasabi) para hacerlo MUCHO más rápido.
    """
    s = get_object_or_404(SesionBilling, pk=sesion_id)

    # 🔒 Si ya está aprobado, bloqueamos
    if s.estado in {"aprobado_supervisor", "aprobado_pm"}:
        return JsonResponse(
            {"ok": False, "error": "Locked. Approved sessions cannot delete photos."},
            status=403,
        )

    # Solo JSON
    try:
        raw = request.body.decode("utf-8") if request.body else ""
        data = json.loads(raw or "{}")
    except Exception:
        return JsonResponse({"ok": False, "error": "Invalid JSON body."}, status=400)

    ids = data.get("ids") or []
    if not isinstance(ids, list) or not ids:
        return JsonResponse({"ok": False, "error": "No ids provided."}, status=400)

    # Normalizar ids a int
    norm_ids = []
    for x in ids:
        try:
            norm_ids.append(int(x))
        except Exception:
            continue

    if not norm_ids:
        return JsonResponse({"ok": False, "error": "No valid ids provided."}, status=400)

    # Query SOLO de esta sesión
    qs = EvidenciaFotoBilling.objects.filter(
        id__in=norm_ids,
        tecnico_sesion__sesion_id=s.id,
    )

    # ids reales encontrados
    found = list(qs.values_list("id", flat=True))
    missing = sorted(list(set(norm_ids) - set(found)))

    # Guardar keys del storage ANTES de borrar rows
    # (ojo: values_list("imagen") devuelve el name del FileField)
    files = [f for f in qs.values_list("imagen", flat=True) if f]

    deleted_count = 0
    file_errors = []

    with transaction.atomic():
        deleted_count, _ = qs.delete()

    # ===========================
    # Borrado rápido en lote (S3/Wasabi)
    # ===========================
    def _try_bulk_delete_s3(keys):
        """
        Intenta usar boto3 delete_objects si el storage lo expone.
        Retorna True si lo intentó (aunque S3 reporte errores parciales),
        False si no pudo usar S3 client.
        """
        if not keys:
            return True

        # Muchos storages S3 exponen .bucket / .bucket_name / .connection / .client
        storage = default_storage

        # Intentar obtener client boto3
        client = None
        bucket = None

        # Caso común django-storages S3Boto3Storage:
        # storage.bucket_name y storage.connection.meta.client
        try:
            bucket = getattr(storage, "bucket_name", None) or getattr(storage, "bucket", None)
        except Exception:
            bucket = None

        try:
            conn = getattr(storage, "connection", None)
            if conn is not None and getattr(conn, "meta", None) is not None:
                client = conn.meta.client
        except Exception:
            client = None

        # Alternativas
        if client is None:
            try:
                client = getattr(storage, "client", None)
            except Exception:
                client = None

        # bucket a string
        if bucket is not None and not isinstance(bucket, str):
            # algunos storages guardan bucket como objeto; intentamos sacar name
            bucket = getattr(bucket, "name", None) or str(bucket)

        if not client or not bucket:
            return False

        # delete_objects soporta hasta 1000 por request
        try:
            CHUNK = 1000
            for i in range(0, len(keys), CHUNK):
                chunk = keys[i:i + CHUNK]
                resp = client.delete_objects(
                    Bucket=bucket,
                    Delete={"Objects": [{"Key": k} for k in chunk], "Quiet": True},
                )
                # Si hay errores, S3 los devuelve en "Errors"
                for err in (resp.get("Errors") or []):
                    # Key puede venir, Code, Message
                    k = err.get("Key")
                    code = err.get("Code")
                    msg = err.get("Message")
                    file_errors.append(f"{k or ''} {code or ''} {msg or ''}".strip())
            return True
        except Exception as e:
            file_errors.append(str(e))
            return True  # lo intentó, pero falló

    tried_s3_bulk = _try_bulk_delete_s3(files)

    # Fallback: borrar uno por uno si no se pudo usar S3 bulk
    if not tried_s3_bulk:
        for name in files:
            try:
                default_storage.delete(name)  # idempotente: si no existe, idealmente no rompe
            except Exception as e:
                file_errors.append(str(e))

    return JsonResponse(
        {
            "ok": True,
            "deleted": deleted_count,
            "missing": missing,
            "file_errors": file_errors[:5],  # no spamear
        }
    )


@login_required
@rol_requerido("admin", "pm", "supervisor")
@csrf_protect
def billing_update_creado_en(request, sesion_id: int):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    s = get_object_or_404(SesionBilling, pk=sesion_id)

    raw = (request.POST.get("creado_en") or "").strip()
    if not raw:
        return JsonResponse({"ok": False, "error": "Date is required."}, status=400)

    try:
        from datetime import datetime

        d = datetime.strptime(raw, "%Y-%m-%d").date()
    except Exception:
        return JsonResponse(
            {"ok": False, "error": "Invalid format. Use: YYYY-MM-DD"}, status=400
        )

    tz = timezone.get_current_timezone()

    current_local = timezone.localtime(s.creado_en, tz) if s.creado_en else None
    if current_local:
        dt = datetime.combine(d, current_local.time().replace(second=0, microsecond=0))
    else:
        dt = datetime.combine(d, datetime.min.time())

    dt = timezone.make_aware(dt, tz)

    s.creado_en = dt
    s.save(update_fields=["creado_en"])

    display = timezone.localtime(s.creado_en, tz).strftime("%Y-%m-%d")
    return JsonResponse(
        {
            "ok": True,
            "display": display,
            "search_value": display,
        }
    )

@login_required
@rol_requerido('admin', 'pm', 'supervisor')
@csrf_protect
def billing_update_project_id(request, sesion_id: int):
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    s = get_object_or_404(SesionBilling, pk=sesion_id)

    proyecto_id = (request.POST.get("proyecto_id") or "").strip()
    if not proyecto_id:
        return JsonResponse({"ok": False, "error": "Project ID is required."}, status=400)

    s.proyecto_id = proyecto_id
    s.save(update_fields=["proyecto_id"])

    return JsonResponse({
        "ok": True,
        "proyecto_id": s.proyecto_id,
    })


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion")
@csrf_protect
def billing_update_project_week(request, sesion_id: int):
    """
    Cambio REAL de week desde listados (Billing List / Invoice List).

    Regla:
    - Cambia semana_pago_proyectada
    - Si la sesión tiene snapshots productivos:
        * mueve semana_base a la nueva week
        * recalcula semana_resultado = semana_base + payment_weeks
        * recalcula semana_pago_real resumen
    - Si la sesión es legacy (sin snapshots productivos):
        * mueve semana_pago_real a la nueva week
    - Re-sincroniza WeeklyPayment en semanas viejas y nuevas

    Importante:
    - No permite cambios si hay weekly payments PAID, salvo admin real.
    - Responde siempre JSON.
    """
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    s = get_object_or_404(SesionBilling, pk=sesion_id)

    if not access_user_can(request.user, "billing.edit_real_week"):
        return JsonResponse(
            {
                "ok": False,
                "error": "FORBIDDEN",
                "message": "You do not have permission to edit the real pay week.",
            },
            status=403,
        )

    semana = (request.POST.get("semana_pago_proyectada") or "").strip().upper()
    if not semana:
        return JsonResponse({"ok": False, "error": "Week is required."}, status=400)

    if not WEEK_RE.match(semana):
        return JsonResponse(
            {"ok": False, "error": "Invalid format. Use YYYY-W##"},
            status=400,
        )

    year, week = _parse_iso_week(semana)
    if not year or not week:
        return JsonResponse(
            {"ok": False, "error": "Invalid ISO week."},
            status=400,
        )

    is_admin = _is_admin(request.user)

    # 🔒 bloqueo real por pagos PAID
    if _session_is_paid_locked(s) and not is_admin:
        return JsonResponse(
            {
                "ok": False,
                "error": "LOCKED_PAID",
                "message": "This session has PAID weekly payments. Only admins can change the pay week.",
            },
            status=403,
        )

    snapshot_model = BillingPayWeekSnapshot
    snapshot_fields = {f.name for f in snapshot_model._meta.get_fields()}

    has_is_adjustment = "is_adjustment" in snapshot_fields
    has_adjustment_of = "adjustment_of" in snapshot_fields

    def _productivo_qs():
        qs = snapshot_model.objects.filter(sesion=s)
        if has_is_adjustment:
            qs = qs.filter(is_adjustment=False)
        elif has_adjustment_of:
            qs = qs.filter(adjustment_of__isnull=True)
        return qs

    # semanas anteriores afectadas
    old_weeks = set(
        snapshot_model.objects.filter(sesion=s)
        .exclude(semana_resultado__isnull=True)
        .exclude(semana_resultado__exact="")
        .values_list("semana_resultado", flat=True)
    )

    old_legacy_week = (getattr(s, "semana_pago_real", "") or "").strip().upper()
    if old_legacy_week:
        old_weeks.add(old_legacy_week)

    snaps = list(_productivo_qs())
    changed_weeks = set()
    summary_week = ""

    with transaction.atomic():
        # siempre guardar la proyectada/base
        s.semana_pago_proyectada = semana

        # ---------------------------------------
        # CASO 1: sesión con snapshots productivos
        # ---------------------------------------
        if snaps:
            for snap in snaps:
                payment_weeks = int(getattr(snap, "payment_weeks", 0) or 0)
                recalculated_week = _add_weeks_to_iso_week(semana, payment_weeks)

                update_fields = []

                if (snap.semana_base or "") != semana:
                    snap.semana_base = semana
                    update_fields.append("semana_base")

                if (snap.semana_resultado or "") != recalculated_week:
                    snap.semana_resultado = recalculated_week
                    update_fields.append("semana_resultado")

                if update_fields:
                    snap.save(update_fields=update_fields)

                if recalculated_week:
                    changed_weeks.add(recalculated_week)

            # semana resumen = mayor semana_resultado de snapshots productivos
            all_result_weeks = list(
                _productivo_qs()
                .exclude(semana_resultado__isnull=True)
                .exclude(semana_resultado__exact="")
                .values_list("semana_resultado", flat=True)
            )

            parsed = []
            for w in all_result_weeks:
                y, wk = _parse_iso_week(w)
                if y and wk:
                    parsed.append((y, wk, w))

            if parsed:
                parsed.sort(key=lambda x: (x[0], x[1]))
                summary_week = parsed[-1][2]

            s.semana_pago_real = summary_week

        # ---------------------------------------
        # CASO 2: sesión legacy sin snapshots
        # ---------------------------------------
        else:
            s.semana_pago_real = semana
            summary_week = semana
            changed_weeks.add(semana)

        s.save(update_fields=["semana_pago_proyectada", "semana_pago_real"])

    # resincronizar weeks afectadas
    try:
        weeks_to_sync = set(filter(None, old_weeks | changed_weeks))
        for wk in weeks_to_sync:
            _sync_weekly_totals(week=wk)
    except Exception:
        pass

    return JsonResponse(
        {
            "ok": True,
            "semana_pago_proyectada": s.semana_pago_proyectada or "",
            "semana_pago_real": s.semana_pago_real or "",
            "summary_week": summary_week or "",
        }
    )


@login_required
@rol_requerido("usuario")
def upload_evidencias_dispatch(request, pk):
    a = get_object_or_404(
        SesionBillingTecnico.objects.select_related("sesion", "tecnico"),
        pk=pk,
    )

    if a.tecnico_id != request.user.id:
        return HttpResponseForbidden(
            "You do not have permission to access this assignment."
        )

    if getattr(a.sesion, "is_cable_installation", False):
        return redirect(
            "cable_installation:technician_requirements", assignment_id=a.pk
        )

    return redirect("operaciones:upload_evidencias", pk=a.pk)
