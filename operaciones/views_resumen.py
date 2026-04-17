from decimal import Decimal
from urllib.parse import urlencode

from django.contrib.auth.decorators import login_required
from django.db.models import Max, Q
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from core.permissions import filter_queryset_by_access
from facturacion.models import Proyecto
from operaciones.models import SesionBilling
from usuarios.decoradores import rol_requerido

try:
    from usuarios.models import ProyectoAsignacion
except Exception:
    ProyectoAsignacion = None


# ==========================================================
# Configuración de estados del resumen operativo
# ==========================================================
# Nota:
# - El flujo real de Billing termina en supervisor.
# - Si existen registros históricos con aprobado_pm/rechazado_pm,
#   los absorbemos visualmente en Approved/Rejected para no perderlos.
ESTADOS_RESUMEN = [
    {
        "key": "asignado",
        "label": "Assigned",
        "statuses": ["asignado"],
        "badge": "slate",
    },
    {
        "key": "en_proceso",
        "label": "In progress",
        "statuses": ["en_proceso"],
        "badge": "amber",
    },
    {
        "key": "en_revision_supervisor",
        "label": "Submitted — supervisor review",
        "statuses": ["en_revision_supervisor"],
        "badge": "blue",
    },
    {
        "key": "rechazado_supervisor",
        "label": "Rejected by supervisor",
        "statuses": ["rechazado_supervisor", "rechazado_pm"],
        "badge": "red",
    },
    {
        "key": "aprobado_supervisor",
        "label": "Approved by supervisor",
        "statuses": ["aprobado_supervisor", "aprobado_pm"],
        "badge": "green",
    },
]

FINANCE_LABELS = dict(SesionBilling._meta.get_field("finance_status").choices)


# ==========================================================
# Helpers de acceso / proyecto
# ==========================================================
def _safe_decimal(val) -> Decimal:
    try:
        return Decimal(str(val or "0.00"))
    except Exception:
        return Decimal("0.00")


def _week_sort_key(week_str: str):
    """
    Convierte YYYY-W## a una llave de orden.
    Si no es válida, la deja al final.
    """
    s = (week_str or "").strip().upper()
    if not s:
        return (-1, -1)

    try:
        year, week = s.split("-W")
        return (int(year), int(week))
    except Exception:
        return (-1, -1)


def _weeks_disponibles_ordenadas(base_qs):
    """
    Semanas proyectadas disponibles, ordenadas por última creación.
    """
    qs = (
        base_qs.exclude(semana_pago_proyectada__isnull=True)
        .exclude(semana_pago_proyectada="")
        .values("semana_pago_proyectada")
        .annotate(last_created=Max("creado_en"))
        .order_by("-last_created")
    )
    return [r["semana_pago_proyectada"] for r in qs if r["semana_pago_proyectada"]]


def _project_label_for_session(sesion, proyectos_qs):
    """
    Resuelve el nombre visible del proyecto con la misma lógica defensiva
    que ya usas en otras vistas.
    """
    proyecto_sel = None
    raw = (getattr(sesion, "proyecto", "") or "").strip()

    if raw:
        try:
            pid = int(raw)
        except (TypeError, ValueError):
            proyecto_sel = proyectos_qs.filter(
                Q(nombre__iexact=raw) | Q(codigo__iexact=raw)
            ).first()
        else:
            proyecto_sel = proyectos_qs.filter(pk=pid).first()

    if not proyecto_sel and getattr(sesion, "proyecto_id", None):
        code = str(sesion.proyecto_id).strip()
        if code:
            proyecto_sel = proyectos_qs.filter(
                Q(codigo__iexact=code) | Q(nombre__icontains=code)
            ).first()

    if proyecto_sel:
        return getattr(proyecto_sel, "nombre", str(proyecto_sel))

    return (
        getattr(sesion, "proyecto", None) or getattr(sesion, "proyecto_id", "") or ""
    ).strip()


def _base_resumen_qs(user):
    """
    Base visible del resumen operativo.
    Replica la lógica principal de acceso por proyecto/ventana temporal
    usada en Billing.
    """
    qs = SesionBilling.objects.all().order_by("-creado_en")

    can_view_legacy_history = user.is_superuser or getattr(
        user, "es_usuario_historial", False
    )

    if not can_view_legacy_history:
        proyectos_user = filter_queryset_by_access(
            Proyecto.objects.all(),
            user,
            "id",
        )

        if proyectos_user.exists():
            allowed_keys = set()

            for p in proyectos_user:
                nombre = (getattr(p, "nombre", "") or "").strip()
                if nombre:
                    allowed_keys.add(nombre)

                codigo = getattr(p, "codigo", None)
                if codigo:
                    allowed_keys.add(str(codigo).strip())

                allowed_keys.add(str(p.id).strip())

            qs = qs.filter(proyecto__in=allowed_keys)

            if ProyectoAsignacion is not None:
                try:
                    asignaciones = list(
                        ProyectoAsignacion.objects.filter(
                            usuario=user, proyecto__in=proyectos_user
                        ).select_related("proyecto")
                    )
                except Exception:
                    asignaciones = []

                if asignaciones:
                    window_q = Q()

                    for a in asignaciones:
                        p = getattr(a, "proyecto", None)
                        if not p:
                            continue

                        keys = set()
                        nombre = (getattr(p, "nombre", "") or "").strip()
                        if nombre:
                            keys.add(nombre)

                        codigo = getattr(p, "codigo", None)
                        if codigo:
                            keys.add(str(codigo).strip())

                        keys.add(str(p.id).strip())

                        if getattr(a, "include_history", False) or not getattr(
                            a, "start_at", None
                        ):
                            window_q |= Q(proyecto__in=keys)
                        else:
                            window_q |= Q(proyecto__in=keys) & Q(
                                creado_en__gte=a.start_at
                            )

                    qs = qs.filter(window_q) if window_q else qs.none()
        else:
            qs = qs.none()

    return qs


def _build_resumen_data(request):
    """
    Devuelve estructura completa del resumen operativo.
    Filtro principal:
      - week = semana_pago_proyectada
    """
    base_all = _base_resumen_qs(request.user)
    semanas_disponibles = _weeks_disponibles_ordenadas(base_all)

    if "week" in request.GET:
        week_sel = (request.GET.get("week") or "").strip()
    else:
        week_sel = semanas_disponibles[0] if semanas_disponibles else ""

    base = base_all
    if week_sel:
        base = base.filter(semana_pago_proyectada=week_sel)

    # Query de proyectos visibles para resolver nombres amigables
    proyectos_qs = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        "id",
    )

    sesiones = list(base.order_by("-creado_en", "-id"))

    # Estructura de grupos
    secciones = []
    total_global_count = 0
    total_global_company = Decimal("0.00")
    total_global_technical = Decimal("0.00")

    approved_count = 0
    review_count = 0

    for conf in ESTADOS_RESUMEN:
        statuses = set(conf["statuses"])
        items = []

        count = 0
        total_company = Decimal("0.00")
        total_technical = Decimal("0.00")

        for s in sesiones:
            if (s.estado or "") not in statuses:
                continue

            company = _safe_decimal(getattr(s, "subtotal_empresa", 0))
            technical = _safe_decimal(getattr(s, "subtotal_tecnico", 0))

            item = {
                "id": s.id,
                "created_at": timezone.localtime(s.creado_en) if s.creado_en else None,
                "project_id": s.proyecto_id or "",
                "project_label": _project_label_for_session(s, proyectos_qs),
                "client": s.cliente or "",
                "city": s.ciudad or "",
                "office": s.oficina or "",
                "projected_week": s.semana_pago_proyectada or "",
                "real_week": s.semana_pago_real or "",
                "subtotal_company": company,
                "subtotal_technical": technical,
                "finance_status": s.finance_status or "none",
                "finance_label": FINANCE_LABELS.get(
                    s.finance_status, s.finance_status or "—"
                ),
                "is_direct_discount": bool(getattr(s, "is_direct_discount", False)),
                "is_cable_installation": bool(
                    getattr(s, "is_cable_installation", False)
                ),
                "is_split_child": bool(getattr(s, "is_split_child", False)),
                "proyecto_especial": bool(getattr(s, "proyecto_especial", False)),
            }
            items.append(item)

            count += 1
            total_company += company
            total_technical += technical

        total_global_count += count
        total_global_company += total_company
        total_global_technical += total_technical

        if conf["key"] == "aprobado_supervisor":
            approved_count = count
        if conf["key"] == "en_revision_supervisor":
            review_count = count

        secciones.append(
            {
                "key": conf["key"],
                "label": conf["label"],
                "badge": conf["badge"],
                "count": count,
                "total_company": total_company,
                "total_technical": total_technical,
                "items": items,
            }
        )

    context = {
        "weeks_disponibles": semanas_disponibles,
        "week_sel": week_sel,
        "secciones": secciones,
        "total_global_count": total_global_count,
        "total_global_company": total_global_company,
        "total_global_technical": total_global_technical,
        "approved_count": approved_count,
        "review_count": review_count,
        "current_query": (
            urlencode({"week": week_sel}) if week_sel or ("week" in request.GET) else ""
        ),
    }
    return context


# ==========================================================
# Views
# ==========================================================
@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion")
def resumen_operativo(request):
    ctx = _build_resumen_data(request)
    return render(request, "operaciones/resumen_operativo.html", ctx)


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion")
def export_resumen_operativo_xlsx(request):
    ctx = _build_resumen_data(request)

    wb = Workbook()
    ws = wb.active
    ws.title = "Operational Summary"

    # Estilos
    thin = Side(style="thin", color="D1D5DB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_fill = PatternFill("solid", fgColor="E8EEF7")
    header_fill = PatternFill("solid", fgColor="1F2937")
    subheader_fill = PatternFill("solid", fgColor="F3F4F6")

    title_font = Font(bold=True, size=14)
    white_bold = Font(bold=True, color="FFFFFF")
    bold_font = Font(bold=True)

    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")
    right = Alignment(horizontal="right", vertical="center")

    row = 1

    # Título
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=10)
    c = ws.cell(row=row, column=1, value="Operational Summary — Billing by stage")
    c.font = title_font
    c.alignment = center
    c.fill = title_fill
    c.border = border
    row += 1

    ws.cell(row=row, column=1, value="Projected week").font = bold_font
    ws.cell(row=row, column=2, value=ctx["week_sel"] or "All")
    ws.cell(row=row, column=4, value="Total billings").font = bold_font
    ws.cell(row=row, column=5, value=ctx["total_global_count"])
    ws.cell(row=row, column=7, value="Total company").font = bold_font
    ws.cell(row=row, column=8, value=float(ctx["total_global_company"]))
    ws.cell(row=row, column=9, value="Total technical").font = bold_font
    ws.cell(row=row, column=10, value=float(ctx["total_global_technical"]))

    for col in range(1, 11):
        ws.cell(row=row, column=col).border = border
    row += 2

    # Secciones
    for sec in ctx["secciones"]:
        if not sec["count"]:
            continue

        # Cabecera sección
        ws.cell(row=row, column=1, value=sec["label"]).font = bold_font
        ws.cell(row=row, column=2, value="Billings").font = bold_font
        ws.cell(row=row, column=3, value=sec["count"])
        ws.cell(row=row, column=4, value="Company").font = bold_font
        ws.cell(row=row, column=5, value=float(sec["total_company"]))
        ws.cell(row=row, column=6, value="Technical").font = bold_font
        ws.cell(row=row, column=7, value=float(sec["total_technical"]))

        for col in range(1, 8):
            ws.cell(row=row, column=col).fill = subheader_fill
            ws.cell(row=row, column=col).border = border
        row += 1

        headers = [
            "Billing",
            "Created",
            "Project ID",
            "Project",
            "Client",
            "City",
            "Office",
            "Projected week",
            "Real pay week",
            "Company",
            "Technical",
            "Finance",
            "Flags",
        ]
        for idx, h in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=idx, value=h)
            cell.font = white_bold
            cell.fill = header_fill
            cell.alignment = center
            cell.border = border
        row += 1

        for item in sec["items"]:
            flags = []
            if item["is_direct_discount"]:
                flags.append("Direct discount")
            if item["is_cable_installation"]:
                flags.append("Cable")
            if item["is_split_child"]:
                flags.append("Split")
            if item["proyecto_especial"]:
                flags.append("Special")

            values = [
                item["id"],
                (
                    item["created_at"].strftime("%Y-%m-%d %H:%M")
                    if item["created_at"]
                    else ""
                ),
                item["project_id"],
                item["project_label"],
                item["client"],
                item["city"],
                item["office"],
                item["projected_week"],
                item["real_week"],
                float(item["subtotal_company"]),
                float(item["subtotal_technical"]),
                item["finance_label"],
                ", ".join(flags),
            ]

            for idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row, column=idx, value=val)
                cell.border = border
                if idx in (10, 11):
                    cell.alignment = right
                else:
                    cell.alignment = left
            row += 1

        ws.cell(row=row, column=9, value=f"Subtotal — {sec['label']}").font = bold_font
        ws.cell(row=row, column=10, value=float(sec["total_company"])).font = bold_font
        ws.cell(row=row, column=11, value=float(sec["total_technical"])).font = (
            bold_font
        )

        for col in range(9, 12):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).fill = subheader_fill
        row += 2

    # Ajuste ancho columnas
    for col in range(1, 14):
        max_len = 0
        for r in ws.iter_rows(min_col=col, max_col=col, min_row=1, max_row=ws.max_row):
            val = r[0].value
            if val is None:
                continue
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(
            max(max_len + 2, 12), 40
        )

    # Formato numérico
    for col in (8, 10, 11):
        for r in range(2, ws.max_row + 1):
            ws.cell(r, col).number_format = "$#,##0.00"

    filename_week = (ctx["week_sel"] or "All").replace(" ", "_")
    filename = f"OperationalSummary_{filename_week}.xlsx"

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response
