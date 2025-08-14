import io
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from django.core.files.base import ContentFile
from django.db.models import Sum


def generar_excel_reporte(tec_sesion):
    """
    Construye un Excel con:
      - Portada: datos de sesión/tecnico/fechas/estado
      - Items: SOLO precios del técnico (ItemBillingTecnico)
      - Fotos: miniaturas con requisito, nota y timestamp
    Devuelve ContentFile listo para guardar en tec_sesion.reporte_fotografico.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    s = tec_sesion.sesion
    u = tec_sesion.tecnico
    ws.append(["Project ID", s.proyecto_id])
    ws.append(["Client", s.cliente])
    ws.append(["City", s.ciudad])
    ws.append(["Project", s.proyecto])
    ws.append(["Office", s.oficina])
    ws.append(["Technician", u.get_full_name() or u.username])
    ws.append(["Status", tec_sesion.get_estado_display()])
    ws.append(["Assigned %", float(tec_sesion.porcentaje)])
    ws.append([])

    # Tabla de items SOLO del técnico
    ws.append(["Job Code", "Description", "Qty",
              "My Rate (eff.)", "Line Total"])
    from operaciones.models import ItemBillingTecnico
    filas = (ItemBillingTecnico.objects
             .filter(item__sesion=s, tecnico=u)
             .select_related("item")
             .order_by("item__id"))

    total = 0.0
    for it_tec in filas:
        it = it_tec.item
        rate = float(it_tec.tarifa_efectiva)
        line = float(it_tec.subtotal)
        ws.append([it.codigo_trabajo, it.descripcion,
                  float(it.cantidad), rate, line])
        total += line
    ws.append(["", "", "", "My Subtotal", total])

    # Hoja de fotos
    photos = wb.create_sheet("Photos")
    photos.append(["Order", "Requirement", "Description",
                  "Taken At", "Note", "Preview"])
    row = 2
    for ev in tec_sesion.evidencias.select_related("requisito"):
        photos.cell(row=row, column=1, value=(
            ev.requisito.orden if ev.requisito else ""))
        photos.cell(row=row, column=2, value=(
            ev.requisito.titulo if ev.requisito else "Extra"))
        photos.cell(row=row, column=3, value=(
            ev.requisito.descripcion if ev.requisito else ""))
        photos.cell(row=row, column=4, value=str(ev.tomada_en))
        photos.cell(row=row, column=5, value=ev.nota or "")
        try:
            buf = io.BytesIO(ev.imagen.read())
            img = XLImage(buf)
            img.width = 240
            img.height = 180
            photos.add_image(img, f"F{row}")
        except Exception:
            pass
        row += 12

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return ContentFile(out.read(), name="photo_report.xlsx")
