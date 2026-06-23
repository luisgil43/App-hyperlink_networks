import csv
import io
import re
from decimal import Decimal, InvalidOperation
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify
from django.views.decorators.http import require_GET, require_POST
from openpyxl import Workbook, load_workbook

from facturacion.models import Proyecto

from .forms_requirement_lists import (RequirementListForm,
                                      RequirementListItemFormSet)
from .models import RequirementList

try:
    from usuarios.decoradores import rol_requerido
except Exception:

    def rol_requerido(*roles):
        def decorator(fn):
            return fn

        return decorator


def _req_list_to_bool(value, default=True):
    if value in (None, ""):
        return default

    text = str(value).strip().lower()

    yes_values = {"1", "true", "t", "yes", "y", "si", "sí", "s"}
    no_values = {"0", "false", "f", "no", "n"}

    if text in yes_values:
        return True

    if text in no_values:
        return False

    return default


def _req_list_to_decimal_string(value, default="0"):
    if value in (None, ""):
        return default

    try:
        number = Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, ValueError, TypeError):
        return default

    if number < 0:
        return default

    text = format(number, "f")
    if "." in text:
        text = text.rstrip("0").rstrip(".")

    return text or default


def _normalize_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _read_requirement_import_rows(uploaded):
    filename = (getattr(uploaded, "name", "") or "").strip()
    ext = (filename.rsplit(".", 1)[-1] or "").lower()

    rows = []

    if ext == "csv":
        raw = uploaded.read().decode("utf-8-sig", errors="ignore")
        if not raw.strip():
            return rows

        reader = csv.DictReader(io.StringIO(raw))
        for row in reader:
            cleaned = {
                _normalize_header(k): (v if v is not None else "")
                for k, v in row.items()
            }
            rows.append(cleaned)

        return rows

    if ext in ("xlsx", "xls"):
        wb = load_workbook(uploaded, data_only=True)
        ws = wb.active

        data = list(ws.iter_rows(values_only=True))
        if not data:
            return rows

        headers = [_normalize_header(x) for x in data[0]]

        for values in data[1:]:
            item = {}

            for idx, header in enumerate(headers):
                if not header:
                    continue

                value = values[idx] if idx < len(values) else ""
                item[header] = "" if value is None else value

            if any(str(v).strip() for v in item.values()):
                rows.append(item)

        return rows

    raise ValueError("Unsupported file type. Use CSV or XLSX.")


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@require_POST
def requirement_list_import_preview(request):
    uploaded = request.FILES.get("file")
    list_type = (
        request.POST.get("list_type") or RequirementList.LIST_TYPE_FIBER
    ).strip()

    if list_type not in (
        RequirementList.LIST_TYPE_FIBER,
        RequirementList.LIST_TYPE_CABLE,
    ):
        list_type = RequirementList.LIST_TYPE_FIBER

    if not uploaded:
        return JsonResponse(
            {
                "ok": False,
                "error": "Please select a CSV or XLSX file.",
            },
            status=400,
        )

    try:
        raw_rows = _read_requirement_import_rows(uploaded)
    except Exception as exc:
        return JsonResponse(
            {
                "ok": False,
                "error": f"Could not read file: {exc}",
            },
            status=400,
        )

    imported = []
    errors = []

    for index, row in enumerate(raw_rows, start=2):
        if list_type == RequirementList.LIST_TYPE_CABLE:
            handhole = row.get("handhole") or row.get("name") or row.get("title") or ""
            handhole = str(handhole or "").strip()

            if not handhole:
                errors.append(f"Row {index}: handhole is required.")
                continue

            planned = _req_list_to_decimal_string(
                row.get("planned_reserve_ft")
                or row.get("planned_slack")
                or row.get("planned_slack_ft")
                or row.get("reserve")
                or "0"
            )

            required = _req_list_to_bool(
                row.get("required") or row.get("mandatory") or row.get("obligatorio"),
                default=True,
            )

            warning = str(row.get("warning") or "").strip()

            try:
                order = int(row.get("order") or len(imported))
            except Exception:
                order = len(imported)

            imported.append(
                {
                    "handhole": handhole,
                    "planned_reserve_ft": planned,
                    "required": required,
                    "warning": warning,
                    "order": order,
                }
            )

        else:
            name = row.get("name") or row.get("title") or row.get("requirement") or ""
            name = str(name or "").strip()

            if not name:
                errors.append(f"Row {index}: name is required.")
                continue

            required = _req_list_to_bool(
                row.get("mandatory") or row.get("required") or row.get("obligatorio"),
                default=True,
            )

            try:
                order = int(row.get("order") or len(imported))
            except Exception:
                order = len(imported)

            imported.append(
                {
                    "name": name,
                    "required": required,
                    "order": order,
                }
            )

    return JsonResponse(
        {
            "ok": True,
            "list_type": list_type,
            "items": imported,
            "errors": errors,
            "count": len(imported),
        }
    )


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
def requirement_list_list(request):
    qs = (
        RequirementList.objects.select_related("project", "created_by")
        .prefetch_related("items")
        .order_by("project__nombre", "list_type", "name")
    )

    q = (request.GET.get("q") or "").strip()
    project_id = (request.GET.get("project") or "").strip()
    status = (request.GET.get("status") or "").strip()
    list_type = (request.GET.get("type") or "").strip()

    if q:
        qs = qs.filter(name__icontains=q)

    if project_id.isdigit():
        qs = qs.filter(project_id=int(project_id))

    if list_type in (
        RequirementList.LIST_TYPE_FIBER,
        RequirementList.LIST_TYPE_CABLE,
    ):
        qs = qs.filter(list_type=list_type)

    if status == "active":
        qs = qs.filter(is_active=True)
    elif status == "paused":
        qs = qs.filter(is_active=False)

    projects = Proyecto.objects.all().order_by("nombre")

    return render(
        request,
        "operaciones/requirement_lists/list.html",
        {
            "items": qs,
            "projects": projects,
            "q": q,
            "project_id": project_id,
            "status": status,
            "list_type": list_type,
        },
    )


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@transaction.atomic
def requirement_list_create(request):
    if request.method == "POST":
        form = RequirementListForm(request.POST)

        selected_type = (
            request.POST.get("list_type") or RequirementList.LIST_TYPE_FIBER
        ).strip()

        formset = RequirementListItemFormSet(
            request.POST,
            list_type=selected_type,
        )

        if form.is_valid():
            selected_type = (
                form.cleaned_data.get("list_type") or RequirementList.LIST_TYPE_FIBER
            )

            formset = RequirementListItemFormSet(
                request.POST,
                list_type=selected_type,
            )

            if formset.is_valid():
                obj = form.save(commit=False)
                obj.created_by = request.user
                obj.save()

                formset.instance = obj
                formset.save()

                messages.success(request, "Requirement list created successfully.")
                return redirect("operaciones:requirement_list_list")
        else:
            formset.is_valid()

    else:
        form = RequirementListForm(
            initial={
                "is_active": True,
                "list_type": RequirementList.LIST_TYPE_FIBER,
            }
        )
        formset = RequirementListItemFormSet(
            list_type=RequirementList.LIST_TYPE_FIBER,
        )

    return render(
        request,
        "operaciones/requirement_lists/form.html",
        {
            "form": form,
            "formset": formset,
            "object": None,
            "mode": "create",
        },
    )


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@transaction.atomic
def requirement_list_edit(request, pk):
    obj = get_object_or_404(RequirementList, pk=pk)

    if request.method == "POST":
        form = RequirementListForm(request.POST, instance=obj)

        formset = RequirementListItemFormSet(
            request.POST,
            instance=obj,
            list_type=obj.list_type,
        )

        if form.is_valid() and formset.is_valid():
            form.save()
            formset.save()

            messages.success(request, "Requirement list updated successfully.")
            return redirect("operaciones:requirement_list_list")

    else:
        form = RequirementListForm(instance=obj)

        formset = RequirementListItemFormSet(
            instance=obj,
            list_type=obj.list_type,
        )

    return render(
        request,
        "operaciones/requirement_lists/form.html",
        {
            "form": form,
            "formset": formset,
            "object": obj,
            "mode": "edit",
        },
    )


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@transaction.atomic
def requirement_list_toggle(request, pk):
    obj = get_object_or_404(RequirementList, pk=pk)

    obj.is_active = not obj.is_active
    obj.save(update_fields=["is_active", "updated_at"])

    if obj.is_active:
        messages.success(request, f"Requirement list '{obj.name}' activated.")
    else:
        messages.success(request, f"Requirement list '{obj.name}' paused.")

    return redirect("operaciones:requirement_list_list")


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@transaction.atomic
def requirement_list_delete(request, pk):
    obj = get_object_or_404(RequirementList, pk=pk)

    if request.method == "POST":
        name = obj.name
        obj.delete()

        messages.success(request, f"Requirement list '{name}' deleted.")
        return redirect("operaciones:requirement_list_list")

    return render(
        request,
        "operaciones/requirement_lists/confirm_delete.html",
        {
            "object": obj,
        },
    )


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@require_GET
def requirement_list_template_download(request, ext):
    list_type = (
        (request.GET.get("type") or RequirementList.LIST_TYPE_FIBER).strip().lower()
    )
    ext = (ext or "").strip().lower()

    if list_type not in (
        RequirementList.LIST_TYPE_FIBER,
        RequirementList.LIST_TYPE_CABLE,
    ):
        list_type = RequirementList.LIST_TYPE_FIBER

    if ext not in ("csv", "xlsx", "xls"):
        return HttpResponse("Unsupported format. Use csv or xlsx.", status=400)

    if list_type == RequirementList.LIST_TYPE_CABLE:
        filename_base = "cable_requirement_list_template"
        headers = ["handhole", "planned_reserve_ft", "required", "warning", "order"]
        sample_rows = [
            ["1000-044", 30, "Yes", "", 0],
            ["1000-044-1", 20, "YES", "", 1],
            ["1000-044-2", 10, "si", "Low clearance", 2],
            ["1000-044-3", 0, "No", "", 3],
        ]
    else:
        filename_base = "fiber_requirement_list_template"
        headers = ["name", "order", "mandatory"]
        sample_rows = [
            ["Front door", 0, 1],
            ["Back door", 1, 1],
            ["Panorama of site", 2, 0],
        ]

    if ext == "csv":
        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(headers)
        writer.writerows(sample_rows)

        response = HttpResponse(
            output.getvalue(), content_type="text/csv; charset=utf-8"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
        return response

    wb = Workbook()
    ws = wb.active
    ws.title = "Requirements"
    ws.append(headers)

    for row in sample_rows:
        ws.append(row)

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
    return response


def _requirement_list_import_bool(value, default=True):
    if value in (None, "", " "):
        return default

    text = str(value).strip().lower()

    yes_values = {"1", "true", "t", "yes", "y", "si", "sí", "s"}
    no_values = {"0", "false", "f", "no", "n"}

    if text in yes_values:
        return True
    if text in no_values:
        return False

    return default


def _requirement_list_import_decimal(value):
    if value in (None, "", " "):
        return Decimal("0")

    try:
        val = Decimal(str(value).replace(",", ".").strip())
    except (InvalidOperation, ValueError, TypeError):
        return None

    if val < 0:
        return None

    return val


def _requirement_list_import_order(value, fallback):
    if value in (None, "", " "):
        return fallback

    try:
        val = int(str(value).strip())
    except Exception:
        return fallback

    if val < 0:
        return fallback

    return val


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@require_POST
def requirement_list_import_parse(request):
    list_type = (
        (request.POST.get("list_type") or RequirementList.LIST_TYPE_FIBER)
        .strip()
        .lower()
    )

    if list_type not in (
        RequirementList.LIST_TYPE_FIBER,
        RequirementList.LIST_TYPE_CABLE,
    ):
        return JsonResponse(
            {"ok": False, "errors": ["Invalid requirement list type."]},
            status=400,
        )

    uploaded = request.FILES.get("file")

    if not uploaded:
        return JsonResponse(
            {"ok": False, "errors": ["Please select a CSV or XLSX file."]},
            status=400,
        )

    filename = (getattr(uploaded, "name", "") or "").strip()
    ext = (filename.rsplit(".", 1)[-1] or "").lower()

    if ext not in ("csv", "xlsx", "xls"):
        return JsonResponse(
            {"ok": False, "errors": ["Unsupported file type. Use .csv or .xlsx."]},
            status=400,
        )

    def normalize_header(value):
        text = str(value or "").strip().lower()
        text = text.replace("\ufeff", "")
        text = text.replace("(ft)", "_ft")
        text = text.replace("ft.", "ft")
        text = re.sub(r"[^a-z0-9]+", "_", text)
        text = re.sub(r"_+", "_", text).strip("_")
        return text

    def pick(data, *keys):
        for key in keys:
            val = data.get(key)
            if val not in (None, ""):
                return val
        return ""

    def decimal_to_text(value):
        if value in (None, "", " "):
            return "0"

        try:
            number = Decimal(str(value).replace(",", ".").strip())
        except (InvalidOperation, ValueError, TypeError):
            return None

        if number < 0:
            return None

        text = format(number, "f")
        if "." in text:
            text = text.rstrip("0").rstrip(".")

        return text or "0"

    def bool_value(value, default=True):
        if value in (None, "", " "):
            return default

        text = str(value).strip().lower()

        yes_values = {"1", "true", "t", "yes", "y", "si", "sí", "s"}
        no_values = {"0", "false", "f", "no", "n"}

        if text in yes_values:
            return True
        if text in no_values:
            return False

        return default

    def order_value(value, fallback):
        if value in (None, "", " "):
            return fallback

        try:
            number = int(str(value).strip())
        except Exception:
            return fallback

        if number < 0:
            return fallback

        return number

    raw_rows = []
    warnings = []
    errors = []

    try:
        if ext == "csv":
            raw = uploaded.read().decode("utf-8-sig", errors="ignore")

            if not raw.strip():
                return JsonResponse(
                    {"ok": False, "errors": ["The file is empty."]},
                    status=400,
                )

            first_line = raw.splitlines()[0].lower() if raw.splitlines() else ""

            if list_type == RequirementList.LIST_TYPE_CABLE:
                has_header = "handhole" in first_line or "planned" in first_line
            else:
                has_header = "name" in first_line or "title" in first_line

            if has_header:
                reader = csv.DictReader(io.StringIO(raw))

                for row_num, row in enumerate(reader, start=2):
                    data = {}

                    for key, value in (row or {}).items():
                        data[normalize_header(key)] = "" if value is None else value

                    raw_rows.append(
                        {
                            "row": row_num,
                            "data": data,
                        }
                    )
            else:
                reader = csv.reader(io.StringIO(raw))

                for row_num, row in enumerate(reader, start=1):
                    if not row:
                        continue

                    if list_type == RequirementList.LIST_TYPE_CABLE:
                        raw_rows.append(
                            {
                                "row": row_num,
                                "data": {
                                    "handhole": row[0] if len(row) > 0 else "",
                                    "planned_reserve_ft": (
                                        row[1] if len(row) > 1 else ""
                                    ),
                                    "required": row[2] if len(row) > 2 else "",
                                    "warning": row[3] if len(row) > 3 else "",
                                    "order": row[4] if len(row) > 4 else "",
                                },
                            }
                        )
                    else:
                        raw_rows.append(
                            {
                                "row": row_num,
                                "data": {
                                    "name": row[0] if len(row) > 0 else "",
                                    "order": row[1] if len(row) > 1 else "",
                                    "mandatory": row[2] if len(row) > 2 else "",
                                },
                            }
                        )

        else:
            wb = load_workbook(uploaded, data_only=True)
            ws = wb.active
            values_rows = list(ws.iter_rows(values_only=True))

            if not values_rows:
                return JsonResponse(
                    {"ok": False, "errors": ["The spreadsheet is empty."]},
                    status=400,
                )

            headers = [normalize_header(x) for x in values_rows[0]]

            if list_type == RequirementList.LIST_TYPE_CABLE:
                has_header = "handhole" in headers or "planned_reserve_ft" in headers
            else:
                has_header = "name" in headers or "title" in headers

            if has_header:
                for row_num, values in enumerate(values_rows[1:], start=2):
                    data = {}

                    for idx, header in enumerate(headers):
                        if not header:
                            continue

                        data[header] = values[idx] if idx < len(values) else ""

                    if any(str(v or "").strip() for v in data.values()):
                        raw_rows.append(
                            {
                                "row": row_num,
                                "data": data,
                            }
                        )
            else:
                for row_num, values in enumerate(values_rows, start=1):
                    if not values:
                        continue

                    if list_type == RequirementList.LIST_TYPE_CABLE:
                        raw_rows.append(
                            {
                                "row": row_num,
                                "data": {
                                    "handhole": values[0] if len(values) > 0 else "",
                                    "planned_reserve_ft": (
                                        values[1] if len(values) > 1 else ""
                                    ),
                                    "required": values[2] if len(values) > 2 else "",
                                    "warning": values[3] if len(values) > 3 else "",
                                    "order": values[4] if len(values) > 4 else "",
                                },
                            }
                        )
                    else:
                        raw_rows.append(
                            {
                                "row": row_num,
                                "data": {
                                    "name": values[0] if len(values) > 0 else "",
                                    "order": values[1] if len(values) > 1 else "",
                                    "mandatory": values[2] if len(values) > 2 else "",
                                },
                            }
                        )

    except Exception as exc:
        return JsonResponse(
            {"ok": False, "errors": [f"Could not parse the file: {exc}"]},
            status=400,
        )

    if not raw_rows:
        return JsonResponse(
            {"ok": False, "errors": ["No valid rows found in the file."]},
            status=400,
        )

    rows = []
    seen = {}

    for item in raw_rows:
        row_num = item.get("row")
        data = item.get("data") or {}

        if list_type == RequirementList.LIST_TYPE_CABLE:
            handhole = pick(
                data,
                "handhole",
                "hh",
                "hh_number",
                "name",
                "title",
            )
            handhole = str(handhole or "").strip()

            if not handhole:
                continue

            key = slugify(handhole)

            if key in seen:
                warnings.append(
                    f"Row {row_num}: duplicated handhole in file — '{handhole}' — duplicates row {seen[key]}."
                )
                continue

            seen[key] = row_num

            reserve_raw = pick(
                data,
                "planned_reserve_ft",
                "planned_slack_ft",
                "planned_slack",
                "planned",
                "reserve",
                "slack",
            )

            reserve_text = decimal_to_text(reserve_raw)

            if reserve_text is None:
                errors.append(f"Row {row_num}: invalid planned_reserve_ft.")
                continue

            required = bool_value(
                pick(data, "required", "mandatory", "obligatorio"),
                default=True,
            )

            order = order_value(
                pick(data, "order", "orden"),
                fallback=len(rows),
            )

            rows.append(
                {
                    "handhole": handhole,
                    "title": handhole,
                    "planned_reserve_ft": reserve_text,
                    "required": required,
                    "warning": str(
                        pick(data, "warning", "note", "notes") or ""
                    ).strip(),
                    "order": order,
                }
            )

        else:
            name = pick(
                data,
                "name",
                "title",
                "requirement",
                "requisito",
            )
            name = str(name or "").strip()

            if not name:
                continue

            key = slugify(name)

            if key in seen:
                warnings.append(
                    f"Row {row_num}: duplicated name in file — '{name}' — duplicates row {seen[key]}."
                )
                continue

            seen[key] = row_num

            required = bool_value(
                pick(data, "mandatory", "required", "obligatorio"),
                default=True,
            )

            order = order_value(
                pick(data, "order", "orden"),
                fallback=len(rows),
            )

            rows.append(
                {
                    "name": name,
                    "title": name,
                    "required": required,
                    "order": order,
                }
            )

    if not rows and not errors:
        errors.append("No valid rows found in the file.")

    if errors:
        return JsonResponse(
            {
                "ok": False,
                "errors": errors,
                "warnings": warnings,
            },
            status=400,
        )

    return JsonResponse(
        {
            "ok": True,
            "rows": rows,
            "warnings": warnings,
            "count": len(rows),
        }
    )
