import json
import unicodedata
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from decimal import ROUND_HALF_UP, Decimal, InvalidOperation
from io import BytesIO

from django import forms
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.core.cache import cache
from django.db import transaction
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.text import slugify
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from access_control.services import user_can as access_user_can
from facturacion.models import Proyecto
from usuarios.models import CustomUser

from .forms_billing_masivo import BillingMasivoUploadForm
from .models import (BillingPayWeekSnapshot, ItemBilling, ItemBillingTecnico,
                     PrecioActividadTecnico, RequirementList,
                     RequisitoFotoBilling, RequisitoFotoBillingPlantilla,
                     SesionBilling, SesionBillingTecnico)

try:
    from usuarios.decoradores import rol_requerido
except Exception:

    def rol_requerido(*roles):
        def decorator(fn):
            return fn

        return decorator


def _bulk_billing_price_permissions(user):
    """
    Permisos visuales para preview de Billing Masivo.

    Usa la misma Access Matrix que Billing List:
    - billing.view_technical_amounts
    - billing.view_company_amounts
    """
    return {
        "can_view_tech_prices": access_user_can(
            user,
            "billing.view_technical_amounts",
        ),
        "can_view_company_prices": access_user_can(
            user,
            "billing.view_company_amounts",
        ),
    }


def _norm_text(value) -> str:
    """
    Normalización controlada para comparar textos:
    - Limpia espacios al inicio/final.
    - Convierte a minúsculas.
    - Quita acentos.
    - Colapsa espacios internos múltiples.
    - NO cambia guiones, puntos ni underscores.
    """
    raw = _clean_cell(value)

    raw = " ".join(raw.split())

    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))

    return raw.lower()


def _same_text(a, b) -> bool:
    return _norm_text(a) == _norm_text(b)


def _find_price_robust(*, tech_id, project, client, city, office, job_code):
    """
    Busca precio de forma robusta para letras/mayúsculas/espacios,
    pero sin transformar símbolos como -, _, .
    """

    qs = PrecioActividadTecnico.objects.filter(
        tecnico_id=tech_id,
        proyecto=project,
    )

    matches = []

    for price in qs:
        if (
            _same_text(price.cliente, client)
            and _same_text(price.ciudad, city)
            and _same_text(price.oficina, office)
            and _same_text(price.codigo_trabajo, job_code)
        ):
            matches.append(price)

    if len(matches) == 1:
        return matches[0], None

    if len(matches) > 1:
        return None, (
            f"More than one price matches Job Code '{job_code}' for this technician/project. "
            "Please clean duplicate prices before importing."
        )

    return None, None


# =============================================================================
# CONFIGURACIÓN DEL TEMPLATE
# =============================================================================

SHEET_BILLINGS = "Billings"
SHEET_TECHNICIANS = "Technicians"
SHEET_ITEMS = "Items"

BILLINGS_HEADERS = [
    "bulk_key",
    "project_id",
    "client",
    "city",
    "project",
    "office",
    "project_address",
    "projected_week",
    "tech_payment_mode",
    "direct_discount",
    "cable_installation",
    "requirement_type",
    "requirement_list",
]

TECHNICIANS_HEADERS = [
    "bulk_key",
    "technician_username",
]

ITEMS_HEADERS = [
    "bulk_key",
    "job_code",
    "quantity",
]

YES_VALUES = {"yes", "y", "true", "1", "si", "sí"}
NO_VALUES = {"no", "n", "false", "0", ""}

VALID_PAYMENT_MODES = {"full", "split"}

VALID_REQUIREMENT_TYPES = {"none", "fiber", "cable", ""}
# =============================================================================
# DATACLASSES DE PREVIEW
# =============================================================================


@dataclass
class CellError:
    sheet: str
    row: int
    field: str
    message: str


@dataclass
class PreviewBilling:

    bulk_key: str

    source_row: int

    project_id: str = ""

    client: str = ""

    city: str = ""

    project: str = ""

    office: str = ""

    project_address: str = ""

    projected_week: str = ""

    tech_payment_mode: str = "full"

    direct_discount: bool = False

    cable_installation: bool = False

    requirement_type: str = "none"

    requirement_list: str = ""

    requirement_list_id: int | None = None

    requirement_list_label: str = ""

    requirement_count: int = 0

    technicians: list = field(default_factory=list)
    items: list = field(default_factory=list)

    subtotal_tecnico: Decimal = Decimal("0.00")
    subtotal_empresa: Decimal = Decimal("0.00")

    errors: list = field(default_factory=list)


@dataclass
class PreviewTechnician:
    source_row: int
    username: str
    user_id: int | None = None
    display_name: str = ""


@dataclass
class PreviewItem:
    source_row: int
    job_code: str
    quantity_raw: str
    quantity: Decimal | None = None

    tipo_trabajo: str = ""
    descripcion: str = ""
    unidad_medida: str = ""
    precio_empresa: Decimal = Decimal("0.00")

    subtotal_tecnico: Decimal = Decimal("0.00")
    subtotal_empresa: Decimal = Decimal("0.00")

    desglose_tecnico: list = field(default_factory=list)
    errors: list = field(default_factory=list)


# =============================================================================
# HELPERS GENERALES
# =============================================================================


def _clean_cell(value) -> str:
    """
    Regla acordada:
    - Se permite limpiar espacios al inicio/final.
    - No se corrigen códigos.
    - No se cambia punto por guion.
    - No se hace fuzzy match.
    """
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value) -> str:
    return _clean_cell(value).lower()


def _to_decimal(value):
    raw = _clean_cell(value)

    if raw == "":
        return None, raw

    try:
        dec = Decimal(raw).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        return dec, raw
    except (InvalidOperation, ValueError):
        return None, raw


def _parse_bool(value):
    raw = _clean_cell(value).lower()

    if raw in YES_VALUES:
        return True, None

    if raw in NO_VALUES:
        return False, None

    return False, "Use YES or NO."


def _cell_error(sheet, row, field, message):
    return {
        "sheet": sheet,
        "row": row,
        "field": field,
        "message": message,
    }


def _read_sheet_rows(wb, sheet_name, expected_headers):
    """
    Retorna:
    - rows: lista de dicts con __rownum
    - errors: errores de estructura de hoja/header
    """
    errors = []

    if sheet_name not in wb.sheetnames:
        errors.append(
            _cell_error(
                sheet_name,
                1,
                "sheet",
                f"Missing sheet '{sheet_name}'.",
            )
        )
        return [], errors

    ws = wb[sheet_name]
    header_values = [_normalize_header(c.value) for c in ws[1]]
    expected_normalized = [_normalize_header(h) for h in expected_headers]

    for idx, expected in enumerate(expected_normalized, start=1):
        actual = header_values[idx - 1] if idx - 1 < len(header_values) else ""

        if actual != expected:
            col = get_column_letter(idx)
            errors.append(
                _cell_error(
                    sheet_name,
                    1,
                    expected_headers[idx - 1],
                    f"Invalid header in cell {col}1. Expected '{expected_headers[idx - 1]}'.",
                )
            )

    if errors:
        return [], errors

    rows = []

    for row_idx in range(2, ws.max_row + 1):
        values = {}
        is_empty = True

        for col_idx, header in enumerate(expected_headers, start=1):
            val = _clean_cell(ws.cell(row=row_idx, column=col_idx).value)
            values[header] = val

            if val != "":
                is_empty = False

        if is_empty:
            continue

        values["__rownum"] = row_idx
        rows.append(values)

    return rows, []


def _format_money(value):
    try:
        return f"{Decimal(value or 0).quantize(Decimal('0.01'))}"
    except Exception:
        return "0.00"


def _display_user(user):
    if not user:
        return ""
    full_name = (user.get_full_name() or "").strip()
    return full_name or user.username


def _iso_week_is_valid(value):
    value = _clean_cell(value).upper()

    if not value:
        return False

    if len(value) != 8:
        return False

    if value[4:6] != "-W":
        return False

    year = value[:4]
    week = value[6:]

    if not year.isdigit() or not week.isdigit():
        return False

    week_num = int(week)

    return 1 <= week_num <= 53


# =============================================================================
# TEMPLATE EXCEL
# =============================================================================


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
def billing_masivo_template(request):
    wb = Workbook()

    ws_b = wb.active
    ws_b.title = SHEET_BILLINGS

    ws_t = wb.create_sheet(SHEET_TECHNICIANS)
    ws_i = wb.create_sheet(SHEET_ITEMS)
    ws_help = wb.create_sheet("Instructions")

    _write_sheet_header(ws_b, BILLINGS_HEADERS)
    _write_sheet_header(ws_t, TECHNICIANS_HEADERS)
    _write_sheet_header(ws_i, ITEMS_HEADERS)

    # ==========================================================
    # Examples
    # ==========================================================
    ws_b.append(
        [
            "BILL-001",
            "0913UA_02_1000-012",
            "ITG",
            "Chile",
            "Underground",
            "PC676",
            "123 Main St",
            "2026-W20",
            "full",
            "NO",
            "NO",
            "fiber",
            "B8G Fiber / Photo",
        ]
    )

    # Nueva forma recomendada: varios técnicos en una misma celda.
    ws_t.append(["BILL-001", "tech1, tech2, tech3"])

    # Forma anterior sigue funcionando:
    ws_t.append(["BILL-001", "another.tech"])

    ws_i.append(["BILL-001", "C-123", "1"])

    # ==========================================================
    # Instructions sheet - visual
    # ==========================================================
    ws_help.sheet_view.showGridLines = False

    dark_fill = PatternFill("solid", fgColor="1F2937")
    blue_fill = PatternFill("solid", fgColor="DBEAFE")
    green_fill = PatternFill("solid", fgColor="DCFCE7")
    amber_fill = PatternFill("solid", fgColor="FEF3C7")
    red_fill = PatternFill("solid", fgColor="FEE2E2")
    gray_fill = PatternFill("solid", fgColor="F3F4F6")

    title_font = Font(color="FFFFFF", bold=True, size=16)
    section_font = Font(color="111827", bold=True, size=12)
    bold_font = Font(bold=True)
    normal_font = Font(color="374151", size=11)
    warning_font = Font(color="92400E", bold=True)
    error_font = Font(color="991B1B", bold=True)

    ws_help.merge_cells("A1:F1")
    ws_help["A1"] = "Bulk Billing Import Guide"
    ws_help["A1"].fill = dark_fill
    ws_help["A1"].font = title_font
    ws_help["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws_help.row_dimensions[1].height = 28

    row = 3

    def section(title, fill):
        nonlocal row
        ws_help.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        cell = ws_help.cell(row=row, column=1)
        cell.value = title
        cell.fill = fill
        cell.font = section_font
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws_help.row_dimensions[row].height = 22
        row += 1

    def line(label, value="", note=""):
        nonlocal row
        ws_help.cell(row=row, column=1).value = label
        ws_help.cell(row=row, column=1).font = bold_font
        ws_help.cell(row=row, column=2).value = value
        ws_help.cell(row=row, column=2).font = normal_font

        if note:
            ws_help.merge_cells(
                start_row=row, start_column=3, end_row=row, end_column=6
            )
            ws_help.cell(row=row, column=3).value = note
            ws_help.cell(row=row, column=3).font = normal_font

        for col in range(1, 7):
            ws_help.cell(row=row, column=col).alignment = Alignment(
                vertical="top",
                wrap_text=True,
            )

        row += 1

    def blank():
        nonlocal row
        row += 1

    section("1. General workflow", blue_fill)
    line("Step 1", "Fill the Billings sheet.", "One row per billing.")
    line(
        "Step 2",
        "Fill the Technicians sheet.",
        "You can use one row per technician or many technicians in one cell.",
    )
    line(
        "Step 3",
        "Fill the Items sheet.",
        "Each item must use a valid Job Code and quantity.",
    )
    line(
        "Step 4",
        "Upload the file.",
        "The system validates everything before creating any billing.",
    )
    line(
        "Important",
        "If one row has an error, nothing will be created.",
        "Fix the file and upload it again.",
    )
    blank()

    section("2. Billings sheet", green_fill)
    line("bulk_key", "Required", "Unique key inside the file. Example: BILL-001.")
    line("project_id", "Required", "Final Project ID visible in Billing List.")
    line("client", "Required", "Must match Technician Prices.")
    line("city", "Required", "Must match Technician Prices.")
    line(
        "project", "Required", "Must match the Project value used in Technician Prices."
    )
    line("office", "Required", "Must match Technician Prices.")
    line("project_address", "Optional", "Address or Google Maps link.")
    line("projected_week", "Required", "ISO format YYYY-W##. Example: 2026-W20.")
    line("tech_payment_mode", "Required", "Only full or split.")
    line("direct_discount", "Required", "YES or NO.")
    line("cable_installation", "Required", "YES or NO.")
    line("requirement_type", "Optional", "Use none, fiber or cable.")
    line(
        "requirement_list",
        "Optional",
        "Exact active Requirement List name for the selected project and type.",
    )
    blank()

    section("3. Technicians sheet", amber_fill)
    line("Option A", "One technician per row", "Example: BILL-001 | tech1")
    line(
        "Option B",
        "Many technicians in one cell",
        "Example: BILL-001 | tech1, tech2, tech3",
    )
    line(
        "Accepted separators",
        "Comma or semicolon",
        "Examples: tech1, tech2, tech3  OR  tech1; tech2; tech3",
    )
    line(
        "Rule",
        "Username must match exactly",
        "The user must exist and must have technician/user role.",
    )
    line(
        "Do not duplicate",
        "Same technician cannot repeat in the same billing.",
        "The preview will show an error.",
    )
    blank()

    section("4. Items sheet", blue_fill)
    line("bulk_key", "Required", "Must match one billing from the Billings sheet.")
    line("job_code", "Required", "Must match exactly what exists in Technician Prices.")
    line("quantity", "Required", "Cannot be zero.")
    line("Normal billing", "Positive quantity", "Example: 1, 2, 3.")
    line("Direct discount", "Negative quantity", "Example: -1, -2.")
    blank()

    section("5. Technician payment mode", green_fill)
    line(
        "full",
        "Full amount for each technician",
        "Example: 2 technicians, qty 1, rate 100. Each technician receives 100. Tech total = 200.",
    )
    line(
        "split",
        "Split between technicians",
        "Example: 2 technicians, qty 1, rate 100. Each technician receives 50. Tech total = 100.",
    )
    line(
        "Do not write",
        "Full amount / Split between technicians / yes / no",
        "Only full or split are valid.",
    )
    blank()

    section("6. Requirement lists", amber_fill)
    line("none", "No requirements loaded", "Leave requirement_list empty.")
    line(
        "fiber",
        "Loads Fiber / Photo requirements",
        "requirement_list must match an active Fiber / Photo Requirement List.",
    )
    line(
        "cable",
        "Loads Cable requirements",
        "cable_installation must be YES and requirement_list must match an active Cable Requirement List.",
    )
    line(
        "Exact name",
        "Requirement list name must match exactly.",
        "The preview validates the list before creating billings.",
    )
    blank()

    section("7. Common errors", red_fill)
    line(
        "Invalid Job Code",
        "Job Code does not match Technician Prices.",
        "Example: C-123 is not the same as C.123.",
    )
    line(
        "Missing technician",
        "The username does not exist.",
        "Check spelling and spaces.",
    )
    line(
        "Missing price",
        "Technician has no matching price.",
        "Client, City, Project, Office and Job Code must match.",
    )
    line(
        "Wrong quantity",
        "Quantity cannot be zero.",
        "Discounts require negative quantities.",
    )
    line(
        "Wrong requirement list",
        "List does not exist or is inactive.",
        "Check Project, requirement_type and requirement_list.",
    )
    blank()

    # Bottom warning box
    ws_help.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws_help.cell(row=row, column=1).value = (
        "IMPORTANT: Do not rename sheets or headers. "
        "If a single row has an error, no billing will be created."
    )
    ws_help.cell(row=row, column=1).fill = red_fill
    ws_help.cell(row=row, column=1).font = error_font
    ws_help.cell(row=row, column=1).alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )
    ws_help.row_dimensions[row].height = 35

    # Sheet formatting
    for ws in [ws_b, ws_t, ws_i, ws_help]:
        _autosize_sheet(ws)

    ws_help.column_dimensions["A"].width = 24
    ws_help.column_dimensions["B"].width = 32
    ws_help.column_dimensions["C"].width = 24
    ws_help.column_dimensions["D"].width = 24
    ws_help.column_dimensions["E"].width = 24
    ws_help.column_dimensions["F"].width = 24

    # Highlight example rows
    for ws in [ws_b, ws_t, ws_i]:
        for cell in ws[2]:
            cell.fill = gray_fill
            cell.alignment = Alignment(wrap_text=True)

        if ws.max_row >= 3:
            for cell in ws[3]:
                cell.fill = gray_fill
                cell.alignment = Alignment(wrap_text=True)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = (
        'attachment; filename="bulk_billing_template.xlsx"'
    )
    return response


def _write_sheet_header(ws, headers):
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F2937")
    font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"


def _autosize_sheet(ws):
    for column_cells in ws.columns:
        max_len = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = _clean_cell(cell.value)
            max_len = max(max_len, len(value))

        ws.column_dimensions[column_letter].width = min(max(max_len + 2, 14), 45)

# =============================================================================
# CACHE TEMPORAL PARA PREVIEW MASIVO
# =============================================================================

BULK_BILLING_CACHE_TIMEOUT = 60 * 60  # 1 hora


def _bulk_billing_cache_key(user_id, token):
    return f"billing_masivo_preview:{user_id}:{token}"


def _save_bulk_billing_preview(request, payload):
    """
    Guarda el preview pesado en cache y deja en sesión solo un token liviano.
    Esto evita que archivos grandes rompan o vacíen la sesión.
    """
    token = uuid.uuid4().hex
    key = _bulk_billing_cache_key(request.user.id, token)

    cache.set(key, payload, BULK_BILLING_CACHE_TIMEOUT)

    request.session["billing_masivo_preview_token"] = token
    request.session.modified = True


def _get_bulk_billing_preview(request):
    """
    Recupera el preview usando el token guardado en sesión.
    """
    token = request.session.get("billing_masivo_preview_token")

    if not token:
        return None

    key = _bulk_billing_cache_key(request.user.id, token)
    return cache.get(key)


def _clear_bulk_billing_preview(request):
    """
    Limpia el preview temporal después de confirmar o cuando ya no se necesita.
    """
    token = request.session.get("billing_masivo_preview_token")

    if token:
        key = _bulk_billing_cache_key(request.user.id, token)
        cache.delete(key)

    request.session.pop("billing_masivo_preview_token", None)
    request.session.modified = True

# =============================================================================
# UPLOAD + PREVIEW
# =============================================================================


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
def billing_masivo_upload(request):
    if request.method == "POST":
        form = BillingMasivoUploadForm(request.POST, request.FILES)

        if form.is_valid():
            archivo = form.cleaned_data["archivo"]
            preview_payload = _build_preview_from_excel(archivo, request.user)

            _save_bulk_billing_preview(request, preview_payload)

            return redirect("operaciones:billing_masivo_preview")

        messages.error(
            request,
            "The uploaded file is not valid. Please choose a valid .xlsx file.",
        )
    else:
        form = BillingMasivoUploadForm()

    return render(
        request,
        "operaciones/billing_masivo/upload.html",
        {
            "form": form,
            "template_url": reverse("operaciones:billing_masivo_template"),
        },
    )


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
def billing_masivo_preview(request):
    payload = _get_bulk_billing_preview(request)

    if not payload:
        messages.warning(request, "Please upload a bulk billing file first.")
        return redirect("operaciones:billing_masivo_upload")

    return render(
        request,
        "operaciones/billing_masivo/preview.html",
        {
            "payload": payload,
            "has_errors": bool(payload.get("has_errors")),
            "template_url": reverse("operaciones:billing_masivo_template"),
        },
    )


# =============================================================================
# VALIDACIÓN PRINCIPAL
# =============================================================================


def _build_preview_from_excel(archivo, user=None):
    global_errors = []

    price_perms = _bulk_billing_price_permissions(user)

    try:
        wb = load_workbook(archivo, data_only=True)
    except Exception:
        return {
            "ok": False,
            "has_errors": True,
            "permissions": price_perms,
            "global_errors": [
                _cell_error(
                    "Workbook",
                    1,
                    "archivo",
                    "The file could not be read. Please upload a valid .xlsx file.",
                )
            ],
            "billings": [],
            "summary": {
                "billing_count": 0,
                "item_count": 0,
                "technician_count": 0,
                "subtotal_tecnico": "0.00",
                "subtotal_empresa": "0.00",
            },
        }

    billing_rows, errors_b = _read_sheet_rows(wb, SHEET_BILLINGS, BILLINGS_HEADERS)
    tech_rows, errors_t = _read_sheet_rows(wb, SHEET_TECHNICIANS, TECHNICIANS_HEADERS)
    item_rows, errors_i = _read_sheet_rows(wb, SHEET_ITEMS, ITEMS_HEADERS)

    global_errors.extend(errors_b)
    global_errors.extend(errors_t)
    global_errors.extend(errors_i)

    if global_errors:
        return {
            "ok": False,
            "has_errors": True,
            "permissions": price_perms,
            "global_errors": global_errors,
            "billings": [],
            "summary": {
                "billing_count": 0,
                "item_count": 0,
                "technician_count": 0,
                "subtotal_tecnico": "0.00",
                "subtotal_empresa": "0.00",
            },
        }

    billings_by_key = {}

    for row in billing_rows:
        bulk_key = _clean_cell(row.get("bulk_key"))

        if not bulk_key:
            global_errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "bulk_key",
                    "bulk_key is required.",
                )
            )
            continue

        if bulk_key in billings_by_key:
            global_errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "bulk_key",
                    f"Duplicate bulk_key '{bulk_key}'. Each billing must have a unique bulk_key.",
                )
            )
            continue

        direct_discount, dd_error = _parse_bool(row.get("direct_discount"))
        cable_installation, cable_error = _parse_bool(row.get("cable_installation"))

        payment_mode = _clean_cell(row.get("tech_payment_mode")).lower() or "full"

        requirement_type = _clean_cell(row.get("requirement_type")).lower() or "none"
        requirement_list_name = _clean_cell(row.get("requirement_list"))

        preview = PreviewBilling(
            bulk_key=bulk_key,
            source_row=row["__rownum"],
            project_id=_clean_cell(row.get("project_id")),
            client=_clean_cell(row.get("client")),
            city=_clean_cell(row.get("city")),
            project=_clean_cell(row.get("project")),
            office=_clean_cell(row.get("office")),
            project_address=_clean_cell(row.get("project_address")),
            projected_week=_clean_cell(row.get("projected_week")).upper(),
            tech_payment_mode=payment_mode,
            direct_discount=direct_discount,
            cable_installation=cable_installation,
            requirement_type=requirement_type,
            requirement_list=requirement_list_name,
        )

        required_fields = [
            "project_id",
            "client",
            "city",
            "project",
            "office",
            "projected_week",
            "tech_payment_mode",
        ]

        for field_name in required_fields:
            if not _clean_cell(row.get(field_name)):
                preview.errors.append(
                    _cell_error(
                        SHEET_BILLINGS,
                        row["__rownum"],
                        field_name,
                        "This field is required.",
                    )
                )

        if payment_mode not in VALID_PAYMENT_MODES:
            preview.errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "tech_payment_mode",
                    "Use full or split.",
                )
            )

        if dd_error:
            preview.errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "direct_discount",
                    dd_error,
                )
            )

        if cable_error:
            preview.errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "cable_installation",
                    cable_error,
                )
            )

        if requirement_type not in VALID_REQUIREMENT_TYPES:
            preview.errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "requirement_type",
                    "Use none, fiber or cable.",
                )
            )

        if requirement_type in ("", "none"):
            preview.requirement_type = "none"
            preview.requirement_list = ""
        elif not requirement_list_name:
            preview.errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "requirement_list",
                    "Requirement list is required when requirement_type is fiber or cable.",
                )
            )

        if requirement_type == "cable" and not cable_installation:
            preview.errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "cable_installation",
                    "Cable requirement lists require cable_installation = YES.",
                )
            )

        if preview.projected_week and not _iso_week_is_valid(preview.projected_week):
            preview.errors.append(
                _cell_error(
                    SHEET_BILLINGS,
                    row["__rownum"],
                    "projected_week",
                    "Use ISO week format YYYY-W##. Example: 2026-W20.",
                )
            )

        billings_by_key[bulk_key] = preview

    tech_rows_by_key = defaultdict(list)
    item_rows_by_key = defaultdict(list)

    for row in tech_rows:
        key = _clean_cell(row.get("bulk_key"))
        tech_rows_by_key[key].append(row)

    for row in item_rows:
        key = _clean_cell(row.get("bulk_key"))
        item_rows_by_key[key].append(row)

    for row in tech_rows:
        key = _clean_cell(row.get("bulk_key"))

        if not key:
            global_errors.append(
                _cell_error(
                    SHEET_TECHNICIANS,
                    row["__rownum"],
                    "bulk_key",
                    "bulk_key is required.",
                )
            )
            continue

        if key not in billings_by_key:
            global_errors.append(
                _cell_error(
                    SHEET_TECHNICIANS,
                    row["__rownum"],
                    "bulk_key",
                    f"bulk_key '{key}' does not exist in Billings sheet.",
                )
            )

    for row in item_rows:
        key = _clean_cell(row.get("bulk_key"))

        if not key:
            global_errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    row["__rownum"],
                    "bulk_key",
                    "bulk_key is required.",
                )
            )
            continue

        if key not in billings_by_key:
            global_errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    row["__rownum"],
                    "bulk_key",
                    f"bulk_key '{key}' does not exist in Billings sheet.",
                )
            )

    for bulk_key, preview in billings_by_key.items():
        _attach_and_validate_technicians(preview, tech_rows_by_key.get(bulk_key, []))
        _attach_and_validate_items(preview, item_rows_by_key.get(bulk_key, []))
        _validate_project_and_prices(preview)

    billings_preview = list(billings_by_key.values())

    total_tecnico = Decimal("0.00")
    total_empresa = Decimal("0.00")
    total_items = 0
    total_tech_rows = 0

    for b in billings_preview:
        total_tecnico += b.subtotal_tecnico or Decimal("0.00")
        total_empresa += b.subtotal_empresa or Decimal("0.00")
        total_items += len(b.items)
        total_tech_rows += len(b.technicians)

    payload = {
        "ok": True,
        "has_errors": bool(
            global_errors
            or any(b.errors or any(i.errors for i in b.items) for b in billings_preview)
        ),
        "permissions": price_perms,
        "global_errors": global_errors,
        "billings": [_billing_to_dict(b) for b in billings_preview],
        "summary": {
            "billing_count": len(billings_preview),
            "item_count": total_items,
            "technician_count": total_tech_rows,
            "subtotal_tecnico": _format_money(total_tecnico),
            "subtotal_empresa": _format_money(total_empresa),
        },
    }

    return payload


def _split_technician_usernames(value):
    """
    Permite cargar técnicos de dos formas:

    1) Una fila por técnico:
       BILL-001 | tech1

    2) Varios técnicos en el mismo campo:
       BILL-001 | tech1, tech2, tech3
       BILL-001 | tech1; tech2; tech3

    No modifica usernames internamente, solo separa por coma, punto y coma o salto de línea.
    """
    raw = _clean_cell(value)

    if not raw:
        return []

    parts = []

    for chunk in raw.replace("\n", ",").replace(";", ",").split(","):
        username = _clean_cell(chunk)

        if username:
            parts.append(username)

    return parts


def _attach_and_validate_technicians(preview: PreviewBilling, rows):
    if not rows:
        preview.errors.append(
            _cell_error(
                SHEET_TECHNICIANS,
                preview.source_row,
                "technician_username",
                "At least one technician is required for this billing.",
            )
        )
        return

    seen = set()

    for row in rows:
        raw_usernames = _clean_cell(row.get("technician_username"))

        if not raw_usernames:
            preview.errors.append(
                _cell_error(
                    SHEET_TECHNICIANS,
                    row["__rownum"],
                    "technician_username",
                    "technician_username is required.",
                )
            )

            preview.technicians.append(
                PreviewTechnician(
                    source_row=row["__rownum"],
                    username="",
                )
            )
            continue

        usernames = (
            raw_usernames.replace(";", ",")
            .replace("\n", ",")
            .replace("\r", ",")
            .split(",")
        )

        usernames = [
            _clean_cell(username) for username in usernames if _clean_cell(username)
        ]

        if not usernames:
            preview.errors.append(
                _cell_error(
                    SHEET_TECHNICIANS,
                    row["__rownum"],
                    "technician_username",
                    "technician_username is required.",
                )
            )
            continue

        for username in usernames:
            tech = PreviewTechnician(
                source_row=row["__rownum"],
                username=username,
            )

            normalized_username = username.strip().lower()

            if normalized_username in seen:
                preview.errors.append(
                    _cell_error(
                        SHEET_TECHNICIANS,
                        row["__rownum"],
                        "technician_username",
                        f"Duplicate technician '{username}' in this billing.",
                    )
                )
                preview.technicians.append(tech)
                continue

            seen.add(normalized_username)

            user = CustomUser.objects.filter(username__iexact=username).first()

            if not user:
                preview.errors.append(
                    _cell_error(
                        SHEET_TECHNICIANS,
                        row["__rownum"],
                        "technician_username",
                        f"Technician username '{username}' does not exist.",
                    )
                )
                preview.technicians.append(tech)
                continue

            try:
                is_user_role = user.roles.filter(nombre__iexact="usuario").exists()
            except Exception:
                is_user_role = True

            if not is_user_role:
                preview.errors.append(
                    _cell_error(
                        SHEET_TECHNICIANS,
                        row["__rownum"],
                        "technician_username",
                        f"User '{username}' is not a technician user.",
                    )
                )

            tech.user_id = user.id
            tech.display_name = _display_user(user)
            preview.technicians.append(tech)


def _attach_and_validate_items(preview: PreviewBilling, rows):
    if not rows:
        preview.errors.append(
            _cell_error(
                SHEET_ITEMS,
                preview.source_row,
                "job_code",
                "At least one item is required for this billing.",
            )
        )
        return

    for row in rows:
        job_code = _clean_cell(row.get("job_code"))
        quantity, quantity_raw = _to_decimal(row.get("quantity"))

        item = PreviewItem(
            source_row=row["__rownum"],
            job_code=job_code,
            quantity_raw=quantity_raw,
            quantity=quantity,
        )

        if not job_code:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    row["__rownum"],
                    "job_code",
                    "job_code is required.",
                )
            )

        if quantity is None:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    row["__rownum"],
                    "quantity",
                    "Quantity must be a valid number.",
                )
            )
        elif quantity == 0:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    row["__rownum"],
                    "quantity",
                    "Quantity cannot be zero.",
                )
            )
        elif preview.direct_discount and quantity >= 0:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    row["__rownum"],
                    "quantity",
                    "Direct discount requires a negative quantity.",
                )
            )
        elif not preview.direct_discount and quantity <= 0:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    row["__rownum"],
                    "quantity",
                    "Normal billing requires a positive quantity.",
                )
            )

        preview.items.append(item)


def _validate_requirement_list_for_preview(preview: PreviewBilling, project):
    requirement_type = (preview.requirement_type or "none").strip().lower()
    requirement_list_name = (preview.requirement_list or "").strip()

    if requirement_type in ("", "none"):
        preview.requirement_type = "none"
        preview.requirement_list = ""
        preview.requirement_list_id = None
        preview.requirement_list_label = ""
        preview.requirement_count = 0
        return

    if requirement_type not in (
        RequirementList.LIST_TYPE_FIBER,
        RequirementList.LIST_TYPE_CABLE,
    ):
        preview.errors.append(
            _cell_error(
                SHEET_BILLINGS,
                preview.source_row,
                "requirement_type",
                "Use none, fiber or cable.",
            )
        )
        return

    if not requirement_list_name:
        preview.errors.append(
            _cell_error(
                SHEET_BILLINGS,
                preview.source_row,
                "requirement_list",
                "Requirement list is required when requirement_type is fiber or cable.",
            )
        )
        return

    qs = (
        RequirementList.objects.filter(
            project=project,
            list_type=requirement_type,
            is_active=True,
            name__iexact=requirement_list_name,
        )
        .prefetch_related("items")
    )

    count = qs.count()

    if count == 0:
        preview.errors.append(
            _cell_error(
                SHEET_BILLINGS,
                preview.source_row,
                "requirement_list",
                (
                    f"Active Requirement List '{requirement_list_name}' was not found "
                    f"for Project '{project.nombre}' and type '{requirement_type}'."
                ),
            )
        )
        return

    if count > 1:
        preview.errors.append(
            _cell_error(
                SHEET_BILLINGS,
                preview.source_row,
                "requirement_list",
                (
                    f"More than one active Requirement List named '{requirement_list_name}' "
                    f"was found for Project '{project.nombre}' and type '{requirement_type}'."
                ),
            )
        )
        return

    req_list = qs.first()
    req_count = req_list.items.count()

    if req_count <= 0:
        preview.errors.append(
            _cell_error(
                SHEET_BILLINGS,
                preview.source_row,
                "requirement_list",
                f"Requirement List '{req_list.name}' has no requirements.",
            )
        )
        return

    preview.requirement_type = requirement_type
    preview.requirement_list = req_list.name
    preview.requirement_list_id = req_list.id
    preview.requirement_list_label = f"{req_list.name} ({req_count} item(s))"
    preview.requirement_count = req_count


def _validate_project_and_prices(preview: PreviewBilling):
    """
    Valida:
    - que exista el Proyecto indicado en project
    - que existan precios para:
      Technician + Client + City + Proyecto FK + Office + Job Code
    - que exista Requirement List si el Excel trae requirement_type/list.
    """

    if preview.errors:
        return

    project, project_error = _find_project_robust(preview)

    if not project:
        preview.errors.append(
            _cell_error(
                SHEET_BILLINGS,
                preview.source_row,
                "project",
                project_error,
            )
        )
        return

    preview.project = project.nombre

    _validate_requirement_list_for_preview(preview, project)

    if preview.errors:
        return

    valid_techs = [t for t in preview.technicians if t.user_id]

    if not valid_techs:
        return

    for item in preview.items:
        if item.errors:
            continue

        _hydrate_item_prices(preview, item, project, valid_techs)


def _hydrate_item_prices(
    preview: PreviewBilling, item: PreviewItem, project, valid_techs
):
    qty = item.quantity or Decimal("0.00")

    prices = []

    # Estos valores vienen del Excel y deben calzar contra PrecioActividadTecnico
    client = _clean_cell(preview.client)
    city = _clean_cell(preview.city)
    office = _clean_cell(preview.office)

    for tech in valid_techs:
        price, price_error = _find_price_robust(
            tech_id=tech.user_id,
            project=project,
            client=client,
            city=city,
            office=office,
            job_code=item.job_code,
        )

        if price_error:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    item.source_row,
                    "job_code",
                    price_error,
                )
            )
            continue

        if not price:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    item.source_row,
                    "job_code",
                    (
                        f"Technician '{tech.username}' does not have a matching price for "
                        f"Client '{client}', City '{city}', Project '{project.nombre}', "
                        f"Office '{office}', Job Code '{item.job_code}'."
                    ),
                )
            )
            continue

        prices.append((tech, price))

    if item.errors:
        return

    if not prices:
        item.errors.append(
            _cell_error(
                SHEET_ITEMS,
                item.source_row,
                "job_code",
                f"Job Code '{item.job_code}' does not exist for the selected configuration.",
            )
        )
        return

    first_price = prices[0][1]

    # Canonicalizar header desde la tabla de precios
    preview.client = first_price.cliente
    preview.city = first_price.ciudad
    preview.office = first_price.oficina
    preview.project = project.nombre

    item.tipo_trabajo = first_price.tipo_trabajo
    item.descripcion = first_price.descripcion
    item.unidad_medida = first_price.unidad_medida
    item.precio_empresa = first_price.precio_empresa or Decimal("0.00")

    for tech, price in prices:
        if not _same_text(price.tipo_trabajo, first_price.tipo_trabajo):
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    item.source_row,
                    "job_code",
                    f"Job Code '{item.job_code}' has different Work Type for technician '{tech.username}'.",
                )
            )

        if not _same_text(price.descripcion, first_price.descripcion):
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    item.source_row,
                    "job_code",
                    f"Job Code '{item.job_code}' has different Description for technician '{tech.username}'.",
                )
            )

        if not _same_text(price.unidad_medida, first_price.unidad_medida):
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    item.source_row,
                    "job_code",
                    f"Job Code '{item.job_code}' has different UOM for technician '{tech.username}'.",
                )
            )

        if price.precio_empresa != first_price.precio_empresa:
            item.errors.append(
                _cell_error(
                    SHEET_ITEMS,
                    item.source_row,
                    "job_code",
                    f"Job Code '{item.job_code}' has different Company Price for technician '{tech.username}'.",
                )
            )

    if item.errors:
        return

    # Canonicalizar Job Code desde la tabla de precios.
    item.job_code = first_price.codigo_trabajo

    percentage = Decimal("100.00")

    if preview.tech_payment_mode == "split":
        percentage = (Decimal("100.00") / Decimal(len(prices))).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

    total_tech = Decimal("0.00")

    for tech, price in prices:
        tarifa_base = price.precio_tecnico or Decimal("0.00")

        if preview.tech_payment_mode == "full":
            tarifa_efectiva = tarifa_base
            pct = Decimal("100.00")
        else:
            pct = percentage
            tarifa_efectiva = (tarifa_base * pct / Decimal("100.00")).quantize(
                Decimal("0.01"),
                rounding=ROUND_HALF_UP,
            )

        subtotal = (qty * tarifa_efectiva).quantize(
            Decimal("0.01"),
            rounding=ROUND_HALF_UP,
        )

        total_tech += subtotal

        item.desglose_tecnico.append(
            {
                "tecnico_id": tech.user_id,
                "tecnico_username": tech.username,
                "tecnico_nombre": tech.display_name,
                "tarifa_base": _format_money(tarifa_base),
                "porcentaje": _format_money(pct),
                "tarifa_efectiva": _format_money(tarifa_efectiva),
                "subtotal": _format_money(subtotal),
                "payment_weeks": int(price.payment_weeks or 0),
            }
        )

    item.subtotal_tecnico = total_tech.quantize(Decimal("0.01"))
    item.subtotal_empresa = (qty * item.precio_empresa).quantize(
        Decimal("0.01"),
        rounding=ROUND_HALF_UP,
    )

    preview.subtotal_tecnico += item.subtotal_tecnico
    preview.subtotal_empresa += item.subtotal_empresa


def _billing_to_dict(preview: PreviewBilling):
    return {
        "bulk_key": preview.bulk_key,
        "source_row": preview.source_row,
        "project_id": preview.project_id,
        "client": preview.client,
        "city": preview.city,
        "project": preview.project,
        "office": preview.office,
        "project_address": preview.project_address,
        "projected_week": preview.projected_week,
        "tech_payment_mode": preview.tech_payment_mode,
        "direct_discount": preview.direct_discount,
        "cable_installation": preview.cable_installation,
        "requirement_type": preview.requirement_type,
        "requirement_list": preview.requirement_list,
        "requirement_list_id": preview.requirement_list_id,
        "requirement_list_label": preview.requirement_list_label,
        "requirement_count": preview.requirement_count,
        "subtotal_tecnico": _format_money(preview.subtotal_tecnico),
        "subtotal_empresa": _format_money(preview.subtotal_empresa),
        "errors": preview.errors,
        "technicians": [
            {
                "source_row": t.source_row,
                "username": t.username,
                "user_id": t.user_id,
                "display_name": t.display_name,
            }
            for t in preview.technicians
        ],
        "items": [
            {
                "source_row": item.source_row,
                "job_code": item.job_code,
                "quantity_raw": item.quantity_raw,
                "quantity": (
                    _format_money(item.quantity or Decimal("0.00"))
                    if item.quantity is not None
                    else ""
                ),
                "tipo_trabajo": item.tipo_trabajo,
                "descripcion": item.descripcion,
                "unidad_medida": item.unidad_medida,
                "precio_empresa": _format_money(item.precio_empresa),
                "subtotal_tecnico": _format_money(item.subtotal_tecnico),
                "subtotal_empresa": _format_money(item.subtotal_empresa),
                "desglose_tecnico": item.desglose_tecnico,
                "errors": item.errors,
            }
            for item in preview.items
        ],
    }


# =============================================================================
# CONFIRMAR CREACIÓN
# =============================================================================


@login_required
@rol_requerido("admin", "pm", "supervisor", "facturacion", "emision_facturacion")
@transaction.atomic
def billing_masivo_confirm(request):
    if request.method != "POST":
        return redirect("operaciones:billing_masivo_upload")

    payload = _get_bulk_billing_preview(request)

    if not payload:
        messages.warning(request, "Please upload a bulk billing file first.")
        return redirect("operaciones:billing_masivo_upload")

    if payload.get("has_errors"):
        messages.error(
            request, "The file still has validation errors. No billing was created."
        )
        return redirect("operaciones:billing_masivo_preview")

    billings = payload.get("billings") or []

    if not billings:
        messages.error(request, "There are no billings to create.")
        return redirect("operaciones:billing_masivo_upload")

    created_ids = []

    for b in billings:
        sesion = SesionBilling.objects.create(
            creado_en=timezone.now(),
            is_direct_discount=bool(b.get("direct_discount")),
            is_cable_installation=bool(b.get("cable_installation")),
            tech_payment_mode=b.get("tech_payment_mode") or "full",
            proyecto_id=b.get("project_id") or "",
            cliente=b.get("client") or "",
            ciudad=b.get("city") or "",
            proyecto=b.get("project") or "",
            oficina=b.get("office") or "",
            direccion_proyecto=b.get("project_address") or "",
            semana_pago_proyectada=b.get("projected_week") or "",
            estado="asignado",
            subtotal_tecnico=Decimal(str(b.get("subtotal_tecnico") or "0.00")),
            subtotal_empresa=Decimal(str(b.get("subtotal_empresa") or "0.00")),
        )

        sesion.cliente = b.get("client") or ""
        sesion.ciudad = b.get("city") or ""
        sesion.proyecto = b.get("project") or ""
        sesion.oficina = b.get("office") or ""
        sesion.subtotal_tecnico = Decimal(str(b.get("subtotal_tecnico") or "0.00"))
        sesion.subtotal_empresa = Decimal(str(b.get("subtotal_empresa") or "0.00"))

        if sesion.is_direct_discount:
            sesion.finance_status = "review_discount"

        sesion.save(
            update_fields=[
                "cliente",
                "ciudad",
                "proyecto",
                "oficina",
                "subtotal_tecnico",
                "subtotal_empresa",
                "finance_status",
            ]
        )

        created_ids.append(sesion.id)

        technicians = b.get("technicians") or []
        tech_count = max(len(technicians), 1)
        created_tech_sessions = []

        for t in technicians:
            user_id = t.get("user_id")

            if not user_id:
                continue

            porcentaje = Decimal("100.00")

            if sesion.tech_payment_mode == "split":
                porcentaje = (Decimal("100.00") / Decimal(tech_count)).quantize(
                    Decimal("0.01"),
                    rounding=ROUND_HALF_UP,
                )

            tecnico_sesion = SesionBillingTecnico.objects.create(
                sesion=sesion,
                tecnico_id=user_id,
                porcentaje=porcentaje,
                estado="asignado",
                is_active=True,
            )

            created_tech_sessions.append(tecnico_sesion)

        _apply_requirement_list_to_sesion(
            sesion=sesion,
            requirement_list_id=b.get("requirement_list_id"),
            requirement_type=b.get("requirement_type"),
            tecnico_sesiones=created_tech_sessions,
        )

        for item_data in b.get("items") or []:
            item = ItemBilling.objects.create(
                sesion=sesion,
                codigo_trabajo=item_data.get("job_code") or "",
                tipo_trabajo=item_data.get("tipo_trabajo") or "",
                descripcion=item_data.get("descripcion") or "",
                unidad_medida=item_data.get("unidad_medida") or "",
                cantidad=Decimal(str(item_data.get("quantity") or "0.00")),
                precio_empresa=Decimal(str(item_data.get("precio_empresa") or "0.00")),
                subtotal_empresa=Decimal(
                    str(item_data.get("subtotal_empresa") or "0.00")
                ),
                subtotal_tecnico=Decimal(
                    str(item_data.get("subtotal_tecnico") or "0.00")
                ),
            )

            for d in item_data.get("desglose_tecnico") or []:
                ItemBillingTecnico.objects.create(
                    item=item,
                    tecnico_id=d.get("tecnico_id"),
                    tarifa_base=Decimal(str(d.get("tarifa_base") or "0.00")),
                    porcentaje=Decimal(str(d.get("porcentaje") or "0.00")),
                    tarifa_efectiva=Decimal(str(d.get("tarifa_efectiva") or "0.00")),
                    subtotal=Decimal(str(d.get("subtotal") or "0.00")),
                )

                _create_pay_week_snapshot(
                    sesion=sesion,
                    item=item,
                    tecnico_id=d.get("tecnico_id"),
                    codigo_trabajo=item.codigo_trabajo,
                    tipo_trabajo=item.tipo_trabajo,
                    payment_weeks=int(d.get("payment_weeks") or 0),
                    semana_base=sesion.semana_pago_proyectada,
                    tarifa_base=Decimal(str(d.get("tarifa_base") or "0.00")),
                    porcentaje=Decimal(str(d.get("porcentaje") or "0.00")),
                    tarifa_efectiva=Decimal(str(d.get("tarifa_efectiva") or "0.00")),
                    subtotal=Decimal(str(d.get("subtotal") or "0.00")),
                )

    _clear_bulk_billing_preview(request)

    messages.success(
        request,
        f"{len(created_ids)} billing(s) created successfully.",
    )

    return redirect("operaciones:listar_billing")


def _create_pay_week_snapshot(
    sesion,
    item,
    tecnico_id,
    codigo_trabajo,
    tipo_trabajo,
    payment_weeks,
    semana_base,
    tarifa_base,
    porcentaje,
    tarifa_efectiva,
    subtotal,
):
    semana_resultado = _add_weeks_to_iso_week(semana_base, payment_weeks)

    BillingPayWeekSnapshot.objects.create(
        sesion=sesion,
        tecnico_id=tecnico_id,
        item=item,
        codigo_trabajo=codigo_trabajo or "",
        tipo_trabajo=tipo_trabajo or "",
        payment_weeks=payment_weeks or 0,
        semana_base=semana_base or "",
        semana_resultado=semana_resultado or semana_base or "",
        tarifa_base=tarifa_base or Decimal("0.00"),
        porcentaje=porcentaje or Decimal("0.00"),
        tarifa_efectiva=tarifa_efectiva or Decimal("0.00"),
        subtotal=subtotal or Decimal("0.00"),
        payment_status="pending",
    )


def _add_weeks_to_iso_week(iso_week, weeks_to_add):
    iso_week = _clean_cell(iso_week).upper()

    if not _iso_week_is_valid(iso_week):
        return iso_week

    try:
        import datetime

        year = int(iso_week[:4])
        week = int(iso_week[6:])

        monday = datetime.date.fromisocalendar(year, week, 1)
        result = monday + datetime.timedelta(weeks=int(weeks_to_add or 0))
        y, w, _ = result.isocalendar()

        return f"{y}-W{int(w):02d}"
    except Exception:
        return iso_week


def _find_project_robust(preview: PreviewBilling):
    """
    Busca Proyecto de forma robusta pero segura.

    En el import masivo, project puede venir como:
    - Proyecto.nombre
    - Proyecto.codigo
    - Proyecto.id

    Para mantener simetría con Technician Prices, lo normal es usar
    el nombre visible del proyecto, ejemplo: Underground.
    """

    project_value = _clean_cell(preview.project)

    if not project_value:
        return None, "Project is required."

    candidates = list(Proyecto.objects.all())

    if project_value.isdigit():
        by_id = [p for p in candidates if str(p.id) == project_value]

        if len(by_id) == 1:
            return by_id[0], None

        if len(by_id) > 1:
            return (
                None,
                f"Project ID '{preview.project}' matched more than one project.",
            )

    by_name = [
        p for p in candidates if _same_text(getattr(p, "nombre", ""), project_value)
    ]

    if len(by_name) == 1:
        return by_name[0], None

    if len(by_name) > 1:
        return None, (
            f"Project name '{preview.project}' matched more than one project. "
            "Use Proyecto.codigo or Proyecto.id to avoid ambiguity."
        )

    by_code = [
        p for p in candidates if _same_text(getattr(p, "codigo", ""), project_value)
    ]

    if len(by_code) == 1:
        return by_code[0], None

    if len(by_code) > 1:
        return None, (
            f"Project code '{preview.project}' matched more than one project. "
            "Use the exact Project name from the system."
        )

    return None, (
        f"Project '{preview.project}' does not exist. "
        "Use the Project value shown in Technician Prices."
    )


def _apply_requirement_list_to_sesion(
    *,
    sesion,
    requirement_list_id,
    requirement_type,
    tecnico_sesiones,
):
    requirement_type = (requirement_type or "none").strip().lower()

    if not requirement_list_id or requirement_type in ("", "none"):
        return

    req_list = (
        RequirementList.objects.filter(
            id=requirement_list_id,
            is_active=True,
        )
        .prefetch_related("items")
        .first()
    )

    if not req_list:
        return

    if req_list.list_type == RequirementList.LIST_TYPE_FIBER:
        _apply_fiber_requirement_list_to_sesion(
            sesion=sesion,
            req_list=req_list,
            tecnico_sesiones=tecnico_sesiones,
        )
        return

    if req_list.list_type == RequirementList.LIST_TYPE_CABLE:
        _apply_cable_requirement_list_to_sesion(
            sesion=sesion,
            req_list=req_list,
            tecnico_sesiones=tecnico_sesiones,
        )
        return


def _apply_fiber_requirement_list_to_sesion(
    *,
    sesion,
    req_list,
    tecnico_sesiones,
):
    items = list(req_list.items.all().order_by("order", "id"))

    for item in items:
        title = (item.title or "").strip()

        if not title:
            continue

        plantilla, _ = RequisitoFotoBillingPlantilla.objects.update_or_create(
            sesion=sesion,
            slug=slugify(title),
            defaults={
                "titulo": title,
                "descripcion": item.description or "",
                "obligatorio": bool(item.required),
                "orden": item.order or 0,
                "needs_power_reading": bool(item.needs_power_reading),
                "needs_light_source_reading": bool(item.needs_light_source_reading),
                "power_port_no": item.power_port_no,
            },
        )

        for tecnico_sesion in tecnico_sesiones:
            RequisitoFotoBilling.objects.update_or_create(
                tecnico_sesion=tecnico_sesion,
                titulo=plantilla.titulo,
                defaults={
                    "descripcion": plantilla.descripcion or "",
                    "obligatorio": bool(plantilla.obligatorio),
                    "orden": plantilla.orden or 0,
                    "needs_power_reading": bool(plantilla.needs_power_reading),
                    "needs_light_source_reading": bool(
                        plantilla.needs_light_source_reading
                    ),
                    "power_port_no": plantilla.power_port_no,
                },
            )


def _apply_cable_requirement_list_to_sesion(
    *,
    sesion,
    req_list,
    tecnico_sesiones,
):
    """
    Crea los Cable Requirements del billing y los asigna a cada técnico.

    Fuente:
    - RequirementListItem.handhole
    - RequirementListItem.planned_reserve_ft
    - RequirementListItem.required
    - RequirementListItem.warning
    - RequirementListItem.order

    Destino:
    - cable_installation.CableRequirement
    - cable_installation.CableAssignmentRequirement
    """

    from cable_installation.models import (CableAssignmentRequirement,
                                           CableRequirement)

    items = list(req_list.items.all().order_by("order", "id"))

    if not items:
        return

    next_sequence = CableRequirement.next_sequence_for_billing(sesion)

    for idx, item in enumerate(items):
        handhole = (item.handhole or item.title or "").strip()

        if not handhole:
            continue

        sequence_no = next_sequence + idx

        cable_requirement, _ = CableRequirement.objects.update_or_create(
            billing=sesion,
            sequence_no=sequence_no,
            defaults={
                "handhole": handhole,
                "planned_reserve_ft": item.planned_reserve_ft or Decimal("0.00"),
                "warning": item.warning or "",
                "required": bool(item.required),
                "order": item.order or 0,
            },
        )

        for tecnico_sesion in tecnico_sesiones:
            CableAssignmentRequirement.objects.update_or_create(
                assignment=tecnico_sesion,
                requirement=cable_requirement,
                defaults={
                    "status": CableAssignmentRequirement.STATUS_PENDING,
                    "note": "",
                    "supervisor_note": "",
                },
            )
