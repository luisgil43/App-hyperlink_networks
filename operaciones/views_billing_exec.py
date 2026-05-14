import csv
import hashlib
import io
import json
import logging
import os
import re
import tempfile
import time
import uuid
from datetime import timedelta
from decimal import Decimal
from io import BytesIO
from tempfile import NamedTemporaryFile
from urllib.parse import urlencode

import boto3
import xlsxwriter
from botocore.client import Config
from django.conf import settings
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.views import redirect_to_login
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage as storage
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import (Case, Count, DecimalField, Exists, F, FloatField,
                              IntegerField, OuterRef, Prefetch, Q, Subquery,
                              Sum, Value, When)
from django.db.models.functions import Coalesce
from django.http import (FileResponse, Http404, HttpResponse,
                         HttpResponseBadRequest, HttpResponseForbidden,
                         HttpResponseNotAllowed, HttpResponseRedirect,
                         JsonResponse)
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.utils.dateparse import parse_datetime
from django.utils.html import strip_tags
from django.utils.http import http_date
from django.utils.text import slugify
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_GET, require_POST
from openpyxl import load_workbook  # asegúrate de tener openpyxl instalado
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from PIL import ExifTags, Image, ImageFile
from pillow_heif import register_heif_opener

from cable_installation.models import CableAssignmentRequirement, CableEvidence
from core.decorators import project_object_access_required
from core.permissions import (filter_queryset_by_access, projects_ids_for_user,
                              user_has_project_access)
from facturacion.models import CartolaMovimiento, Proyecto
from operaciones.excel_images import tmp_jpeg_from_filefield
from usuarios.decoradores import rol_requerido

from .models import (EvidenciaFotoBilling, ItemBillingTecnico,
                     ReporteFotograficoJob, RequisitoFotoBilling,
                     SesionBilling, SesionBillingTecnico)

log = logging.getLogger(__name__)


ImageFile.LOAD_TRUNCATED_IMAGES = True
register_heif_opener()  # habilita abrir .heic/.heif en Pillow
# ============================
# UTIL
# ============================

POWER_PORT_RE = re.compile(r"^\s*power\s*port\s*(\d+)\s*$", re.IGNORECASE)


def _power_meta_from_title(title: str):
    t = (title or "").strip()
    m = POWER_PORT_RE.match(t)
    if not m:
        return (False, None)
    try:
        n = int(m.group(1))
    except Exception:
        n = None
    if n is not None and not (1 <= n <= 8):
        n = None
    return (True, n)


def storage_file_exists(filefield) -> bool:
    if not filefield or not getattr(filefield, "name", ""):
        return False
    try:
        return filefield.storage.exists(filefield.name)
    except Exception:
        return False


def _has_ops_role(u):
    return (
        getattr(u, "es_pm", False) or
        getattr(u, "es_facturacion", False) or
        getattr(u, "es_admin_general", False) or
        u.is_superuser
    )

# ============================
# TÉCNICO
# ============================

def _is_asig_active(asig) -> bool:
    # Compat: si no existe el campo, asumimos activo (para no romper)
    return getattr(asig, "is_active", True) is True

def _get_my_active_assignment_or_404(request, pk: int):
    a = get_object_or_404(SesionBillingTecnico, pk=pk, tecnico=request.user)
    if not _is_asig_active(a):
        raise Http404()
    return a


def _cable_required_shots():
    return [
        CableEvidence.SHOT_START_CABLE,
        CableEvidence.SHOT_END_CABLE,
        CableEvidence.SHOT_HANDHOLE,
    ]


def _cable_assignment_rows(assignment):
    return list(
        CableAssignmentRequirement.objects.filter(assignment=assignment)
        .select_related("requirement", "assignment")
        .order_by("requirement__order", "requirement__sequence_no", "id")
    )


def _cable_present_shots_for_row(row):
    return set(
        CableEvidence.objects.filter(assignment_requirement=row)
        .exclude(shot_type="")
        .exclude(review_status=CableEvidence.REVIEW_REJECTED)
        .values_list("shot_type", flat=True)
        .distinct()
    )


def _cable_row_has_rejected_photo(row):
    return CableEvidence.objects.filter(
        assignment_requirement=row,
        review_status=CableEvidence.REVIEW_REJECTED,
    ).exists()


def _cable_row_latest_rejection_comment(row):
    ev = (
        CableEvidence.objects.filter(
            assignment_requirement=row,
            review_status=CableEvidence.REVIEW_REJECTED,
        )
        .exclude(review_comment="")
        .order_by("-reviewed_at", "-id")
        .first()
    )
    return ev.review_comment if ev else ""


def _cable_row_is_complete(row):
    req = row.requirement

    if req.start_ft is None:
        return False
    if req.planned_reserve_ft is None:
        return False
    if req.end_ft is None:
        return False
    if _cable_row_has_rejected_photo(row):
        return False

    present = _cable_present_shots_for_row(row)
    missing = [shot for shot in _cable_required_shots() if shot not in present]
    return len(missing) == 0


def _cable_assignment_missing_labels(assignment):
    labels = []
    rows = _cable_assignment_rows(assignment)
    row_by_req = {row.requirement_id: row for row in rows}

    requirements = assignment.sesion.cable_requirements.filter(required=True).order_by(
        "order", "sequence_no", "id"
    )

    for req in requirements:
        row = row_by_req.get(req.id)

        missing = []

        if req.start_ft is None or req.planned_reserve_ft is None or req.end_ft is None:
            missing.append("measurement")

        if not row:
            missing.append("photos")
        else:
            present = _cable_present_shots_for_row(row)
            pending = [shot for shot in _cable_required_shots() if shot not in present]
            if pending:
                missing.append("photos")
            if _cable_row_has_rejected_photo(row):
                missing.append("review")

        if missing:
            labels.append(f"PK {req.sequence_no} - {req.handhole} ({', '.join(missing)})")

    return labels


def _cable_assignment_has_rejected_photo(assignment):
    return CableEvidence.objects.filter(
        assignment_requirement__assignment=assignment,
        review_status=CableEvidence.REVIEW_REJECTED,
    ).exists()


def _cable_assignment_latest_rejection_comment(assignment):
    ev = (
        CableEvidence.objects.filter(
            assignment_requirement__assignment=assignment,
            review_status=CableEvidence.REVIEW_REJECTED,
        )
        .exclude(review_comment="")
        .order_by("-reviewed_at", "-id")
        .first()
    )
    return ev.review_comment if ev else ""


def _cable_assignment_can_finish(assignment):
    pendientes = []

    qs_asg = assignment.sesion.tecnicos_sesion.select_related("tecnico").all()
    try:
        SesionBillingTecnico._meta.get_field("is_active")
        qs_asg = qs_asg.filter(is_active=True)
    except Exception:
        pass

    for asg in qs_asg:
        accepted = bool(asg.aceptado_en) or asg.estado != "asignado"
        if not accepted:
            name = (
                getattr(asg.tecnico, "get_full_name", lambda: "")()
                or asg.tecnico.username
            )
            pendientes.append(name)

    faltantes = _cable_assignment_missing_labels(assignment)
    return not pendientes and not faltantes


@login_required
@rol_requerido("usuario", "admin", "pm", "supervisor")
def mis_assignments(request):
    import json
    from decimal import Decimal

    from django.core.paginator import Paginator
    from django.db.models import (Case, DecimalField, IntegerField, OuterRef,
                                  Q, Subquery, Sum, Value, When)
    from django.db.models.functions import Coalesce

    try:
        from cable_installation.models import (CableAssignmentRequirement,
                                               CableEvidence, CableRequirement)
    except Exception:
        CableAssignmentRequirement = None
        CableEvidence = None
        CableRequirement = None

    visibles = [
        "asignado",
        "en_proceso",
        "en_revision_supervisor",
        "rechazado_supervisor",
        "rechazado_pm",
        "rechazado_finanzas",
    ]

    base_qs = SesionBillingTecnico.objects.select_related("sesion", "tecnico").filter(
        tecnico=request.user,
        estado__in=visibles,
        sesion__is_direct_discount=False,
    )

    try:
        SesionBillingTecnico._meta.get_field("is_active")
        base_qs = base_qs.filter(is_active=True)
    except Exception:
        pass

    ibt = (
        ItemBillingTecnico.objects.filter(
            item__sesion=OuterRef("sesion_id"), tecnico=request.user
        )
        .values("tecnico")
        .annotate(total=Sum("subtotal"))
        .values("total")
    )

    dec_field = DecimalField(max_digits=12, decimal_places=2)

    asignaciones_qs = base_qs.annotate(
        my_total=Coalesce(
            Subquery(ibt, output_field=dec_field),
            Value(Decimal("0.00"), output_field=dec_field),
            output_field=dec_field,
        ),
        estado_priority=Case(
            When(estado="asignado", then=Value(1)),
            When(estado="en_proceso", then=Value(2)),
            When(estado="en_revision_supervisor", then=Value(3)),
            When(estado="rechazado_supervisor", then=Value(4)),
            When(estado="rechazado_pm", then=Value(5)),
            When(estado="rechazado_finanzas", then=Value(6)),
            default=Value(999),
            output_field=IntegerField(),
        ),
    ).order_by("estado_priority", "-sesion__creado_en", "-id")

    asignaciones = list(asignaciones_qs)

    proyectos_qs = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        "id",
    )

    for a in asignaciones:
        s = a.sesion
        proyecto_sel = None
        raw = (getattr(s, "proyecto", "") or "").strip()

        if raw:
            try:
                pid = int(raw)
            except (TypeError, ValueError):
                proyecto_sel = proyectos_qs.filter(
                    Q(nombre__iexact=raw) | Q(codigo__iexact=raw)
                ).first()
            else:
                proyecto_sel = proyectos_qs.filter(pk=pid).first()

        if not proyecto_sel and getattr(s, "proyecto_id", None):
            code = str(s.proyecto_id).strip()
            proyecto_sel = proyectos_qs.filter(
                Q(codigo__iexact=code) | Q(nombre__icontains=code)
            ).first()

        if proyecto_sel:
            a.proyecto_label = getattr(proyecto_sel, "nombre", str(proyecto_sel))
        else:
            a.proyecto_label = (
                getattr(s, "proyecto", None) or getattr(s, "proyecto_id", "") or ""
            ).strip()

    def _norm_title(x: str) -> str:
        return (x or "").strip().lower()

    def _cable_label(req):
        return f"PK {req.sequence_no} - {req.handhole}"

    by_session = {}
    for a in asignaciones:
        by_session.setdefault(a.sesion_id, []).append(a)

    sesion_ids = list(by_session.keys())

    pending_accept_names = {sid: [] for sid in sesion_ids}
    qs_asg = (
        SesionBillingTecnico.objects.filter(sesion_id__in=sesion_ids)
        .select_related("tecnico")
        .only(
            "sesion_id",
            "estado",
            "aceptado_en",
            "tecnico__username",
            "tecnico__first_name",
            "tecnico__last_name",
        )
    )
    try:
        SesionBillingTecnico._meta.get_field("is_active")
        qs_asg = qs_asg.filter(is_active=True)
    except Exception:
        pass

    for asg in qs_asg:
        sid = asg.sesion_id
        accepted = bool(asg.aceptado_en) or asg.estado != "asignado"
        if not accepted:
            name = (
                getattr(asg.tecnico, "get_full_name", lambda: "")()
                or asg.tecnico.username
            )
            pending_accept_names.setdefault(sid, []).append(name)

    cable_session_ids = [
        a.sesion_id
        for a in asignaciones
        if getattr(a.sesion, "is_cable_installation", False)
    ]
    normal_session_ids = [
        a.sesion_id
        for a in asignaciones
        if not getattr(a.sesion, "is_cable_installation", False)
    ]

    sample_map_by_sesion = {}
    req_titles_by_sesion = {}
    covered_by_sesion = {}

    if normal_session_ids:
        qs_sample = RequisitoFotoBilling.objects.filter(
            tecnico_sesion__sesion_id__in=normal_session_ids, titulo__isnull=False
        ).values_list("tecnico_sesion__sesion_id", "titulo")
        for sid, t in qs_sample:
            if not t:
                continue
            sample_map_by_sesion.setdefault(sid, {})
            sample_map_by_sesion[sid][_norm_title(t)] = t

        qs_req = RequisitoFotoBilling.objects.filter(
            tecnico_sesion__sesion_id__in=normal_session_ids, obligatorio=True
        ).values_list("tecnico_sesion__sesion_id", "titulo")
        for sid, t in qs_req:
            if t:
                req_titles_by_sesion.setdefault(sid, set()).add(_norm_title(t))

        qs_cov = (
            EvidenciaFotoBilling.objects.filter(
                tecnico_sesion__sesion_id__in=normal_session_ids,
                requisito__isnull=False,
            )
            .values_list("tecnico_sesion__sesion_id", "requisito__titulo")
            .distinct()
        )
        for sid, t in qs_cov:
            if t:
                covered_by_sesion.setdefault(sid, set()).add(_norm_title(t))

    cable_missing_by_sesion = {}
    cable_has_rejected_by_sesion = {}
    cable_rejection_comment_by_sesion = {}

    if (
        cable_session_ids
        and CableRequirement
        and CableAssignmentRequirement
        and CableEvidence
    ):
        requirements_by_session = {}
        qs_cable_req = CableRequirement.objects.filter(
            billing_id__in=cable_session_ids, required=True
        ).order_by("billing_id", "order", "sequence_no", "id")
        for req in qs_cable_req:
            requirements_by_session.setdefault(req.billing_id, []).append(req)

        rows_by_session = {}
        qs_rows = (
            CableAssignmentRequirement.objects.filter(
                assignment__sesion_id__in=cable_session_ids
            )
            .select_related("requirement", "assignment")
            .order_by(
                "assignment__sesion_id",
                "requirement__order",
                "requirement__sequence_no",
                "id",
            )
        )
        for row in qs_rows:
            rows_by_session.setdefault(row.assignment.sesion_id, {})
            rows_by_session[row.assignment.sesion_id].setdefault(
                row.requirement_id, row
            )

        present_shots_by_row = {}
        qs_present = (
            CableEvidence.objects.filter(
                assignment_requirement__assignment__sesion_id__in=cable_session_ids
            )
            .exclude(shot_type="")
            .exclude(review_status=CableEvidence.REVIEW_REJECTED)
            .values_list("assignment_requirement_id", "shot_type")
            .distinct()
        )
        for row_id, shot_type in qs_present:
            present_shots_by_row.setdefault(row_id, set()).add(shot_type)

        rejected_comments_qs = (
            CableEvidence.objects.filter(
                assignment_requirement__assignment__sesion_id__in=cable_session_ids,
                review_status=CableEvidence.REVIEW_REJECTED,
            )
            .exclude(review_comment="")
            .select_related(
                "assignment_requirement", "assignment_requirement__assignment"
            )
            .order_by(
                "assignment_requirement__assignment__sesion_id", "-reviewed_at", "-id"
            )
        )

        rejected_comments_map = {}
        for ev in rejected_comments_qs:
            sid = ev.assignment_requirement.assignment.sesion_id
            if sid not in rejected_comments_map:
                rejected_comments_map[sid] = ev.review_comment

        rejected_exists_qs = (
            CableEvidence.objects.filter(
                assignment_requirement__assignment__sesion_id__in=cable_session_ids,
                review_status=CableEvidence.REVIEW_REJECTED,
            )
            .values_list("assignment_requirement__assignment__sesion_id", flat=True)
            .distinct()
        )
        rejected_session_ids = set(rejected_exists_qs)

        required_shots = [
            CableEvidence.SHOT_START_CABLE,
            CableEvidence.SHOT_END_CABLE,
            CableEvidence.SHOT_HANDHOLE,
        ]

        for sid, reqs in requirements_by_session.items():
            faltantes = []
            row_map = rows_by_session.get(sid, {})

            for req in reqs:
                missing = []

                if (
                    req.start_ft is None
                    or req.planned_reserve_ft is None
                    or req.end_ft is None
                ):
                    missing.append("measurement")

                row = row_map.get(req.id)
                if not row:
                    missing.append("photos")
                else:
                    present = present_shots_by_row.get(row.id, set())
                    pending = [shot for shot in required_shots if shot not in present]
                    if pending:
                        missing.append("photos")

                    row_has_rejected = CableEvidence.objects.filter(
                        assignment_requirement=row,
                        review_status=CableEvidence.REVIEW_REJECTED,
                    ).exists()
                    if row_has_rejected:
                        missing.append("review")

                if missing:
                    faltantes.append(f"{_cable_label(req)} ({', '.join(missing)})")

            cable_missing_by_sesion[sid] = faltantes
            cable_has_rejected_by_sesion[sid] = sid in rejected_session_ids
            cable_rejection_comment_by_sesion[sid] = rejected_comments_map.get(sid, "")

    for a in asignaciones:
        sid = a.sesion_id

        if getattr(a.sesion, "is_cable_installation", False):
            a.faltantes_global_labels = cable_missing_by_sesion.get(sid, [])
            a.has_cable_rejected_photo = cable_has_rejected_by_sesion.get(sid, False)
            a.cable_rejection_comment = cable_rejection_comment_by_sesion.get(sid, "")
            a.can_open_cable_report = (
                a.estado
                in ["en_proceso", "rechazado_supervisor", "en_revision_supervisor"]
                or getattr(a, "reintento_habilitado", False)
                or a.has_cable_rejected_photo
            )
        else:
            required = req_titles_by_sesion.get(sid, set())
            covered = covered_by_sesion.get(sid, set())
            faltan_keys = required - covered

            smap = sample_map_by_sesion.get(sid, {})
            a.faltantes_global_labels = [smap.get(k, k) for k in sorted(faltan_keys)]
            a.has_cable_rejected_photo = False
            a.cable_rejection_comment = ""
            a.can_open_cable_report = False

        a.pendientes_aceptar_names = pending_accept_names.get(sid, [])

        if getattr(a.sesion, "is_cable_installation", False):
            a.can_finish = (
                a.estado in ["en_proceso", "rechazado_supervisor"]
                and not a.faltantes_global_labels
                and not a.pendientes_aceptar_names
            )
        else:
            a.can_finish = (
                a.estado == "en_proceso"
                and not a.faltantes_global_labels
                and not a.pendientes_aceptar_names
            )

    def _cell_value(a, col_idx: int) -> str:
        s = a.sesion

        def vac(x):
            x = (x or "").strip()
            return x if x else "(Vacías)"

        if col_idx == 0:
            return vac(s.creado_en.strftime("%Y-%m-%d"))
        if col_idx == 1:
            return vac(getattr(s, "proyecto_id", "") or "")
        if col_idx == 2:
            addr = (getattr(s, "direccion_proyecto", "") or "").strip()
            href = (getattr(s, "maps_href", "") or "").strip()
            if not addr and not href:
                return "(Vacías)"
            if addr and href and addr == href:
                return "Address"
            return vac(addr)
        if col_idx == 3:
            return vac(getattr(s, "cliente", "") or "")
        if col_idx == 4:
            return vac(getattr(s, "ciudad", "") or "")
        if col_idx == 5:
            return vac(getattr(a, "proyecto_label", "") or "")
        if col_idx == 6:
            return vac(getattr(s, "oficina", "") or "")
        if col_idx == 7:
            try:
                val = getattr(a, "my_total", Decimal("0.00")) or Decimal("0.00")
                return f"${val:.2f}"
            except Exception:
                return "$0.00"
        if col_idx == 8:
            estado = getattr(a, "estado", "") or ""
            if estado == "asignado":
                return "Pending acceptance"
            if estado == "en_proceso":
                return "In progress"
            if estado == "en_revision_supervisor":
                return "Submitted — supervisor review"
            if estado == "rechazado_supervisor":
                return "Rejected by supervisor"
            if estado == "rechazado_pm":
                return "Rejected by PM"
            if estado == "rechazado_finanzas":
                return "Rejected by Finance"
            return vac(estado)
        if col_idx == 9:
            comentario = getattr(a, "tecnico_comentario", "") or ""
            if getattr(a, "sesion", None) and getattr(
                a.sesion, "is_cable_installation", False
            ):
                rechazo = getattr(a, "cable_rejection_comment", "") or ""
                if comentario and rechazo:
                    return f"{comentario} | Review: {rechazo}"
                if rechazo:
                    return rechazo
            return vac(comentario)
        if col_idx == 10:
            if getattr(s, "is_cable_installation", False):
                return "Cable report"
            return "Download" if getattr(s, "reporte_fotografico", None) else "(Vacías)"

        return "(Vacías)"

    excel_filters_raw = (request.GET.get("excel_filters") or "").strip()
    active_excel_filters = {}
    if excel_filters_raw:
        try:
            active_excel_filters = json.loads(excel_filters_raw) or {}
        except Exception:
            active_excel_filters = {}

    if isinstance(active_excel_filters, dict) and active_excel_filters:
        filtered = []
        for a in asignaciones:
            ok = True
            for col_str, allowed_list in active_excel_filters.items():
                try:
                    col = int(col_str)
                except Exception:
                    continue
                allowed_set = set((allowed_list or []))
                if not allowed_set:
                    continue
                val = _cell_value(a, col)
                if val not in allowed_set:
                    ok = False
                    break
            if ok:
                filtered.append(a)
        asignaciones = filtered

    excel_global = {}
    MAX_COLS = 11
    for i in range(MAX_COLS):
        vals = set()
        for a in asignaciones:
            vals.add(_cell_value(a, i))
        excel_global[str(i)] = sorted(vals, key=lambda x: (x == "(Vacías)", x.lower()))

    excel_global_json = json.dumps(excel_global, ensure_ascii=False)

    cantidad = (request.GET.get("cantidad") or "20").strip()
    try:
        per_page = int(cantidad)
    except Exception:
        per_page = 20
    if per_page not in (5, 10, 20, 50, 100):
        per_page = 20

    paginator = Paginator(asignaciones, per_page)
    page_num = request.GET.get("page") or "1"
    pagina = paginator.get_page(page_num)

    qs_keep = request.GET.copy()
    qs_keep.pop("page", None)
    base_qs = qs_keep.urlencode()

    return render(
        request,
        "operaciones/billing_mis_asignaciones.html",
        {
            "asignaciones": pagina.object_list,
            "pagina": pagina,
            "cantidad": str(per_page),
            "base_qs": base_qs,
            "excel_global_json": excel_global_json,
        },
    )


@login_required
@rol_requerido("usuario")
def detalle_assignment(request, pk):
    a = get_object_or_404(SesionBillingTecnico, pk=pk, tecnico=request.user)

    if not _is_asig_active(a):
        raise Http404()

    items = (
        ItemBillingTecnico.objects.filter(item__sesion=a.sesion, tecnico=request.user)
        .select_related("item")
        .order_by("item__id")
    )

    s = a.sesion

    pendientes_aceptar = []
    qs_asg = s.tecnicos_sesion.select_related("tecnico").all()
    try:
        SesionBillingTecnico._meta.get_field("is_active")
        qs_asg = qs_asg.filter(is_active=True)
    except Exception:
        pass

    for asg in qs_asg:
        accepted = bool(asg.aceptado_en) or asg.estado != "asignado"
        if not accepted:
            name = (
                getattr(asg.tecnico, "get_full_name", lambda: "")()
                or asg.tecnico.username
            )
            pendientes_aceptar.append(name)

    if getattr(s, "is_cable_installation", False):
        try:
            from cable_installation.models import (CableAssignmentRequirement,
                                                   CableEvidence,
                                                   CableRequirement)
        except Exception:
            CableAssignmentRequirement = None
            CableEvidence = None
            CableRequirement = None

        cable_missing_labels = []
        has_cable_rejected_photo = False
        cable_rejection_comment = ""
        can_open_cable_report = False

        if CableAssignmentRequirement and CableEvidence and CableRequirement:
            rows = list(
                CableAssignmentRequirement.objects.filter(assignment=a)
                .select_related("requirement", "assignment")
                .order_by("requirement__order", "requirement__sequence_no", "id")
            )
            row_by_req = {row.requirement_id: row for row in rows}

            requirements = list(
                s.cable_requirements.filter(required=True).order_by(
                    "order", "sequence_no", "id"
                )
            )

            for req in requirements:
                row = row_by_req.get(req.id)
                missing = []

                if (
                    req.start_ft is None
                    or req.planned_reserve_ft is None
                    or req.end_ft is None
                ):
                    missing.append("measurement")

                if not row:
                    missing.append("photos")
                else:
                    present = set(
                        CableEvidence.objects.filter(assignment_requirement=row)
                        .exclude(shot_type="")
                        .exclude(review_status=CableEvidence.REVIEW_REJECTED)
                        .values_list("shot_type", flat=True)
                        .distinct()
                    )

                    required_shots = [
                        CableEvidence.SHOT_START_CABLE,
                        CableEvidence.SHOT_END_CABLE,
                        CableEvidence.SHOT_HANDHOLE,
                    ]
                    pending = [shot for shot in required_shots if shot not in present]
                    if pending:
                        missing.append("photos")

                    row_has_rejected = CableEvidence.objects.filter(
                        assignment_requirement=row,
                        review_status=CableEvidence.REVIEW_REJECTED,
                    ).exists()
                    if row_has_rejected:
                        missing.append("review")

                if missing:
                    cable_missing_labels.append(
                        f"PK {req.sequence_no} - {req.handhole} ({', '.join(missing)})"
                    )

            rejected_ev = (
                CableEvidence.objects.filter(
                    assignment_requirement__assignment=a,
                    review_status=CableEvidence.REVIEW_REJECTED,
                )
                .exclude(review_comment="")
                .order_by("-reviewed_at", "-id")
                .first()
            )
            if rejected_ev:
                has_cable_rejected_photo = True
                cable_rejection_comment = rejected_ev.review_comment or ""
            else:
                has_cable_rejected_photo = CableEvidence.objects.filter(
                    assignment_requirement__assignment=a,
                    review_status=CableEvidence.REVIEW_REJECTED,
                ).exists()

            can_open_cable_report = (
                a.estado
                in ["en_proceso", "rechazado_supervisor", "en_revision_supervisor"]
                or getattr(a, "reintento_habilitado", False)
                or has_cable_rejected_photo
            )

        can_finish = (
            a.estado in ["en_proceso", "rechazado_supervisor"]
            and not cable_missing_labels
            and not pendientes_aceptar
        )

        return render(
            request,
            "operaciones/billing_detalle_asignacion.html",
            {
                "a": a,
                "items": items,
                "can_finish": can_finish,
                "cable_missing_labels": cable_missing_labels,
                "has_cable_rejected_photo": has_cable_rejected_photo,
                "cable_rejection_comment": cable_rejection_comment,
                "can_open_cable_report": can_open_cable_report,
            },
        )

    def _norm_title(s: str) -> str:
        return (s or "").strip().lower()

    required_titles = RequisitoFotoBilling.objects.filter(
        tecnico_sesion__sesion=s, obligatorio=True
    ).values_list("titulo", flat=True)
    required_key_set = {_norm_title(t) for t in required_titles if t}

    taken_titles = (
        EvidenciaFotoBilling.objects.filter(
            tecnico_sesion__sesion=s, requisito__isnull=False
        )
        .values_list("requisito__titulo", flat=True)
        .distinct()
    )
    covered_key_set = {_norm_title(t) for t in taken_titles if t}

    missing_keys = required_key_set - covered_key_set

    can_finish = (
        a.estado == "en_proceso" and not missing_keys and not pendientes_aceptar
    )

    return render(
        request,
        "operaciones/billing_detalle_asignacion.html",
        {
            "a": a,
            "items": items,
            "can_finish": can_finish,
            "cable_missing_labels": [],
            "has_cable_rejected_photo": False,
            "cable_rejection_comment": "",
            "can_open_cable_report": False,
        },
    )


@login_required
@rol_requerido('usuario')
@require_POST
def start_assignment(request, pk):
    """
    El técnico acepta la tarea y la pone en 'en_proceso'.
    El proyecto pasa a 'en_proceso' si estaba 'rechazado_supervisor' o 'asignado'.
    """
    a = get_object_or_404(SesionBillingTecnico, pk=pk, tecnico=request.user)

    # ✅ NUEVO: bloquear si la asignación está inactiva
    if not _is_asig_active(a):
        messages.error(request, "This assignment is no longer available.")
        return redirect("operaciones:mis_assignments")

    if a.estado not in {"asignado", "rechazado_supervisor"} and not a.reintento_habilitado:
        messages.error(request, "This assignment cannot be started.")
        return redirect("operaciones:mis_assignments")

    a.estado = "en_proceso"
    a.aceptado_en = timezone.now()
    a.reintento_habilitado = False
    a.save(update_fields=["estado", "aceptado_en", "reintento_habilitado"])

    s = a.sesion
    if s.estado in {"rechazado_supervisor", "asignado"}:
        s.estado = "en_proceso"
        s.save(update_fields=["estado"])

    messages.success(request, "Assignment started.")
    return redirect("operaciones:mis_assignments")


def _to_jpeg_if_needed(uploaded_file):
    """
    Si es HEIC/HEIF (o un formato no-JPEG) lo convierte a JPEG (quality 92)
    conservando EXIF cuando exista. Devuelve un ContentFile listo para asignar
    a un ImageField/FileField (con nombre .jpg).
    """
    uploaded_file.seek(0)
    im = Image.open(uploaded_file)
    fmt = (im.format or "").upper()
    exif = im.info.get("exif")

    if fmt in {"HEIC", "HEIF"}:
        bio = BytesIO()
        im = im.convert("RGB")
        if exif:
            im.save(bio, format="JPEG", quality=92, exif=exif)
        else:
            im.save(bio, format="JPEG", quality=92)
        bio.seek(0)
        name = (uploaded_file.name.rsplit(".", 1)[0]) + ".jpg"
        return ContentFile(bio.read(), name=name)

    # Si es otra cosa (PNG, WEBP, etc.) lo dejamos igual
    uploaded_file.seek(0)
    return uploaded_file


def _exif_to_latlng_taken_at(image):
    """
    Extrae (lat, lng, taken_at) de EXIF si existen.
    Retorna (lat, lng, dt) o (None, None, None).
    """
    try:
        exif = getattr(image, "_getexif", lambda: None)()
        if not exif:
            return None, None, None

        tagmap = {ExifTags.TAGS.get(k, k): v for k, v in exif.items()}
        # Fecha/hora
        dt_raw = tagmap.get("DateTimeOriginal") or tagmap.get("DateTime")
        taken_at = None
        if dt_raw:
            from datetime import datetime
            try:
                taken_at = timezone.make_aware(
                    datetime.strptime(dt_raw, "%Y:%m:%d %H:%M:%S"))
            except Exception:
                taken_at = None

        # GPS
        gps_info = tagmap.get("GPSInfo")
        if not gps_info:
            return None, None, taken_at

        def _ratio_to_float(r):
            try:
                return float(r[0]) / float(r[1])
            except Exception:
                return float(r)

        def _dms_to_deg(dms, ref):
            deg = _ratio_to_float(dms[0])
            minutes = _ratio_to_float(dms[1])
            seconds = _ratio_to_float(dms[2])
            value = deg + (minutes / 60.0) + (seconds / 3600.0)
            if ref in ['S', 'W']:
                value = -value
            return value

        gps_tagmap = {ExifTags.GPSTAGS.get(
            k, k): v for k, v in gps_info.items()}
        lat = lng = None
        if all(k in gps_tagmap for k in ["GPSLatitude", "GPSLatitudeRef", "GPSLongitude", "GPSLongitudeRef"]):
            lat = _dms_to_deg(
                gps_tagmap["GPSLatitude"], gps_tagmap["GPSLatitudeRef"])
            lng = _dms_to_deg(
                gps_tagmap["GPSLongitude"], gps_tagmap["GPSLongitudeRef"])

        return lat, lng, taken_at
    except Exception:
        return None, None, None


# --- VISTAS: copiar/pegar reemplazando las actuales ---

@login_required
@rol_requerido('usuario')
def upload_evidencias(request, pk):
    """
    Carga de evidencias con 'lock' por TÍTULO compartido a nivel de sesión.
    """
    a = get_object_or_404(SesionBillingTecnico, pk=pk, tecnico=request.user)

    # ✅ NUEVO: bloquear si la asignación está inactiva
    if not _is_asig_active(a):
        messages.error(request, "This assignment is no longer available.")
        return redirect("operaciones:mis_assignments")

    # ---------- helpers ----------
    def _norm_title(s: str) -> str:
        return (s or "").strip().lower()

    def _is_safe_wasabi_key(key: str) -> bool:
        return bool(key) and ".." not in key and not key.startswith("/")

    def _create_evidencia_from_key(
        req_id, key, nota, lat, lng, acc, taken_dt,
        titulo_manual: str = "", direccion_manual: str = ""
    ):
        return EvidenciaFotoBilling.objects.create(
            tecnico_sesion=a,
            requisito_id=req_id,
            imagen=key,
            nota=nota or "",
            lat=lat or None,
            lng=lng or None,
            gps_accuracy_m=acc or None,
            client_taken_at=taken_dt,
            titulo_manual=titulo_manual or "",
            direccion_manual=direccion_manual or "",
        )

    def _boolish(v):
        if isinstance(v, bool):
            return v
        if v is None:
            return False
        if isinstance(v, int):
            return v != 0
        if isinstance(v, str):
            return v.strip().lower() in {"1", "true", "t", "yes", "y", "on", "si", "sí"}
        return bool(v)

    # Flag ROBUSTO de proyecto especial:
    def _es_proyecto_especial(asig: SesionBillingTecnico) -> bool:
        s = asig.sesion
        candidatos = [
            getattr(s, "proyecto_especial", None),
            getattr(getattr(s, "servicio", None), "proyecto_especial", None),
            getattr(getattr(asig, "servicio", None),
                    "proyecto_especial", None),
            getattr(getattr(s, "proyecto", None), "proyecto_especial", None),
        ]
        for v in candidatos:
            if v is not None:
                return _boolish(v)
        # Heurística: si la sesión no tiene REQUISITOS, tratar como especial
        no_reqs = not RequisitoFotoBilling.objects.filter(
            tecnico_sesion__sesion=s
        ).exists()
        return no_reqs

    # Si el técnico fue agregado tarde y no tiene requisitos, clonar de la sesión
    def _ensure_requisitos_para_asignacion():
        if a.requisitos.exists():
            return
        base_qs = (RequisitoFotoBilling.objects
                   .filter(tecnico_sesion__sesion=a.sesion)
                   .order_by("orden", "id")
                   .select_related("tecnico_sesion"))
        to_create, seen = [], set()
        orden_fallback = 0
        for br in base_qs:
            key = _norm_title(br.titulo)
            if not key or key in seen:
                continue
            orden_fallback += 1
            to_create.append(RequisitoFotoBilling(
                tecnico_sesion=a,
                titulo=br.titulo,
                descripcion=br.descripcion,
                obligatorio=br.obligatorio,
                orden=br.orden or orden_fallback,
            ))
            seen.add(key)
        if to_create:
            RequisitoFotoBilling.objects.bulk_create(to_create)

    _ensure_requisitos_para_asignacion()

    # Permisos para subir según estado
    puede_subir = (a.estado == "en_proceso") or (
        a.estado == "rechazado_supervisor" and a.reintento_habilitado
    )
    if not puede_subir and request.method != "GET":
        messages.info(request, "This assignment is not open for uploads.")
        return redirect("operaciones:mis_assignments")

    s = a.sesion
    is_especial = _es_proyecto_especial(a)

    # -------------------- POST -------------------- (fallback no-AJAX)
    if request.method == "POST":
        req_id = request.POST.get("req_id") or None
        nota = (request.POST.get("nota") or "").strip()

        files = request.FILES.getlist("imagenes[]")
        wasabi_keys = request.POST.getlist(
            "wasabi_keys[]") if settings.DIRECT_UPLOADS_ENABLED else []

        lat = request.POST.get("lat") or None
        lng = request.POST.get("lng") or None
        acc = request.POST.get("acc") or None
        taken = request.POST.get("client_taken_at")
        taken_dt = parse_datetime(taken) if taken else None

        # Campos manuales para Extra en proyecto especial
        titulo_manual = (request.POST.get("titulo_manual") or "").strip()
        direccion_manual = (request.POST.get("direccion_manual") or "").strip()

        if is_especial and not req_id:
            if not titulo_manual:
                messages.error(
                    request, "Please enter a Title for the photo (special project).")
                return redirect("operaciones:upload_evidencias", pk=a.pk)
            if not direccion_manual:
                messages.error(
                    request, "Please enter an Address for the photo (special project).")
                return redirect("operaciones:upload_evidencias", pk=a.pk)

        # Lock por título (si es requisito)
        if req_id:
            req = get_object_or_404(
                RequisitoFotoBilling, pk=req_id, tecnico_sesion=a)
            shared_key = _norm_title(req.titulo)
            taken_titles = (EvidenciaFotoBilling.objects
                            .filter(tecnico_sesion__sesion=s, requisito__isnull=False)
                            .values_list("requisito__titulo", flat=True))
            locked_title_set = {_norm_title(t) for t in taken_titles if t}
            if shared_key in locked_title_set:
                messages.warning(
                    request,
                    "This requirement is already covered by the team. "
                    "Remove the existing photo to re-activate it."
                )
                return redirect("operaciones:upload_evidencias", pk=a.pk)

        # Wasabi keys
        n = 0
        for key in wasabi_keys:
            if _is_safe_wasabi_key(key):
                _create_evidencia_from_key(
                    req_id, key, nota, lat, lng, acc, taken_dt,
                    titulo_manual=titulo_manual, direccion_manual=direccion_manual
                )
                n += 1

        # Archivos
        for f in files:
            f_conv = _to_jpeg_if_needed(f)
            try:
                f_conv.seek(0)
                im = Image.open(f_conv)
                exif_lat, exif_lng, exif_dt = _exif_to_latlng_taken_at(im)
            except Exception:
                exif_lat = exif_lng = exif_dt = None
            finally:
                f_conv.seek(0)

            use_lat = lat or exif_lat
            use_lng = lng or exif_lng
            use_taken = taken_dt or exif_dt

            EvidenciaFotoBilling.objects.create(
                tecnico_sesion=a,
                requisito_id=req_id,
                imagen=f_conv,
                nota=nota,
                lat=use_lat,
                lng=use_lng,
                gps_accuracy_m=acc,
                client_taken_at=use_taken,
                titulo_manual=titulo_manual,
                direccion_manual=direccion_manual,
            )
            n += 1

        messages.success(request, f"{n} photo(s) uploaded.") if n else messages.info(
            request, "No files selected."
        )
        return redirect("operaciones:upload_evidencias", pk=a.pk)

    # -------------------- GET --------------------
    requisitos = (
        a.requisitos
         .annotate(uploaded=Count("evidencias"))
         .order_by("orden", "id")
    )

    taken_titles = (EvidenciaFotoBilling.objects
                    .filter(tecnico_sesion__sesion=s, requisito__isnull=False)
                    .values_list("requisito__titulo", flat=True))
    locked_title_set = {_norm_title(t) for t in taken_titles if t}
    locked_ids = [r.id for r in requisitos if _norm_title(
        r.titulo) in locked_title_set]

    required_titles = (RequisitoFotoBilling.objects
                       .filter(tecnico_sesion__sesion=s, obligatorio=True)
                       .values_list("titulo", flat=True))
    required_key_set = {_norm_title(t) for t in required_titles if t}
    covered_key_set = locked_title_set
    missing_keys = required_key_set - covered_key_set

    sample_titles = list(
        RequisitoFotoBilling.objects
        .filter(tecnico_sesion__sesion=s, titulo__isnull=False)
        .values_list("titulo", flat=True)
    )
    sample_map = {_norm_title(t): t for t in sample_titles if t}
    faltantes_global = [sample_map.get(k, k) for k in sorted(missing_keys)]

    qs_asg = s.tecnicos_sesion.select_related("tecnico").all()
    try:
        SesionBillingTecnico._meta.get_field("is_active")
        qs_asg = qs_asg.filter(is_active=True)
    except Exception:
        pass

    asignaciones = list(qs_asg)
    pendientes_aceptar = []
    for asg in asignaciones:
        accepted = bool(asg.aceptado_en) or asg.estado != "asignado"
        if not accepted:
            name = getattr(asg.tecnico, "get_full_name",
                           lambda: "")() or asg.tecnico.username
            pendientes_aceptar.append(name)
    all_accepted = (len(pendientes_aceptar) == 0)

    can_finish = (a.estado == "en_proceso" and len(
        faltantes_global) == 0 and all_accepted)

    evidencias = (
        a.evidencias
         .select_related("requisito")
         .order_by("requisito__orden", "tomada_en", "id")
    )

    can_delete = puede_subir

    proj_id = (a.sesion.proyecto_id or "project").strip()
    proj_slug = slugify(proj_id) or "project"
    sess_tag = f"{proj_slug}-{a.sesion_id}"

    tech = a.tecnico
    tech_name = (
        getattr(tech, "get_full_name", lambda: "")()
        or getattr(tech, "username", "")
        or f"user-{tech.id}"
    )
    tech_slug = slugify(tech_name) or f"user-{tech.id}"

    direct_uploads_folder = f"operaciones/reporte_fotografico/{sess_tag}/{tech_slug}/evidencia/"

    return render(
        request,
        "operaciones/billing_upload_evidencias.html",
        {
            "a": a,
            "requisitos": requisitos,
            "evidencias": evidencias,
            "can_delete": can_delete,

            "locked_ids": locked_ids,
            "faltantes_global": faltantes_global,
            "pendientes_aceptar": pendientes_aceptar,
            "can_finish": can_finish,

            "direct_uploads_enabled": settings.DIRECT_UPLOADS_ENABLED,
            "direct_uploads_max_mb": getattr(settings, "DIRECT_UPLOADS_MAX_MB", 15),
            "direct_uploads_folder": direct_uploads_folder,
            "project_id": a.sesion.proyecto_id,
            "current_user_name": tech_name,

            # ✅ viene del helper robusto
            "is_proyecto_especial": is_especial,
        },
    )


@login_required
@rol_requerido("usuario")
@require_POST
def upload_evidencias_ajax(request, pk):
    """
    Subida AJAX (una imagen por request) al estilo GZ Services.

    - Si la evidencia corresponde a POWER PORT, intenta extraer automáticamente
      la potencia dBm al momento de guardar la foto.
    - Si es Extra, solo intenta extraer automáticamente si el título/nota sugiere
      potencia, para no gastar IA en fotos normales.
    - Si no logra extraer, NO bloquea la subida.
      La foto queda subida y el botón manual "Extract power" sigue disponible.
    """
    a = get_object_or_404(SesionBillingTecnico, pk=pk, tecnico=request.user)

    if not _is_asig_active(a):
        return JsonResponse(
            {"ok": False, "error": "Assignment no longer available."},
            status=404,
        )

    s = a.sesion

    def _boolish(v):
        if isinstance(v, bool):
            return v
        if v is None:
            return False
        if isinstance(v, int):
            return v != 0
        if isinstance(v, str):
            return v.strip().lower() in {"1", "true", "t", "yes", "y", "on", "si", "sí"}
        return bool(v)

    def _es_proyecto_especial(asig: SesionBillingTecnico) -> bool:
        sess = asig.sesion
        candidatos = [
            getattr(sess, "proyecto_especial", None),
            getattr(getattr(sess, "servicio", None), "proyecto_especial", None),
            getattr(getattr(asig, "servicio", None), "proyecto_especial", None),
            getattr(getattr(sess, "proyecto", None), "proyecto_especial", None),
        ]
        for v in candidatos:
            if v is not None:
                return _boolish(v)

        return not RequisitoFotoBilling.objects.filter(
            tecnico_sesion__sesion=sess
        ).exists()

    is_especial = _es_proyecto_especial(a)

    puede_subir = (a.estado == "en_proceso") or (
        a.estado == "rechazado_supervisor" and a.reintento_habilitado
    )
    if not puede_subir:
        return JsonResponse(
            {"ok": False, "error": "Asignación no abierta para subir fotos."},
            status=400,
        )

    req_id = request.POST.get("req_id") or None
    nota = (request.POST.get("nota") or "").strip()
    lat = request.POST.get("lat") or None
    lng = request.POST.get("lng") or None
    acc = request.POST.get("acc") or None
    taken = request.POST.get("client_taken_at")
    taken_dt = parse_datetime(taken) if taken else None
    titulo_manual = (request.POST.get("titulo_manual") or "").strip()
    direccion_manual = (request.POST.get("direccion_manual") or "").strip()

    if is_especial and not req_id:
        if not titulo_manual:
            return JsonResponse(
                {"ok": False, "error": "Ingresa un Título (proyecto especial)."},
                status=400,
            )
        if not direccion_manual:
            return JsonResponse(
                {"ok": False, "error": "Ingresa una Dirección (proyecto especial)."},
                status=400,
            )

    if not req_id:
        total_extra = EvidenciaFotoBilling.objects.filter(
            tecnico_sesion__sesion=s,
            requisito__isnull=True,
        ).count()
        if total_extra >= 1000:
            return JsonResponse(
                {
                    "ok": False,
                    "error": "Límite alcanzado: máximo 1000 fotos extra por proyecto.",
                },
                status=400,
            )

    file = request.FILES.get("imagen")
    if not file:
        return JsonResponse({"ok": False, "error": "No llegó la imagen."}, status=400)

    f_conv = _to_jpeg_if_needed(file)

    try:
        f_conv.seek(0)
        im = Image.open(f_conv)
        exif_lat, exif_lng, exif_dt = _exif_to_latlng_taken_at(im)
    except Exception:
        exif_lat = exif_lng = exif_dt = None
    finally:
        f_conv.seek(0)

    use_lat = lat or exif_lat
    use_lng = lng or exif_lng
    use_taken = taken_dt or exif_dt

    ev = a.evidencias.create(
        requisito_id=req_id,
        imagen=f_conv,
        nota=nota,
        lat=use_lat,
        lng=use_lng,
        gps_accuracy_m=acc,
        client_taken_at=use_taken,
        titulo_manual=titulo_manual,
        direccion_manual=direccion_manual or "",
    )

    auto_power = {
        "attempted": False,
        "ok": False,
        "power_dbm": "",
        "port_no": None,
        "error": "",
    }

    try:
        should_try_auto_power = False

        if ev.requisito_id:
            req = ev.requisito
            titulo_req = (req.titulo or "").strip().upper()

            should_try_auto_power = bool(
                getattr(req, "needs_power_reading", False)
            ) or titulo_req.startswith("POWER PORT")
        else:
            extra_hint = f"{titulo_manual} {nota}".strip().lower()
            should_try_auto_power = any(
                x in extra_hint
                for x in ["power", "port", "dbm", "opm", "light level", "light"]
            )

        if should_try_auto_power:
            auto_power["attempted"] = True
            result = _extract_power_dbm_for_evidence(ev, user=request.user)
            auto_power["ok"] = True
            auto_power["power_dbm"] = result.get("power_dbm", "")
            auto_power["port_no"] = result.get("port_no")

    except Exception as e:
        auto_power["ok"] = False
        auto_power["error"] = str(e)[:255]

    extras_left = max(
        0,
        1000
        - EvidenciaFotoBilling.objects.filter(
            tecnico_sesion__sesion=s,
            requisito__isnull=True,
        ).count(),
    )

    ev.refresh_from_db()

    titulo = ev.requisito.titulo if ev.requisito_id else (ev.titulo_manual or "Extra")

    fecha_txt = timezone.localtime(ev.client_taken_at or ev.tomada_en).strftime(
        "%Y-%m-%d %H:%M"
    )

    return JsonResponse(
        {
            "ok": True,
            "evidencia": {
                "id": ev.id,
                "url": ev.imagen.url,
                "titulo": titulo,
                "fecha": fecha_txt,
                "lat": ev.lat,
                "lng": ev.lng,
                "acc": ev.gps_accuracy_m,
                "req_id": int(req_id) if req_id else None,
                "power_dbm": f"{ev.power_dbm:.2f}" if ev.power_dbm is not None else "",
            },
            "auto_power": auto_power,
            "extras_left": extras_left,
            "max_extra": 1000,
        }
    )


@rol_requerido('usuario')
@login_required
def fotos_status_json(request, asig_id: int):
    """
    JSON para el polling del front (GZ-style):
    - can_finish
    - faltantes_global (por título)
    - requisitos (estado global/my_count)
    - evidencias_nuevas (id > after)
    - extras_left / max_extra
    """
    a = get_object_or_404(SesionBillingTecnico,
                          pk=asig_id, tecnico=request.user)

    # ✅ NUEVO: bloquear si la asignación está inactiva
    if not _is_asig_active(a):
        return JsonResponse({"ok": False, "error": "Assignment no longer available."}, status=404)

    s = a.sesion

    # ⛳️ MISMO FIX: si el técnico no tiene requisitos aún, clonarlos de la sesión.
    def _norm_title(s: str) -> str:
        return (s or "").strip().lower()

    if not a.requisitos.exists():
        base_qs = (RequisitoFotoBilling.objects
                   .filter(tecnico_sesion__sesion=s)
                   .order_by("orden", "id"))
        to_create, seen = [], set()
        orden_fallback = 0
        for br in base_qs:
            key = _norm_title(br.titulo)
            if not key or key in seen:
                continue
            orden_fallback += 1
            to_create.append(RequisitoFotoBilling(
                tecnico_sesion=a,
                titulo=br.titulo,
                descripcion=br.descripcion,
                obligatorio=br.obligatorio,
                orden=br.orden or orden_fallback,
            ))
            seen.add(key)
        if to_create:
            RequisitoFotoBilling.objects.bulk_create(to_create)

    after = int(request.GET.get("after", "0") or 0)

    # Requisitos de esta asignación (ya garantizados)
    reqs = list(
        a.requisitos
        .order_by("orden")
        .values("id", "titulo", "obligatorio")
    )

    # Conteo propio del técnico
    my_counts = {
        x["requisito_id"]: x["c"]
        for x in (EvidenciaFotoBilling.objects
                  .filter(tecnico_sesion=a, requisito_id__isnull=False)
                  .values("requisito_id")
                  .annotate(c=Count("id")))
    }

    # Títulos ya cubiertos por el EQUIPO en la sesión
    titles_done = {
        _norm_title(t)
        for t in (EvidenciaFotoBilling.objects
                  .filter(tecnico_sesion__sesion=s, requisito__isnull=False)
                  .values_list("requisito__titulo", flat=True)
                  .distinct())
        if t
    }

    requisitos_json, faltantes = [], []
    for r in reqs:
        titulo = r["titulo"] or ""
        global_done = (_norm_title(titulo) in titles_done)
        my_count = my_counts.get(r["id"], 0)
        if r["obligatorio"] and not global_done:
            faltantes.append(titulo)
        requisitos_json.append({
            "id": r["id"],
            "titulo": titulo,
            "obligatorio": r["obligatorio"],
            "team_count": 1 if global_done else 0,
            "my_count": my_count,
            "global_done": global_done,
        })

    # Evidencias nuevas desde 'after'
    nuevas_qs = (EvidenciaFotoBilling.objects
                 .filter(tecnico_sesion__sesion=s, id__gt=after)
                 .order_by("id"))
    evidencias_nuevas = [{
        "id": ev.id,
        "url": ev.imagen.url,
        "req_id": ev.requisito_id,
        "titulo": (ev.requisito.titulo if ev.requisito_id else (ev.titulo_manual or "Extra")),
        "fecha": timezone.localtime(ev.client_taken_at or ev.tomada_en).strftime("%Y-%m-%d %H:%M"),
        "lat": ev.lat, "lng": ev.lng, "acc": ev.gps_accuracy_m,
    } for ev in nuevas_qs]

    # Cupo global de extras (1000 por sesión)
    total_extra = EvidenciaFotoBilling.objects.filter(
        tecnico_sesion__sesion=s, requisito__isnull=True
    ).count()
    extras_left = max(0, 1000 - total_extra)

    # ¿Faltan aceptaciones? ✅ solo asignaciones activas
    pendientes_aceptar = []

    qs_asg = s.tecnicos_sesion.select_related("tecnico")
    try:
        SesionBillingTecnico._meta.get_field("is_active")
        qs_asg = qs_asg.filter(is_active=True)
    except Exception:
        pass

    for asg in qs_asg:
        accepted = bool(asg.aceptado_en) or asg.estado != "asignado"
        if not accepted:
            nombre = getattr(asg.tecnico, "get_full_name",
                             lambda: "")() or asg.tecnico.username
            pendientes_aceptar.append(nombre)

    # Finish (mismo criterio que la página)
    can_finish = (
        a.estado == "en_proceso" and not faltantes and not pendientes_aceptar)

    return JsonResponse({
        "ok": True,
        "can_finish": can_finish,
        "faltantes_global": faltantes,
        "requisitos": requisitos_json,
        "evidencias_nuevas": evidencias_nuevas,
        "extras_left": extras_left,
        "max_extra": 1000,
    })


@login_required
@rol_requerido('usuario')
def finish_assignment(request, pk):
    """
    Finalización en equipo:
    - Requiere que esta asignación esté en 'en_proceso'.
    - Calcula los requisitos obligatorios vigentes como la INTERSECCIÓN de títulos
      (normalizados) entre TODAS las asignaciones de la sesión.
    - Verifica que esos títulos tengan al menos una foto (de cualquiera del equipo).
    - Verifica que todos hayan aceptado (Start).
    - Si todo OK: pasa TODAS las asignaciones y la sesión a 'en_revision_supervisor'.
    - NUEVO: exige comentario y lo guarda en la asignación del técnico que presiona Finish.
    """
    a = get_object_or_404(SesionBillingTecnico, pk=pk, tecnico=request.user)

    # ✅ NUEVO: bloquear si la asignación está inactiva
    if not _is_asig_active(a):
        messages.error(request, "This assignment is no longer available.")
        return redirect("operaciones:mis_assignments")

    if a.estado != "en_proceso":
        messages.error(request, "This assignment is not in progress.")
        return redirect("operaciones:mis_assignments")

    # ✅ NUEVO (comentario obligatorio desde el modal)
    if request.method != "POST":
        messages.error(request, "Comment is required to finish.")
        return redirect("operaciones:mis_assignments")

    comentario = (request.POST.get("comentario") or "").strip()
    if not comentario:
        messages.error(request, "Please enter a comment to finish.")
        return redirect("operaciones:mis_assignments")

    def _norm_title(s: str) -> str:
        return (s or "").strip().lower()

    s = a.sesion

    # --- Recolectar títulos obligatorios por asignación (normalizados) ✅ solo activas
    qs_asg = (
        s.tecnicos_sesion
        .select_related("tecnico")
        .prefetch_related("requisitos")
        .all()
    )
    try:
        SesionBillingTecnico._meta.get_field("is_active")
        qs_asg = qs_asg.filter(is_active=True)
    except Exception:
        pass

    asignaciones = list(qs_asg)

    per_asg_required_sets = []
    sample_titles = set()  # para nombres bonitos
    for asg in asignaciones:
        # Solo títulos OBLIGATORIOS de esta asignación
        titles = [
            r.titulo for r in asg.requisitos.all()
            if getattr(r, "obligatorio", True)
        ]
        sample_titles.update([t for t in titles if t])
        keyset = {_norm_title(t) for t in titles if t}
        per_asg_required_sets.append(keyset)

    # Si no hay requisitos cargados en ninguna asignación, no se bloquea por fotos
    if not per_asg_required_sets or all(len(sset) == 0 for sset in per_asg_required_sets):
        required_key_set = set()
    else:
        # INTERSECCIÓN entre todas las asignaciones: lo común es lo realmente "vigente"
        required_key_set = set.intersection(*per_asg_required_sets) if len(per_asg_required_sets) > 1 else per_asg_required_sets[0]

    # Map para mostrar nombres con mayúsculas originales
    sample_map = {_norm_title(t): t for t in sample_titles if t}

    # --- Títulos ya cubiertos (algún miembro subió foto para ese requisito)
    taken_titles = (
        EvidenciaFotoBilling.objects
        .filter(tecnico_sesion__sesion=s, requisito__isnull=False)
        .values_list("requisito__titulo", flat=True)
    )
    covered_key_set = {_norm_title(t) for t in taken_titles if t}

    # Lo faltante es la intersección menos lo cubierto
    missing_keys = required_key_set - covered_key_set
    if missing_keys:
        pretty_missing = [sample_map.get(k, k) for k in sorted(missing_keys)]
        messages.error(request, "Missing required photos: " + ", ".join(pretty_missing))
        return redirect("operaciones:upload_evidencias", pk=a.pk)

    # --- Validar que todos hayan dado Start
    pendientes_aceptar = []
    for asg in asignaciones:
        accepted = bool(asg.aceptado_en) or asg.estado != "asignado"
        if not accepted:
            name = getattr(asg.tecnico, "get_full_name", lambda: "")() or asg.tecnico.username
            pendientes_aceptar.append(name)

    if pendientes_aceptar:
        messages.error(request, "Pending acceptance (Start): " + ", ".join(pendientes_aceptar))
        return redirect("operaciones:upload_evidencias", pk=a.pk)

    # --- Transición a revisión de supervisor + guardar comentario
    now = timezone.now()
    with transaction.atomic():
        # ✅ NUEVO: guardar comentario en la asignación que está finalizando
        a.tecnico_comentario = comentario
        a.save(update_fields=["tecnico_comentario"])

        s.tecnicos_sesion.update(
            estado="en_revision_supervisor",
            finalizado_en=now,
        )
        s.estado = "en_revision_supervisor"
        s.save(update_fields=["estado"])

    messages.success(request, "Submitted for supervisor review for all assignees.")
    return redirect("operaciones:mis_assignments")

@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def revisar_assignment(request, pk):
    """
    Compat: antes se revisaba por asignación.
    Ahora redirigimos a la revisión unificada por PROYECTO.
    """
    a = get_object_or_404(SesionBillingTecnico, pk=pk)
    return redirect("operaciones:revisar_sesion", sesion_id=a.sesion_id)


ALLOWED_MIME = {"image/jpeg", "image/png", "image/webp"}


def _safe_prefix() -> str:
    return getattr(settings, "DIRECT_UPLOADS_SAFE_PREFIX", "operaciones/reporte_fotografico/")


def _build_key(folder: str, filename: str) -> str:
    """
    Genera una key segura bajo el prefijo permitido, manteniendo tu estructura.
    - folder debe comenzar con DIRECT_UPLOADS_SAFE_PREFIX (p.ej. operaciones/reporte_fotografico/<proj>/<tech>/evidencia/)
    - filename solo aporta la extensión; el nombre es uuid para evitar colisiones.
    """
    ext = (filename.rsplit(".", 1)[-1] or "jpg").lower()
    base = (folder or "").strip().lstrip("/")
    if not base.startswith(_safe_prefix()):
        # Fuerza a prefijo seguro si el cliente envía algo fuera de rango
        base = _safe_prefix().rstrip("/") + "/evidencia/"
    return f"{base.rstrip('/')}/{uuid.uuid4().hex}.{ext}"


@login_required
@require_POST
def presign_wasabi(request):
    if not getattr(settings, "DIRECT_UPLOADS_ENABLED", False):
        return HttpResponseBadRequest("Direct uploads disabled.")

    try:
        data = json.loads(request.body.decode("utf-8"))
    except Exception:
        return HttpResponseBadRequest("Invalid JSON.")

    filename = (data.get("filename") or "").strip()
    content_type = (data.get("contentType") or "").strip()
    size_bytes = int(data.get("sizeBytes") or 0)
    folder = (data.get("folder") or "").strip()

    if not filename or content_type not in ALLOWED_MIME:
        return HttpResponseBadRequest("Invalid file type.")

    max_bytes = int(getattr(settings, "DIRECT_UPLOADS_MAX_MB", 15)) * 1024 * 1024
    if size_bytes <= 0 or size_bytes > max_bytes:
        return HttpResponseBadRequest("File too large.")

    key = _build_key(folder, filename)

    s3 = boto3.client(
        "s3",
        endpoint_url=getattr(settings, "WASABI_ENDPOINT_URL", "https://s3.us-east-1.wasabisys.com"),
        region_name=getattr(settings, "WASABI_REGION_NAME", "us-east-1"),
        aws_access_key_id=getattr(settings, "WASABI_ACCESS_KEY_ID"),
        aws_secret_access_key=getattr(settings, "WASABI_SECRET_ACCESS_KEY"),
        config=Config(signature_version="s3v4", s3={"addressing_style": "path"}),
        verify=getattr(settings, "AWS_S3_VERIFY", True),
    )

    bucket = getattr(settings, "WASABI_BUCKET_NAME")

    # ✅ IMPORTANTE:
    # No enviamos x-amz-meta-* en el POST presignado porque Wasabi es sensible
    # a los "eq" estrictos en policy (PolicyConditionFailed).
    # Los metadatos reales (lat/lng/taken_at/address/title) los guardas en tu DB via confirmKey().
    fields = {
        "acl": "private",
        "Content-Type": content_type,
        "success_action_status": "201",
    }

    conditions = [
        {"bucket": bucket},
        # Permitimos el prefijo del folder (no forzamos el key exacto)
        ["starts-with", "$key", key.rsplit("/", 1)[0] + "/"],
        {"acl": "private"},
        {"Content-Type": content_type},
        {"success_action_status": "201"},
        ["content-length-range", 1, max_bytes],
    ]

    presigned = s3.generate_presigned_post(
        Bucket=bucket,
        Key=key,
        Fields=fields,
        Conditions=conditions,
        ExpiresIn=300,
    )

    # Fuerza URL path-style
    presigned["url"] = f"{settings.WASABI_ENDPOINT_URL.rstrip('/')}/{bucket}"

    return JsonResponse({"url": presigned["url"], "fields": presigned["fields"], "key": key})


SAFE_EVIDENCE_PREFIX = getattr(
    settings, "DIRECT_UPLOADS_SAFE_PREFIX", "operaciones/reporte_fotografico/")


def _is_safe_wasabi_key(key: str) -> bool:
    """Acepta solo claves dentro del prefijo seguro y sin '..'."""
    return isinstance(key, str) and key.startswith(SAFE_EVIDENCE_PREFIX) and ".." not in key


def _create_evidencia_from_key(a, req_id, key, nota, lat, lng, acc, taken_dt,
                               titulo_manual="", direccion_manual=""):
    """
    Create EvidenciaFotoBilling pointing to an object ALREADY uploaded to Wasabi.
    Doesn't re-upload bytes: assigns .name to the FileField and saves.
    """
    ev = EvidenciaFotoBilling(
        tecnico_sesion=a,
        requisito_id=req_id or None,
        nota=nota or "",
        lat=lat, lng=lng, gps_accuracy_m=acc,
        client_taken_at=taken_dt or None,
        titulo_manual=titulo_manual or "",
        direccion_manual=direccion_manual or "",
    )
    ev.imagen.name = key.strip()
    ev.save()
    return ev


# ============================
# SUPERVISOR — Revisión POR PROYECTO (unificada)
# ============================


def _project_report_key(sesion: SesionBilling) -> str:
    """
    Ruta determinística para el reporte por PROYECTO **por sesión**.
    Ej: operaciones/reporte_fotografico/<proj>-<sesion_id>/project/<proj>-<sesion_id>.xlsx
    """
    proj_slug = slugify(
        sesion.proyecto_id or f"billing-{sesion.id}") or f"billing-{sesion.id}"
    sess_tag = f"{proj_slug}-{sesion.id}"
    return f"operaciones/reporte_fotografico/{sess_tag}/project/{sess_tag}.xlsx"


# ...tus otros imports (decoradores, modelos usados en el template, etc.)


# ---------- revisar_sesion ----------
@login_required
@rol_requerido("supervisor", "admin", "pm")
def revisar_sesion(request, sesion_id):
    """
    Revisión por PROYECTO.
    - APPROVE: encola job para generar el XLSX final.
    - REJECT: marca rechazado (sin tocar Wasabi).
    """
    s = get_object_or_404(SesionBilling, pk=sesion_id)

    asignaciones = (
        s.tecnicos_sesion.select_related("tecnico")
        .prefetch_related("evidencias__requisito")
        .all()
    )

    # Mantén sincronizado el estado a partir de las asignaciones
    s.recomputar_estado_desde_asignaciones()

    can_review = s.estado in {"en_revision_supervisor"}

    if request.method == "POST":
        accion = (request.POST.get("accion") or "").strip().lower()
        comentario = (request.POST.get("comentario") or "").strip()

        if not can_review and accion in {"aprobar", "approve", "rechazar", "reject"}:
            messages.error(request, "This project is not ready for supervisor review.")
            return redirect("operaciones:revisar_sesion", sesion_id=s.id)

        if accion in {"aprobar", "approve"}:
            from usuarios.schedulers import enqueue_reporte_fotografico

            last_job = (
                ReporteFotograficoJob.objects.filter(sesion=s)
                .exclude(log__icontains="[partial]")  # solo FINAL
                .order_by("-creado_en")
                .first()
            )
            if last_job and last_job.estado in ("pendiente", "procesando"):
                messages.info(
                    request,
                    "Photographic report is already being generated in background. It will be attached automatically when it’s ready.",
                )
                return redirect("operaciones:revisar_sesion", sesion_id=s.id)

            job = ReporteFotograficoJob.objects.create(sesion=s)
            enqueue_reporte_fotografico(job.id)

            messages.info(
                request,
                "Generating photographic report in background. It will be attached automatically when it’s ready.",
            )
            return redirect("operaciones:revisar_sesion", sesion_id=s.id)

        elif accion in {"rechazar", "reject"}:
            now = timezone.now()
            with transaction.atomic():
                s.estado = "rechazado_supervisor"
                s.save(update_fields=["estado"])
                for a in asignaciones:
                    a.estado = "rechazado_supervisor"
                    a.supervisor_comentario = comentario or "Rejected."
                    a.supervisor_revisado_en = now
                    a.reintento_habilitado = True
                    a.save(
                        update_fields=[
                            "estado",
                            "supervisor_comentario",
                            "supervisor_revisado_en",
                            "reintento_habilitado",
                        ]
                    )

            messages.warning(
                request, "Project rejected. Reupload enabled for technicians."
            )
            return redirect("operaciones:revisar_sesion", sesion_id=s.id)

        messages.error(request, "Unknown action.")
        return redirect("operaciones:revisar_sesion", sesion_id=s.id)

    # GET: datos para template
    evidencias_por_asig = []
    for a in asignaciones:
        evs_qs = a.evidencias.select_related("requisito").order_by(
            "requisito__orden", "tomada_en", "id"
        )

        evs = list(evs_qs)

        for ev in evs:
            is_power_candidate = False

            if ev.requisito_id:
                titulo_req = (ev.requisito.titulo or "").strip()
                needs_power, _port_no = _power_meta_from_title(titulo_req)

                is_power_candidate = (
                    bool(getattr(ev.requisito, "needs_power_reading", False))
                    or needs_power
                    or ev.power_dbm is not None
                    or "port=" in ((ev.power_extract_note or "").lower())
                )
            else:
                # Extras antiguas:
                # Permitimos botón para que IA detecte potencia + puerto.
                # Si ya tiene potencia o título POWER PORT, seguro es candidata.
                titulo_manual = (ev.titulo_manual or "").strip()
                needs_power, _port_no = _power_meta_from_title(titulo_manual)

                is_power_candidate = (
                    needs_power
                    or ev.power_dbm is not None
                    or "port=" in ((ev.power_extract_note or "").lower())
                    or titulo_manual.lower() in {"extra", ""}
                )

            ev.is_power_candidate = is_power_candidate
            ev.power_port_no_display = _power_port_no_from_evidence(ev)

        evidencias_por_asig.append((a, evs))

    # Archivo final existente (en storage)
    project_report_exists = bool(
        s.reporte_fotografico and storage_file_exists(s.reporte_fotografico)
    )

    # Job FINAL en curso
    last_job = (
        ReporteFotograficoJob.objects.filter(sesion=s)
        .exclude(log__icontains="[partial]")  # solo FINAL
        .order_by("-creado_en")
        .first()
    )
    job_running = bool(last_job and last_job.estado in ("pendiente", "procesando"))

    # Solo consideramos "ready" si HOY el servidor dice que está aprobado
    server_approved = s.estado in {"aprobado_supervisor", "aprobado_pm"}
    project_report_effective_ready = (
        server_approved and project_report_exists and not job_running
    )

    status_url = reverse(
        "operaciones:project_report_status", kwargs={"sesion_id": s.id}
    )

    # ========= Resolver etiqueta legible del proyecto (para el header) =========
    proyectos_qs = filter_queryset_by_access(
        Proyecto.objects.all(),
        request.user,
        "id",
    )

    proyecto_sel = None
    raw = (s.proyecto or "").strip()

    if raw:
        try:
            # si s.proyecto es el PK (nuevo flujo)
            pid = int(raw)
        except (TypeError, ValueError):
            # datos viejos: nombre/código en texto
            proyecto_sel = proyectos_qs.filter(
                Q(nombre__iexact=raw) | Q(codigo__iexact=raw)
            ).first()
        else:
            proyecto_sel = proyectos_qs.filter(pk=pid).first()

    # si no encontramos nada con s.proyecto, probamos con s.proyecto_id (NB3231, etc.)
    if not proyecto_sel and s.proyecto_id:
        code = str(s.proyecto_id).strip()
        proyecto_sel = proyectos_qs.filter(
            Q(codigo__iexact=code) | Q(nombre__icontains=code)
        ).first()

    if proyecto_sel:
        proyecto_label = getattr(proyecto_sel, "nombre", str(proyecto_sel))
    else:
        # fallback para sesiones antiguas / casos raros
        proyecto_label = (s.proyecto or s.proyecto_id or "").strip()

    # =============================== RENDER =============================== #
    return render(
        request,
        "operaciones/billing_revisar_sesion.html",
        {
            "s": s,
            "evidencias_por_asig": evidencias_por_asig,
            "can_review": can_review,
            "project_report_exists": project_report_effective_ready,
            "job_running": job_running,
            "project_report_url": (
                s.reporte_fotografico.url if project_report_effective_ready else ""
            ),
            "status_url": status_url,
            "poll_ms": 1000,
            # 👈 para que el JS no pinte aprobado si no lo está
            "server_approved": server_approved,
            # 👈 NUEVO: nombre legible del proyecto
            "proyecto_label": proyecto_label,
        },
    )


# operaciones/views_billing.py


@login_required
@rol_requerido('supervisor', 'pm', 'admin')
@require_POST
def cancelar_reporte_proyecto(request, sesion_id: int):
    s = get_object_or_404(SesionBilling, pk=sesion_id)
    job = (
        ReporteFotograficoJob.objects
        .filter(sesion=s, estado__in=("pendiente", "procesando"))
        .order_by("-creado_en")
        .first()
    )
    if not job:
        return JsonResponse({"ok": False, "message": "No running job."}, status=404)

    job.cancel_requested = True
    job.save(update_fields=["cancel_requested"])
    return JsonResponse({"ok": True})


@login_required
@require_GET
@never_cache
def project_report_status(request, sesion_id: int):
    s = get_object_or_404(SesionBilling, pk=sesion_id)
    approved = s.estado in ("aprobado_supervisor", "aprobado_pm")

    job = (
        ReporteFotograficoJob.objects
        .filter(sesion_id=sesion_id)
        .exclude(log__icontains="[partial]")   # solo FINAL
        .order_by("-creado_en")
        .first()
    )

    # Antes de aprobación
    if not approved:
        if job and job.estado in ("pendiente", "procesando"):
            state_map = {"pendiente": "pending", "procesando": "processing"}
            return JsonResponse({
                "state": state_map[job.estado],
                "processed": job.procesadas or 0,
                "total": job.total or 0,
                "error": job.error or "",
                "cancel_requested": bool(getattr(job, "cancel_requested", False)),
                "approved": False,
            })
        return JsonResponse({"state": "none", "approved": False})

    # Ya aprobado
    if not job:
        return JsonResponse({"state": "none", "approved": True})

    state_map = {"pendiente": "pending",
                 "procesando": "processing", "ok": "ok", "error": "error"}
    return JsonResponse({
        "state": state_map.get(job.estado, job.estado),
        "processed": job.procesadas or 0,
        "total": job.total or 0,
        "error": job.error or "",
        "cancel_requested": bool(getattr(job, "cancel_requested", False)),
        "approved": True,
    })
# ============================
# REPORTE FOTOGRÁFICO — PROYECTO
# ============================


# --- helper: construir XLSX a DISCO desde un queryset de evidencias ---
def _xlsx_path_from_evqs(sesion: SesionBilling, ev_qs, progress_cb=None, should_cancel=None):
    """
    Construye XLSX en disco (streaming) con progreso y cancelación opcional.
    """
    from tempfile import NamedTemporaryFile

    import xlsxwriter

    tmp_xlsx = NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_xlsx.close()
    wb = xlsxwriter.Workbook(tmp_xlsx.name, {"in_memory": False})
    ws = wb.add_worksheet("PHOTOGRAPHIC REPORT")
    ws.hide_gridlines(2)

    fmt_title = wb.add_format({"bold": True, "align": "center",
                              "valign": "vcenter", "border": 1, "bg_color": "#E8EEF7"})
    fmt_head = wb.add_format({"border": 1, "align": "center", "valign": "vcenter",
                             "bold": True, "text_wrap": True, "bg_color": "#F5F7FB", "font_size": 11})
    fmt_box = wb.add_format({"border": 1})
    fmt_info = wb.add_format({"border": 1, "align": "center",
                             "valign": "vcenter", "text_wrap": True, "font_size": 9})

    # layout
    BLOCK_COLS, SEP_COLS = 6, 1
    LEFT_COL = 0
    RIGHT_COL = LEFT_COL + BLOCK_COLS + SEP_COLS

    HEAD_ROWS, ROWS_IMG, ROW_INFO, ROW_SPACE = 1, 12, 1, 1
    BLOCK_ROWS = HEAD_ROWS + ROWS_IMG + ROW_INFO

    # px helpers
    COL_W = 13
    IMG_ROW_H = 18
    def col_px(w): return int(w * 7 + 5)
    def row_px(h): return int(h * 4 / 3)
    max_w_px = BLOCK_COLS * col_px(COL_W)
    max_h_px = ROWS_IMG * row_px(IMG_ROW_H)

    # cols
    for c in range(LEFT_COL, LEFT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)
    ws.set_column(LEFT_COL + BLOCK_COLS, LEFT_COL + BLOCK_COLS, 2)
    for c in range(RIGHT_COL, RIGHT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)

    ws.merge_range(0, 0, 0, RIGHT_COL + BLOCK_COLS - 1,
                   f"ID PROJECT: {sesion.proyecto_id}", fmt_title)
    cur_row = 2

    def draw_block(r, c, ev):
        # header
        if sesion.proyecto_especial and ev.requisito_id is None:
            titulo_req = (ev.titulo_manual or "").strip() or "Title (missing)"
        else:
            titulo_req = ((getattr(ev.requisito, "titulo", "")
                          or "").strip() or "Extra")
        ws.merge_range(r, c, r + HEAD_ROWS - 1, c +
                       BLOCK_COLS - 1, titulo_req, fmt_head)
        for rr in range(r, r + HEAD_ROWS):
            ws.set_row(rr, 20)

        # image frame
        img_top = r + HEAD_ROWS
        for rr in range(img_top, img_top + ROWS_IMG):
            ws.set_row(rr, IMG_ROW_H)
        ws.merge_range(img_top, c, img_top + ROWS_IMG -
                       1, c + BLOCK_COLS - 1, "", fmt_box)

        # image
        try:
            tmp_img_path, w, h = tmp_jpeg_from_filefield(
                ev.imagen, max_side_px=1600, quality=75)
            sx = max_w_px / float(w)
            sy = max_h_px / float(h)
            scale = min(sx, sy, 1.0)
            scaled_w = int(w * scale)
            scaled_h = int(h * scale)
            x_off = max((max_w_px - scaled_w)//2, 0)
            y_off = max((max_h_px - scaled_h)//2, 0)
            ws.insert_image(img_top, c, tmp_img_path, {
                "x_scale": scale, "y_scale": scale,
                "x_offset": x_off, "y_offset": y_off,
                "object_position": 1,
            })
        except Exception:
            pass

        # info row
        info_row = img_top + ROWS_IMG
        dt = ev.client_taken_at or ev.tomada_en
        taken_txt = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
        lat_txt = f"{float(ev.lat):.6f}" if ev.lat is not None else ""
        lng_txt = f"{float(ev.lng):.6f}" if ev.lng is not None else ""
        addr_txt = (ev.direccion_manual or "").strip()

        if sesion.proyecto_especial and ev.requisito_id is None:
            ws.merge_range(info_row, c,     info_row, c + 2,
                           f"Taken at\n{taken_txt}", fmt_info)
            ws.merge_range(info_row, c + 3, info_row, c + 5,
                           f"Address\n{addr_txt}",   fmt_info)
        else:
            ws.merge_range(info_row, c,     info_row, c + 1,
                           f"Taken at\n{taken_txt}", fmt_info)
            ws.merge_range(info_row, c + 2, info_row, c + 3,
                           f"Lat\n{lat_txt}",       fmt_info)
            ws.merge_range(info_row, c + 4, info_row, c + 5,
                           f"Lng\n{lng_txt}",       fmt_info)
        ws.set_row(info_row, 30)

    # iteración + progreso + cancelación
    idx = 0
    for ev in ev_qs.iterator(chunk_size=100):
        # cancel?
        if callable(should_cancel) and should_cancel(idx):
            raise ReportCancelled()

        if idx % 2 == 0:
            draw_block(cur_row, LEFT_COL, ev)
        else:
            draw_block(cur_row, RIGHT_COL, ev)
            cur_row += BLOCK_ROWS + ROW_SPACE
        idx += 1

        if callable(progress_cb):
            try:
                progress_cb(idx)
            except ReportCancelled:
                raise
            except Exception:
                pass

    if idx % 2 == 1:
        cur_row += BLOCK_ROWS + ROW_SPACE

    wb.close()

    if idx == 0 and callable(progress_cb):
        try:
            progress_cb(0)
        except Exception:
            pass

    return tmp_xlsx.name


def _xlsx_path_reporte_fotografico_qs(sesion: SesionBilling, ev_qs=None, progress_cb=None, should_cancel=None) -> str:
    if ev_qs is None:
        ev_qs = (
            EvidenciaFotoBilling.objects
            .filter(tecnico_sesion__sesion=sesion)
            .select_related("requisito")
            .order_by("requisito__orden", "tomada_en", "id")
        )
    return _xlsx_path_from_evqs(sesion, ev_qs, progress_cb=progress_cb, should_cancel=should_cancel)


@login_required
@require_POST
@rol_requerido('supervisor', 'admin', 'pm')
def generar_reporte_parcial_proyecto(request, sesion_id):
    """
    Encola un NUEVO job PARCIAL y marca cancelación de parciales previos.
    Responde de inmediato (jamás bloquea el request).
    """
    from usuarios.schedulers import enqueue_reporte_parcial

    s = get_object_or_404(SesionBilling, pk=sesion_id)

    # Cancela cualquier parcial previo en curso
    (ReporteFotograficoJob.objects
        .filter(sesion=s, estado__in=("pendiente", "procesando"), log__icontains="[partial]")
        .update(cancel_requested=True))

    # Crea nuevo job parcial
    job = ReporteFotograficoJob.objects.create(
        sesion=s, log="[partial] queued\n", total=0, procesadas=0
    )

    # Arranca SOLO cuando el insert haya sido confirmado
    def _start():
        enqueue_reporte_parcial(job.id)

    transaction.on_commit(_start)

    messages.info(
        request,
        "Generating partial photographic report in background. It will be available to download when it’s ready."
    )
    return redirect("operaciones:revisar_sesion", sesion_id=s.id)


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
@never_cache
def estado_reporte_parcial(request, sesion_id):
    """
    Estado del ÚLTIMO job PARCIAL (los que tienen log con '[partial]').
    """
    s = get_object_or_404(SesionBilling, pk=sesion_id)
    job = (ReporteFotograficoJob.objects
           .filter(sesion=s, log__icontains="[partial]")
           .order_by("-creado_en").first())

    if not job:
        return JsonResponse({"state": "none"})

    state_map = {"pendiente": "pending",
                 "procesando": "processing", "ok": "ok", "error": "error"}
    log_tail = (job.log or "").splitlines()[-5:]

    return JsonResponse({
        "state": state_map.get(job.estado, job.estado),
        "processed": job.procesadas or 0,
        "total": job.total or 0,
        "log_tail": log_tail,
        "error": job.error or "",
        "cancel_requested": bool(getattr(job, "cancel_requested", False)),
    })


class ReportCancelled(Exception):
    pass


def _cache_key_for_ff(ff) -> str:
    """
    Intenta generar un key de cache estable por archivo + last_modified.
    Si el storage no soporta get_modified_time, usamos el nombre.
    """
    base = getattr(ff, "name", str(ff))
    try:
        mtime = storage.get_modified_time(ff.name)
        base = f"{base}:{int(mtime.timestamp())}"
    except Exception:
        pass
    return hashlib.sha1(base.encode("utf-8")).hexdigest()


def tmp_jpeg_from_filefield(ff, max_side_px=1600, quality=75):
    """
    Descarga/convierte a JPEG optimizado y devuelve (path, width, height).
    - Usa thumbnail() que es muy rápida y conserva proporción.
    - Progressive + optimize para tamaño/velocidad.
    - Cache local en /tmp/reporte_cache para no reconvertir en regeneraciones.
    """
    cache_dir = os.path.join(tempfile.gettempdir(), "reporte_cache")
    os.makedirs(cache_dir, exist_ok=True)
    key = _cache_key_for_ff(ff)
    cached_path = os.path.join(cache_dir, f"{key}.jpg")
    if os.path.exists(cached_path):
        with Image.open(cached_path) as im:
            w, h = im.size
        return cached_path, w, h

    # leer datos del storage
    ff.open("rb")
    raw = ff.read()
    ff.close()

    im = Image.open(io.BytesIO(raw))
    im = im.convert("RGB")
    im.draft("RGB", (max_side_px, max_side_px))  # acelera decode de JPEG
    im.thumbnail((max_side_px, max_side_px), Image.LANCZOS)

    tmp_path = cached_path  # guardamos directo en cache
    im.save(tmp_path, "JPEG", quality=quality, optimize=True,
            progressive=True, subsampling="4:2:0")

    w, h = im.size
    return tmp_path, w, h


@login_required
@require_POST
@rol_requerido('supervisor', 'admin', 'pm')
def generar_reporte_parcial_asignacion(request, asig_id):
    """Compat: generate partial report by assignment -> redirect to project version."""
    a = get_object_or_404(SesionBillingTecnico, pk=asig_id)
    return redirect('operaciones:generar_reporte_parcial_proyecto', sesion_id=a.sesion_id)


def _open_file_with_retries(ff, attempts=3, delay=1.0):
    """
    Intenta abrir el FieldFile del storage con pequeños reintentos.
    Devuelve un file-like abierto o levanta la última excepción.
    """
    last = None
    for _ in range(attempts):
        try:
            return ff.open("rb")
        except Exception as e:
            last = e
            time.sleep(delay)
    raise last


@login_required
def descargar_reporte_fotos_proyecto(request, sesion_id):
    s = get_object_or_404(SesionBilling, pk=sesion_id)
    allowed = (getattr(request.user, "rol", "") in ("supervisor", "pm", "admin")) \
        or s.tecnicos_sesion.filter(tecnico=request.user).exists()
    if not allowed:
        raise Http404()

    if not s.reporte_fotografico or not storage_file_exists(s.reporte_fotografico):
        messages.warning(
            request, "The photo report is not available. You can regenerate it now.")
        return redirect("operaciones:regenerar_reporte_fotografico_proyecto", sesion_id=s.id)

    # 1) intentamos abrir con reintentos
    try:
        _open_file_with_retries(s.reporte_fotografico, attempts=3, delay=0.8)
        f = s.reporte_fotografico  # ya está abierto en modo rb
        filename = f'PHOTOGRAPHIC REPORT {s.proyecto_id}.xlsx'
        resp = FileResponse(f, as_attachment=True, filename=filename)
        resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp["Pragma"] = "no-cache"
        resp["Expires"] = http_date(0)
        return resp
    except Exception:
        # 2) Fallback opcional: URL presignada corta (no carga el web worker)
        try:
            # django-storages S3: .url(expire=...)
            url = s.reporte_fotografico.storage.url(
                s.reporte_fotografico.name, expire=600)
            return HttpResponseRedirect(url)
        except Exception:
            messages.error(
                request, "Could not open the report right now. Please try again.")
            return redirect("operaciones:revisar_sesion", sesion_id=s.id)


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def descargar_reporte_parcial_proyecto(request, sesion_id):
    import os
    s = get_object_or_404(SesionBilling, pk=sesion_id)

    job = (ReporteFotograficoJob.objects
           .filter(sesion=s, estado="ok", log__icontains="[partial]")
           .order_by("-creado_en").first())

    if not job or not job.resultado_key:
        messages.warning(
            request, "Partial report is not available. Please generate it again.")
        return redirect("operaciones:revisar_sesion", sesion_id=s.id)

    key_or_path = job.resultado_key
    if os.path.exists(key_or_path):
        f = open(key_or_path, "rb")
    else:
        from django.core.files.storage import default_storage as storage
        if storage.exists(key_or_path):
            f = storage.open(key_or_path, "rb")
        else:
            messages.warning(
                request, "Partial report not found. Please generate it again.")
            return redirect("operaciones:revisar_sesion", sesion_id=s.id)

    proj_slug = slugify(
        s.proyecto_id or f"billing-{s.id}") or f"billing-{s.id}"
    filename = f"PHOTOGRAPHIC REPORT (partial) {proj_slug}-{s.id}.xlsx"
    resp = FileResponse(f, as_attachment=True, filename=filename)
    resp["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp["Pragma"] = "no-cache"
    resp["Expires"] = http_date(0)
    return resp


def _bytes_excel_reporte_fotografico_qs(sesion: SesionBilling, ev_qs=None) -> bytes:
    """
    Igual que antes pero con centrado exacto de imagen.
    Mantiene in_memory=True (para este caso) y sin cambios en el orden/iteración.
    """
    import io

    import xlsxwriter

    from .models import EvidenciaFotoBilling

    if ev_qs is None:
        ev_qs = (
            EvidenciaFotoBilling.objects
            .filter(tecnico_sesion__sesion=sesion)
            .select_related("requisito")
            .order_by("requisito__orden", "tomada_en", "id")
        )

    bio = io.BytesIO()
    wb = xlsxwriter.Workbook(bio, {"in_memory": True})
    ws = wb.add_worksheet("PHOTOGRAPHIC REPORT")
    ws.hide_gridlines(2)

    fmt_title = wb.add_format({"bold": True, "align": "center",
                              "valign": "vcenter", "border": 1, "bg_color": "#E8EEF7"})
    fmt_head = wb.add_format({"border": 1, "align": "center", "valign": "vcenter",
                             "bold": True, "text_wrap": True, "bg_color": "#F5F7FB", "font_size": 11})
    fmt_box = wb.add_format({"border": 1})
    fmt_info = wb.add_format({"border": 1, "align": "center",
                             "valign": "vcenter", "text_wrap": True, "font_size": 9})

    BLOCK_COLS, SEP_COLS = 6, 1
    LEFT_COL = 0
    RIGHT_COL = LEFT_COL + BLOCK_COLS + SEP_COLS

    # Filas/constantes ANTES de calcular píxeles
    HEAD_ROWS, ROWS_IMG, ROW_INFO, ROW_SPACE = 1, 12, 1, 1
    BLOCK_ROWS = HEAD_ROWS + ROWS_IMG + ROW_INFO

    # Conversión a píxeles
    COL_W = 13
    IMG_ROW_H = 18
    def col_px(w): return int(w * 7 + 5)
    def row_px(h): return int(h * 4 / 3)
    max_w_px = BLOCK_COLS * col_px(COL_W)
    max_h_px = ROWS_IMG * row_px(IMG_ROW_H)

    # Columnas
    for c in range(LEFT_COL, LEFT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)
    ws.set_column(LEFT_COL + BLOCK_COLS, LEFT_COL + BLOCK_COLS, 2)
    for c in range(RIGHT_COL, RIGHT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)

    ws.merge_range(0, 0, 0, RIGHT_COL + BLOCK_COLS - 1,
                   f"ID PROJECT: {sesion.proyecto_id}", fmt_title)
    cur_row = 2

    def draw_block(r, c, ev):
        if sesion.proyecto_especial and ev.requisito_id is None:
            titulo_req = (ev.titulo_manual or "").strip() or "Title (missing)"
        else:
            titulo_req = ((getattr(ev.requisito, "titulo", "")
                          or "").strip() or "Extra")
        ws.merge_range(r, c, r + HEAD_ROWS - 1, c +
                       BLOCK_COLS - 1, titulo_req, fmt_head)
        for rr in range(r, r + HEAD_ROWS):
            ws.set_row(rr, 20)

        img_top = r + HEAD_ROWS
        for rr in range(img_top, img_top + ROWS_IMG):
            ws.set_row(rr, IMG_ROW_H)
        ws.merge_range(img_top, c, img_top + ROWS_IMG -
                       1, c + BLOCK_COLS - 1, "", fmt_box)

        # Escala + centrado
        image_data = None
        x_scale = y_scale = 1.0
        scaled_w = scaled_h = None
        try:
            from PIL import Image
            ev.imagen.open("rb")
            raw = ev.imagen.read()
            image_data = io.BytesIO(raw)
            with Image.open(io.BytesIO(raw)) as im:
                w, h = im.size
            sx = max_w_px / float(w)
            sy = max_h_px / float(h)
            scale = min(sx, sy, 1.0)
            x_scale = y_scale = scale
            scaled_w = int(w * scale)
            scaled_h = int(h * scale)
        except Exception:
            try:
                ev.imagen.open("rb")
                image_data = io.BytesIO(ev.imagen.read())
                scaled_w = max_w_px
                scaled_h = max_h_px
            except Exception:
                image_data = None

        if image_data:
            x_off = max((max_w_px - (scaled_w or max_w_px)) // 2, 0)
            y_off = max((max_h_px - (scaled_h or max_h_px)) // 2, 0)
            ws.insert_image(img_top, c, ev.imagen.name, {
                "image_data": image_data,
                "x_scale": x_scale, "y_scale": y_scale,
                "x_offset": x_off, "y_offset": y_off,
                "object_position": 1,
            })

        info_row = img_top + ROWS_IMG
        dt = ev.client_taken_at or ev.tomada_en
        taken_txt = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
        lat_txt = f"{float(ev.lat):.6f}" if ev.lat is not None else ""
        lng_txt = f"{float(ev.lng):.6f}" if ev.lng is not None else ""
        addr_txt = (ev.direccion_manual or "").strip()

        if sesion.proyecto_especial and ev.requisito_id is None:
            ws.merge_range(info_row, c, info_row, c + 2,
                           f"Taken at\n{taken_txt}", fmt_info)
            ws.merge_range(info_row, c + 3, info_row, c + 5,
                           f"Address\n{addr_txt}",   fmt_info)
        else:
            ws.merge_range(info_row, c, info_row, c + 1,
                           f"Taken at\n{taken_txt}", fmt_info)
            ws.merge_range(info_row, c + 2, info_row, c + 3,
                           f"Lat\n{lat_txt}",        fmt_info)
            ws.merge_range(info_row, c + 4, info_row, c + 5,
                           f"Lng\n{lng_txt}",        fmt_info)
        ws.set_row(info_row, 30)

    idx = 0
    for ev in ev_qs:
        if idx % 2 == 0:
            draw_block(cur_row, LEFT_COL, ev)
        else:
            draw_block(cur_row, RIGHT_COL, ev)
            cur_row += BLOCK_ROWS + ROW_SPACE
        idx += 1
    if idx % 2 == 1:
        cur_row += BLOCK_ROWS + ROW_SPACE

    wb.close()
    return bio.getvalue()


def _bytes_excel_reporte_fotografico(sesion: SesionBilling) -> bytes:
    """
    XLSX con imágenes embebidas (2 por fila) y centradas.
    Mantiene el uso de memoria/flujo original.
    """
    import io

    import xlsxwriter

    from .models import EvidenciaFotoBilling

    evs = (
        EvidenciaFotoBilling.objects
        .filter(tecnico_sesion__sesion=sesion)
        .select_related("requisito")
        .order_by("requisito__orden", "tomada_en", "id")
    )

    bio = io.BytesIO()
    wb = xlsxwriter.Workbook(bio, {"in_memory": True})
    ws = wb.add_worksheet("PHOTOGRAPHIC REPORT")
    ws.hide_gridlines(2)

    fmt_title = wb.add_format({
        "bold": True, "align": "center", "valign": "vcenter",
        "border": 1, "bg_color": "#E8EEF7"
    })
    fmt_head = wb.add_format({
        "border": 1, "align": "center", "valign": "vcenter",
        "bold": True, "text_wrap": True, "bg_color": "#F5F7FB", "font_size": 11
    })
    fmt_box = wb.add_format({"border": 1})
    fmt_info = wb.add_format({
        "border": 1, "align": "center", "valign": "vcenter",
        "text_wrap": True, "font_size": 9
    })

    BLOCK_COLS, SEP_COLS = 6, 1
    LEFT_COL = 0
    RIGHT_COL = LEFT_COL + BLOCK_COLS + SEP_COLS

    # Filas/constantes ANTES de calcular píxeles
    HEAD_ROWS, ROWS_IMG, ROW_INFO, ROW_SPACE = 1, 12, 1, 1
    BLOCK_ROWS = HEAD_ROWS + ROWS_IMG + ROW_INFO

    # Conversión a píxeles
    COL_W = 13
    IMG_ROW_H = 18
    def col_px(w): return int(w * 7 + 5)
    def row_px(h): return int(h * 4 / 3)
    max_w_px = BLOCK_COLS * col_px(COL_W)
    max_h_px = ROWS_IMG * row_px(IMG_ROW_H)

    # Columnas
    for c in range(LEFT_COL, LEFT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)
    ws.set_column(LEFT_COL + BLOCK_COLS, LEFT_COL + BLOCK_COLS, 2)
    for c in range(RIGHT_COL, RIGHT_COL + BLOCK_COLS):
        ws.set_column(c, c, COL_W)

    ws.merge_range(0, 0, 0, RIGHT_COL + BLOCK_COLS - 1,
                   f"ID PROJECT: {sesion.proyecto_id}", fmt_title)
    cur_row = 2

    def draw_block(r, c, ev):
        if sesion.proyecto_especial and ev.requisito_id is None:
            titulo_req = (ev.titulo_manual or "").strip() or "Extra"
        else:
            titulo_req = ((getattr(ev.requisito, "titulo", "")
                          or "").strip() or "Extra")
        ws.merge_range(r, c, r + HEAD_ROWS - 1, c +
                       BLOCK_COLS - 1, titulo_req, fmt_head)
        for rr in range(r, r + HEAD_ROWS):
            ws.set_row(rr, 20)

        img_top = r + HEAD_ROWS
        for rr in range(img_top, img_top + ROWS_IMG):
            ws.set_row(rr, IMG_ROW_H)
        ws.merge_range(img_top, c, img_top + ROWS_IMG -
                       1, c + BLOCK_COLS - 1, "", fmt_box)

        # Escala + centrado
        image_data = None
        x_scale = y_scale = 1.0
        scaled_w = scaled_h = None
        try:
            from PIL import Image
            ev.imagen.open("rb")
            raw = ev.imagen.read()
            image_data = io.BytesIO(raw)
            with Image.open(io.BytesIO(raw)) as im:
                w, h = im.size
            sx = max_w_px / float(w)
            sy = max_h_px / float(h)
            scale = min(sx, sy, 1.0)
            x_scale = y_scale = scale
            scaled_w = int(w * scale)
            scaled_h = int(h * scale)
        except Exception:
            try:
                ev.imagen.open("rb")
                image_data = io.BytesIO(ev.imagen.read())
                scaled_w = max_w_px
                scaled_h = max_h_px
            except Exception:
                image_data = None

        if image_data:
            x_off = max((max_w_px - (scaled_w or max_w_px)) // 2, 0)
            y_off = max((max_h_px - (scaled_h or max_h_px)) // 2, 0)
            ws.insert_image(img_top, c, ev.imagen.name, {
                "image_data": image_data,
                "x_scale": x_scale, "y_scale": y_scale,
                "x_offset": x_off, "y_offset": y_off,
                "object_position": 1,
            })

        info_row = img_top + ROWS_IMG
        dt = ev.client_taken_at or ev.tomada_en
        taken_txt = dt.strftime("%Y-%m-%d %H:%M") if dt else ""
        lat_txt = f"{float(ev.lat):.6f}" if ev.lat is not None else ""
        lng_txt = f"{float(ev.lng):.6f}" if ev.lng is not None else ""
        addr_txt = (ev.direccion_manual or "").strip()

        if sesion.proyecto_especial and ev.requisito_id is None:
            ws.merge_range(info_row, c,         info_row, c + 2,
                           f"Taken at\n{taken_txt}", fmt_info)
            ws.merge_range(info_row, c + 3,     info_row, c + 5,
                           f"Address\n{addr_txt}",   fmt_info)
        else:
            ws.merge_range(info_row, c,         info_row, c + 1,
                           f"Taken at\n{taken_txt}", fmt_info)
            ws.merge_range(info_row, c + 2,     info_row, c + 3,
                           f"Lat\n{lat_txt}",        fmt_info)
            ws.merge_range(info_row, c + 4,     info_row, c + 5,
                           f"Lng\n{lng_txt}",        fmt_info)
        ws.set_row(info_row, 30)

    idx = 0
    for ev in evs:
        if idx % 2 == 0:
            draw_block(cur_row, LEFT_COL, ev)
        else:
            draw_block(cur_row, RIGHT_COL, ev)
            cur_row += BLOCK_ROWS + ROW_SPACE
        idx += 1
    if idx % 2 == 1:
        cur_row += BLOCK_ROWS + ROW_SPACE

    wb.close()
    return bio.getvalue()


@login_required
@rol_requerido('supervisor', 'pm', 'admin')
@require_POST
def regenerar_reporte_fotografico_proyecto(request, sesion_id):
    """
    Encola la regeneración del REPORTE FINAL (nunca bloquea).
    Si ya hay uno en curso, solo informa y redirige.
    """
    from usuarios.schedulers import enqueue_reporte_fotografico

    s = get_object_or_404(SesionBilling, pk=sesion_id)

    last_job = (ReporteFotograficoJob.objects
                .filter(sesion=s)
                .exclude(log__icontains="[partial]")
                .order_by("-creado_en").first())
    if last_job and last_job.estado in ("pendiente", "procesando"):
        messages.info(
            request,
            "Photographic report is already being generated in background. It will replace the current file when it’s ready."
        )
        return redirect("operaciones:revisar_sesion", sesion_id=s.id)

    job = ReporteFotograficoJob.objects.create(
        sesion=s, log="[regen] queued\n")

    # Arranca el job solo después del commit de la creación
    def _start():
        enqueue_reporte_fotografico(job.id)

    transaction.on_commit(_start)

    messages.info(
        request,
        "Regenerating photographic report in background. It will replace the current file when it’s ready."
    )
    return redirect("operaciones:revisar_sesion", sesion_id=s.id)


def _power_port_no_from_evidence(ev):
    """
    Devuelve puerto 1..8 para una evidencia de potencia.

    Regla correcta:
    1) Si la foto pertenece a un requisito POWER PORT X, SIEMPRE manda el puerto del requisito.
       Ejemplo: requisito "POWER PORT 5" => Port 5, aunque la IA haya guardado port=1.
    2) Solo para fotos Extra se usa el port=N guardado por IA o edición manual.
    3) Para Extra también se permite detectar el puerto desde titulo_manual si dice POWER PORT X.
    """

    # ==========================================================
    # 1) Requisito cargado: manda SIEMPRE el requisito
    # ==========================================================
    if ev.requisito_id:
        port_no = getattr(ev.requisito, "power_port_no", None)

        if port_no:
            try:
                port_no = int(port_no)
                if 1 <= port_no <= 8:
                    return port_no
            except Exception:
                pass

        title = (ev.requisito.titulo or "").strip()
        _needs_power, port_no = _power_meta_from_title(title)

        if port_no:
            try:
                port_no = int(port_no)
                if 1 <= port_no <= 8:
                    return port_no
            except Exception:
                pass

        return None

    # ==========================================================
    # 2) Extra: aquí sí manda IA/manual desde power_extract_note
    # ==========================================================
    note = (getattr(ev, "power_extract_note", "") or "").strip()
    m = re.search(r"\bport\s*=\s*([1-8])\b", note, flags=re.IGNORECASE)
    if m:
        try:
            return int(m.group(1))
        except Exception:
            pass

    # ==========================================================
    # 3) Extra con título manual tipo POWER PORT X
    # ==========================================================
    title_manual = (getattr(ev, "titulo_manual", "") or "").strip()
    _needs_power, port_no = _power_meta_from_title(title_manual)

    if port_no:
        try:
            port_no = int(port_no)
            if 1 <= port_no <= 8:
                return port_no
        except Exception:
            pass

    return None


def _build_light_level_workbook(row_count=1):
    """
    Crea el Excel de Light Level con el MISMO formato para individual y masivo.

    Formato:
    - A1:N1: DFN / Column2..Column14
    - A2: Structure ID
    - B2: Light level
    - B3:I3: PORT 1..PORT 8
    - Desde A4 hacia abajo: proyectos seleccionados
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "DFN"
    ws.sheet_view.showGridLines = False

    total_rows = max(7, 3 + int(row_count or 1))

    thick = Side(style="medium", color="000000")
    border = Border(left=thick, right=thick, top=thick, bottom=thick)

    fill_blue = PatternFill("solid", fgColor="C0E6F5")
    font_normal = Font(name="Calibri", size=11, color="000000")

    headers_row_1 = [
        "DFN",
        "Column2",
        "Column3",
        "Column4",
        "Column5",
        "Column6",
        "Column7",
        "Column8",
        "Column9",
        "Column10",
        "Column11",
        "Column12",
        "Column13",
        "Column14",
    ]

    for col_idx, value in enumerate(headers_row_1, start=1):
        ws.cell(row=1, column=col_idx).value = value

    ws["A2"] = "Structure ID"
    ws["B2"] = "Light level"

    ws["B3"] = "PORT 1"
    ws["C3"] = "PORT 2"
    ws["D3"] = "PORT 3"
    ws["E3"] = "PORT 4"
    ws["F3"] = "PORT 5"
    ws["G3"] = "PORT 6"
    ws["H3"] = "PORT 7"
    ws["I3"] = "PORT 8"

    for row in range(1, total_rows + 1):
        for col in range(1, 15):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.font = font_normal
            if row >= 2:
                cell.fill = fill_blue

    ws.column_dimensions["A"].width = 32
    for col in ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N"]:
        ws.column_dimensions[col].width = 13

    for row in range(1, total_rows + 1):
        ws.row_dimensions[row].height = 22

    ws.row_dimensions[3].height = 26

    return wb

def _write_light_level_row(ws, row, structure_id, port_values):
    """
    Escribe una fila de light level usando el formato base:
    A = Structure ID
    B:I = PORT 1..PORT 8
    """
    ws.cell(row=row, column=1).value = structure_id or ""

    for port_no in range(1, 9):
        value = port_values.get(port_no)
        if value is not None:
            ws.cell(row=row, column=port_no + 1).value = float(value)

@login_required
@rol_requerido("supervisor", "admin", "pm")
def export_light_levels_xlsx(request, sesion_id):
    """
    Exporta Excel individual de potencias con el MISMO formato usado en masivo:
    - A2: Structure ID
    - B2: Light level
    - B3:I3: PORT 1..PORT 8
    - A4: ID proyecto
    - B4:I4: potencia de PORT 1..8
    """
    s = get_object_or_404(SesionBilling, pk=sesion_id)

    evidencias = (
        EvidenciaFotoBilling.objects.filter(
            tecnico_sesion__sesion=s,
            power_dbm__isnull=False,
        )
        .select_related("requisito", "tecnico_sesion", "tecnico_sesion__sesion")
        .order_by("client_taken_at", "tomada_en", "id")
    )

    port_values = {}

    for ev in evidencias:
        port_no = _power_port_no_from_evidence(ev)

        if not port_no:
            continue

        try:
            port_no = int(port_no)
        except Exception:
            continue

        if not (1 <= port_no <= 8):
            continue

        port_values[port_no] = ev.power_dbm

    wb = _build_light_level_workbook(row_count=1)
    ws = wb.active

    _write_light_level_row(
        ws=ws,
        row=4,
        structure_id=s.proyecto_id or "",
        port_values=port_values,
    )

    filename = f"LIGHT LEVEL {s.proyecto_id or s.id}.xlsx"

    bio = BytesIO()
    wb.save(bio)
    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@login_required
@rol_requerido("supervisor", "admin", "pm")
def bulk_export_light_levels_xlsx(request):
    """

    Exporta un Excel consolidado de light levels para varios billings seleccionados.

    Usa el MISMO formato del export individual:

    - A2: Structure ID

    - B2: Light level

    - B3:I3: PORT 1..PORT 8

    - Desde A4 hacia abajo: una fila por proyecto/billing.

    """

    ids_raw = (request.GET.get("ids") or request.POST.get("ids") or "").strip()

    ids = []

    for x in ids_raw.split(","):

        x = (x or "").strip()

        if not x:

            continue

        try:

            ids.append(int(x))

        except Exception:

            pass

    ids = list(dict.fromkeys(ids))

    if not ids:

        messages.error(request, "Please select at least one billing.")

        return redirect("operaciones:listar_billing")

    sesiones = list(
        SesionBilling.objects.filter(id__in=ids).order_by("proyecto_id", "id")
    )

    if not sesiones:

        messages.error(request, "No selected billings were found.")

        return redirect("operaciones:listar_billing")

    session_ids = [s.id for s in sesiones]

    evidencias = (
        EvidenciaFotoBilling.objects.filter(
            tecnico_sesion__sesion_id__in=session_ids,
            power_dbm__isnull=False,
        )
        .select_related(
            "requisito",
            "tecnico_sesion",
            "tecnico_sesion__sesion",
        )
        .order_by(
            "tecnico_sesion__sesion_id",
            "client_taken_at",
            "tomada_en",
            "id",
        )
    )

    values_by_session = {sid: {} for sid in session_ids}

    for ev in evidencias:

        port_no = _power_port_no_from_evidence(ev)

        if not port_no:

            continue

        try:

            port_no = int(port_no)

        except Exception:

            continue

        if not (1 <= port_no <= 8):

            continue

        sid = ev.tecnico_sesion.sesion_id

        values_by_session.setdefault(sid, {})[port_no] = ev.power_dbm

    wb = _build_light_level_workbook(row_count=len(sesiones))

    ws = wb.active

    row = 4

    for s in sesiones:

        port_values = values_by_session.get(s.id, {})

        _write_light_level_row(
            ws=ws,
            row=row,
            structure_id=s.proyecto_id or f"Billing #{s.id}",
            port_values=port_values,
        )

        row += 1

    filename = "LIGHT LEVELS SELECTED PROJECTS.xlsx"

    bio = BytesIO()

    wb.save(bio)

    bio.seek(0)

    response = HttpResponse(
        bio.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    return response


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def backfill_light_levels_project(request, sesion_id):
    """
    Procesa fotos históricas ya cargadas para extraer:
    - power_dbm
    - port=1..8 en power_extract_note

    Sirve para proyectos ya finalizados/aprobados.
    No borra fotos, no cambia estados, solo completa metadata de potencia.

    force=1:
    - Reprocesa incluso fotos que ya tienen power_dbm + puerto.
    """
    s = get_object_or_404(SesionBilling, pk=sesion_id)

    force = (request.POST.get("force") or "").strip().lower() in {
        "1",
        "true",
        "yes",
        "y",
        "on",
        "si",
        "sí",
    }

    evidencias = (
        EvidenciaFotoBilling.objects.filter(
            tecnico_sesion__sesion=s,
        )
        .select_related("requisito", "tecnico_sesion", "tecnico_sesion__sesion")
        .order_by("id")
    )

    total = 0
    procesadas = 0
    extraidas = 0
    omitidas = 0
    errores = 0

    for ev in evidencias:
        total += 1

        # Si ya tiene potencia y puerto, no gastamos IA salvo que el usuario fuerce.
        if not force and ev.power_dbm is not None and _power_port_no_from_evidence(ev):
            omitidas += 1
            continue

        should_try = False

        if ev.requisito_id:
            titulo_req = (ev.requisito.titulo or "").strip()
            needs_power, _port_no = _power_meta_from_title(titulo_req)

            should_try = (
                bool(getattr(ev.requisito, "needs_power_reading", False))
                or needs_power
                or titulo_req.upper().startswith("POWER PORT")
            )
        else:
            titulo_manual = (ev.titulo_manual or "").strip()
            nota = (ev.nota or "").strip()
            hint = f"{titulo_manual} {nota}".lower()

            should_try = (
                force
                or ev.power_dbm is not None
                or titulo_manual.lower() in {"", "extra"}
                or any(
                    x in hint
                    for x in ["power", "port", "dbm", "opm", "light level", "light"]
                )
            )

        if not should_try:
            omitidas += 1
            continue

        procesadas += 1

        try:
            result = _extract_power_dbm_for_evidence(
                ev,
                user=request.user,
                allow_extra=True,
                allow_locked=True,
            )

            if result.get("power_dbm"):
                extraidas += 1
            else:
                errores += 1

        except Exception:
            errores += 1
            continue

    if procesadas == 0 and omitidas > 0:
        messages.info(
            request,
            (
                f"Light levels already completed or skipped. "
                f"Total photos: {total}. "
                f"Processed: {procesadas}. "
                f"Extracted: {extraidas}. "
                f"Skipped: {omitidas}. "
                f"Errors: {errores}."
            ),
        )
    else:
        messages.success(
            request,
            (
                f"Light levels backfill completed. "
                f"Total photos: {total}. "
                f"Processed: {procesadas}. "
                f"Extracted: {extraidas}. "
                f"Skipped: {omitidas}. "
                f"Errors: {errores}."
            ),
        )

    return redirect("operaciones:revisar_sesion", sesion_id=s.id)




def _extract_power_dbm_for_evidence(
    ev, user=None, allow_extra=True, allow_locked=False
):
    """
    Extrae automáticamente:
    - potencia dBm del Optical Power Meter
    - puerto físico 1..8 donde está conectado el jumper en la caja

    Regla de puerto:
    - Si la evidencia tiene requisito POWER PORT X, se guarda y se muestra SIEMPRE ese puerto X.
    - Si la evidencia es Extra, se usa el puerto detectado por IA.
    """
    import base64
    import json
    import os
    import re
    from decimal import Decimal, InvalidOperation

    from django.conf import settings
    from django.utils import timezone

    ses = ev.tecnico_sesion.sesion

    if (
        getattr(ses, "estado", "") in ("aprobado_supervisor", "aprobado_pm")
        and not allow_locked
    ):
        raise ValueError("Locked after approval.")

    if not ev.requisito_id and not allow_extra:
        raise ValueError("This evidence has no requirement.")

    api_key = getattr(settings, "OPENAI_API_KEY", "") or os.getenv("OPENAI_API_KEY", "")
    if not api_key:
        raise ValueError("OPENAI_API_KEY is not configured on the server.")

    model_name = getattr(settings, "OPENAI_VISION_MODEL", "gpt-4o-mini")

    try:
        from openai import OpenAI
    except Exception:
        raise ValueError("openai package is not installed. Run: pip install openai")

    def _read_image_bytes(evidence):
        try:
            evidence.imagen.open("rb")
            return evidence.imagen.read()
        finally:
            try:
                evidence.imagen.close()
            except Exception:
                pass

    def _normalize_dbm_value(raw):
        txt = (raw or "").strip()
        txt = txt.replace(",", ".")
        txt = txt.replace("−", "-").replace("–", "-").replace("—", "-")

        m = re.search(r"-\s*\d{1,2}(?:\.\d{1,2})?", txt)
        if not m:
            return None

        val_txt = m.group(0).replace(" ", "")

        try:
            val = Decimal(val_txt)
        except InvalidOperation:
            return None

        if not (Decimal("-60.00") <= val <= Decimal("0.00")):
            return None

        return val.quantize(Decimal("0.01"))

    def _normalize_port(raw):
        if raw in (None, "", "null"):
            return None

        try:
            port = int(str(raw).strip())
        except Exception:
            return None

        if 1 <= port <= 8:
            return port

        return None

    raw_bytes = _read_image_bytes(ev)
    if not raw_bytes:
        raise ValueError("Image file is empty.")

    image_b64 = base64.b64encode(raw_bytes).decode("utf-8")

    name = (getattr(ev.imagen, "name", "") or "").lower()
    if name.endswith(".png"):
        mime = "image/png"
    elif name.endswith(".webp"):
        mime = "image/webp"
    else:
        mime = "image/jpeg"

    known_port = None
    titulo = ""

    if ev.requisito_id:
        titulo = ev.requisito.titulo or ""

        known_port = getattr(ev.requisito, "power_port_no", None)
        if not known_port:
            _needs_power, known_port = _power_meta_from_title(titulo)

        try:
            known_port = int(known_port) if known_port else None
        except Exception:
            known_port = None

        if known_port and not (1 <= known_port <= 8):
            known_port = None
    else:
        titulo = ev.titulo_manual or "Extra"

    prompt = f"""
You are analyzing a telecom field photo for a light level report.

The photo usually contains:
- An Optical Power Meter with an LCD screen showing a dBm value.
- A fiber terminal box with 8 green adapter ports.
- A yellow or green fiber jumper connected from the power meter to one of the green ports.

Return the result as strict JSON only.

Tasks:
1. Read ONLY the main optical power value shown on the Optical Power Meter LCD screen in dBm.
2. Detect the physical port number on the terminal box where the active yellow/green jumper is connected.

Important context:
- Evidence title: {titulo}
- Known requirement port: {known_port or "unknown"}
- If Known requirement port is available, it belongs to the original loaded requirement.
- If this is a normal requirement photo, the system will use the known requirement port.
- If this is an Extra photo, the system will use your detected physical port.

How to count ports:
- Look at the row of green adapter ports on the terminal box.
- Count the green ports from LEFT to RIGHT as:
  PORT 1, PORT 2, PORT 3, PORT 4, PORT 5, PORT 6, PORT 7, PORT 8.
- The active port is the green port where the jumper from the power meter is plugged in.
- Follow the yellow/green jumper coming from the Optical Power Meter.
- Do NOT use the GPS text, date, structure ID, meter serial number, address, or any overlay numbers as the port.
- Do NOT use the dBm value as the port.
- If several yellow fibers are visible, choose the one physically connected to the power meter/test lead.

Rules:
- Do not guess the dBm value if unreadable.
- If the dBm value is readable, return found true.
- If the port is visible but not perfectly clear, return your best port number and lower port_confidence.
- Use null for port_no only when the terminal box port connection is not visible at all.

Respond ONLY as strict JSON:
{{
  "found": true,
  "value_dbm": "-30.21",
  "port_no": 5,
  "confidence": 0.0,
  "port_confidence": 0.0,
  "reason": "short reason"
}}
"""

    client = OpenAI(api_key=api_key)

    resp = client.responses.create(
        model=model_name,
        input=[
            {
                "role": "user",
                "content": [
                    {
                        "type": "input_text",
                        "text": prompt,
                    },
                    {
                        "type": "input_image",
                        "image_url": f"data:{mime};base64,{image_b64}",
                        "detail": "high",
                    },
                ],
            }
        ],
        temperature=0,
    )

    content = (getattr(resp, "output_text", "") or "").strip()

    cleaned = content.strip()
    cleaned = re.sub(r"^```json\s*", "", cleaned, flags=re.IGNORECASE)
    cleaned = re.sub(r"^```\s*", "", cleaned)
    cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        data = json.loads(cleaned)
    except Exception:
        raise ValueError("Vision response was not valid JSON.")

    found = bool(data.get("found"))
    confidence = data.get("confidence", 0)
    port_confidence = data.get("port_confidence", 0)
    raw_value = data.get("value_dbm", "")

    detected_port_no = _normalize_port(data.get("port_no"))

    try:
        confidence_decimal = Decimal(str(confidence))
    except Exception:
        confidence_decimal = Decimal("0")

    try:
        port_confidence_decimal = Decimal(str(port_confidence))
    except Exception:
        port_confidence_decimal = Decimal("0")

    val = _normalize_dbm_value(raw_value)

    if not found or val is None:
        raise ValueError("Could not find a valid dBm value.")

    if confidence_decimal < Decimal("0.70"):
        raise ValueError(
            "The dBm value was detected but confidence is too low. Please review manually."
        )

    # ==========================================================
    # REGLA FINAL DEL PUERTO
    # ==========================================================
    if ev.requisito_id and known_port:
        # Si ya viene de un requisito POWER PORT X, manda ese puerto.
        port_no = known_port
    else:
        # Solo extras usan el puerto detectado por IA.
        port_no = detected_port_no

    ev.power_dbm = val
    ev.power_extracted_at = timezone.now()
    ev.power_extracted_by = user

    note_parts = [
        "Vision OCR",
        f"model={model_name}",
        f"confidence={confidence_decimal}",
        f"raw={raw_value}",
    ]

    if ev.requisito_id and known_port:
        note_parts.append(f"port={known_port}")
        note_parts.append("port_source=requirement")
        if detected_port_no:
            note_parts.append(f"detected_port={detected_port_no}")
            note_parts.append(f"detected_port_confidence={port_confidence_decimal}")
    else:
        if port_no:
            note_parts.append(f"port={port_no}")
            note_parts.append("port_source=vision_extra")
            note_parts.append(f"port_confidence={port_confidence_decimal}")
        else:
            note_parts.append("port=null")
            note_parts.append("port_source=vision_extra")
            note_parts.append(f"port_confidence={port_confidence_decimal}")

    reason = (data.get("reason") or "")[:100]
    if reason:
        note_parts.append(f"reason={reason}")

    ev.power_extract_note = " | ".join(note_parts)[:255]

    ev.save(
        update_fields=[
            "power_dbm",
            "power_extracted_at",
            "power_extracted_by",
            "power_extract_note",
        ]
    )

    return {
        "power_dbm": f"{val:.2f}",
        "port_no": port_no,
        "confidence": str(confidence_decimal),
        "port_confidence": str(port_confidence_decimal),
        "method": "vision",
        "raw_response": data,
    }


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def extract_power_from_evidence(request, evidencia_id: int):
    """
    Botón manual para extraer potencia dBm y puerto.
    Sirve para requisitos POWER PORT y también para Extras.
    """
    ev = get_object_or_404(
        EvidenciaFotoBilling.objects.select_related(
            "requisito",
            "tecnico_sesion",
            "tecnico_sesion__sesion",
        ),
        pk=evidencia_id,
    )

    try:
        result = _extract_power_dbm_for_evidence(ev, user=request.user)
    except ValueError as e:
        return JsonResponse(
            {
                "ok": False,
                "error": str(e),
            },
            status=422,
        )
    except Exception as e:
        return JsonResponse(
            {
                "ok": False,
                "error": f"Vision extraction failed: {e}",
            },
            status=500,
        )

    return JsonResponse(
        {
            "ok": True,
            "power_dbm": result.get("power_dbm", ""),
            "port_no": result.get("port_no", None),
            "confidence": result.get("confidence", ""),
            "port_confidence": result.get("port_confidence", ""),
            "method": result.get("method", "vision"),
            "raw_response": result.get("raw_response", {}),
        }
    )


# ============================
# Editar potencia en la linea
# ============================
@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def update_power_from_evidence(request, evidencia_id: int):
    import json
    import re
    from decimal import Decimal, InvalidOperation

    ev = get_object_or_404(
        EvidenciaFotoBilling.objects.select_related(
            "requisito",
            "tecnico_sesion",
            "tecnico_sesion__sesion",
        ),
        pk=evidencia_id,
    )

    try:
        payload = json.loads(request.body.decode("utf-8") or "{}")
    except Exception:
        payload = {}

    raw_power = str(payload.get("power_dbm", "")).strip()
    raw_port = str(payload.get("port_no", "")).strip()

    # ==========================================================
    # Detectar si es evidencia de Power Port
    # ==========================================================
    is_power_port = False

    if ev.requisito_id:
        titulo_req = (ev.requisito.titulo or "").strip()
        needs_power, _port_no = _power_meta_from_title(titulo_req)

        is_power_port = (
            bool(getattr(ev.requisito, "needs_power_reading", False))
            or needs_power
            or ev.power_dbm is not None
            or _power_port_no_from_evidence(ev)
        )
    else:
        titulo_manual = (ev.titulo_manual or "").strip()
        nota = (ev.nota or "").strip()
        needs_power, _port_no = _power_meta_from_title(titulo_manual)

        # ✅ FIX:
        # Si es Extra, permitimos edición manual cuando el usuario envía power o port.
        # Esto evita el error:
        # "This evidence is not marked as Power Port."
        is_power_port = (
            needs_power
            or ev.power_dbm is not None
            or _power_port_no_from_evidence(ev)
            or bool(raw_power)
            or bool(raw_port)
            or titulo_manual.lower() in {"extra", ""}
            or any(
                x in f"{titulo_manual} {nota}".lower()
                for x in ["power", "port", "dbm", "opm", "light level", "light"]
            )
        )

    if not is_power_port:
        return JsonResponse(
            {"ok": False, "error": "This evidence is not marked as Power Port."},
            status=400,
        )

    # ==========================
    # Validar puerto
    # ==========================
    port_no = None

    if raw_port:
        try:
            port_no = int(raw_port)
        except Exception:
            return JsonResponse(
                {"ok": False, "error": "Port must be a number from 1 to 8."},
                status=400,
            )

        if not (1 <= port_no <= 8):
            return JsonResponse(
                {"ok": False, "error": "Port must be between 1 and 8."},
                status=400,
            )
    else:
        port_no = _power_port_no_from_evidence(ev)

    # ✅ Si pertenece a requisito POWER PORT, actualizamos también el requisito.
    # En Extra no hay requisito, por eso el puerto queda guardado en power_extract_note.
    if ev.requisito_id and port_no:
        req = ev.requisito
        req.power_port_no = port_no
        req.needs_power_reading = True
        req.save(update_fields=["power_port_no", "needs_power_reading"])

    # ==========================
    # Limpiar potencia
    # ==========================
    raw = raw_power
    raw = raw.replace(",", ".")
    raw = raw.replace("−", "-").replace("–", "-").replace("—", "-")
    raw = raw.replace("dbm", "").replace("dBm", "").replace("DBM", "").strip()

    # ==========================
    # Si viene vacío, limpiar power pero conservar port si existe
    # ==========================
    if raw == "":
        ev.power_dbm = None
        ev.power_extracted_at = timezone.now()
        ev.power_extracted_by = request.user

        note_parts = [f"Manual edit | cleared | user={request.user.username}"]

        if port_no:
            note_parts.append(f"port={port_no}")
            if ev.requisito_id:
                note_parts.append("port_source=manual_requirement")
            else:
                note_parts.append("port_source=manual_extra")
            note_parts.append("port_confidence=manual")

        ev.power_extract_note = " | ".join(note_parts)[:255]

        ev.save(
            update_fields=[
                "power_dbm",
                "power_extracted_at",
                "power_extracted_by",
                "power_extract_note",
            ]
        )

        return JsonResponse(
            {
                "ok": True,
                "power_dbm": "",
                "port_no": port_no,
                "display": "pending",
            }
        )

    m = re.search(r"-?\d{1,2}(?:\.\d{1,2})?", raw)
    if not m:
        return JsonResponse(
            {"ok": False, "error": "Invalid value. Example: -28.03"},
            status=400,
        )

    try:
        val = Decimal(m.group(0))
    except InvalidOperation:
        return JsonResponse(
            {"ok": False, "error": "Invalid value. Example: -28.03"},
            status=400,
        )

    if not (Decimal("-60.00") <= val <= Decimal("0.00")):
        return JsonResponse(
            {"ok": False, "error": "Power must be between -60.00 and 0.00 dBm."},
            status=400,
        )

    val = val.quantize(Decimal("0.01"))

    ev.power_dbm = val
    ev.power_extracted_at = timezone.now()
    ev.power_extracted_by = request.user

    note_parts = [f"Manual edit | user={request.user.username}"]

    if port_no:
        note_parts.append(f"port={port_no}")
        if ev.requisito_id:
            note_parts.append("port_source=manual_requirement")
        else:
            note_parts.append("port_source=manual_extra")
        note_parts.append("port_confidence=manual")

    ev.power_extract_note = " | ".join(note_parts)[:255]

    ev.save(
        update_fields=[
            "power_dbm",
            "power_extracted_at",
            "power_extracted_by",
            "power_extract_note",
        ]
    )

    return JsonResponse(
        {
            "ok": True,
            "power_dbm": f"{val:.2f}",
            "port_no": port_no,
            "display": f"{val:.2f} dBm",
        }
    )


# ============================
# CONFIGURAR REQUISITOS (¡la que faltaba!)
# ============================


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def confirmar_importar_requisitos(request, sesion_id):
    """
    CONFIRM:
    - Lee preview desde session
    - Crea SOLO requisitos nuevos en TODAS las asignaciones
    - NO modifica los existentes (por tu regla)
    - Marca automáticamente POWER PORT X como lectura de potencia
    """
    from django.contrib import messages
    from django.db import transaction
    from django.shortcuts import get_object_or_404, redirect
    from django.utils.text import slugify

    s = get_object_or_404(SesionBilling, pk=sesion_id)

    payload = request.session.get("req_import_preview") or {}
    if not payload or payload.get("sesion_id") != s.id:
        messages.error(
            request, "No preview data to confirm. Please re-upload the file."
        )
        return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

    to_create = payload.get("to_create") or []
    if not isinstance(to_create, list) or not to_create:
        messages.info(request, "Nothing to create.")
        request.session.pop("req_import_preview", None)
        return redirect("operaciones:configurar_requisitos", sesion_id=sesion_id)

    try:
        with transaction.atomic():
            asignaciones = list(
                s.tecnicos_sesion.select_related("tecnico")
                .prefetch_related("requisitos")
                .all()
            )

            created_total = 0

            for a in asignaciones:
                existentes_por_slug = {
                    slugify((r.titulo or "").strip()): r for r in a.requisitos.all()
                }

                # ✅ fallback de orden estable por asignación
                try:
                    current_max_order = max(
                        (r.orden for r in a.requisitos.all()), default=-1
                    )
                except Exception:
                    current_max_order = -1
                next_order = current_max_order + 1

                for r in to_create:
                    name = (r.get("name") or "").strip()
                    if not name:
                        continue

                    key = slugify(name)
                    if not key:
                        continue

                    if key in existentes_por_slug:
                        continue  # seguridad extra

                    # order: si viene bien en el archivo, úsalo. Si no, usa next_order incremental.
                    raw_order = r.get("order", None)
                    order = None
                    if raw_order not in (None, "", "null"):
                        try:
                            order = int(raw_order)
                        except Exception:
                            order = None
                    if order is None:
                        order = next_order
                        next_order += 1

                    mandatory = bool(r.get("mandatory", True))

                    # ✅ NUEVO: detectar POWER PORT 1..8
                    needs_power, port_no = _power_meta_from_title(name)

                    RequisitoFotoBilling.objects.create(
                        tecnico_sesion=a,
                        titulo=name,
                        descripcion="",
                        obligatorio=mandatory,
                        orden=order,
                        needs_power_reading=bool(needs_power),
                        power_port_no=port_no,
                    )

                    existentes_por_slug[key] = True
                    created_total += 1

        messages.success(
            request,
            f"Created {created_total} new requirement(s). Existing identical requirements were kept unchanged.",
        )
    except Exception as e:
        messages.error(request, f"Could not apply imported requirements: {e}")
        return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

    request.session.pop("req_import_preview", None)
    return redirect("operaciones:configurar_requisitos", sesion_id=sesion_id)




@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def configurar_requisitos(request, sesion_id):
    """
    Configura la lista compartida de requisitos a nivel de proyecto
    y la sincroniza con TODAS las asignaciones SIN borrar estados previos.
    + NUEVO: marca automáticamente needs_power_reading/power_port_no si el título es POWER PORT X.
    """
    s = get_object_or_404(SesionBilling, pk=sesion_id)

    asignaciones = list(
        s.tecnicos_sesion.select_related("tecnico")
         .prefetch_related("requisitos")
         .all()
    )
    canonical = []
    if asignaciones and asignaciones[0].requisitos.exists():
        canonical = list(asignaciones[0].requisitos.order_by("orden", "id"))

    if request.method == "POST":
        names = request.POST.getlist("name[]")
        orders = request.POST.getlist("order[]")
        mand = request.POST.getlist("mandatory[]")
        ids = request.POST.getlist("id[]")
        to_del = set(request.POST.getlist("delete_id[]"))
        s.proyecto_especial = bool(request.POST.get("proyecto_especial"))

        normalized = []  # [(req_id, orden, name, mandatory)]
        for i, raw_name in enumerate(names):
            name = (raw_name or "").strip()
            if not name:
                continue
            try:
                orden = int(orders[i]) if i < len(orders) else i
            except Exception:
                orden = i
            mandatory = (mand[i] == "1") if i < len(mand) else True
            req_id = (ids[i].strip() or None) if i < len(ids) else None
            normalized.append((req_id, orden, name, mandatory))

        try:
            with transaction.atomic():
                s.save(update_fields=["proyecto_especial"])

                for a in asignaciones:
                    existentes = {str(r.id): r for r in a.requisitos.all()}
                    existentes_por_slug = {
                        slugify((r.titulo or "").strip()): r for r in a.requisitos.all()
                    }

                    if to_del:
                        del_objs = [existentes[x] for x in to_del if x in existentes]
                        if del_objs:
                            RequisitoFotoBilling.objects.filter(
                                id__in=[d.id for d in del_objs]
                            ).delete()

                    for req_id, orden, name, mandatory in normalized:
                        needs_power, port_no = _power_meta_from_title(name)

                        if req_id and req_id in existentes:
                            r = existentes[req_id]
                            changed = (
                                r.titulo != name
                                or r.orden != orden
                                or r.obligatorio != mandatory
                                or r.tecnico_sesion_id != a.id
                                or bool(getattr(r, "needs_power_reading", False)) != bool(needs_power)
                                or getattr(r, "power_port_no", None) != port_no
                            )
                            if changed:
                                r.titulo = name
                                r.orden = orden
                                r.obligatorio = mandatory
                                r.tecnico_sesion = a
                                r.needs_power_reading = bool(needs_power)
                                r.power_port_no = port_no
                                r.save(update_fields=[
                                    "titulo", "orden", "obligatorio", "tecnico_sesion",
                                    "needs_power_reading", "power_port_no",
                                ])
                        else:
                            key = slugify(name)
                            r = existentes_por_slug.get(key)

                            if r and str(r.id) not in to_del:
                                # rename/match por slug
                                changed = (
                                    r.titulo != name
                                    or r.orden != orden
                                    or r.obligatorio != mandatory
                                    or bool(getattr(r, "needs_power_reading", False)) != bool(needs_power)
                                    or getattr(r, "power_port_no", None) != port_no
                                )
                                if changed:
                                    r.titulo = name
                                    r.orden = orden
                                    r.obligatorio = mandatory
                                    r.needs_power_reading = bool(needs_power)
                                    r.power_port_no = port_no
                                    r.save(update_fields=[
                                        "titulo", "orden", "obligatorio",
                                        "needs_power_reading", "power_port_no",
                                    ])
                            else:
                                RequisitoFotoBilling.objects.create(
                                    tecnico_sesion=a,
                                    titulo=name,
                                    descripcion="",
                                    obligatorio=mandatory,
                                    orden=orden,
                                    needs_power_reading=bool(needs_power),
                                    power_port_no=port_no,
                                )

            messages.success(request, "Photo requirements saved (project-wide).")
            return redirect("operaciones:listar_billing")

        except Exception as e:
            messages.error(request, f"Could not save requirements: {e}")

        class _Row:
            def __init__(self, orden, titulo, obligatorio):
                self.orden = orden
                self.titulo = titulo
                self.obligatorio = obligatorio

        canonical = [_Row(o, n, m) for _, o, n, m in normalized]

    return render(
        request,
        "operaciones/billing_configurar_requisitos.html",
        {
            "sesion": s,
            "requirements": canonical,
            "is_special": bool(s.proyecto_especial),
        },
    )

@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def import_requirements_page(request, sesion_id):
    """
    Shows the import screen with download links for the template and
    a file input to upload the CSV/XLSX.
    """
    s = get_object_or_404(SesionBilling, pk=sesion_id)
    return render(
        request,
        "operaciones/billing_import_requisitos.html",
        {"sesion": s},
    )


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
def download_requirements_template(request, sesion_id, ext):
    """
    Returns a requirements template as CSV or XLSX.
    Columns: name, order, mandatory
    - name: string (required)
    - order: integer (optional)
    - mandatory: 1/0 or true/false (optional; defaults to 1/true)
    """
    ext = (ext or "").lower()
    filename_base = f"requirements_template_billing_{sesion_id}"

    if ext == "csv":
        content = (
            "name,order,mandatory\n"
            "Front door,0,1\n"
            "Back door,1,1\n"
            "Panorama of site,2,0\n"
        )
        resp = HttpResponse(content, content_type="text/csv; charset=utf-8")
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.csv"'
        return resp

    if ext in ("xlsx", "xls"):
        wb = Workbook()
        ws = wb.active
        ws.title = "Requirements"
        ws.append(["name", "order", "mandatory"])
        ws.append(["Front door", 0, 1])
        ws.append(["Back door", 1, 1])
        ws.append(["Panorama of site", 2, 0])

        from io import BytesIO
        bio = BytesIO()
        wb.save(bio)
        bio.seek(0)
        resp = HttpResponse(
            bio.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        resp["Content-Disposition"] = f'attachment; filename="{filename_base}.xlsx"'
        return resp

    messages.error(request, "Unsupported format. Use csv or xlsx.")
    return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
@require_POST
def importar_requisitos(request, sesion_id):
    """
    PREVIEW (sin escribir DB):
    - Parse CSV/XLSX
    - Detecta duplicados dentro del archivo (por slug) y reporta:
        * file row duplicada
        * nombre duplicado
        * file row original que está duplicando
    - Detecta duplicados contra requisitos existentes (por slug) y reporta:
        * file row entrante
        * existing row number (1..N) del listado actual del billing (orden/id)
    - Guarda preview en session
    - Renderiza pantalla preview para confirmar
    """
    import csv
    import io

    from django.contrib import messages
    from django.shortcuts import get_object_or_404, redirect, render
    from django.utils.text import slugify
    from openpyxl import load_workbook

    s = get_object_or_404(SesionBilling, pk=sesion_id)
    f = request.FILES.get("file")

    if not f:
        messages.error(request, "Please select a CSV or XLSX file.")
        return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

    filename = (getattr(f, "name", "") or "").strip()
    ext = (filename.rsplit(".", 1)[-1] or "").lower()

    def _to_bool(v, default=True):
        if v is None or v == "":
            return default
        t = str(v).strip().lower()
        return t in ("1", "true", "yes", "y", "si", "sí")

    normalized = []   # [{'order':int,'name':str,'mandatory':bool,'slug':str,'row':int}]
    warnings = []
    errors = []

    # ---------------- PARSEO ----------------
    try:
        if ext == "csv":
            raw = f.read().decode("utf-8", errors="ignore")
            if not raw.strip():
                messages.warning(request, "The file is empty.")
                return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

            lines = raw.splitlines()
            header_line = (lines[0].lower() if lines else "")
            has_header = "name" in header_line

            if has_header:
                reader = csv.DictReader(io.StringIO(raw))
                for idx, row in enumerate(reader, start=2):  # header is row 1
                    name = (row.get("name") or "").strip()
                    if not name:
                        continue
                    try:
                        order = int(row.get("order")) if (row.get("order") not in (None, "")) else len(normalized)
                    except Exception:
                        order = len(normalized)
                    mandatory = _to_bool(row.get("mandatory"), default=True)
                    normalized.append({
                        "order": order,
                        "name": name,
                        "mandatory": mandatory,
                        "slug": slugify(name),
                        "row": idx,   # ✅ file row number
                    })
            else:
                reader = csv.reader(io.StringIO(raw))
                for idx, row in enumerate(reader, start=1):
                    if not row:
                        continue
                    name = (row[0] or "").strip()
                    if not name:
                        continue
                    normalized.append({
                        "order": len(normalized),
                        "name": name,
                        "mandatory": True,
                        "slug": slugify(name),
                        "row": idx,   # ✅ file row number
                    })

        elif ext in ("xlsx", "xls"):
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messages.warning(request, "The spreadsheet is empty.")
                return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

            header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            has_header = "name" in header
            start = 1 if has_header else 0

            if has_header:
                i_name = header.index("name")
                i_order = header.index("order") if "order" in header else None
                i_mand = header.index("mandatory") if "mandatory" in header else None

                for ridx, r in enumerate(rows[start:], start=2):  # header is row 1
                    name = (str(r[i_name]) if i_name < len(r) and r[i_name] is not None else "").strip()
                    if not name:
                        continue

                    if i_order is not None and i_order < len(r) and r[i_order] not in (None, ""):
                        try:
                            order = int(r[i_order])
                        except Exception:
                            order = len(normalized)
                    else:
                        order = len(normalized)

                    if i_mand is not None and i_mand < len(r):
                        mandatory = _to_bool(r[i_mand], default=True)
                    else:
                        mandatory = True

                    normalized.append({
                        "order": order,
                        "name": name,
                        "mandatory": mandatory,
                        "slug": slugify(name),
                        "row": ridx,  # ✅ file row number
                    })
            else:
                for ridx, r in enumerate(rows, start=1):
                    if not r:
                        continue
                    name = (str(r[0]) if r[0] is not None else "").strip()
                    if not name:
                        continue
                    normalized.append({
                        "order": len(normalized),
                        "name": name,
                        "mandatory": True,
                        "slug": slugify(name),
                        "row": ridx,  # ✅ file row number
                    })
        else:
            messages.error(request, "Unsupported file type. Use .csv or .xlsx.")
            return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

    except Exception as e:
        messages.error(request, f"Could not parse the file: {e}")
        return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

    if not normalized:
        messages.warning(request, "No valid rows found in the file.")
        return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

    # --------- Dedup dentro del archivo por slug (con referencia a la fila original) ----------
    seen = {}  # slug -> {"row": int, "name": str}
    cleaned = []
    file_duplicates = []  # para tabla opcional

    for r in normalized:
        slug = r.get("slug") or ""
        rownum = r.get("row")
        nm = (r.get("name") or "").strip()

        if not slug:
            errors.append(f"Row {rownum}: invalid name.")
            continue

        if slug in seen:
            first = seen[slug]
            warnings.append(
                f"Row {rownum}: duplicated name in file — '{nm}' — duplicates row {first['row']} ('{first['name']}')."
            )
            file_duplicates.append({
                "row": rownum,
                "name": nm,
                "dup_of_row": first["row"],
                "dup_of_name": first["name"],
                "slug": slug,
            })
            continue

        seen[slug] = {"row": rownum, "name": nm}
        cleaned.append(r)

    normalized = cleaned

    if not normalized and not errors:
        messages.warning(request, "No valid rows found in the file.")
        return redirect("operaciones:import_requirements_page", sesion_id=sesion_id)

    # --------- EXISTENTES (canónico = primera asignación) + row number visual (1..N) ----------
    asignaciones = list(
        s.tecnicos_sesion.select_related("tecnico").prefetch_related("requisitos").all()
    )

    existing_by_slug = {}  # slug -> {"req": obj, "row_num": int}
    if asignaciones:
        canonical_qs = asignaciones[0].requisitos.order_by("orden", "id")
        for idx, req in enumerate(canonical_qs, start=1):
            key = slugify((req.titulo or "").strip())
            if key and key not in existing_by_slug:
                existing_by_slug[key] = {"req": req, "row_num": idx}

    duplicates = []
    to_create = []

    for r in normalized:
        ex_info = existing_by_slug.get(r["slug"])
        if ex_info:
            ex = ex_info["req"]
            duplicates.append({
                "incoming_row": r.get("row"),                 # ✅ file row
                "existing_row_num": ex_info.get("row_num"),   # ✅ row in current billing list (1..N)
                "name": r["name"],
                "slug": r["slug"],
                "existing_order": getattr(ex, "orden", 0),
                "existing_mandatory": bool(getattr(ex, "obligatorio", True)),
                "incoming_order": r["order"],
                "incoming_mandatory": r["mandatory"],
            })
        else:
            to_create.append({
                "row": r.get("row"),          # ✅ file row
                "name": r["name"],
                "slug": r["slug"],
                "order": r["order"],
                "mandatory": r["mandatory"],
            })

    request.session["req_import_preview"] = {
        "sesion_id": s.id,
        "source_filename": filename,
        "to_create": to_create,
        "duplicates": duplicates,
        "file_duplicates": file_duplicates,
        "warnings": warnings,
        "errors": errors,
    }

    return render(
        request,
        "operaciones/preview_import_requirements.html",
        {
            "sesion": s,
            "source_filename": filename,
            "duplicates": duplicates,
            "to_create": to_create,
            "file_duplicates": file_duplicates,
            "warnings": warnings,
            "errors": errors,
        },
    )


@login_required
@rol_requerido('supervisor', 'admin', 'pm')
@require_POST
def bulk_importar_requisitos_preview(request):
    """
    BULK PREVIEW (sin escribir DB):
    - Recibe ids (seleccionados) + file
    - Filtra sesiones por estado permitido: asignado/en_proceso/en_revision_supervisor
      (NO aprobado supervisor/pm)
    - Parsea el archivo una sola vez (igual que importar_requisitos)
    - Para cada sesión elegible:
        * detecta duplicados vs requisitos existentes (canonical = primera asignación)
        * arma to_create / duplicates
    - Guarda preview en session y renderiza template de preview (reutiliza preview_import_requirements.html)
    """
    ids_raw = (request.POST.get("ids") or "").strip()
    f = request.FILES.get("file")

    if not ids_raw:
        messages.error(request, "Please select at least one billing.")
        return redirect("operaciones:listar_billing")

    if not f:
        messages.error(request, "Please select a CSV or XLSX file.")
        return redirect("operaciones:listar_billing")

    # -------- Parse ids --------
    ids = []
    for x in ids_raw.split(","):
        x = (x or "").strip()
        if not x:
            continue
        try:
            ids.append(int(x))
        except Exception:
            pass

    if not ids:
        messages.error(request, "Invalid selection.")
        return redirect("operaciones:listar_billing")

    # -------- Helpers parse (MISMO comportamiento que importar_requisitos) --------
    filename = (getattr(f, "name", "") or "").strip()
    ext = (filename.rsplit(".", 1)[-1] or "").lower()

    def _to_bool(v, default=True):
        if v is None or v == "":
            return default
        t = str(v).strip().lower()
        return t in ("1", "true", "yes", "y", "si", "sí")

    normalized = []   # [{'order':int,'name':str,'mandatory':bool,'slug':str,'row':int}]
    warnings = []
    errors = []

    # ---------------- PARSEO ----------------
    try:
        if ext == "csv":
            raw = f.read().decode("utf-8", errors="ignore")
            if not raw.strip():
                messages.warning(request, "The file is empty.")
                return redirect("operaciones:listar_billing")

            lines = raw.splitlines()
            header_line = (lines[0].lower() if lines else "")
            has_header = "name" in header_line

            if has_header:
                reader = csv.DictReader(io.StringIO(raw))
                for idx, row in enumerate(reader, start=2):  # header row is 1
                    name = (row.get("name") or "").strip()
                    if not name:
                        continue
                    try:
                        order = int(row.get("order")) if (row.get("order") not in (None, "")) else len(normalized)
                    except Exception:
                        order = len(normalized)
                    mandatory = _to_bool(row.get("mandatory"), default=True)
                    normalized.append({
                        "order": order,
                        "name": name,
                        "mandatory": mandatory,
                        "slug": slugify(name),
                        "row": idx,
                    })
            else:
                reader = csv.reader(io.StringIO(raw))
                for idx, row in enumerate(reader, start=1):
                    if not row:
                        continue
                    name = (row[0] or "").strip()
                    if not name:
                        continue
                    normalized.append({
                        "order": len(normalized),
                        "name": name,
                        "mandatory": True,
                        "slug": slugify(name),
                        "row": idx,
                    })

        elif ext in ("xlsx", "xls"):
            wb = load_workbook(f, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                messages.warning(request, "The spreadsheet is empty.")
                return redirect("operaciones:listar_billing")

            header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]
            has_header = "name" in header
            start = 1 if has_header else 0

            if has_header:
                i_name = header.index("name")
                i_order = header.index("order") if "order" in header else None
                i_mand = header.index("mandatory") if "mandatory" in header else None

                for ridx, r in enumerate(rows[start:], start=2):  # header row is 1
                    name = (str(r[i_name]) if i_name < len(r) and r[i_name] is not None else "").strip()
                    if not name:
                        continue

                    if i_order is not None and i_order < len(r) and r[i_order] not in (None, ""):
                        try:
                            order = int(r[i_order])
                        except Exception:
                            order = len(normalized)
                    else:
                        order = len(normalized)

                    if i_mand is not None and i_mand < len(r):
                        mandatory = _to_bool(r[i_mand], default=True)
                    else:
                        mandatory = True

                    normalized.append({
                        "order": order,
                        "name": name,
                        "mandatory": mandatory,
                        "slug": slugify(name),
                        "row": ridx,
                    })
            else:
                for ridx, r in enumerate(rows, start=1):
                    if not r:
                        continue
                    name = (str(r[0]) if r[0] is not None else "").strip()
                    if not name:
                        continue
                    normalized.append({
                        "order": len(normalized),
                        "name": name,
                        "mandatory": True,
                        "slug": slugify(name),
                        "row": ridx,
                    })
        else:
            messages.error(request, "Unsupported file type. Use .csv or .xlsx.")
            return redirect("operaciones:listar_billing")

    except Exception as e:
        messages.error(request, f"Could not parse the file: {e}")
        return redirect("operaciones:listar_billing")

    if not normalized:
        messages.warning(request, "No valid rows found in the file.")
        return redirect("operaciones:listar_billing")

    # --------- Dedup dentro del archivo (igual que tu lógica mejorada) --------
    seen = {}  # slug -> {"row": int, "name": str}
    cleaned = []
    file_duplicates = []

    for r in normalized:
        slug = r.get("slug") or ""
        rownum = r.get("row")
        nm = (r.get("name") or "").strip()

        if not slug:
            errors.append(f"Row {rownum}: invalid name.")
            continue

        if slug in seen:
            first = seen[slug]
            warnings.append(
                f"Row {rownum}: duplicated name in file — '{nm}' — duplicates row {first['row']} ('{first['name']}')."
            )
            file_duplicates.append({
                "row": rownum,
                "name": nm,
                "dup_of_row": first["row"],
                "dup_of_name": first["name"],
                "slug": slug,
            })
            continue

        seen[slug] = {"row": rownum, "name": nm}
        cleaned.append(r)

    normalized = cleaned

    if not normalized and not errors:
        messages.warning(request, "No valid rows found in the file.")
        return redirect("operaciones:listar_billing")

    # -------- Cargar sesiones + filtrar por estado permitido --------
    sesiones = list(SesionBilling.objects.filter(id__in=ids).order_by("-creado_en"))

    allowed_states = {"asignado", "en_proceso", "en_revision_supervisor"}
    blocked_states = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}  # por seguridad

    eligible = []
    skipped = []

    for s in sesiones:
        est = (getattr(s, "estado", "") or "").strip()
        if est in blocked_states:
            skipped.append({"id": s.id, "proyecto_id": s.proyecto_id, "estado": est, "reason": "Already approved."})
            continue
        if est not in allowed_states:
            skipped.append({"id": s.id, "proyecto_id": s.proyecto_id, "estado": est, "reason": "Status not allowed."})
            continue
        eligible.append(s)

    if not eligible:
        messages.error(request, "None of the selected billings can receive requirements (status not allowed or already approved).")
        return redirect("operaciones:listar_billing")

    # -------- Por sesión: duplicates vs existentes + to_create --------
    bulk_preview = []
    total_create = 0
    total_dup = 0

    for s in eligible:
        asignaciones = list(
            s.tecnicos_sesion.select_related("tecnico").prefetch_related("requisitos").all()
        )

        existing_by_slug = {}  # slug -> {"row_num":int,"name":str,"orden":int,"obligatorio":bool}
        if asignaciones:
            canonical_qs = asignaciones[0].requisitos.order_by("orden", "id")
            for idx, req in enumerate(canonical_qs, start=1):
                key = slugify((req.titulo or "").strip())
                if key and key not in existing_by_slug:
                    existing_by_slug[key] = {
                        "row_num": idx,
                        "name": (req.titulo or "").strip(),
                        "orden": getattr(req, "orden", 0),
                        "obligatorio": bool(getattr(req, "obligatorio", True)),
                    }

        duplicates = []
        to_create = []

        for r in normalized:
            ex = existing_by_slug.get(r["slug"])
            if ex:
                duplicates.append({
                    "incoming_row": r.get("row"),
                    "existing_row_num": ex["row_num"],
                    "name": r["name"],
                    "existing_name": ex["name"],
                    "existing_order": ex["orden"],
                    "existing_mandatory": ex["obligatorio"],
                    "incoming_order": r["order"],
                    "incoming_mandatory": r["mandatory"],
                })
            else:
                to_create.append({
                    "row": r.get("row"),
                    "name": r["name"],
                    "order": r["order"],
                    "mandatory": r["mandatory"],
                    "slug": r["slug"],
                })

        total_create += len(to_create)
        total_dup += len(duplicates)

        bulk_preview.append({
            "sesion_id": s.id,
            "proyecto_id": s.proyecto_id,
            "estado": getattr(s, "estado", ""),
            "to_create": to_create,
            "duplicates": duplicates,
        })

    # -------- Guardar preview en session --------
    request.session["bulk_req_import_preview"] = {
        "source_filename": filename,
        "eligible_ids": [s.id for s in eligible],
        "skipped": skipped,
        "file_duplicates": file_duplicates,
        "warnings": warnings,
        "errors": errors,
        "bulk_preview": bulk_preview,
    }

    # ✅ IMPORTANTE: reutilizamos el template que ya existe (preview_import_requirements.html)
    return render(
        request,
        "operaciones/preview_import_requirements.html",
        {
            "is_bulk": True,
            "source_filename": filename,
            "skipped": skipped,
            "file_duplicates": file_duplicates,
            "warnings": warnings,
            "errors": errors,
            "bulk_preview": bulk_preview,
            "total_create": total_create,
            "total_duplicates": total_dup,
        },
    )


@login_required
@rol_requerido("supervisor", "admin", "pm")
@require_POST
def bulk_confirmar_importar_requisitos(request):
    """
    BULK CONFIRM:
    - Lee preview desde session
    - Para cada sesión elegible:
        crea SOLO requisitos nuevos en TODAS las asignaciones
        NO modifica existentes
    - Bloquea si alguna sesión cambió a estado no permitido entre preview y confirm
    - Marca automáticamente POWER PORT X como lectura de potencia
    """
    payload = request.session.get("bulk_req_import_preview") or {}
    bulk_preview = payload.get("bulk_preview") or []

    if not bulk_preview:
        messages.error(
            request, "No preview data to confirm. Please re-upload the file."
        )
        return redirect("operaciones:listar_billing")

    allowed_states = {"asignado", "en_proceso", "en_revision_supervisor"}
    blocked_states = {"aprobado_supervisor", "aprobado_pm", "aprobado_finanzas"}

    created_total = 0
    affected_sessions = 0
    blocked_now = []

    try:
        with transaction.atomic():
            for block in bulk_preview:
                sesion_id = block.get("sesion_id")
                to_create = block.get("to_create") or []

                if not sesion_id or not isinstance(to_create, list):
                    continue

                s = get_object_or_404(SesionBilling, pk=int(sesion_id))
                est = (getattr(s, "estado", "") or "").strip()

                if est in blocked_states or est not in allowed_states:
                    blocked_now.append(
                        {
                            "id": s.id,
                            "proyecto_id": s.proyecto_id,
                            "estado": est,
                        }
                    )
                    continue

                asignaciones = list(
                    s.tecnicos_sesion.select_related("tecnico")
                    .prefetch_related("requisitos")
                    .all()
                )

                for a in asignaciones:
                    existentes_por_slug = {
                        slugify((r.titulo or "").strip()): True
                        for r in a.requisitos.all()
                    }

                    for r in to_create:
                        name = (r.get("name") or "").strip()
                        if not name:
                            continue

                        key = slugify(name)
                        if not key:
                            continue

                        if key in existentes_por_slug:
                            continue

                        try:
                            order = int(r.get("order"))
                        except Exception:
                            order = len(existentes_por_slug)

                        mandatory = bool(r.get("mandatory", True))

                        # ✅ NUEVO: detectar POWER PORT 1..8
                        needs_power, port_no = _power_meta_from_title(name)

                        RequisitoFotoBilling.objects.create(
                            tecnico_sesion=a,
                            titulo=name,
                            descripcion="",
                            obligatorio=mandatory,
                            orden=order,
                            needs_power_reading=bool(needs_power),
                            power_port_no=port_no,
                        )

                        created_total += 1
                        existentes_por_slug[key] = True

                affected_sessions += 1

    except Exception as e:
        messages.error(request, f"Could not apply bulk requirements: {e}")
        return redirect("operaciones:listar_billing")
    finally:
        request.session.pop("bulk_req_import_preview", None)

    if blocked_now:
        messages.warning(
            request,
            f"Created {created_total} requirements. Some billings were skipped because their status changed.",
        )
    else:
        messages.success(
            request,
            f"Created {created_total} new requirement(s) across {affected_sessions} billing(s). Existing identical requirements were kept unchanged.",
        )

    return redirect("operaciones:listar_billing")


# ============================
# PM — Aprobación/Rechazo PROYECTO
# ============================

@login_required
@rol_requerido('pm', 'admin')
def pm_aprobar_proyecto(request, sesion_id):
    s = get_object_or_404(SesionBilling, pk=sesion_id)
    if s.estado not in ("aprobado_supervisor",):
        messages.error(
            request, "El proyecto aún no está aprobado por Supervisor.")
        return redirect("operaciones:revisar_sesion", sesion_id=s.id)
    s.estado = "aprobado_pm"
    s.save(update_fields=["estado"])
    messages.success(request, "Proyecto aprobado por PM.")
    return redirect("operaciones:listar_billing")


@login_required
@rol_requerido('pm', 'admin')
def pm_rechazar_proyecto(request, sesion_id):
    s = get_object_or_404(SesionBilling, pk=sesion_id)
    s.estado = "rechazado_pm"
    s.save(update_fields=["estado"])
    messages.warning(request, "Proyecto rechazado por PM.")
    return redirect("operaciones:revisar_sesion", sesion_id=s.id)


# ============================
# ELIMINAR EVIDENCIA (corregido)
# ============================

@login_required
@rol_requerido('usuario', 'supervisor', 'admin', 'pm')
@require_POST
def eliminar_evidencia(request, pk, evidencia_id):
    """
    El técnico puede borrar en 'en_proceso' o si fue rechazado con reintento.
    Supervisor/Admin/PM pueden borrar mientras el proyecto NO esté aprobado por supervisor/PM.
    Una vez que el supervisor aprueba (o PM aprueba), no se permite borrar.
    """
    a = get_object_or_404(SesionBillingTecnico, pk=pk)
    s = a.sesion

    # ✅ NUEVO: si es el dueño y la asignación está inactiva, NO puede borrar
    if a.tecnico_id == request.user.id and not _is_asig_active(a):
        return HttpResponseForbidden("This assignment is no longer available.")

    # 🔒 Candado por estado del proyecto: si ya fue aprobado por supervisor o PM, no se permite borrar
    if s.estado in ("aprobado_supervisor", "aprobado_pm"):
        messages.error(
            request, "Photos cannot be deleted after supervisor approval.")
        next_url = (
            request.POST.get("next")
            or (reverse("operaciones:upload_evidencias", args=[a.pk]) if a.tecnico_id == request.user.id else reverse("operaciones:revisar_sesion", args=[s.pk]))
        )
        return redirect(next_url)

    # ¿Quién es?
    is_owner = (a.tecnico_id == request.user.id)
    is_staff_role = getattr(request.user, "rol", None) in {
        "supervisor", "admin", "pm"}

    # Reglas para técnico: sólo en proceso o rechazado con reintento habilitado
    can_owner_delete = (
        a.estado == "en_proceso"
        or (a.estado == "rechazado_supervisor" and a.reintento_habilitado)
    )

    # Staff puede borrar mientras NO esté aprobado (ya validado arriba)
    if not (is_staff_role or (is_owner and can_owner_delete)):
        return HttpResponseForbidden("You can't delete photos at this stage.")

    ev = get_object_or_404(EvidenciaFotoBilling,
                           pk=evidencia_id, tecnico_sesion=a)

    # Eliminar archivo físico si existe (ignorar errores del storage)
    try:
        ev.imagen.delete(save=False)
    except Exception:
        pass

    # Eliminar registro
    ev.delete()

    # Mensaje al usuario (en inglés)
    messages.success(request, "Photo deleted.")

    # Redirección: usar 'next' si viene, si no, a la vista apropiada (técnico vs staff)
    next_url = (
        request.POST.get("next")
        or (reverse("operaciones:upload_evidencias", args=[a.pk]) if is_owner else reverse("operaciones:revisar_sesion", args=[s.pk]))
    )
    return redirect(next_url)


# views.py

# asume tu modelo


@csrf_protect
def update_semana_pago_real(request, sesion_id):
    """
    Inline update for 'Real pay week' (YYYY-W##).
    Returns JSON always, with user-facing messages in English.
    """
    # --- Method check ---
    if request.method != "POST":
        return HttpResponseNotAllowed(["POST"])

    # --- Detect AJAX/XHR ---
    is_ajax = request.headers.get("x-requested-with") == "XMLHttpRequest"

    # --- Auth check (avoid 302 for AJAX) ---
    if not request.user.is_authenticated:
        if is_ajax:
            return JsonResponse(
                {"ok": False, "error": "Your session has expired. Please sign in again."},
                status=401,
            )
        return redirect_to_login(request.get_full_path())

    # --- Role check (admin | pm | facturacion). Ajusta a tu helper real ---
    allowed = False
    for attr in ("tiene_rol", "has_role"):
        fn = getattr(request.user, attr, None)
        if callable(fn) and fn("admin", "pm", "facturacion"):
            allowed = True
            break
    if request.user.is_superuser:
        allowed = True
    if not allowed:
        return JsonResponse(
            {"ok": False, "error": "You do not have permission to edit the real pay week."},
            status=403,
        )

    # --- Load session ---
    s = get_object_or_404(SesionBilling, pk=sesion_id)

    # --- Business lock: PAID only admin/superuser can change ---
    if getattr(s, "finance_status", None) == "paid" and not request.user.is_superuser:
        return JsonResponse(
            {"ok": False, "error": "Locked (PAID). Only admins can edit."},
            status=403,
        )

    # --- Read value ---
    raw = (request.POST.get("semana") or "").strip()

    # 1) Empty => clear
    if raw == "":
        s.semana_pago_real = ""
        s.save(update_fields=["semana_pago_real"])
        return JsonResponse({"ok": True, "semana": ""})

    # 2) Normalization
    v = raw.lower().replace(" ", "")
    now = timezone.now()
    cur_year = now.isocalendar().year

    # Parse
    if re.fullmatch(r"\d{4}-w?\d{1,2}", v):            # 2025-w3, 2025-W34
        y, w = re.split(r"-w?", v)
        year = int(y)
        week = int(w)
    elif re.fullmatch(r"w?\d{1,2}", v):                # w34, 34 -> current year
        year = cur_year
        week = int(v.lstrip("w"))
    elif re.fullmatch(r"\d{1,2}/\d{4}", v):            # 34/2025
        w, y = v.split("/")
        year = int(y)
        week = int(w)
    elif re.fullmatch(r"\d{4}/\d{1,2}", v):            # 2025/34
        y, w = v.split("/")
        year = int(y)
        week = int(w)
    elif re.fullmatch(r"\d{4}-W\d{2}", raw):           # already correct
        s.semana_pago_real = raw
        s.save(update_fields=["semana_pago_real"])
        return JsonResponse({"ok": True, "semana": s.semana_pago_real})
    else:
        return JsonResponse(
            {
                "ok": False,
                "error": "Invalid format. Use: 2025-W34, W34, 34, 34/2025, or 2025/34.",
            },
            status=400,
        )

    # 3) Range check
    if not (1 <= week <= 53):
        return JsonResponse(
            {"ok": False, "error": "Week must be between 1 and 53."},
            status=400,
        )

    # 4) Save normalized YYYY-W##
    value_norm = f"{year}-W{week:02d}"
    s.semana_pago_real = value_norm
    s.save(update_fields=["semana_pago_real"])
    return JsonResponse({"ok": True, "semana": value_norm})
